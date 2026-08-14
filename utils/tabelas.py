"""
utils/tabelas.py — tabelas de leitura em HTML, com exportação espelhada.

Uma tabela é uma matriz de CÉLULAS (`cel`), cada uma com valor, formato e estilo.
A mesma matriz alimenta duas renderizações: `tabela_html` desenha a tela e
`tabela_excel` gera o arquivo. Assim o Excel sai igual ao que se vê — o que não
era possível quando a tela vinha de cellStyle em JavaScript e o arquivo de
openpyxl, dois caminhos que divergem sem avisar.

Usado pela página de Head-to-Head e pela de Análise Conjunta. Se precisar mudar
cor, formato ou espaçamento, mude aqui: as duas páginas seguem juntas.
"""

import io

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

MIME_XLSX = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")

FMT_XL = {"pct1": '0.0"%"', "pct0": '0"%"', "num1": '0.0', "num0": '0',
          "sinal1": '+0.0;-0.0', "txt": None}


def _texto(val, tipo):
    if val is None or (isinstance(val, float) and np.isnan(val)) or val is pd.NA:
        return ""
    if tipo == "pct1":
        return f"{val:.1f}".replace(".", ",") + "%"
    if tipo == "pct0":
        return f"{val:.0f}%"
    if tipo == "num1":
        return f"{val:.1f}".replace(".", ",")
    if tipo == "num0":
        return f"{val:.0f}"
    if tipo == "sinal1":
        return ("+" if val >= 0 else "−") + f"{abs(val):.1f}".replace(".", ",")
    return str(val)


def cel(val, tipo="txt", cor="#1A1A1A", bg=None, bold=False, align=None, barra=None):
    """Uma célula. `barra` (0–100) desenha a barra proporcional atrás do número."""
    return {"val": val, "tipo": tipo, "txt": _texto(val, tipo), "cor": cor, "bg": bg,
            "bold": bold, "align": align or ("left" if tipo == "txt" else "right"),
            "barra": barra}


def hdr(txt, bg=None, cor=None):
    """Cabeçalho com cor própria — usado para repetir na coluna a cor da
    classificação, o que dispensa legenda para aquelas colunas."""
    return {"txt": txt, "bg": bg or "#4A4A4A", "cor": cor or "#FFFFFF"}


def tabela_html(headers, linhas, largura_1a=190, altura_max=560):
    """HTML da tabela. Primeira coluna fixa à esquerda; numeral tabular."""
    headers = [h if isinstance(h, dict) else hdr(h) for h in headers]
    borda = "1px solid #E5E7EB"
    th = ('padding:9px 10px;font-weight:700;'
          'font-size:12.5px;text-align:center;position:sticky;top:0;z-index:3;'
          'white-space:nowrap;')
    out = [f'<div style="overflow:auto;max-height:{altura_max}px;border:{borda};'
           f'border-radius:8px;">'
           '<table style="border-collapse:separate;border-spacing:0;width:100%;'
           'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;font-size:13px;'
           'font-variant-numeric:tabular-nums;">']
    out.append("<thead><tr>")
    for i, h in enumerate(headers):
        extra = (f'left:0;z-index:4;min-width:{largura_1a}px;text-align:left;' if i == 0 else '')
        cor_h = f'background:{h["bg"]};color:{h["cor"]};'
        out.append(f'<th style="{th}{cor_h}{extra}">{h["txt"]}</th>')
    out.append("</tr></thead><tbody>")
    for ri, linha in enumerate(linhas):
        fundo = "#FFFFFF" if ri % 2 == 0 else "#FAFAFA"
        out.append("<tr>")
        for ci, c in enumerate(linha):
            bg = c["bg"] or fundo
            peso = "700" if c["bold"] else "400"
            fixa = (f'position:sticky;left:0;z-index:2;min-width:{largura_1a}px;'
                    if ci == 0 else '')
            estilo = (f'padding:7px 10px;border-bottom:{borda};background:{bg};'
                      f'color:{c["cor"]};font-weight:{peso};text-align:{c["align"]};'
                      f'white-space:nowrap;{fixa}')
            if c["barra"] is not None:
                larg = max(0.0, min(100.0, float(c["barra"])))
                conteudo = (
                    '<div style="display:flex;align-items:center;gap:8px;">'
                    f'<span style="flex:0 0 46px;text-align:right;">{c["txt"]}</span>'
                    '<span style="flex:1;min-width:52px;background:#EDEFF2;height:9px;'
                    'border-radius:5px;display:inline-block;overflow:hidden;">'
                    f'<span style="display:block;width:{larg:.1f}%;height:100%;'
                    f'background:{c["cor"]};"></span></span></div>')
            else:
                conteudo = c["txt"]
            out.append(f'<td style="{estilo}">{conteudo}</td>')
        out.append("</tr>")
    out.append("</tbody></table></div>")
    return "".join(out)


def tabela_excel(headers, linhas) -> bytes:
    """Mesma matriz de células, agora em xlsx: mesmas cores, mesmo alinhamento e
    mesmo formato. Números continuam números — o símbolo vem do formato."""
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "H2H"
    thin = Side(style="thin", color="D9D9D9")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [h if isinstance(h, dict) else hdr(h) for h in headers]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=str(h["txt"]))
        c.font = Font(bold=True, name="Arial", size=10, color=h["cor"].replace("#", ""))
        c.fill = PatternFill("solid", start_color=h["bg"].replace("#", ""))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = \
            max(11, min(30, len(str(h["txt"])) + 4))
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "B2"

    for ri, linha in enumerate(linhas, start=2):
        for ci, cl in enumerate(linha, start=1):
            escrever = cl["val"]
            if cl["tipo"] == "txt" or escrever is None or escrever is pd.NA:
                escrever = cl["txt"] or None
            elif isinstance(escrever, float) and np.isnan(escrever):
                escrever = None
            c = ws.cell(row=ri, column=ci, value=escrever)
            c.font = Font(name="Arial", size=10, bold=cl["bold"],
                          color=cl["cor"].replace("#", ""))
            if cl["bg"]:
                c.fill = PatternFill("solid", start_color=cl["bg"].replace("#", ""))
            fmt = FMT_XL.get(cl["tipo"])
            if fmt:
                c.number_format = fmt
            c.alignment = Alignment(horizontal=cl["align"], vertical="center")
            c.border = brd
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# Linhas e colunas de resumo (média do local, aproveitamento da linha, totais)
# saem em cinza: são contexto de leitura, não dado comparável com o resto.
BG_RESUMO = "#F0F1F3"
FG_RESUMO = "#4A4A4A"


def cel_resumo(val, tipo="txt", align=None, bold=True):
    return cel(val, tipo, cor=FG_RESUMO, bg=BG_RESUMO, bold=bold,
               align=align or ("left" if tipo == "txt" else "center"))


def legenda_cores(itens):
    """itens: lista de (cor, rótulo, modo) — modo 'bg' pinta o quadrado, 'txt' pinta
    o texto de exemplo. Devolve o HTML da legenda."""
    if not itens:
        return ""
    partes = ['<div style="display:flex;flex-wrap:wrap;gap:16px;align-items:center;'
              'margin:8px 0 2px;font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;'
              'font-size:12.5px;color:#374151;">'
              '<span style="font-weight:700;color:#6B7280;">LEGENDA</span>']
    for cor, rotulo, modo in itens:
        if modo == "bg":
            amostra = (f'<span style="display:inline-block;width:15px;height:15px;'
                       f'border-radius:3px;background:{cor};border:1px solid #D9DCE1;'
                       f'vertical-align:-3px;"></span>')
        elif modo == "barra":
            amostra = ('<span style="display:inline-block;width:34px;height:9px;'
                       'border-radius:5px;background:#EDEFF2;vertical-align:0px;'
                       'overflow:hidden;">'
                       f'<span style="display:block;width:62%;height:100%;background:{cor};">'
                       '</span></span>')
        else:
            amostra = (f'<span style="color:{cor};font-weight:700;">123</span>')
        partes.append(f'<span style="display:inline-flex;align-items:center;gap:6px;">'
                      f'{amostra}<span>{rotulo}</span></span>')
    partes.append('</div>')
    return "".join(partes)


def render_tabela(headers, linhas, nome_arquivo, key, largura_1a=190, altura_max=560,
                  label="⬇️ Exportar Excel (igual ao que está na tela)", legenda=None,
                  sufixo=""):
    """Desenha a tabela, a legenda e o botão de exportar. `sufixo` entra no nome do
    arquivo para marcar o recorte (ex.: "_MT_2526")."""
    st.markdown(tabela_html(headers, linhas, largura_1a, altura_max), unsafe_allow_html=True)
    if legenda:
        st.markdown(legenda_cores(legenda), unsafe_allow_html=True)
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    st.download_button(label, data=tabela_excel(headers, linhas),
                       file_name=f"{nome_arquivo}{sufixo}.xlsx",
                       mime=MIME_XLSX, key=key)


