"""
pages/5_Perdas.py — Perdas e fenômenos da colheita (milho)

Decomposição das perdas e fenômenos registrados na av4, por híbrido. Fonte:
tabela_analitica_faixa das safras 2024/25 e 2025/26 (só Faixa). Segue Better Data
Visualization (Schwabish). A sanidade foliar (doenças) fica na página própria de Sanidade.

Cobre av4:
  - perdas: acamadas, quebradas, dominadas, colmo podre e o total — média COM zero (taxa na rede);
  - fenômenos: green snap, morte prematura, má formação de espigas, enfezamento-contagem —
    média SÓ ONDE OCORREU (o zero fica de fora, para não achatar; fenômeno é raro).

Cuidados de leitura (valem para a página inteira):
  - a nota de doença é INVERSA (9 é bom); perdas e fenômenos são percentuais (alto é ruim).
    Nunca usar a mesma paleta nos dois blocos;
  - enfezamento é medido DUAS vezes (nota na av2, contagem na av4). São escalas e momentos
    diferentes: não somar, não plotar na mesma série;
  - fenômenos só existem em 2025 (ajuste de protocolo). Em 24/25 as colunas não vêm;
  - perda e fenômeno usam a MESMA régua e vêm da MESMA subamostra no app: média das contagens
    ÷ estande final × 100. A contagem 0 conta (o avaliador percorreu os 10 m e não achou); só a
    subamostra não avaliada fica de fora. O percentual é a taxa do plot.

Seções (construídas incrementalmente):
  1. Auditoria — tabela analítica de sanidade por ensaio
"""
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt

# ── Cores por status do material (milho: CHECK / STINE / EXP / DP2) ──────────
COR_STATUS_PLOT = {
    "CHECK": "#F4B184",   # testemunha externa (laranja)
    "STINE": "#2976B6",   # comercial Stine (azul)
    "EXP":   "#00FF00",   # experimental / em avaliação (verde vibrante)
    "DP2":   "#C4DFB4",   # duplo propósito / segundo ano (verde claro)
}
COR_TEXTO_STATUS = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                    "DP2": "#1A1A1A", "": "#000000"}
COR_BORDA = {
    "CHECK": "#C46A3A",
    "STINE": "#1A4F7A",
    "EXP":   "#009900",
    "DP2":   "#7AAF6A",
}

# ── Classes de reação a doenças (escala INVERSA: 9 = mais resistente) ────────
ORDEM_CLASS = ["AS", "S", "MT", "T", "R"]
COR_CLASS = {
    "AS": "#8B0000",   # vermelho escuro — altamente suscetível
    "S":  "#E63946",   # vermelho        — suscetível
    "MT": "#FFD600",   # amarelo         — moderadamente tolerante
    "T":  "#70C96E",   # verde claro     — tolerante
    "R":  "#1E7A34",   # verde escuro    — resistente
}
COR_TEXTO_CLASS = {"AS": "#FFFFFF", "S": "#FFFFFF", "MT": "#1A1A1A",
                   "T": "#1A1A1A", "R": "#FFFFFF"}
LABEL_CLASS = {
    "AS": "AS — Altamente suscetível (nota 1–2)",
    "S":  "S — Suscetível (3–4)",
    "MT": "MT — Moderadamente tolerante (5–6)",
    "T":  "T — Tolerante (7–8)",
    "R":  "R — Resistente (9)",
}


def nota_para_classe(nota):
    """Converte nota 1-9 em sigla de classe. Mesma régua do pipeline (_classificar_doenca)."""
    if nota is None or (isinstance(nota, float) and np.isnan(nota)):
        return None
    n = float(nota)
    if n <= 2:
        return "AS"
    if n <= 4:
        return "S"
    if n <= 6:
        return "MT"
    if n <= 8:
        return "T"
    return "R"


def resumo_doenca(grp: pd.DataFrame, col_nota: str) -> tuple:
    """Nota típica, incidência e classe de UMA doença num recorte de plots.

    Fonte única das Seções 2 e 4 — se cada uma calculasse por conta, uma hora divergiriam.
      - Nota  = moda das notas válidas; em caso de EMPATE fica a MENOR nota (lado suscetível);
      - Inc.  = locais com detecção (nota 1–5) ÷ locais em que a doença foi avaliada;
      - Classe = régua do pipeline aplicada à nota.
    Nota 0 é "não avaliado" e sai de todas as contas. Devolve (None, None, None) sem dado."""
    if col_nota not in grp.columns:
        return None, None, None
    g = grp[["cod_fazenda", col_nota]].copy()
    g[col_nota] = pd.to_numeric(g[col_nota], errors="coerce").where(lambda x: x > 0)
    g = g.dropna(subset=[col_nota])
    if g.empty:
        return None, None, None
    nota = round(float(g[col_nota].mode().min()), 1)
    n_aval = g["cod_fazenda"].nunique()
    # o local conta como "teve" se ao menos um plot dele detectou a doença
    n_com = g.loc[g[col_nota].between(1, 5), "cod_fazenda"].nunique()
    inc = round(n_com / n_aval * 100, 1) if n_aval else None
    return nota, inc, nota_para_classe(nota)


# ── Doenças da av2 (6) ───────────────────────────────────────────────────────
# `inc` não vem do pipeline (não está em MET_AV2): é recalculado aqui a partir da nota.
DOENCAS = {
    "Turcicum":          {"nota": "nota_turcicum",          "class": "class_nota_turcicum"},
    "Cercospora":        {"nota": "nota_cercospora",        "class": "class_nota_cercospora"},
    "Mancha branca":     {"nota": "nota_mancha_branca",     "class": "class_nota_mancha_branca"},
    "Bipolaris":         {"nota": "nota_bipolaris",         "class": "class_nota_bipolaris"},
    "Ferrugem tropical": {"nota": "nota_ferrugem_tropical", "class": "class_nota_ferrugem_tropical"},
    "Enfezamento":       {"nota": "nota_enfezamento",       "class": "class_nota_enfezamento"},
}

# ── Perdas da av4 (média das subamostras avaliadas ÷ estande final; somam no total) ──
PERDAS = {
    "Acamamento":  "pct_acamadas",
    "Quebramento": "pct_quebradas",
    "Dominadas":   "pct_dominadas",
    "Colmo Podre": "pct_colmo_podre",
}
PERDA_TOTAL = "pct_perda_total"

# ── Fenômenos da av4 (mesma régua e mesma subamostra das perdas; só 2025) ──
# NÃO somam com pct_perda_total — denominador diferente.
FENOMENOS = {
    "Green snap":             {"pct": "pct_green_snap",           "n": "green_snap_plantas"},
    "Morte prematura":        {"pct": "pct_morte_prematura",      "n": "morte_prematura_plantas"},
    "Má formação de espigas": {"pct": "pct_ma_formacao_espigas",  "n": "ma_formacao_espigas_plantas"},
    "Enfezamento (contagem)": {"pct": "pct_enfezamento",          "n": "enfezamento_plantas"},
}

# ordem de exibição dos materiais por status (CHECK primeiro, DP2 por último)
ORDEM_STATUS = ["CHECK", "STINE", "EXP", "DP2"]

# ── Sanidade medida só em 2024 (mudou de instrumento em 2025) ────────────────
# tombamento verde virou a contagem de green snap na av4; grãos ardidos saiu do protocolo.
LEGADO_2024 = {
    "Nota tombamento verde (24/25)": "nota_tombamento_verde",
    "Grãos ardidos % (24/25)":       "graos_ardidos_pct",
}

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_multisafra
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

st.set_page_config(
    page_title="Perdas · JAUM DTC",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

# ── Helper AgGrid (mesma configuração da soja: header escuro, menu funcionando) ─
def ag_table(df, height=400, estilos_col=None, renderers_col=None):
    """Tabela AgGrid padrão. `estilos_col`: {coluna: JsCode} para cellStyle ·
    `renderers_col`: {coluna: JsCode} para cellRenderer (permite barras na célula)."""
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    for _col, _estilo in (estilos_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellStyle=_estilo)
    for _col, _rend in (renderers_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellRenderer=_rend, minWidth=170)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                   {"background-color": "#4A4A4A !important"},
            ".ag-header-row":               {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":              {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":        {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":         {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                     {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":  {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-icon-menu":                {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # sem "color" aqui: um !important sobreporia a cor inline do cellStyle
            # (era o que impedia o verde/vermelho dos desvios de aparecer)
            ".ag-cell":                     {"font-size": "13px !important"},
            ".ag-row":                      {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True,
    )


# ── Helper exportar Excel (mesmo padrão da soja) ──────────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = border

    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


# ── Header ────────────────────────────────────────────────────────────────────
page_header(
    "Sanidade",
    "Doenças na lavoura (av2) e o que apareceu na colheita (av4) — perdas e fenômenos. "
    "Comece conferindo os dados por ensaio na Auditoria; as análises agregadas vêm em seguida.",
    imagem="Researchers-pana.png",
)

# ── Carregamento: analítica de Faixa das duas safras, já reconciliada ─────────
# Usa carregar_multisafra(), que aplica o depara_mestre — sem ele, o mesmo híbrido
# aparece com nomes diferentes em cada safra.
@st.cache_data(show_spinner=False)
def carregar_concat():
    d = carregar_multisafra()
    df = d.get("tabela_analitica_faixa")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    df = df.copy()
    # produtividade canônica: usa a válida (kg/ha e sacas), com fallback para a bruta
    if "produtividade_valida_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_valida_kg_ha"], errors="coerce")
        if "produtividade_kg_ha" in df.columns:
            df["kg_ha"] = df["kg_ha"].fillna(pd.to_numeric(df["produtividade_kg_ha"], errors="coerce"))
    elif "produtividade_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_kg_ha"], errors="coerce")
    if "produtividade_valida_sacas_ha" in df.columns:
        df["sc_ha"] = pd.to_numeric(df["produtividade_valida_sacas_ha"], errors="coerce")
    elif "kg_ha" in df.columns:
        df["sc_ha"] = (df["kg_ha"] / 60).round(1)   # 1 saca = 60 kg
    # altura em metros (o pipeline entrega em cm) — mesma régua da Análise Conjunta
    if "altura_planta_cm" in df.columns:
        df["altura_planta_m"] = (pd.to_numeric(df["altura_planta_cm"], errors="coerce") / 100).round(1)
    if "altura_espiga_cm" in df.columns:
        df["altura_espiga_m"] = (pd.to_numeric(df["altura_espiga_cm"], errors="coerce") / 100).round(1)
    return df


with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

# Só interessam os plots com ALGUMA medida de sanidade (doença, perda ou fenômeno).
# Um plot só com produtividade não tem o que auditar aqui.
COLS_PERDAS_FEN = (list(PERDAS.values()) + [PERDA_TOTAL]
                   + [f["pct"] for f in FENOMENOS.values()])
_cols_pf_existentes = [c for c in COLS_PERDAS_FEN if c in ta_raw.columns]
# mantém linhas com produtividade OU algum dado de perda/fenômeno — a Auditoria de perdas precisa
# conferir a produção junto, então não exige que a perda exista para a linha aparecer.
if _cols_pf_existentes and "kg_ha" in ta_raw.columns:
    _tem_prod = pd.to_numeric(ta_raw["kg_ha"], errors="coerce").notna()
    _tem_pf = ta_raw[_cols_pf_existentes].notna().any(axis=1)
    ta_raw = ta_raw[_tem_prod | _tem_pf].copy()

if ta_raw.empty:
    st.error("Nenhum registro com produtividade ou dados de perda/fenômeno. Verifique a página de Diagnóstico.")
    st.stop()


# ── Sidebar — filtros encadeados (prefixo `sn_`, independente das outras páginas) ─
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>', unsafe_allow_html=True)

    if st.button("Limpar filtros", use_container_width=True, key="sn_btn_limpar"):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["sn_safra_", "sn_macro_", "sn_micro_", "sn_estado_",
                                               "sn_cidade_", "sn_fazenda_", "sn_resp_", "sn_status_",
                                               "sn_cult_", "sn_busca_", "__opts_sn_"]):
                del st.session_state[key]
        st.rerun()

    def _podar_keys(prefix, opcoes, molde):
        """Remove o estado de checkboxes de opções que saíram da cascata.
        Sem isso, ao reaparecerem elas voltam marcadas e o filtro se reaplica sozinho."""
        antigas = st.session_state.get(f"__opts_{prefix}", [])
        atuais = set(map(str, opcoes))
        for o in antigas:
            if str(o) not in atuais:
                st.session_state.pop(molde(o), None)
        st.session_state[f"__opts_{prefix}"] = list(opcoes)

    def checkboxes(opcoes, default_all=True, defaults=None, prefix=""):
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_{o}")
        sel = []
        for o in opcoes:
            key = f"{prefix}_{o}"
            if key in st.session_state:          # já existe: o widget é a verdade
                marcado = st.checkbox(str(o), key=key)
            else:                                 # primeira criação: aplica o padrão
                checked = (o in defaults) if defaults is not None else default_all
                marcado = st.checkbox(str(o), value=checked, key=key)
            if marcado:
                sel.append(o)
        return sel

    def filtro_busca(opcoes, prefix):
        """Busca textual + seleção persistente. A seleção é lida direto dos checkboxes
        (fonte única), inclusive dos itens ocultos pela busca."""
        if f"{prefix}_reset" not in st.session_state:
            st.session_state[f"{prefix}_reset"] = 0
        r = st.session_state[f"{prefix}_reset"]
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_chk_{r}_{o}")

        busca = st.text_input("Buscar", value="", key=f"sn_busca_{prefix}",
                              placeholder="Digite para filtrar...")
        filtradas = [c for c in opcoes if busca.strip().lower() in str(c).lower()] if busca.strip() else opcoes

        if st.button("Limpar seleção", key=f"{prefix}_limpar", use_container_width=True):
            for o in opcoes:
                st.session_state.pop(f"{prefix}_chk_{r}_{o}", None)
            st.session_state.pop(f"__opts_{prefix}", None)
            st.session_state[f"{prefix}_reset"] = r + 1
            st.rerun()

        for c in filtradas:
            st.checkbox(str(c), key=f"{prefix}_chk_{r}_{c}")

        sel = [o for o in opcoes if st.session_state.get(f"{prefix}_chk_{r}_{o}", False)]
        return sel or opcoes

    # 1. Safra — padrão safra atual (na base o valor é "25/26")
    with st.expander("Safra", expanded=True):
        safras_all = sorted(ta_raw["safra"].dropna().unique().tolist())
        SAFRA_PADRAO = ("25/26", "2025/26")   # aceita os dois formatos
        safra_default = [s for s in safras_all if str(s) in SAFRA_PADRAO] or safras_all[-1:]
        # inicializa o estado UMA vez: garante só a safra padrão marcada
        if "sn_safra_init" not in st.session_state:
            for o in safras_all:
                st.session_state[f"sn_safra_{o}"] = (o in safra_default)
            st.session_state["sn_safra_init"] = True
        safras_sel = checkboxes(safras_all, defaults=safra_default, prefix="sn_safra")
    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("Região Macro", expanded=False):
        macros_sel = checkboxes(sorted(ta_f1["regiao_macro"].dropna().unique().tolist()), prefix="sn_macro")
    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("Região Micro", expanded=False):
        micros_sel = checkboxes(sorted(ta_f2["regiao_micro"].dropna().unique().tolist()), prefix="sn_micro")
    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("Estado", expanded=False):
        estados_sel = filtro_busca(sorted(ta_f3["estado_sigla"].dropna().unique().tolist()), "sn_estado")
    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("Cidade", expanded=False):
        cidades_sel = filtro_busca(sorted(ta_f4["cidade_nome"].dropna().unique().tolist()), "sn_cidade")
    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("Fazenda", expanded=False):
        fazendas_sel = filtro_busca(sorted(ta_f5["nomeFazenda"].dropna().unique().tolist()), "sn_fazenda")
    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Responsável
    with st.expander("Responsável", expanded=False):
        resps_sel = filtro_busca(sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist()), "sn_resp")
    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # 8. Status do híbrido
    with st.expander("Status do Híbrido", expanded=False):
        status_sel = checkboxes(sorted(ta_f7["status_material"].dropna().unique().tolist()), prefix="sn_status")
    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    # 9. Híbrido (dePara)
    with st.expander("Híbrido", expanded=False):
        materiais_sel = filtro_busca(sorted(ta_f8["dePara"].dropna().unique().tolist()), "sn_cult")
    ta_filtrado = ta_f8[ta_f8["dePara"].isin(materiais_sel)] if materiais_sel else ta_f8.iloc[0:0]


def _aplicar_filtros_local(df):
    """Aplica só os filtros de tempo e geografia (nunca híbrido/status) — a referência de um
    ensaio existe independentemente de quais híbridos foram selecionados. Base da produção
    relativa POR LOCAL com referência fixa, igual à Análise Conjunta."""
    if df.empty:
        return df
    d = df
    for _col, _sel in [("safra", safras_sel), ("regiao_macro", macros_sel),
                       ("regiao_micro", micros_sel), ("estado_sigla", estados_sel),
                       ("cidade_nome", cidades_sel), ("nomeFazenda", fazendas_sel),
                       ("nomeResponsavel", resps_sel)]:
        if _sel and _col in d.columns:
            d = d[d[_col].isin(_sel)]
    return d


def _descreve_base(base, testemunha=None, escopo="local"):
    """Texto curto do critério de referência da produção relativa, para o aviso da tabela."""
    if escopo == "material":
        if base == "Maior produtividade":
            return "o **maior** rendimento entre os materiais filtrados (o líder = 100%)"
        if base == "Testemunha" and testemunha:
            return f"a média da testemunha **{testemunha}**"
        return "a **média** dos materiais filtrados"
    if base == "Maior produtividade":
        return "o **maior** rendimento do local (o líder = 100%)"
    if base == "Testemunha" and testemunha:
        return f"a média da testemunha **{testemunha}** no local"
    return "a **média** de todos os híbridos do local"


if ta_filtrado.empty:
    st.warning("Nenhum dado para os filtros selecionados.")
    st.stop()


# ── Contexto: filtros ativos (usado nos subtítulos das seções) ────────────────
def _linha_safra(s, g):
    """Uma linha de contexto por safra, com as regiões/estados que ela realmente contém."""
    partes = [f"<b>Safra:</b> {s}"]
    if "regiao_macro" in g.columns and g["regiao_macro"].notna().any():
        partes.append("Macro: " + ", ".join(sorted(g["regiao_macro"].dropna().unique())))
    if "regiao_micro" in g.columns and g["regiao_micro"].notna().any():
        partes.append("Micro: " + ", ".join(sorted(g["regiao_micro"].dropna().unique())))
    if "estado_sigla" in g.columns and g["estado_sigla"].notna().any():
        partes.append(", ".join(sorted(g["estado_sigla"].dropna().unique())))
    partes.append(f"{g['cidade_nome'].nunique()} cidades")
    partes.append(f"{g['cod_fazenda'].nunique()} locais")
    return " · ".join(partes)


_grupos_safra = list(ta_filtrado.groupby("safra"))
_linhas_safra = [_linha_safra(s, g) for s, g in _grupos_safra]
if len(_grupos_safra) > 1:
    _linhas_safra = ["<b>Análise multissafra</b>"] + _linhas_safra
contexto_str = "<br>".join(_linhas_safra)




# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA (produtividade + agronômicos + perdas + fenômenos, plot a plot)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados de colheita por ensaio?",
    "Visão individual de cada plot: produtividade e agronômicos ao lado das perdas e dos fenômenos. "
    "Use para conferir os dados e cruzar perda com produção antes das análises agregadas.",
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada linha é **um plot** (uma parcela num local). É a tabela mais crua da página — serve para
conferir os dados e cruzar a perda com a produtividade no mesmo plot, antes de olhar as médias.

**Produção Relativa (%) — como é calculada**

Compara o rendimento do plot com uma referência **do próprio local**, escolhida no seletor acima:

```
Prod. Relativa (%) = kg/ha do plot ÷ referência do local × 100
```

A referência pode ser a média do local, o maior rendimento do local, ou uma testemunha. O ponto
importante: ela é calculada sobre **todos os híbridos do ensaio naquele local** (referência fixa),
não sobre os que você filtrou. Assim o filtro de híbrido muda só o que aparece na tabela, nunca o
"100% do local" — o desempenho de um plot é sempre medido contra o ensaio inteiro, não contra os
pares que sobraram no filtro. Um valor acima de 100% significa que o plot rendeu mais que a
referência do seu local; abaixo, menos.

**Cuidado com os filtros da barra lateral**

A produção relativa é **por local e com referência fixa**, então filtrar híbrido não a distorce.
Mas os filtros de **local, estado e safra** mudam quais plots aparecem — e a ordenação e as médias
que você tira de olho mudam junto. Para comparar rendimento ou perda entre híbridos, garanta que
estão no **mesmo conjunto de locais**; senão, você compara ambientes diferentes, não materiais.

**As colunas de perda e fenômeno**

São a **% daquele plot** na av4 (colheita), sem média — é o dado bruto de cada parcela. Para a
leitura por híbrido, com a régua de cada família (perda com zero, fenômeno e qualidade só onde
ocorreu), use a **tabela-resumo** logo abaixo. Aqui é o caso a caso: útil para achar o plot
específico onde uma perda foi alta e ver o que aconteceu com a produtividade dele.

**Ordenação:** por Região → Estado → Cidade → Fazenda e, dentro de cada local, do maior para o
menor kg/ha. Assim os plots de um mesmo ensaio ficam juntos, do melhor ao pior.
""")

col_ref, col_test, _ = st.columns([2, 2, 3])
with col_ref:
    base_rel = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0, key="perdas_base_rel")
with col_test:
    if base_rel == "Testemunha":
        testemunhas = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist())
        testemunha_sel = st.selectbox("Selecione a testemunha", options=testemunhas,
                                      key="perdas_test") if testemunhas else None
        if not testemunhas:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
    else:
        testemunha_sel = None

# produção relativa POR LOCAL, referência fixa do ensaio inteiro (mesma régua da Conjunta)
df_tabela = ta_filtrado.copy()
LOCAL = [c for c in ["safra", "cod_fazenda"] if c in df_tabela.columns] or ["cod_fazenda"]
_ref_scope = _aplicar_filtros_local(ta_raw)
_ref_scope = _ref_scope[pd.to_numeric(_ref_scope["kg_ha"], errors="coerce") > 0]

if base_rel == "Maior produtividade":
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].max()
elif base_rel == "Testemunha" and testemunha_sel:
    _ref = _ref_scope[_ref_scope["dePara"] == testemunha_sel].groupby(LOCAL)["kg_ha"].mean()
else:
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].mean()

_chave = df_tabela.set_index(LOCAL).index
ref_por_local = pd.Series(_chave.map(_ref).to_numpy(), index=df_tabela.index)
df_tabela["prod_relativa_pct"] = ((df_tabela["kg_ha"] / ref_por_local) * 100).round(1)

# col_map = agronômicos da Conjunta + perdas + FENÔMENOS (a novidade da página de Perdas)
col_map = {
    "safra": "Safra", "cod_fazenda": "Cód. Local", "nomeFazenda": "Fazenda",
    "cidade_nome": "Cidade", "estado_sigla": "Estado", "regiao_macro": "Região Macro",
    "regiao_micro": "Região Micro", "nomeResponsavel": "Responsável", "dePara": "Híbrido",
    "status_material": "Status", "indexTratamento": "Trat.",
    "dataPlantioMilho": "Plantio", "dataColheitaMilho": "Colheita",
    "kg_ha": "kg/ha", "sc_ha": "sc/ha", "prod_relativa_pct": "Prod. Relativa (%)",
    "umidade_pct": "Umidade (%)", "populacao_real_plantas_ha": "Pop. Real (pl/ha)",
    "pmg_corrigido_g": "PMG (g)", "fileiras_media": "Fileiras",
    "graos_fileira_media": "Grãos/Fileira", "graos_ardidos_pct": "Ardidos (%)",
    # perdas
    "pct_acamadas": "Acamamento (%)", "pct_quebradas": "Quebramento (%)",
    "pct_dominadas": "Dominadas (%)", "pct_colmo_podre": "Colmo Podre (%)",
    "pct_perda_total": "Perda Total (%)",
    # fenômenos (o que a Conjunta não tem — aqui é o diferencial)
    "pct_green_snap": "Green snap (%)", "pct_morte_prematura": "Morte prematura (%)",
    "pct_ma_formacao_espigas": "Má formação (%)", "pct_enfezamento": "Enfezamento (%)",
}
cols_disp = [c for c in col_map if c in df_tabela.columns]

_hier = [c for c in ["regiao_macro", "regiao_micro", "estado_sigla", "cidade_nome", "nomeFazenda"]
         if c in df_tabela.columns]
if _hier and "kg_ha" in df_tabela.columns:
    df_tabela = df_tabela.sort_values(_hier + ["kg_ha"],
                                      ascending=[True] * len(_hier) + [False],
                                      na_position="last").reset_index(drop=True)

df_show = df_tabela[cols_disp].rename(columns=col_map)
if "Pop. Real (pl/ha)" in df_show.columns:
    df_show["Pop. Real (pl/ha)"] = pd.to_numeric(df_show["Pop. Real (pl/ha)"], errors="coerce").round(0).astype("Int64")
for _c in ["Plantio", "Colheita"]:
    if _c in df_show.columns:
        df_show[_c] = pd.to_datetime(df_show[_c], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")

ag_table(df_show, height=min(560, 40 + 32 * min(len(df_show), 15) + 20))
exportar_excel(df_show, nome_arquivo="auditoria_perdas.xlsx",
               label="⬇️ Exportar Auditoria", key="exp_auditoria_perdas")

st.caption(f"{len(df_show)} observações · {df_tabela['dePara'].nunique()} híbridos · "
           f"{df_tabela['cod_fazenda'].nunique()} locais.")
st.info(f"📐 **Produção Relativa — método POR LOCAL.** Cada plot é comparado com "
        f"{_descreve_base(base_rel, testemunha_sel)}, calculado **dentro de cada local** "
        f"(referência fixa do ensaio inteiro, não muda com o filtro de híbrido). "
        f"As colunas de perda e fenômeno são a % de cada plot na av4 — veja a régua de cada uma "
        f"na tabela-resumo abaixo.")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# PERDAS E FENÔMENOS DA COLHEITA (matriz híbrido × variável da av4)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Resumo de perdas e fenômenos",
    "O que cada híbrido perdeu na colheita?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Para cada híbrido, o percentual médio de cada perda, da qualidade do grão e de cada fenômeno.
**As famílias usam réguas diferentes**, porque se comportam de forma diferente no campo:

- **Perdas** (acamamento, quebramento, dominadas, colmo podre, Perda Total): média de **todos os
  plots avaliados** — o zero entra. É a taxa média da perda na rede, como na Auditoria.
- **Qualidade do grão** (ardidos): média **só onde ocorreu** (plots com valor > 0). Ardido é
  pontual — muitos plots vêm zerados; incluir o zero achataria a média e esconderia quem teve o
  problema. Aqui **0** = avaliado e nunca teve; **vazio** = não avaliado.
- **Fenômenos** (green snap, morte prematura, má formação, enfezamento): média **só onde
  ocorreu**. Fenômeno é raro (aparece em menos de 10% dos plots); se o zero entrasse, a coluna
  ficaria achatada perto de zero e esconderia quem tem o problema. Aqui **0** = avaliado e nunca
  ocorreu; **vazio** = não avaliado.

**Atenção à direção da escala.** Aqui a unidade é **percentual de plantas**: quanto maior, pior.
É o contrário da tabela de doenças, onde a nota vai de 1 a 9 e o 9 é o melhor.

**O que está em cada coluna**

- As quatro perdas — acamamento, quebramento, dominadas e colmo podre — e a **Perda Total**,
  que é a soma das quatro.
- Os quatro fenômenos da colheita: green snap, morte prematura, má formação de espigas e
  enfezamento.

**Como ler os números**

- **Nos fenômenos, o número não diz a frequência.** Um híbrido com green snap "16%" pode ter tido
  o problema em 4 de 35 locais — o 16% é a média onde ocorreu, não em toda a rede. A coluna
  *Locais* mostra o total do híbrido; para ver em quantos ocorreu e o caso a caso, use a Auditoria.
- **Perda Total** é a soma das quatro perdas dentro de cada plot, e sua média segue a régua das
  perdas (com zero). Como cada perda tem o seu conjunto de plots, a média da Perda Total pode não
  ser a soma exata das médias das colunas — é esperado.
- **Perda e fenômeno usam a mesma régua e a mesma subamostra**, mas não somam entre si: as quatro
  perdas já estão somadas no Perda Total, e o fenômeno é outra contagem sobre o mesmo estande.
- **Fenômeno só existe em 2025.** Foi ajuste de protocolo; em 24/25 não houve coleta, e por isso
  essas colunas somem quando o recorte é só daquela safra. O equivalente ao green snap então era
  uma nota de tombamento verde na av2, que está na Auditoria e não pode entrar aqui — nota de 1
  a 9 não se mistura com percentual.
- **A média dilui evento isolado.** Green snap e morte prematura costumam vir de um episódio num
  local só. Um híbrido com 30% de quebra numa fazenda e nada nas outras vinte aparece aqui com
  pouco mais de 1%. Para ver caso a caso, use a Auditoria; para alertas, o Diagnóstico.

**Cor do cabeçalho:** **perdas** em **azul-petróleo**, **qualidade** em **marrom** e **fenômenos**
em **vinho** — famílias diferentes, com réguas diferentes; a cor ajuda a não confundir ao ler.

---

**Cuidado com os filtros da barra lateral**

Os números mudam conforme o recorte ativo. Ao filtrar por local, estado ou safra, a média é
recalculada só com os plots que restaram — um híbrido pode parecer melhor ou pior só porque você
mudou o conjunto de ambientes. Para uma leitura estável, compare híbridos **dentro do mesmo
recorte**, não entre recortes diferentes.

O seletor **Selecionar colunas** (acima da tabela) liga e desliga as variáveis — use para focar no
que interessa sem poluir a tela. Colunas sem nenhum dado no recorte somem sozinhas (é o caso dos
fenômenos em 24/25, que não existiam no protocolo).

**Como interpretar em uma frase:** a tabela responde "quando o problema apareceu, quão forte foi"
para qualidade e fenômenos, e "qual a taxa média na rede" para perdas. Para o caso a caso (em quais
locais, quanto em cada um), use a **Auditoria** no topo da página.
""")

# ── Resumo por híbrido × variável da colheita ────────────────────────────────
# REGRA: a média é SÓ dos plots ONDE O PROBLEMA OCORREU (valor > 0). Responde "quando a perda
# aconteceu / o fenômeno apareceu, quanto foi em média". Os plots em 0 (avaliados, sem ocorrência)
# NÃO entram na média — senão a média achata para perto de zero e esconde quem tem o problema
# (green snap é 0 em 92% dos plots). Três estados distintos por célula:
#   - teve ocorrência → média dos plots com valor > 0;
#   - avaliado e NUNCA teve (só zeros) → 0 (resistente, informação boa — não pode ficar vazio);
#   - não avaliado (sem plot com a métrica) → vazio.
VARS_PERDAS = [(f"{n} (%)", c) for n, c in PERDAS.items()] + [("Perda Total (%)", PERDA_TOTAL)]
VARS_FENOMENOS = [(f"{n} (%)", d["pct"]) for n, d in FENOMENOS.items()]
# qualidade do grão: ardidos. Régua = só onde ocorreu (como fenômeno), mas família própria —
# conceitualmente é qualidade do grão, não perda física de plantas nem fenômeno de estande.
VARS_QUALIDADE = [("Ardidos (%)", "graos_ardidos_pct")]
VARS_COLHEITA = VARS_PERDAS + VARS_QUALIDADE + VARS_FENOMENOS

# rótulos de cada família — usados para montar o cabeçalho e escolher a régua
# régua: perda inclui o zero (taxa média na rede); fenômeno e qualidade usam só onde ocorreu
# (severidade — o problema é pontual e o zero achataria a coluna).
_ROTULOS_FENOMENO = ({rotulo for rotulo, _ in VARS_FENOMENOS}
                     | {rotulo for rotulo, _ in VARS_QUALIDADE})

GRUPO_PERDAS = "Perdas na colheita (% médio na rede)"
GRUPO_FENOMENOS = "Fenômenos (% médio ONDE OCORREU)"
GRUPO_QUALIDADE = "Qualidade do grão (% ONDE OCORREU)"

# ── Seletor de variáveis: liga/desliga colunas, como o "Selecionar doenças" da Sanidade ──
# só entram no seletor as variáveis com algum dado no recorte
_vars_com_dado = [(rot, col) for rot, col in VARS_COLHEITA
                  if col in ta_filtrado.columns
                  and pd.to_numeric(ta_filtrado[col], errors="coerce").notna().any()]
with st.expander("Selecionar colunas", expanded=False):
    st.caption("Escolha quais perdas, qualidade e fenômenos aparecem na tabela.")
    _sel_cols = st.columns(3)
    _vars_sel = []
    for i, (rot, col) in enumerate(_vars_com_dado):
        if _sel_cols[i % 3].checkbox(rot, value=True, key=f"perdas_col_{col}"):
            _vars_sel.append((rot, col))
if not _vars_sel:
    st.info("Selecione ao menos uma coluna para gerar a tabela.")
    st.stop()

# seletor de produtividade: sc/ha sempre visível; kg/ha opcional (igual à Apresentação da Conjunta)
mostrar_kg_perdas = st.checkbox("Mostrar kg/ha", value=True, key="perdas_chk_kg")

perdas_rows = []
for hibrido, grp in ta_filtrado.groupby("dePara", dropna=True):
    _modo_status = grp["status_material"].mode()
    row = {
        "Híbrido": hibrido,
        "Status":  _modo_status.iloc[0] if not _modo_status.empty else "",
        "Locais":  grp["cod_fazenda"].nunique(),
        "kg/ha":   round(grp["kg_ha"].dropna().mean(), 1) if "kg_ha" in grp.columns and grp["kg_ha"].notna().any() else None,
        "sc/ha":   round(grp["sc_ha"].dropna().mean(), 1) if "sc_ha" in grp.columns and grp["sc_ha"].notna().any() else None,
    }
    for rotulo, col in _vars_sel:
        valor = None                         # None = não avaliado → célula vazia
        if col in grp.columns:
            s = pd.to_numeric(grp[col], errors="coerce").dropna()
            if not s.empty:                  # foi avaliado em ao menos um plot
                if rotulo in _ROTULOS_FENOMENO:
                    # FENÔMENO: média só ONDE OCORREU (>0). Fenômeno é raro (zero em ~90% dos
                    # plots); incluir o zero achataria tudo perto de zero e esconderia quem tem o
                    # problema. 0 = avaliado e nunca ocorreu; vazio = não avaliado.
                    com_ocorrencia = s[s > 0]
                    valor = round(float(com_ocorrencia.mean()), 1) if not com_ocorrencia.empty else 0.0
                else:
                    # PERDA: média de TODOS os plots avaliados (o zero entra). Mantém a perda como
                    # taxa média do híbrido na rede, coerente com a Auditoria e a Conjunta. O zero
                    # é medição (plot avaliado sem aquela perda).
                    valor = round(float(s.mean()), 1)
        row[rotulo] = valor
    perdas_rows.append(row)

df_perdas = pd.DataFrame(perdas_rows)

# variáveis sem nenhum dado no recorte saem da tabela (é o caso dos fenômenos em 24/25)
# variáveis sem informação real no recorte saem da tabela:
#  - perdas: só somem se totalmente NaN (o zero é dado real — taxa na rede);
#  - qualidade/fenômenos: somem se NaN OU tudo zero. Um zero aqui pode ser "não ocorreu" ou
#    "dado ainda não veio" (safra em andamento); coluna toda zerada não informa nada e só polui.
#    Quando surgir o primeiro valor > 0, a coluna reaparece sozinha.
_vars_vazias = []
for _rot, _c in _vars_sel:
    if _rot not in df_perdas.columns:
        continue
    _s = pd.to_numeric(df_perdas[_rot], errors="coerce")
    if _s.isna().all():
        _vars_vazias.append(_rot)
    elif _rot in _ROTULOS_FENOMENO and (_s.fillna(0) == 0).all():
        _vars_vazias.append(_rot)   # qualidade/fenômeno todo zero = ainda sem ocorrência real
if _vars_vazias:
    df_perdas = df_perdas.drop(columns=_vars_vazias)

_ordem_p = [s for s in ORDEM_STATUS if s in set(df_perdas["Status"])]
_ordem_p += sorted(set(df_perdas["Status"]) - set(_ordem_p))
df_perdas["Status"] = pd.Categorical(df_perdas["Status"], categories=_ordem_p, ordered=True)
df_perdas = (df_perdas.sort_values(["Status", "sc/ha"], ascending=[True, False], na_position="last")
             .reset_index(drop=True))
df_perdas["Status"] = df_perdas["Status"].astype(str)


# ── AgGrid: número puro, sem classificação de severidade ────────────────────
def ag_table_perdas(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36, groupHeaderHeight=34, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        enableRangeSelection=True,
    )

    _centro = {"fontSize": "13px", "color": "#000000",
               "fontFamily": "Helvetica Neue, sans-serif", "textAlign": "center"}

    for col in df.columns:
        if col.endswith("(%)"):
            gb.configure_column(col, width=130, cellStyle=_centro, headerClass="ag-header-center")

    gb.configure_column("Híbrido", pinned="left", width=170)
    # coluna Status colorida pela cor do material (CHECK laranja, STINE azul, EXP verde, DP2 verde claro)
    _js_bg = ";".join(f"if(v==='{k}')return{{background:'{COR_STATUS_PLOT[k]}',color:'{COR_TEXTO_STATUS.get(k, '#000')}',fontWeight:'700',textAlign:'center'}}"
                      for k in COR_STATUS_PLOT)
    gb.configure_column("Status", width=90,
        cellStyle=JsCode("function(p){var v=p.value;" + _js_bg + ";return{textAlign:'center'};}"))
    gb.configure_column("Locais", width=80, cellStyle=_centro, headerClass="ag-header-center")
    gb.configure_column("kg/ha", width=90, cellStyle=_centro, headerClass="ag-header-center")
    gb.configure_column("sc/ha", width=90, cellStyle=_centro, headerClass="ag-header-center")

    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    # sem sizeColumnsToFit — ver comentário na tabela de doenças

    # ── Cabeçalho de dois níveis: separa visualmente perda de fenômeno ──
    # As duas famílias são percentuais construídos de formas diferentes; o agrupamento
    # deixa isso explícito na tela em vez de depender só do modal.
    _campos_perdas = [rot for rot, _c in VARS_PERDAS if rot in df.columns]
    _campos_fen = [rot for rot, _c in VARS_FENOMENOS if rot in df.columns]
    _por_campo = {cd.get("field"): cd for cd in go["columnDefs"]}
    _defs = []
    for cd in go["columnDefs"]:
        campo = cd.get("field")
        if campo in _campos_perdas or campo in _campos_fen:
            continue                      # entram dentro do grupo, mais abaixo
        _defs.append(cd)
    if _campos_perdas:
        _defs.append({"headerName": GRUPO_PERDAS, "headerClass": "grp-perdas",
                      "children": [dict(_por_campo[c], headerClass="grp-perdas") for c in _campos_perdas]})
    if _campos_fen:
        _defs.append({"headerName": GRUPO_FENOMENOS, "headerClass": "grp-fenomenos",
                      "children": [dict(_por_campo[c], headerClass="grp-fenomenos") for c in _campos_fen]})
    go["columnDefs"] = _defs

    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span": {"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-row":                         {"font-size": "13px !important"},
            ".ag-header-center .ag-header-cell-label": {"justify-content": "center !important"},
            # linha de grupo um tom mais escura, para separar as duas famílias
            ".ag-header-group-cell":           {"background-color": "#333333 !important",
                                                "border-left": "1px solid #6B6B6B !important"},
            ".ag-header-group-cell-label":     {"color": "#FFFFFF !important",
                                                "font-weight": "700 !important",
                                                "font-size": "12px !important",
                                                "justify-content": "center !important"},
            # cor por família: PERDAS em azul-petróleo, FENÔMENOS em vinho — o usuário bate o
            # olho e distingue as duas famílias sem ler o cabeçalho. Vale no grupo e nas colunas.
            ".ag-header-cell.grp-perdas":      {"background-color": "#1F5673 !important"},
            ".ag-header-group-cell.grp-perdas": {"background-color": "#1F5673 !important"},
            ".grp-perdas .ag-header-cell-text": {"color": "#FFFFFF !important"},
            ".grp-perdas .ag-header-group-cell-label": {"color": "#FFFFFF !important"},
            ".ag-header-cell.grp-fenomenos":   {"background-color": "#7A3B4E !important"},
            ".ag-header-group-cell.grp-fenomenos": {"background-color": "#7A3B4E !important"},
            ".grp-fenomenos .ag-header-cell-text": {"color": "#FFFFFF !important"},
            ".grp-fenomenos .ag-header-group-cell-label": {"color": "#FFFFFF !important"},
        },
        theme="streamlit", use_container_width=True,
    )


def tabela_perdas_html(df, mostrar_kg=True):
    """Tabela HTML colorida por status do material, no mesmo padrão visual da Apresentação da
    Conjunta (linha inteira na cor do material). Preserva a distinção perda × fenômeno com o
    cabeçalho de cada família colorido (perdas azul-petróleo, fenômenos vinho) e uma borda
    vertical grossa separando os dois blocos."""
    import streamlit.components.v1 as components

    campos_perdas = [rot for rot, _c in VARS_PERDAS if rot in df.columns]
    campos_qual = [rot for rot, _c in VARS_QUALIDADE if rot in df.columns]
    campos_fen = [rot for rot, _c in VARS_FENOMENOS if rot in df.columns]

    # colunas base: Híbrido, Status, Locais, [kg/ha], sc/ha, [Ardidos logo após sc/ha]
    campos_base = [c for c in ["Híbrido", "Status", "Locais"] if c in df.columns]
    if mostrar_kg and "kg/ha" in df.columns:
        campos_base.append("kg/ha")
    if "sc/ha" in df.columns:
        campos_base.append("sc/ha")
    # Ardidos (qualidade) vem logo depois de sc/ha, junto da produtividade
    campos_base += campos_qual
    cols_show = campos_base + campos_perdas + campos_fen

    COR_INT = {"Locais"}  # inteiros sem casa decimal
    primeiro_qual = campos_qual[0] if campos_qual else None   # borda antes de Ardidos
    primeiro_fen = campos_fen[0] if campos_fen else None      # borda antes dos fenômenos

    def _fmt(c, v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        if c in COR_INT:
            try:
                return f"{int(round(float(v), 0))}"
            except (ValueError, TypeError):
                return v
        return v

    html = """
<style>
.tb-perdas { width:100%; border-collapse:collapse; font-size:14px; font-family:'Helvetica Neue',sans-serif; }
.tb-perdas th { color:#FFF !important; padding:8px 10px; text-align:center; border:1px solid #ccc;
    white-space:nowrap; font-weight:700; font-size:13px; background:#4A4A4A; }
.tb-perdas th.grp-perdas { background:#1F5673 !important; }
.tb-perdas th.grp-qualidade { background:#5B4A2E !important; }
.tb-perdas th.grp-fenomenos { background:#7A3B4E !important; }
.tb-perdas th:first-child { text-align:left; }
.tb-perdas td { padding:7px 10px; border:1px solid #ddd; text-align:center; white-space:nowrap; font-size:14px; }
.tb-perdas td:first-child { text-align:left; font-weight:600; }
.tb-perdas td[data-fg="white"], .tb-perdas td[data-fg="white"] * { color:#FFF !important; }
.tb-perdas td[data-fg="dark"], .tb-perdas td[data-fg="dark"] * { color:#1A1A1A !important; }
.tb-perdas .sep-qual { border-left:3px solid #5B4A2E !important; }
.tb-perdas .sep-fen { border-left:3px solid #7A3B4E !important; }
</style>
<table class="tb-perdas"><thead><tr>"""
    for c in cols_show:
        cls = ""
        if c in campos_perdas:
            cls = "grp-perdas"
        elif c in campos_qual:
            cls = "grp-qualidade"
        elif c in campos_fen:
            cls = "grp-fenomenos"
        sep = ""
        if c == primeiro_qual:
            sep = " sep-qual"
        elif c == primeiro_fen:
            sep = " sep-fen"
        html += f'<th class="{cls}{sep}">{c}</th>'
    html += "</tr></thead><tbody>"

    for _, row in df.iterrows():
        status = row.get("Status", "")
        bg = COR_STATUS_PLOT.get(status, "#FFFFFF")
        fg = COR_TEXTO_STATUS.get(status, "#000000")
        data_fg = "white" if fg == "#FFFFFF" else "dark"
        html += "<tr>"
        for c in cols_show:
            sep = ""
            if c == primeiro_qual:
                sep = "sep-qual"
            elif c == primeiro_fen:
                sep = "sep-fen"
            html += f'<td class="{sep}" data-fg="{data_fg}" style="background:{bg};">{_fmt(c, row.get(c))}</td>'
        html += "</tr>"
    html += "</tbody></table>"

    altura = 60 + (len(df) + 2) * 40
    components.html(html, height=min(altura, 720), scrolling=True)


tabela_perdas_html(df_perdas, mostrar_kg=mostrar_kg_perdas)
exportar_excel(df_perdas, nome_arquivo="resumo_perdas_fenomenos.xlsx",
               label="⬇️ Exportar Perdas e Fenômenos", key="exp_perdas_sn")

_cap_pf = ("**Perdas**: média de todos os plots avaliados (o zero entra) — taxa média na rede. "
           "**Fenômenos**: média só onde ocorreu (o zero fica de fora, para não achatar); neles, "
           "**0** = avaliado e nunca ocorreu, **vazio** = não avaliado. As quatro perdas somam no "
           "Perda Total dentro de cada plot; fenômeno não entra nessa soma. Para o caso a caso, "
           "use a Auditoria; os cortes de alerta ficam no Diagnóstico.")
if _vars_vazias:
    _cap_pf += (" Sem medição neste recorte, fora da tabela: "
                + ", ".join(sorted(_vars_vazias)) + ".")
st.caption(_cap_pf)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — PERDAS POR LOCAL (gráfico de linhas, seletor de perda/fenômeno)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Perdas por Local",
    "Como cada híbrido perdeu em cada local?",
    contexto_str,
)


def _valor_por_familia(serie, eh_fenomeno):
    """Aplica a régua por família a um recorte de plots (mesma da tabela-resumo):
    fenômeno = média só onde ocorreu (>0); perda = média de todos (o zero entra)."""
    s = pd.to_numeric(serie, errors="coerce").dropna()
    if s.empty:
        return None
    if eh_fenomeno:
        com = s[s > 0]
        return round(float(com.mean()), 1) if not com.empty else 0.0
    return round(float(s.mean()), 1)


with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Escolha uma perda ou fenômeno. Cada linha é um híbrido; o eixo Y mostra a **% de perda** naquele
local. Aqui, ao contrário das notas de doença, **quanto mais baixa a linha, melhor** — menos perda.
Uma linha colada no chão é um híbrido que quase não perdeu naquele ambiente.

**A régua do valor por local** segue a mesma lógica da tabela-resumo:
- **Perdas** (acamamento, quebramento, dominadas, colmo podre, total) → média de **todos** os plots
  do híbrido no local, o zero entra. É a taxa média de perda na rede.
- **Fenômenos** (green snap, morte prematura, má formação, enfezamento) → média **só dos plots onde
  ocorreu**. Fenômeno é raro; incluir o zero achataria tudo e esconderia quem teve o problema.

**Faixa sombreada em vermelho** → o local onde algum híbrido teve o **pior caso** (a maior % de
perda), a linha que mais sobe. O rótulo mostra o valor e o híbrido.

Passe o mouse sobre a linha para ver híbrido, local, status e a % de perda.
""")

if not VARS_COLHEITA:
    st.info("Nenhuma variável de perda ou fenômeno disponível.")
else:
    _opts_local = [rot for rot, col in VARS_COLHEITA if col in ta_filtrado.columns
                   and pd.to_numeric(ta_filtrado[col], errors="coerce").notna().any()]
    if not _opts_local:
        st.info("Nenhuma perda ou fenômeno com dado nos filtros ativos.")
    else:
        perda_graf = st.selectbox("Selecione a perda ou fenômeno", options=_opts_local,
                                  key="perdas_graf_sel")
        col_perda_g = dict(VARS_COLHEITA)[perda_graf]
        eh_fenomeno_g = perda_graf in _ROTULOS_FENOMENO

        df_g = ta_filtrado[["dePara", "status_material", "cod_fazenda", "cidade_nome",
                            "estado_sigla", col_perda_g]].copy()
        df_g[col_perda_g] = pd.to_numeric(df_g[col_perda_g], errors="coerce")
        df_g = df_g.dropna(subset=[col_perda_g])
        # para fenômeno, só interessa quem tem ocorrência em algum plot do recorte
        if df_g.empty:
            st.info("Nenhuma avaliação disponível para esta variável nos filtros ativos.")
        else:
            # valor por híbrido × local, aplicando a régua por família
            reg_rows = []
            for (hib, st_mat, loc, cid, uf), grp in df_g.groupby(
                    ["dePara", "status_material", "cod_fazenda", "cidade_nome", "estado_sigla"]):
                val = _valor_por_familia(grp[col_perda_g], eh_fenomeno_g)
                if val is not None:
                    reg_rows.append({"dePara": hib, "status_material": st_mat, "cod_fazenda": loc,
                                     "cidade_nome": cid, "estado_sigla": uf, "valor": val})
            df_g_agg = pd.DataFrame(reg_rows)
            # fenômeno: remove híbrido×local que ficou 0 (não ocorreu) para não poluir com linhas no chão
            if eh_fenomeno_g:
                df_g_agg = df_g_agg[df_g_agg["valor"] > 0]

            if df_g_agg.empty:
                st.info(f"Nenhuma ocorrência de {perda_graf} registrada nos filtros ativos.")
            else:
                _n_hib_g = df_g_agg["dePara"].nunique()
                if _n_hib_g > 12:
                    st.info(
                        f"São **{_n_hib_g} híbridos** neste recorte — o gráfico de linhas fica mais "
                        "legível com menos materiais. Use o filtro de híbrido na barra lateral.")

                locais_ord = (df_g_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                              .drop_duplicates()
                              .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                              ["cod_fazenda"].tolist())
                hibridos_g = sorted(df_g_agg["dePara"].unique().tolist())
                palette = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                           "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
                           "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
                           "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5",
                           "#393B79", "#637939", "#8C6D31", "#843C39", "#7B4173"]
                cor_hib = {c: palette[i % len(palette)] for i, c in enumerate(hibridos_g)}

                fig_g = go_plt.Figure()
                for hibrido in hibridos_g:
                    df_c = df_g_agg[df_g_agg["dePara"] == hibrido].sort_values(
                        "cod_fazenda", key=lambda s: s.map(lambda x: locais_ord.index(x)))
                    if df_c.empty:
                        continue
                    status = df_c["status_material"].mode()
                    status = status.iloc[0] if not status.empty else ""
                    fig_g.add_trace(go_plt.Scatter(
                        x=df_c["cod_fazenda"], y=df_c["valor"], mode="lines+markers", name=hibrido,
                        line=dict(color=cor_hib[hibrido], width=2, shape="spline", smoothing=0.6),
                        marker=dict(size=6, color=cor_hib[hibrido]),
                        customdata=[[status]] * len(df_c),
                        hovertemplate=(f"<b>{hibrido}</b> · %{{x}}<br>Status: %{{customdata[0]}}<br>"
                                       f"{perda_graf}: %{{y:.1f}}%<extra></extra>")))

                # destaque do PIOR caso (maior % de perda individual) — a linha que mais sobe
                _idx_pior = df_g_agg["valor"].idxmax()
                local_crit = df_g_agg.loc[_idx_pior, "cod_fazenda"]
                val_crit = df_g_agg.loc[_idx_pior, "valor"]
                hib_crit = df_g_agg.loc[_idx_pior, "dePara"]
                cid_crit = df_g_agg.loc[_idx_pior, "cidade_nome"]
                if local_crit in locais_ord and val_crit > 0:
                    fig_g.add_shape(type="rect", xref="x", yref="paper",
                                    x0=locais_ord.index(local_crit) - 0.5,
                                    x1=locais_ord.index(local_crit) + 0.5,
                                    y0=0, y1=1, fillcolor="rgba(231,76,60,0.07)", line=dict(width=0))
                    fig_g.add_annotation(
                        x=local_crit, y=val_crit, yref="y",
                        text=f"pior caso · {val_crit:.1f}%<br>{cid_crit} · {hib_crit}",
                        showarrow=False, xanchor="center", yanchor="bottom", yshift=6,
                        font=dict(size=10, color="#E74C3C"), bgcolor="rgba(255,255,255,0.85)",
                        bordercolor="#E74C3C", borderwidth=1, borderpad=3)

                n_locais_g = len(locais_ord)
                altura_g = max(450, min(700, 350 + n_locais_g * 8))
                fig_g.update_layout(
                    height=altura_g,
                    yaxis=dict(title=dict(text=f"<b>{perda_graf} (menor = melhor)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               rangemode="tozero", tickfont=dict(size=12, color="#111111", weight="bold"),
                               gridcolor="#EEEEEE", zeroline=True, zerolinecolor="#CCCCCC"),
                    xaxis=dict(title=dict(text="<b>Local (código)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickangle=-45, tickfont=dict(size=11, color="#111111", weight="bold"),
                               categoryorder="array", categoryarray=locais_ord, gridcolor="#EEEEEE"),
                    legend=dict(title=dict(text="<b>Híbrido</b>", font=dict(size=12, color="#111111")),
                                font=dict(size=12, color="#111111", weight="bold"),
                                bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                    plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                    margin=dict(t=40, b=100, l=60, r=20),
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                    hovermode="closest")

                n_locais_total_g = ta_filtrado["cod_fazenda"].nunique()
                st.plotly_chart(fig_g, use_container_width=True)
                _regua_txt = ("média só dos plots onde ocorreu (fenômeno)" if eh_fenomeno_g
                              else "média de todos os plots do local, o zero entra (perda)")
                _cap_g = (f"Eixo Y = {perda_graf}, quanto menor melhor. Valor por local = {_regua_txt}. "
                          f"{len(hibridos_g)} híbridos · {n_locais_g} locais com registro")
                _cap_g += (f" (de {n_locais_total_g} ativos)." if n_locais_g < n_locais_total_g else ".")
                st.caption(_cap_g)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — DELTA VS REFERÊNCIA (quanto o híbrido perde a mais/menos que a testemunha)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Delta vs Referência",
    "Cada híbrido perde mais ou menos que o CHECK/STINE de referência?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Escolha uma **perda ou fenômeno** e um **híbrido de referência** (um CHECK ou STINE). Cada barra
mostra a diferença de perda entre um híbrido e essa referência **no mesmo local** — assim você
compara os materiais contra a testemunha, ambiente por ambiente, tirando o efeito do local.

**Atenção à direção — aqui é o contrário das doenças.** Na perda, o bom é perder **menos**:

- **Barra verde (delta negativo)** — o híbrido perdeu **menos** que a referência naquele local.
  É o resultado desejado: menos acamamento, menos quebra, menos perda.
- **Barra vermelha (delta positivo)** — o híbrido perdeu **mais** que a referência ali. Sinal de
  atenção.
- Só entram **locais onde os dois foram avaliados** — a comparação é sempre no mesmo ambiente.

**O cálculo**

```
delta = perda do híbrido no local − perda da referência no mesmo local
```

O valor por local segue a régua da família: **perda** usa a média de todos os plots (o zero entra);
**fenômeno e ardidos** usam a média só onde ocorreu.

**A tabela abaixo** resume: em quantos locais o híbrido perdeu menos, mais ou igual à referência, e
o delta médio. Delta médio **negativo** é bom (perde menos que a testemunha em média).
""")

_opts_delta = [rot for rot, col in VARS_COLHEITA if col in ta_filtrado.columns
               and pd.to_numeric(ta_filtrado[col], errors="coerce").notna().any()]

if not _opts_delta:
    st.info("Nenhuma perda ou fenômeno com dado nos filtros ativos.")
else:
    _c_d, _c_ref = st.columns(2)
    with _c_d:
        perda_delta = st.selectbox("Perda ou fenômeno", options=_opts_delta, key="perdas_delta_var")
    _refs_disp = sorted(
        ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist())
    if not _refs_disp:
        st.info("Nenhum híbrido CHECK ou STINE disponível nos filtros ativos para usar como referência.")
    else:
        with _c_ref:
            ref_delta = st.selectbox("Referência (CHECK / STINE)", options=_refs_disp, key="perdas_delta_ref")
        col_delta = dict(VARS_COLHEITA)[perda_delta]
        eh_fen_delta = perda_delta in _ROTULOS_FENOMENO

        df_d = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                            "cidade_nome", "estado_sigla", col_delta]].copy()
        df_d[col_delta] = pd.to_numeric(df_d[col_delta], errors="coerce")
        df_d = df_d.dropna(subset=[col_delta])

        # valor por híbrido × local com a régua da família
        reg_rows = []
        for (hib, st_mat, loc, cid, uf), grp in df_d.groupby(
                ["dePara", "status_material", "cod_fazenda", "cidade_nome", "estado_sigla"]):
            val = _valor_por_familia(grp[col_delta], eh_fen_delta)
            if val is not None:
                reg_rows.append({"dePara": hib, "status_material": st_mat, "cod_fazenda": loc,
                                 "cidade_nome": cid, "estado_sigla": uf, "valor": val})
        df_d_agg = pd.DataFrame(reg_rows)

        df_ref = df_d_agg[df_d_agg["dePara"] == ref_delta][["cod_fazenda", "valor"]].rename(
            columns={"valor": "_val_ref"})

        if df_ref.empty:
            st.info(f"**{ref_delta}** não tem registro de {perda_delta} nos filtros ativos.")
        else:
            df_d_agg = df_d_agg.merge(df_ref, on="cod_fazenda", how="inner")
            df_d_agg["delta"] = (df_d_agg["valor"] - df_d_agg["_val_ref"]).round(1)
            df_plot = df_d_agg[df_d_agg["dePara"] != ref_delta].copy()

            if df_plot.empty:
                st.info("Nenhum outro híbrido foi avaliado nos mesmos locais que a referência.")
            else:
                _n_hib_d = df_plot["dePara"].nunique()
                if _n_hib_d > 8:
                    st.info(
                        f"São **{_n_hib_d} híbridos** comparados — o gráfico de barras fica cheio. "
                        "Use o filtro de híbrido na barra lateral; a tabela resumo abaixo mostra "
                        "todos de forma compacta.")

                locais_d = (df_plot[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                            .drop_duplicates()
                            .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                            ["cod_fazenda"].tolist())
                # ordena híbridos pelo delta médio: menor (perde menos) primeiro = melhor
                ordem_hib = (df_plot.groupby("dePara")["delta"].mean()
                             .sort_values(ascending=True).index.tolist())

                fig_d = go_plt.Figure()
                for hibrido in ordem_hib:
                    df_c = df_plot[df_plot["dePara"] == hibrido].set_index("cod_fazenda")
                    dx, dy, hover = [], [], []
                    for local in locais_d:
                        if local in df_c.index:
                            dv = df_c.loc[local, "delta"]
                            dx.append(local); dy.append(dv)
                            hover.append(f"<b>{hibrido}</b> · {local}<br>"
                                         f"{perda_delta}: {df_c.loc[local, 'valor']:.1f}% · "
                                         f"Ref: {df_c.loc[local, '_val_ref']:.1f}%<br>Delta: {dv:+.1f}")
                    if not dx:
                        continue
                    # verde = delta NEGATIVO (perde menos = bom); vermelho = positivo (perde mais)
                    cores = ["#1E7A34" if v <= 0 else "#E63946" for v in dy]
                    fig_d.add_trace(go_plt.Bar(
                        name=hibrido, x=dx, y=dy, marker_color=cores,
                        text=[f"{v:+.1f}" for v in dy], textposition="outside",
                        textfont=dict(size=10, color="#333333"),
                        hovertext=hover, hoverinfo="text", offsetgroup=hibrido))

                fig_d.add_hline(y=0, line_color="#333333", line_width=1.5)
                # faixa: abaixo de zero é bom (verde), acima é ruim (vermelho) — inverso da Sanidade
                fig_d.add_hrect(y0=-100, y1=0, fillcolor="rgba(30,122,52,0.04)", line_width=0)
                fig_d.add_hrect(y0=0, y1=100, fillcolor="rgba(230,57,70,0.04)", line_width=0)

                altura_d = max(420, min(680, 350 + len(locais_d) * 12))
                fig_d.update_layout(
                    height=altura_d,
                    title=dict(text=f"Delta de {perda_delta} vs <b>{ref_delta}</b> "
                                    f"(negativo = perde menos, melhor)",
                               font=dict(size=13, color="#111111"), x=0, xanchor="left"),
                    barmode="group", bargap=0.15, bargroupgap=0.05,
                    xaxis=dict(title=dict(text="<b>Local (código)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickangle=-45, tickfont=dict(size=11, color="#111111", weight="bold"),
                               categoryorder="array", categoryarray=locais_d, gridcolor="#EEEEEE"),
                    yaxis=dict(title=dict(text="<b>Delta (p.p. de perda)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickfont=dict(size=12, color="#111111", weight="bold"),
                               gridcolor="#EEEEEE", zeroline=False),
                    legend=dict(title=dict(text="<b>Híbrido</b>",
                                           font=dict(size=12, color="#111111", weight="bold")),
                                font=dict(size=11, color="#111111", weight="bold"),
                                bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                    plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                    margin=dict(t=50, b=120, l=60, r=20),
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                    hovermode="x unified")

                st.plotly_chart(fig_d, use_container_width=True)
                _regua_d = ("média só onde ocorreu" if eh_fen_delta else "média de todos os plots (com zero)")
                st.caption(f"Delta = {perda_delta} do híbrido − de {ref_delta} no mesmo local "
                           f"({_regua_d}). Verde = perde menos que a referência · vermelho = perde mais. "
                           f"{len(ordem_hib)} híbridos · {len(locais_d)} locais em comum.")

                # tabela resumo
                res_rows = []
                for hibrido in ordem_hib:
                    df_c = df_plot[df_plot["dePara"] == hibrido]
                    status = df_c["status_material"].mode()
                    status = status.iloc[0] if not status.empty else ""
                    res_rows.append({
                        "Híbrido": hibrido, "Status": status, "Locais": len(df_c),
                        "▼ Perde menos": int((df_c["delta"] < 0).sum()),
                        "▲ Perde mais": int((df_c["delta"] > 0).sum()),
                        "= Igual": int((df_c["delta"] == 0).sum()),
                        "Delta médio": round(df_c["delta"].mean(), 2)})
                df_res = pd.DataFrame(res_rows)

                gb_d = GridOptionsBuilder.from_dataframe(df_res)
                gb_d.configure_default_column(
                    resizable=True, sortable=True, filter=True,
                    cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
                gb_d.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
                gb_d.configure_column("Híbrido", pinned="left", width=170)
                gb_d.configure_column("Status", width=90)
                gb_d.configure_column("▼ Perde menos", width=120,
                    cellStyle=JsCode("function(p){if(p.value>0)return{background:'#D5F5D5',color:'#1A1A1A',fontWeight:'700'};return{};}"))
                gb_d.configure_column("▲ Perde mais", width=120,
                    cellStyle=JsCode("function(p){if(p.value>0)return{background:'#FDDCDE',color:'#1A1A1A',fontWeight:'700'};return{};}"))
                gb_d.configure_column("= Igual", width=90)
                # delta médio: NEGATIVO é bom (verde), positivo é ruim (vermelho) — inverso da nota
                gb_d.configure_column("Delta médio", width=120,
                    cellStyle=JsCode("function(p){var v=p.value; if(v<0)return{background:'#1E7A34',color:'#FFFFFF',fontWeight:'700'}; if(v>0)return{background:'#E63946',color:'#FFFFFF',fontWeight:'700'}; return{background:'#FFD600',color:'#1A1A1A'};}"))

                go_d = gb_d.build()
                go_d["defaultColDef"]["headerClass"] = "ag-header-black"
                AgGrid(df_res, gridOptions=go_d, height=min(500, 36 + 32 * len(df_res) + 20),
                       update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
                       allow_unsafe_jscode=True, enable_enterprise_modules=True,
                       custom_css={
                           ".ag-header":            {"background-color": "#4A4A4A !important"},
                           ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                           ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                           ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                           ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                           ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                           ".ag-row":               {"font-size": "13px !important"},
                       },
                       theme="streamlit", use_container_width=True)
                st.caption(f"**▼ Perde menos** = nº de locais em que o híbrido perdeu menos que "
                           f"{ref_delta}. **Delta médio negativo é bom** (perde menos na média).")
                exportar_excel(df_res, nome_arquivo="delta_perdas.xlsx",
                               label="⬇️ Exportar Delta vs Referência", key="exp_delta_perdas")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5 — ANÁLISE DE SOBREVIVÊNCIA + MAPA DE COLAPSO (invertida para perda)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Análise de Sobrevivência",
    "Qual híbrido mantém a perda baixa na maioria dos locais?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada curva mostra, para um nível de perda (eixo X), **em quantos % dos locais o híbrido teve perda
igual ou maior que aquele valor**. Todas começam em 100% (no limite 0%, todo local tem perda ≥ 0) e
**caem** conforme a perda aumenta. Um híbrido que perde pouco **despenca cedo** — em poucos locais a
perda dele chega a valores altos. Um que perde muito **se sustenta no alto** — a perda é alta em
muitos locais.

**Como ler:** curva que **cai rápido = híbrido bom** (a perda raramente chega alto). Curva que fica
alta por mais tempo = perde muito, em mais locais.

**Eixos**
- **X — Perda (%)** → de 0 até a maior perda observada.
- **Y — % de locais com perda ≥ aquele valor** → começa em 100% e cai; queda precoce é melhor.

**Destaques**
- **▲ melhor (perde pouco)** → curva que cai mais cedo, com a **menor** área sob a curva (ASC).
- **▼ pior (perde muito)** → curva que se sustenta, com a **maior** ASC.

**Tabela ASC** → resume num número. Aqui, **ASC baixa é bom** — quer dizer que a curva caiu cedo, ou
seja, a perda raramente atingiu valores altos.

**Mapa de Colapso** (abaixo) → em quantos locais cada híbrido teve perda **acima do limite crítico**
que você define no slider. Ajuste conforme o tipo de perda — 20% pode ser crítico para perda total,
mas pouco para dominadas.

O valor por local segue a régua da família: perda com o zero, fenômeno e ardidos só onde ocorreu.
""")

_opts_surv = [rot for rot, col in VARS_COLHEITA if col in ta_filtrado.columns
              and pd.to_numeric(ta_filtrado[col], errors="coerce").gt(0).any()]

if not _opts_surv:
    st.info("Nenhuma perda ou fenômeno com ocorrência nos filtros ativos.")
else:
    perda_surv = st.selectbox("Selecione a perda ou fenômeno", options=_opts_surv, key="perdas_surv_var")
    col_surv = dict(VARS_COLHEITA)[perda_surv]
    eh_fen_surv = perda_surv in _ROTULOS_FENOMENO

    df_s = ta_filtrado[["dePara", "status_material", "cod_fazenda", col_surv]].copy()
    df_s[col_surv] = pd.to_numeric(df_s[col_surv], errors="coerce")
    df_s = df_s.dropna(subset=[col_surv])

    # valor por híbrido × local com a régua da família
    reg_rows = []
    for (hib, st_mat, loc), grp in df_s.groupby(["dePara", "status_material", "cod_fazenda"]):
        val = _valor_por_familia(grp[col_surv], eh_fen_surv)
        if val is not None:
            reg_rows.append({"dePara": hib, "status_material": st_mat, "cod_fazenda": loc, "valor": val})
    df_s_agg = pd.DataFrame(reg_rows)
    # fenômeno: só locais com ocorrência entram na curva
    if eh_fen_surv:
        df_s_agg = df_s_agg[df_s_agg["valor"] > 0]

    hibridos_sv = sorted(df_s_agg["dePara"].unique().tolist()) if not df_s_agg.empty else []

    if df_s_agg.empty or not hibridos_sv:
        st.info(f"Nenhuma ocorrência de {perda_surv} nos filtros ativos.")
    else:
        perda_max = float(df_s_agg["valor"].max())
        # eixo X até o p95, não o máximo: um outlier estica o eixo e esmaga a curva contra a
        # esquerda. O p95 deixa a curva usar o espaço todo e cair suave. O cálculo usa TODOS os
        # valores; só a visualização é cortada (a cauda além do p95 já rasteja no chão).
        _p95 = float(np.percentile(df_s_agg["valor"], 95))
        eixo_max = max(5.0, np.ceil(_p95))
        limite_max = max(5.0, np.ceil(perda_max))       # thresholds calculam a curva inteira
        thresholds = list(np.linspace(0, limite_max, 120))

        # slider do limite crítico ANTES da curva: controla a linha vertical no gráfico E o mapa
        # de colapso (abaixo). Um controle só para os dois — mexer no slider move a linha e
        # recalcula o colapso juntos.
        _lim_default = float(min(eixo_max, max(5.0, round(_p95 * 0.5, 0)))) or 10.0
        limite_colapso = st.slider(
            "Limite de perda crítica (%)", min_value=1.0,
            max_value=float(eixo_max),
            value=_lim_default, step=1.0, key="perdas_lim_colapso",
            help="A linha vertical na curva marca este limite. Um local com perda acima dele conta "
                 "como colapso na tabela abaixo. Ajuste conforme o tipo de perda.")

        palette_sv = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                      "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
                      "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
                      "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5",
                      "#393B79", "#637939", "#8C6D31", "#843C39", "#7B4173"]
        cor_sv = {c: palette_sv[i % len(palette_sv)] for i, c in enumerate(hibridos_sv)}

        curvas, asc_vals = {}, {}
        for hibrido in hibridos_sv:
            vals = df_s_agg[df_s_agg["dePara"] == hibrido]["valor"].tolist()
            n = len(vals)
            # % de locais com perda <= t (mantém a perda abaixo do limite)
            # % de locais com perda >= t. Começa em 100% (limite 0) e CAI conforme a perda
            # aumenta: híbrido que perde pouco despenca cedo (poucos locais com perda alta).
            sobrev = [sum(1 for v in vals if v >= t) / n * 100 for t in thresholds]
            curvas[hibrido] = sobrev
            asc_vals[hibrido] = float(np.trapezoid(sobrev, thresholds))

        # ASC BAIXA é boa aqui: curva que cai cedo = perde pouco. Ordena do menor (melhor) ao maior.
        df_asc = (pd.DataFrame.from_dict(asc_vals, orient="index", columns=["Área sob a curva"])
                  .reset_index().rename(columns={"index": "Híbrido"})
                  .sort_values("Área sob a curva", ascending=True).reset_index(drop=True))
        df_asc["Área sob a curva"] = df_asc["Área sob a curva"].round(1)

        if len(hibridos_sv) > 12:
            st.info(f"São **{len(hibridos_sv)} híbridos** — muitas curvas juntas. Use o filtro de "
                    "híbrido na barra lateral; a tabela ASC e o Mapa de Colapso cobrem todos.")

        fig_s = go_plt.Figure()
        for hibrido in hibridos_sv:
            status = df_s_agg[df_s_agg["dePara"] == hibrido]["status_material"].mode()
            status = status.iloc[0] if not status.empty else ""
            fig_s.add_trace(go_plt.Scatter(
                x=thresholds, y=curvas[hibrido], mode="lines", name=hibrido,
                line=dict(color=cor_sv[hibrido], width=2, shape="spline", smoothing=0.5),
                customdata=[[status, round(asc_vals[hibrido], 1)]] * len(thresholds),
                hovertemplate=(f"<b>{hibrido}</b><br>Status: %{{customdata[0]}}<br>"
                               f"Perda ≥ %{{x:.1f}}%: %{{y:.1f}}% dos locais<br>"
                               f"ASC: %{{customdata[1]}}<extra></extra>")))

        # linha vertical do limite crítico (segue o slider) — onde cada curva cruza esta linha
        # é quantos % dos locais estão em colapso para aquele híbrido
        fig_s.add_vline(x=limite_colapso, line_dash="dash", line_color="#111111", line_width=1.5,
                        annotation_text=f"limite {limite_colapso:.0f}%",
                        annotation_position="top",
                        annotation_font=dict(size=12, color="#111111", weight="bold"))

        melhor_sv = df_asc.iloc[0]["Híbrido"]   # menor ASC = cai cedo = perde pouco
        pior_sv = df_asc.iloc[-1]["Híbrido"]     # maior ASC = se sustenta = perde muito
        # em vez de caixas flutuantes no gráfico, a info de melhor/pior vai no TÍTULO, com o
        # nome de cada híbrido na cor da sua curva — mais limpo, sem poluir a área de plotagem.
        _cor_melhor = "#1E7A34"   # verde = melhor
        _cor_pior = "#C0201E"     # vermelho = pior
        titulo_curva = (
            f"<b>Sobrevivência — {perda_surv}</b><br>"
            f"<span style='font-size:15px;color:{_cor_melhor}'><b>{melhor_sv}</b></span>"
            f"<span style='font-size:11px;color:{_cor_melhor}'> (melhor · perde pouco)</span>"
            f"<span style='font-size:12px;color:#888888'>    ·    </span>"
            f"<span style='font-size:15px;color:{_cor_pior}'><b>{pior_sv}</b></span>"
            f"<span style='font-size:11px;color:{_cor_pior}'> (pior · perde muito)</span>")

        top10 = df_asc.head(10)
        # gradiente verde (melhor, ASC baixa) -> vermelho (pior, ASC alta) pela posição na lista,
        # que já está ordenada por ASC crescente. A cor carrega a ordem (Schwabish).
        def _grad(frac):
            # frac 0 = verde, 1 = vermelho, passando por amarelo no meio
            if frac <= 0.5:
                t = frac / 0.5
                r, g, b = int(0xD5 + (0xFF - 0xD5) * t), int(0xF5 + (0xF5 - 0xF5) * t), int(0xD5 + (0xC8 - 0xD5) * t)
            else:
                t = (frac - 0.5) / 0.5
                r, g, b = int(0xFF), int(0xF5 - (0xF5 - 0xC0) * t), int(0xC8 - (0xC8 - 0xC0) * t)
            return f"rgb({r},{g},{b})"

        _n = len(top10)
        _fill = [_grad(i / max(_n - 1, 1)) for i in range(_n)]
        # nomes com rótulo nas pontas: (melhor) no primeiro, (pior) no último
        _nomes_asc = top10["Híbrido"].tolist()
        if _nomes_asc:
            _nomes_asc[0] = f"{_nomes_asc[0]}  (melhor)"
            if _n > 1:
                _nomes_asc[-1] = f"{_nomes_asc[-1]}  (pior)"
        fig_s.add_trace(go_plt.Table(
            domain=dict(x=[0.63, 0.90], y=[0.40, 0.90]), columnwidth=[110, 55],
            header=dict(values=["<b>Híbrido</b>", "<b>ASC</b>"], fill_color="#4A4A4A",
                        font=dict(color="white", size=11), align="center", height=26),
            cells=dict(values=[_nomes_asc, top10["Área sob a curva"].tolist()],
                       fill_color=[_fill, _fill], font=dict(color="#111111", size=11),
                       align=["left", "center"], height=22)))

        fig_s.update_layout(
            height=540,
            title=dict(text=titulo_curva, font=dict(size=15, color="#111111"),
                       x=0, xanchor="left", y=0.98, yanchor="top"),
            xaxis=dict(title=dict(text="<b>Perda (%)</b>",
                                  font=dict(size=14, color="#111111", weight="bold")),
                       tickfont=dict(size=12, color="#111111", weight="bold"),
                       range=[-eixo_max * 0.02, eixo_max], autorange=False,
                       gridcolor="#CCCCCC", griddash="dot", gridwidth=1, domain=[0, 0.60]),
            yaxis=dict(title=dict(text="<b>% de locais com perda ≥ limite</b>",
                                  font=dict(size=14, color="#111111", weight="bold")),
                       tickformat=".0f", ticksuffix="%", range=[-5, 105], autorange=False,
                       tickfont=dict(size=12, color="#111111", weight="bold"),
                       gridcolor="#CCCCCC", griddash="dot", gridwidth=1, zeroline=False),
            legend=dict(title=dict(text="<b>Híbrido</b>", font=dict(size=12, color="#111111", weight="bold")),
                        font=dict(size=11, color="#111111", weight="bold"), x=1.02, y=1,
                        xanchor="left", bgcolor="rgba(255,255,255,0.85)",
                        bordercolor="#DDDDDD", borderwidth=1),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            margin=dict(t=90, b=60, l=60, r=160),
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
            hovermode="x unified")

        n_locais_sv = df_s_agg["cod_fazenda"].nunique()
        st.plotly_chart(fig_s, use_container_width=True)
        st.caption(f"{perda_surv} · {len(hibridos_sv)} híbridos · {n_locais_sv} locais. "
                   f"Curva que cai cedo = perde pouco (bom). ASC baixa é melhor. "
                   f"Eixo cortado em {eixo_max:.0f}% (p95) para a curva não ficar espremida por "
                   f"outliers; a cauda além disso são poucos locais e já rasteja no chão.")

        st.divider()
        st.markdown(f"#### Mapa de Colapso — locais com perda acima de {limite_colapso:.0f}%")
        st.caption("O limite é o mesmo do slider acima (a linha vertical na curva). "
                   "Ajuste lá para recalcular.")

        colapso_rows = []
        for hibrido in hibridos_sv:
            df_c = df_s_agg[df_s_agg["dePara"] == hibrido]
            status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
            n_total = len(df_c)
            n_colapso = int((df_c["valor"] > limite_colapso).sum())
            n_ok = int((df_c["valor"] <= limite_colapso).sum())
            locais_col = df_c[df_c["valor"] > limite_colapso]["cod_fazenda"].tolist()
            colapso_rows.append({
                "Híbrido": hibrido, "Status": status, "Locais Aval.": n_total,
                f"Colapso (>{limite_colapso:.0f}%)": n_colapso,
                f"OK (≤{limite_colapso:.0f}%)": n_ok,
                "% Colapso": round(n_colapso / n_total * 100, 1) if n_total else 0,
                "Locais em Colapso": ", ".join(locais_col) if locais_col else "—"})

        df_colapso = (pd.DataFrame(colapso_rows)
                      .sort_values(["% Colapso", f"Colapso (>{limite_colapso:.0f}%)"],
                                   ascending=[True, True]).reset_index(drop=True))

        _col_colapso = f"Colapso (>{limite_colapso:.0f}%)"
        _col_ok = f"OK (≤{limite_colapso:.0f}%)"
        gb_c = GridOptionsBuilder.from_dataframe(df_colapso)
        gb_c.configure_default_column(
            resizable=True, sortable=True, filter=True,
            cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
        gb_c.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
        gb_c.configure_column("Híbrido", pinned="left", width=170)
        gb_c.configure_column("Status", width=90)
        gb_c.configure_column("Locais Aval.", width=100)
        gb_c.configure_column(_col_colapso, width=130,
            cellStyle=JsCode("function(p){if(p.value>0)return{background:'#E63946',color:'#FFFFFF',fontWeight:'700'};return{background:'#D5F5D5',color:'#1A1A1A'};}"))
        gb_c.configure_column(_col_ok, width=120,
            cellStyle=JsCode("function(p){if(p.value>0)return{background:'#70C96E',color:'#1A1A1A'};return{};}"))
        gb_c.configure_column("% Colapso", width=110,
            cellStyle=JsCode("function(p){var v=p.value; if(v>=50)return{background:'#8B0000',color:'#FFFFFF',fontWeight:'700'}; if(v>=20)return{background:'#E63946',color:'#FFFFFF',fontWeight:'700'}; if(v>0)return{background:'#FFD600',color:'#1A1A1A'}; return{background:'#D5F5D5',color:'#1A1A1A'};}"))
        gb_c.configure_column("Locais em Colapso", width=300)

        go_c = gb_c.build()
        go_c["defaultColDef"]["headerClass"] = "ag-header-black"
        AgGrid(df_colapso, gridOptions=go_c, height=min(560, 36 + 32 * len(df_colapso) + 20),
               update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
               allow_unsafe_jscode=True, enable_enterprise_modules=True,
               custom_css={
                   ".ag-header":            {"background-color": "#4A4A4A !important"},
                   ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                   ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                   ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                   ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                   ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                   ".ag-row":               {"font-size": "13px !important"},
               },
               theme="streamlit", use_container_width=True)
        st.caption(f"**Colapso** = local com {perda_surv} acima de {limite_colapso:.0f}%. "
                   f"Ordenado do menor para o maior % de colapso (melhores primeiro).")
        exportar_excel(df_colapso, nome_arquivo="colapso_perdas.xlsx",
                       label="⬇️ Exportar Mapa de Colapso", key="exp_colapso_perdas")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — HEATMAP HÍBRIDO × LOCAL (perda, cor por gradiente fixo)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Heatmap por Local",
    "Onde cada híbrido perdeu mais?",
    contexto_str,
)


def _cor_perda(v, teto=25.0):
    """Gradiente fixo: 0% verde -> meio amarelo -> teto (25%) vermelho. Acima do teto satura no
    vermelho. Escala igual para toda perda, então a cor significa o mesmo em qualquer célula."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "#E0E0E0", "#888888"
    frac = min(max(v / teto, 0.0), 1.0)
    if frac <= 0.5:
        t = frac / 0.5
        r, g, b = int(0x1E + (0xFF - 0x1E) * t), int(0x7A + (0xD6 - 0x7A) * t), int(0x34 + (0x00 - 0x34) * t)
    else:
        t = (frac - 0.5) / 0.5
        # amarelo (255,214,0) -> vermelho pleno (200,32,30)
        r, g, b = int(0xFF - (0xFF - 0xC8) * t), int(0xD6 - (0xD6 - 0x20) * t), int(0x00 + (0x1E) * t)
    fg = "#FFFFFF" if frac > 0.55 or frac < 0.12 else "#1A1A1A"
    return f"rgb({r},{g},{b})", fg


with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Para a perda ou fenômeno escolhido, cada célula mostra a **% de perda** do híbrido naquele local. A
cor vai de **verde (perda baixa)** a **vermelho (perda alta)**, numa escala fixa — 25% ou mais já é
vermelho pleno. Como a escala é a mesma para toda perda, a cor significa o mesmo em qualquer célula:
mais vermelho = perdeu mais, em qualquer coluna.

O valor por célula segue a régua da família: perda com o zero; fenômeno e ardidos só onde ocorreu.

**Destaques automáticos**
- **Borda vermelha + ⚠️** → local de maior perda (a perda média do grupo ali é a mais alta).
- **★ verde** → híbrido com o melhor desempenho geral (menor perda média entre os locais).

Célula cinza (—) = híbrido não avaliado naquele local. A linha preta separa grupos de status.
""")

with st.popover(f"📍 Dicionário de locais ({ta_filtrado['cod_fazenda'].nunique()})",
                use_container_width=False):
    _df_dic_hp = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                  .drop_duplicates()
                  .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                  .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                   "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                  .reset_index(drop=True))
    st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
    st.dataframe(_df_dic_hp, hide_index=True, use_container_width=True)

_opts_hp = [rot for rot, col in VARS_COLHEITA if col in ta_filtrado.columns
            and pd.to_numeric(ta_filtrado[col], errors="coerce").gt(0).any()]

if not _opts_hp:
    st.info("Nenhuma perda ou fenômeno com ocorrência nos filtros ativos.")
else:
    perda_hp = st.selectbox("Selecione a perda ou fenômeno", options=_opts_hp, key="perdas_hp_var")
    col_hp = dict(VARS_COLHEITA)[perda_hp]
    eh_fen_hp = perda_hp in _ROTULOS_FENOMENO

    df_hp = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                         "cidade_nome", "estado_sigla", col_hp]].copy()
    df_hp[col_hp] = pd.to_numeric(df_hp[col_hp], errors="coerce")
    df_hp = df_hp.dropna(subset=[col_hp])

    # valor por híbrido × local com a régua da família
    reg_rows = []
    for (hib, st_mat, loc, cid, uf), grp in df_hp.groupby(
            ["dePara", "status_material", "cod_fazenda", "cidade_nome", "estado_sigla"]):
        val = _valor_por_familia(grp[col_hp], eh_fen_hp)
        if val is not None:
            reg_rows.append({"dePara": hib, "status_material": st_mat, "cod_fazenda": loc,
                             "cidade_nome": cid, "estado_sigla": uf, "valor": val})
    df_hp_agg = pd.DataFrame(reg_rows)

    if df_hp_agg.empty:
        st.info(f"Nenhum registro de {perda_hp} nos filtros ativos.")
    else:
        locais_hp = (df_hp_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                     .drop_duplicates()
                     .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                     ["cod_fazenda"].tolist())
        cult_status_hp = (df_hp_agg[["dePara", "status_material"]].drop_duplicates()
                          .assign(_ord=lambda d: d["status_material"].apply(
                              lambda s: ORDEM_STATUS.index(s) if s in ORDEM_STATUS else 99))
                          .sort_values(["_ord", "dePara"]))
        hibridos_hp = cult_status_hp["dePara"].tolist()
        status_map_hp = cult_status_hp.set_index("dePara")["status_material"].to_dict()

        pivot_hp = df_hp_agg.pivot_table(index="dePara", columns="cod_fazenda",
                                         values="valor", aggfunc="mean").reindex(
            index=hibridos_hp, columns=locais_hp)

        # matriz de valores (z) — fundo colorido pelo go.Heatmap NATIVO (leve, não trava).
        # Os NÚMEROS vêm de annotations, com cor decidida por luminância (branco no escuro, escuro
        # no claro) — mesma ideia do data-fg da Conjunta, mas para gradiente contínuo.
        TETO_HP = 25.0
        z = [[(None if (v is None or (isinstance(v, float) and np.isnan(v)))
               else min(float(v), TETO_HP)) for v in pivot_hp.loc[h].tolist()]
             for h in hibridos_hp]

        # colorscale verde escuro (perda baixa) -> amarelo -> vermelho (perda alta). Verde escuro
        # de volta como você preferiu; a legibilidade vem da cor do texto, não de clarear o fundo.
        colorscale_hp = [
            [0.00, "#1E7A34"], [0.22, "#5F9E3A"], [0.45, "#B7D64C"],
            [0.60, "#F4D03F"], [0.78, "#E8850D"], [1.00, "#C0201E"]]

        def _cor_fundo_hp(frac):
            """Interpola a colorscale para a fração dada (0-1) e devolve (r,g,b)."""
            stops = [(0.00, (0x1E, 0x7A, 0x34)), (0.22, (0x5F, 0x9E, 0x3A)),
                     (0.45, (0xB7, 0xD6, 0x4C)), (0.60, (0xF4, 0xD0, 0x3F)),
                     (0.78, (0xE8, 0x85, 0x0D)), (1.00, (0xC0, 0x20, 0x1E))]
            frac = min(max(frac, 0.0), 1.0)
            for k in range(len(stops) - 1):
                f0, c0 = stops[k]
                f1, c1 = stops[k + 1]
                if f0 <= frac <= f1:
                    t = (frac - f0) / (f1 - f0) if f1 > f0 else 0
                    return tuple(int(c0[i] + (c1[i] - c0[i]) * t) for i in range(3))
            return stops[-1][1]

        def _texto_para_fundo(frac):
            """Branco se o fundo for escuro, quase-preto se claro (contraste)."""
            r, g, b = _cor_fundo_hp(frac)
            lum = (0.299 * r + 0.587 * g + 0.114 * b) / 255
            return lum < 0.55  # True = fundo escuro (texto branco)

        # duas matrizes de texto: números escuros nas células claras, brancos nas escuras.
        # cada uma vai num heatmap com texttemplate (renderiza tudo numa passada, SEM annotation
        # por célula — que era o que travava com muitas células).
        txt_escuro, txt_branco = [], []
        for h in hibridos_hp:
            lin_e, lin_b = [], []
            for loc in locais_hp:
                v = pivot_hp.loc[h, loc] if (h in pivot_hp.index and loc in pivot_hp.columns) else None
                if v is None or (isinstance(v, float) and np.isnan(v)):
                    lin_e.append(""); lin_b.append("")
                else:
                    frac = min(float(v), TETO_HP) / TETO_HP
                    s = f"{v:.1f}"
                    if _texto_para_fundo(frac):   # fundo escuro -> texto branco
                        lin_b.append(s); lin_e.append("")
                    else:                          # fundo claro -> texto escuro
                        lin_e.append(s); lin_b.append("")
            txt_escuro.append(lin_e); txt_branco.append(lin_b)

        fig_hp = go_plt.Figure()
        # trace 1: fundo colorido + números escuros (nas células claras)
        fig_hp.add_trace(go_plt.Heatmap(
            z=z, x=locais_hp, y=hibridos_hp, text=txt_escuro, texttemplate="%{text}",
            textfont=dict(size=10, color="#1A1A1A"),
            zmin=0, zmax=TETO_HP, colorscale=colorscale_hp,
            xgap=1, ygap=1, hoverongaps=False,
            colorbar=dict(title=dict(text="% perda", side="right"), thickness=12, len=0.6),
            hovertemplate="%{y} · %{x}<br>" + perda_hp + ": %{z:.1f}%<extra></extra>"))
        # trace 2: transparente + números brancos (nas células escuras) — sobreposto
        fig_hp.add_trace(go_plt.Heatmap(
            z=z, x=locais_hp, y=hibridos_hp, text=txt_branco, texttemplate="%{text}",
            textfont=dict(size=10, color="#FFFFFF"),
            zmin=0, zmax=TETO_HP, colorscale=[[0, "rgba(0,0,0,0)"], [1, "rgba(0,0,0,0)"]],
            showscale=False, hoverinfo="skip"))

        # separadores de status: linha entre grupos. Com eixo categórico, usa a posição via paper
        # não é trivial; usamos as categorias com deslocamento em coordenada de dados do heatmap.
        for i in range(len(hibridos_hp) - 1):
            if status_map_hp.get(hibridos_hp[i], "") != status_map_hp.get(hibridos_hp[i + 1], ""):
                fig_hp.add_shape(type="line", xref="paper", x0=0, x1=1,
                                 y0=i + 0.5, y1=i + 0.5, line=dict(color="#333333", width=2))

        # destaque do local de MAIOR perda média (o pior local) — coluna inteira
        media_local_hp = df_hp_agg.groupby("cod_fazenda")["valor"].mean()
        if not media_local_hp.empty and len(locais_hp) > 0:
            local_crit_hp = media_local_hp.idxmax()
            perda_crit_hp = round(media_local_hp.max(), 1)
            if local_crit_hp in locais_hp:
                j_crit = locais_hp.index(local_crit_hp)
                fig_hp.add_shape(type="rect", xref="x", yref="paper",
                                 x0=j_crit - 0.5, x1=j_crit + 0.5, y0=0, y1=1,
                                 fillcolor="rgba(0,0,0,0)", line=dict(color="#E74C3C", width=2.5))
                fig_hp.add_annotation(x=local_crit_hp, xref="x", y=1.0, yref="paper",
                                      text=f"⚠️ {perda_crit_hp:.0f}%",
                                      showarrow=False, xanchor="center", yanchor="bottom", yshift=4,
                                      font=dict(size=10, color="#E74C3C"),
                                      bgcolor="rgba(255,255,255,0.9)", bordercolor="#E74C3C",
                                      borderwidth=1, borderpad=2)

        # híbrido de MENOR perda média (o melhor) — marca ★ na linha dele
        media_cult_hp = df_hp_agg.groupby("dePara")["valor"].mean()
        if not media_cult_hp.empty and len(hibridos_hp) > 0:
            melhor_hp = media_cult_hp.idxmin()
            perda_melhor_hp = round(media_cult_hp.min(), 1)
            if melhor_hp in hibridos_hp:
                fig_hp.add_annotation(x=1.0, xref="paper", y=melhor_hp, yref="y",
                                      text=f"★ menor perda ({perda_melhor_hp:.1f}%)", showarrow=False,
                                      xanchor="left", yanchor="middle", xshift=8,
                                      font=dict(size=11, color="#1E7A34", weight="bold"),
                                      bgcolor="rgba(255,255,255,0.85)", bordercolor="#1E7A34",
                                      borderwidth=1, borderpad=2)

        # rótulos dos híbridos coloridos por status (à esquerda), via categoria
        for hibrido in hibridos_hp:
            cor_s = COR_STATUS_PLOT.get(status_map_hp.get(hibrido, ""), "#333333")
            if status_map_hp.get(hibrido, "") == "EXP":
                cor_s = "#1A7A1A"
            fig_hp.add_annotation(x=-0.01, xref="paper", y=hibrido, yref="y",
                                  text=f"<b>{hibrido}</b>", showarrow=False,
                                  xanchor="right", yanchor="middle",
                                  font=dict(size=11, color=cor_s, weight="bold"))

        altura_hp = max(400, len(hibridos_hp) * 28 + 90)
        fig_hp.update_layout(
            height=altura_hp,
            xaxis=dict(side="bottom", tickfont=dict(size=11, color="#111111", weight="bold"),
                       tickangle=-45, title=dict(text="<b>Local (cod_fazenda)</b>",
                                                 font=dict(size=14, color="#111111", weight="bold")),
                       categoryorder="array", categoryarray=locais_hp),
            yaxis=dict(tickfont=dict(size=11, color="#111111", weight="bold"),
                       autorange="reversed", showticklabels=False),
            margin=dict(t=30, b=100, l=200, r=160),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"))

        st.plotly_chart(fig_hp, use_container_width=True)
        n_locais_total_hp = ta_filtrado["cod_fazenda"].nunique()
        _regua_hp = ("média só onde ocorreu" if eh_fen_hp else "média de todos os plots (com zero)")
        _cap_hp = (f"Cor da célula = % de {perda_hp} ({_regua_hp}), verde (baixa) → vermelho (alta, "
                   f"25%+). Cinza = não avaliado. **⚠️** = local de maior perda · **★** = híbrido de "
                   f"menor perda. {len(hibridos_hp)} híbridos · {len(locais_hp)} locais")
        _cap_hp += (f" (de {n_locais_total_hp} ativos)." if len(locais_hp) < n_locais_total_hp else ".")
        st.caption(_cap_hp)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — PIORES CASOS (máximo de perda antes de diluir na média)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Piores Casos",
    "Qual foi o pior caso de cada híbrido, antes de diluir na média?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
A média (usada na tabela-resumo) é a taxa **na rede** — boa para comparar híbridos no geral, mas ela
**esconde os picos**. Um híbrido pode ter média baixa e mesmo assim ter tido um pico forte num local
específico: aquele valor alto dilui entre os plots limpos e some da média.

Vale para **perdas e fenômenos**. Para uma perda (acamamento, quebramento…), o pico é o local onde
mais plantas se perderam; para um fenômeno (green snap, morte prematura…), é o local de maior
severidade. Em ambos, é o pior caso que a média não mostra.

Esta tabela mostra os dois lado a lado, para expor o risco pontual:

- **Média (rede)** → a taxa média, com a régua da família (perda com o zero; fenômeno e ardidos só
  onde ocorreu). A mesma da tabela-resumo.
- **Máximo** → o **pior plot** do híbrido: onde a perda foi mais alta. É o valor cru, sem diluir.
- **Local do máx** → onde aconteceu o pior caso.
- **Δ (máx − média)** → o tamanho do salto. Um Δ grande significa que o híbrido é irregular: vai
  bem na média, mas teve um pico forte em algum lugar. Δ pequeno = comportamento uniforme.
- **Locais críticos (> limite)** → em **quantos** locais a perda passou do limite que você define no
  campo acima. Enquanto o **máximo** mede a *intensidade* do pior caso (um número), esta coluna
  mede a *frequência* dos casos ruins. A distinção importa: um híbrido com máximo alto mas só
  **1** local crítico teve um pico isolado (pode ter sido azar daquele local); um com vários
  locais críticos tem um problema recorrente — mais preocupante, mesmo que o pico seja menor.

**Como usar:** um híbrido com média baixa **e** máximo baixo é consistente. Um com média baixa mas
máximo alto tem um risco escondido — vale investigar o que houve naquele local (vento, manejo,
suscetibilidade pontual). A tabela ordena pelo maior máximo: quem teve o pior caso aparece no topo.

O **máximo** é sempre o pior plot real; a **média** segue a régua da família (perda com o zero;
fenômeno e ardidos só onde ocorreu).
""")

_opts_pc = [rot for rot, col in VARS_COLHEITA if col in ta_filtrado.columns
            and pd.to_numeric(ta_filtrado[col], errors="coerce").gt(0).any()]

if not _opts_pc:
    st.info("Nenhuma perda ou fenômeno com ocorrência nos filtros ativos.")
else:
    _c_pc1, _c_pc2 = st.columns([2, 2])
    with _c_pc1:
        perda_pc = st.selectbox("Perda ou fenômeno", options=_opts_pc, key="perdas_pc_var")
    col_pc = dict(VARS_COLHEITA)[perda_pc]
    eh_fen_pc = perda_pc in _ROTULOS_FENOMENO

    df_pc = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                         "cidade_nome", "estado_sigla", col_pc]].copy()
    df_pc[col_pc] = pd.to_numeric(df_pc[col_pc], errors="coerce")
    df_pc = df_pc.dropna(subset=[col_pc])

    with _c_pc2:
        _lim_pc = st.number_input("Limite p/ 'locais críticos' (%)", min_value=1.0, max_value=100.0,
                                  value=20.0, step=1.0, key="perdas_pc_lim",
                                  help="Conta em quantos locais o híbrido passou deste valor.")

    if df_pc.empty:
        st.info(f"Nenhum registro de {perda_pc} nos filtros ativos.")
    else:
        # valor por híbrido × local (para a média com a régua da família e para achar o máximo)
        reg_rows = []
        for (hib, st_mat, loc, cid, uf), grp in df_pc.groupby(
                ["dePara", "status_material", "cod_fazenda", "cidade_nome", "estado_sigla"]):
            s = grp[col_pc]
            # valor do local: régua da família (para a média)
            val_fam = _valor_por_familia(s, eh_fen_pc)
            # máximo do local: sempre o pior plot real
            val_max = float(s.max())
            reg_rows.append({"dePara": hib, "status_material": st_mat, "cod_fazenda": loc,
                             "cidade_nome": cid, "estado_sigla": uf,
                             "val_fam": val_fam if val_fam is not None else 0.0, "val_max": val_max})
        df_pc_agg = pd.DataFrame(reg_rows)

        pc_rows = []
        for hib, g in df_pc_agg.groupby("dePara"):
            status = g["status_material"].mode()
            status = status.iloc[0] if not status.empty else ""
            # média na rede: média dos valores-de-local (que já vêm com a régua da família)
            media = round(g["val_fam"].mean(), 1)
            # máximo: o pior local, e onde foi
            idx_max = g["val_max"].idxmax()
            maximo = round(g.loc[idx_max, "val_max"], 1)
            local_max = g.loc[idx_max, "cod_fazenda"]
            cidade_max = g.loc[idx_max, "cidade_nome"]
            n_criticos = int((g["val_max"] > _lim_pc).sum())
            pc_rows.append({
                "Híbrido": hib, "Status": status, "Locais": len(g),
                "Média (rede)": media, "Máximo": maximo,
                "Local do máx": f"{local_max} ({cidade_max})",
                "Δ (máx − média)": round(maximo - media, 1),
                f"Locais críticos (> {_lim_pc:.0f}%)": n_criticos})

        df_pior = (pd.DataFrame(pc_rows)
                   .sort_values("Máximo", ascending=False).reset_index(drop=True))

        gb_pc = GridOptionsBuilder.from_dataframe(df_pior)
        gb_pc.configure_default_column(
            resizable=True, sortable=True, filter=True,
            cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
        gb_pc.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
        gb_pc.configure_column("Híbrido", pinned="left", width=160)
        gb_pc.configure_column("Status", width=90)
        gb_pc.configure_column("Locais", width=80)
        gb_pc.configure_column("Média (rede)", width=110)
        gb_pc.configure_column("Máximo", width=100,
            cellStyle=JsCode("function(p){var v=p.value; if(v>=25)return{background:'#C8201E',color:'#FFF',fontWeight:'700'}; if(v>=15)return{background:'#E8850D',color:'#1A1A1A',fontWeight:'700'}; if(v>=8)return{background:'#F6D202',color:'#1A1A1A'}; return{background:'#D5F5D5',color:'#1A1A1A'};}"))
        gb_pc.configure_column("Local do máx", width=200)
        gb_pc.configure_column("Δ (máx − média)", width=130,
            cellStyle=JsCode("function(p){var v=p.value; if(v>=20)return{background:'#C8201E',color:'#FFF',fontWeight:'700'}; if(v>=10)return{background:'#E8850D',color:'#1A1A1A'}; return{};}"))
        gb_pc.configure_column(f"Locais críticos (> {_lim_pc:.0f}%)", width=160,
            cellStyle=JsCode("function(p){if(p.value>0)return{background:'#FDDCDE',color:'#1A1A1A',fontWeight:'700'};return{background:'#D5F5D5',color:'#1A1A1A'};}"))

        go_pc = gb_pc.build()
        go_pc["defaultColDef"]["headerClass"] = "ag-header-black"
        AgGrid(df_pior, gridOptions=go_pc, height=min(560, 36 + 32 * len(df_pior) + 20),
               update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
               allow_unsafe_jscode=True, enable_enterprise_modules=True,
               custom_css={
                   ".ag-header":            {"background-color": "#4A4A4A !important"},
                   ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                   ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                   ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                   ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                   ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                   ".ag-row":               {"font-size": "13px !important"},
               },
               theme="streamlit", use_container_width=True)
        st.caption(f"**Máximo** = pior plot de {perda_pc}, sem diluir na média. "
                   f"**Δ** grande = híbrido irregular (média baixa mas com um pico forte). "
                   f"Ordenado pelo maior máximo. {len(df_pior)} híbridos.")
        exportar_excel(df_pior, nome_arquivo="piores_casos_perda.xlsx",
                       label="⬇️ Exportar Piores Casos", key="exp_pc_perdas")


st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — SIMULADOR DE PERDAS (e se a perda fosse menor? efeito na produtividade)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Simulador de Perdas",
    "Quanto o híbrido produziria com menos perda?",
    contexto_str,
)

with st.popover("ℹ️ Como funciona", use_container_width=False):
    st.markdown("""
Esta é uma ferramenta de **"e se"**: você ajusta as perdas de um híbrido e vê quanto de
produtividade ele recuperaria (ou perderia). Serve para enxergar quanto de produção as perdas de
colheita estão custando.

---

#### De onde vêm os valores dos controles

Quando você escolhe um híbrido, cada slider já começa no **valor médio real dele na rede**, contando
só os plots que passaram pelos filtros ativos na barra lateral. Ou seja: se o slider de acamamento
mostra 2,8%, é porque, na média dos locais filtrados, esse híbrido acamou 2,8%. A produtividade de
partida também é a média da rede sob os mesmos filtros.

Nada é inventado ou estimado por fora — o ponto de partida é sempre o que os dados mostram. Mudar os
filtros (região, safra, local) muda os valores iniciais, porque muda o conjunto de plots.

---

#### O passo a passo do cálculo

**1. A produção que você vê já é a que sobrou.** As plantas que acamaram, quebraram ou foram
dominadas não entraram na produtividade colhida. Então a produção da rede já embute as perdas.

**2. Descobrimos a produção potencial** — o que teria sido colhido sem nenhuma perda:

```
potencial = produção média ÷ (1 − perda total ÷ 100)
```

**3. Você ajusta as perdas nos sliders.** A perda total é a soma dos quatro componentes (acamamento
+ quebramento + dominadas + colmo podre) — confirmado nos dados, a soma bate exato com a perda total
medida.

**4. A produção simulada recalcula sobre o potencial:**

```
produção simulada = potencial × (1 − nova perda total ÷ 100)
```

Reduziu a perda, recupera parte do potencial; aumentou, perde mais. A conta é a mesma que a
agronomia diz: perda física é planta que não produziu, então recuperá-la recupera produção na mesma
proporção.

---

#### Cuidados com a análise

- **É uma estimativa de teto, não uma promessa.** Ela isola só o efeito das perdas físicas,
  mantendo todo o resto igual. A produtividade real depende de muito mais — genética, clima, solo,
  manejo, população. Reduzir o acamamento não garante o ganho todo; ele mostra o **máximo** que se
  recuperaria se só a perda mudasse.
- **Compare dentro do mesmo recorte.** Os valores partem da média sob os filtros ativos. Não faz
  sentido comparar um cenário filtrado por uma região com outro de outra região — são conjuntos de
  plots diferentes.
- **Os fenômenos não entram.** Green snap, morte prematura e os outros têm base de cálculo
  diferente (não somam com a perda total), então ficam de fora desta conta para não misturar réguas.
- **Cuidado com híbridos de poucos locais.** Se o híbrido aparece em poucos plots no recorte, a
  média de partida é frágil e o cenário fica menos confiável.

---

#### Como interpretar o resultado

Os três cartões mostram a **perda total simulada**, a **produção simulada** (em kg/ha e sacas) e os
**deltas** em relação ao valor atual. A barra compara três situações:

- **Atual** (cinza) → o que o híbrido produz hoje, com as perdas que teve.
- **Simulada** (azul se ganhou, vermelho se perdeu) → o cenário que você montou.
- **Potencial** (verde) → o teto, sem perda nenhuma.

A distância entre "Atual" e "Potencial" é **todo o custo das perdas de colheita** naquele híbrido. A
distância entre "Atual" e "Simulada" é **quanto o seu cenário recupera desse custo**. Se o simulado
encosta no potencial, você zerou quase tudo; se fica perto do atual, o ganho é pequeno.
""")

_hib_sim = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
if not _hib_sim:
    st.info("Nenhum híbrido disponível nos filtros ativos.")
else:
    hib_sim = st.selectbox("Selecione o híbrido", options=_hib_sim, key="perdas_sim_hib")
    g_sim = ta_filtrado[ta_filtrado["dePara"] == hib_sim]

    prod_atual = pd.to_numeric(g_sim["kg_ha"], errors="coerce").mean()
    _comp_medio = {}
    for nome, col in PERDAS.items():
        if col in g_sim.columns:
            _comp_medio[nome] = round(float(pd.to_numeric(g_sim[col], errors="coerce").fillna(0).mean()), 1)
        else:
            _comp_medio[nome] = 0.0
    perda_total_atual = round(sum(_comp_medio.values()), 1)

    if pd.isna(prod_atual) or prod_atual <= 0:
        st.warning(f"**{hib_sim}** não tem produtividade registrada nos filtros ativos — "
                   "não dá para simular sem uma produção de base.")
    else:
        potencial = prod_atual / (1 - perda_total_atual / 100) if perda_total_atual < 100 else prod_atual

        st.markdown(f"**{hib_sim}** — média da rede sob os filtros atuais: "
                    f"produção **{prod_atual:.0f} kg/ha** ({prod_atual/60:.1f} sc/ha), "
                    f"perda total **{perda_total_atual:.1f}%**. "
                    f"Produção potencial sem perdas: **{potencial:.0f} kg/ha**.")

        st.markdown("##### Ajuste as perdas e veja o efeito")
        _cols_sim = st.columns(len(PERDAS))
        _novo_comp = {}
        for i, (nome, col) in enumerate(PERDAS.items()):
            atual = _comp_medio.get(nome, 0.0)
            _max_slider = float(max(10.0, np.ceil(atual * 1.5 / 5) * 5))
            _novo_comp[nome] = _cols_sim[i].slider(
                f"{nome}", min_value=0.0, max_value=_max_slider,
                value=float(atual), step=0.5, key=f"sim_{col}", format="%.1f%%",
                help=f"Percentual de perda por {nome.lower()}. Média na rede: {atual:.1f}%.")

        perda_total_nova = round(sum(_novo_comp.values()), 1)
        prod_simulada = potencial * (1 - perda_total_nova / 100)
        delta_kg = prod_simulada - prod_atual
        delta_sc = delta_kg / 60
        delta_pct = (delta_kg / prod_atual * 100) if prod_atual else 0

        st.markdown("##### Resultado")
        _r1, _r2, _r3 = st.columns(3)
        _r1.metric("Perda total simulada", f"{perda_total_nova:.1f}%",
                   f"{perda_total_nova - perda_total_atual:+.1f} pp", delta_color="inverse")
        _r2.metric("Produção simulada", f"{prod_simulada:.0f} kg/ha",
                   f"{delta_kg:+.0f} kg/ha")
        _r3.metric("Em sacas", f"{prod_simulada/60:.1f} sc/ha",
                   f"{delta_sc:+.1f} sc/ha")

        fig_sim = go_plt.Figure()
        _cats = ["Atual", "Simulada", "Potencial (sem perda)"]
        _vals = [prod_atual, prod_simulada, potencial]
        _cores = ["#7F7F7F", "#2976B6" if delta_kg >= 0 else "#C0201E", "#1E7A34"]
        fig_sim.add_trace(go_plt.Bar(
            x=_cats, y=_vals, marker_color=_cores,
            text=[f"{v:.0f} kg/ha<br>{v/60:.1f} sc/ha" for v in _vals],
            textposition="outside", textfont=dict(size=12, color="#111111", weight="bold"),
            width=0.5))
        fig_sim.update_layout(
            height=380,
            yaxis=dict(title=dict(text="<b>Produtividade (kg/ha)</b>",
                                  font=dict(size=13, color="#111111", weight="bold")),
                       range=[0, potencial * 1.15], tickfont=dict(size=11, color="#111111"),
                       gridcolor="#EEEEEE"),
            xaxis=dict(tickfont=dict(size=12, color="#111111", weight="bold")),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            margin=dict(t=30, b=40, l=60, r=20), showlegend=False,
            font=dict(family="Helvetica Neue, sans-serif"))
        st.plotly_chart(fig_sim, use_container_width=True)

        if abs(delta_kg) < 1:
            st.info("Os valores estão iguais aos da média da rede — ajuste as perdas acima para simular.")
        elif delta_kg > 0:
            st.success(f"Reduzindo as perdas de {perda_total_atual:.1f}% para {perda_total_nova:.1f}%, "
                       f"**{hib_sim}** recuperaria **{delta_kg:+.0f} kg/ha** ({delta_sc:+.1f} sc/ha, "
                       f"{delta_pct:+.1f}%) — o que a perda de colheita está custando.")
        else:
            st.warning(f"Aumentando as perdas para {perda_total_nova:.1f}%, a produção cairia "
                       f"**{delta_kg:.0f} kg/ha** ({delta_sc:.1f} sc/ha).")

        st.caption("Estimativa de teto: isola o efeito das perdas físicas, mantendo tudo o mais "
                   "constante. Produção e perdas de base são a média da rede sob os filtros ativos.")

st.divider()

rodape()
