"""
pages/2_Analise_Conjunta.py — Análise Conjunta de Produtividade (milho)

Adaptada do painel de soja, seguindo Better Data Visualization (Schwabish).
Fonte: tabela_analitica_faixa das safras 2024/25 e 2025/26.

Seções (construídas incrementalmente):
  1. Auditoria — tabela analítica por ensaio + produção relativa
"""
import numpy as np
import pandas as pd
import streamlit as st
from scipy import stats
import plotly.graph_objects as go_plt

# ── Cores por status do material (milho: CHECK / STINE / EXP / DP2) ──────────
COR_STATUS_PLOT = {
    "CHECK": "#F4B184",   # testemunha externa (laranja)
    "STINE": "#2976B6",   # comercial Stine (azul)
    "EXP":   "#00FF00",   # experimental / em avaliação (verde vibrante)
    "DP2":   "#C4DFB4",   # duplo propósito / segundo ano (verde claro)
}
COR_BORDA = {
    "CHECK": "#C46A3A",
    "STINE": "#1A4F7A",
    "EXP":   "#009900",
    "DP2":   "#7AAF6A",
}

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_multisafra
from utils.tabelas import cel, render_tabela
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

st.set_page_config(
    page_title="Análise Conjunta · JAUM DTC",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()


# ── Helper AgGrid (mesma configuração da soja: header escuro, menu funcionando) ─
def ag_table(df, height=400, estilos_col=None, renderers_col=None):
    """Tabela AgGrid padrão. `estilos_col`: {coluna: JsCode} para cellStyle ·
    `renderers_col`: {coluna: JsCode} para cellRenderer (permite barras na célula)."""
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    for _col in df.columns:
        if str(_col).startswith("_"):          # colunas técnicas (ex.: _st) não aparecem
            gb.configure_column(_col, hide=True)
    for _col, _estilo in (estilos_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellStyle=_estilo)
    for _col, _rend in (renderers_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellRenderer=_rend, minWidth=170)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                   {"background-color": "#4A4A4A !important"},
            ".ag-header-row":               {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":              {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":        {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":         {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                     {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":  {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-icon-menu":                {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # sem "color" aqui: um !important sobreporia a cor inline do cellStyle
            # (era o que impedia o verde/vermelho dos desvios de aparecer)
            ".ag-cell":                     {"font-size": "13px !important"},
            ".ag-row":                      {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True,
    )


# ── Helper exportar Excel (mesmo padrão da soja) ──────────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = border

    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


# ── LSD via ANOVA conjunta (material × local), OLS numpy ──────────────────────
def calcular_lsd(df, col="kg_ha", fator="dePara", bloco="cod_fazenda", alpha=0.05):
    """Diferença Mínima Significativa (5%) da análise conjunta.
    Modelo: y = µ + material + local + erro. np.nan se sem graus de liberdade."""
    try:
        d = df[[col, fator, bloco]].dropna().copy()
        d = d[d[col] > 0].reset_index(drop=True)
        if d.empty or d[fator].nunique() < 2 or d[bloco].nunique() < 2:
            return np.nan
        y = d[col].values.astype(float)
        X_cult = pd.get_dummies(d[fator], drop_first=True).values.astype(float)
        X_local = pd.get_dummies(d[bloco], drop_first=True).values.astype(float)
        X = np.hstack([np.ones((len(y), 1)), X_cult, X_local])
        beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
        ss_res = np.sum((y - X @ beta) ** 2)
        gl_res = len(y) - rank
        if gl_res <= 0:
            return np.nan
        qmr = ss_res / gl_res
        n_bloco = d[bloco].nunique()
        t_crit = stats.t.ppf(1 - alpha / 2, df=gl_res)
        return round(t_crit * np.sqrt(2 * qmr / n_bloco), 1)
    except Exception:
        return np.nan


# CSS para tabelas nativas legíveis (padrão da soja)
st.markdown("""
<style>
[data-testid="stDataFrame"] td, [data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"] span {
    font-size: 13px !important; font-weight: 600 !important; color: #000000 !important;
}
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] {
    color: #374151 !important; opacity: 1 !important;
}
</style>
""", unsafe_allow_html=True)

page_header(
    "Análise Conjunta",
    "Avalie o desempenho produtivo dos materiais em faixa considerando múltiplos locais e safras. "
    "Comece conferindo os dados por ensaio na Auditoria; as análises agregadas vêm em seguida.",
    imagem="Business mission-amico.png",
)


# ── Carregamento: analítica de Faixa das duas safras, já reconciliada ─────────
# Usa carregar_multisafra(), que aplica o depara_mestre — sem ele, o mesmo híbrido
# aparece com nomes diferentes em cada safra (ex.: 9505VTPRO4 em 24/25 e 9505PRO4 em 25/26).
@st.cache_data(show_spinner=False)
def carregar_concat():
    d = carregar_multisafra()
    df = d.get("tabela_analitica_faixa")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    df = df.copy()
    # produtividade canônica: usa a válida (kg/ha e sacas), com fallback para a bruta
    if "produtividade_valida_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_valida_kg_ha"], errors="coerce")
        if "produtividade_kg_ha" in df.columns:
            df["kg_ha"] = df["kg_ha"].fillna(pd.to_numeric(df["produtividade_kg_ha"], errors="coerce"))
    elif "produtividade_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_kg_ha"], errors="coerce")
    if "produtividade_valida_sacas_ha" in df.columns:
        df["sc_ha"] = pd.to_numeric(df["produtividade_valida_sacas_ha"], errors="coerce")
    elif "kg_ha" in df.columns:
        df["sc_ha"] = (df["kg_ha"] / 60).round(1)   # 1 saca = 60 kg
    # altura em metros (o pipeline entrega em cm)
    if "altura_planta_cm" in df.columns:
        df["altura_planta_m"] = (pd.to_numeric(df["altura_planta_cm"], errors="coerce") / 100).round(1)
    if "altura_espiga_cm" in df.columns:
        df["altura_espiga_m"] = (pd.to_numeric(df["altura_espiga_cm"], errors="coerce") / 100).round(1)
    return df


@st.cache_data(show_spinner=False)
def carregar_densidade():
    """Analítica de Densidade — usada apenas para contar ensaios por tipo nos cards.
    As análises desta página continuam exclusivamente com Faixa."""
    d = carregar_multisafra()
    df = d.get("tabela_analitica_densidade")
    return df.copy() if isinstance(df, pd.DataFrame) and not df.empty else pd.DataFrame()


with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()


# ── Sidebar — filtros encadeados ──────────────────────────────────────────────
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>', unsafe_allow_html=True)

    if st.button("Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["safra_", "macro_", "micro_", "estado_", "cidade_",
                                               "fazenda_", "resp_", "status_", "cult_",
                                               "busca_", "__opts_"]):
                del st.session_state[key]
        st.rerun()

    def _podar_keys(prefix, opcoes, molde):
        """Remove o estado de checkboxes de opções que saíram da cascata.
        Sem isso, ao reaparecerem elas voltam marcadas e o filtro se reaplica sozinho."""
        antigas = st.session_state.get(f"__opts_{prefix}", [])
        atuais = set(map(str, opcoes))
        for o in antigas:
            if str(o) not in atuais:
                st.session_state.pop(molde(o), None)
        st.session_state[f"__opts_{prefix}"] = list(opcoes)

    def checkboxes(opcoes, default_all=True, defaults=None, prefix=""):
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_{o}")
        sel = []
        for o in opcoes:
            key = f"{prefix}_{o}"
            if key in st.session_state:          # já existe: o widget é a verdade
                marcado = st.checkbox(str(o), key=key)
            else:                                 # primeira criação: aplica o padrão
                checked = (o in defaults) if defaults is not None else default_all
                marcado = st.checkbox(str(o), value=checked, key=key)
            if marcado:
                sel.append(o)
        return sel

    def filtro_busca(opcoes, prefix):
        """Busca textual + seleção persistente. A seleção é lida direto dos checkboxes
        (fonte única), inclusive dos itens ocultos pela busca."""
        if f"{prefix}_reset" not in st.session_state:
            st.session_state[f"{prefix}_reset"] = 0
        r = st.session_state[f"{prefix}_reset"]
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_chk_{r}_{o}")

        busca = st.text_input("Buscar", value="", key=f"busca_{prefix}", placeholder="Digite para filtrar...")
        filtradas = [c for c in opcoes if busca.strip().lower() in str(c).lower()] if busca.strip() else opcoes

        if st.button("Limpar seleção", key=f"{prefix}_limpar", use_container_width=True):
            for o in opcoes:
                st.session_state.pop(f"{prefix}_chk_{r}_{o}", None)
            st.session_state.pop(f"__opts_{prefix}", None)
            st.session_state[f"{prefix}_reset"] = r + 1
            st.rerun()

        for c in filtradas:
            st.checkbox(str(c), key=f"{prefix}_chk_{r}_{c}")

        sel = [o for o in opcoes if st.session_state.get(f"{prefix}_chk_{r}_{o}", False)]
        return sel or opcoes

    # 1. Safra — padrão safra atual (na base o valor é "25/26")
    with st.expander("Safra", expanded=True):
        safras_all = sorted(ta_raw["safra"].dropna().unique().tolist())
        SAFRA_PADRAO = ("25/26", "2025/26")   # aceita os dois formatos
        safra_default = [s for s in safras_all if str(s) in SAFRA_PADRAO] or safras_all[-1:]
        # inicializa o estado UMA vez: garante só a safra padrão marcada
        if "safra_init_v2" not in st.session_state:
            for o in safras_all:
                st.session_state[f"safra_{o}"] = (o in safra_default)
            st.session_state["safra_init_v2"] = True
        safras_sel = checkboxes(safras_all, defaults=safra_default, prefix="safra")
    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("Região Macro", expanded=False):
        macros_sel = checkboxes(sorted(ta_f1["regiao_macro"].dropna().unique().tolist()), prefix="macro")
    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("Região Micro", expanded=False):
        micros_sel = checkboxes(sorted(ta_f2["regiao_micro"].dropna().unique().tolist()), prefix="micro")
    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("Estado", expanded=False):
        estados_sel = filtro_busca(sorted(ta_f3["estado_sigla"].dropna().unique().tolist()), "estado")
    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("Cidade", expanded=False):
        cidades_sel = filtro_busca(sorted(ta_f4["cidade_nome"].dropna().unique().tolist()), "cidade")
    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("Fazenda", expanded=False):
        fazendas_sel = filtro_busca(sorted(ta_f5["nomeFazenda"].dropna().unique().tolist()), "fazenda")
    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Responsável
    with st.expander("Responsável", expanded=False):
        resps_sel = filtro_busca(sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist()), "resp")
    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # 8. Status do material
    with st.expander("Status do Híbrido", expanded=False):
        status_sel = checkboxes(sorted(ta_f7["status_material"].dropna().unique().tolist()), prefix="status")
    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    # 9. Material (dePara)
    with st.expander("Híbrido", expanded=False):
        materiais_sel = filtro_busca(sorted(ta_f8["dePara"].dropna().unique().tolist()), "cult")
    ta_filtrado = ta_f8[ta_f8["dePara"].isin(materiais_sel)] if materiais_sel else ta_f8.iloc[0:0]


def _aplicar_filtros_local(df):
    """Aplica só os filtros de tempo e geografia — um ensaio existe independentemente
    de quais híbridos ou status foram selecionados para a análise. Definida aqui, logo
    após os filtros, porque a produção relativa da Auditoria já a usa."""
    if df.empty:
        return df
    d = df
    for _col, _sel in [("safra", safras_sel), ("regiao_macro", macros_sel),
                       ("regiao_micro", micros_sel), ("estado_sigla", estados_sel),
                       ("cidade_nome", cidades_sel), ("nomeFazenda", fazendas_sel),
                       ("nomeResponsavel", resps_sel)]:
        if _sel and _col in d.columns:
            d = d[d[_col].isin(_sel)]
    return d


def _descreve_base(base, testemunha=None, escopo="local"):
    """Texto curto do critério de referência, para o aviso das tabelas.
    escopo='local' (Auditoria: referência dentro de cada local) ou
    escopo='material' (Desempenho/Apresentação: referência sobre os materiais filtrados)."""
    if escopo == "material":
        if base == "Maior produtividade":
            return "o **maior** rendimento entre os materiais filtrados (o líder = 100%)"
        if base == "Testemunha" and testemunha:
            return f"a média da testemunha **{testemunha}**"
        return "a **média** dos materiais filtrados"
    # escopo local (Auditoria)
    if base == "Maior produtividade":
        return "o **maior** rendimento do local (o líder = 100%)"
    if base == "Testemunha" and testemunha:
        return f"a média da testemunha **{testemunha}** no local"
    return "a **média** de todos os híbridos do local"


if ta_filtrado.empty:
    st.warning("Nenhum dado para os filtros selecionados.")
    st.stop()


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA (tabela analítica por ensaio + produção relativa)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação (a tabela analítica de Faixa). Use para conferência dos "
    "dados antes das análises agregadas.",
)

# ── Seletor de base da Produção Relativa (3 bases, como na soja) ──────────────
col_ref, col_test, _ = st.columns([2, 2, 3])
with col_ref:
    base_rel = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0,
    )
with col_test:
    if base_rel == "Testemunha":
        testemunhas = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist())
        testemunha_sel = st.selectbox("Selecione a testemunha", options=testemunhas) if testemunhas else None
        if not testemunhas:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
    else:
        testemunha_sel = None

# ── Produção relativa (%) — POR LOCAL, com referência FIXA do ensaio inteiro ──
#    A referência de cada local (média/máximo/testemunha) é calculada sobre TODOS os híbridos
#    do ensaio — via ta_raw + filtros de local — e NÃO sobre os híbridos filtrados. Assim o
#    filtro de material/status muda só o que aparece na tabela, não o "100% do local": o
#    desempenho relativo de um híbrido é sempre contra o ensaio todo, não contra os pares que
#    sobraram no filtro. LOCAL = (safra, cod_fazenda) para não juntar safras de um mesmo local.
df_tabela = ta_filtrado.copy()
LOCAL = [c for c in ["safra", "cod_fazenda"] if c in df_tabela.columns] or ["cod_fazenda"]

# universo de referência: ensaio inteiro (só filtros de tempo/geografia, sem filtro de híbrido)
_ref_scope = _aplicar_filtros_local(ta_raw)
_ref_scope = _ref_scope[pd.to_numeric(_ref_scope["kg_ha"], errors="coerce") > 0]

if base_rel == "Maior produtividade":
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].max()
elif base_rel == "Testemunha" and testemunha_sel:
    _ref = (_ref_scope[_ref_scope["dePara"] == testemunha_sel].groupby(LOCAL)["kg_ha"].mean())
else:  # "Média geral do ensaio" (padrão)
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].mean()

# casa cada linha exibida com a referência do seu (safra, local)
_chave = df_tabela.set_index(LOCAL).index
ref_por_local = pd.Series(_chave.map(_ref).to_numpy(), index=df_tabela.index)
df_tabela["prod_relativa_pct"] = ((df_tabela["kg_ha"] / ref_por_local) * 100).round(1)

# ── Colunas para exibir (analítica do milho, rótulos amigáveis) ───────────────
col_map = {
    "safra":                      "Safra",
    "cod_fazenda":                "Cód. Local",
    "nomeFazenda":                "Fazenda",
    "cidade_nome":                "Cidade",
    "estado_sigla":               "Estado",
    "regiao_macro":               "Região Macro",
    "regiao_micro":               "Região Micro",
    "nomeResponsavel":            "Responsável",
    "dePara":                     "Híbrido",
    "status_material":            "Status",
    "indexTratamento":            "Trat.",
    "dataPlantioMilho":           "Plantio",
    "dataColheitaMilho":          "Colheita",
    "kg_ha":                      "kg/ha",
    "sc_ha":                      "sc/ha",
    "prod_relativa_pct":          "Prod. Relativa (%)",
    "umidade_pct":                "Umidade (%)",
    "populacao_real_plantas_ha":  "Pop. Real (pl/ha)",
    "altura_planta_m":            "Alt. Planta (m)",
    "altura_espiga_m":            "Alt. Espiga (m)",
    "pmg_corrigido_g":            "PMG (g)",
    "fileiras_media":             "Fileiras",
    "graos_fileira_media":        "Grãos/Fileira",
    "graos_ardidos_pct":          "Ardidos (%)",
    "pct_acamadas":               "Acamamento (%)",
    "pct_colmo_podre":            "Colmo Podre (%)",
    "pct_quebradas":              "Quebramento (%)",
    "pct_dominadas":              "Dominadas (%)",
    "pct_perda_total":            "Perda Total (%)",
}
cols_disp = [c for c in col_map if c in df_tabela.columns]

# ordenação hierárquica: Macro → Micro → Estado → Cidade → Fazenda; dentro da fazenda, maior→menor kg/ha
_hier = [c for c in ["regiao_macro", "regiao_micro", "estado_sigla", "cidade_nome", "nomeFazenda"]
         if c in df_tabela.columns]
if _hier and "kg_ha" in df_tabela.columns:
    df_tabela = df_tabela.sort_values(
        _hier + ["kg_ha"],
        ascending=[True] * len(_hier) + [False],
        na_position="last",
    ).reset_index(drop=True)

df_show = df_tabela[cols_disp].rename(columns=col_map)
# população como número inteiro (não 60942.0)
if "Pop. Real (pl/ha)" in df_show.columns:
    df_show["Pop. Real (pl/ha)"] = pd.to_numeric(df_show["Pop. Real (pl/ha)"], errors="coerce").round(0).astype("Int64")
# datas em dd/mm/aaaa (sem hora)
for _c in ["Plantio", "Colheita"]:
    if _c in df_show.columns:
        df_show[_c] = pd.to_datetime(df_show[_c], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")

ag_table(df_show, height=min(560, 40 + 32 * min(len(df_show), 15) + 20))
exportar_excel(df_show, nome_arquivo="auditoria_milho.xlsx",
               label="⬇️ Exportar Auditoria", key="exp_auditoria")

st.caption(f"{len(df_show)} observações · {df_tabela['dePara'].nunique()} híbridos · "
           f"{df_tabela['cod_fazenda'].nunique()} locais.")
st.info(f"📐 **Produção Relativa — método POR LOCAL.** Cada plot é comparado com "
        f"{_descreve_base(base_rel, testemunha_sel)}, calculado **dentro de cada local**. "
        "A referência é fixa no ensaio inteiro: mudar o filtro de híbrido muda só o que aparece, "
        "não o denominador.")

st.divider()

# ── Mapa dos locais do recorte ────────────────────────────────────────────────
secao_titulo("Localização", "Onde estão os ensaios de desenvolvimento de produtos?",
             "Distribuição geográfica da rede de ensaios que compõe as análises acima. Os pontos "
             "seguem os filtros da barra lateral.")

# ── Cards: rede de ensaios do recorte ──
# Um ensaio é um local × tipo de teste. As datas de plantio e colheita são da fazenda,
# então valem para os dois tipos de ensaio daquele local.
_faixa_e = _aplicar_filtros_local(ta_raw).assign(_tipo="Faixa")
_dens_e = _aplicar_filtros_local(carregar_densidade()).assign(_tipo="Densidade")
_cols_e = [c for c in ["cod_fazenda", "dataPlantioMilho", "dataColheitaMilho", "_tipo"]
           if c in _faixa_e.columns or c in _dens_e.columns]
_ens = pd.concat(
    [d[[c for c in _cols_e if c in d.columns]] for d in [_faixa_e, _dens_e] if not d.empty],
    ignore_index=True) if (not _faixa_e.empty or not _dens_e.empty) else pd.DataFrame()

if not _ens.empty:
    _ens = _ens.drop_duplicates(subset=["cod_fazenda", "_tipo"])
    _ens["_plant"] = pd.to_datetime(_ens.get("dataPlantioMilho"), errors="coerce")
    _ens["_colh"] = pd.to_datetime(_ens.get("dataColheitaMilho"), errors="coerce")

    _n_plant = int(_ens["_plant"].notna().sum())
    _n_faixa = int(((_ens["_tipo"] == "Faixa") & _ens["_plant"].notna()).sum())
    _n_dens = int(((_ens["_tipo"] == "Densidade") & _ens["_plant"].notna()).sum())
    _n_colh = int((_ens["_plant"].notna() & _ens["_colh"].notna()).sum())

    def _pct(n):
        return f"{n / _n_plant * 100:.0f}%" if _n_plant else "—"

    _card_css = ("border:1px solid #E5E7EB;border-radius:10px;padding:12px 16px;background:#FFFFFF;"
                 "text-align:center;box-shadow:0 1px 4px rgba(0,0,0,0.07);")
    _c1, _c2, _c3, _c4 = st.columns(4)
    for _col_ui, _lbl, _n, _sub, _cor in [
            (_c1, "Ensaios plantados", _n_plant, "local × tipo de ensaio", "#1A1A1A"),
            (_c2, "Faixa", _n_faixa, _pct(_n_faixa), "#2976B6"),
            (_c3, "Densidade", _n_dens, _pct(_n_dens), "#7C3AED"),
            (_c4, "Colhidos", _n_colh, _pct(_n_colh), "#1E8449")]:
        _col_ui.markdown(
            f'<div style="{_card_css}">'
            f'<p style="margin:0;font-size:13px;color:#6B7280;font-weight:600;">{_lbl}</p>'
            f'<p style="margin:4px 0 0;font-size:2rem;font-weight:700;color:{_cor};">{_n}</p>'
            f'<p style="margin:0;font-size:12px;color:#6B7280;">{_sub}</p></div>',
            unsafe_allow_html=True)
    st.caption("ℹ️ Um ensaio é um **local × tipo** (a mesma fazenda com Faixa e Densidade conta duas "
               "vezes). As datas são da fazenda, então valem para os dois tipos. Estes números "
               "respondem aos filtros de safra, região, cidade, fazenda e responsável — **não** aos "
               "de híbrido e status, porque o ensaio existe independentemente do material analisado.")
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

_cols_geo = ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
             "latitude", "longitude", "dataPlantioMilho", "dataColheitaMilho", "safra"]
if not {"latitude", "longitude"}.issubset(ta_filtrado.columns):
    st.info("Coordenadas não disponíveis na base. Atualize o pipeline e limpe o cache para "
            "habilitar o mapa.")
else:
    _geo = (ta_filtrado[[c for c in _cols_geo if c in ta_filtrado.columns]]
            .dropna(subset=["latitude", "longitude"])
            .sort_values("safra").drop_duplicates("cod_fazenda", keep="last").copy())
    _sem_geo = ta_filtrado["cod_fazenda"].nunique() - len(_geo)

    # cor pelo estágio: colhido, plantado à espera de colheita, ou sem data lançada
    _dt_p = pd.to_datetime(_geo.get("dataPlantioMilho"), errors="coerce")
    _dt_c = pd.to_datetime(_geo.get("dataColheitaMilho"), errors="coerce")
    _geo["_estagio"] = np.where(_dt_p.notna() & _dt_c.notna(), "colhido",
                                np.where(_dt_p.notna(), "plantado", "sem_data"))
    _COR_ESTAGIO = {"colhido": "#1E8449", "plantado": "#D97706", "sem_data": "#9CA3AF"}

    if _geo.empty:
        st.info("Nenhum local do recorte tem coordenada cadastrada. O Diagnóstico lista quais "
                "fazendas precisam de latitude e longitude.")
    else:
        try:
            import folium
            from streamlit_folium import st_folium
        except ModuleNotFoundError:
            folium = None
            st.warning("Mapa indisponível: rode `pip install folium==0.20.0 streamlit-folium==0.27.1` "
                       "e reinicie o app.")

        if folium is not None:
            import math

            # locais que dividem a mesma coordenada ficariam empilhados: espalha num círculo pequeno
            _geo = _geo.reset_index(drop=True)
            _geo["_dlat"] = 0.0
            _geo["_dlon"] = 0.0
            _chave = (_geo["latitude"].round(4).astype(str) + "_"
                      + _geo["longitude"].round(4).astype(str))
            _n_emp = 0
            for _, _idx in _geo.groupby(_chave).groups.items():
                _idx = list(_idx)
                if len(_idx) > 1:
                    _n_emp += len(_idx)
                    for _k, _i in enumerate(_idx):
                        _ang = 2 * math.pi * _k / len(_idx)
                        _geo.at[_i, "_dlat"] = 0.03 * math.cos(_ang)
                        _geo.at[_i, "_dlon"] = 0.03 * math.sin(_ang)

            _m = folium.Map(location=[_geo["latitude"].mean(), _geo["longitude"].mean()],
                            zoom_start=5, tiles="OpenStreetMap")

            def _dt_br(v):
                _d = pd.to_datetime(v, errors="coerce")
                return _d.strftime("%d/%m/%Y") if pd.notna(_d) else "—"

            def _svg_ponto(cor):
                return ('<svg xmlns="http://www.w3.org/2000/svg" width="20" height="20">'
                        f'<circle cx="10" cy="10" r="7" fill="{cor}" stroke="white" stroke-width="2"/></svg>')
            for _, _r in _geo.iterrows():
                _desl = (_r["_dlat"] != 0) or (_r["_dlon"] != 0)
                _nota = ("<br><i style='font-size:11px;color:#6B7280;'>ponto deslocado para não "
                         "sobrepor outro local na mesma coordenada</i>" if _desl else "")
                folium.Marker(
                    location=[_r["latitude"] + _r["_dlat"], _r["longitude"] + _r["_dlon"]],
                    popup=folium.Popup(
                        f"<b>{_r['cod_fazenda']}</b> — {_r.get('nomeFazenda', '')}<br>"
                        f"{_r.get('cidade_nome', '')} — {_r.get('estado_sigla', '')}<br>"
                        f"Safra: {_r.get('safra', '—')}<br>"
                        f"Plantio: {_dt_br(_r.get('dataPlantioMilho'))}<br>"
                        f"Colheita: {_dt_br(_r.get('dataColheitaMilho'))}{_nota}", max_width=280),
                    tooltip=f"{_r['cod_fazenda']} · {_r.get('cidade_nome', '')}",
                    icon=folium.DivIcon(html=_svg_ponto(_COR_ESTAGIO[_r["_estagio"]]),
                                        icon_size=(20, 20),
                                        icon_anchor=(10, 10))).add_to(_m)

            _n_colh_m = int((_geo["_estagio"] == "colhido").sum())
            _n_plant_m = int((_geo["_estagio"] == "plantado").sum())
            _n_sd_m = int((_geo["_estagio"] == "sem_data").sum())
            _leg = [("Plantado e colhido", "#1E8449", _n_colh_m),
                    ("Plantado, ainda não colhido", "#D97706", _n_plant_m)]
            if _n_sd_m:
                _leg.append(("Sem data lançada", "#9CA3AF", _n_sd_m))
            st.markdown(
                '<div style="display:flex;gap:22px;align-items:center;margin:2px 0 8px;">'
                + "".join(
                    f'<span style="display:inline-flex;align-items:center;gap:7px;font-size:13px;'
                    f'color:#374151;"><span style="width:13px;height:13px;border-radius:50%;'
                    f'background:{c};border:1px solid #fff;box-shadow:0 0 0 1px #ccc;"></span>'
                    f'{t} <b>({n})</b></span>' for t, c, n in _leg)
                + '</div>', unsafe_allow_html=True)

            st_folium(_m, use_container_width=True, height=690, returned_objects=[])
            _txt_sem = (f" {_sem_geo} locais do recorte não aparecem por não terem coordenada "
                        "cadastrada (veja o Diagnóstico)." if _sem_geo > 0 else "")
            _txt_emp = (f" {_n_emp} locais dividem a mesma coordenada e foram afastados alguns "
                        "quilômetros entre si só para ficarem visíveis — o popup avisa quando o "
                        "ponto está deslocado." if _n_emp else "")
            _tot_rec = ta_filtrado["cod_fazenda"].nunique()
            st.caption(f"ℹ️ **{len(_geo)} de {_tot_rec} locais do recorte** aparecem no mapa, coloridos "
                       f"pelo estágio. Clique no ponto para ver fazenda, cidade e as datas de plantio "
                       f"e colheita.{_txt_sem}{_txt_emp}")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — MARCHA DE PLANTIO E COLHEITA (avanço acumulado por local)
# ════════════════════════════════════════════════════════════════════════════════
def _marcha(df_datas, col_data, verbo_plural, cor_linha):
    """Gráfico de avanço acumulado por semana, contando LOCAIS (deduplicados por fazenda).
    verbo_plural: 'plantados'/'colhidos' (para o hover)."""
    d = df_datas.copy()
    d["_dt"] = pd.to_datetime(d[col_data], errors="coerce")
    # um local por fazenda — a data do local é a data (mesma época para todo o ensaio)
    d = d.dropna(subset=["_dt"]).drop_duplicates(subset="cod_fazenda", keep="first")
    if d.empty:
        return None, 0
    total = len(d)
    d["semana"] = d["_dt"].dt.to_period("W").dt.start_time
    sem = (d.groupby("semana").size().reset_index(name="qtd").sort_values("semana"))
    sem["acum"] = sem["qtd"].cumsum()
    sem["pct"] = (sem["acum"] / total * 100).round(1)
    sem["rotulo"] = sem["semana"].dt.strftime("%d/%b").str.lstrip("0")

    fig = go_plt.Figure()
    # zonas de fundo (início 0-50, progresso 50-90, fim 90-100)
    for y0, y1, cor in [(0, 50, "rgba(220,38,38,0.07)"), (50, 90, "rgba(217,119,6,0.07)"),
                        (90, 100, "rgba(126,211,33,0.10)")]:
        fig.add_shape(type="rect", xref="paper", yref="y", x0=0, x1=1, y0=y0, y1=y1,
                      fillcolor=cor, line_width=0, layer="below")
    for y_ref, cor in [(50, "rgba(217,119,6,0.5)"), (90, "rgba(126,211,33,0.6)")]:
        fig.add_shape(type="line", xref="paper", yref="y", x0=0, x1=1, y0=y_ref, y1=y_ref,
                      line=dict(color=cor, width=1.5, dash="dot"), layer="below")
    for y_a, txt, cor in [(25, "Início", "rgba(220,38,38,0.5)"), (70, "Progresso", "rgba(217,119,6,0.6)"),
                          (95, "Fim", "rgba(100,180,50,0.8)")]:
        fig.add_annotation(x=1, xref="paper", y=y_a, yref="y", text=f"<b>{txt}</b>",
                           showarrow=False, xanchor="left", font=dict(size=11, color=cor), xshift=8)

    fig.add_trace(go_plt.Scatter(x=sem["semana"], y=sem["pct"], mode="lines",
                                 line=dict(color=cor_linha, width=2), showlegend=False, hoverinfo="skip"))

    q_max, q_min = sem["qtd"].max(), sem["qtd"].min()
    # tamanho mínimo maior que antes: a bola precisa comportar o número acumulado por dentro
    def _tam(q):
        return 34 if q_max == q_min else 24 + (q - q_min) / (q_max - q_min) * 30
    def _cor(pct):
        return "#7ED321" if pct >= 90 else ("#D97706" if pct >= 50 else "#DC2626")
    sem["tam"] = sem["qtd"].apply(_tam)
    sem["cor"] = sem["pct"].apply(_cor)

    fig.add_trace(go_plt.Scatter(
        x=sem["semana"], y=sem["pct"], mode="markers+text",
        marker=dict(size=sem["tam"], color=sem["cor"], opacity=0.90, line=dict(color="white", width=2)),
        text=sem["acum"], textposition="middle center",
        textfont=dict(size=11, color="#FFFFFF", weight="bold"),
        customdata=sem[["qtd", "acum", "rotulo"]],
        hovertemplate=("<b>Semana de %{customdata[2]}</b><br>%{customdata[0]} locais " + verbo_plural +
                       "<br>Acumulado: %{customdata[1]} locais (%{y}%)<extra></extra>"),
        showlegend=False))

    # percentual acima da bola (o número dentro é o acumulado em locais)
    for _, _r in sem.iterrows():
        fig.add_annotation(x=_r["semana"], y=_r["pct"], text=f"{_r['pct']:.0f}%",
                           showarrow=False, yshift=_r["tam"] / 2 + 11,
                           font=dict(size=11, color="#374151"))

    fig.update_layout(
        height=420, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=20, r=80, t=40, b=60),
        xaxis=dict(title="", tickformat="%d/%b", tickfont=dict(color="black", size=11),
                   showgrid=False, zeroline=False),
        yaxis=dict(title="% acumulado", range=[-5, 112], ticksuffix="%",
                   tickvals=[0, 20, 40, 60, 80, 100], tickfont=dict(color="black", size=11),
                   showgrid=True, gridcolor="rgba(0,0,0,0.05)", zeroline=False))
    return fig, total

# ── Marcha de Plantio ──
secao_titulo("Avanço de plantio", "Como foi a marcha de plantio?",
             "Evolução semanal do plantio. O número dentro da bola é o total de locais plantados até aquela semana; o tamanho indica quantos entraram na semana; e a altura, o percentual acumulado.")
if "dataPlantioMilho" in ta_filtrado.columns:
    fig_pl, n_pl = _marcha(ta_filtrado, "dataPlantioMilho", "plantados", "rgba(0,95,174,0.3)")
    if fig_pl is None:
        st.info("Nenhum local com data de plantio registrada para os filtros selecionados.")
    else:
        st.plotly_chart(fig_pl, use_container_width=True)
        _tot_pl = ta_filtrado["cod_fazenda"].nunique()
        _falta_pl = _tot_pl - n_pl
        st.caption(f"ℹ️ {n_pl} de {_tot_pl} locais do recorte têm data de plantio lançada"
                   + (f" ({_falta_pl} sem data — veja o Diagnóstico)." if _falta_pl > 0 else ".")
                   + " Cada bola é uma semana: o número dentro é o **acumulado de locais** até ali, o tamanho "
                   "reflete quantos entraram naquela semana e o percentual acima é a mesma leitura "
                   "do eixo vertical. As zonas de fundo marcam início (até 50%), progresso (50–90%) "
                   "e fim (acima de 90%) do plantio.")
else:
    st.info("Coluna de data de plantio não disponível.")

st.divider()

# ── Marcha de Colheita ──
secao_titulo("Avanço de colheita", "Como está a marcha de colheita?",
             "Evolução semanal da colheita. O número dentro da bola é o total de locais colhidos até aquela semana; o tamanho indica quantos entraram na semana; e a altura, o percentual acumulado.")
if "dataColheitaMilho" in ta_filtrado.columns:
    fig_co, n_co = _marcha(ta_filtrado, "dataColheitaMilho", "colhidos", "rgba(0,157,87,0.3)")
    if fig_co is None:
        st.info("Nenhum local com data de colheita registrada para os filtros selecionados.")
    else:
        st.plotly_chart(fig_co, use_container_width=True)
        _tot_co = ta_filtrado["cod_fazenda"].nunique()
        _falta_co = _tot_co - n_co
        st.caption(f"ℹ️ {n_co} de {_tot_co} locais do recorte têm data de colheita lançada"
                   + (f" ({_falta_co} sem data — veja o Diagnóstico)." if _falta_co > 0 else ".")
                   + " Cada bola é uma semana: o número dentro é o **acumulado de locais** até ali, o tamanho "
                   "reflete quantos entraram naquela semana e o percentual acima é a mesma leitura "
                   "do eixo vertical. As zonas de fundo marcam início (até 50%), progresso (50–90%) "
                   "e fim (acima de 90%) da colheita.")
else:
    st.info("Coluna de data de colheita não disponível.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — DESCRITIVA GERAL DO CONJUNTO
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Conjunto",
    "Como está o experimento como um todo?",
    "Estatísticas descritivas por variável considerando todos os ensaios filtrados.",
)

vars_desc = {
    "kg_ha":                     "kg/ha",
    "sc_ha":                     "sc/ha",
    "umidade_pct":               "Umidade (%)",
    "populacao_real_plantas_ha": "Pop. Real",
    "altura_espiga_m":           "Alt. Espiga (m)",
    "altura_planta_m":           "Alt. Planta (m)",
    "pmg_corrigido_g":           "PMG (g)",
    "fileiras_media":            "Fileiras",
    "graos_fileira_media":       "Grãos/Fileira",
    "graos_ardidos_pct":         "Ardidos (%)",
    "pct_acamadas":              "Acamamento (%)",
    "pct_colmo_podre":           "Colmo Podre (%)",
    "pct_quebradas":             "Quebramento (%)",
    "pct_dominadas":             "Dominadas (%)",
    "pct_perda_total":           "Perda Total (%)",
}

medidas = ["Total de Observações", "Média", "Desvio Padrão", "Mínimo",
           "1º Quartil", "Mediana", "3º Quartil", "Máximo", "CV (%)", "LSD (5%)", "Locais"]

rows_geral = {m: {} for m in medidas}
lsd_kg = calcular_lsd(ta_filtrado)   # calculado uma vez (produtividade)

for col, label in vars_desc.items():
    if col not in ta_filtrado.columns:
        continue
    serie = pd.to_numeric(ta_filtrado[col], errors="coerce").dropna()
    # perdas e ardidos têm 0 legítimo (avaliado, sem ocorrência); as demais descartam 0
    _perdas_ok = ("pct_perda_total", "graos_ardidos_pct",
                  "pct_acamadas", "pct_quebradas", "pct_dominadas", "pct_colmo_podre")
    if col not in _perdas_ok:
        serie = serie[serie > 0]
    if len(serie) == 0:
        for m in medidas:
            rows_geral[m][label] = "—"
        continue

    media = serie.mean()
    dp = serie.std()
    cv = round(dp / media * 100, 1) if media > 0 else np.nan
    q1, q2, q3 = serie.quantile([0.25, 0.50, 0.75])
    if col == "kg_ha":
        lsd = round(lsd_kg, 1) if isinstance(lsd_kg, (int, float)) and not np.isnan(lsd_kg) else "—"
    elif col == "sc_ha":
        lsd = round(lsd_kg / 60, 1) if isinstance(lsd_kg, (int, float)) and not np.isnan(lsd_kg) else "—"
    else:
        lsd = "—"

    # população é inteiro; demais variáveis, 1 casa decimal.
    # devolve TEXTO na população: ao transpor a tabela o pandas promoveria o int a float
    # (58230 viraria 58230.0) por causa das outras variáveis na mesma linha.
    _dec = 0 if col == "populacao_real_plantas_ha" else 1
    def _r(x):
        return f"{int(round(x, 0))}" if _dec == 0 else round(x, 1)

    rows_geral["Total de Observações"][label] = int(len(serie))
    rows_geral["Média"][label]                = _r(media)
    rows_geral["Desvio Padrão"][label]        = _r(dp)
    rows_geral["Mínimo"][label]               = _r(serie.min())
    rows_geral["1º Quartil"][label]           = _r(q1)
    rows_geral["Mediana"][label]              = _r(q2)
    rows_geral["3º Quartil"][label]           = _r(q3)
    rows_geral["Máximo"][label]               = _r(serie.max())
    rows_geral["CV (%)"][label]               = round(cv, 1) if not np.isnan(cv) else "—"
    rows_geral["LSD (5%)"][label]             = lsd
    rows_geral["Locais"][label]               = ta_filtrado["cod_fazenda"].nunique()

df_geral = pd.DataFrame(rows_geral).T.reset_index().rename(columns={"index": "Medida"})
ag_table(df_geral, height=425)
exportar_excel(df_geral, nome_arquivo="descritiva_conjunto.xlsx",
               label="⬇️ Exportar Descritiva Geral", key="exp_geral")
st.caption(
    "ℹ️ **CV (%) desta tabela** = Desvio Padrão ÷ Média × 100, calculado sobre todas as observações "
    "brutas. Inclui a variação entre híbridos, entre locais e o erro experimental — por isso tende "
    "a ser maior. Para avaliar a qualidade do experimento, use o CV da ANOVA da Tabela de "
    "Apresentação, que desconta os efeitos de híbrido e local, retendo apenas o erro residual.")
st.caption(
    "ℹ️ **LSD (5%)** — Se a diferença de produtividade entre dois híbridos for maior que este valor, "
    "ela é real e não fruto do acaso (95% de confiança). Tecnicamente: Diferença Mínima "
    "Significativa = t(α/2, gl_resíduo) × √(2 × QMR / nº de locais), onde QMR é o Quadrado Médio do "
    "Resíduo da ANOVA conjunta (modelo: y = μ + híbrido + local + erro).")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — UMIDADE × PRODUTIVIDADE
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Umidade × Produtividade",
    "Como cada híbrido se posiciona em relação à umidade de colheita?",
    "Compare híbridos pela umidade média de colheita — um indicador de ciclo/precocidade — e "
    "identifique quais se destacam para o seu ponto de colheita.",
)

with st.popover("ℹ️ Como interpretar · Umidade × Produtividade", use_container_width=False):
    st.markdown("""
**📌 O que este gráfico mostra**

Cada ponto é um híbrido. A posição horizontal é a **umidade média de colheita** (%) do híbrido no
conjunto filtrado — um indicador de ciclo: híbridos mais precoces tendem a colher mais secos, os
mais tardios mais úmidos. A posição vertical é a **produtividade média (sc/ha)**.

---

**📐 Como ler**

- **Eixo X → Umidade (%)** — média de colheita do híbrido nos locais filtrados.
- **Eixo Y → sc/ha** — produtividade média no conjunto de locais filtrado.
- **Cor do ponto** → status do híbrido (CHECK, STINE, EXP).
- **Nome** → identificação do híbrido.

---

**💡 O que observar**

- **Tendência ascendente** → híbridos que colhem mais úmidos (mais tardios) produzem mais neste
  conjunto — o ciclo mais longo está sendo aproveitado.
- **Tendência plana ou descendente** → a umidade de colheita não explica a produtividade; outros
  fatores dominam.
- **Pontos acima da tendência** → híbridos que produzem mais que o esperado para a sua umidade —
  candidatos a destaque (bom desempenho independente do ciclo).
- **Pontos abaixo da tendência** → ficaram aquém do esperado para o seu ponto de colheita.
""")

unid_um = st.radio("Unidade", ["sc/ha", "kg/ha"], horizontal=True, key="unid_um")
_cum = "sc_ha" if unid_um == "sc/ha" else "kg_ha"
_dum = 1 if unid_um == "sc/ha" else 0

# ── Cálculo: média de produtividade e de umidade por híbrido (conforme filtros) ──
# ESCALA ABSOLUTA vs RELATIVA AO LOCAL
# Na absoluta, a média de cada híbrido carrega o efeito de ONDE ele foi testado: um material
# avaliado só em locais fracos aparece embaixo mesmo tendo ido bem em todos eles. Como a rede é
# desbalanceada (nem todo híbrido está em todo local), isso distorce a comparação entre pontos.
# Na relativa, cada parcela é comparada com o SEU ensaio antes de virar média — o efeito de local
# sai, e o que sobra é desempenho.
# SÓ A PRODUTIVIDADE é relativizada. A umidade continua em % de colheita, absoluta: é assim que
# ela é lida em campo ("colhe a 18%"), e é o número que decide secagem e logística. Relativizada,
# viraria "1,8 ponto abaixo da referência" — perde o valor que interessa e ainda esbarra em
# "umidade relativa" já significar outra coisa.
# A referência é a DO PRÓPRIO LOCAL (média, maior ou testemunha daquele ensaio), calculada sobre
# todos os híbridos avaliados ali. Difere da Descritiva de propósito: lá a referência é única
# para o conjunto e só renumera o eixo; aqui ela muda a posição de cada ponto, que é o que
# desconta o ambiente.
_esc_c1, _esc_c2, _esc_c3 = st.columns([2, 2, 2])
with _esc_c1:
    # contra QUEM (média/maior/testemunha) é o seletor ao lado
    _esc_um = st.radio(
        "Escala da produtividade", ["Absoluta", "Relativa"], horizontal=True, key="escala_um",
        help="Relativa: a produtividade vira índice contra a referência escolhida ao lado "
             "(100 = referência). A umidade continua em % de colheita nas duas escalas. "
             "Vale só para este gráfico.")
_rel_um = (_esc_um == "Relativa")

with _esc_c2:
    if _rel_um:
        # seletor PRÓPRIO, independente do da Auditoria e do Desempenho: são perguntas
        # diferentes e mudar um não deve mexer nos outros. As bases são as mesmas, para a
        # leitura ser a mesma.
        _base_um_sel = st.selectbox("Relativizar pela produção do local",
                                    ["Média do ensaio", "Maior produtividade", "Testemunha"],
                                    index=0, key="base_um")
    else:
        _base_um_sel = None
with _esc_c3:
    if _rel_um and _base_um_sel == "Testemunha":
        _tests_um = sorted(ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]
                           ["dePara"].dropna().unique().tolist())
        _test_um = st.selectbox("Testemunha", _tests_um, key="test_um") if _tests_um else None
    else:
        _test_um = None
    # só na escala relativa, junto dos demais controles de relativização. Multiselect em vez de
    # checkbox: com os quatro status são oito linhas, e quase sempre a comparação que interessa
    # é de dois grupos (o seu contra os CHECK), não de todos contra todos.
    if _rel_um:
        _st_disp = [s for s in COR_STATUS_PLOT
                    if s in set(ta_filtrado["status_material"].dropna().unique())]
        _status_linhas = st.multiselect(
            "Linhas de média por status", _st_disp, default=_st_disp, key="lin_status_um",
            help="Traça a média do grupo nos dois eixos: a horizontal na produtividade relativa "
                 "e a vertical na umidade média. O cruzamento é o centro de massa do status.")
    else:
        _status_linhas = []

    # DESTACAR HÍBRIDO: mesmas linhas de média das do status, um nível abaixo — a horizontal na
    # produtividade relativa do híbrido e a vertical na umidade dele. Sólidas, para separar de
    # grupo (tracejado) e material (sólido) num relance.
    _hib_detalhe = st.multiselect(
        "Linhas de média por híbrido",
        # opções de `ta_filtrado`: o widget é montado ANTES de `_base_um` existir
        options=sorted(ta_filtrado.loc[ta_filtrado[_cum] > 0, "dePara"]
                       .dropna().unique().tolist()),
        default=[], key="det_hib_um", max_selections=6,
        help="Traça a média do híbrido nos dois eixos, em linha sólida, para comparar com a "
             "média do status ou de outro material.")

_base_um = ta_filtrado[ta_filtrado[_cum] > 0].copy()

if _rel_um:
    # Referência DE CADA LOCAL, como na Auditoria: cada parcela é comparada com o próprio ensaio
    # ANTES de virar média do híbrido. É o que corrige rede desbalanceada — um material avaliado
    # só em locais fracos deixa de aparecer embaixo por causa de onde foi testado. Relativizar
    # por uma referência única do conjunto não faria isso: seria dividir todos os pontos pela
    # mesma constante, ou seja, o mesmo gráfico com outra numeração no eixo.
    # A referência sai do ENSAIO INTEIRO (todos os híbridos daquele local, via ta_raw + filtros
    # de local), não dos filtrados: o filtro de material muda quem aparece, não o 100%.
    if _base_um_sel == "Maior produtividade":
        _ref_um = _ref_scope.groupby(LOCAL)[_cum].max()
    elif _base_um_sel == "Testemunha" and _test_um:
        _ref_um = _ref_scope[_ref_scope["dePara"] == _test_um].groupby(LOCAL)[_cum].mean()
    else:
        _ref_um = _ref_scope.groupby(LOCAL)[_cum].mean()
    _k_um = _base_um.set_index(LOCAL).index
    _base_um["_ref_prod"] = pd.Series(_k_um.map(_ref_um).to_numpy(), index=_base_um.index)
    _base_um["_y_um"] = (_base_um[_cum] / _base_um["_ref_prod"]) * 100
    _base_um["_x_um"] = _base_um["umidade_pct"]      # umidade fica ABSOLUTA (ver acima)

else:
    _base_um["_y_um"] = _base_um[_cum]
    _base_um["_x_um"] = _base_um["umidade_pct"]

df_um = (
    _base_um
    .groupby("dePara")
    .agg(
        media_sc = ("_y_um", "mean"),
        umidade  = ("_x_um", "mean"),
        status   = ("status_material", "first"),
    )
    .reset_index()
    .dropna(subset=["umidade"])
)
_dec_um = 1 if _rel_um else _dum
df_um["media_sc"] = df_um["media_sc"].round(_dec_um)
df_um["umidade"] = df_um["umidade"].round(1)

_alvo_ref = {"Maior produtividade": "ao maior produtor",   # do local, não da rede
             "Testemunha": f"à testemunha ({_test_um})" if _test_um else "à testemunha"
             }.get(_base_um_sel, "à média do ensaio")
_nome_ref = f"{_alvo_ref} de cada local"
_tit_x = "Umidade média de colheita (%)"       # sempre absoluta, nas duas escalas
_tit_y = (f"Produtividade relativa {_nome_ref} (100 = referência)" if _rel_um
          else f"Produtividade média ({unid_um})")

if df_um.empty:
    st.info("Sem dados de umidade para os filtros selecionados.")
else:
    # ONDE AS LINHAS HORIZONTAIS VÃO PASSAR — calculado ANTES dos pontos porque o rótulo de cada
    # ponto precisa saber disso. O nome fica acima do marcador por padrão; quando uma linha corta
    # justamente essa faixa, o texto vai para baixo. Sem isso a linha risca o nome no meio, que é
    # o que acontecia com os materiais de produtividade próxima à média do grupo.
    _linhas_y = []
    if _status_linhas:
        _linhas_y += df_um[df_um["status"].isin(_status_linhas)].groupby("status")["media_sc"] \
                     .mean().tolist()
    if _hib_detalhe:
        _linhas_y += df_um[df_um["dePara"].isin(_hib_detalhe)]["media_sc"].tolist()
    if _rel_um:
        _linhas_y.append(100.0)
    # a banda é a altura do rótulo (~26px) convertida em unidades do eixo: depende da altura do
    # gráfico e da faixa de valores, então é calculada, não chutada em porcentagem
    _ALT_FIG, _PX_ROTULO = 980, 26
    _alt_plot = _ALT_FIG - 60 - (70 + (95 if _status_linhas else 0) + (60 if _hib_detalhe else 0))
    _faixa_y = (df_um["media_sc"].max() - df_um["media_sc"].min()) or 1
    _banda = _faixa_y * (_PX_ROTULO / max(_alt_plot, 200))

    def _pos_rotulo(_yv):
        """'bottom center' quando há linha na faixa onde o nome cairia; senão 'top center'."""
        return ("bottom center"
                if any(_yv < _ly <= _yv + _banda for _ly in _linhas_y) else "top center")

    fig_um = go_plt.Figure()
    for status, cor in COR_STATUS_PLOT.items():
        df_s = df_um[df_um["status"] == status]
        if df_s.empty:
            continue
        fig_um.add_trace(go_plt.Scatter(
            x=df_s["umidade"], y=df_s["media_sc"],
            mode="markers+text", name=status,
            text=df_s["dePara"],
            textposition=[_pos_rotulo(_v) for _v in df_s["media_sc"]],
            textfont=dict(size=13, color="#333333", weight="bold"),
            marker=dict(color=cor, size=14,
                        line=dict(color=COR_BORDA.get(status, "#888"), width=1.5), opacity=0.9),
            hovertemplate=("<b>%{text}</b><br>Umidade: %{x:.1f}%<br>"
                           + ("Produtividade: %{y:.1f}% da referência" if _rel_um else
                              "Média: %{y:,." + str(_dum) + "f} " + unid_um)
                           + "<extra></extra>"),
        ))

    # Linha de tendência (regressão linear simples)
    if len(df_um) >= 3:
        x_tr = df_um["umidade"].values
        y_tr = df_um["media_sc"].values
        z = np.polyfit(x_tr, y_tr, 1)
        p = np.poly1d(z)
        x_line = np.linspace(x_tr.min(), x_tr.max(), 100)
        fig_um.add_trace(go_plt.Scatter(
            x=x_line, y=p(x_line), mode="lines", name="Tendência",
            line=dict(color="#AAAAAA", width=1.5, dash="dash"), hoverinfo="skip",
        ))

    # FAIXA FIXA DOS EIXOS: mantida porque as linhas de média (status e híbrido) são desenhadas
    # como shapes e não entram no autorange do Plotly — sem faixa definida, uma média fora da
    # nuvem de pontos ficaria pela metade, cortada na borda.
    _px = (df_um["umidade"].max() - df_um["umidade"].min()) * 0.06 or 0.5
    _py = (df_um["media_sc"].max() - df_um["media_sc"].min()) * 0.08 or 1.0
    _xs_ref = [df_um["umidade"].min(), df_um["umidade"].max()]
    _ys_ref = [df_um["media_sc"].min(), df_um["media_sc"].max()]
    if _rel_um:
        _ys_ref.append(100)                     # a linha de referência também tem de caber
    _xr = [min(_xs_ref) - _px, max(_xs_ref) + _px]
    _yr = [min(_ys_ref) - _py, max(_ys_ref) + _py]

    fig_um.update_layout(
        # gráfico alto: são 43 pontos com rótulo, e a altura é o que separa materiais de
        # produtividade parecida — a largura já vem do container
        height=_ALT_FIG, plot_bgcolor="#F5F5F5", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=13, color="#111111"),
        xaxis=dict(
            # standoff afasta o título dos rótulos de umidade dos status, que ficam logo abaixo
            # dos ticks — sem ele, os dois se sobrepõem
            title=dict(text=f"<b>{_tit_x}</b>", font=dict(size=14, color="#111111"),
                       standoff=(74 if _status_linhas else 12) + (52 if _hib_detalhe else 0)),
            tickfont=dict(size=12, color="#111111", weight="bold"),
            gridcolor="#FFFFFF", gridwidth=1.5, zeroline=False,
            showline=True, linecolor="#CCCCCC", linewidth=1, range=_xr),
        yaxis=dict(
            title=dict(text=f"<b>{_tit_y}</b>", font=dict(size=14, color="#111111")),
            tickfont=dict(size=12, color="#111111", weight="bold"),
            gridcolor="#FFFFFF", gridwidth=1.5, zeroline=False,
            showline=True, linecolor="#CCCCCC", linewidth=1, range=_yr),
        legend=dict(
            title=dict(text="<b>Status</b>", font=dict(size=13, color="#111111")),
            orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
            font=dict(size=13, color="#111111", weight="bold")),
        # b: rótulos de umidade em duas alturas + título afastado · r: os de produtividade
        margin=dict(t=60,
                    b=70 + (95 if _status_linhas else 0) + (60 if _hib_detalhe else 0),
                    l=75, r=110),
    )

    if _rel_um:
        # os dois eixos ganham origem com significado: 100 = produz como a referência,
        # 0 pp = colhe na mesma umidade dela
        fig_um.add_hline(y=100, line=dict(color="#9AA5B1", width=1.5, dash="dot"))
        fig_um.add_trace(go_plt.Scatter(   # só para entrar na legenda
            x=[None], y=[None], mode="lines", name="referência = 100",
            line=dict(color="#9AA5B1", width=1.5, dash="dot"), hoverinfo="skip"))

    # rótulos da margem direita (status e híbridos): recolhidos numa lista e posicionados de uma
    # vez, mais abaixo — duas médias podem cair a meio ponto uma da outra, e foi o que aconteceu
    # com EXP 99,1 e STINE 98,6 escrevendo um por cima do outro
    _rot_dir = []

    def _fmt_y(_v):
        """Valor do eixo Y no formato da escala ativa — índice com % ou sc/ha (kg/ha)."""
        return f"{_v:.1f}%" if _rel_um else f"{_v:,.{_dum}f} {unid_um}".replace(",", ".")

    # ── médias por status: compara GRUPOS, não pontos ──────────────────────────
    # A horizontal é a média de produtividade do status; a vertical, a de umidade. O cruzamento
    # das duas é o centro de massa daquele grupo — dá para ver de relance se o portfólio STINE
    # está, em média, acima e mais seco que os concorrentes.
    if _status_linhas:
        _med_st = (df_um[df_um["status"].isin(_status_linhas)].groupby("status")
                   .agg(_y=("media_sc", "mean"), _x=("umidade", "mean"), _n=("dePara", "size"))
                   .reset_index())
        for _, _r in _med_st.iterrows():
            _cor_st = COR_STATUS_PLOT.get(_r["status"], "#888888")
            _cor_ln = COR_BORDA.get(_r["status"], _cor_st)   # borda: legível sobre fundo claro
            fig_um.add_hline(y=_r["_y"], line=dict(color=_cor_ln, width=3, dash="dash"))
            fig_um.add_vline(x=_r["_x"], line=dict(color=_cor_ln, width=2, dash="dot"))

        # VALORES NA MARGEM, cada um no eixo a que pertence: produtividade relativa na lateral
        # direita (é leitura do eixo Y) e umidade embaixo (leitura do eixo X). Fora da área de
        # plotagem eles não cobrem ponto nenhum. Como dois status podem ter umidade quase igual,
        # os rótulos de baixo alternam de altura em vez de se sobreporem.
        _ordem_x = _med_st.sort_values("_x").reset_index(drop=True)
        for _i, _r in _ordem_x.iterrows():
            _c = COR_BORDA.get(_r["status"], COR_STATUS_PLOT.get(_r["status"], "#888888"))
            fig_um.add_annotation(
                x=_r["_x"], xref="x", y=0, yref="paper", yanchor="top",
                # abaixo dos NÚMEROS do eixo (que ocupam ~20px), não na mesma linha deles:
                # a -6 o rótulo caía em cima do tick quando a média batia perto de um valor
                # inteiro. Duas alturas alternadas, para status com umidade parecida não colidirem.
                yshift=-46 if _i % 2 else -24, showarrow=False,
                text=f"<b>{_r['status']} {_r['_x']:.1f}%</b>",
                font=dict(size=12, color=_c))
        for _, _r in _med_st.iterrows():
            # cor recalculada aqui: `_c` do laço acima é o do ÚLTIMO status daquele laço, e
            # todos os rótulos sairiam da mesma cor
            _rot_dir.append(dict(
                y=float(_r["_y"]), txt=f"<b>{_r['status']} {_fmt_y(_r['_y'])}</b>",
                cor=COR_BORDA.get(_r["status"], COR_STATUS_PLOT.get(_r["status"], "#888888"))))

    # ── híbridos detalhados: mesmas linhas dos status, SÓLIDAS ─────────────────
    # Mesma leitura das tracejadas de status, um nível abaixo: a horizontal na produtividade
    # média do híbrido e a vertical na umidade média dele. Sólidas para diferenciar do grupo.
    # Os rótulos ficam do lado OPOSTO aos de status — produtividade à esquerda, umidade em cima —
    # para as duas famílias não disputarem a mesma margem.
    if _hib_detalhe:
        _med_hib_det = (df_um[df_um["dePara"].isin(_hib_detalhe)]
                        .set_index("dePara").reindex(_hib_detalhe).dropna(how="all"))
        for _i, (_h, _r) in enumerate(_med_hib_det.iterrows()):
            # cor do STATUS do material, não de uma paleta própria: assim a linha do híbrido e a
            # tracejada do grupo dele são da mesma família, e a comparação que interessa (o
            # material contra a média do próprio status) se lê pela distância entre duas linhas
            # da mesma cor. Sólida contra tracejada é o que separa uma da outra.
            _cor_h = COR_BORDA.get(_r["status"], COR_STATUS_PLOT.get(_r["status"], "#555555"))
            # de ponta a ponta, não só até o ponto: a linha é referência de leitura do gráfico
            # inteiro — dá para conferir quais materiais ficam acima ou abaixo dela
            fig_um.add_hline(y=_r["media_sc"], line=dict(color=_cor_h, width=3), layer="below")
            fig_um.add_vline(x=_r["umidade"], line=dict(color=_cor_h, width=2), layer="below")
            # mesma coluna e mesmo formato dos de status; o espaçamento entre TODOS eles é
            # resolvido de uma vez, mais abaixo
            _rot_dir.append(dict(y=float(_r["media_sc"]),
                                 txt=f"<b>{_h} {_fmt_y(_r['media_sc'])}</b>", cor=_cor_h))
            fig_um.add_annotation(
                x=_r["umidade"], xref="x", y=0, yref="paper", yanchor="top",
                # abaixo dos rótulos de status quando eles estão ligados, senão colidem: as duas
                # famílias marcam umidade e podem cair no mesmo ponto do eixo
                yshift=(-70 if _status_linhas else -24) - (22 if _i % 2 else 0),
                showarrow=False,
                text=f"<b>{_h} {_r['umidade']:.1f}%</b>",
                font=dict(size=12, color=_cor_h))

    # AFASTAMENTO MÍNIMO: ordena por altura e empurra para cima quem ficaria colado no anterior.
    # A LINHA continua no valor real; só o rótulo se desloca — e como o valor vai escrito nele,
    # a leitura não se perde.
    if _rot_dir:
        _rot_dir.sort(key=lambda d: d["y"])
        _sep, _ult = _banda * 1.15, None
        for _d in _rot_dir:
            _yrot = _d["y"] if _ult is None else max(_d["y"], _ult + _sep)
            _ult = _yrot
            fig_um.add_annotation(
                x=1.005, xref="paper", xanchor="left", y=_yrot, yref="y",
                showarrow=False, text=_d["txt"], font=dict(size=12, color=_d["cor"]))

        # margem direita pelo MAIOR rótulo: "DKB360PRO3 98.6%" precisa de mais espaço que
        # "EXP 99.1%", e com a margem fixa o nome comprido saía cortado na borda
        _n_car = max(len(_d["txt"].replace("<b>", "").replace("</b>", "")) for _d in _rot_dir)
        fig_um.update_layout(margin_r=max(110, int(_n_car * 7.2) + 24))

    st.plotly_chart(fig_um, use_container_width=True)
    if _rel_um:
        st.caption(
            f"ℹ️ Eixo Y = produtividade **relativa {_nome_ref}** (100 = referência). "
            f"Eixo X = umidade de colheita **absoluta**, como sempre. "
            f"**Acima de 100 e à esquerda é o canto confortável**: rende mais que a referência e "
            f"colhe mais seco. Tracejado cinza = tendência"
            + ("; tracejado colorido = média do status (produtividade na **margem direita**, "
               "umidade **abaixo do eixo**)" if _status_linhas else "")
            + ("; linha sólida = média do híbrido selecionado, na cor do status dele "
               "(produtividade na **margem direita**, umidade **abaixo do eixo**)."
               if _hib_detalhe else ".")
            + " Como cada parcela é comparada com o próprio ensaio, as posições **mudam** em "
              "relação à escala absoluta: sai o efeito de o híbrido ter sido testado em locais "
              "melhores ou piores. Por isso estes valores não batem com a coluna Prod. Relativa "
              "da Descritiva, que usa referência única do conjunto — as perguntas são outras.")
    else:
        st.caption(
            f"ℹ️ Cada ponto = média do híbrido no conjunto filtrado, em **{unid_um}** (produtividade e "
            "umidade de colheita). "
            "Linha tracejada = tendência linear entre umidade e produtividade. Pontos acima da linha "
            "produzem mais que o esperado para a sua umidade de colheita. "
            "⚠️ Como a rede é desbalanceada, parte da diferença entre pontos é **onde** cada híbrido "
            "foi testado. Troque a escala para **Relativa** para descontar esse efeito: cada "
            "parcela passa a ser comparada com o próprio ensaio antes de virar média.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3B — REMOVIDA (umidade × produtividade dentro do local)
#
# Cruzava umidade e produtividade dos híbridos de UM ensaio, com reta de tendência. Saiu porque
# a reta não sustentava a leitura que induzia ("acima da linha produziu mais que o esperado
# para o ciclo"):
#   · a correlação era ENTRE genótipos, não dentro de um: 35 materiais com potenciais
#     diferentes, então o que a reta capturava era "os mais tardios deste ensaio renderam
#     mais", e não um efeito de ciclo sobre produtividade;
#   · sem repetição, cada ponto é UMA parcela — na faixa de 18 a 19% de umidade a dispersão
#     vertical chegava a 40 sc/ha, boa parte dela erro experimental.
#
# A pergunta "quem seca antes e ainda produz" continua respondida pela SEÇÃO 3C (Secagem
# relativa), que regride com a rede inteira: dezenas de pontos por híbrido, coeficiente
# estimável e efeito de local descontado.
# ════════════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3C — SECAGEM RELATIVA (umidade por local)
#
# Por que por LOCAL: a data de plantio e a de colheita são da fazenda, iguais para
# todos os híbridos do ensaio. Então a diferença de umidade DENTRO de um local é
# diferença de ciclo, limpa de ambiente. Comparar a umidade média de dois híbridos
# na rede inteira mistura ciclo com onde cada um foi testado; aqui não.
#
# NÃO é dry down: dry down exige medir a mesma parcela em datas sucessivas. Aqui a
# variação vem de locais diferentes — serve para ordenar híbridos, não para estimar
# perda de umidade por dia.
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Secagem relativa",
    "Quais híbridos acompanham o ambiente na hora de secar, e quais não?",
    "Em cada local todos os híbridos foram plantados e colhidos no mesmo dia — então a diferença "
    "de umidade entre eles ali é diferença de ciclo, sem efeito de ambiente. Esta seção mede como "
    "cada híbrido responde a locais que colhem mais secos ou mais úmidos.",
)

with st.popover("ℹ️ Como interpretar · Secagem relativa", use_container_width=False):
    st.markdown("""
**📌 A pergunta**

> **Existe híbrido que chega ao ponto de colheita mais cedo? E ele se comporta assim em todo
> lugar, ou depende do ano e do local?**

Todo produtor conhece a primeira parte: uns materiais colhem mais secos que outros. O que quase
ninguém tem é a segunda — saber se essa vantagem se repete ou se some quando o ano muda.

---

**🌽 O ponto de partida, sem estatística**

Num ensaio, **todo mundo é plantado no mesmo dia e colhido no mesmo dia**. A colhedora passa uma
vez só. Então, se um híbrido sai da lavoura com 18% de umidade e o vizinho com 22%, essa
diferença não é do solo, nem da chuva, nem da data — é do próprio material. Ele fecha o ciclo
antes.

É por isso que esta seção olha **local por local** e não pela média da rede. Na média, um híbrido
pode parecer mais seco só porque foi testado em lugares que colheram cedo.

---

**🔢 As três contas, uma de cada vez**

**1. Umidade do híbrido no local.** Média das parcelas daquele material naquele ensaio. Descarta
as leituras fora da faixa de 10% a 40%, que o próprio sistema já marca como suspeitas.

**2. Umidade do local.** Média de **todos** os híbridos do ensaio. É o "quão úmido esse ensaio
colheu no geral" — o retrato do ambiente. Um local que colheu a 25% de média foi um local úmido,
por chuva, por altitude, por atraso na colheita, não importa a razão.

**3. A reta de cada híbrido.** Junta-se um ponto por local: de um lado a umidade do ensaio, do
outro a umidade daquele material ali. Traça-se a reta que melhor passa por esses pontos. Precisa
de pelo menos 3 locais.

---

**📋 O que é cada coluna da tabela**

| Coluna | Em linguagem simples | Como sai da conta |
|---|---|---|
| **Híbrido** | o material. A cor do nome é o tipo: azul-escuro Stine, laranja concorrente, verde experimental | — |
| **Locais** | em quantos ensaios ele entrou nesta conta | nº de locais com umidade válida. **Vermelho** = base curta, menos da metade do mais testado |
| **b (secagem)** | **o quanto ele acompanha o ambiente.** Se o ensaio veio 2 pontos mais úmido que o normal, um híbrido com b = 1 também vem 2 pontos mais úmido; com b = 1,2 vem 2,4; com b = 0,8 vem só 1,6 | inclinação da reta |
| **R²** | **o quanto dá para confiar na reta.** Perto de 1, os pontos estão quase em cima dela. Perto de 0, o material faz o que quer | de 0 a 1. **Vermelho** abaixo de 0,70 |
| **Precocidade (pp)** | **quantos pontos de umidade ele fica abaixo ou acima da média do ensaio.** −2,0 quer dizer: onde o ensaio colheu a 20%, ele colheu a 18% | média da diferença entre a umidade dele e a do local. "pp" = pontos percentuais |
| **Produtividade (sc/ha)** | quanto ele produziu, em média, nesses mesmos locais | **negrito** = acima da mediana do recorte |

---

**🎨 A cor, em uma frase**

**Azul = mais seco e mais estável. Laranja = mais úmido e mais dependente do ambiente.**

Não existe verde nem vermelho de aprovação nesta tabela, de propósito. Colher mais seco não é
"certo": é uma característica que resolve a vida de um produtor com janela apertada e é
indiferente para quem tem secador na fazenda.

---

**📐 Juntando b e precocidade**

São duas perguntas diferentes e vale ler as duas:

- **Precocidade** responde *"ele colhe seco?"* — é a altura da reta.
- **b** responde *"ele colhe seco em qualquer ano?"* — é a inclinação.

| Combinação | O que significa na prática |
|---|---|
| precocidade negativa **e** b abaixo de 1 | colhe seco e continua colhendo seco mesmo em ano úmido. É o mais previsível para programar colheita |
| precocidade negativa **e** b acima de 1 | colhe seco em ano bom, mas em ano úmido perde a vantagem |
| precocidade positiva **e** b abaixo de 1 | é mais tardio, mas de forma constante — dá para planejar |
| precocidade positiva **e** b acima de 1 | mais tardio e imprevisível: em ano úmido, fica muito úmido |

---

**🧭 Como usar — um caso**

O produtor tem uma colhedora só e quer começar cedo para liberar área.

1. Na tabela, procure **precocidade negativa** — ele colhe abaixo da média do ensaio.
2. Dentro desses, prefira **b abaixo de 1**: a vantagem não some no ano ruim.
3. Cheque o **R²**. Se estiver em vermelho, os pontos estão espalhados e a reta não descreve nada
   — não use esse material como argumento de secagem.
4. Cheque **Locais**. Em vermelho, foram poucos ensaios; o número pode mudar na próxima safra.
5. Por último, olhe a **produtividade**. Secar cedo só vale se não custar saca. Se o mais seco
   for também o menos produtivo, a conversa vira uma escolha entre logística e rendimento — e
   isso o produtor decide, não a tabela.

---

**⚠️ O que esta seção NÃO é**

- **Não é dry down.** Dry down é acompanhar a mesma lavoura ao longo dos dias e medir a perda de
  umidade. Aqui não existe medição repetida no tempo: a variação vem de locais diferentes. Serve
  para **ordenar** híbridos, não para dizer quantos pontos de umidade se perdem por dia.
- **Não mede ciclo em dias.** Data de plantio e de colheita são do ensaio, iguais para todos os
  materiais — ninguém foi colhido antes. O que se mede é o estado em que cada um chegou no dia
  comum.
- **A inclinação absoluta carrega efeito de região e de clima.** O que se compara com segurança é
  o b **entre híbridos**, porque todos passaram pelos mesmos locais.
- **O índice do local muda com o filtro.** Ao comparar duas leituras, confirme que o recorte é o
  mesmo.
""")

# ── Cálculo ────────────────────────────────────────────────────────────────────
_MIN_LOC_SEC = 3

_base_sec = ta_filtrado.copy()
if "umidade_pct" in _base_sec.columns:
    _base_sec = _base_sec[pd.to_numeric(_base_sec["umidade_pct"], errors="coerce").notna()]
    # descarta leitura fora da faixa sã (o pipeline já sinaliza)
    if "flags_produtividade" in _base_sec.columns:
        _fl = _base_sec["flags_produtividade"].astype(str)
        _base_sec = _base_sec[~_fl.str.contains("umidade_baixa|umidade_alta", na=False)]
else:
    _base_sec = _base_sec.iloc[0:0]

if _base_sec.empty:
    st.info("Sem dados de umidade para os filtros selecionados.")
else:
    _hl = (_base_sec.groupby(["dePara", "cod_fazenda"], as_index=False)
           .agg(umid=("umidade_pct", "mean"),
                prod=("sc_ha", "mean"),
                status=("status_material", "first")))
    _idx_loc = _hl.groupby("cod_fazenda")["umid"].mean().rename("idx")
    _hl = _hl.join(_idx_loc, on="cod_fazenda")

    _n_loc_hib = _hl.groupby("dePara")["cod_fazenda"].nunique()
    _elegiveis = sorted(_n_loc_hib[_n_loc_hib >= _MIN_LOC_SEC].index)

    if len(_elegiveis) < 2 or _hl["idx"].nunique() < 3:
        st.info(f"Poucos locais para estimar secagem relativa — são necessários ao menos "
                f"{_MIN_LOC_SEC} locais por híbrido e 3 locais no recorte.")
    else:
        # coeficientes por híbrido
        _coef = []
        for _h in _elegiveis:
            _g = _hl[_hl["dePara"] == _h]
            _x, _y = _g["idx"].values, _g["umid"].values
            if len(_g) < _MIN_LOC_SEC or np.ptp(_x) == 0:
                continue
            _X = np.column_stack([np.ones(len(_y)), _x])
            (_a, _b), *_ = np.linalg.lstsq(_X, _y, rcond=None)
            _pred = _a + _b * _x
            _ss_res = float(((_y - _pred) ** 2).sum())
            _ss_tot = float(((_y - _y.mean()) ** 2).sum())
            _coef.append(dict(
                dePara=_h, status=_g["status"].iloc[0], b=float(_b), a=float(_a),
                r2=(1 - _ss_res / _ss_tot) if _ss_tot > 0 else np.nan,
                precoc=float((_g["umid"] - _g["idx"]).mean()),
                prod=float(_g["prod"].mean()), n=int(_g["cod_fazenda"].nunique())))
        df_sec = pd.DataFrame(_coef)

        if df_sec.empty:
            st.info("Não foi possível estimar as retas com os filtros atuais.")
        else:
            _stine_def = df_sec[df_sec["status"] == "STINE"]["dePara"].tolist()
            _default = _stine_def[:5] or df_sec.sort_values("prod", ascending=False)["dePara"].head(5).tolist()
            sel_sec = st.multiselect(
                "Híbridos em destaque (os demais ficam em cinza, como contexto):",
                options=sorted(df_sec["dePara"].tolist()), default=_default, key="sel_secagem",
                max_selections=6)
            if not sel_sec:
                st.info("Selecione ao menos um híbrido.")
            else:
                _PAL_SEC = ["#9B59B6", "#E91E63", "#00BCD4", "#795548", "#FF5722", "#673AB7"]
                _cor_de = {h: _PAL_SEC[i % len(_PAL_SEC)] for i, h in enumerate(sel_sec)}

                # Um gráfico EMBAIXO do outro, não lado a lado: com a legenda na lateral
                # (padrão do Índice Ambiental) cada um precisa da largura inteira — em duas
                # colunas, a legenda comeria metade da área de plotagem.
                # ── gráfico 1: retas de secagem ────────────────────────────────
                with st.container():
                    fig_sec = go_plt.Figure()
                    _xs = np.linspace(_hl["idx"].min(), _hl["idx"].max(), 50)

                    # contexto: todos os demais híbridos em cinza claro
                    _outros = _hl[~_hl["dePara"].isin(sel_sec)]
                    if not _outros.empty:
                        fig_sec.add_trace(go_plt.Scatter(
                            x=_outros["idx"], y=_outros["umid"], mode="markers",
                            name="demais híbridos", legendgroup="ctx",
                            marker=dict(color="#D5D8DC", size=6, opacity=0.55),
                            hovertemplate="%{customdata}<br>local %{x:.1f}% · híbrido %{y:.1f}%<extra></extra>",
                            customdata=_outros["dePara"]))

                    # diagonal b=1
                    fig_sec.add_trace(go_plt.Scatter(
                        x=_xs, y=_xs, mode="lines", name="b = 1 (acompanha o ambiente)",
                        line=dict(color="#9AA5B1", width=1.5, dash="dash"), hoverinfo="skip"))

                    for _h in sel_sec:
                        _g = _hl[_hl["dePara"] == _h]
                        _r = df_sec[df_sec["dePara"] == _h].iloc[0]
                        _cor = _cor_de[_h]
                        # como no Índice Ambiental: quem entra na legenda é a RETA, não a nuvem
                        # de pontos — assim o traço colorido da legenda corresponde ao que o
                        # olho procura no gráfico
                        fig_sec.add_trace(go_plt.Scatter(
                            x=_g["idx"], y=_g["umid"], mode="markers", name=_h,
                            legendgroup=_h, showlegend=False,
                            marker=dict(color=_cor, size=7, opacity=0.6,
                                        line=dict(color="#FFFFFF", width=0.8)),
                            hovertemplate=(f"<b>{_h}</b><br>local: %{{x:.1f}}%<br>"
                                           "híbrido: %{y:.1f}%<extra></extra>")))
                        fig_sec.add_trace(go_plt.Scatter(
                            x=_xs, y=_r["a"] + _r["b"] * _xs, mode="lines", name=_h,
                            legendgroup=_h, showlegend=True,
                            line=dict(color=_cor, width=2),
                            hovertemplate=(f"<b>{_h}</b><br>b = {_r['b']:.2f} · "
                                           f"R² = {_r['r2']:.2f}<extra></extra>")))

                    # mesmo padrão visual do Índice Ambiental (fundo branco, grade #E5E5E5,
                    # legenda vertical à direita): as três leituras de adaptabilidade da página
                    # passam a ser lidas do mesmo jeito
                    fig_sec.update_layout(
                        height=520, margin=dict(t=40, b=70, l=75, r=160),
                        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                        font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
                        legend=dict(orientation="v", x=1.01, y=1, xanchor="left",
                                    font=dict(size=13, color="#111111"), itemsizing="constant"),
                        xaxis=dict(title=dict(text="Umidade média do local (%)",
                                              font=dict(size=16, color="#000000", weight="bold")),
                                   tickfont=dict(size=14, color="#000000", weight="bold"), gridcolor="#E5E5E5"),
                        yaxis=dict(title=dict(text="Umidade do híbrido (%)",
                                              font=dict(size=16, color="#000000", weight="bold")),
                                   tickfont=dict(size=14, color="#000000", weight="bold"), gridcolor="#E5E5E5"))
                    st.plotly_chart(fig_sec, use_container_width=True)
                    st.caption("ℹ️ Cada ponto é um híbrido em um local. Cinza = demais híbridos do "
                               "recorte, como contexto. Tracejado = diagonal b = 1.")

                st.write("")      # respiro entre os dois gráficos empilhados

                # ── gráfico 2: onde cada híbrido cai ──────────────────────────
                # Trocado o dot plot de duas colunas por um plano: com 35 híbridos
                # aquilo virava uma lista de pontos sem eixo legível. Aqui b e
                # produtividade são os dois eixos, os quadrantes têm nome em
                # português e só os destacados são rotulados.
                with st.container():
                    _med_prod = float(df_sec["prod"].median())
                    _dst = df_sec[df_sec["dePara"].isin(sel_sec)]
                    _ctx = df_sec[~df_sec["dePara"].isin(sel_sec)]

                    fig_co = go_plt.Figure()
                    fig_co.add_vline(x=1, line=dict(color="#9AA5B1", width=1.5, dash="dash"))
                    fig_co.add_hline(y=_med_prod, line=dict(color="#9AA5B1", width=1.5, dash="dash"))

                    if not _ctx.empty:
                        fig_co.add_trace(go_plt.Scatter(
                            x=_ctx["b"], y=_ctx["prod"], mode="markers",
                            name="demais híbridos", legendgroup="ctx", showlegend=True,
                            marker=dict(color="#D5D8DC", size=9,
                                        line=dict(color="#FFFFFF", width=1)),
                            customdata=np.stack([_ctx["dePara"], _ctx["r2"], _ctx["n"]], axis=-1),
                            hovertemplate=("<b>%{customdata[0]}</b><br>b = %{x:.2f} · "
                                           "R² = %{customdata[1]:.2f}<br>%{y:.1f} sc/ha · "
                                           "%{customdata[2]} locais<extra></extra>")))
                    # UM TRACE POR HÍBRIDO (não um só com vetor de cores): é o que permite a
                    # mesma legenda vertical do gráfico de cima. `legendgroup` repete o do gráfico
                    # 1, então a ordem e as cores dos dois batem.
                    for _h in sel_sec:
                        _p = _dst[_dst["dePara"] == _h]
                        if _p.empty:
                            continue
                        fig_co.add_trace(go_plt.Scatter(
                            x=_p["b"], y=_p["prod"], mode="markers", name=_h,
                            legendgroup=_h, showlegend=True,
                            marker=dict(color=_cor_de[_h], size=15,
                                        line=dict(color="#FFFFFF", width=1.5)),
                            customdata=np.stack([_p["r2"], _p["n"], _p["precoc"]], axis=-1),
                            hovertemplate=(f"<b>{_h}</b><br>b = %{{x:.2f}} · "
                                           "R² = %{customdata[0]:.2f}<br>%{y:.1f} sc/ha · "
                                           "%{customdata[1]} locais<br>"
                                           "precocidade %{customdata[2]:+.1f} pp<extra></extra>")))

                    # nomes dos quadrantes, em português
                    _xmin, _xmax = df_sec["b"].min(), df_sec["b"].max()
                    _ymin, _ymax = df_sec["prod"].min(), df_sec["prod"].max()
                    _pad_x, _pad_y = (_xmax - _xmin) * 0.06, (_ymax - _ymin) * 0.08
                    for _tx, _ty, _txt, _anc in [
                        (_xmin - _pad_x / 2, _ymax + _pad_y / 2, "produz e é previsível", "left"),
                        (_xmax + _pad_x / 2, _ymax + _pad_y / 2, "produz, mas depende da janela", "right"),
                    ]:
                        fig_co.add_annotation(x=_tx, y=_ty, text=f"<b>{_txt}</b>", showarrow=False,
                                              xanchor=_anc, font=dict(size=12, color="#6B7280"))
                    fig_co.add_annotation(
                        x=1, y=_ymin - _pad_y, text="acompanha o ambiente", showarrow=False,
                        yanchor="top", font=dict(size=11.5, color="#6B7280"))

                    fig_co.update_layout(
                        title=dict(text="<b>Quem seca junto com o ambiente, e quem produz</b>",
                                   font=dict(size=15, color="#111111"), x=0, xanchor="left"),
                        height=520, margin=dict(t=80, b=80, l=75, r=160),
                        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                        font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
                        legend=dict(orientation="v", x=1.01, y=1, xanchor="left",
                                    font=dict(size=13, color="#111111"), itemsizing="constant"),
                        xaxis=dict(
                            title=dict(text="← umidade mais estável &nbsp;&nbsp;|&nbsp;&nbsp; "
                                            "amplifica o ambiente →",
                                       font=dict(size=16, color="#000000", weight="bold")),
                            tickfont=dict(size=14, color="#000000", weight="bold"), gridcolor="#E5E5E5",
                            range=[_xmin - _pad_x * 2, _xmax + _pad_x * 2]),
                        yaxis=dict(
                            title=dict(text="Produtividade média (sc/ha)",
                                       font=dict(size=16, color="#000000", weight="bold")),
                            tickfont=dict(size=14, color="#000000", weight="bold"), gridcolor="#E5E5E5",
                            range=[_ymin - _pad_y * 2.2, _ymax + _pad_y * 2]))
                    st.plotly_chart(fig_co, use_container_width=True)
                    st.caption(
                        f"ℹ️ Cada ponto é um híbrido, identificado pela cor na legenda ao lado. "
                        f"Em cinza, os que não estão em destaque — o nome aparece ao passar o "
                        f"mouse. A vertical marca **b = 1**, o híbrido "
                        f"que varia igual ao ambiente; a horizontal marca a **mediana de "
                        f"produtividade do recorte ({_med_prod:.1f} sc/ha)**. O canto superior "
                        f"esquerdo é a posição mais confortável: produz acima da mediana e a "
                        f"umidade não depende do ambiente.")

                # ── tabela de apoio ────────────────────────────────────────────
                # Em HTML, não em AgGrid: a mesma matriz de células desenha a tela e
                # gera o Excel, então o arquivo sai idêntico ao que se vê. Com AgGrid
                # a tela vinha de cellStyle em JS e o arquivo de openpyxl — dois
                # caminhos que divergem sem avisar (e divergiam: o Excel saía branco).
                _med_prod_tab = float(df_sec["prod"].median())
                _n_ref = int(df_sec["n"].max())

                _c_ord, _ = st.columns([2, 3])
                with _c_ord:
                    _ordem_tab = st.selectbox(
                        "Ordenar por",
                        ["Secagem (b), do mais estável ao mais dependente",
                         "Precocidade, do mais seco ao mais úmido",
                         "Produtividade, da maior para a menor",
                         "Nome do híbrido"],
                        key="ord_tab_secagem")
                _chave = {"Secagem (b), do mais estável ao mais dependente": ("b", True),
                          "Precocidade, do mais seco ao mais úmido": ("precoc", True),
                          "Produtividade, da maior para a menor": ("prod", False),
                          "Nome do híbrido": ("dePara", True)}[_ordem_tab]
                _dt = df_sec.sort_values(_chave[0], ascending=_chave[1])

                # Uma linguagem de cor só: AZUL = mais seco / mais estável ·
                # LARANJA = mais úmido / mais dependente. Verde e vermelho de
                # aprovação ficaram de fora — colher mais seco não é "certo", é uma
                # característica. Vermelho fica reservado ao que é aviso de fragilidade.
                _ESC_B = [(0.85, "#B7D7EF"), (0.97, "#DCEBF7"), (1.03, "#F0F1F3"),
                          (1.15, "#FBE2CC"), (99, "#F4B184")]
                _ESC_P = [(-1.5, "#B7D7EF"), (-0.3, "#DCEBF7"), (0.3, "#F0F1F3"),
                          (1.5, "#FBE2CC"), (99, "#F4B184")]

                def _bg(v, escala):
                    for _lim, _c in escala:
                        if v <= _lim:
                            return _c
                    return escala[-1][1]

                _COR_ST = {"CHECK": "#C46A3A", "STINE": "#1A4F7A",
                           "EXP": "#009900", "DP2": "#7AAF6A"}
                _headers = ["Híbrido", "Locais", "b (secagem)", "R²",
                            "Precocidade (pp)", "Produtividade (sc/ha)"]
                _linhas = []
                for _r in _dt.itertuples():
                    _linhas.append([
                        cel(_r.dePara, cor=_COR_ST.get(_r.status, "#1A1A1A"), bold=True),
                        cel(_r.n, "num0", align="center",
                            cor="#C0201E" if _r.n < _n_ref * 0.5 else "#6B7280",
                            bold=_r.n < _n_ref * 0.5),
                        cel(f"{_r.b:.2f}".replace(".", ","), align="center", bold=True,
                            bg=_bg(_r.b, _ESC_B)),
                        cel(f"{_r.r2:.2f}".replace(".", ","), align="center",
                            cor="#C0201E" if _r.r2 < 0.70 else "#6B7280",
                            bold=_r.r2 < 0.70),
                        cel(_r.precoc, "sinal1", align="center", bold=True,
                            bg=_bg(_r.precoc, _ESC_P)),
                        cel(_r.prod, "num1", align="center",
                            bold=_r.prod > _med_prod_tab,
                            cor="#1A1A1A" if _r.prod > _med_prod_tab else "#6B7280"),
                    ])

                st.markdown("###### Todos os híbridos")
                render_tabela(
                    _headers, _linhas, "secagem_relativa", "exp_secagem",
                    largura_1a=200, altura_max=560,
                    legenda=[("#B7D7EF", "seca mais e varia menos que o ambiente", "bg"),
                             ("#F0F1F3", "acompanha o ambiente", "bg"),
                             ("#F4B184", "fica mais úmido e depende mais do ambiente", "bg"),
                             ("#C0201E", "número frágil — reta fraca ou poucos locais", "txt"),
                             ("#1A4F7A", "STINE", "txt"), ("#C46A3A", "concorrente", "txt"),
                             ("#009900", "experimental", "txt")])
                st.caption(
                    "ℹ️ **Azul = mais seco e mais estável · laranja = mais úmido e mais "
                    "dependente do ambiente.** Não há verde nem vermelho de aprovação: colher "
                    "mais seco não é melhor nem pior, é uma característica que serve para um "
                    "produtor e não serve para outro. "
                    f"**R² em vermelho** abaixo de 0,70 — a reta explica pouco. **Locais em "
                    f"vermelho** — menos da metade da cobertura do híbrido mais testado "
                    f"({_n_ref} locais). **Produtividade em negrito** — acima da mediana do "
                    f"recorte ({_med_prod_tab:.1f} sc/ha). Não é dry down: ver o modal da seção.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — DESEMPENHO POR HÍBRIDO (agregação no conjunto)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Desempenho",
    "Como cada híbrido se comportou no conjunto?",
    "Estatísticas agregadas por híbrido sobre todos os locais filtrados. Avalie consistência, "
    "variabilidade e desempenho relativo entre locais.",
)

with st.popover("ℹ️ Como interpretar · Legenda", use_container_width=False):
    st.markdown("""
**📌 Como interpretar esta tabela**

Cada linha é um híbrido, com estatísticas calculadas sobre **todos os locais nos filtros ativos**.
Os híbridos são ordenados por **Média (kg/ha) decrescente**. Por ser uma visão do conjunto, a
**Produção Relativa aqui é sobre a base geral** (não por local como na Auditoria) — responde "como
o híbrido se posiciona no conjunto todo".

---

**📋 Glossário das colunas**
""")
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("""
| Coluna | Descrição |
|---|---|
| **Status** | Categoria do híbrido (CHECK, STINE, EXP) |
| **Umidade (%)** | Umidade média de colheita — indicador de ciclo |
| **Locais** | Nº de locais onde o híbrido foi avaliado |
| **N** | Nº de observações válidas (kg/ha > 0) |
| **Média (kg/ha)** | Média de produtividade entre os locais |
| **Prod. Relativa (%)** | Média do híbrido ÷ base escolhida × 100 (sobre o conjunto) |

**⚠️ Duas produções relativas na página — não confunda:**
- **Aqui (Desempenho)** o método é **por material**: a média do híbrido é comparada com a
  referência do **conjunto de materiais filtrados**. Responde "entre os materiais em análise,
  quanto cada um rende". O filtro de material muda a referência — é o grupo de comparação.
- **Na Auditoria** o método é **por local**: cada plot é comparado com a referência **do seu
  próprio local** (todos os híbridos daquele local, referência fixa). Responde "como foi neste local".

Em ambas, o filtro de **local** define o ensaio. A diferença é o filtro de **material**: na
Auditoria ele não entra na referência (é por local); aqui ele define o grupo comparado. Por isso
os números diferem entre as seções — as perguntas são diferentes.
""")
    with col2:
        st.markdown("""
| Coluna | Descrição |
|---|---|
| **DP** | Desvio Padrão (kg/ha) — dispersão absoluta entre locais |
| **CV (%)** | Coeficiente de Variação — dispersão relativa (DP ÷ Média × 100) |
| **Mínimo / Máximo** | Menor e maior valor entre os locais |
| **Q1 / Mediana / Q3** | Faixa: Q1 = piso (75% dos locais acima), Mediana = central, Q3 = teto |
""")
    st.markdown("""
---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana podem contar histórias diferentes sobre o mesmo híbrido.** Quando próximas,
o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes puxam a média.

- **Média:** soma das produtividades ÷ nº de locais. Pode ser puxada por uma área excepcional ou
  muito ruim — nem sempre representa o comportamento típico.
- **Mediana (Q2):** o valor "do meio" com os resultados em ordem — metade das áreas produz menos,
  metade mais. Não sofre distorção por áreas extremas.
- **Q1 (primeiro quartil):** separa os 25% piores dos 75% restantes. Um Q1 alto significa que mesmo
  nos piores ambientes o híbrido entrega uma produtividade razoável — segurança na recomendação.
- **Q3 (terceiro quartil):** separa os 75% melhores dos 25% mais altos.

---

**🎯 O que observar**

- **CV baixo (< 15%)** → híbrido consistente entre locais — menor risco de recomendação.
- **CV alto (> 25%)** → adaptação específica — bom em alguns locais, ruim em outros.
- **DP baixo** → pouca variação absoluta entre locais — comportamento previsível.
- **Mediana próxima da média** → distribuição equilibrada, sem locais discrepantes.
- **Q1 alto** → piso de produtividade elevado — híbrido seguro mesmo nos piores ambientes.
""")

# ── Seletor PRÓPRIO da base de produção relativa (independente da Auditoria) ──
# O Desempenho tem seu próprio critério: você pode ver a Auditoria por média geral e esta
# tabela por testemunha ao mesmo tempo. Keys distintas (base_rel_desemp) para não colidir.
_cbd, _ctd, _ = st.columns([2, 2, 3])
with _cbd:
    base_rel_desemp = st.selectbox(
        "Base da Produção Relativa (desta tabela)",
        options=["Média geral dos materiais", "Maior produtividade", "Testemunha"],
        index=0, key="base_rel_desemp")
with _ctd:
    if base_rel_desemp == "Testemunha":
        _tests_d = sorted(_scope_desemp[_scope_desemp["status_material"].isin(["CHECK", "STINE"])]
                          ["dePara"].dropna().unique().tolist()) \
            if "status_material" in _scope_desemp.columns else []
        testemunha_desemp = st.selectbox("Selecione a testemunha", options=_tests_d,
                                         key="testemunha_desemp") if _tests_d else None
    else:
        testemunha_desemp = None

# ── Referência da produção relativa (visão de conjunto, POR MATERIAL) ─────────
# A referência é o CONJUNTO DE MATERIAIS que aparece: numa análise conjunta a comparação é entre
# os materiais selecionados, então usa ta_filtrado (já com os filtros de local e de material).
# Trocar o filtro de material MUDA a referência de propósito — é o grupo de comparação. Difere da
# Auditoria, que é por local (referência do ensaio inteiro, sem o filtro de material).
# A referência é a MÉDIA DOS VALORES DA COLUNA — isto é, a média das médias por híbrido (cada
# material pesa igual, coerente com a tabela ser por material). Assim a linha "Média Geral" fecha
# em 100 na base padrão. Média por híbrido em kg/ha (independe da unidade de exibição):
_scope_desemp = ta_filtrado[pd.to_numeric(ta_filtrado["kg_ha"], errors="coerce") > 0]
_medias_hib_desemp = _scope_desemp.groupby("dePara")["kg_ha"].mean()
if base_rel_desemp == "Maior produtividade":
    ref_global = _medias_hib_desemp.max()
elif base_rel_desemp == "Testemunha" and testemunha_desemp:
    _mt = _medias_hib_desemp.get(testemunha_desemp, np.nan)
    ref_global = _mt if not pd.isna(_mt) else np.nan
else:  # "Média geral dos materiais" (padrão) — média das médias por híbrido
    ref_global = _medias_hib_desemp.mean()

unid_desemp = st.radio("Unidade", ["sc/ha", "kg/ha"], horizontal=True, key="unid_desemp")
_col_desemp = "sc_ha" if unid_desemp == "sc/ha" else "kg_ha"
_lbl_media = f"Média ({unid_desemp})"

desc_rows = []
for hibrido, grp in ta_filtrado.groupby("dePara", dropna=True):
    kg = pd.to_numeric(grp[_col_desemp], errors="coerce").dropna()
    kg = kg[kg > 0]
    if len(kg) == 0:
        continue
    q1, q2, q3 = kg.quantile([0.25, 0.50, 0.75])
    media = kg.mean()
    dp = kg.std()
    cv = (dp / media * 100) if media > 0 else np.nan
    umid = pd.to_numeric(grp["umidade_pct"], errors="coerce")
    umid = umid[umid > 0]
    _ref_u = ref_global if _col_desemp == "kg_ha" else (ref_global / 60)
    prod_rel = ((media / _ref_u) * 100) if _ref_u else np.nan
    desc_rows.append({
        "Híbrido":            hibrido,
        "Status":             grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else "",
        "Umidade (%)":        round(umid.mean(), 1) if len(umid) > 0 else "—",
        "Locais":             grp["cod_fazenda"].nunique(),
        "N":                  int(len(kg)),
        _lbl_media:           round(media, 1),
        "Prod. Relativa (%)": round(prod_rel, 1) if not np.isnan(prod_rel) else np.nan,
        "DP":                 round(dp, 1) if not np.isnan(dp) else "—",
        "CV (%)":             round(cv, 1) if not np.isnan(cv) else "—",
        "Mínimo":             round(kg.min(), 1),
        "Q1":                 round(q1, 1),
        "Mediana":            round(q2, 1),
        "Q3":                 round(q3, 1),
        "Máximo":             round(kg.max(), 1),
    })

if not desc_rows:
    st.info("Nenhum híbrido com produtividade válida para os filtros selecionados. "
            "Pode ser que a colheita ainda não tenha sido lançada neste recorte.")
else:
    df_desc = pd.DataFrame(desc_rows).sort_values(_lbl_media, ascending=False).reset_index(drop=True)
    ag_table(df_desc, height=min(600, 36 + 32 * len(df_desc) + 20))
    exportar_excel(df_desc, nome_arquivo="desempenho_hibridos.xlsx",
                   label="⬇️ Exportar Desempenho", key="exp_desc")
    st.caption(f"ℹ️ Todas as medidas (média, DP, mínimo, quartis e máximo) estão em **{unid_desemp}**. "
               "A Produção Relativa não muda com a unidade, por ser uma razão.")
    st.info(f"📐 **Produção Relativa — método POR MATERIAL (visão de conjunto).** A média de cada "
            f"híbrido é comparada com {_descreve_base(base_rel_desemp, testemunha_desemp, escopo='material')}. "
            "É diferente da Auditoria (que é por local): aqui a pergunta é quanto o material rende "
            "na média do grupo em análise. Mudar o filtro de híbrido muda o grupo de comparação de "
            "propósito. (O filtro de local já define o ensaio.)")

st.divider()

# ── Base de apresentação: média das métricas por híbrido (ranking + apresentação) ──
# ordem da tabela de apresentação: sc/ha, umidade, (prod. rel. inserida depois), pop,
# alturas, acamamento/quebramento/dominadas, PMG, ardidos
_metricas_apres = {
    "sc_ha":                     "sc/ha",
    "kg_ha":                     "kg/ha",          # oculta na tabela; usada só p/ ordenar e prod. rel.
    "umidade_pct":               "Umidade (%)",
    "populacao_real_plantas_ha": "Pop. Real",
    "altura_planta_m":           "Alt. Planta (m)",
    "altura_espiga_m":           "Alt. Espiga (m)",
    "pct_acamadas":              "Acamamento (%)",
    "pct_colmo_podre":           "Colmo Podre (%)",
    "pct_quebradas":             "Quebramento (%)",
    "pct_dominadas":             "Dominadas (%)",
    "pmg_corrigido_g":           "PMG (g)",
    "graos_ardidos_pct":         "Ardidos (%)",
}
_apres_rows = []
for _hib, _grp in ta_filtrado.groupby("dePara", dropna=True):
    _row = {"Híbrido": _hib,
            "status_material": _grp["status_material"].mode()[0] if not _grp["status_material"].mode().empty else ""}
    for _col, _label in _metricas_apres.items():
        if _col not in _grp.columns:
            _row[_label] = None
            continue
        _serie = pd.to_numeric(_grp[_col], errors="coerce").dropna()
        # perdas e ardidos têm 0 legítimo (avaliado, sem ocorrência); as demais descartam 0
        if _col not in ("pct_perda_total", "graos_ardidos_pct",
                        "pct_acamadas", "pct_quebradas", "pct_dominadas", "pct_colmo_podre"):
            _serie = _serie[_serie > 0]
        if len(_serie) == 0:
            _row[_label] = None
        elif _col == "populacao_real_plantas_ha":
            _row[_label] = int(round(_serie.mean(), 0))
        else:
            _row[_label] = round(_serie.mean(), 1)
    _apres_rows.append(_row)

df_apres = pd.DataFrame(_apres_rows)
# descarta híbridos sem produtividade (sc/ha None) — evita quebra no ranking/apresentação
if not df_apres.empty and "sc/ha" in df_apres.columns:
    df_apres = df_apres[df_apres["sc/ha"].notna()].copy()
if not df_apres.empty:
    df_apres = df_apres.sort_values("sc/ha", ascending=False).reset_index(drop=True)

# LSD pré-calculado (usado no ranking; em sacas)
_lsd_apres = calcular_lsd(ta_filtrado, col="kg_ha")
lsd_sc = round(_lsd_apres / 60, 2) if isinstance(_lsd_apres, (int, float)) and not np.isnan(_lsd_apres) else None

# ── Contexto: filtros ativos abaixo da pergunta do ranking ────────────────────
_all_safras = sorted(ta_raw["safra"].dropna().unique().tolist())
_all_macros = sorted(ta_raw["regiao_macro"].dropna().unique().tolist()) if "regiao_macro" in ta_raw.columns else []
_all_micros = sorted(ta_raw["regiao_micro"].dropna().unique().tolist()) if "regiao_micro" in ta_raw.columns else []
_all_estados = sorted(ta_raw["estado_sigla"].dropna().unique().tolist()) if "estado_sigla" in ta_raw.columns else []
_all_cidades = sorted(ta_raw["cidade_nome"].dropna().unique().tolist()) if "cidade_nome" in ta_raw.columns else []

def _linha_safra(s, g):
    """Uma linha de contexto por safra, com as regiões/estados que ela realmente contém."""
    partes = [f"<b>Safra:</b> {s}"]
    if "regiao_macro" in g.columns and g["regiao_macro"].notna().any():
        partes.append("Macro: " + ", ".join(sorted(g["regiao_macro"].dropna().unique())))
    if "regiao_micro" in g.columns and g["regiao_micro"].notna().any():
        partes.append("Micro: " + ", ".join(sorted(g["regiao_micro"].dropna().unique())))
    if "estado_sigla" in g.columns and g["estado_sigla"].notna().any():
        partes.append(", ".join(sorted(g["estado_sigla"].dropna().unique())))
    partes.append(f"{g['cidade_nome'].nunique()} cidades")
    partes.append(f"{g['cod_fazenda'].nunique()} locais")
    return " · ".join(partes)


_grupos_safra = list(ta_filtrado.groupby("safra"))
_linhas_safra = [_linha_safra(s, g) for s, g in _grupos_safra]
if len(_grupos_safra) > 1:
    _linhas_safra = ["<b>Análise multissafra</b>"] + _linhas_safra
n_fazendas_ctx = ta_filtrado["cod_fazenda"].nunique()
n_cidades_ctx = ta_filtrado["cidade_nome"].nunique()
contexto_str = "<br>".join(_linhas_safra)

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5 — RANKING (LOLLIPOP)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo("Visualização", "Quem lidera o ranking de produtividade?", contexto_str)

with st.popover("ℹ️ Como interpretar · Ranking", use_container_width=False):
    st.markdown("""
**📌 Como ler este gráfico**

Cada ponto é a **média de sc/ha** de um híbrido em todos os locais filtrados. Os híbridos são
ordenados do mais produtivo (topo) ao menos produtivo (base).

---

**🔲 Elementos do gráfico**
""")
    c1, c2, c3 = st.columns(3)
    c1.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Ponto colorido</b><br>Média sc/ha do híbrido<br>(cor = status)</div>', unsafe_allow_html=True)
    c2.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha tracejada cinza</b><br>Média geral do conjunto<br>(referência visual)</div>', unsafe_allow_html=True)
    c3.markdown('<div style="background:#fff;border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha vermelha pontilhada</b><br>Corte pelo LSD (5%)<br>separa grupos estatísticos</div>', unsafe_allow_html=True)
    st.markdown("""
---

**🎨 Legenda de cores**
""")
    c1, c2, c3 = st.columns(3)
    c1.markdown('<div style="background:#00FF00;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>EXP</b><br>híbridos em avaliação</div>', unsafe_allow_html=True)
    c2.markdown('<div style="background:#F4B184;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>CHECK</b><br>testemunhas externas</div>', unsafe_allow_html=True)
    c3.markdown('<div style="background:#2976B6;color:#fff;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>STINE</b><br>híbridos comerciais Stine</div>', unsafe_allow_html=True)
    st.markdown("""
---

**📊 Interpretação prática**

- Híbridos **acima da linha vermelha** não diferem estatisticamente do melhor — **candidatos ao avanço**.
- Híbridos **entre linhas vermelhas** formam um grupo intermediário.
- Híbridos **abaixo da última linha vermelha** são estatisticamente inferiores ao grupo de elite.
- Quando **não há linha vermelha**, não há diferença estatística significativa entre os híbridos.
""")

if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    unid_lol = st.radio("Unidade", ["sc/ha", "kg/ha"], horizontal=True, key="unid_lol")
    _cu = unid_lol                       # nome da coluna em df_apres coincide com o rótulo
    _lsd_u = lsd_sc if unid_lol == "sc/ha" else (
        round(_lsd_apres, 1) if isinstance(_lsd_apres, (int, float)) and not np.isnan(_lsd_apres) else None)
    _dec = 1 if unid_lol == "sc/ha" else 0

    df_plot = df_apres[["Híbrido", _cu, "status_material"]].dropna(subset=[_cu]).copy()
    df_plot = df_plot.sort_values(_cu, ascending=False).reset_index(drop=True)

    n_obs = (ta_filtrado[ta_filtrado["sc_ha"] > 0].groupby("dePara")["sc_ha"].count().rename("n_obs"))
    df_plot = df_plot.merge(n_obs, left_on="Híbrido", right_index=True, how="left")
    df_plot["n_obs"] = df_plot["n_obs"].fillna(0).astype(int)

    media_plot = df_plot[_cu].mean()

    fig = go_plt.Figure()
    for _, row in df_plot.iterrows():
        fig.add_shape(type="line", x0=0, x1=row[_cu],
                      y0=row["Híbrido"], y1=row["Híbrido"],
                      line=dict(color="#DDDDDD", width=1.5))

    cores_pt = [COR_STATUS_PLOT.get(s, "#AAAAAA") for s in df_plot["status_material"]]
    bordas_pt = [COR_BORDA.get(s, "#888888") for s in df_plot["status_material"]]
    fig.add_trace(go_plt.Scatter(
        x=df_plot[_cu], y=df_plot["Híbrido"], mode="markers+text", name="", showlegend=False,
        marker=dict(color=cores_pt, size=16, line=dict(color=bordas_pt, width=1.5)),
        text=[f"  {v:,.{_dec}f} ({n})".replace(",", ".") for v, n in zip(df_plot[_cu], df_plot["n_obs"])],
        textposition="middle right", textfont=dict(size=13, color="#000000"),
        hovertemplate="<b>%{y}</b><br>" + unid_lol + ": %{x:,." + str(_dec) + "f}<extra></extra>",
    ))
    for status, cor in COR_STATUS_PLOT.items():
        if status in df_plot["status_material"].values:
            fig.add_trace(go_plt.Scatter(
                x=[None], y=[None], mode="markers", name=status,
                marker=dict(color=cor, size=12, line=dict(color=COR_BORDA.get(status, "#888"), width=1.5))))

    col_chk1, col_chk2 = st.columns(2)
    mostrar_lsd_lol = col_chk1.checkbox("Mostrar linhas de corte LSD", value=True, key="chk_lsd_lol")
    mostrar_media_lol = col_chk2.checkbox("Mostrar linha de média", value=True, key="chk_media_lol")

    if mostrar_media_lol:
        fig.add_vline(x=media_plot, line=dict(color="#888888", width=1.5, dash="dash"),
                      annotation_text=f"Média: {media_plot:,.{_dec}f}".replace(",", "."),
                      annotation_position="top",
                      annotation_font=dict(size=13, color="#333333", weight="bold"))

    # Linhas de corte LSD: separam grupos que diferem do líder por mais que o LSD
    cultivares_asc = df_plot["Híbrido"].tolist()
    if _lsd_u and mostrar_lsd_lol:
        lider_p2 = df_apres[_cu].max()
        sc_desc2 = df_apres.sort_values(_cu, ascending=False)[_cu].tolist()
        cult_desc2 = df_apres.sort_values(_cu, ascending=False)["Híbrido"].tolist()
        for i in range(1, len(sc_desc2)):
            v = sc_desc2[i]
            if v is not None and (lider_p2 - v) > _lsd_u:
                y_corte = cult_desc2[i - 1]
                idx_y = cultivares_asc.index(y_corte) if y_corte in cultivares_asc else None
                if idx_y is not None and idx_y > 0:
                    fig.add_hline(y=idx_y - 0.5, line=dict(color="#FF0000", width=2.5, dash="dot"))
                    fig.add_annotation(x=0.02, xref="paper", y=idx_y - 0.5, yref="y",
                                       text=f"LSD: {_lsd_u:,.{_dec}f}".replace(",", "."),
                                       showarrow=False,
                                       xanchor="left", yanchor="bottom",
                                       font=dict(size=12, color="#FF0000", weight="bold"))
                lider_p2 = v

    x_max = df_plot[_cu].max()
    x_range_max = round(x_max * 1.18, 1)
    altura_fig = max(400, len(df_plot) * 28 + 80)
    fig.update_layout(
        height=altura_fig, margin=dict(l=180, r=40, t=40, b=60),
        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=15, color="#000000"),
        xaxis=dict(title=dict(text=unid_lol, font=dict(size=16, color="#000000", weight="bold")),
                   showgrid=False, zeroline=False, showline=False,
                   tickfont=dict(size=14, color="#000000", weight="bold"), range=[0, x_range_max]),
        yaxis=dict(showgrid=True, gridcolor="#EEEEEE", gridwidth=1, zeroline=False, showline=False,
                   tickfont=dict(size=14, color="#000000", weight="bold"), categoryorder="array",
                   categoryarray=df_plot["Híbrido"].tolist()[::-1], showticklabels=False),
        legend=dict(title=dict(text="Status", font=dict(size=14, color="#000000", weight="bold")),
                    orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                    font=dict(size=14, color="#000000", weight="bold")),
        showlegend=True,
    )
    fig.update_traces(textfont=dict(size=14, color="#000000", weight="bold"), selector=dict(mode="markers+text"))

    COR_STATUS_TEXTO = {"CHECK": "#C46A3A", "STINE": "#2976B6", "EXP": "#009900", "DP2": "#5A8A4A"}
    for hib in df_plot["Híbrido"].tolist():
        status = df_plot[df_plot["Híbrido"] == hib]["status_material"].iloc[0]
        cor = COR_STATUS_TEXTO.get(status, "#333333")
        fig.add_annotation(x=0, xref="paper", y=hib, yref="y", text=f"<b>{hib}</b>",
                           showarrow=False, xanchor="right", yanchor="middle",
                           font=dict(size=13, color=cor, weight="bold"))

    st.plotly_chart(fig, use_container_width=True)
    st.caption(
        f"ℹ️ **LSD (5%)** — valores em **{unid_lol}**. Se a diferença de produtividade entre dois "
        "híbridos for maior que este "
        "valor, ela é real e não fruto do acaso (95% de confiança). A linha vermelha pontilhada marca "
        "onde começa essa diferença significativa em relação ao melhor híbrido do grupo. Tecnicamente: "
        "t(α/2, gl_resíduo) × √(2 × QMR / nº de locais).")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — TABELA DE APRESENTAÇÃO (com agrupamento LSD e exportação formatada)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Apresentação",
    "Quais híbridos estão no grupo de elite?",
    "Médias por híbrido com agrupamento estatístico pelo LSD. É a tabela usada para montar a "
    "apresentação — a exportação em Excel sai neste formato, com as cores e os cortes.",
)

with st.popover("ℹ️ Como interpretar · Legenda", use_container_width=False):
    st.markdown("""
**📌 Como interpretar esta tabela**

Cada linha é a **média** de um híbrido em todos os locais filtrados, ordenada por **sc/ha
decrescente**. A **linha vermelha** separa grupos estatisticamente distintos pelo **LSD (5%)**:
híbridos **acima da linha** não diferem do melhor — são os **candidatos ao avanço**.

---

**🎨 Legenda de cores**
""")
    c1, c2, c3 = st.columns(3)
    c1.markdown('<div style="background:#00FF00;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>EXP</b><br>híbridos em avaliação</div>', unsafe_allow_html=True)
    c2.markdown('<div style="background:#F4B184;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>CHECK</b><br>testemunhas externas</div>', unsafe_allow_html=True)
    c3.markdown('<div style="background:#2976B6;color:#fff;padding:6px 10px;border-radius:4px;text-align:center;font-size:13px;"><b>STINE</b><br>híbridos comerciais Stine</div>', unsafe_allow_html=True)
    st.markdown("""
---

**📊 Produção Relativa (%)** — método **POR MATERIAL** (visão de conjunto), o mesmo do
Desempenho: a média do híbrido é comparada com a referência do **conjunto de materiais filtrados**
(média, maior produtividade ou testemunha, conforme o seletor **desta tabela**). Acima de 100% =
produziu acima da referência do grupo.

Numa análise conjunta a comparação é entre os materiais selecionados, então trocar o filtro de
híbrido muda a referência — de propósito, é o grupo de comparação. (O filtro de local já define
o ensaio.)

**⚠️ Não confunda com a Auditoria.** Lá a produção relativa é **por local** (cada plot vs. todos
os híbridos do próprio local, referência fixa). Aqui é **por material** (o híbrido vs. o grupo em
análise). Perguntas diferentes, números diferentes — de propósito. Use **Mostrar kg/ha** para
exibir a coluna em quilos.
""")

if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    df_ap = df_apres.sort_values("sc/ha", ascending=False).reset_index(drop=True)

    # ── Seletor PRÓPRIO da base (independente da Auditoria e do Desempenho) ──
    # Apresentação é visão de CONJUNTO, como o Desempenho: produção relativa POR MATERIAL, referência
    # referência sobre os materiais filtrados (grupo de comparação da análise conjunta).
    _cba, _cta, _ = st.columns([2, 2, 3])
    with _cba:
        base_rel_apres = st.selectbox(
            "Base da Produção Relativa (desta tabela)",
            options=["Média geral dos materiais", "Maior produtividade", "Testemunha"],
            index=0, key="base_rel_apres")
    with _cta:
        if base_rel_apres == "Testemunha":
            _tests_a = sorted(_aplicar_filtros_local(ta_raw)
                              .pipe(lambda d: d[d["status_material"].isin(["CHECK", "STINE"])])
                              ["dePara"].dropna().unique().tolist()) \
                if "status_material" in ta_raw.columns else []
            testemunha_apres = st.selectbox("Selecione a testemunha", options=_tests_a,
                                            key="testemunha_apres") if _tests_a else None
        else:
            testemunha_apres = None

    # referência = MÉDIA DOS VALORES DA COLUNA kg/ha do df_ap (cada linha já é a média de um
    # híbrido). Assim cada material pesa igual e a linha "Média Geral" fecha em 100 na base padrão.
    # É a mesma lógica por-material do Desempenho, mas lida direto da coluna exibida.
    _kg_col = pd.to_numeric(df_ap.get("kg/ha"), errors="coerce").dropna() \
        if "kg/ha" in df_ap.columns else pd.Series([], dtype=float)
    if base_rel_apres == "Maior produtividade":
        ref_apres = _kg_col.max() if len(_kg_col) else np.nan
    elif base_rel_apres == "Testemunha" and testemunha_apres:
        _row_t = df_ap[df_ap["Híbrido"] == testemunha_apres]
        ref_apres = pd.to_numeric(_row_t.get("kg/ha"), errors="coerce").dropna().mean() \
            if len(_row_t) else np.nan
    else:  # média das médias por híbrido (a média da coluna)
        ref_apres = _kg_col.mean() if len(_kg_col) else np.nan

    if "kg/ha" in df_ap.columns and ref_apres and not np.isnan(ref_apres):
        df_ap.insert(df_ap.columns.get_loc("sc/ha") + 1, "Prod. Rel. (%)",
                     (pd.to_numeric(df_ap["kg/ha"], errors="coerce") / ref_apres * 100).round(1))
    else:
        df_ap.insert(df_ap.columns.get_loc("sc/ha") + 1, "Prod. Rel. (%)", np.nan)

    st.info(f"📐 **Produção Relativa — método POR MATERIAL (visão de conjunto)**, igual ao Desempenho. "
            f"A média de cada híbrido é comparada com {_descreve_base(base_rel_apres, testemunha_apres, escopo='material')}. "
            "Diferente da Auditoria (por local): a referência é o grupo de materiais filtrados, "
            "então o filtro de híbrido muda o denominador de propósito. (O filtro de local já "
            "define o ensaio.)")

    # CV da ANOVA (desconta híbrido e local — o CV correto da precisão do experimento)
    try:
        d_anova = ta_filtrado[["kg_ha", "dePara", "cod_fazenda"]].dropna().copy()
        d_anova = d_anova[d_anova["kg_ha"] > 0].reset_index(drop=True)
        y_a = d_anova["kg_ha"].values.astype(float)
        grand_mean = y_a.mean()
        X_c = pd.get_dummies(d_anova["dePara"], drop_first=True).values.astype(float)
        X_l = pd.get_dummies(d_anova["cod_fazenda"], drop_first=True).values.astype(float)
        X_a = np.hstack([np.ones((len(y_a), 1)), X_c, X_l])
        beta_a, _, rank_a, _ = np.linalg.lstsq(X_a, y_a, rcond=None)
        ss_res_a = np.sum((y_a - X_a @ beta_a) ** 2)
        gl_res_a = len(y_a) - rank_a
        qmr_a = ss_res_a / gl_res_a if gl_res_a > 0 else np.nan
        cv_anova = round(np.sqrt(qmr_a) / grand_mean * 100, 1) if not np.isnan(qmr_a) else "—"
    except Exception:
        cv_anova = "—"

    # Linhas de corte entre grupos (LSD)
    linhas_corte = set()
    if lsd_sc is not None:
        sc_vals = df_ap["sc/ha"].tolist()
        lider = sc_vals[0]
        for i in range(1, len(sc_vals)):
            v = sc_vals[i]
            if v is None:
                continue
            if (lider - v) > lsd_sc:
                linhas_corte.add(i - 1)
                lider = v

    COR_STATUS = {"CHECK": "#F4B184", "STINE": "#2976B6", "EXP": "#00FF00", "DP2": "#C4DFB4"}
    COR_TEXTO = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A", "DP2": "#1A1A1A", "": "#000000"}

    # seletor: mostrar/ocultar a coluna kg/ha (visível por padrão)
    mostrar_kg = st.checkbox("Mostrar kg/ha", value=True, key="chk_kg_apres")

    # ordem: Híbrido, [kg/ha], sc/ha, Prod. Rel., Umidade, ... (kg/ha antes de sc/ha quando visível)
    _demais = [c for c in df_ap.columns
               if c not in ("Híbrido", "status_material", "kg/ha", "sc/ha")]
    cols_show = ["Híbrido"]
    if mostrar_kg and "kg/ha" in df_ap.columns:
        cols_show.append("kg/ha")
    cols_show.append("sc/ha")
    cols_show += _demais

    # Rodapé: média geral por coluna
    medias_rodape = {"Híbrido": "Média Geral", "status_material": ""}
    for c in cols_show[1:]:
        vals = pd.to_numeric(df_ap[c], errors="coerce").dropna() if c in df_ap.columns else pd.Series([], dtype=float)
        if c == "Pop. Real":
            medias_rodape[c] = int(round(vals.mean(), 0)) if len(vals) > 0 else "—"
        else:
            # Média Geral = média SIMPLES da coluna, igual para todas (inclusive Prod. Rel.).
            # Consistência acima de tudo: a linha é a média do que aparece na tabela, verificável
            # à mão. A Prod. Rel. pode ficar um pouco acima/abaixo de 100 — é a média das razões
            # dos híbridos (N desigual), não um 100 "arredondado". Correto e transparente.
            medias_rodape[c] = round(vals.mean(), 1) if len(vals) > 0 else "—"

    # ── Tabela HTML colorida ──────────────────────────────────────────────────
    html = """
<style>
.tb-apres { width:100%; border-collapse:collapse; font-size:15px; font-family:'Helvetica Neue',sans-serif; }
.tb-apres th { background:#F2F2F2; color:#000 !important; padding:8px 10px; text-align:center;
    border:1px solid #ccc; white-space:nowrap; font-weight:700; font-size:15px; }
.tb-apres th:first-child { text-align:left; }
.tb-apres td { padding:7px 10px; border:1px solid #ddd; text-align:center; white-space:nowrap; font-size:15px; }
.tb-apres td:first-child { text-align:left; font-weight:500; }
.tb-apres td[data-fg="white"], .tb-apres td[data-fg="white"] * { color:#FFF !important; }
.tb-apres td[data-fg="dark"], .tb-apres td[data-fg="dark"] * { color:#1A1A1A !important; }
.tb-apres tr.corte td { border-bottom:10px solid #FF0000 !important; }
.tb-apres tr.rodape td { background:#D9D9D9 !important; font-weight:700; border-top:2px solid #888; color:#000 !important; }
.tb-apres tr.rodape-info td { font-weight:700 !important; border-top:none; color:#000 !important; }
.tb-apres tr.rodape-info td:first-child { background:#D9D9D9 !important; border:1px solid #ddd; }
.tb-apres tr.rodape-info td:nth-child(2) { background:#D9D9D9 !important; border:1px solid #ddd; }
.tb-apres tr.rodape-info td:nth-child(n+3) { background:#FFF !important; border:none; }
</style>
<table class="tb-apres"><thead><tr>"""
    for c in cols_show:
        html += f"<th>{c}</th>"
    html += "</tr></thead><tbody>"

    # colunas que devem exibir como inteiro (sem .0)
    COLS_INT = {"Pop. Real"}

    def _fmt_cel(c, val):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return "—"
        if c in COLS_INT:
            try:
                return f"{int(round(float(val), 0))}"
            except (ValueError, TypeError):
                return val
        return val

    for i, row in df_ap.iterrows():
        status = row.get("status_material", "")
        bg = COR_STATUS.get(status, "#FFFFFF")
        fg = COR_TEXTO.get(status, "#000000")
        data_fg = "white" if fg == "#FFFFFF" else "dark"
        classe = "corte" if i in linhas_corte else ""
        html += f'<tr class="{classe}">'
        for c in cols_show:
            val = _fmt_cel(c, row.get(c, ""))
            html += f'<td data-fg="{data_fg}" style="background:{bg};">{val}</td>'
        html += "</tr>"

    html += '<tr class="rodape">'
    for c in cols_show:
        html += f"<td>{_fmt_cel(c, medias_rodape.get(c, ''))}</td>"
    html += "</tr>"

    lsd_fmt = round(lsd_sc, 2) if lsd_sc else "—"
    n_locais = ta_filtrado["cod_fazenda"].nunique()

    def _rodape_info_row(label, valor, n_cols):
        cells = f'<td style="background:#D9D9D9;border:1px solid #ddd;font-weight:700;color:#000 !important;">{label}</td>'
        cells += f'<td style="background:#D9D9D9;border:1px solid #ddd;font-weight:700;text-align:left;color:#000 !important;">{valor}</td>'
        cells += f'<td colspan="{n_cols-2}" style="background:#FFF;border:none;"></td>'
        return f'<tr class="rodape-info">{cells}</tr>'

    html += _rodape_info_row("CV (%)", f"{cv_anova}%", len(cols_show))
    html += _rodape_info_row("LSD sc/ha (5%)", f"{lsd_fmt}", len(cols_show))
    html += _rodape_info_row("Locais", n_locais, len(cols_show))
    html += "</tbody></table>"

    import streamlit.components.v1 as components
    iframe_height = 60 + (len(df_ap) + 5) * 38
    components.html(html, height=iframe_height, scrolling=True)
    st.caption(
        "ℹ️ **CV (%) desta tabela** = √QMR ÷ Média Geral × 100, onde QMR é o Quadrado Médio do "
        "Resíduo da ANOVA conjunta (modelo: y = μ + híbrido + local + erro). Desconta os efeitos de "
        "híbrido e local, restando apenas o erro experimental — é o indicador correto da precisão.")
    st.caption(
        "ℹ️ **LSD sc/ha (5%)** — Se a diferença de produtividade entre dois híbridos for maior que "
        "este valor, ela é real e não fruto do acaso (95% de confiança). A linha vermelha marca onde "
        "começa essa diferença em relação ao melhor híbrido.")
    st.caption(
        "ℹ️ **Linha Média Geral** — é a média simples de cada coluna (a média do que aparece na "
        "tabela). A Produção Relativa da linha pode ficar pouco acima ou abaixo de 100%: é a média "
        "das razões dos híbridos, que não fecha em 100 exato quando eles têm nº de locais diferente. "
        "Não é erro — é a média fiel da coluna.")

    # ── Exportação Excel com a mesma formatação (cores + cortes) ──────────────
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("⬇️ Exportar Excel com formatação", type="primary", key="btn_exp_apres"):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Apresentação"
        thin = Side(style="thin", color="CCCCCC")
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ci, col in enumerate(cols_show, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = Font(bold=True, color="1A1A1A", name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color="F2F2F2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = max(12, len(col) + 2)
        ws.row_dimensions[1].height = 28

        for ri, row in df_ap.iterrows():
            status = row.get("status_material", "")
            bg_hex = COR_STATUS.get(status, "#FFFFFF").replace("#", "")
            fg_hex = COR_TEXTO.get(status, "#1A1A1A").replace("#", "")
            is_corte = ri in linhas_corte
            for ci, col in enumerate(cols_show, 1):
                val = row.get(col, None)
                if isinstance(val, float) and np.isnan(val):
                    val = None
                elif col in COLS_INT and val is not None:
                    try:
                        val = int(round(float(val), 0))   # população como inteiro, sem .0
                    except (ValueError, TypeError):
                        pass
                cell = ws.cell(row=ri + 2, column=ci, value=val)
                cell.font = Font(name="Arial", size=10, color=fg_hex)
                cell.fill = PatternFill("solid", start_color=bg_hex)
                cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                b_bottom = Side(style="thick", color="FF0000") if is_corte else thin
                cell.border = Border(left=thin, right=thin, top=thin, bottom=b_bottom)

        n_data = len(df_ap)
        rodape_rows = [
            ("Média Geral", {c: medias_rodape.get(c, "") for c in cols_show}),
            ("CV", {cols_show[0]: "CV (%)", cols_show[1]: f"{cv_anova}%"}),
            ("LSD", {cols_show[0]: "LSD sc/ha (5%)", cols_show[1]: f"{lsd_fmt}"}),
            ("Locais", {cols_show[0]: "Locais", cols_show[1]: n_locais}),
        ]
        for rj, (_, rdata) in enumerate(rodape_rows):
            for ci, col in enumerate(cols_show, 1):
                _v = rdata.get(col, None)
                if col in COLS_INT and isinstance(_v, (int, float)) and not (isinstance(_v, float) and np.isnan(_v)):
                    _v = int(round(float(_v), 0))
                cell = ws.cell(row=n_data + 2 + rj, column=ci, value=_v)
                cell.font = Font(name="Arial", size=10, bold=(rj == 0))
                cell.fill = PatternFill("solid", start_color="F0F0F0")
                cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                cell.border = BORDER

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(label="📥 Baixar Excel", data=buf, file_name="tabela_apresentacao_milho.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.divider()


# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9 — DISTRIBUIÇÃO POR HÍBRIDO (range + IQR + média + pontos por local)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo("Visualização", "Como é a distribuição de produtividade por híbrido?", contexto_str)

with st.popover("ℹ️ Como interpretar · Distribuição", use_container_width=False):
    st.markdown("""
**📌 Como ler este gráfico**

Cada barra representa a distribuição de produtividade (sc/ha) de um híbrido considerando todos os
locais nos filtros ativos.

---

**🔲 Elementos da barra**
""")
    c1, c2, c3 = st.columns(3)
    c1.markdown('<div style="background:rgba(93,174,139,0.12);border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Barra clara</b><br>Range total<br>(mín → máx)</div>', unsafe_allow_html=True)
    c2.markdown('<div style="background:rgba(93,174,139,0.45);border:1px solid #ccc;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Barra escura</b><br>Onde a maioria dos locais se concentra<br><span style="font-size:11px;color:#555;">(intervalo interquartil Q1→Q3, 50% dos dados)</span></div>', unsafe_allow_html=True)
    c3.markdown('<div style="background:#fff;border:2px solid #333;padding:8px;border-radius:4px;text-align:center;font-size:13px;"><b>Linha preta</b><br>Média<br>com valor e nº de observações</div>', unsafe_allow_html=True)
    st.markdown("""
---

**⚫ Pontos individuais**

Cada ponto é um local (fazenda). Passe o mouse para ver o **código do local** e o valor de sc/ha.
Pontos muito afastados da barra indicam ambientes atípicos — muito favoráveis ou muito adversos.

> **n** = número de observações com produtividade válida (sc/ha > 0) do híbrido no conjunto filtrado.

---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana podem contar histórias diferentes sobre o mesmo híbrido.** Quando próximas,
o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes puxam a média.

- **Média:** soma das produtividades ÷ nº de locais. Pode ser puxada por uma área excepcional ou
  muito ruim — nem sempre representa o comportamento típico.
- **Mediana (Q2):** o valor "do meio" com os resultados em ordem — metade das áreas produz menos,
  metade mais. Não sofre distorção por áreas extremas.
- **Q1 (primeiro quartil):** separa os 25% piores dos 75% restantes. Um Q1 alto significa que mesmo
  nos piores ambientes o híbrido entrega produtividade razoável — segurança na recomendação.
  *(início da barra escura)*
- **Q3 (terceiro quartil):** separa os 75% melhores dos 25% mais altos. *(fim da barra escura)*

---

**📊 Interpretação prática**

- **Barra curta + média alta** → híbrido consistente e produtivo, ideal para recomendação ampla.
- **Barra longa** → alta variabilidade entre locais, pode ter adaptação específica.
- **Média alta mas barra escura larga** → desempenho instável, maior risco na recomendação.
""")

if df_apres.empty:
    st.warning("⚠️ Nenhum dado para exibir.")
else:
    unid_box = st.radio("Unidade", ["sc/ha", "kg/ha"], horizontal=True, key="unid_box")
    _cbox = "sc_ha" if unid_box == "sc/ha" else "kg_ha"
    _dbox = 1 if unid_box == "sc/ha" else 0

    df_box = ta_filtrado[["dePara", _cbox, "status_material", "cod_fazenda"]].copy()
    df_box = df_box.rename(columns={"dePara": "Híbrido", _cbox: "valor"})
    df_box["valor"] = pd.to_numeric(df_box["valor"], errors="coerce")
    df_box = df_box[df_box["valor"] > 0].dropna(subset=["valor"])

    if df_box.empty:
        st.info("Nenhum híbrido com produtividade válida para os filtros selecionados.")
    else:
        ordem_box = (df_box.groupby("Híbrido")["valor"].mean()
                     .sort_values(ascending=False).index.tolist())

        COR_BOX_ESC = {"CHECK": "rgba(244,177,132,0.5)",  "STINE": "rgba(41,118,182,0.45)",
                       "EXP":   "rgba(0,255,0,0.45)",     "DP2":   "rgba(196,223,180,0.45)"}
        COR_BOX_CLA = {"CHECK": "rgba(244,177,132,0.15)", "STINE": "rgba(41,118,182,0.12)",
                       "EXP":   "rgba(0,255,0,0.12)",     "DP2":   "rgba(196,223,180,0.12)"}
        COR_PONTO   = {"CHECK": "#C46A3A", "STINE": "#1A4F7A", "EXP": "#009900", "DP2": "#7AAF6A"}

        fig2 = go_plt.Figure()
        n_hib = len(ordem_box)

        for hib in ordem_box:
            grp = df_box[df_box["Híbrido"] == hib]["valor"].dropna()
            if len(grp) < 2:
                continue
            status = df_box[df_box["Híbrido"] == hib]["status_material"].mode()[0]
            q1, q3 = grp.quantile(0.25), grp.quantile(0.75)
            vmin, vmax, media, n = grp.min(), grp.max(), grp.mean(), len(grp)

            fig2.add_trace(go_plt.Bar(
                x=[vmax - vmin], base=vmin, y=[hib], orientation="h",
                marker_color=COR_BOX_CLA.get(status, "rgba(150,150,150,0.2)"),
                marker_line_width=0, width=0.35, showlegend=False, hoverinfo="skip"))
            fig2.add_trace(go_plt.Bar(
                x=[q3 - q1], base=q1, y=[hib], orientation="h",
                marker_color=COR_BOX_ESC.get(status, "#888888"),
                marker_line_width=0, width=0.35, showlegend=False, hoverinfo="skip"))

            idx = n_hib - 1 - ordem_box.index(hib)
            fig2.add_shape(type="line", x0=media, x1=media, y0=idx - 0.18, y1=idx + 0.18,
                           line=dict(color="#000000", width=2.5))
            fig2.add_annotation(x=media, y=hib, text=f"<b>{media:,.{_dbox}f}</b> ({n})".replace(",", "."),
                                showarrow=False, xanchor="center", yanchor="bottom", yshift=22,
                                font=dict(size=13, color="#000000", weight="bold"))

        for hib in ordem_box:
            grp2 = df_box[df_box["Híbrido"] == hib].dropna(subset=["valor"])
            if grp2.empty:
                continue
            status = grp2["status_material"].mode()[0]
            fig2.add_trace(go_plt.Scatter(
                x=grp2["valor"], y=[hib] * len(grp2), mode="markers",
                showlegend=False, legendgroup=status,
                marker=dict(color=COR_PONTO.get(status, "#555555"), size=8, opacity=0.85,
                            line=dict(color="#FFFFFF", width=0.8)),
                customdata=grp2["cod_fazenda"].tolist(),
                hovertemplate="<b>%{customdata}</b><br>" + unid_box + ": %{x:,." + str(_dbox) + "f}<extra></extra>"))

        for status, cor in COR_STATUS_PLOT.items():
            if status in df_box["status_material"].values:
                fig2.add_trace(go_plt.Scatter(
                    x=[None], y=[None], mode="markers", name=status, legendgroup=status,
                    marker=dict(color=cor, size=12), showlegend=True))

        lsd_box = calcular_lsd(ta_filtrado, col="kg_ha")
        _lsd_ok = isinstance(lsd_box, (int, float)) and not np.isnan(lsd_box)
        lsd_box_sc = (round(lsd_box / 60, 2) if unid_box == "sc/ha" else round(lsd_box, 1)) if _lsd_ok else None
        mostrar_lsd_box = st.checkbox("Mostrar linhas de corte LSD", value=True, key="chk_lsd_box")

        if lsd_box_sc is not None and mostrar_lsd_box:
            medias_box = df_box.groupby("Híbrido")["valor"].mean()
            lider_box = medias_box[ordem_box[0]]
            for i in range(1, len(ordem_box)):
                v = medias_box.get(ordem_box[i], None)
                if v is None:
                    continue
                if (lider_box - v) > lsd_box_sc:
                    idx_corte = n_hib - 1 - i + 0.5
                    fig2.add_shape(type="line", x0=0, x1=1, xref="paper",
                                   y0=idx_corte, y1=idx_corte,
                                   line=dict(color="#FF0000", width=2, dash="dot"))
                    fig2.add_annotation(x=0.02, xref="paper", y=idx_corte, yref="y",
                                        text=f"LSD: {lsd_box_sc:,.{_dbox}f}".replace(",", "."), showarrow=False,
                                        xanchor="left", yanchor="bottom",
                                        font=dict(size=11, color="#FF0000", weight="bold"))
                    lider_box = v

        fig2.update_layout(
            barmode="overlay", height=max(500, n_hib * 75 + 80),
            margin=dict(l=180, r=60, t=40, b=60),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=14, color="#000000"),
            xaxis=dict(title=dict(text=unid_lol, font=dict(size=16, color="#000000", weight="bold")),
                       showgrid=False, zeroline=False, showline=False,
                       tickfont=dict(size=14, color="#000000", weight="bold"), rangemode="tozero"),
            yaxis=dict(showgrid=True, gridcolor="#EEEEEE", gridwidth=1, zeroline=False, showline=False,
                       tickfont=dict(size=14, color="#000000", weight="bold"),
                       categoryorder="array", categoryarray=ordem_box[::-1], showticklabels=False),
            legend=dict(title=dict(text="Status", font=dict(size=14, color="#000000", weight="bold")),
                        orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                        font=dict(size=14, color="#000000", weight="bold")),
            showlegend=True)

        COR_TXT_BOX = {"CHECK": "#C46A3A", "STINE": "#2976B6", "EXP": "#009900", "DP2": "#5A8A4A"}
        status_box_map = df_box.drop_duplicates("Híbrido").set_index("Híbrido")["status_material"].to_dict()
        for hib in ordem_box:
            fig2.add_annotation(x=0, xref="paper", y=hib, yref="y", text=f"<b>{hib}</b>",
                                showarrow=False, xanchor="right", yanchor="middle",
                                font=dict(size=13, color=COR_TXT_BOX.get(status_box_map.get(hib, ""), "#333333"),
                                          weight="bold"))

        st.plotly_chart(fig2, use_container_width=True)

        df_dic_box = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                      .drop_duplicates().sort_values("cod_fazenda")
                      .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                       "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                      .reset_index(drop=True))
        with st.popover(f"Dicionário de locais ({len(df_dic_box)} locais)", use_container_width=False):
            st.markdown("Passe o mouse sobre os pontos do gráfico para identificar o local.")
            _b = st.text_input("Buscar", value="", key="busca_dic_box",
                               placeholder="Código, fazenda, cidade...")
            _df_f = (df_dic_box[df_dic_box.apply(
                lambda r: _b.strip().lower() in " ".join(r.astype(str).str.lower()), axis=1)]
                if _b.strip() else df_dic_box)
            st.dataframe(_df_f, hide_index=True, use_container_width=True)

st.divider()
# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 10 — ESTABILIDADE E ÍNDICE DE CONFIANÇA
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Estabilidade",
    "Qual o comportamento do híbrido e sua previsibilidade entre locais e safras?",
    "Híbridos com alto índice de confiança mantêm desempenho consistente independente do ambiente "
    "— menor risco na recomendação.",
)

with st.popover("ℹ️ Como interpretar · Índice de Confiança", use_container_width=False):
    st.markdown("""
**📌 Como ler esta tabela**

Cada linha é um híbrido com seu desempenho consolidado no conjunto filtrado (locais e safras ativos).
Como os ensaios de faixa são **sem repetição**, a variação observada reflete diferenças entre
ambientes — não erro experimental.

---

**📋 Colunas**
""")
    c1, c2 = st.columns(2)
    c1.markdown("""
| Coluna | Significado |
|--------|-------------|
| **Híbrido** | Nome do material |
| **[Safra]** | Nº de observações válidas naquela safra (— = não avaliado) |
| **N Total** | Total de observações válidas no conjunto filtrado |
| **Média (sc/ha)** | Média de produtividade no conjunto |
| **Mín / Máx** | Menor e maior produtividade observada |
| **Amplitude** | Máx − Mín — variação total entre ambientes |
""")
    c2.markdown("""
| Coluna | Significado |
|--------|-------------|
| **Q1 / Mediana / Q3** | Quartis — faixa central de 50% dos locais |
| **IQR (sc/ha)** | Q3 − Q1 — dispersão dos 50% centrais |
| **CV (%)** | Desvio padrão ÷ Média × 100 |
| **IQR Rel. (%)** | IQR ÷ Mediana × 100 — base do índice de confiança |
| **Índice de Confiança** | Classificação de 1 a 5 estrelas |
""")
    st.markdown("""
---

**📐 Entendendo cada medida**

> ⚠️ **Média e Mediana podem contar histórias diferentes sobre o mesmo híbrido.** Quando próximas,
o desempenho é equilibrado. Quando divergirem, vale investigar quais ambientes puxam a média.

- **Média:** soma das produtividades ÷ nº de locais. Pode ser puxada por uma área excepcional ou
  muito ruim — nem sempre representa o comportamento típico.
- **Mediana (Q2):** o valor "do meio" com os resultados em ordem — metade das áreas produz menos,
  metade mais. Não sofre distorção por áreas extremas.
- **Q1 (primeiro quartil):** separa os 25% piores dos 75% restantes. Um Q1 alto significa que mesmo
  nos piores ambientes o híbrido entrega produtividade razoável — segurança na recomendação.
- **Q3 (terceiro quartil):** separa os 75% melhores dos 25% mais altos.
- **IQR:** distância entre Q1 e Q3 — a faixa onde estão os **50% centrais** dos locais. Quanto
  menor, mais concentrado e previsível é o desempenho.

---

**⭐ Como interpretar as estrelas**

Baseado no **IQR Relativo (%)** — mais robusto para ensaios sem repetição, pois ignora valores
extremos e representa os 50% centrais dos ambientes:

| Estrelas | IQR Rel. (%) | Interpretação |
|----------|-------------|---------------|
| ⭐⭐⭐⭐⭐ | < 15% | Altamente consistente — previsível em qualquer ambiente |
| ⭐⭐⭐⭐ | 15 – 25% | Consistente — pequena variação entre locais |
| ⭐⭐⭐ | 25 – 35% | Moderado — variação aceitável, atenção ao ambiente |
| ⭐⭐ | 35 – 45% | Instável — desempenho dependente do ambiente |
| ⭐ | > 45% | Imprevisível — alto risco de recomendação ampla |

> **⚠️** indica N Total < 5 — índice calculado com poucos pontos, interpretar com cautela.

---

**🔢 Como é calculado**

> **IQR Rel. (%)** = (Q3 − Q1) ÷ Mediana × 100

> **CV (%)** = Desvio Padrão ÷ Média × 100

Com múltiplas safras ativas, a variação entre safras também entra no cálculo — o índice fica mais
exigente e representa estabilidade ampla entre locais **e** anos.
""")

df_estab = ta_filtrado[ta_filtrado["sc_ha"] > 0][["dePara", "sc_ha", "safra"]].dropna().copy()

if df_estab.empty:
    st.info("Nenhum híbrido com produtividade válida para os filtros selecionados.")
else:
    safras_ativas = sorted(df_estab["safra"].unique().tolist())

    def _estrelas(iqr_rel, n):
        if pd.isna(iqr_rel):
            return "—"
        aviso = " ⚠️" if n < 5 else ""
        if iqr_rel < 15:
            return f"⭐⭐⭐⭐⭐{aviso}"
        if iqr_rel < 25:
            return f"⭐⭐⭐⭐{aviso}"
        if iqr_rel < 35:
            return f"⭐⭐⭐{aviso}"
        if iqr_rel < 45:
            return f"⭐⭐{aviso}"
        return f"⭐{aviso}"

    rows_estab = []
    for hib, grp in df_estab.groupby("dePara"):
        s = grp["sc_ha"]
        n = len(s)
        media = s.mean()
        dp = s.std()
        cv = (dp / media * 100) if media > 0 else np.nan
        q1, med, q3 = s.quantile(0.25), s.median(), s.quantile(0.75)
        iqr = q3 - q1
        iqr_rel = (iqr / med * 100) if med > 0 else np.nan

        row = {"Híbrido": hib}
        for safra in safras_ativas:
            n_safra = len(grp[grp["safra"] == safra])
            row[str(safra)] = n_safra if n_safra > 0 else "—"
        row.update({
            "N Total":             n,
            "Média (sc/ha)":       round(media, 1),
            "Mín":                 round(s.min(), 1),
            "Máx":                 round(s.max(), 1),
            "Amplitude":           round(s.max() - s.min(), 1),
            "Q1":                  round(q1, 1),
            "Mediana":             round(med, 1),
            "Q3":                  round(q3, 1),
            "IQR (sc/ha)":         round(iqr, 1),
            "CV (%)":              round(cv, 1) if not np.isnan(cv) else None,
            "IQR Rel. (%)":        round(iqr_rel, 1) if not np.isnan(iqr_rel) else None,
            "Índice de Confiança": _estrelas(iqr_rel, n),
        })
        rows_estab.append(row)

    df_ic = (pd.DataFrame(rows_estab)
             .sort_values("IQR Rel. (%)", ascending=True, na_position="last")
             .reset_index(drop=True))

    ag_table(df_ic, height=min(600, 80 + len(df_ic) * 36))
    exportar_excel(df_ic, nome_arquivo="estabilidade.xlsx",
                   label="⬇️ Exportar Estabilidade", key="exp_estab")
    st.caption(
        "ℹ️ Ensaios de faixa são sem repetição — a variação reflete diferenças entre ambientes. "
        "O **Índice de Confiança** usa o IQR Relativo (%), mais robusto a valores extremos que o CV. "
        "Com múltiplas safras ativas, a variação entre anos também entra no cálculo.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 11 — ADAPTABILIDADE E ESTABILIDADE (Eberhart & Russell)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Adaptabilidade e Estabilidade",
    "Como cada híbrido reage aos diferentes ambientes?",
    "Híbridos responsivos se destacam em ambientes favoráveis. Híbridos estáveis mantêm desempenho "
    "mesmo em condições adversas. Os melhores combinam as duas características.",
)

df_er = ta_filtrado[ta_filtrado["sc_ha"] > 0][
    ["dePara", "cod_fazenda", "sc_ha", "status_material"]].dropna().copy()

if df_er.empty:
    st.info("Nenhum híbrido com produtividade válida para os filtros selecionados.")
else:
    # Índice ambiental: média do local − média geral
    media_geral_er = df_er["sc_ha"].mean()
    media_local_er = df_er.groupby("cod_fazenda")["sc_ha"].mean()
    df_er["idx_amb"] = df_er["cod_fazenda"].map(media_local_er) - media_geral_er

    MIN_LOCAIS_ER = 4
    contagem_er = df_er.groupby("dePara")["cod_fazenda"].nunique()
    hibridos_validos = contagem_er[contagem_er >= MIN_LOCAIS_ER].index.tolist()
    df_er = df_er[df_er["dePara"].isin(hibridos_validos)]

    rows_er = []
    for hib, grp in df_er.groupby("dePara"):
        y = grp["sc_ha"].values
        x = grp["idx_amb"].values
        n = len(y)
        X = np.column_stack([np.ones(n), x])
        try:
            beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
            a, b = beta
            ss_res = np.sum((y - X @ beta) ** 2)
            gl_res = n - 2
            s2 = (ss_res / gl_res - 1) if gl_res > 0 else np.nan   # desvio dos desvios (E&R)
            s2 = max(s2, 0) if not np.isnan(s2) else np.nan
        except Exception:
            b, s2 = np.nan, np.nan
        rows_er.append({
            "dePara": hib,
            "status_material": grp["status_material"].iloc[0],
            "b": round(b, 3) if not np.isnan(b) else None,
            "s2": round(s2, 3) if not np.isnan(s2) else None,
            "n_locais": grp["cod_fazenda"].nunique(),
        })

    if not rows_er:
        st.warning(f"⚠️ Nenhum híbrido avaliado em pelo menos {MIN_LOCAIS_ER} locais no recorte atual. "
                   "Amplie a seleção de locais para calcular adaptabilidade e estabilidade.")
    else:
        df_er_res = pd.DataFrame(rows_er).dropna(subset=["b", "s2"])

        # Lin & Binns — Pi por híbrido (usado no hover)
        maximo_local = df_er.groupby("cod_fazenda")["sc_ha"].max().rename("max_local")
        df_pi = df_er.join(maximo_local, on="cod_fazenda")
        df_pi["diff"] = df_pi["max_local"] - df_pi["sc_ha"]
        pi_rows = []
        for hib, grp in df_pi.groupby("dePara"):
            n_loc = grp["cod_fazenda"].nunique()
            pi_rows.append({"dePara": hib,
                            "Pi": round(grp["diff"].sum() / (2 * n_loc), 2),
                            "delta_lider": round(grp["diff"].mean(), 1)})
        df_final_er = df_er_res.merge(pd.DataFrame(pi_rows), on="dePara", how="left")

        # classificação por quadrante: b vs 1,0 e s² vs a mediana do conjunto
        s2_medio = df_final_er["s2"].median()

        def _quadrante(b, s2):
            if pd.isna(b) or pd.isna(s2):
                return "—"
            if b >= 1.0 and s2 <= s2_medio:
                return "🟢 Alta Performance"
            if b >= 1.0 and s2 > s2_medio:
                return "🟡 Ambiente Favorável"
            if b < 1.0 and s2 <= s2_medio:
                return "🔵 Ampla Adaptação"
            return "🔴 Atenção"

        df_final_er["Quadrante"] = df_final_er.apply(lambda r: _quadrante(r["b"], r["s2"]), axis=1)

        # tamanho do ponto: menor Pi (mais próximo do líder) = ponto maior
        _pi_max, _pi_min = df_final_er["Pi"].max(), df_final_er["Pi"].min()
        df_final_er["pi_size"] = 8 + 24 * (1 - (df_final_er["Pi"] - _pi_min) / (_pi_max - _pi_min + 1e-9))

        st.markdown("#### Índice Ambiental (Regressão de Eberhart & Russell)")
        st.caption("Cada reta representa um híbrido. A inclinação (b) indica adaptabilidade — retas "
                   "mais inclinadas respondem mais a ambientes favoráveis.")

        with st.popover("ℹ️ Como interpretar · Regressão E&R", use_container_width=False):
            st.markdown("""
**📌 O que este gráfico mostra**

Cada reta representa a relação entre a produtividade de um híbrido e a qualidade do ambiente —
quanto melhor o ambiente (eixo X), quanto mais o híbrido produz (eixo Y).

---

**📐 Como ler os elementos**

- **Eixo X** → produtividade média do ambiente (sc/ha); ambientes à direita são mais favoráveis.
- **Eixo Y** → produtividade do híbrido (sc/ha).
- **Cada reta** → comportamento de um híbrido ao longo dos ambientes.
- **Pontos** → observações reais de cada local.
- **Linha tracejada cinza** → referência b = 1,0 (acompanha exatamente a média dos ambientes).
- **Linha vertical pontilhada** → média geral do conjunto.

---

**📐 Como interpretar a inclinação (b)**

- **b > 1 (reta mais inclinada)** → híbrido **responsivo**: ganha mais que a média em ambientes
  favoráveis, mas perde mais nos ruins.
- **b ≈ 1 (paralela à referência)** → **adaptação ampla**: acompanha a média em qualquer ambiente.
- **b < 1 (reta mais plana)** → **pouco responsivo**: estável, mas não aproveita ambientes de alto
  potencial.

---

**📐 Posição da reta**

A altura da reta em relação à tracejada indica o nível geral de produtividade:

| Posição | Inclinação | Interpretação |
|---------|-----------|---------------|
| Acima + b > 1 | mais íngreme | Alto potencial — lidera e cresce mais em bons ambientes |
| Acima + b < 1 | mais plana | Produtivo e estável — bom piso mesmo em ambientes adversos |
| Abaixo + b > 1 | mais íngreme | Começa atrás mas se aproxima dos líderes em ambientes favoráveis |
| Abaixo + b < 1 | mais plana | Abaixo da média e pouco responsivo — perfil de risco |

---

**📐 Dispersão dos pontos**

- **Pontos próximos da reta** → híbrido previsível (s² baixo): o que a reta promete é o que entrega.
- **Pontos espalhados** → comportamento imprevisível (s² alto): mesmo em ambientes similares, reage
  de forma diferente.

> Só entram híbridos avaliados em pelo menos **{min_loc} locais** — abaixo disso a regressão não é
confiável.
""".replace("{min_loc}", str(MIN_LOCAIS_ER)))

        top10_default = (df_er.groupby("dePara")["sc_ha"].mean()
                         .sort_values(ascending=False).head(10).index.tolist())
        hibridos_er = sorted(df_final_er["dePara"].tolist())
        sel_hibridos_er = st.multiselect(
            "Selecione os híbridos para exibir:",
            options=hibridos_er,
            default=[c for c in top10_default if c in hibridos_er],
            key="sel_er_reg")

        if not sel_hibridos_er:
            st.info("Selecione ao menos um híbrido para exibir o gráfico.")
        else:
            fig_reg = go_plt.Figure()
            x_idx = np.linspace(df_er["idx_amb"].min(), df_er["idx_amb"].max(), 100)
            x_scha = x_idx + media_geral_er

            # paleta que não colide com as cores de status (laranja/azul/verde vibrante/verde claro)
            _PALETA_ER = ["#9B59B6", "#E91E63", "#00BCD4", "#795548", "#607D8B", "#FF5722",
                          "#673AB7", "#F06292", "#4DB6AC", "#A1887F", "#78909C", "#5D4037"]

            for _i, hib in enumerate(sel_hibridos_er):
                grp = df_er[df_er["dePara"] == hib]
                cor = _PALETA_ER[_i % len(_PALETA_ER)]
                y = grp["sc_ha"].values
                x = grp["idx_amb"].values
                X = np.column_stack([np.ones(len(y)), x])
                beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
                a, b = beta
                linha_res = df_final_er[df_final_er["dePara"] == hib]
                s2_val = linha_res["s2"].iloc[0] if not linha_res.empty else np.nan
                pi_val = linha_res["Pi"].iloc[0] if not linha_res.empty else np.nan

                fig_reg.add_trace(go_plt.Scatter(
                    x=grp["idx_amb"].values + media_geral_er, y=y, mode="markers", name=hib,
                    marker=dict(color=cor, size=7, opacity=0.6, line=dict(color="#FFFFFF", width=0.8)),
                    legendgroup=hib, showlegend=False,
                    hovertemplate=(f"<b>{hib}</b><br>Média do local: %{{x:.1f}} sc/ha<br>"
                                   "Produtividade: %{y:.1f} sc/ha<extra></extra>")))
                fig_reg.add_trace(go_plt.Scatter(
                    x=x_scha, y=a + b * x_idx, mode="lines", name=hib,
                    line=dict(color=cor, width=2), legendgroup=hib, showlegend=True,
                    hovertemplate=(f"<b>{hib}</b><br>b: {b:.3f} · s²: {s2_val:.3f} · "
                                   f"Pi: {pi_val:.1f}<extra></extra>")))

            fig_reg.add_trace(go_plt.Scatter(
                x=x_scha, y=media_geral_er + 1.0 * x_idx, mode="lines", name="Referência (b=1)",
                line=dict(color="#555555", width=1.8, dash="dash"), hoverinfo="skip"))
            fig_reg.add_vline(x=media_geral_er, line=dict(color="#444444", width=1.5, dash="dot"),
                              annotation_text=f"Média geral: {media_geral_er:.1f}",
                              annotation_position="top",
                              annotation_font=dict(size=12, color="#222222", weight="bold"))

            fig_reg.update_layout(
                height=520, margin=dict(t=40, b=60, l=60, r=160),
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
                legend=dict(orientation="v", x=1.01, y=1, xanchor="left",
                            font=dict(size=13, color="#111111"), itemsizing="constant"),
                xaxis=dict(title=dict(text="Produtividade média do ambiente (sc/ha)",
                                      font=dict(size=15, color="#111111", weight="bold")),
                           tickfont=dict(size=13, color="#111111"), gridcolor="#E5E5E5"),
                yaxis=dict(title=dict(text="Produtividade do híbrido (sc/ha)",
                                      font=dict(size=15, color="#111111", weight="bold")),
                           tickfont=dict(size=13, color="#111111"), gridcolor="#E5E5E5"))

            st.plotly_chart(fig_reg, use_container_width=True)
            st.caption("ℹ️ Reta mais inclinada que a referência (b > 1) = responsivo · reta mais plana "
                       "(b < 1) = estável · pontos espalhados = comportamento imprevisível (s² alto).")

            # ── Régua de Adaptabilidade ───────────────────────────────────────
            import streamlit.components.v1 as _comp

            # gradiente azul (neutro): a régua mede sensibilidade ao ambiente, não qualidade —
            # vermelho→verde induziria a ler um extremo como ruim e o outro como bom
            _AZUL = [(0, "#E3EDF6"), (30, "#B3CFE6"), (50, "#7FAED2"), (70, "#4A87B8"), (100, "#1A4F7A")]
            _STOPS = "".join(f'<stop offset="{p}%" stop-color="{c}"/>' for p, c in _AZUL)

            def _classe_b(b_val):
                if b_val < 0.7:
                    return "Pouco sensível ao ambiente — produção estável entre locais"
                if b_val < 0.9:
                    return "Levemente estável — responde menos às variações do ambiente"
                if b_val < 1.1:
                    return "Acompanha o ambiente — comportamento médio do conjunto"
                if b_val < 1.3:
                    return "Responsivo — responde proporcionalmente mais às melhorias do ambiente"
                return "Altamente responsivo — maior amplitude entre ambientes bons e ruins"

            def _regua_svg(hib, b_val, media_sc, n_loc, cor_hib, W=580):
                b_pct = max(0.0, min(1.0, b_val / 2.0))
                H, pad_l, pad_r = 130, 50, 50
                bar_w, bar_h, bar_y = W - pad_l - pad_r, 30, 62
                mx = pad_l + b_pct * bar_w
                gid = hib.replace(" ", "_").replace("/", "_")
                return f"""<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:{W}px;">
  <defs>
    <linearGradient id="g_{gid}" x1="0%" y1="0%" x2="100%" y2="0%">{_STOPS}</linearGradient>
    <filter id="sh_{gid}"><feDropShadow dx="0" dy="2" stdDeviation="3" flood-color="rgba(0,0,0,0.15)"/></filter>
  </defs>
  <text x="{W//2}" y="18" text-anchor="middle" font-family="Helvetica Neue,sans-serif" font-size="13" font-weight="700" fill="#1A1A1A">{hib}</text>
  <text x="{W//2}" y="36" text-anchor="middle" font-family="Helvetica Neue,sans-serif" font-size="13" fill="#1A1A1A">b = {b_val:.3f}  ·  {media_sc:.1f} sc/ha  ·  {n_loc} locais</text>
  <rect x="{pad_l}" y="{bar_y}" width="{bar_w}" height="{bar_h}" rx="15" ry="15" fill="url(#g_{gid})" filter="url(#sh_{gid})"/>
  <line x1="{pad_l + bar_w//2}" y1="{bar_y+4}" x2="{pad_l + bar_w//2}" y2="{bar_y+bar_h-4}" stroke="#374151" stroke-width="1.5" stroke-dasharray="4,3" opacity="0.85"/>
  <polygon points="{mx},{bar_y-3} {mx-10},{bar_y-16} {mx+10},{bar_y-16}" fill="{cor_hib}" stroke="white" stroke-width="1.5"/>
  <circle cx="{mx}" cy="{bar_y-3}" r="3.5" fill="white"/>
  <text x="{W//2}" y="{bar_y+bar_h+18}" text-anchor="middle" font-family="Helvetica Neue,sans-serif" font-size="12" font-weight="600" fill="#1A1A1A">{_classe_b(b_val)}</text>
</svg>"""

            def _regua_referencia_svg(W=900):
                H, pad_l, pad_r = 140, 80, 80
                bar_w, bar_h, bar_y = W - pad_l - pad_r, 32, 60
                cx = pad_l + bar_w // 2
                zonas = [
                    (pad_l,                    "Pouco\nsensível",        "#4A7FA8"),
                    (pad_l + bar_w * 0.375,    "Estável",                "#3D6E96"),
                    (cx,                       "Referência\nb=1",        "#374151"),
                    (pad_l + bar_w * 0.625,    "Responsivo",             "#2C5F8A"),
                    (pad_l + bar_w,            "Altamente\nresponsivo",  "#1A4F7A"),
                ]
                ticks = ""
                for tx, lbl, cor in zonas:
                    y0 = bar_y + bar_h + 14
                    for li, ln in enumerate(lbl.split("\n")):
                        ticks += (f'<text x="{tx}" y="{y0 + li*15}" text-anchor="middle" '
                                  f'font-family="Helvetica Neue,sans-serif" font-size="12" '
                                  f'fill="{cor}" font-weight="700">{ln}</text>')
                    ticks += (f'<line x1="{tx}" y1="{bar_y+bar_h}" x2="{tx}" y2="{bar_y+bar_h+8}" '
                              f'stroke="{cor}" stroke-width="2"/>')
                return f"""<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="width:100%;max-width:{W}px;">
  <defs>
    <linearGradient id="ref_ruler" x1="0%" y1="0%" x2="100%" y2="0%">{_STOPS}</linearGradient>
    <filter id="sh_ref"><feDropShadow dx="0" dy="2" stdDeviation="4" flood-color="rgba(0,0,0,0.12)"/></filter>
  </defs>
  <text x="{W//2}" y="20" text-anchor="middle" font-family="Helvetica Neue,sans-serif" font-size="14" font-weight="700" fill="#1A1A1A">Como interpretar a régua</text>
  <text x="{pad_l}" y="42" text-anchor="start" font-family="Helvetica Neue,sans-serif" font-size="12" fill="#6B7280">Menos sensível ao ambiente</text>
  <text x="{W-pad_r}" y="42" text-anchor="end" font-family="Helvetica Neue,sans-serif" font-size="12" fill="#6B7280">Mais sensível ao ambiente</text>
  <rect x="{pad_l}" y="{bar_y}" width="{bar_w}" height="{bar_h}" rx="16" ry="16" fill="url(#ref_ruler)" filter="url(#sh_ref)"/>
  <line x1="{cx}" y1="{bar_y+5}" x2="{cx}" y2="{bar_y+bar_h-5}" stroke="#374151" stroke-width="1.5" stroke-dasharray="4,3" opacity="0.85"/>
  {ticks}
</svg>"""

            st.markdown("#### Régua de Adaptabilidade")
            st.caption("Posição do marcador baseada no coeficiente b da regressão · b = 1 = referência (centro)")

            with st.popover("ℹ️ Como interpretar · Régua de Adaptabilidade", use_container_width=False):
                st.markdown("""
**📌 O que esta régua mostra**

A régua posiciona cada híbrido no espectro de **sensibilidade ao ambiente**, com base no
coeficiente **b** da regressão de Eberhart & Russell.

**O que é o coeficiente b?** É a inclinação da reta — mede o quanto a produtividade do híbrido
varia conforme o ambiente melhora ou piora.

---

**📐 Cada posição**

| Posição na régua | b | Significado prático |
|---|---|---|
| Extremo esquerdo | b < 0,7 | Pouco sensível ao ambiente — produção estável entre locais |
| Esquerda | 0,7 – 0,9 | Levemente estável — responde menos às variações do ambiente |
| Centro | ≈ 1,0 | Acompanha o ambiente — comportamento médio do conjunto |
| Direita | 1,1 – 1,3 | Responsivo — responde mais às melhorias do ambiente |
| Extremo direito | b > 1,3 | Altamente responsivo — maior amplitude entre ambientes bons e ruins |

---

**⚠️ Nenhum extremo é "melhor"**

A escala é azul de propósito: ela mede **sensibilidade**, não qualidade. Um híbrido à esquerda não
é pior — é mais estável. Um à direita não é melhor — é mais dependente do ambiente. O b **não mede
produtividade**: um híbrido com b baixo pode ser muito ou pouco produtivo. Combine sempre com a
média de sc/ha para uma recomendação completa.

> Na prática: em regiões heterogêneas, com muita variação de ambiente, híbridos com b próximo de 1
ou menor tendem a ser mais seguros. Em áreas de alto potencial, híbridos com b > 1 se destacam mais.
""")

            _comp.html(f"<div style='display:flex;justify-content:center;padding:8px 0 4px;'>{_regua_referencia_svg()}</div>", height=155)
            st.markdown("<div style='margin:8px 0;'></div>", unsafe_allow_html=True)

            _b_map = {}
            for _hib in sel_hibridos_er:
                _grp = df_er[df_er["dePara"] == _hib]
                if len(_grp) >= 2:
                    _y = _grp["sc_ha"].values
                    _x = _grp["idx_amb"].values
                    _X = np.column_stack([np.ones(len(_y)), _x])
                    _beta, _, _, _ = np.linalg.lstsq(_X, _y, rcond=None)
                    _b_map[_hib] = {"b": _beta[1], "media": _grp["sc_ha"].mean(),
                                    "n": _grp["cod_fazenda"].nunique()}

            _hibs_reg = [c for c in sel_hibridos_er if c in _b_map]
            for _i in range(0, len(_hibs_reg), 2):
                _par = _hibs_reg[_i:_i + 2]
                _cols = st.columns(len(_par))
                for _ci, _hib in enumerate(_par):
                    _info = _b_map[_hib]
                    _cor = _PALETA_ER[sel_hibridos_er.index(_hib) % len(_PALETA_ER)]
                    _svg = _regua_svg(_hib, _info["b"], _info["media"], _info["n"], _cor)
                    _nome_arq = _hib.replace(" ", "_").replace("/", "_")
                    _svg_id = f"regua_{_nome_arq}"
                    _svg_com_id = _svg.replace("<svg ", f'<svg id="svg_{_svg_id}" ', 1)
                    with _cols[_ci]:
                        _comp.html(f"""
<div id="{_svg_id}" style="background:#FFFFFF;padding:12px 8px 8px;border-radius:8px;width:100%;box-sizing:border-box;">
  {_svg_com_id}
</div>
<div style="margin-top:6px;">
  <button onclick="(function(){{
    var svgEl = document.getElementById('svg_{_svg_id}');
    var svgData = new XMLSerializer().serializeToString(svgEl);
    var encoded = btoa(unescape(encodeURIComponent(svgData)));
    var url = 'data:image/svg+xml;base64,' + encoded;
    var img = new Image();
    img.onload = function(){{
      var canvas = document.createElement('canvas');
      canvas.width  = svgEl.viewBox.baseVal.width  * 3;
      canvas.height = svgEl.viewBox.baseVal.height * 3;
      var ctx = canvas.getContext('2d');
      ctx.fillStyle = '#FFFFFF';
      ctx.fillRect(0, 0, canvas.width, canvas.height);
      ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
      var a = document.createElement('a');
      a.download = 'regua_{_nome_arq}.png';
      a.href = canvas.toDataURL('image/png');
      a.click();
    }};
    img.src = url;
  }})();"
    style="background:#F3F4F6;color:#374151;border:1px solid #D1D5DB;
           padding:6px 16px;border-radius:6px;font-size:13px;font-weight:600;
           cursor:pointer;font-family:'Helvetica Neue',sans-serif;">
    Baixar
  </button>
</div>""", height=195)

        # ── Dispersão por Quadrantes ──────────────────────────────────────────
        st.divider()
        st.markdown("#### Dispersão por Quadrantes — Adaptabilidade × Estabilidade")
        st.caption("Posição = b (adaptabilidade) × s² (estabilidade) · tamanho do ponto = "
                   "superioridade de Lin & Binns (maior = mais próximo do líder de cada local).")

        st.info(
            "💡 **As cores aqui têm sentido diferente da Régua de Adaptabilidade.** Na régua, a escala "
            "azul indica apenas *onde* o híbrido está no espectro de sensibilidade — nenhum extremo é "
            "ruim. Aqui as cores indicam o **desempenho combinado**: verde = responsivo E previsível "
            "(ideal) · vermelho = pouco responsivo E imprevisível (cautela). O que muda é o **s²** — "
            "híbridos com s² alto são difíceis de recomendar por serem imprevisíveis entre locais.")

        with st.popover("ℹ️ Como interpretar · Quadrantes", use_container_width=False):
            st.markdown("""
**📌 O que este gráfico mostra**

Responde duas perguntas ao mesmo tempo sobre cada híbrido:
1. **Ele aproveita bem ambientes favoráveis?** (adaptabilidade)
2. **Ele é previsível entre os locais?** (estabilidade)

---

**📐 Como ler os eixos**

- **Eixo X → Adaptabilidade (b)** — o quanto o híbrido reage quando o ambiente melhora ou piora.
  - À **direita de 1,0**: cresce mais que a média em bons ambientes, mas cai mais nos ruins — **responsivo**.
  - À **esquerda de 1,0**: varia menos — não sobe tanto nos bons, nem cai tanto nos ruins.
- **Eixo Y → Estabilidade (s²)** — o quanto o híbrido "sai da linha" além do que a adaptabilidade explica.
  - **Abaixo** da tracejada: **previsível** — entrega o que a reta promete.
  - **Acima** da tracejada: **imprevisível** — em ambientes parecidos, reage de forma diferente.

---

**🟩 Os 4 perfis**

| Perfil | Posição | O que significa na prática |
|---|---|---|
| 🟢 **Alta Performance** | direita + baixo | Responsivo e previsível — melhor perfil para recomendação ampla |
| 🟡 **Ambiente Favorável** | direita + alto | Bom potencial, mas imprevisível — apostar em áreas de alto potencial |
| 🔵 **Ampla Adaptação** | esquerda + baixo | Consistente e seguro — bom para ambientes adversos ou baixo apetite a risco |
| 🔴 **Atenção** | esquerda + alto | Não aproveita bons ambientes e ainda é imprevisível — recomendar com cautela |

---

**⚪ O tamanho do ponto — superioridade (Lin & Binns)**

Representa o quão próximo o híbrido ficou do melhor resultado **em cada local**:

- **Ponto maior** → esteve perto do líder na maioria dos locais.
- **Ponto menor** → ficou consistentemente atrás dos melhores.

> ⚠️ Um híbrido pode estar em 🟡 Ambiente Favorável com ponto **grande** — foi responsivo,
imprevisível, mas ainda competitivo. A combinação conta a história completa.

---

**📊 Ranking e Δ Líder na tabela**

- **Ranking** → posição pela proximidade ao melhor de cada local (índice Pi). O 1º é o que mais
  vezes esteve perto do vencedor local — não necessariamente o de maior média geral.
- **Δ Líder (sc/ha)** → quanto o híbrido ficou abaixo do melhor de cada local, em média.
  **0,0** = foi o melhor em todos; **−4,6** = perdeu em média 4,6 sc/ha para o líder local.

> **Exemplo:** média geral alta mas Δ Líder de −12 indica um híbrido bom no geral, mas que em cada
local havia sempre alguém claramente melhor. Já um Δ de −3 com média menor pode indicar um material
mais consistentemente competitivo.
""")

        fig_er = go_plt.Figure()
        for _status, _cor in COR_STATUS_PLOT.items():
            _df_s = df_final_er[df_final_er["status_material"] == _status]
            if _df_s.empty:
                continue
            fig_er.add_trace(go_plt.Scatter(
                x=_df_s["b"], y=_df_s["s2"], mode="markers+text", name=_status,
                text=_df_s["dePara"], textposition="top center",
                textfont=dict(size=13, color="#111111", weight="bold"),
                marker=dict(color=_cor, size=_df_s["pi_size"],
                            line=dict(color=COR_BORDA.get(_status, "#888"), width=1.5), opacity=0.90),
                customdata=_df_s[["dePara", "b", "s2", "Pi", "n_locais", "Quadrante"]].values,
                hovertemplate=("<b>%{customdata[0]}</b><br>b: %{customdata[1]:.3f}<br>"
                               "s²: %{customdata[2]:.3f}<br>Pi: %{customdata[3]:.1f}<br>"
                               "Locais: %{customdata[4]}<br>Perfil: %{customdata[5]}<extra></extra>")))

        _x_min_p = df_final_er["b"].min() - 0.1
        _x_max_p = df_final_er["b"].max() + 0.1
        _s2_max_p = df_final_er["s2"].max()
        _s2_min_p = min(0.0, float(df_final_er["s2"].min()))
        _s2_top_p = _s2_max_p * 1.10 if _s2_max_p > 0 else 1.0

        # sombra de fundo por quadrante, na cor da classificação (bem clara, só para orientar)
        for _qx0, _qx1, _qy0, _qy1, _qcor in [
            (1.0, _x_max_p, _s2_min_p, s2_medio,   "rgba(39,174,96,0.07)"),    # Alta Performance
            (1.0, _x_max_p, s2_medio,  _s2_top_p,  "rgba(243,156,18,0.07)"),   # Ambiente Favorável
            (_x_min_p, 1.0, _s2_min_p, s2_medio,   "rgba(41,128,185,0.07)"),   # Ampla Adaptação
            (_x_min_p, 1.0, s2_medio,  _s2_top_p,  "rgba(231,76,60,0.07)"),    # Atenção
        ]:
            fig_er.add_shape(type="rect", x0=_qx0, x1=_qx1, y0=_qy0, y1=_qy1,
                             fillcolor=_qcor, line_width=0, layer="below")

        fig_er.add_vline(x=1.0, line=dict(color="#888", width=1.2, dash="dash"))
        fig_er.add_hline(y=s2_medio, line=dict(color="#888", width=1.2, dash="dash"))

        for _x, _y, _xa, _ya, _txt, _c in [
            (_x_max_p, _s2_min_p, "right", "bottom", "🟢 Alta Performance",  "#27AE60"),
            (_x_max_p, _s2_top_p, "right", "top",    "🟡 Ambiente Favorável", "#F39C12"),
            (_x_min_p, _s2_min_p, "left",  "bottom", "🔵 Ampla Adaptação",    "#2980B9"),
            (_x_min_p, _s2_top_p, "left",  "top",    "🔴 Atenção",            "#E74C3C"),
        ]:
            fig_er.add_annotation(x=_x, y=_y, xanchor=_xa, yanchor=_ya, text=_txt, showarrow=False,
                                  font=dict(size=13, color=_c, weight="bold"),
                                  bgcolor="rgba(255,255,255,0.8)")

        fig_er.update_layout(
            height=550, margin=dict(t=60, b=60, l=60, r=40),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=14, color="#111111"),
            showlegend=False,
            xaxis=dict(title=dict(text="b — Adaptabilidade (b = 1: adaptação ampla)",
                                  font=dict(size=15, color="#111111", weight="bold")),
                       tickfont=dict(size=13, color="#111111", weight="bold"), gridcolor="#E5E5E5",
                       range=[_x_min_p, _x_max_p]),
            yaxis=dict(title=dict(text="s² — Estabilidade (menor = mais estável)",
                                  font=dict(size=15, color="#111111", weight="bold")),
                       tickfont=dict(size=13, color="#111111", weight="bold"), gridcolor="#E5E5E5",
                       range=[_s2_min_p, _s2_top_p]))

        st.plotly_chart(fig_er, use_container_width=True)
        st.caption(
            f"ℹ️ **Eberhart & Russell:** b = adaptabilidade (referência 1,0) · s² = desvio da regressão "
            f"· s² mediano do conjunto = {s2_medio:.3f}. **Lin & Binns:** tamanho do ponto proporcional "
            f"ao Pi — pontos maiores = maior superioridade. Híbridos com menos de {MIN_LOCAIS_ER} "
            "locais ficam fora da análise.")

        def _label_resp(b):
            if pd.isna(b):
                return "—"
            if b > 1.15:
                return f"Alta ({b:.2f})"
            if b >= 0.85:
                return f"Ampla ({b:.2f})"
            return f"Baixa ({b:.2f})"

        def _label_estab(s2):
            if pd.isna(s2):
                return "—"
            if s2 <= s2_medio * 0.5:
                return f"Alta ({s2:.2f})"
            if s2 <= s2_medio:
                return f"Média ({s2:.2f})"
            return f"Baixa ({s2:.2f})"

        _tab = df_final_er.sort_values("Pi", ascending=True).reset_index(drop=True)
        _medalhas = {1: "🥇", 2: "🥈", 3: "🥉"}
        _tab["Ranking"] = [f"{_medalhas.get(i+1, '')} {i+1}º".strip() for i in range(len(_tab))]
        _tab["Δ Líder (sc/ha)"] = _tab["delta_lider"].apply(lambda d: 0.0 if d < 0.01 else round(-d, 1))

        df_tabela_er = pd.DataFrame({
            "Híbrido":         _tab["dePara"],
            "Locais":          _tab["n_locais"],
            "Responsividade":  _tab["b"].apply(_label_resp),
            "Estabilidade":    _tab["s2"].apply(_label_estab),
            "Ranking":         _tab["Ranking"],
            "Δ Líder (sc/ha)": _tab["Δ Líder (sc/ha)"],
            "Perfil":          _tab["Quadrante"],
        })

        # fundo da coluna Perfil na mesma cor da sombra do quadrante correspondente
        _estilo_perfil = JsCode("""
        function(params) {
            const v = String(params.value || '');
            const base = {'fontWeight':'600','textAlign':'left'};
            if (v.includes('Alta Performance'))   return Object.assign({}, base, {'backgroundColor':'rgba(39,174,96,0.18)'});
            if (v.includes('Ambiente Favorável')) return Object.assign({}, base, {'backgroundColor':'rgba(243,156,18,0.18)'});
            if (v.includes('Ampla Adaptação'))    return Object.assign({}, base, {'backgroundColor':'rgba(41,128,185,0.18)'});
            if (v.includes('Atenção'))            return Object.assign({}, base, {'backgroundColor':'rgba(231,76,60,0.18)'});
            return base;
        }
        """)
        ag_table(df_tabela_er, height=min(600, 80 + len(df_tabela_er) * 36),
                 estilos_col={"Perfil": _estilo_perfil})
        exportar_excel(df_tabela_er, nome_arquivo="adaptabilidade_estabilidade.xlsx",
                       label="⬇️ Exportar Adaptabilidade e Estabilidade", key="exp_er")
        st.caption(
            "ℹ️ **Δ Líder (sc/ha)** = quanto o híbrido ficou abaixo do melhor resultado em cada local, "
            "em média — 0,0 significa que foi o melhor em todos. **Ranking** pelo índice Pi de Lin & "
            "Binns. **Responsividade** e **Estabilidade** pelos coeficientes b e s² de Eberhart & Russell.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — ÉPOCA DE PLANTIO × PRODUÇÃO (por local)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo("Época de plantio", "A época de plantio influenciou a produção?",
             "Como a produção varia conforme a data de plantio. A análise é por local (cada ensaio "
             "é plantado numa data), não por híbrido — a época é característica do ambiente.")

with st.popover("ℹ️ Como ler e interpretar esta análise", use_container_width=False):
    st.markdown("""
**📌 A pergunta que esta análise responde**

> **A janela de plantio afetou a produção nos nossos ensaios? Em que direção e com que força?**

É uma pergunta sobre a **safra e o ambiente**, não sobre híbridos. A expectativa agronômica no
safrinha é que plantios mais tardios encontrem menos chuva no enchimento de grãos — mas o quanto
isso pesou nesta safra e neste recorte é o que o gráfico mostra, e pode variar bastante entre
regiões e anos. Qualquer efeito aqui atinge todos os materiais do ensaio, não um em particular.

---

**🔢 Como é calculado, passo a passo**

1. Cada ponto do gráfico é um **local** (uma fazenda), não um híbrido.
2. Para cada local, calcula-se a **produtividade média de todos os híbridos** avaliados ali.
3. Esse valor é posicionado na data em que aquele local foi plantado.
4. A linha tracejada é a tendência: se ela desce, os plantios mais tardios renderam menos.

> **Por que por local e não por híbrido?** Porque a data de plantio é do ensaio: todos os híbridos
de uma fazenda foram plantados no mesmo dia. Se cruzássemos época diretamente com híbrido, um
material que por acaso só foi avaliado nos plantios cedo pareceria melhor — quando o mérito seria
do ambiente, não dele.

---

**📐 Como ler o gráfico**

- **Eixo X** → data de plantio do local.
- **Eixo Y** → produtividade média do local, na unidade escolhida no seletor.
- **Código sobre o ponto** → identifica o local (a correspondência está no dicionário abaixo).
- **Linha tracejada** → tendência linear entre data e produção.

---

**📅 Como ler a tabela de locais**

- **Dezena** → divisão do mês em três partes: dias 1 a 10, 11 a 20, e 21 até o fim.
- **% plantado até aqui** → o acumulado de locais na ordem de plantio. Usa as mesmas cores da marcha
  de plantio: vermelho é o início (até 50%), laranja o progresso (50–90%) e verde o fim (acima de
  90%). Mostra em que momento da janela cada ensaio entrou.

---

**🔀 A diferença para o gráfico parecido do Head-to-Head**

Existe no H2H um gráfico com o mesmo eixo horizontal (data de plantio), mas que responde a **outra
pergunta**. Confundir os dois leva a recomendações erradas, então vale a distinção:

| | Este gráfico (Análise Conjunta) | Gráfico do H2H |
|---|---|---|
| **Eixo vertical** | Produção média do local (absoluta) | Desvio do híbrido em relação à média do local |
| **Cada ponto é** | Um local | Um local, para um híbrido específico |
| **Responde** | A época afetou a produção da safra? | Este híbrido perde ou ganha terreno para os concorrentes quando o plantio atrasa? |
| **É sobre** | O ambiente | O posicionamento do material |

**O impacto de confundir:**

- Ler **este** gráfico como se fosse sobre híbridos leva a concluir "plantio tardio foi ruim, logo
  o híbrido avaliado lá é ruim" — mas aqui todos os híbridos daquele local sofreram igual. O gráfico
  não separa material nenhum.
- Ler o **do H2H** como se fosse produção absoluta é o erro mais caro: um híbrido cuja linha de
  desvio **sobe** no plantio tardio não produz mais quando se planta tarde. Ele produz menos, como
  todos — só **perde menos** que os concorrentes. Recomendá-lo como "bom para plantio tardio" está
  correto; prometer produtividade alta em plantio tardio, não.

> Regra prática: use **este** gráfico para entender a safra e o efeito da janela, **sem filtrar
híbrido**; use o **do H2H** para decidir o posicionamento de um material dentro dessa janela.

---

**⚠️ Cuidados com os filtros**

- **Não filtre por híbrido nesta análise.** Aqui a pergunta é sobre o ambiente, e o ponto de cada
  local representa o nível daquele ensaio. A média do local é sempre calculada sobre **todos os
  híbridos avaliados ali**, mesmo os fora do filtro — então o ponto não muda de altura. Mas
  selecionar um híbrido **reduz os locais** ao subconjunto onde ele foi avaliado, e isso distorce a
  leitura da janela de plantio: você passa a ver só as datas em que aquele material entrou, não a
  janela da safra.
- **Para analisar um híbrido específico, use o gráfico do Head-to-Head** (aba Desvios por Ambiente).
  Ele foi feito para isso: desconta o efeito do ambiente e mostra o comportamento do material.
- Os filtros de **safra, região, estado e cidade mudam quais locais aparecem**, e portanto mudam a
  tendência. Ao comparar duas leituras, confirme que o recorte é o mesmo.
- **Locais sem data de plantio lançada ficam de fora.** O texto abaixo do gráfico informa quantos
  são; o Diagnóstico lista quais, para o cadastro ser corrigido.
- Com **duas safras** marcadas, os pontos misturam anos diferentes — datas próximas podem pertencer
  a safras distintas, com climas distintos. Para ler o efeito da época com clareza, prefira uma
  safra por vez.

---

**⚠️ Cuidados na interpretação**

- **Tendência não é causa.** A data de plantio anda junto com outras coisas: quem planta tarde
  muitas vezes está em outra região, com outro solo ou outro manejo. A inclinação mostra
  associação, não prova de causa.
- **Poucos locais numa ponta da janela** deixam a reta instável. Se há só dois ou três ensaios no
  fim do plantio, eles puxam a tendência sozinhos.
- **Dispersão vertical na mesma data** significa que outros fatores pesaram mais que a época
  naquele conjunto — solo, manejo, chuva local.
- **Uma safra é um ano de clima.** Um efeito forte em 25/26 pode não se repetir. Confirmar em mais
  de uma safra dá segurança à conclusão.

---

**💡 Como usar na prática**

- Tendência claramente descendente e com muitos locais → argumento sólido para antecipar a janela
  de plantio na recomendação técnica.
- Tendência plana → naquele recorte, a data não foi o fator limitante; procure a explicação em
  outro lugar (região, manejo, sanidade).
- Combine com a **marcha de plantio**: se o grosso da área foi plantada justamente na janela de
  menor produção, isso tem consequência direta na média da safra.
""")

if "dataPlantioMilho" not in ta_filtrado.columns:
    st.info("Coluna de data de plantio não disponível.")
else:
    unidade_ep = st.radio("Unidade", ["sc/ha", "kg/ha"], horizontal=True, key="unid_epoca")
    col_prod_ep = "sc_ha" if unidade_ep == "sc/ha" else "kg_ha"

    def _dezena(dia):
        return "1ª dezena" if dia <= 10 else ("2ª dezena" if dia <= 20 else "3ª dezena")

    # base por local: data de plantio/colheita + produtividade média do local.
    # A média usa TODOS os híbridos avaliados no local (parte de ta_raw), não só os filtrados —
    # senão, ao filtrar um híbrido, o ponto deixaria de ser o nível do ambiente e viraria a
    # produtividade daquele material, confundindo efeito de ambiente com efeito de híbrido.
    _locais_ep = set(ta_filtrado["cod_fazenda"].dropna().unique())
    base_ep = ta_raw[ta_raw["cod_fazenda"].isin(_locais_ep)].copy()
    base_ep["_plantio"] = pd.to_datetime(base_ep["dataPlantioMilho"], errors="coerce")
    base_ep["_colheita"] = pd.to_datetime(base_ep.get("dataColheitaMilho"), errors="coerce")
    base_ep = base_ep.dropna(subset=["_plantio"])
    prod_ep = pd.to_numeric(base_ep[col_prod_ep], errors="coerce")
    base_ep = base_ep[prod_ep.notna() & (prod_ep > 0)]

    if base_ep.empty:
        st.info("Nenhum local com data de plantio e produtividade para os filtros selecionados.")
    else:
        agg_local = (
            base_ep.groupby("cod_fazenda")
            .agg(**{
                "regiao_macro":  ("regiao_macro", "first"),
                "regiao_micro":  ("regiao_micro", "first"),
                "estado":        ("estado_sigla", "first"),
                "cidade":        ("cidade_nome", "first"),
                "fazenda":       ("nomeFazenda", "first"),
                "plantio":       ("_plantio", "first"),
                "colheita":      ("_colheita", "first"),
                "prod":          (col_prod_ep, "mean"),
            })
            .reset_index()
        )
        agg_local["prod"] = agg_local["prod"].round(1)
        agg_local["dezena"] = agg_local["plantio"].dt.day.apply(_dezena)
        agg_local["mes"] = agg_local["plantio"].dt.strftime("%b/%Y")
        agg_local = agg_local.sort_values("plantio").reset_index(drop=True)

        # ── Gráfico: produtividade do local × data de plantio ──
        fig_ep = go_plt.Figure()
        fig_ep.add_trace(go_plt.Scatter(
            x=agg_local["plantio"], y=agg_local["prod"], mode="markers+text",
            marker=dict(size=12, color="#475569", opacity=0.8, line=dict(color="#334155", width=1)),
            text=agg_local["cod_fazenda"], textposition="top center",
            textfont=dict(size=10, color="#333333"),
            customdata=agg_local[["fazenda", "cidade", "estado", "dezena"]],
            hovertemplate=("<b>%{text}</b> — %{customdata[0]}<br>%{customdata[1]}/%{customdata[2]}<br>"
                           "Plantio: %{x|%d/%m/%Y} (%{customdata[3]})<br>"
                           f"Produção: %{{y:.1f}} {unidade_ep}<extra></extra>"),
            showlegend=False))
        if len(agg_local) >= 3:
            x_num = agg_local["plantio"].map(pd.Timestamp.toordinal).values.astype(float)
            y_v = agg_local["prod"].values
            z = np.polyfit(x_num, y_v, 1)
            p = np.poly1d(z)
            # reta linear: duas pontas bastam (converter muitos pontos com int() criaria degraus)
            x_line = np.array([x_num.min(), x_num.max()])
            fig_ep.add_trace(go_plt.Scatter(
                x=[agg_local["plantio"].min(), agg_local["plantio"].max()], y=p(x_line),
                mode="lines", line=dict(color="#AAAAAA", width=1.5, dash="dash"),
                name="Tendência", hoverinfo="skip"))
        fig_ep.update_layout(
            height=440, plot_bgcolor="#F5F5F5", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif", size=13, color="#111111"),
            xaxis=dict(title=dict(text="<b>Data de plantio</b>", font=dict(size=14)),
                       tickformat="%d/%b", tickfont=dict(size=12, color="#111111"),
                       gridcolor="#FFFFFF", showline=True, linecolor="#CCCCCC"),
            yaxis=dict(title=dict(text=f"<b>Produção média do local ({unidade_ep})</b>", font=dict(size=14)),
                       tickfont=dict(size=12, color="#111111"),
                       gridcolor="#FFFFFF", showline=True, linecolor="#CCCCCC"),
            margin=dict(t=40, b=60, l=70, r=40))
        st.plotly_chart(fig_ep, use_container_width=True)

        # Dicionário de locais (código → fazenda/cidade/estado)
        df_dic = (
            agg_local[["cod_fazenda", "fazenda", "cidade", "estado"]]
            .drop_duplicates().sort_values("cod_fazenda")
            .rename(columns={"cod_fazenda": "Código", "fazenda": "Local",
                             "cidade": "Cidade", "estado": "Estado"})
            .reset_index(drop=True)
        )
        with st.popover(f"Dicionário de locais ({len(df_dic)} locais)", use_container_width=False):
            st.markdown("Código do local exibido nos pontos do gráfico. Busque por código, fazenda ou cidade.")
            _busca_dic = st.text_input("Buscar", value="", key="busca_dic_epoca",
                                       placeholder="Código, fazenda, cidade...")
            if _busca_dic.strip():
                _mask = df_dic.apply(
                    lambda r: _busca_dic.strip().lower() in " ".join(r.astype(str).str.lower()), axis=1)
                _df_dic_f = df_dic[_mask]
            else:
                _df_dic_f = df_dic
            st.dataframe(_df_dic_f, hide_index=True, use_container_width=True)

        n_locais_recorte = ta_filtrado["cod_fazenda"].nunique()
        n_sem_data = n_locais_recorte - len(agg_local)
        _txt_sem = (f" Outros {n_sem_data} locais do recorte não aparecem por não terem data de "
                    "plantio lançada (veja o Diagnóstico).") if n_sem_data > 0 else ""
        st.caption(f"ℹ️ {len(agg_local)} de {n_locais_recorte} locais do recorte têm data de plantio "
                   f"e produção — cada ponto é um deles, em {unidade_ep}. O código sobre o ponto "
                   "identifica o local (veja o dicionário acima). A linha tracejada mostra a tendência: "
                   f"se inclina para baixo, plantios mais tardios renderam menos.{_txt_sem}")

        # ── Resumo por dezena: onde o plantio se concentrou e como rendeu ──
        _res = agg_local.copy()
        _res["_ord"] = _res["plantio"].apply(lambda d: (d.year, d.month,
                                                        1 if d.day <= 10 else (2 if d.day <= 20 else 3)))
        _MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        _res["_lbl"] = _res["plantio"].apply(
            lambda d: f"{_MESES[d.month - 1]} · {1 if d.day <= 10 else (2 if d.day <= 20 else 3)}ª dezena")
        _dz = (_res.groupby(["_ord", "_lbl"])
               .agg(Locais=("cod_fazenda", "nunique"),
                    _ini=("plantio", "min"), _fim=("plantio", "max"),
                    Prod=("prod", "mean"))
               .reset_index().sort_values("_ord"))
        _tot_dz = int(_dz["Locais"].sum())
        _tab_dz = pd.DataFrame({
            "Época de plantio": _dz["_lbl"],
            "Período de plantio": _dz.apply(lambda r: f"{r['_ini']:%d/%m} a {r['_fim']:%d/%m}", axis=1),
            "Locais": _dz["Locais"],
            "% acumulado": (_dz["Locais"].cumsum() / _tot_dz * 100).round(0).astype(int) if _tot_dz else 0,
            f"Produção média ({unidade_ep})": _dz["Prod"].round(1),
        })

        # a barra vai na própria coluna Locais: o número e a proporção são a mesma informação
        _max_loc = max(int(_tab_dz["Locais"].max()), 1)
        _barra = JsCode(f"""
        function(params) {{
            const v = Number(params.value);
            if (isNaN(v)) return {{'textAlign':'center'}};
            const w = Math.max(3, v / {_max_loc} * 100);
            return {{
                'background': 'linear-gradient(to right, rgba(100,116,139,0.38) 0%, rgba(100,116,139,0.38) '
                              + w + '%, rgba(226,232,240,0.55) ' + w + '%, rgba(226,232,240,0.55) 100%)',
                'color': '#111827', 'fontWeight': '700', 'textAlign': 'center'
            }};
        }}
        """)
        _zona = JsCode("""
        function(params) {
            const v = Number(params.value);
            if (isNaN(v)) return {};
            if (v < 50) return {'backgroundColor':'rgba(220,38,38,0.13)','textAlign':'center','fontWeight':'600'};
            if (v < 90) return {'backgroundColor':'rgba(217,119,6,0.15)','textAlign':'center','fontWeight':'600'};
            return {'backgroundColor':'rgba(126,211,33,0.20)','textAlign':'center','fontWeight':'600'};
        }
        """)

        st.markdown("**Distribuição do plantio por época**")
        ag_table(_tab_dz, height=min(360, 40 + 32 * len(_tab_dz) + 20),
                 estilos_col={"% acumulado": _zona, "Locais": _barra})
        st.caption("ℹ️ Em **Locais**, a barra mostra a proporção de ensaios de cada janela — a mais longa "
                   "marca onde o plantio se concentrou. **% acumulado** situa a janela dentro da safra, "
                   "com as cores da marcha de plantio. Compare a concentração com a produção média: se o "
                   "grosso da área caiu numa janela de produção baixa, isso puxa a média da safra.")
        exportar_excel(_tab_dz, nome_arquivo="epoca_por_dezena.xlsx",
                       label="⬇️ Exportar Distribuição por Época", key="exp_epoca_dz")

        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

        # ── Tabela de referência por local ──
        st.markdown("**Locais por data de plantio**")
        tab_ep = agg_local[["regiao_macro", "regiao_micro", "estado", "cidade", "fazenda",
                            "plantio", "colheita", "dezena", "mes", "prod"]].copy()
        # % acumulado de locais plantados até cada data (mesmas faixas da marcha de plantio)
        tab_ep = tab_ep.sort_values("plantio").reset_index(drop=True)
        tab_ep["pct_acum"] = ((tab_ep.index + 1) / len(tab_ep) * 100).round(0).astype(int)
        tab_ep["plantio"] = tab_ep["plantio"].dt.strftime("%d/%m/%Y")
        tab_ep["colheita"] = tab_ep["colheita"].dt.strftime("%d/%m/%Y").fillna("—")
        tab_ep = tab_ep.rename(columns={
            "regiao_macro": "Macro", "regiao_micro": "Micro", "estado": "Estado",
            "cidade": "Cidade", "fazenda": "Fazenda", "plantio": "Plantio",
            "colheita": "Colheita", "dezena": "Dezena", "mes": "Mês",
            "pct_acum": "% plantado até aqui",
            "prod": f"Produção ({unidade_ep})"})

        _estilo_zona = JsCode("""
        function(params) {
            const v = Number(params.value);
            if (isNaN(v)) return {};
            if (v < 50)  return {'backgroundColor':'rgba(220,38,38,0.13)','textAlign':'center'};
            if (v < 90)  return {'backgroundColor':'rgba(217,119,6,0.15)','textAlign':'center'};
            return {'backgroundColor':'rgba(126,211,33,0.20)','textAlign':'center'};
        }
        """)
        ag_table(tab_ep, height=min(560, 40 + 32 * min(len(tab_ep), 15) + 20),
                 estilos_col={"% plantado até aqui": _estilo_zona})
        st.caption("ℹ️ A coluna **% plantado até aqui** acumula os locais na ordem de plantio e usa as "
                   "mesmas faixas da marcha — vermelho: início (até 50%) · laranja: progresso (50–90%) · "
                   "verde: fim (acima de 90%). Ajuda a ver em que momento da janela cada local entrou.")
        exportar_excel(tab_ep, nome_arquivo="epoca_plantio.xlsx",
                       label="⬇️ Exportar Época de Plantio", key="exp_epoca")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 12 — RESPOSTA À ÉPOCA DE PLANTIO (por híbrido)
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Resposta à época de plantio",
    "Quais híbridos seguram o plantio tardio?",
    "Mede, para cada híbrido, se ele perde ou ganha terreno em relação aos concorrentes conforme a "
    "janela de plantio avança. É a mesma ideia da adaptabilidade, trocando a qualidade do ambiente "
    "pela data de plantio.",
)

MIN_LOCAIS_EPOCA = 4      # abaixo disso a reta de um híbrido não é confiável
JANELA_MIN_DIAS = 14      # janela de plantio mínima para ajustar a reta

with st.popover("ℹ️ Como ler e interpretar esta análise", use_container_width=False):
    st.markdown(f"""
**📌 As perguntas que esta análise ajuda a responder**

> **1. A época de plantio influenciou a produção neste conjunto de ensaios?**
> **2. Se influenciou, todos os híbridos responderam da mesma forma?**

A primeira é sobre o **ambiente** e se verifica na seção **Época de plantio**, mais acima: lá cada
ponto é um local, e a tendência mostra se e quanto a janela pesou nesta safra e neste recorte.

A segunda é a desta seção, e é sobre o **material**: mesmo que o conjunto todo se comporte de um
jeito, cada híbrido pode reagir de forma diferente ao avanço da janela. Alguns mantêm a posição,
outros perdem, e há os que **melhoram** conforme o plantio avança.

> Vale começar pela primeira pergunta. Sem saber se a época pesou naquele recorte, a leitura desta
seção fica solta.

---

**🔢 Como é calculado, passo a passo**

1. Em cada local, calcula-se a **média de produtividade de todos os híbridos** avaliados ali — o
   nível daquele ambiente.
2. O **desvio** de um híbrido é quanto ele ficou acima ou abaixo dessa média. Exemplo: média do
   local 120 sc/ha, híbrido fez 128 → desvio **+8**.
3. Como o desvio já desconta o ambiente, ele pode ser comparado entre locais bons e ruins.
4. Para cada híbrido, ajusta-se uma reta entre o desvio e a **data de plantio** do local.
5. A inclinação é traduzida para **sc/ha a cada 10 dias de atraso**.

> Só entram híbridos avaliados em pelo menos **{MIN_LOCAIS_EPOCA} locais com data de plantio** e com
uma janela de no mínimo **{JANELA_MIN_DIAS} dias** entre o primeiro e o último plantio. Abaixo disso
a reta seria ruído.

---

**📋 Como ler a tabela**

- **Efeito por 10 dias** → quanto o híbrido ganha (+) ou perde (−) de desvio a cada 10 dias de
  atraso. **Verde = segura o atraso**; **vermelho = perde terreno**.
- **R²** → o quanto a época explica o desvio daquele híbrido. Alto significa padrão consistente;
  baixo significa que os pontos estão espalhados e a reta é fraca.
- **Locais** → quantos ensaios com data entraram no ajuste. Poucos locais deixam a reta instável.
- **Produção média** → a produtividade média do híbrido no recorte, para você não confundir
  "segura o atraso" com "é produtivo".

---

**⚠️ O que o efeito significa — e o que não significa**

Um efeito **positivo** quer dizer que o híbrido **melhora sua posição relativa** conforme o plantio
avança. Se isso corresponde a produzir mais em números absolutos **depende do que aconteceu com o
conjunto naquele recorte**:

- Se a produção geral **caiu** com o atraso, um efeito positivo significa que o híbrido **perdeu
  menos** que os concorrentes — continua sendo a melhor escolha para a janela tardia, mas sem
  promessa de produtividade alta.
- Se a produção geral **não caiu**, o híbrido de fato **entregou mais** naquela janela.

As duas situações existem e mudam a conversa comercial. Por isso vale conferir a seção **Época de
plantio** antes: é lá que se vê o que aconteceu com o conjunto.

A tabela traz a produção média ao lado justamente para isso: o material ideal para uma janela é o
que combina **efeito favorável** com **produção que se sustenta**.

---

**⚠️ Cuidados**

- **A média do local usa todos os híbridos avaliados ali**, inclusive os fora do filtro — assim o
  desvio de um material não muda conforme você seleciona outros.
- Os filtros de **safra, região e local mudam quais ensaios entram**, e portanto mudam as retas. Ao
  comparar leituras, confirme o mesmo recorte.
- **Locais sem data de plantio ficam de fora.** O Diagnóstico lista quais são.
- **Época e ambiente andam juntos no safrinha**: plantios tardios tendem a cair em ambientes
  piores. O desvio desconta o nível do ambiente, mas se um híbrido só foi avaliado tarde, a leitura
  fica frágil de qualquer forma.
- **Uma safra é um ano de clima.** Confirmar o padrão em mais de uma safra dá muito mais segurança.

---

**💡 Como usar na prática**

- Ordene pela coluna **Efeito por 10 dias**: no topo estão os materiais que mais **ganham posição**
  conforme a janela avança; no fim, os que mais perdem. Confira sempre a produção média ao lado.
- **Efeito próximo de zero** também é informação: indica material com desempenho relativo estável em
  qualquer janela — previsível para recomendação ampla.
- Cruze com a seção **Época de plantio**: se a produção do conjunto caiu pouco naquele recorte, a
  escolha da janela pesa menos na decisão do que a escolha do material.
- Use o **Head-to-Head** quando quiser o detalhe de um confronto específico, local a local.
""")

if "dataPlantioMilho" not in ta_filtrado.columns:
    st.info("Data de plantio não disponível na base — análise indisponível.")
else:
    _loc_ep2 = set(ta_filtrado["cod_fazenda"].dropna().unique())
    _base_ep2 = ta_raw[(ta_raw["cod_fazenda"].isin(_loc_ep2)) & (ta_raw["sc_ha"] > 0)].copy()
    _dt2 = (ta_raw[["cod_fazenda", "dataPlantioMilho"]].dropna().drop_duplicates("cod_fazenda"))
    _dt2["plantio"] = pd.to_datetime(_dt2["dataPlantioMilho"], errors="coerce")
    _dt2 = _dt2.dropna(subset=["plantio"])[["cod_fazenda", "plantio"]]

    if _base_ep2.empty or _dt2.empty:
        st.info("Sem dados suficientes para os filtros selecionados.")
    else:
        _media_loc2 = _base_ep2.groupby("cod_fazenda")["sc_ha"].mean().rename("media_local")
        _hl = (_base_ep2.groupby(["dePara", "cod_fazenda"], as_index=False)["sc_ha"].mean()
               .join(_media_loc2, on="cod_fazenda")
               .merge(_dt2, on="cod_fazenda", how="inner"))
        _hl["desvio"] = _hl["sc_ha"] - _hl["media_local"]
        _hl["_x"] = _hl["plantio"].map(pd.Timestamp.toordinal).astype(float)

        _hib_filtrados = set(ta_filtrado["dePara"].dropna().unique())
        _linhas_ep2, _coef_ep2 = [], {}
        for _hib, _g in _hl.groupby("dePara"):
            if _hib not in _hib_filtrados or len(_g) < MIN_LOCAIS_EPOCA:
                continue
            if (_g["plantio"].max() - _g["plantio"].min()).days < JANELA_MIN_DIAS:
                continue
            _x, _y = _g["_x"].values, _g["desvio"].values
            _X = np.column_stack([np.ones(len(_x)), _x])
            _bb, _, _, _ = np.linalg.lstsq(_X, _y, rcond=None)
            _ssr = np.sum((_y - _X @ _bb) ** 2)
            _sst = np.sum((_y - _y.mean()) ** 2)
            _r2 = 1 - _ssr / _sst if _sst > 0 else np.nan
            _coef_ep2[_hib] = _bb
            _linhas_ep2.append({
                "Híbrido": _hib,
                "Status": (ta_filtrado[ta_filtrado["dePara"] == _hib]["status_material"].mode()[0]
                           if not ta_filtrado[ta_filtrado["dePara"] == _hib].empty else ""),
                "Efeito por 10 dias (sc/ha)": round(_bb[1] * 10, 1),
                "R²": round(_r2, 2) if not np.isnan(_r2) else None,
                "Locais": int(len(_g)),
                "Produção média (sc/ha)": round(_g["sc_ha"].mean(), 1),
            })

        if not _linhas_ep2:
            st.info(f"Nenhum híbrido atende aos mínimos ({MIN_LOCAIS_EPOCA} locais com data de "
                    f"plantio e janela de {JANELA_MIN_DIAS} dias) no recorte atual.")
        else:
            _tab_ep2 = (pd.DataFrame(_linhas_ep2)
                        .sort_values("Efeito por 10 dias (sc/ha)", ascending=False)
                        .reset_index(drop=True))

            st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

            _top_ep2 = _tab_ep2.head(4)["Híbrido"].tolist()
            _sel_ep2 = st.multiselect("Selecione até 4 híbridos para exibir no gráfico:",
                                      options=_tab_ep2["Híbrido"].tolist(),
                                      default=_top_ep2, max_selections=4, key="sel_ep2")

            if not _sel_ep2:
                st.info("Selecione ao menos um híbrido para exibir o gráfico.")
            else:
                _PAL_EP2 = ["#9B59B6", "#E91E63", "#00BCD4", "#795548", "#607D8B", "#FF5722",
                            "#673AB7", "#F06292", "#4DB6AC", "#A1887F", "#78909C", "#5D4037"]
                _cores_sel = {h: _PAL_EP2[i % len(_PAL_EP2)] for i, h in enumerate(_sel_ep2)}
                _info_sel = _tab_ep2.set_index("Híbrido")

                # caixa-tradução: nome na cor da reta + quanto ganha ou perde
                _itens = []
                for _h in _sel_ep2:
                    if _h not in _info_sel.index:
                        continue
                    _d10 = float(_info_sel.loc[_h, "Efeito por 10 dias (sc/ha)"])
                    _nm = f'<span style="color:{_cores_sel[_h]};font-weight:700;">{_h}</span>'
                    _vb = "ganha" if _d10 >= 0 else "perde"
                    _itens.append(f"{_nm} {_vb} <b>{abs(_d10):.1f} sc/ha</b>")
                st.markdown(
                    '<div style="background:#F8FAF9;border:1px solid #E5E7EB;border-radius:8px;'
                    'padding:14px 18px;margin:4px 0 14px;">'
                    '<p style="margin:0;font-size:15.5px;line-height:1.8;color:#1A1A1A;'
                    "font-family:'Helvetica Neue',Helvetica,Arial,sans-serif;\">"
                    'A cada <b>10 dias de atraso</b> no plantio, em relação à média do local:<br>'
                    + "<br>".join(_itens) + '</p></div>', unsafe_allow_html=True)

                _meta_ep2 = (ta_filtrado[["cod_fazenda", "nomeFazenda"]]
                             .drop_duplicates("cod_fazenda").set_index("cod_fazenda"))
                fig_ep2 = go_plt.Figure()
                for _i, _hib in enumerate(_sel_ep2):
                    _g = _hl[_hl["dePara"] == _hib]
                    if _g.empty or _hib not in _coef_ep2:
                        continue
                    _cor = _cores_sel[_hib]
                    _gm = _g.join(_meta_ep2, on="cod_fazenda", how="left")
                    fig_ep2.add_trace(go_plt.Scatter(
                        x=_gm["plantio"], y=_gm["desvio"], mode="markers", name=_hib,
                        marker=dict(color=_cor, size=8, opacity=0.6,
                                    line=dict(color="#FFFFFF", width=0.8)),
                        customdata=_gm[["cod_fazenda", "nomeFazenda"]].values,
                        hovertemplate=(f"<b>{_hib}</b><br><b>%{{customdata[0]}}</b> — %{{customdata[1]}}<br>"
                                       "Plantio: %{x|%d/%m/%Y}<br>Desvio: %{y:+.1f} sc/ha<extra></extra>"),
                        showlegend=False))
                    _bb = _coef_ep2[_hib]
                    _xp = np.array([_g["_x"].min(), _g["_x"].max()])
                    _yp = _bb[0] + _bb[1] * _xp
                    _xd = [_g["plantio"].min(), _g["plantio"].max()]
                    _r2h = (_info_sel.loc[_hib, "R²"] if _hib in _info_sel.index else np.nan)
                    _d10h = _bb[1] * 10
                    _r2s = f" · R²={_r2h:.2f}" if pd.notna(_r2h) else ""
                    _lbl = f"{_hib}   {_d10h:+.1f} sc/ha/10d{_r2s}"
                    fig_ep2.add_trace(go_plt.Scatter(
                        x=_xd, y=_yp, mode="lines", name=_lbl,
                        line=dict(color=_cor, width=2.5), showlegend=True,
                        hovertemplate=(f"<b>{_hib}</b><br>{_d10h:+.1f} sc/ha a cada 10 dias"
                                       "<extra></extra>")))

                fig_ep2.add_hline(y=0, line=dict(color="#444444", width=1.5, dash="dot"))
                fig_ep2.add_annotation(x=0, xref="paper", y=0, yref="y", text="média do local",
                                       showarrow=False, xanchor="left", yanchor="bottom", yshift=4,
                                       font=dict(size=12, color="#6B7280"))
                fig_ep2.update_layout(
                    height=520, plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif", color="#1A1A1A"),
                    showlegend=True,
                    legend=dict(orientation="v", x=1.01, y=1, xanchor="left", yanchor="top",
                                font=dict(size=12, color="#111111"), itemsizing="constant",
                                bgcolor="rgba(255,255,255,0.85)", bordercolor="#E5E7EB", borderwidth=1),
                    margin=dict(t=40, b=60, l=80, r=260),
                    xaxis=dict(title=dict(text="<b>Data de plantio</b>",
                                          font=dict(size=14, color="#1A1A1A", weight="bold")),
                               tickformat="%d/%b", tickfont=dict(size=12, color="#1A1A1A"),
                               showgrid=False, zeroline=False),
                    yaxis=dict(title=dict(text="<b>Desvio em relação à média do local (sc/ha)</b>",
                                          font=dict(size=14, color="#1A1A1A", weight="bold")),
                               tickfont=dict(size=12, color="#1A1A1A"),
                               showgrid=True, gridcolor="#EEEEEE", zeroline=False))
                st.plotly_chart(fig_ep2, use_container_width=True)
                st.caption("ℹ️ Cada ponto é um local: quando foi plantado e quanto o híbrido ficou "
                           "acima ou abaixo da média daquele local. Reta subindo = segura o atraso; "
                           "descendo = perde terreno. A linha pontilhada marca a média do local.")

                # ── Desvio médio por janela de plantio (dezena) ───────────────
                _MES_EP = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
                           "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

                def _dez_ep(dt):
                    _d = 1 if dt.day <= 10 else (2 if dt.day <= 20 else 3)
                    return f"{_MES_EP[dt.month - 1]} · {_d}ª dezena", (dt.year, dt.month, _d)

                _hs = _hl[_hl["dePara"].isin(_sel_ep2)].copy()
                _hs[["_lbl", "_ord"]] = _hs["plantio"].apply(
                    lambda x: pd.Series(_dez_ep(x), index=["_lbl", "_ord"]))

                # locais e período vêm de todos os ensaios com data no recorte (não de um híbrido)
                _loc_dez = (_hs.drop_duplicates("cod_fazenda")
                            .groupby(["_ord", "_lbl"])
                            .agg(Locais=("cod_fazenda", "nunique"),
                                 _ini=("plantio", "min"), _fim=("plantio", "max"))
                            .reset_index().sort_values("_ord"))
                _tot_dz2 = int(_loc_dez["Locais"].sum())

                _piv_ep = pd.DataFrame({
                    "Época de plantio": _loc_dez["_lbl"],
                    "Período de plantio": _loc_dez.apply(
                        lambda r: f"{r['_ini']:%d/%m} a {r['_fim']:%d/%m}", axis=1),
                    "Locais": _loc_dez["Locais"],
                    "% acumulado": ((_loc_dez["Locais"].cumsum() / _tot_dz2 * 100).round(0).astype(int)
                                    if _tot_dz2 else 0),
                })
                for _h in _sel_ep2:
                    _m = (_hs[_hs["dePara"] == _h].groupby("_lbl")["desvio"].mean().round(1))
                    _piv_ep[_h] = _piv_ep["Época de plantio"].map(_m)

                _max_dz2 = max(int(_piv_ep["Locais"].max()), 1)
                _barra_dz2 = JsCode(f"""
                function(params) {{
                    const v = Number(params.value);
                    if (isNaN(v)) return {{'textAlign':'center'}};
                    const w = Math.max(3, v / {_max_dz2} * 100);
                    return {{
                        'background': 'linear-gradient(to right, rgba(100,116,139,0.38) 0%, rgba(100,116,139,0.38) '
                                      + w + '%, rgba(226,232,240,0.55) ' + w + '%, rgba(226,232,240,0.55) 100%)',
                        'color': '#111827', 'fontWeight': '700', 'textAlign': 'center'
                    }};
                }}
                """)
                _zona_dz2 = JsCode("""
                function(params) {
                    const v = Number(params.value);
                    if (isNaN(v)) return {};
                    if (v < 50) return {'backgroundColor':'rgba(220,38,38,0.13)','textAlign':'center','fontWeight':'600'};
                    if (v < 90) return {'backgroundColor':'rgba(217,119,6,0.15)','textAlign':'center','fontWeight':'600'};
                    return {'backgroundColor':'rgba(126,211,33,0.20)','textAlign':'center','fontWeight':'600'};
                }
                """)
                def _js_sinal():
                    """Nova instância a cada chamada: o streamlit-aggrid substitui o marcador de
                    cada JsCode uma única vez, então reusar o mesmo objeto em várias colunas
                    faz o estilo se perder (ou ir para a coluna errada)."""
                    return JsCode("""
                    function(params) {
                        const v = Number(params.value);
                        if (isNaN(v)) return {'textAlign':'center'};
                        if (v > 0) return {'color':'#15803D','fontWeight':'800','textAlign':'center'};
                        if (v < 0) return {'color':'#B91C1C','fontWeight':'800','textAlign':'center'};
                        return {'textAlign':'center'};
                    }
                    """)

                st.markdown("**Desvio médio por janela de plantio**")
                ag_table(_piv_ep, height=min(400, 40 + 32 * len(_piv_ep) + 20),
                         estilos_col={"Locais": _barra_dz2, "% acumulado": _zona_dz2,
                                      **{_h: _js_sinal() for _h in _sel_ep2}})
                exportar_excel(_piv_ep, nome_arquivo="desvio_por_janela.xlsx",
                               label="⬇️ Exportar Desvio por Janela", key="exp_dz2")
                st.caption(
                    "ℹ️ Desvio médio de cada híbrido dentro da janela: **verde** = acima da média do "
                    "local, **vermelho** = abaixo. Em **Locais**, a barra mostra a proporção de ensaios "
                    "de cada janela — a mais longa marca onde o plantio se concentrou. **% acumulado** "
                    "situa a janela dentro da safra, com as cores da marcha de plantio. Uma vantagem "
                    "grande numa janela de poucos locais vale menos na média geral.")

            _est_efeito = JsCode("""
            function(params) {
                const v = Number(params.value);
                if (isNaN(v)) return {'textAlign':'center'};
                if (v > 0.5)  return {'color':'#15803D','fontWeight':'800','textAlign':'center'};
                if (v < -0.5) return {'color':'#B91C1C','fontWeight':'800','textAlign':'center'};
                return {'color':'#6B7280','fontWeight':'600','textAlign':'center'};
            }
            """)
            ag_table(_tab_ep2, height=min(560, 40 + 32 * min(len(_tab_ep2), 15) + 20),
                     estilos_col={"Efeito por 10 dias (sc/ha)": _est_efeito})
            exportar_excel(_tab_ep2, nome_arquivo="resposta_epoca_plantio.xlsx",
                           label="⬇️ Exportar Resposta à Época", key="exp_ep2")
            st.caption(
                f"ℹ️ **Efeito por 10 dias** = quanto o híbrido ganha (verde) ou perde (vermelho) de "
                f"desvio a cada 10 dias de atraso no plantio. Positivo **não** significa produzir mais "
                f"no plantio tardio — significa perder menos que os concorrentes. Só entram híbridos "
                f"com ao menos {MIN_LOCAIS_EPOCA} locais com data e {JANELA_MIN_DIAS} dias de janela.")

st.divider()

# ════════════════════════════════════════════════════════════════════════════════
# SEÇÃO 13 — HEATMAP DE DESEMPENHO POR LOCAL
# ════════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Heatmap",
    "Como cada híbrido se saiu em cada local?",
    "Visualize padrões de desempenho — onde cada híbrido foi forte, onde ficou atrás e quais locais "
    "são mais exigentes.",
)

with st.popover("ℹ️ Como interpretar · Heatmap", use_container_width=False):
    st.markdown("""
**📌 O que este gráfico mostra**

Cada célula é o desempenho de um híbrido em um local específico. A cor indica se ele foi bem ou mal
**em relação aos outros híbridos naquele mesmo local** — não em termos absolutos.

---

**🎨 Como ler as cores**

- **Verde escuro** → desempenho excelente naquele local, próximo do melhor.
- **Verde claro** → acima da média do local.
- **Amarelo** → na média do local.
- **Laranja / vermelho** → abaixo da média, ficou para trás naquele ambiente.
- **Célula vazia** → híbrido não avaliado naquele local.

> A cor é sempre **relativa ao local** — um híbrido pode ser verde num local difícil e vermelho num
local fácil. O que importa é a posição relativa **dentro de cada coluna**.

---

**📊 Os dois modos**

- **Produção Relativa (%)** → quanto o híbrido produziu em relação ao melhor do local (100% = foi o
  melhor). 87% significa que produziu 87% do que o líder daquele local produziu.
- **Ranking por local** → a posição do híbrido naquele local (1º, 2º, 3º...). Mais intuitivo para
  comparar posições, mas não mostra a distância entre os colocados.

---

**📐 Linhas e colunas**

- **Linhas (híbridos)** → agrupados por status, em ordem alfabética dentro de cada grupo.
- **Colunas (locais)** → ordenados por estado → cidade → código do local.
- **Linha toda verde** → híbrido consistente, bom em muitos locais.
- **Verde isolado numa coluna** → local muito específico, poucos híbridos foram bem ali.
- **Linha toda amarela/vermelha** → híbrido abaixo da média em todos os ambientes.

---

**💡 Dica de leitura**

Compare as colunas: se uma tem muitos verdes, aquele local favoreceu quase todos (ambiente fácil).
Se tem muitos vermelhos, foi um ambiente exigente onde só os melhores se destacaram.
""")

_cols_scope_hm = ["dePara", "status_material", "cod_fazenda", "estado_sigla", "cidade_nome"]
_cols_scope_hm += [c for c in ["regiao_macro", "regiao_micro"] if c in ta_filtrado.columns]
df_hm_scope = ta_filtrado[ta_filtrado["sc_ha"] > 0][_cols_scope_hm].dropna(
    subset=["dePara", "cod_fazenda", "estado_sigla", "cidade_nome"]).drop_duplicates()

if df_hm_scope.empty:
    st.info("Nenhum dado com produtividade válida para montar o heatmap.")
else:
    hibridos_no_filtro = df_hm_scope["dePara"].unique().tolist()
    locais_no_filtro = df_hm_scope["cod_fazenda"].unique().tolist()

    # Base da produção relativa POR LOCAL — mesmos 3 critérios da Auditoria, seletor próprio.
    # A referência é fixa dentro de cada local (ensaio inteiro): o filtro de híbrido não a altera.
    _cbh, _cth = st.columns([2, 3])
    with _cbh:
        base_rel_hm = st.selectbox(
            "Base da Produção Relativa (por local)",
            options=["Maior produtividade", "Média geral do local", "Testemunha"],
            index=0, key="base_rel_hm")
    with _cth:
        if base_rel_hm == "Testemunha":
            _tests_hm = sorted(ta_raw[ta_raw["status_material"].isin(["CHECK", "STINE"])]
                               ["dePara"].dropna().unique().tolist()) \
                if "status_material" in ta_raw.columns else []
            testemunha_hm = st.selectbox("Selecione a testemunha", options=_tests_hm,
                                         key="testemunha_hm") if _tests_hm else None
        else:
            testemunha_hm = None

    # Agrupamento dos locais no eixo: por estado, macrorregião ou microrregião. Macro/micro cruzam
    # estado (uma macro abrange vários estados), então cada opção REORDENA os locais para que o
    # grupo escolhido fique contíguo, e os rótulos/linhas seguem esse grupo.
    _op_grp = [("Estado", "estado_sigla")]
    if "regiao_macro" in df_hm_scope.columns:
        _op_grp.append(("Macrorregião", "regiao_macro"))
    if "regiao_micro" in df_hm_scope.columns:
        _op_grp.append(("Microrregião", "regiao_micro"))
    _lbl_grp = st.radio("Agrupar locais por:", options=[o[0] for o in _op_grp],
                        horizontal=True, key="grp_heatmap")
    col_grupo = dict(_op_grp)[_lbl_grp]   # coluna usada para agrupar/ordenar o eixo

    # valores calculados sobre TODOS os híbridos dos mesmos locais (contexto completo do local).
    # A referência é FIXA (ensaio inteiro): parte de ta_raw, não de ta_filtrado — o filtro de
    # material muda só o que aparece, não o "100% do local". Agrupa por (safra, local) para
    # não juntar safras de um mesmo local num recorte multissafra.
    _cols_hm = ["dePara", "cod_fazenda", "sc_ha"] + (["safra"] if "safra" in ta_raw.columns else [])
    df_hm_full = ta_raw[(ta_raw["sc_ha"] > 0) &
                        (ta_raw["cod_fazenda"].isin(locais_no_filtro))][
        _cols_hm].dropna().copy()
    _LOC_HM = [c for c in ["safra", "cod_fazenda"] if c in df_hm_full.columns] or ["cod_fazenda"]

    # max_local: sempre o maior do local (usado no ranking e na diferença para o líder).
    df_hm_full["max_local"] = df_hm_full.groupby(_LOC_HM)["sc_ha"].transform("max")

    # referência da produção relativa POR LOCAL, conforme a base escolhida (fixa no ensaio inteiro):
    if base_rel_hm == "Média geral do local":
        _ref_hm = df_hm_full.groupby(_LOC_HM)["sc_ha"].transform("mean")
    elif base_rel_hm == "Testemunha" and testemunha_hm:
        _rt = (df_hm_full[df_hm_full["dePara"] == testemunha_hm]
               .groupby(_LOC_HM)["sc_ha"].mean())
        _ref_hm = df_hm_full.set_index(_LOC_HM).index.map(_rt).to_numpy()
        _ref_hm = pd.Series(_ref_hm, index=df_hm_full.index)
    else:  # "Maior produtividade" (padrão) — o líder do local = 100%
        _ref_hm = df_hm_full["max_local"]

    df_hm_full["prod_rel"] = (df_hm_full["sc_ha"] / _ref_hm * 100).round(1)
    df_hm_full["ranking_local"] = (df_hm_full.groupby(_LOC_HM)["sc_ha"]
                                   .rank(ascending=False, method="min").astype(int))
    df_hm_full["total_local"] = df_hm_full.groupby(_LOC_HM)["dePara"].transform("nunique")
    # diferença para o líder — sempre contra o máximo do local (independe da base de %)
    df_hm_full["diff_sc"] = (df_hm_full["sc_ha"] - df_hm_full["max_local"]).round(1)

    df_hm_plot = df_hm_full[df_hm_full["dePara"].isin(hibridos_no_filtro)].copy()

    # ordena os locais pelo grupo escolhido (para o grupo ficar contíguo), depois cidade e código
    _cols_ord = [col_grupo, "cidade_nome", "cod_fazenda"]
    locais_ordem = (df_hm_scope[["cod_fazenda", col_grupo, "cidade_nome"]].drop_duplicates()
                    .sort_values(_cols_ord)["cod_fazenda"].tolist())

    STATUS_ORDER_HM = ["CHECK", "STINE", "EXP", "DP2"]
    hib_status = (df_hm_scope[["dePara", "status_material"]].drop_duplicates()
                  .assign(_ord=lambda d: d["status_material"].apply(
                      lambda s: STATUS_ORDER_HM.index(s) if s in STATUS_ORDER_HM else 99))
                  .sort_values(["_ord", "dePara"]))
    hibridos_ordem = hib_status["dePara"].tolist()

    _piv = lambda col, agg: (df_hm_plot.pivot_table(index="dePara", columns="cod_fazenda",
                                                    values=col, aggfunc=agg)
                             .reindex(index=hibridos_ordem, columns=locais_ordem))
    pivot_rel = _piv("prod_rel", "mean")
    pivot_rank = _piv("ranking_local", "min")
    pivot_total = _piv("total_local", "first")
    pivot_diff = _piv("diff_sc", "mean")

    col_hm1, col_hm2 = st.columns(2)
    with col_hm1:
        modo_hm = st.radio("Visualizar por:", options=["Produção Relativa (%)", "Ranking por local"],
                           horizontal=True, key="radio_heatmap")
    with col_hm2:
        if modo_hm == "Produção Relativa (%)":
            celula_hm = st.radio("Mostrar na célula:",
                                 options=["Produção relativa (%)", "Diferença para o líder (sc/ha)",
                                          "Ambos"],
                                 index=2, horizontal=True, key="radio_heatmap_celula")
        else:
            celula_hm = None

    text_mat, hover_mat = [], []
    if modo_hm == "Produção Relativa (%)":
        pivot_plot = pivot_rel
        colorscale = [[0, "#d73027"], [0.5, "#fee08b"], [1, "#1a9850"]]
        # com base "Maior produtividade" o teto é 100 (o líder); com média/testemunha os valores
        # podem passar de 100, então o topo da escala acompanha o maior valor observado.
        zmin = 60
        _vmax = float(np.nanmax(pivot_rel.values)) if pivot_rel.notna().any().any() else 100
        zmax = 100 if base_rel_hm == "Maior produtividade" else max(100, round(_vmax))
        colorbar_title = "Prod. Rel. (%)"
        for i in range(len(hibridos_ordem)):
            rt, rh = [], []
            for j in range(len(locais_ordem)):
                v = pivot_rel.iloc[i, j]
                d = pivot_diff.iloc[i, j]
                if pd.isna(v):
                    rt.append(""); rh.append("—")
                    continue
                lider = (not pd.isna(d)) and d >= 0
                # rótulo do líder: "líder" na base do máximo (onde ele é 100%); nas outras bases
                # mostra o valor real, pois o líder pode ter mais de 100% (acima da média/testemunha)
                _lider_lbl = "100%<br>líder" if (lider and base_rel_hm == "Maior produtividade") \
                    else (f"{v:.0f}%<br>líder" if lider else None)
                if celula_hm == "Produção relativa (%)":
                    rt.append(_lider_lbl if lider else f"{v:.0f}%")
                elif celula_hm == "Diferença para o líder (sc/ha)":
                    rt.append("líder" if lider else f"{d:+.1f} sc")
                else:
                    rt.append((_lider_lbl or f"{v:.0f}%<br>líder") if lider
                              else f"{v:.0f}%<br>{d:+.1f} sc")
                rh.append(f"{v:.0f}% · líder do local" if lider
                          else f"{v:.0f}% · {d:+.1f} sc/ha vs líder")
            text_mat.append(rt); hover_mat.append(rh)
    else:
        pivot_plot = pivot_rank
        colorscale = [[0, "#1a9850"], [0.5, "#fee08b"], [1, "#d73027"]]
        zmin = 1
        zmax = int(np.nanmax(pivot_rank.values)) if pivot_rank.notna().any().any() else 10
        colorbar_title = "Ranking"
        for i in range(len(hibridos_ordem)):
            rt, rh = [], []
            for j in range(len(locais_ordem)):
                r = pivot_rank.iloc[i, j]
                t = pivot_total.iloc[i, j]
                if pd.isna(r):
                    rt.append(""); rh.append("—")
                else:
                    rt.append(f"{int(r)}º")
                    rh.append(f"{int(r)}º" + (f" de {int(t)}" if not pd.isna(t) else ""))
            text_mat.append(rt); hover_mat.append(rh)

    row_h = 52 if (modo_hm == "Produção Relativa (%)" and celula_hm == "Ambos") else 38
    fig_hm = go_plt.Figure(go_plt.Heatmap(
        z=pivot_plot.values.tolist(), x=locais_ordem, y=hibridos_ordem,
        text=text_mat, customdata=hover_mat, texttemplate="%{text}",
        textfont=dict(size=11, color="#111111", weight="bold"),
        colorscale=colorscale, zmin=zmin, zmax=zmax, xgap=2, ygap=2,
        colorbar=dict(title=dict(text=colorbar_title, font=dict(size=12)),
                      tickfont=dict(size=11), thickness=14),
        hovertemplate="<b>%{y}</b> · %{x}<br>" + colorbar_title + ": %{customdata}<extra></extra>"))

    hib_status_map = hib_status.set_index("dePara")["status_material"].to_dict()
    for i, hib in enumerate(hibridos_ordem[:-1]):
        if hib_status_map.get(hib, "") != hib_status_map.get(hibridos_ordem[i + 1], ""):
            fig_hm.add_shape(type="line", x0=0, x1=1, xref="paper",
                             y0=i + 0.5, y1=i + 0.5, yref="y",
                             line=dict(color="#333333", width=2))

    # Linha vertical + rótulo separando GRUPOS (estado/macro/micro, conforme o seletor). Os locais
    # já vêm ordenados pelo grupo, então cada grupo é contíguo. A divisa entre a coluna i e i+1
    # fica em x = i + 0.5. Linha mais grossa/escura que o xgap para se destacar.
    _grp_por_local = (df_hm_scope.drop_duplicates("cod_fazenda")
                      .set_index("cod_fazenda")[col_grupo].to_dict())
    _grp_ordem = [str(_grp_por_local.get(loc, "")) for loc in locais_ordem]
    for j in range(len(_grp_ordem) - 1):
        if _grp_ordem[j] != _grp_ordem[j + 1]:
            fig_hm.add_shape(type="line", x0=j + 0.5, x1=j + 0.5, xref="x",
                             y0=0, y1=1, yref="paper",
                             line=dict(color="#222222", width=2.5))

    # Rótulo do grupo, centralizado sobre cada bloco. Micro tem nomes longos (TA01_Chapada_DF...):
    # encurta para não sobrepor. Estado (sigla) e macro cabem inteiros.
    def _rotulo_grupo(txt):
        if col_grupo == "regiao_micro":
            # mostra só o código da micro (antes do primeiro "_") ou trunca
            _cod = txt.split("_")[0] if "_" in txt else txt
            return _cod
        return txt
    _ini = 0
    for j in range(1, len(_grp_ordem) + 1):
        if j == len(_grp_ordem) or _grp_ordem[j] != _grp_ordem[_ini]:
            _centro = (_ini + j - 1) / 2
            _g = _grp_ordem[_ini]
            if _g:
                fig_hm.add_annotation(x=_centro, xref="x", y=1.02, yref="paper",
                                      text=f"<b>{_rotulo_grupo(_g)}</b>", showarrow=False,
                                      xanchor="center", yanchor="bottom",
                                      font=dict(size=13, color="#2976B6", weight="bold"))
            _ini = j

    fig_hm.update_layout(
        height=max(350, len(hibridos_ordem) * row_h + 100),
        xaxis=dict(side="bottom", tickfont=dict(size=12, color="#111111", weight="bold"),
                   title=dict(text=f"<b>Local (código) · agrupado por {_lbl_grp.lower()}</b>",
                              font=dict(size=15, color="#111111"))),
        yaxis=dict(tickfont=dict(size=13, color="#111111", weight="bold"),
                   autorange="reversed", showticklabels=False),
        margin=dict(t=48, b=80, l=180, r=60),
        plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"))

    COR_TXT_HM = {"CHECK": "#C46A3A", "STINE": "#2976B6", "EXP": "#009900", "DP2": "#5A8A4A"}
    for i, hib in enumerate(hibridos_ordem):
        fig_hm.add_annotation(x=-0.01, xref="paper", y=i, yref="y", text=f"<b>{hib}</b>",
                              showarrow=False, xanchor="right", yanchor="middle",
                              font=dict(size=13, weight="bold",
                                        color=COR_TXT_HM.get(hib_status_map.get(hib, ""), "#333333")))

    st.plotly_chart(fig_hm, use_container_width=True)
    st.caption("ℹ️ A produção relativa e o ranking são calculados **por local**, sobre todos os "
               "híbridos avaliados ali — inclusive os que estão fora do filtro. Por isso o 100% é "
               "sempre o líder real do local e a posição não muda conforme você filtra: o filtro "
               "altera apenas quais linhas aparecem. Células vazias = híbrido não avaliado naquele "
               "local. A linha preta horizontal separa os grupos de status; a vertical separa os "
               "estados (os locais estão ordenados por UF).")

    df_dic_hm = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                 .drop_duplicates().sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                 .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                  "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                 .reset_index(drop=True))
    with st.popover(f"Dicionário de locais ({len(df_dic_hm)} locais)", use_container_width=False):
        st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
        _b_hm = st.text_input("Buscar", value="", key="busca_dic_hm",
                              placeholder="Código, fazenda, cidade...")
        _df_hm_f = (df_dic_hm[df_dic_hm.apply(
            lambda r: _b_hm.strip().lower() in " ".join(r.astype(str).str.lower()), axis=1)]
            if _b_hm.strip() else df_dic_hm)
        st.dataframe(_df_hm_f, hide_index=True, use_container_width=True)

st.divider()

rodape()
