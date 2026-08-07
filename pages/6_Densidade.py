"""
pages/6_Densidade.py — Análise de Densidade (milho)

Efeito da população de plantas na produtividade e no comportamento do híbrido.
Fonte: tabela_analitica_densidade das safras 2024/25 e 2025/26 (só tipoTeste = Densidade).
Segue Better Data Visualization (Schwabish).

O agrupamento da população é FLEXÍVEL (ajustável na tela), porque a densidade é um ensaio
planejado e o método certo depende do dado real:
  - por população-ALVO (pop_tratamento): respeita o desenho do ensaio (as densidades que o
    pesquisador escolheu testar). Grupos limpos, sem partir uma densidade em duas.
  - por K-Means na população REAL contada (populacao_real_plantas_ha): encontra os agrupamentos
    naturais da população realizada. Nº de grupos ajustável (o auto força até 6, o que pode criar
    grupos artificiais quando a contagem varia muito em torno de cada alvo).

Estrutura:
  - Auditoria com DUAS abas:
      Resumida  = auditoria por plot (uma linha por parcela; produção + sanidade + perdas + fenômenos)
      Detalhe   = subamostra da população final (os trechos de 10 m que compõem o estande contado)
"""
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_multisafra
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

# ── Cores por status do material (milho) ─────────────────────────────────────
COR_STATUS_PLOT = {
    "CHECK": "#F4B184", "STINE": "#2976B6", "EXP": "#00FF00", "DP2": "#C4DFB4",
}
COR_TEXTO_STATUS = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                    "DP2": "#1A1A1A", "": "#000000"}
COR_BORDA = {"CHECK": "#C46A3A", "STINE": "#1A4F7A", "EXP": "#009900", "DP2": "#7AAF6A"}
ORDEM_STATUS = ["CHECK", "STINE", "EXP", "DP2"]

# ── Classes de reação a doenças (escala INVERSA: 9 = mais resistente) ─────────
COR_CLASS = {"AS": "#8B0000", "S": "#E63946", "MT": "#FFD600", "T": "#70C96E", "R": "#1E7A34"}
COR_TEXTO_CLASS = {"AS": "#FFFFFF", "S": "#FFFFFF", "MT": "#1A1A1A", "T": "#1A1A1A", "R": "#FFFFFF"}


def nota_para_classe(nota):
    if nota is None or (isinstance(nota, float) and np.isnan(nota)):
        return None
    n = float(nota)
    if n <= 2:
        return "AS"
    if n <= 4:
        return "S"
    if n <= 6:
        return "MT"
    if n <= 8:
        return "T"
    return "R"


st.set_page_config(page_title="Densidade · JAUM DTC", page_icon="🌽",
                   layout="wide", initial_sidebar_state="expanded")
aplicar_tema()


def ag_table(df, height=400, estilos_col=None, renderers_col=None):
    """Tabela AgGrid padrão. `estilos_col`: {coluna: JsCode} para cellStyle ·
    `renderers_col`: {coluna: JsCode} para cellRenderer (permite barras na célula)."""
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    for _col, _estilo in (estilos_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellStyle=_estilo)
    for _col, _rend in (renderers_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellRenderer=_rend, minWidth=170)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                   {"background-color": "#4A4A4A !important"},
            ".ag-header-row":               {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":              {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":        {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":         {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                     {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":  {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-icon-menu":                {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # sem "color" aqui: um !important sobreporia a cor inline do cellStyle
            # (era o que impedia o verde/vermelho dos desvios de aparecer)
            ".ag-cell":                     {"font-size": "13px !important"},
            ".ag-row":                      {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True,
    )


# ── Helper exportar Excel (mesmo padrão da soja) ──────────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 28

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = border

    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


page_header(
    "Análise de Densidade",
    "Efeito da população de plantas na produtividade e no comportamento do híbrido. "
    "Compare densidades de plantio e identifique os materiais mais responsivos e estáveis.",
    imagem="Business mission-amico.png",
)


# ══════════════════════════════════════════════════════════════════════════════
# CARREGAMENTO — tabela_analitica_densidade das duas safras
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def carregar_densidade():
    d = carregar_multisafra()
    df = d.get("tabela_analitica_densidade")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    df = df.copy()
    # produtividade canônica
    if "produtividade_valida_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_valida_kg_ha"], errors="coerce")
        if "produtividade_kg_ha" in df.columns:
            df["kg_ha"] = df["kg_ha"].fillna(pd.to_numeric(df["produtividade_kg_ha"], errors="coerce"))
    elif "produtividade_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_kg_ha"], errors="coerce")
    if "produtividade_valida_sacas_ha" in df.columns:
        df["sc_ha"] = pd.to_numeric(df["produtividade_valida_sacas_ha"], errors="coerce")
    elif "kg_ha" in df.columns:
        df["sc_ha"] = (df["kg_ha"] / 60).round(1)
    # população final realizada (contada). No milho vem de populacao_real_plantas_ha.
    for c in ["populacao_real_plantas_ha", "populacao", "pop_tratamento"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    # normaliza populações digitadas em milhares (ex: 55 -> 55000) e remove absurdos
    for c in ["populacao_real_plantas_ha", "populacao", "pop_tratamento"]:
        if c in df.columns:
            _mask = df[c] < 1000
            df.loc[_mask, c] = df.loc[_mask, c] * 1000
            df.loc[df[c] > 1_000_000, c] = pd.NA
    # altura em metros
    if "altura_planta_cm" in df.columns:
        df["altura_planta_m"] = (pd.to_numeric(df["altura_planta_cm"], errors="coerce") / 100).round(1)
    return df


with st.spinner("Carregando dados de densidade..."):
    dens_raw = carregar_densidade()

if dens_raw.empty:
    st.error("Nenhum dado de densidade disponível. Verifique se o pipeline expõe "
             "`tabela_analitica_densidade` (tipoTeste = Densidade).")
    st.stop()

# coluna de população ALVO (o tratamento planejado) — define o grupo dinamicamente por safra
COL_POP_ALVO = next((c for c in ["populacao", "pop_tratamento"] if c in dens_raw.columns), None)
# coluna de população REAL contada (varia no plantio) — informação dentro do grupo
COL_POP_REAL = next((c for c in ["populacao_real_plantas_ha", "pop_plantasFinal_ha"]
                     if c in dens_raw.columns), None)


# ══════════════════════════════════════════════════════════════════════════════
# AGRUPAMENTO — pela POPULAÇÃO REAL (K-Means: silhouette + piso de tamanho)
# ══════════════════════════════════════════════════════════════════════════════
# O grupo reúne parcelas de densidade REAL parecida (o que de fato foi plantado). O K-Means acha as
# faixas naturais na população real; o nº de grupos é escolhido pelo silhouette (a partição mais bem
# separada) e faixas minúsculas de outliers são dissolvidas no vizinho (piso), evitando grupos-
# fantasma de ~1%. Cada grupo é nomeado pela população média da faixa — então o "41k" contém só
# parcelas de ~41k (coeso). Só entram parcelas com contagem de estande (as sem contagem ficam de
# fora — não há densidade real para agrupá-las). Dinâmico por safra.
def _nome_grupo(valor):
    return f"{round(valor / 1000):.0f}k"


if COL_POP_REAL is None:
    st.error("A base não tem a coluna de população real (populacao_real_plantas_ha). Sem como "
             "agrupar pela densidade realizada.")
    st.stop()


def formar_grupos_pop_final(valores, k_fixo=None, k_min=2, k_max=6, piso_frac=0.05):
    """K-Means na população real. Se k_fixo for dado, usa esse nº de grupos (ex.: o nº de
    tratamentos-alvo do ensaio); senão, escolhe por silhouette. Descarta grupos minúsculos.
    Se o sklearn não estiver disponível, cai para um agrupamento por quantis com o mesmo nº de
    grupos (nunca gera um grupo por milhar). Retorna a lista ordenada de centros ou None."""
    v = np.asarray(valores, dtype=float)
    v = v[v > 0]
    if len(v) < 2:
        return None

    def _fallback_quantis(vals, k):
        # divide em k faixas por quantis e usa a média de cada faixa como centro
        k = max(1, min(int(k), len(np.unique(vals))))
        bordas = np.quantile(vals, np.linspace(0, 1, k + 1))
        bordas[0], bordas[-1] = -np.inf, np.inf
        centros = []
        for j in range(k):
            faixa = vals[(vals >= bordas[j]) & (vals < bordas[j + 1])]
            if len(faixa):
                centros.append(float(np.mean(faixa)))
        return sorted(centros) if centros else sorted(np.unique(vals).tolist())

    try:
        from sklearn.cluster import KMeans
        from sklearn.metrics import silhouette_score
    except Exception:
        return _fallback_quantis(v, k_fixo if k_fixo else 4)

    X = v.reshape(-1, 1)
    n_unico = len(np.unique(v))
    if k_fixo and k_fixo >= 1:
        # fixa o nº de grupos (limitado ao nº de valores distintos disponíveis)
        k = min(int(k_fixo), n_unico)
        km = KMeans(n_clusters=k, random_state=42, n_init=10).fit(X)
    else:
        # escolha automática por silhouette
        _melhor = None
        for kk in range(k_min, min(k_max, n_unico) + 1):
            _km = KMeans(n_clusters=kk, random_state=42, n_init=10).fit(X)
            if len(set(_km.labels_)) < kk:
                continue
            sil = silhouette_score(X, _km.labels_)
            if _melhor is None or sil > _melhor[0]:
                _melhor = (sil, kk, _km)
        if _melhor is None:
            return _fallback_quantis(v, k_fixo if k_fixo else 4)
        _, k, km = _melhor
    centros = km.cluster_centers_.flatten()
    piso = max(int(len(v) * piso_frac), 3)  # grupo precisa de ao menos 5% dos plots (mín. 3)
    tam = np.bincount(km.labels_, minlength=k)
    centros_ok = sorted([float(centros[i]) for i in range(k) if tam[i] >= piso])
    return centros_ok if centros_ok else sorted(centros.tolist())


def atribuir_grupo(p, centros):
    if pd.isna(p) or p <= 0 or not centros:
        return None
    return _nome_grupo(centros[int(np.argmin([abs(p - c) for c in centros]))])


dens_raw = dens_raw.copy()  # evita mutar o objeto cacheado
# a coluna de população final numérica (usada para formar e atribuir os grupos)
dens_raw["_pop_final"] = pd.to_numeric(dens_raw[COL_POP_REAL], errors="coerce")

# os grupos são formados DEPOIS do filtro (sobre a seleção ativa), para refletirem a safra/recorte.


# ══════════════════════════════════════════════════════════════════════════════
# FILTROS (sidebar) — encadeados, padrão safra 25/26
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>', unsafe_allow_html=True)

    if st.button("🔄 Limpar filtros", use_container_width=True, key="den_btn_limpar"):
        for key in list(st.session_state.keys()):
            if (str(key).startswith("den_") or str(key).startswith("__opts_den")) \
                    and not str(key).endswith("_btn_limpar") and key != "den_limpar_inline":
                del st.session_state[key]
        st.rerun()

    def _podar_keys(prefix, opcoes, molde):
        """Remove o estado de checkboxes de opções que saíram da cascata — senão, ao reaparecerem,
        voltam marcadas e o filtro se reaplica sozinho."""
        antigas = st.session_state.get(f"__opts_{prefix}", [])
        atuais = set(map(str, opcoes))
        for o in antigas:
            if str(o) not in atuais:
                st.session_state.pop(molde(o), None)
        st.session_state[f"__opts_{prefix}"] = list(opcoes)

    def checkboxes(opcoes, default_all=True, defaults=None, prefix=""):
        """Checkboxes simples (para listas curtas: safra, status, grupo)."""
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_{o}")
        sel = []
        for o in opcoes:
            key = f"{prefix}_{o}"
            if key in st.session_state:
                marcado = st.checkbox(str(o), key=key)
            else:
                checked = (o in defaults) if defaults is not None else default_all
                marcado = st.checkbox(str(o), value=checked, key=key)
            if marcado:
                sel.append(o)
        return sel

    def filtro_busca(opcoes, prefix):
        """Busca textual + seleção persistente (para listas longas: cidade, fazenda, híbrido,
        responsável). A seleção é lida direto dos checkboxes, inclusive os ocultos pela busca."""
        if f"{prefix}_reset" not in st.session_state:
            st.session_state[f"{prefix}_reset"] = 0
        r = st.session_state[f"{prefix}_reset"]
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_chk_{r}_{o}")

        busca = st.text_input("Buscar", value="", key=f"den_busca_{prefix}",
                              placeholder="Digite para filtrar...")
        filtradas = ([c for c in opcoes if busca.strip().lower() in str(c).lower()]
                     if busca.strip() else opcoes)

        if st.button("Limpar seleção", key=f"{prefix}_limpar", use_container_width=True):
            for o in opcoes:
                st.session_state.pop(f"{prefix}_chk_{r}_{o}", None)
            st.session_state.pop(f"__opts_{prefix}", None)
            st.session_state[f"{prefix}_reset"] = r + 1
            st.rerun()

        for c in filtradas:
            st.checkbox(str(c), key=f"{prefix}_chk_{r}_{c}")

        sel = [o for o in opcoes if st.session_state.get(f"{prefix}_chk_{r}_{o}", False)]
        return sel or opcoes

    # trilha de diagnóstico: quantas linhas sobram depois de cada filtro. Serve para achar,
    # quando a tela vier vazia, EXATAMENTE qual filtro zerou a base.
    _trilha = [("base (todas as safras)", len(dens_raw))]

    # 1. Safra — checkboxes simples, padrão a mais recente
    def _ano_safra(s):
        try:
            return int(str(s).split("/")[0])
        except Exception:
            return -1
    safras_all = sorted(dens_raw["safra"].dropna().unique().tolist()) if "safra" in dens_raw.columns else []
    safra_def = sorted(safras_all, key=_ano_safra)[-1:] if safras_all else []
    with st.expander("📅 Safra", expanded=True):
        safras_sel = checkboxes(safras_all, defaults=safra_def, prefix="den_safra")
    # nenhuma safra marcada = filtro não aplicado. Zerar a base aqui (iloc[0:0]) produzia a tela
    # "nenhum registro" toda vez que sobrava estado de sessão inconsistente.
    if safras_sel:
        d1 = dens_raw[dens_raw["safra"].isin(safras_sel)]
    else:
        d1 = dens_raw
        st.caption("Nenhuma safra marcada — mostrando todas.")
    _trilha.append(("safra", len(d1)))

    # grupos pela POPULAÇÃO REAL (K-Means + silhouette + piso): cada grupo reúne parcelas de
    # densidade real parecida, então o 41k contém só quem é ~41k (coeso). O nome do grupo é a
    # população média da faixa. Só entram parcelas com contagem de estande; as sem contagem ficam
    # de fora (não há densidade real para agrupá-las). Dinâmico por safra.
    d1 = d1.copy()
    # nº de grupos = nº de tratamentos-alvo do ensaio (dinâmico por safra: 4 nesta, mas acompanha
    # se uma safra testar 3 ou 5 densidades). Assim os grupos da população real espelham o desenho.
    _n_alvos = None
    if COL_POP_ALVO and COL_POP_ALVO in d1.columns:
        _alvos_distintos = pd.to_numeric(d1[COL_POP_ALVO], errors="coerce").dropna()
        _alvos_distintos = _alvos_distintos[_alvos_distintos > 0].unique()
        _n_alvos = len(_alvos_distintos) if len(_alvos_distintos) else None
    _centros_grupos = (formar_grupos_pop_final(d1["_pop_final"].dropna().values, k_fixo=_n_alvos)
                       if not d1.empty else None)
    if _centros_grupos:
        d1["pop_grupo"] = d1["_pop_final"].apply(lambda p: atribuir_grupo(p, _centros_grupos))
    else:
        d1["pop_grupo"] = None

    # filtros encadeados: listas curtas usam checkboxes; listas longas usam busca
    _config_filtros = [
        ("regiao_macro", "🗺️ Macro", "den_macro", False),
        ("regiao_micro", "📍 Micro", "den_micro", False),
        ("estado_sigla", "🏛️ Estado", "den_estado", False),
        ("cidade_nome", "🏙️ Cidade", "den_cidade", True),
        ("nomeFazenda", "🚜 Fazenda", "den_fazenda", True),
        ("nomeResponsavel", "👤 Responsável", "den_resp", True),
        ("status_material", "🏷️ Status", "den_status", False),
        ("dePara", "🌽 Híbrido", "den_hib", True),
        ("pop_grupo", "📊 Grupo densidade", "den_grupo", False),
    ]
    for _col, _lab, _pref, _usar_busca in _config_filtros:
        if _col in d1.columns:
            _ops = sorted(d1[_col].dropna().unique().tolist(),
                          key=lambda x: (int(str(x).replace("k", "")) if _col == "pop_grupo"
                                         and str(x).replace("k", "").isdigit() else str(x)))
            if not _ops:
                continue  # coluna sem valores na seleção (ex: nenhum grupo formado) — não filtra
            with st.expander(_lab, expanded=False):
                if _usar_busca:
                    _sel = filtro_busca(_ops, _pref)
                else:
                    _sel = checkboxes(_ops, prefix=_pref)
            # só vale o que ainda existe na cascata: uma seleção herdada de outra safra
            # (ex: grupo "82k" que sumiu) não pode zerar a base
            _sel_ok = [v for v in (_sel or []) if v in _ops]
            if _sel_ok:
                d1 = d1[d1[_col].isin(_sel_ok)]
            # nada marcado (ou só opções que não existem mais) = filtro não aplicado
            _trilha.append((_lab.split(" ", 1)[-1], len(d1)))

ta_filtrado = d1.copy()

if ta_filtrado.empty:
    st.warning("Nenhum registro nos filtros ativos. A tabela abaixo mostra quantas linhas "
               "sobraram depois de cada filtro — o primeiro com 0 é o que zerou a base.")
    st.dataframe(pd.DataFrame(_trilha, columns=["Filtro", "Linhas"]), hide_index=True)
    if st.button("🔄 Limpar todos os filtros", key="den_limpar_inline"):
        for key in list(st.session_state.keys()):
            if str(key).startswith("den_") or str(key).startswith("__opts_den"):
                del st.session_state[key]
        st.rerun()
    st.stop()

# contexto para os subtítulos — safra + a cadeia geográfica DO QUE ESTÁ SELECIONADO.
# Todos os níveis aparecem sempre, refletindo o resultado dos filtros: sem filtro nenhum eles
# mostram o ensaio inteiro e vão encolhendo conforme você filtra. Todos os níveis vêm
# DETALHADOS: quando são mais de 3 valores, a contagem entra na frente da lista
# ("9 cidades: Cristalina, ...") para manter o resumo sem perder o detalhe.
_NIVEIS_CTX = [
    ("regiao_macro", "Macro", "macros"),
    ("regiao_micro", "Micro", "micros"),
    ("estado_sigla", "UF", "UFs"),
    ("cidade_nome", "Cidade", "cidades"),
    ("cod_fazenda", "Local", "locais"),
]
_MAX_NOMES_CTX = 3   # acima disso a contagem entra ANTES da lista, que continua inteira

_ctx_partes, _ctx_detalhe = [], []
if "safra" in ta_filtrado.columns:
    _ctx_partes.append("Safra: " + ", ".join(sorted(ta_filtrado["safra"].dropna().unique().astype(str))))

for _cg, _rot_g, _plural_g in _NIVEIS_CTX:
    if _cg not in ta_filtrado.columns:
        continue
    _vals_txt = sorted(ta_filtrado[_cg].dropna().unique().astype(str))
    if not _vals_txt:
        continue
    # o subtítulo de cada seção traz a lista COMPLETA de cada nível. Quando são mais de 3
    # valores, a contagem entra na frente ("20 cidades: Alta Floresta, ...") para o resumo
    # continuar visível junto do detalhe.
    _ctx_partes.append(f"{_rot_g}: " + ", ".join(_vals_txt)
                       if len(_vals_txt) <= _MAX_NOMES_CTX
                       else f"{len(_vals_txt)} {_plural_g}: " + ", ".join(_vals_txt))
    _ctx_detalhe.append((_rot_g, _plural_g, _vals_txt))

contexto_str = " · ".join([p for p in _ctx_partes if p])

# o recorte também fica na sidebar, junto dos filtros que o produziram: lá ele aparece
# quebrado por nível (um parágrafo por camada), mais fácil de conferir do que a linha corrida.
_PLURAL_TIT = {"macros": "Macros", "micros": "Micros", "UFs": "UFs",
               "cidades": "Cidades", "locais": "Locais"}


def _render_recorte():
    if "safra" in ta_filtrado.columns:
        st.markdown("**Safra** — " +
                    ", ".join(sorted(ta_filtrado["safra"].dropna().unique().astype(str))))
    for _rot_d, _plural_d, _vals_d in _ctx_detalhe:
        st.markdown(f"**{_rot_d if len(_vals_d) <= _MAX_NOMES_CTX else _PLURAL_TIT.get(_plural_d, _plural_d.capitalize())}** "
                    f"({len(_vals_d)}) — " + ", ".join(_vals_d))

with st.sidebar:
    with st.expander("📍 Recorte atual", expanded=False):
        _render_recorte()

# grupos de densidade presentes NA SELEÇÃO ATIVA — só os que têm plots de fato (value_counts
# ignora categorias vazias; unique() poderia carregar níveis fantasma de outra safra)
_cont_grupos = ta_filtrado["pop_grupo"].dropna().value_counts()
_grupos_encontrados = sorted([g for g, n in _cont_grupos.items() if n > 0],
                             key=lambda x: int(str(x).replace("k", "")) if str(x).replace("k", "").isdigit() else 0)

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA (tabela por plot: produção, componentes, perdas, fenômenos, sanidade)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Auditoria", "Os dados de cada parcela, como vieram do campo", contexto_str)

col_map = {
    "safra": "Safra", "cod_fazenda": "Cód. Local", "nomeFazenda": "Fazenda",
    "cidade_nome": "Cidade", "estado_sigla": "Estado", "regiao_macro": "Região Macro",
    "regiao_micro": "Região Micro", "nomeResponsavel": "Responsável",
    "dePara": "Híbrido", "status_material": "Status",
    # densidade / produção — alvo, final e grupo lado a lado
    # a população-alvo vem como 'pop_tratamento' na base do pipeline e 'populacao' na amostra —
    # mapeamos os dois nomes para o mesmo rótulo para funcionar em qualquer base.
    "populacao": "Pop. Alvo (pl/ha)",
    "pop_tratamento": "Pop. Alvo (pl/ha)",
    COL_POP_REAL or "populacao_real_plantas_ha": "Pop. Final (pl/ha)",
    "pop_grupo": "Grupo Densidade", "kg_ha": "kg/ha", "sc_ha": "sc/ha",
    "umidade_pct": "Umidade (%)",
    # agronômicas
    "altura_planta_m": "Altura Planta (m)", "altura_espiga_cm": "Altura Espiga (cm)",
    # componentes de produção (usados adiante para ver a influência da densidade)
    "plantas_10m_media": "Plantas/10m (n)",
    "fileiras_media": "Fileiras (n)",
    "graos_fileira_media": "Grãos/Fileira (n)",
    "pmg_corrigido_g": "PMG (g)",
    "prod_estimada_sacas_ha": "Prod. Estimada (sc/ha)",
    "divergencia_prod_pct": "Divergência Prod. (%)",
    # perdas
    "pct_acamadas": "Acamamento (%)", "pct_quebradas": "Quebramento (%)",
    "pct_dominadas": "Dominadas (%)", "pct_colmo_podre": "Colmo Podre (%)",
    "pct_perda_total": "Perda Total (%)",
    # fenômenos
    "pct_green_snap": "Green Snap (%)", "pct_morte_prematura": "Morte Prematura (%)",
    "pct_ma_formacao_espigas": "Má Formação (%)", "pct_enfezamento": "Enfezamento (%)",
}
# sanidade: adiciona dinamicamente as colunas de nota de doença que existirem
_cols_doenca = [c for c in ta_filtrado.columns if c.startswith("nota_") and c != "nota_densidade"]
for _c in _cols_doenca:
    col_map[_c] = _c.replace("nota_", "").replace("_", " ").title() + " (nota)"

cols_disp = [c for c in col_map if c in ta_filtrado.columns]
# ordem por blocos lógicos: (1) onde — geografia + código; (2) o quê — material; (3) quanto
# produziu; (4) em que densidade — populações; (5) o restante (componentes, perdas, sanidade).
_bloco_geo = ["safra", "cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
              "regiao_macro", "regiao_micro", "nomeResponsavel"]
_bloco_material = ["dePara", "status_material"]
_bloco_producao = ["kg_ha", "sc_ha"]
_bloco_populacao = [COL_POP_ALVO or "populacao",
                    COL_POP_REAL or "populacao_real_plantas_ha", "pop_grupo"]
_ordem_blocos = _bloco_geo + _bloco_material + _bloco_producao + _bloco_populacao
_frente = [c for c in _ordem_blocos if c in cols_disp]
_resto = [c for c in cols_disp if c not in _frente]
cols_disp = _frente + _resto
df_show = ta_filtrado[cols_disp].rename(columns=col_map)

with st.popover("ℹ️ Como ler esta tabela", use_container_width=False):
    st.markdown("""
**Esta é a tabela crua: uma linha por parcela colhida.** Todas as outras seções da página são
resumos dela. Se um número parecer estranho lá embaixo, é aqui que você acha a parcela que o
causou.

**Como usar os filtros da barra lateral**

Eles são **encadeados**: escolher um estado reduz as cidades disponíveis, escolher a cidade reduz
as fazendas, e assim por diante. Tudo o que você marcar vale para a **página inteira** — as onze
seções mostram sempre o mesmo recorte, e ele está escrito embaixo do título de cada uma.

- **Nada marcado num filtro = filtro desligado**, mostra tudo. Não é preciso marcar todos.
- **Buscar** aparece nos filtros longos (cidade, fazenda, híbrido, responsável): digite parte do
  nome para achar sem rolar a lista. O que já estava marcado continua marcado, mesmo sumindo da
  busca.
- **Limpar filtros**, no topo da barra, volta tudo ao começo, com a safra mais recente marcada.
- **Recorte atual**, logo abaixo dos filtros, lista tudo o que está selecionado.

**Os blocos de colunas, na ordem em que aparecem**

1. **Onde** — safra, código do local, fazenda, cidade, estado, regiões e responsável. O *Cód.
   Local* é o identificador curto usado nos gráficos e nos mapas do resto da página.
2. **O quê** — híbrido e status (CHECK, STINE, EXP, DP2).
3. **Quanto produziu** — kg/ha e sc/ha, já corrigidos para 13,5% de umidade.
4. **Em que densidade** — três colunas diferentes, e a confusão entre elas é o erro mais comum:
   - **Pop. Alvo** é o que o tratamento pediu no plantio.
   - **Pop. Final** é o que foi **contado** no campo. É esta que explica a produtividade, porque é
     a densidade que a planta viveu.
   - **Grupo Densidade** é a faixa em que a parcela caiu, formada pela **Pop. Final**. Uma parcela
     que mirou 82 mil e emergiu com 41 mil aparece no grupo de 41 mil.
5. **O resto** — componentes de produção (PMG, fileiras, grãos por fileira), alturas, perdas,
   fenômenos e as notas de sanidade.

**Duas direções de escala convivendo na mesma tabela.** Nas notas de doença, a escala vai de 1 a 9
e **9 é o melhor** (planta que resistiu). Nas perdas e fenômenos, a unidade é percentual de plantas
e **quanto maior, pior**. Repare no nome da coluna antes de julgar o número.

**Zero nem sempre significa a mesma coisa.** Numa nota de doença, `0` quer dizer **não avaliado**.
Numa perda, `0` quer dizer **avaliado e sem ocorrência**. As seções seguintes tratam os dois casos
de forma diferente e explicam isso em cada popover.

**Ordenar e filtrar dentro da tabela:** clique no cabeçalho para ordenar, ou use o funil para
filtrar só ali, sem mexer nos filtros da página. O botão de exportar leva o que está na tela.
""")

ag_table(df_show, height=min(460, 36 + 32 * len(df_show) + 20))
exportar_excel(df_show, nome_arquivo="auditoria_densidade.xlsx",
               label="⬇️ Exportar Auditoria", key="den_exp_audit_resumo")
st.caption(f"{len(df_show)} plots · uma linha por parcela, com produção, agronômicas, "
           f"perdas, fenômenos e sanidade juntos. Grupo de densidade pela população real (K-Means).")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — GRUPOS DE DENSIDADE (narrativa Schwabish: Volume + Desempenho)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Grupos de População",
             "Como as parcelas se distribuem entre as populações reais",
             contexto_str)

_CL_VERDE = "#1E7A34"
_CL_CINZA = "#94A3B8"
_CL_TEXTO = "#1A1A1A"
_CL_SUB = "#6B7280"

# ── Visão dos dados: distribuição da população real dentro de cada alvo ────────
# Mostra a matéria-prima ANTES do agrupamento: quanto cada tratamento-alvo variou no plantio.
# É o que justifica, adiante, agrupar pela população final (a variação é grande).
_COL_ALVO_v = COL_POP_ALVO if COL_POP_ALVO and COL_POP_ALVO in ta_filtrado.columns else None
if _COL_ALVO_v and COL_POP_REAL:
    _dfv = ta_filtrado[[_COL_ALVO_v, "_pop_final"]].copy()
    _dfv["_alvo"] = pd.to_numeric(_dfv[_COL_ALVO_v], errors="coerce")
    _dfv = _dfv[(_dfv["_alvo"] > 0) & (_dfv["_pop_final"] > 0)]

    if not _dfv.empty and _dfv["_alvo"].nunique() >= 1:
        st.markdown("##### Como a população real ficou dentro de cada densidade planejada")
        st.caption("Cada linha é um tratamento-alvo (o que se planejou plantar); os pontos são as "
                   "parcelas na sua população real. Faixas largas = plantio variou muito em torno "
                   "do alvo. É a foto crua dos dados, antes de formar os grupos.")

        _alvos_v = sorted(_dfv["_alvo"].unique())
        _alvos_rev = list(reversed(_alvos_v))
        # uma cor por densidade-alvo (do menor ao maior), para distinguir as linhas
        _paleta_dados = ["#2976B6", "#1E7A34", "#E8871E", "#8B5CF6", "#EF476F", "#06A77D"]
        _cor_alvo = {a: _paleta_dados[i % len(_paleta_dados)] for i, a in enumerate(_alvos_v)}
        fig_dados = go_plt.Figure()
        for i, a in enumerate(_alvos_rev):
            v = _dfv[_dfv["_alvo"] == a]["_pop_final"].values
            if len(v) == 0:
                continue
            media, mediana = float(np.mean(v)), float(np.median(v))
            vmin, vmax = float(np.min(v)), float(np.max(v))
            q1, q3 = float(np.percentile(v, 25)), float(np.percentile(v, 75))
            cv = float(np.std(v) / media * 100) if media > 0 else 0
            cor = _cor_alvo[a]
            # faixa amplitude total (clara) e IQR (escura)
            fig_dados.add_shape(type="rect", x0=vmin, x1=vmax, y0=i - 0.32, y1=i + 0.32,
                                fillcolor=cor, opacity=0.10, line=dict(width=0))
            fig_dados.add_shape(type="rect", x0=q1, x1=q3, y0=i - 0.32, y1=i + 0.32,
                                fillcolor=cor, opacity=0.18, line=dict(width=0))
            # média (sólida) e mediana (tracejada)
            fig_dados.add_shape(type="line", x0=media, x1=media, y0=i - 0.38, y1=i + 0.38,
                                line=dict(color=cor, width=2.5))
            fig_dados.add_shape(type="line", x0=mediana, x1=mediana, y0=i - 0.38, y1=i + 0.38,
                                line=dict(color=cor, width=1.6, dash="dash"))
            # referência do alvo (linha pontilhada cinza no valor planejado)
            fig_dados.add_shape(type="line", x0=a, x1=a, y0=i - 0.42, y1=i + 0.42,
                                line=dict(color="#9CA3AF", width=1.2, dash="dot"))
            # pontos com jitter
            jit = np.random.uniform(-0.20, 0.20, size=len(v))
            fig_dados.add_trace(go_plt.Scatter(
                x=v, y=[i + j for j in jit], mode="markers",
                marker=dict(color=cor, size=5, opacity=0.5, line=dict(color="#FFFFFF", width=0.5)),
                hovertemplate=f"Alvo {a/1000:.0f}k<br>Pop. real: %{{x:,.0f}}<extra></extra>",
                showlegend=False))
            # rótulo à direita: média + n + CV
            fig_dados.add_annotation(
                x=vmax, y=i,
                text=(f"<b>{media/1000:.0f}k</b>"
                      f"<span style='color:#6B7280;font-size:10px'> (alvo {a/1000:.0f}k · "
                      f"n={len(v)} · CV {cv:.0f}%)</span>"),
                xanchor="left", yanchor="middle", showarrow=False, xshift=10,
                font=dict(size=12, color="#1A1A1A", weight="bold"))

        _xmn = _dfv["_pop_final"].min()
        _xmx = _dfv["_pop_final"].max()
        _pad = (_xmx - _xmn) * 0.22 + 3000
        fig_dados.update_layout(
            height=max(240, len(_alvos_v) * 70 + 70),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
            margin=dict(t=10, b=45, l=30, r=30),
            xaxis=dict(title=dict(text="<b>População real (pl/ha)</b>",
                                  font=dict(size=12, color="#111111", weight="bold")),
                       range=[_xmn - _pad, _xmx + _pad * 2.8],
                       showgrid=True, gridcolor="#EEEEEE", zeroline=False,
                       tickfont=dict(size=11, color="#111111")),
            yaxis=dict(tickmode="array", tickvals=list(range(len(_alvos_v))),
                       ticktext=[f"alvo {a/1000:.0f}k" for a in _alvos_rev],
                       tickfont=dict(size=12, color="#111111", weight="bold"),
                       range=[-0.6, len(_alvos_v) - 0.4]))
        st.plotly_chart(fig_dados, use_container_width=True)
        st.caption("Linha pontilhada cinza = população-alvo planejada · linha sólida = média real · "
                   "tracejada = mediana · faixa escura = 50% das parcelas (IQR) · faixa clara = "
                   "amplitude total. Só parcelas com contagem de estande.")


# resumo por grupo: nº de plots, produção média, população real média
_res_grupos = []
for g in _grupos_encontrados:
    sub = ta_filtrado[ta_filtrado["pop_grupo"] == g]
    if len(sub) == 0:
        continue  # grupo sem parcelas na seleção ativa não entra (evita grupos-fantasma)
    sc = pd.to_numeric(sub["sc_ha"], errors="coerce").dropna() if "sc_ha" in sub.columns else pd.Series(dtype=float)
    sc = sc[sc > 0]
    real = pd.to_numeric(sub["_pop_final"], errors="coerce").dropna()
    real = real[real > 0]
    _res_grupos.append({
        "grupo": g, "n": len(sub), "n_real": len(real), "n_sc": len(sc),
        "sc_media": float(sc.mean()) if len(sc) else np.nan,
        "real_media": float(real.mean()) if len(real) else np.nan,
        "real_mediana": float(real.median()) if len(real) else np.nan,
        "vmin": float(real.min()) if len(real) else np.nan,
        "vmax": float(real.max()) if len(real) else np.nan,
        # quartis e CV da população real — coesão interna do grupo (usados no resumo numérico)
        "q1": float(real.quantile(0.25)) if len(real) else np.nan,
        "q3": float(real.quantile(0.75)) if len(real) else np.nan,
        "cv": (float(real.std(ddof=1) / real.mean() * 100)
               if len(real) > 1 and real.mean() > 0 else 0.0),
    })
_df_grupos = pd.DataFrame(_res_grupos)

if _df_grupos.empty or _df_grupos["n"].sum() == 0:
    st.info("Nenhum grupo de densidade disponível nos filtros ativos.")
else:
    _grupo_lider = (_df_grupos.dropna(subset=["sc_media"]).sort_values("sc_media", ascending=False)
                    ["grupo"].iloc[0] if _df_grupos["sc_media"].notna().any() else None)

    # ── Cabeçalho + cards de contexto (igual à soja) ──────────────────────────
    _df_ord = _df_grupos.sort_values("grupo", key=lambda s: s.str.replace("k", "").astype(int))
    _tot_p = int(_df_ord["n"].sum())
    _grupo_maior = _df_ord.loc[_df_ord["n"].idxmax(), "grupo"]  # destaque = mais parcelas

    st.markdown(f"""
<div style="margin:0.5rem 0 0.2rem;">
    <h2 style="font-size:1.7rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.2;">
        {len(_df_ord)} faixas de densidade identificadas no ensaio
    </h2>
    <p style="font-size:15px;color:#6B7280;margin:6px 0 0;line-height:1.6;">
        Esta etapa <strong>não analisa nenhum cultivar específico</strong> — ela define as faixas
        de população que existem nos dados e serão usadas em todas as análises seguintes.
        {_tot_p:,} parcelas no total.
    </p>
</div>""", unsafe_allow_html=True)

    with st.popover("ℹ️ Como ler esta seção", use_container_width=False):
        _lista_alvos = ", ".join(_df_ord["grupo"].tolist())
        st.markdown(f"""
**Esta seção não analisa nenhum cultivar específico.** Ela organiza as densidades do ensaio em
grupos e verifica se esse agrupamento ficou bom — uma etapa necessária antes de qualquer análise
por híbrido.

---

**Por que agrupar pela POPULAÇÃO REAL (o que de fato foi plantado)?**

A resposta da planta à densidade depende de quantas plantas ela realmente teve por hectare, não de
quantas se pretendia plantar. Uma parcela planejada para 82k mas que emergiu com 41k competiu como
uma de 41k — é essa a densidade que importa. Por isso o grupo reúne parcelas de densidade real
parecida: o grupo "41k" contém só parcelas de ~41k, coeso. (Consequência: só entram parcelas com
contagem de estande; sem ela, não há densidade real para agrupar.)

**Por que K-Means?**

As populações reais variam continuamente em torno das densidades plantadas. O K-Means acha as
faixas naturais nesses dados — os agrupamentos onde as parcelas se concentram — em vez de impormos
faixas fixas arbitrárias que poderiam cortar um grupo real ao meio. Cada grupo é nomeado pela
população média da sua faixa.

**Por que o número de grupos é automático (silhouette)?**

Deixamos os dados decidirem quantos grupos existem: o silhouette mede a separação de cada partição
e escolhemos a mais nítida. Assim o agrupamento se adapta a cada safra — 3 densidades → 3 grupos;
5 → 5 grupos. E faixas minúsculas de outliers são dissolvidas no vizinho (piso), evitando grupos-
fantasma de ~1%.

Nesta seleção saíram **{len(_df_ord)} grupos**: {_lista_alvos}.

---

**Os três atos verificam se o agrupamento ficou bom:**

**Ato 1 — Volume** → *Cada grupo tem parcelas suficientes?* Usa um lollipop porque o que importa
aqui é comparar o tamanho de cada grupo de relance. Um grupo minúsculo ao lado de um enorme
indicaria agrupamento desbalanceado — médias pouco confiáveis.

**Ato 2 — Dispersão** → *As faixas estão bem definidas e coesas?* Mostra cada parcela como um ponto,
com a faixa do IQR (50% centrais) e a amplitude total. Faixas estreitas = grupos coesos (parcelas
de densidade parecida); largas = o grupo reúne densidades mais distantes.

**Ato 3 — Desempenho** → *Os grupos têm comportamento distinto?* Usa ponto com **erro padrão da
média** (não barra simples) de propósito: o erro padrão mostra a incerteza de cada média. Se as
barras de dois grupos se sobrepõem, a diferença provavelmente não é real. Por isso a leitura é
cautelosa, sem cravar uma "melhor densidade".

Se os três atos mostram grupos equilibrados e com comportamento distinto, o agrupamento é confiável
para as próximas análises.
""")

    _ccols = st.columns(len(_df_ord))
    for _ci, (_, _row) in enumerate(_df_ord.reset_index(drop=True).iterrows()):
        _destaque = _row["grupo"] == _grupo_maior
        _cor_borda = _CL_VERDE if _destaque else "#E5E7EB"
        _cor_num = _CL_VERDE if _destaque else "#1A1A1A"
        _pct = int(_row["n"] / _tot_p * 100)
        # o grupo JÁ é a população real média da faixa; embaixo, a amplitude (menor–maior da faixa)
        _num_grande = _row["grupo"]
        _sub = (f"faixa {_row['vmin']/1000:.0f}–{_row['vmax']/1000:.0f}k pl/ha"
                if not np.isnan(_row.get("vmin", np.nan)) else "população real")
        _ccols[_ci].markdown(f"""
<div style="border:2px solid {_cor_borda};border-radius:12px;padding:14px 12px;
            background:#FAFAFA;text-align:center;">
    <p style="font-size:22px;font-weight:800;color:{_cor_num};margin:0;">{_num_grande}</p>
    <p style="font-size:12px;color:#6B7280;margin:2px 0 0;">{_sub}</p>
    <p style="font-size:12px;color:#6B7280;margin:4px 0 0;"><b style="color:#1A1A1A">{int(_row['n'])}</b> parc. do ensaio · {_pct}%</p>
</div>""", unsafe_allow_html=True)

    st.markdown("<div style='margin-top:2rem;'></div>", unsafe_allow_html=True)

    # ── Ato 1 — Volume (lollipop horizontal, idêntico à soja) ─────────────────
    _df_vol = _df_grupos.sort_values("grupo", key=lambda s: s.str.replace("k", "").astype(int))
    _tot_parc = int(_df_vol["n"].sum())
    st.markdown(
        f'<div style="margin:1rem 0 0.2rem;">'
        f'<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        f'letter-spacing:0.07em;margin:0 0 2px;">Ato 1 — Volume</p>'
        f'<p style="font-size:1.15rem;font-weight:700;color:#1A1A1A;margin:0;">Quantas parcelas '
        f'foram avaliadas em cada densidade?</p>'
        f'<p style="font-size:13px;color:#6B7280;margin:3px 0 0;">Um ensaio balanceado — densidades '
        f'com número parecido de parcelas — dá comparações mais justas.</p>'
        f'</div>', unsafe_allow_html=True)

    # ordem invertida no Y (maior densidade em cima), com índices numéricos como a soja
    _grupos_rev1 = list(reversed(_df_vol["grupo"].tolist()))
    fig_vol = go_plt.Figure()
    for i, g in enumerate(_grupos_rev1):
        r = _df_vol[_df_vol["grupo"] == g].iloc[0]
        n_val = int(r["n"])
        pct_v = int(n_val / _tot_parc * 100)
        is_dest = g == _grupo_maior  # Ato 1 é sobre VOLUME: destaca o de mais parcelas (= cards)
        cor = _CL_VERDE if is_dest else _CL_CINZA
        # linha do lollipop (shape, mais leve)
        fig_vol.add_shape(type="line", x0=0, x1=n_val, y0=i, y1=i,
                          line=dict(color=cor, width=2.5))
        # ponto com borda branca
        fig_vol.add_trace(go_plt.Scatter(
            x=[n_val], y=[i], mode="markers",
            marker=dict(color=cor, size=14, line=dict(color="#FFFFFF", width=2)),
            showlegend=False,
            hovertemplate=f"<b>Faixa {g}</b><br>{n_val} parcelas ({pct_v}%)<extra></extra>"))
        # label do número + % (% em cinza)
        fig_vol.add_annotation(
            x=n_val, y=i,
            text=f"<b>{n_val}</b> parcelas <span style='color:#6B7280'>{pct_v}%</span>",
            xanchor="left", yanchor="middle", showarrow=False, xshift=14,
            font=dict(size=13, color="#1A1A1A" if is_dest else "#6B7280", weight="bold"))
        # label do grupo à esquerda: o grupo já é a população final média da faixa
        _txt_lbl = f"<b>{g}</b>"
        fig_vol.add_annotation(
            x=0, y=i, text=_txt_lbl, xanchor="right", yanchor="middle",
            showarrow=False, xshift=-10, align="right",
            font=dict(size=13, color=_CL_VERDE if is_dest else "#1A1A1A", weight="bold"))

    _x_max1 = int(_df_vol["n"].max() * 1.45)
    fig_vol.update_layout(
        height=max(240, len(_df_vol) * 60 + 80),
        plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
        margin=dict(t=20, b=20, l=80, r=20),
        xaxis=dict(range=[0, _x_max1], showgrid=True, gridcolor="#EEEEEE",
                   zeroline=False, showline=False,
                   tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                   title=dict(text="<b>Número de parcelas</b>",
                              font=dict(size=14, color="#1A1A1A", weight="bold"))),
        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False, showline=False,
                   range=[-0.7, len(_df_vol) - 0.3]))
    st.plotly_chart(fig_vol, use_container_width=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── Ato 2 — Separação (dispersão da população real dentro de cada grupo) ───
    # Usa os MESMOS grupos dos cards (os tratamentos-alvo). Mostra como a população real se
    # distribui dentro de cada tratamento — a variação de plantio em torno do alvo. Um plot que
    # emergiu bem abaixo do alvo aparece à esquerda (falha de estande); é informação real.
    _df_disp = ta_filtrado[ta_filtrado["pop_grupo"].notna()].copy()
    if COL_POP_REAL:
        _df_disp["_real"] = pd.to_numeric(_df_disp[COL_POP_REAL], errors="coerce")
        _df_disp = _df_disp[_df_disp["_real"] > 0]

    if COL_POP_REAL and not _df_disp.empty:
        st.markdown(
            '<div style="margin:0 0 0.2rem;">'
            '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
            'letter-spacing:0.07em;margin:0 0 2px;">Ato 2 — Dispersão</p>'
            '<p style="font-size:1.15rem;font-weight:700;color:#1A1A1A;margin:0;">Como a população '
            'real se distribui dentro de cada grupo?</p>'
            '<p style="font-size:13px;color:#6B7280;margin:3px 0 0;">Cada linha é um grupo (os mesmos '
            'dos cards); os pontos são a população real das parcelas. Faixas estreitas confirmam que '
            'o grupo é coeso — reúne parcelas de densidade parecida.</p>'
            '</div>', unsafe_allow_html=True)

        with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
            st.markdown("""
Este gráfico mostra a **população real** de cada parcela, dentro dos grupos formados pela própria
população real (os mesmos dos cards). Só usa parcelas com contagem de estande.

- **Cada ponto** → uma parcela; posição no eixo X = população real (pl/ha)
- **Faixa clara** → amplitude total do grupo (menor ao maior valor)
- **Faixa escura** → onde estão 50% das parcelas (IQR)
- **Linha sólida** → média · **Linha tracejada** → mediana
- **Label** → n · CV do grupo

**CV (coeficiente de variação)** mede a coesão do grupo:
- **CV baixo** → parcelas de densidade parecida — grupo bem definido
- **CV alto** → o grupo reúne densidades mais distantes

Como o grupo é formado pela densidade real, cada faixa fica coesa: o grupo "41k" contém só parcelas
de ~41k. Passe o mouse sobre os pontos para ver fazenda, cidade e produtividade.
""")

        # resumo por grupo (os mesmos grupos-alvo dos cards)
        _res_disp = {}
        for g in _grupos_encontrados:
            v = _df_disp[_df_disp["pop_grupo"] == g]["_real"].values
            if len(v):
                _res_disp[g] = dict(
                    n=len(v), media=float(np.mean(v)), mediana=float(np.median(v)),
                    cv=float(np.std(v) / np.mean(v) * 100) if np.mean(v) > 0 else 0,
                    vmin=float(np.min(v)), vmax=float(np.max(v)),
                    q1=float(np.percentile(v, 25)), q3=float(np.percentile(v, 75)))

        _grupos_disp = [g for g in _grupos_encontrados if g in _res_disp]
        _grupo_maior_real = _grupo_maior
        _grupos_rev2 = list(reversed(_grupos_disp))

        fig_sep = go_plt.Figure()
        for i, g in enumerate(_grupos_rev2):
            r = _res_disp[g]
            is_dest = g == _grupo_maior_real
            cor = _CL_VERDE if is_dest else "#94A3B8"
            op_area = 0.18 if is_dest else 0.10
            op_iqr = 0.30 if is_dest else 0.16
            # faixa amplitude total
            fig_sep.add_shape(type="rect", x0=r["vmin"], x1=r["vmax"], y0=i - 0.32, y1=i + 0.32,
                              fillcolor=cor, opacity=op_area, line=dict(width=0))
            # faixa IQR
            fig_sep.add_shape(type="rect", x0=r["q1"], x1=r["q3"], y0=i - 0.32, y1=i + 0.32,
                              fillcolor=cor, opacity=op_iqr, line=dict(width=0))
            # média (sólida) e mediana (tracejada)
            fig_sep.add_shape(type="line", x0=r["media"], x1=r["media"], y0=i - 0.38, y1=i + 0.38,
                              line=dict(color=cor, width=3 if is_dest else 2))
            fig_sep.add_shape(type="line", x0=r["mediana"], x1=r["mediana"], y0=i - 0.38, y1=i + 0.38,
                              line=dict(color=cor, width=1.8, dash="dash"))
            # pontos com jitter + hover
            sub_g = _df_disp[_df_disp["pop_grupo"] == g].reset_index(drop=True)
            vals = sub_g["_real"].values
            _cd_cols = [c for c in ["nomeFazenda", "cidade_nome", "estado_sigla", "sc_ha"]
                        if c in sub_g.columns]
            if len(vals):
                jit = np.random.uniform(-0.22, 0.22, size=len(vals))
                if _cd_cols:
                    cd = sub_g[_cd_cols].values
                    _ht = ("<b>%{customdata[0]}</b> · %{customdata[1]}, %{customdata[2]}<br>"
                           "Pop. real: <b>%{x:,.0f}</b> pl/ha<br>sc/ha: %{customdata[3]:.1f}<extra></extra>")
                else:
                    cd, _ht = None, f"<b>Grupo {g}</b><br>Pop. real: %{{x:,.0f}}<extra></extra>"
                fig_sep.add_trace(go_plt.Scatter(
                    x=vals, y=[i + j for j in jit], mode="markers",
                    marker=dict(color=cor, size=6 if is_dest else 5,
                                opacity=0.75 if is_dest else 0.55,
                                line=dict(color="#FFFFFF", width=0.5)),
                    customdata=cd, hovertemplate=_ht, showlegend=False))
            # label: média + n + CV (o grupo/alvo aparece no eixo Y)
            fig_sep.add_annotation(
                x=r["vmax"], y=i,
                text=f"<span style='color:#6B7280;font-size:11px'>n={r['n']} · CV {r['cv']:.0f}%</span>",
                xanchor="left", yanchor="middle", showarrow=False, xshift=10,
                font=dict(size=11, color="#6B7280"))

        _xmin_s = min(r["vmin"] for r in _res_disp.values())
        _xmax_s = max(r["vmax"] for r in _res_disp.values())
        _pad_s = (_xmax_s - _xmin_s) * 0.22 + 3000
        fig_sep.update_layout(
            height=max(240, len(_grupos_disp) * 70 + 70),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
            margin=dict(t=10, b=45, l=75, r=30),
            xaxis=dict(title=dict(text="<b>População real (pl/ha)</b>",
                                  font=dict(size=12, color="#111111", weight="bold")),
                       range=[_xmin_s - _pad_s, _xmax_s + _pad_s * 2.5],
                       showgrid=True, gridcolor="#EEEEEE", zeroline=False,
                       tickfont=dict(size=11, color="#111111")),
            yaxis=dict(tickmode="array", tickvals=list(range(len(_grupos_disp))),
                       ticktext=[f"{g}" for g in _grupos_rev2],
                       tickfont=dict(size=12, color="#111111", weight="bold"),
                       showgrid=False, zeroline=False,
                       range=[-0.7, len(_grupos_disp) - 0.3]))
        st.plotly_chart(fig_sep, use_container_width=True)
        st.caption("Cada ponto = uma parcela · linha sólida = média · linha tracejada = mediana · "
                   "faixa escura = 50% das parcelas (IQR) · faixa clara = amplitude total. Só parcelas "
                   "com contagem de estande entram aqui.")

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── Ato 3 — Desempenho ────────────────────────────────────────────────────
    st.markdown(
        '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.07em;margin:0 0 6px;">Ato 3 — Desempenho</p>', unsafe_allow_html=True)
    st.markdown("Produtividade média por densidade, na média de **todos os híbridos e locais** — "
                "uma visão do ensaio como um todo, não de um material específico. As barras de erro "
                "mostram a incerteza da média.")

    with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
        st.markdown("""
Este gráfico mostra a produtividade média de **todos os híbridos juntos** em cada densidade — é o
ensaio como um todo, não a resposta de um material.

- **Ponto** → média de sc/ha de todas as parcelas da densidade.
- **Barra horizontal** (├──┤) → erro padrão da média: quanto pode variar se o ensaio fosse repetido.
- **Linha tracejada** → média geral do ensaio (referência).

**Como ler com cuidado:**
- Se as barras de erro de duas densidades **se sobrepõem**, a diferença entre elas provavelmente
  **não é significativa** — pode ser variação normal do campo.
- Densidades com **poucas parcelas** (n pequeno) têm médias menos confiáveis — cheque o n.
- Esta é a média geral. **Cada híbrido responde diferente** à densidade — quem ganha com mais
  população e quem satura só aparece na análise por híbrido (próximas seções). Não tire daqui uma
  recomendação de densidade para um material específico.
""")

    _dfd = _df_grupos.dropna(subset=["sc_media"]).copy()
    if _dfd.empty:
        st.info("Sem produtividade registrada para comparar as densidades.")
    else:
        # erro-padrão e n por densidade
        _ep, _n_g = {}, {}
        for g in _dfd["grupo"]:
            v = pd.to_numeric(ta_filtrado[(ta_filtrado["pop_grupo"] == g)
                                          & (ta_filtrado["sc_ha"] > 0)]["sc_ha"], errors="coerce").dropna()
            _n_g[g] = len(v)
            _ep[g] = float(v.std(ddof=1) / np.sqrt(len(v))) if len(v) > 1 else 0.0
        _media_geral = float(_dfd["sc_media"].mean())

        # ordena densidade crescente; Y invertido (maior densidade em cima)
        _dfd = _dfd.sort_values("grupo", key=lambda s: s.str.replace("k", "").astype(int))
        _grupos_y = _dfd["grupo"].tolist()

        fig_des = go_plt.Figure()
        fig_des.add_vline(x=_media_geral, line=dict(color="#374151", width=1.3, dash="dot"))
        fig_des.add_annotation(x=_media_geral, y=len(_grupos_y) - 0.4,
                               text=f"média geral {_media_geral:.1f}", showarrow=False,
                               xanchor="left", xshift=6, font=dict(size=10, color="#374151"))
        for _, r in _dfd.iterrows():
            g = r["grupo"]
            cor = _CL_VERDE if g == _grupo_lider else _CL_CINZA
            fig_des.add_trace(go_plt.Scatter(
                x=[r["sc_media"]], y=[g], mode="markers",
                marker=dict(color=cor, size=11),
                error_x=dict(type="data", array=[_ep[g]], color=cor, thickness=1.5, width=6),
                showlegend=False,
                hovertemplate=f"{g}<br>%{{x:.1f}} sc/ha ± {_ep[g]:.1f}<br>n={_n_g[g]}<extra></extra>"))
            fig_des.add_annotation(x=r["sc_media"] + _ep[g], y=g,
                                   text=f"  <b>{r['sc_media']:.1f}</b> ±{_ep[g]:.1f} (n={_n_g[g]})",
                                   showarrow=False, xanchor="left",
                                   font=dict(size=11, color="#1A1A1A"))

        _xmin = (_dfd["sc_media"] - pd.Series(_ep).reindex(_dfd["grupo"]).values).min()
        _xmax = (_dfd["sc_media"] + pd.Series(_ep).reindex(_dfd["grupo"]).values).max()
        _pad = (_xmax - _xmin) * 0.25 + 5
        fig_des.update_layout(
            height=340,
            xaxis=dict(title=dict(text="<b>Produtividade média (sc/ha)</b>",
                                  font=dict(size=12, color="#111111", weight="bold")),
                       range=[_xmin - _pad, _xmax + _pad * 1.5], gridcolor="#EEEEEE",
                       tickfont=dict(size=11, color="#111111")),
            yaxis=dict(title=dict(text="<b>Densidade</b>", font=dict(size=12, color="#111111", weight="bold")),
                       tickfont=dict(size=12, color="#111111", weight="bold"),
                       categoryorder="array", categoryarray=_grupos_y),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            margin=dict(t=20, b=45, l=60, r=40), showlegend=False,
            font=dict(family="Helvetica Neue, sans-serif"))
        st.plotly_chart(fig_des, use_container_width=True)

        # leitura DESCRITIVA e cautelosa (sem cravar "a melhor")
        _ord = _dfd.copy()
        _cresce = _ord["sc_media"].is_monotonic_increasing
        _decresce = _ord["sc_media"].is_monotonic_decreasing
        _delta = _ord["sc_media"].iloc[-1] - _ord["sc_media"].iloc[0]
        # a "melhor" tem n pequeno? (alerta)
        _lider_n = _n_g.get(_grupo_lider, 0)
        _n_mediano = int(np.median(list(_n_g.values()))) if _n_g else 0
        _alerta_n = _lider_n < _n_mediano * 0.3

        if _cresce:
            _msg = (f"Na média do ensaio, a produtividade **tende a subir com a densidade** "
                    f"(+{_delta:.1f} sc/ha do menor ao maior alvo).")
        elif _decresce:
            _msg = (f"Na média do ensaio, a produtividade **tende a cair com a densidade** "
                    f"({_delta:.1f} sc/ha do menor ao maior alvo).")
        else:
            _msg = ("Na média do ensaio, a resposta à densidade **não é uniforme** — sobe em parte "
                    "da faixa e cai em outra.")
        _msg += (" Esta é a média geral: **cada híbrido pode responder de forma diferente**, e é a "
                 "análise por híbrido que dá a recomendação de densidade.")
        if _alerta_n:
            _msg += (f" ⚠️ A densidade {_grupo_lider} teve poucas parcelas (n={_lider_n}), então sua "
                     f"média é menos confiável — não a tome como a melhor sem olhar o n.")
        st.info(_msg)
        st.caption("Ponto = média da densidade · barra = ± erro padrão da média · linha tracejada = "
                   "média geral do ensaio. Barras que se sobrepõem indicam densidades que "
                   "provavelmente não diferem de verdade.")

    # ── Resumo numérico por grupo (fecha a Seção 2 com os números crus) ────────
    # Os três atos são a leitura visual; esta tabela é a conferência. Traz o tamanho do grupo,
    # a coesão da população real dentro dele (IQR e CV) e o desempenho com a incerteza da média.
    st.markdown(
        f'<p style="font-size:12px;font-weight:600;color:{_CL_SUB};text-transform:uppercase;'
        f'letter-spacing:0.07em;margin:1.4rem 0 0.4rem;">Resumo numérico por grupo</p>',
        unsafe_allow_html=True)

    with st.popover("ℹ️ Como ler a tabela", use_container_width=False):
        st.markdown("""
Esta tabela é a conferência numérica dos três atos acima.

- **n (parcelas)** → quantas parcelas caíram no grupo. Só entram as que têm contagem de estande.
- **n (com produção)** → dessas, quantas têm produtividade registrada. É esse número que sustenta
  a coluna sc/ha — se for muito menor que o n, a média do grupo vem de poucas parcelas.
- **Pop. média** e **Mediana** → o centro do grupo. Muito diferentes entre si indica assimetria
  (algumas parcelas puxando a média para um lado).
- **IQR pop.** → a faixa onde estão os 50% do meio das parcelas. Mostra a largura real do grupo.
  Aparece **—** quando as parcelas têm população praticamente idêntica.
- **CV pop. (%)** → dispersão relativa da população dentro do grupo. **Abaixo de 5%** o grupo é
  coeso (todas as parcelas viveram a mesma densidade); **acima de 10%** o grupo mistura densidades
  diferentes e as comparações que dependem dele ficam menos limpas.
- **sc/ha médio** e **EP** → desempenho e a incerteza da média. Grupos cujas faixas
  `média ± EP` se sobrepõem provavelmente não diferem de verdade.

**Cuidado com grupos pequenos.** Um grupo com poucas parcelas pode aparecer com a maior média só
por acaso, e o EP pequeno dele não significa confiança — significa poucos dados. Confira sempre o
n antes de tratar um grupo como o melhor.
""")

    _dfr = _df_grupos.sort_values("grupo", key=lambda s: s.str.replace("k", "").astype(int)).copy()
    _tot_parc = int(_dfr["n"].sum())

    # erro-padrão por grupo (recalculado aqui para a tabela existir mesmo sem o gráfico do Ato 3)
    _ep_tab, _n_sc_tab = {}, {}
    for _g in _dfr["grupo"]:
        _v = pd.to_numeric(ta_filtrado[(ta_filtrado["pop_grupo"] == _g)
                                       & (ta_filtrado["sc_ha"] > 0)]["sc_ha"],
                           errors="coerce").dropna()
        _n_sc_tab[_g] = len(_v)
        _ep_tab[_g] = float(_v.std(ddof=1) / np.sqrt(len(_v))) if len(_v) > 1 else 0.0

    # destaque = maior sc/ha média, MAS só se o grupo tiver tamanho comparável aos demais.
    # (na soja, um grupo com 1% das parcelas aparecia pintado de verde como "o melhor")
    _n_med = float(np.median(_dfr["n"])) if len(_dfr) else 0
    _lider_tab = _grupo_lider
    _lider_fragil = False
    if _lider_tab is not None:
        _n_lider = int(_dfr.loc[_dfr["grupo"] == _lider_tab, "n"].iloc[0]) if (_dfr["grupo"] == _lider_tab).any() else 0
        _lider_fragil = _n_med > 0 and _n_lider < _n_med * 0.3

    _linhas_tab = []
    for _, _r in _dfr.iterrows():
        _g = _r["grupo"]
        _iqr = ("—" if pd.isna(_r["q1"]) or int(_r["q1"]) == int(_r["q3"])
                else f"{int(_r['q1']):,} – {int(_r['q3']):,}".replace(",", "."))
        _linhas_tab.append({
            "Grupo": _g,
            "n (parcelas)": int(_r["n"]),
            "n (com produção)": int(_n_sc_tab.get(_g, 0)),
            "% do ensaio": f"{_r['n'] / _tot_parc * 100:.1f}%" if _tot_parc else "—",
            "Pop. média (pl/ha)": f"{int(_r['real_media']):,}".replace(",", ".") if pd.notna(_r["real_media"]) else "—",
            "Mediana (pl/ha)": f"{int(_r['real_mediana']):,}".replace(",", ".") if pd.notna(_r["real_mediana"]) else "—",
            "IQR pop.": _iqr,
            "CV pop. (%)": f"{_r['cv']:.1f}%",
            "sc/ha médio": f"{_r['sc_media']:.1f}" if pd.notna(_r["sc_media"]) else "—",
            "EP (sc/ha)": f"±{_ep_tab.get(_g, 0):.1f}" if _ep_tab.get(_g, 0) > 0 else "—",
        })
    _df_tab_res = pd.DataFrame(_linhas_tab)

    _gb_res = GridOptionsBuilder.from_dataframe(_df_tab_res)
    _gb_res.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000",
                   "fontFamily": "Helvetica Neue, sans-serif"})
    _grid_opts_res = dict(headerHeight=36, rowHeight=32, domLayout="normal",
                          suppressMenuHide=True, suppressColumnVirtualisation=True)
    # a linha só é pintada se o líder tiver tamanho de amostra comparável aos outros grupos
    if _lider_tab and not _lider_fragil:
        _grid_opts_res["getRowStyle"] = JsCode(f"""
            function(params) {{
                if (params.data["Grupo"] === "{_lider_tab}") {{
                    return {{'background': '#D5F5E3', 'fontWeight': '600'}};
                }}
            }}""")
    _gb_res.configure_grid_options(**_grid_opts_res)
    _go_res = _gb_res.build()
    _go_res["defaultColDef"]["headerClass"] = "ag-header-black"
    _go_res["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        _df_tab_res, gridOptions=_go_res,
        height=min(340, 36 + 32 * len(_df_tab_res) + 22),
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-row":              {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":             {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":       {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":        {"color": "#FFFFFF !important", "font-size": "13px !important",
                                            "font-weight": "700 !important"},
            ".ag-icon":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button": {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-icon-menu":               {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":             {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-cell":                    {"font-size": "13px !important"},
            ".ag-row":                     {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True)

    if _lider_tab and _lider_fragil:
        st.caption(f"A maior média ficou com o grupo **{_lider_tab}**, que tem poucas parcelas em "
                   f"relação aos demais — por isso não está destacado na tabela. Média de grupo "
                   f"pequeno oscila muito; confira o n antes de tratá-lo como o melhor.")
    else:
        st.caption("Linha destacada = grupo com a maior produtividade média. CV abaixo de 5% indica "
                   "grupo coeso (parcelas na mesma densidade real); acima de 10%, o grupo mistura "
                   "densidades e as comparações ficam menos limpas.")

    exportar_excel(_df_tab_res, "densidade_resumo_grupos.xlsx",
                   "⬇️ Exportar resumo por grupo", key="den_resumo_grupos_xlsx")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — DISTRIBUIÇÃO POR DENSIDADE + LSD (a diferença entre densidades é real?)
# ══════════════════════════════════════════════════════════════════════════════
# O eixo Y são os GRUPOS formados na Seção 2 (K-Means na população REAL contada), não os
# tratamentos-alvo. Uma parcela que mirou 82k e emergiu com 41k entra no grupo 41k — a
# densidade que a planta de fato experimentou. Toda a produtividade aqui é lida por esse
# agrupamento; a população-alvo não entra em nenhuma conta desta seção.
from scipy import stats as _stats

secao_titulo("Distribuição da Produtividade",
             "A produtividade em cada população, e o quanto as diferenças são confiáveis",
             contexto_str)

_hib_dist = sorted(ta_filtrado[ta_filtrado["sc_ha"] > 0]["dePara"].dropna().unique().tolist()) \
    if "sc_ha" in ta_filtrado.columns else []

if not _hib_dist:
    st.info("Nenhum híbrido com produtividade nos filtros ativos.")
else:
    _c_d1, _c_d2, _c_d3 = st.columns([3, 2, 2])
    with _c_d1:
        hib_dist = st.selectbox("Híbrido", options=_hib_dist, key="den_dist_hib",
                                label_visibility="collapsed",
                                placeholder="Selecione o híbrido...")
    with _c_d2:
        mostrar_lsd = st.checkbox("Mostrar corte LSD", value=True, key="den_dist_lsd")
    with _c_d3:
        mostrar_pts = st.checkbox("Mostrar parcelas individuais", value=True, key="den_dist_pts")

    _cols_d = [c for c in ["sc_ha", "pop_grupo", "_pop_final", "cod_fazenda", "nomeFazenda",
                           "cidade_nome", "estado_sigla"] if c in ta_filtrado.columns]
    df_d = ta_filtrado[(ta_filtrado["dePara"] == hib_dist)
                       & (ta_filtrado["sc_ha"] > 0)][_cols_d].copy()

    # os grupos vêm da Seção 2 (população real); só entram os que têm parcelas deste híbrido
    _grupos_h = [g for g in _grupos_encontrados if g in df_d["pop_grupo"].values]
    _dados_pg = {g: df_d[df_d["pop_grupo"] == g]["sc_ha"].values for g in _grupos_h}
    _df_pg = {g: df_d[df_d["pop_grupo"] == g].reset_index(drop=True) for g in _grupos_h}

    if len([v for v in _dados_pg.values() if len(v) >= 2]) < 2:
        st.info(f"**{hib_dist}** não tem densidades suficientes com repetição para comparar.")
    else:
        # ── Estatísticas ──────────────────────────────────────────────────────
        _amostras = [v for v in _dados_pg.values() if len(v) >= 2]
        _f, _p = _stats.f_oneway(*_amostras) if len(_amostras) >= 2 else (np.nan, np.nan)

        def _lsd_dens(df, fator="pop_grupo", bloco="cod_fazenda", col="sc_ha", alpha=0.05):
            """LSD em blocos por local: desconta a variação entre ambientes e isola a densidade."""
            try:
                d = df[[col, fator, bloco]].dropna()
                d = d[d[col] > 0].reset_index(drop=True)
                if d.empty or d[fator].nunique() < 2 or d[bloco].nunique() < 2:
                    return np.nan
                y = d[col].values.astype(float)
                Xf = pd.get_dummies(d[fator], drop_first=True).values.astype(float)
                Xb = pd.get_dummies(d[bloco], drop_first=True).values.astype(float)
                X = np.hstack([np.ones((len(y), 1)), Xf, Xb])
                beta, _, rank, _ = np.linalg.lstsq(X, y, rcond=None)
                ss_res = np.sum((y - X @ beta) ** 2)
                gl_res = len(y) - rank
                if gl_res <= 0:
                    return np.nan
                qmr = ss_res / gl_res
                n_blk = d[bloco].nunique()
                t_crit = _stats.t.ppf(1 - alpha / 2, df=gl_res)
                return round(t_crit * np.sqrt(2 * qmr / n_blk), 1)
            except Exception:
                return np.nan

        _lsd = _lsd_dens(df_d)

        _medias = {g: float(np.mean(v)) for g, v in _dados_pg.items() if len(v) > 0}
        _melhor_g = max(_medias, key=_medias.get) if _medias else None
        _media_lider = _medias[_melhor_g] if _melhor_g else None

        # ── Título dinâmico em linguagem natural ──────────────────────────────
        _p_sig = pd.notna(_p) and _p < 0.05
        if _melhor_g and _media_lider is not None:
            if _p_sig:
                _tit_d = (f"Em <b>{hib_dist}</b>, plantar em "
                          f"<b style='color:{_CL_VERDE}'>{_melhor_g}</b> resultou na maior "
                          f"produtividade — diferença estatisticamente significativa")
                _cor_sub_d = _CL_VERDE
                _sub_d = (f"A densidade influenciou a produtividade (p = {_p:.3f}). "
                          f"Grupo líder: {_melhor_g} com {_media_lider:.1f} sc/ha de média.")
            else:
                _tit_d = (f"Em <b>{hib_dist}</b>, a densidade <b>não alterou</b> "
                          f"significativamente a produtividade")
                _cor_sub_d = "#B45309"
                _sub_d = (f"A diferença entre grupos não é estatisticamente significativa "
                          f"(p = {_p:.3f}). Grupo com maior média: {_melhor_g} "
                          f"({_media_lider:.1f} sc/ha).")
        else:
            _tit_d = f"Distribuição de produtividade por grupo de densidade · {hib_dist}"
            _cor_sub_d = _CL_SUB
            _sub_d = "Dados insuficientes para análise estatística."

        st.markdown(f"""
<div style="margin:1.2rem 0 0.8rem;">
    <p style="font-size:1.15rem;font-weight:700;color:{_CL_TEXTO};margin:0;line-height:1.4;">
        {_tit_d}</p>
    <p style="font-size:13px;color:{_cor_sub_d};margin:5px 0 0;font-weight:500;">
        {_sub_d}{'' if pd.isna(_lsd) else f' · LSD = {_lsd:.1f} sc/ha'}</p>
</div>""", unsafe_allow_html=True)

        with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
            st.markdown("""
**Os três controles acima do gráfico**

- **Híbrido** — o gráfico é de um material por vez.
- **Mostrar corte LSD** — liga a linha vermelha vertical, explicada abaixo.
- **Mostrar parcelas individuais** — liga os pontos. Desligado, ficam só as faixas e as médias.

Cada **ponto** é uma parcela colhida no campo. A posição horizontal é a produtividade (sc/ha).

**O que as formas significam:**
- **Faixa clara** → amplitude total: da menor à maior produtividade do grupo
- **Faixa escura** → onde estão 50% das parcelas (IQR)
- **Linha sólida** → média do grupo
- **Linha tracejada** → mediana (o valor do meio)
- **Linha vermelha** → corte do LSD: grupos à esquerda dela produziram significativamente menos
  que o líder; grupos à direita são estatisticamente equivalentes a ele

**As densidades do eixo são os grupos da Seção 2**, formados pela **população real contada**, não
pelo tratamento planejado. Uma parcela que mirou 82 mil e emergiu com 41 mil está no grupo 41 mil —
é a densidade que a planta de fato viveu que explica a produtividade dela.

**ANOVA** verifica se existe alguma diferença real entre os grupos:
- Subtítulo **verde** → a densidade influenciou a produtividade (p < 0,05)
- Subtítulo **laranja** → diferença não comprovada; pode ser variação normal do ambiente

**LSD** (diferença mínima significativa) é a régua: diferença menor que o LSD está dentro da
margem de erro do ensaio. Ele é calculado em blocos por local, então já desconta a variação
entre ambientes.

**Dica:** faixa IQR larga indica alta variação entre locais — o híbrido pode ser instável
nessa densidade.
""")

        # ── Gráfico: strip horizontal (grupos no Y, produtividade no X) ───────
        fig_d = go_plt.Figure()
        _grupos_rev = list(reversed(_grupos_h))  # maior densidade no topo

        for i, g in enumerate(_grupos_rev):
            vals = _dados_pg.get(g, np.array([]))
            if len(vals) == 0:
                continue
            _is_lider = g == _melhor_g
            _cor = _CL_VERDE if _is_lider else "#94A3B8"
            _op_ampl = 0.18 if _is_lider else 0.10
            _op_iqr = 0.30 if _is_lider else 0.16

            _media = float(np.mean(vals))
            _mediana = float(np.median(vals))
            _q1 = float(np.percentile(vals, 25))
            _q3 = float(np.percentile(vals, 75))
            _n = len(vals)

            # amplitude total
            fig_d.add_shape(type="rect", x0=float(np.min(vals)), x1=float(np.max(vals)),
                            y0=i - 0.32, y1=i + 0.32,
                            fillcolor=_cor, opacity=_op_ampl, line=dict(width=0))
            # IQR
            fig_d.add_shape(type="rect", x0=_q1, x1=_q3, y0=i - 0.32, y1=i + 0.32,
                            fillcolor=_cor, opacity=_op_iqr, line=dict(width=0))
            # média (sólida) e mediana (tracejada)
            fig_d.add_shape(type="line", x0=_media, x1=_media, y0=i - 0.38, y1=i + 0.38,
                            line=dict(color=_cor, width=3 if _is_lider else 2))
            fig_d.add_shape(type="line", x0=_mediana, x1=_mediana, y0=i - 0.38, y1=i + 0.38,
                            line=dict(color=_cor, width=1.8, dash="dash"))

            # parcelas individuais — hover com local e a POPULAÇÃO REAL da parcela
            if mostrar_pts:
                _dfg = _df_pg.get(g, pd.DataFrame())
                _jit = np.random.uniform(-0.22, 0.22, size=len(vals))
                _cd_cols = [c for c in ["cod_fazenda", "cidade_nome", "estado_sigla", "_pop_final"]
                            if c in _dfg.columns]
                if not _dfg.empty and len(_cd_cols) == 4:
                    _cd = _dfg[_cd_cols].values
                    _ht = ("<b>%{customdata[0]}</b> · %{customdata[1]}, %{customdata[2]}<br>"
                           "sc/ha: <b>%{x:.1f}</b><br>"
                           "População real: %{customdata[3]:,.0f} pl/ha<extra></extra>")
                else:
                    _cd = None
                    _ht = f"<b>{g}</b><br>sc/ha: %{{x:.1f}}<extra></extra>"
                fig_d.add_trace(go_plt.Scatter(
                    x=vals, y=[i + j for j in _jit], mode="markers",
                    marker=dict(color=_cor, size=9 if _is_lider else 8,
                                opacity=0.90 if _is_lider else 0.75,
                                line=dict(color="#FFFFFF", width=0.5)),
                    customdata=_cd, hovertemplate=_ht, showlegend=False))

            # rótulo direto: média + n
            _cor_lab = _CL_VERDE if _is_lider else _CL_SUB
            fig_d.add_annotation(
                x=_media, y=i + 0.42,
                text=(f"<b style='color:{_cor_lab};font-size:14px'>{_media:.1f}</b>"
                      f"<b style='color:{_CL_SUB};font-size:14px'> ({_n})</b>"),
                showarrow=False, xanchor="center", yanchor="bottom",
                font=dict(size=14, color=_CL_TEXTO, weight="bold"))

        # ── Corte LSD ─────────────────────────────────────────────────────────
        if mostrar_lsd and pd.notna(_lsd) and _melhor_g:
            _lsd_x = _media_lider - _lsd
            fig_d.add_vline(x=_lsd_x, line=dict(color="#E74C3C", width=1.8, dash="dot"))
            fig_d.add_annotation(
                x=_lsd_x, y=len(_grupos_rev) - 0.1,
                text=f"<b style='color:#E74C3C;font-size:13px'>LSD = {_lsd:.1f} sc/ha</b>",
                showarrow=False, xanchor="right", yanchor="top", xshift=-6,
                font=dict(size=13, color="#E74C3C", weight="bold"), align="right")

        _all_sc = np.concatenate([v for v in _dados_pg.values() if len(v) > 0])
        _x_min_d = float(np.min(_all_sc)) - 2
        _x_max_d = float(np.max(_all_sc)) + 8  # espaço para os rótulos

        fig_d.update_layout(
            height=max(280, len(_grupos_rev) * 95 + 80),
            plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
            margin=dict(t=20, b=50, l=70, r=20),
            xaxis=dict(range=[_x_min_d, _x_max_d],
                       title=dict(text="<b>Produtividade (sc/ha)</b>",
                                  font=dict(size=14, color="#1A1A1A", weight="bold")),
                       tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                       showgrid=True, gridcolor="#EEEEEE", zeroline=False, showline=False),
            yaxis=dict(tickmode="array", tickvals=list(range(len(_grupos_rev))),
                       ticktext=[f"<b style='color:{_CL_VERDE}'>{g}</b>" if g == _melhor_g
                                 else f"<b style='color:{_CL_SUB}'>{g}</b>" for g in _grupos_rev],
                       tickfont=dict(size=14, weight="bold"),
                       showgrid=False, zeroline=False, showline=False))
        st.plotly_chart(fig_d, use_container_width=True)

        # ── Legenda + dicionário de locais ────────────────────────────────────
        _col_cap_d, _col_dic_d = st.columns([3, 1])
        with _col_cap_d:
            st.caption("Cada ponto = uma parcela · linha sólida = média · linha tracejada = mediana · "
                       "faixa escura = 50% das parcelas (IQR) · faixa clara = amplitude total · "
                       "linha vermelha = corte do LSD (5%). Densidades do eixo = grupos formados "
                       "pela população real contada.")
        with _col_dic_d:
            _dic_cols_d = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]
                           if c in df_d.columns]
            if _dic_cols_d:
                _df_dic_d = (df_d[_dic_cols_d].drop_duplicates()
                             .sort_values(_dic_cols_d[0])
                             .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                              "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                             .reset_index(drop=True))
                with st.popover(f"📍 {len(_df_dic_d)} locais", use_container_width=True):
                    st.dataframe(_df_dic_d, hide_index=True, use_container_width=True)

        # ── Leitura automática de equivalência ────────────────────────────────
        if pd.notna(_lsd) and _melhor_g:
            _equiv = [g for g in _grupos_h if g != _melhor_g
                      and (_medias[_melhor_g] - _medias[g]) < _lsd]
            if _equiv:
                _menor_equiv = min(_equiv + [_melhor_g],
                                   key=lambda g: int(str(g).replace("k", "")) if
                                   str(g).replace("k", "").isdigit() else 10**9)
                st.info(f"Para **{hib_dist}**, a densidade **{_melhor_g}** teve a maior média "
                        f"({_medias[_melhor_g]:.1f} sc/ha), mas {', '.join(_equiv)} "
                        f"{'estão' if len(_equiv) > 1 else 'está'} dentro do LSD — ou seja, "
                        f"**estatisticamente equivalente{'s' if len(_equiv) > 1 else ''}** a ela. "
                        f"Na prática, a menor densidade entre as equivalentes (**{_menor_equiv}**) "
                        f"pode ser preferível: mesmo rendimento com menos semente.")
            else:
                st.success(f"Para **{hib_dist}**, a densidade **{_melhor_g}** "
                           f"({_medias[_melhor_g]:.1f} sc/ha) foi significativamente superior às "
                           f"demais — todas as diferenças passaram o LSD de {_lsd:.1f} sc/ha.")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — RESPOSTA À DENSIDADE POR HÍBRIDO (regressão polinomial)
# ══════════════════════════════════════════════════════════════════════════════
# Aqui a população deixa de ser um grupo e vira número contínuo: cada parcela é um ponto
# (x = população REAL contada, y = produtividade) e sobre eles ajustamos uma parábola por
# híbrido. O ponto de máximo (-b/2a) só é anunciado depois de quatro travas: parcelas
# suficientes, curvatura estatisticamente significativa, curva côncava e topo DENTRO da
# faixa testada. Sem isso, uma parábola ajusta em qualquer nuvem e sempre devolve um "ótimo".

# Régua de aceitação do ajuste (o híbrido só é ajustado se passar nas três)
_REG_MIN_PLOTS = 10      # nº mínimo de parcelas com produtividade e estande contado
_REG_MIN_POPS = 5        # nº mínimo de populações reais distintas (a curva precisa de eixo X)
_REG_MIN_AMPL = 8000     # amplitude mínima da população testada, em plantas/ha
_REG_ALFA = 0.05         # nível para a curvatura ser considerada real


def _mil(v):
    """1234567 -> '1.234.567' (separador de milhar BR, sem mexer nas vírgulas do texto)."""
    return f"{v:,.0f}".replace(",", ".")


def _ajuste_densidade(x_plantas, y_scha, alfa=_REG_ALFA):
    """Ajusta produtividade (sc/ha) em função da população REAL contada (plantas/ha).

    O ajuste é feito com x em MILHARES de plantas: com x na casa de 60.000, o termo x²
    chega a 3,6 bilhões, a matriz fica mal condicionada e o coeficiente quadrático sai
    instável. Os resultados são devolvidos já convertidos para plantas/ha.

    Testa a CURVATURA antes de anunciar população ótima: se o termo quadrático não for
    significativo, a curva não existe — o que existe é uma reta (ou nada).

    Devolve também a matriz de covariância (escala de milhares), usada para desenhar o
    intervalo de confiança da curva e para o erro-padrão da população ótima (delta method).
    """
    x = pd.to_numeric(pd.Series(x_plantas), errors="coerce")
    y = pd.to_numeric(pd.Series(y_scha), errors="coerce")
    ok = x.notna() & y.notna() & (x > 0) & (y > 0)
    x, y = x[ok].to_numpy(float), y[ok].to_numpy(float)

    r = {"n": len(x), "n_pops": int(len(np.unique(x))),
         "x_min": float(x.min()) if len(x) else np.nan,
         "x_max": float(x.max()) if len(x) else np.nan,
         "x_medio": float(x.mean()) if len(x) else np.nan,
         "a": np.nan, "b": np.nan, "c": np.nan, "r2": np.nan, "r2_aj": np.nan,
         "p_curvatura": np.nan, "p_linear": np.nan, "inclinacao": np.nan,
         "pop_otima": np.nan, "pop_otima_ep": np.nan, "prod_otima": np.nan,
         "dentro_faixa": False, "resp_marginal": np.nan,
         "veredito": "sem ajuste", "detalhe": "",
         "_coefs_m": None, "_cov_m": None, "_gl": np.nan}

    ampl = (x.max() - x.min()) if len(x) else 0
    if len(x) < _REG_MIN_PLOTS or r["n_pops"] < _REG_MIN_POPS or ampl < _REG_MIN_AMPL:
        r["veredito"] = "dados insuficientes"
        r["detalhe"] = (f"{len(x)} parcelas, {r['n_pops']} populações distintas, "
                        f"amplitude de {_mil(ampl)} plantas/ha")
        return r

    xm = x / 1000.0  # milhares de plantas/ha

    def _ols(X, yv):
        beta, _, rank, _ = np.linalg.lstsq(X, yv, rcond=None)
        gl = len(yv) - rank
        if gl <= 0:
            return None
        res = yv - X @ beta
        qmr = float(res @ res) / gl
        try:
            cov = np.linalg.inv(X.T @ X) * qmr
        except np.linalg.LinAlgError:
            return None
        ep = np.sqrt(np.diag(cov))
        sq_tot = float(((yv - yv.mean()) ** 2).sum())
        r2 = 1 - float(res @ res) / sq_tot if sq_tot > 0 else np.nan
        return beta, ep, gl, r2, rank, cov

    # ── modelo quadrático ─────────────────────────────────────────────────────
    Xq = np.column_stack([np.ones(len(xm)), xm, xm ** 2])
    _q = _ols(Xq, y)
    if _q is None:
        r["veredito"] = "dados insuficientes"
        return r
    beta_q, ep_q, gl_q, r2_q, rank_q, cov_q = _q
    c_, b_, a_ = beta_q
    r["_coefs_m"], r["_cov_m"], r["_gl"] = beta_q, cov_q, gl_q
    r["r2"] = round(float(r2_q), 3)
    r["r2_aj"] = round(1 - (1 - r2_q) * (len(y) - 1) / max(len(y) - rank_q, 1), 3)
    t_a = a_ / ep_q[2] if ep_q[2] > 0 else np.nan
    r["p_curvatura"] = float(2 * (1 - _stats.t.cdf(abs(t_a), df=gl_q))) if pd.notna(t_a) else np.nan
    r["a"], r["b"], r["c"] = a_ / 1e6, b_ / 1e3, float(c_)  # coeficientes em plantas/ha

    # ── modelo linear (referência quando a curvatura não é real) ──────────────
    Xl = np.column_stack([np.ones(len(xm)), xm])
    _l = _ols(Xl, y)
    if _l is not None:
        beta_l, ep_l, gl_l, r2_l, _, _ = _l
        r["inclinacao"] = float(beta_l[1])  # sc/ha por mil plantas
        t_l = beta_l[1] / ep_l[1] if ep_l[1] > 0 else np.nan
        r["p_linear"] = float(2 * (1 - _stats.t.cdf(abs(t_l), df=gl_l))) if pd.notna(t_l) else np.nan
        r["_coefs_lin_m"] = beta_l

    # resposta marginal na população média (sc/ha por mil plantas a mais)
    r["resp_marginal"] = float(2 * a_ * (r["x_medio"] / 1000) + b_)

    # ── leitura ───────────────────────────────────────────────────────────────
    curvatura_real = pd.notna(r["p_curvatura"]) and r["p_curvatura"] < alfa

    if curvatura_real and a_ < 0:
        _ot_m = -b_ / (2 * a_)             # em milhares
        r["pop_otima"] = float(_ot_m * 1000)
        r["prod_otima"] = float(a_ * _ot_m ** 2 + b_ * _ot_m + c_)
        # erro-padrão do ótimo pelo método delta: θ = -b/(2a)
        grad = np.array([0.0, -1.0 / (2 * a_), b_ / (2 * a_ ** 2)])
        var_ot = float(grad @ cov_q @ grad)
        r["pop_otima_ep"] = float(np.sqrt(var_ot) * 1000) if var_ot > 0 else np.nan
        r["dentro_faixa"] = bool(r["x_min"] <= r["pop_otima"] <= r["x_max"])
        if r["dentro_faixa"]:
            r["veredito"] = "ponto de máximo"
            r["detalhe"] = (f"produtividade máxima em {_mil(r['pop_otima'])} plantas/ha, "
                            f"dentro da faixa testada")
        else:
            lado = "acima" if r["pop_otima"] > r["x_max"] else "abaixo"
            r["veredito"] = "máximo fora da faixa"
            r["detalhe"] = (f"o máximo calculado ({_mil(r['pop_otima'])}) cai {lado} das "
                            f"densidades testadas — é extrapolação, não recomendação")
    elif curvatura_real and a_ > 0:
        r["veredito"] = "curva em U"
        r["detalhe"] = ("a curva tem ponto de mínimo, não de máximo — comportamento atípico para "
                        "densidade; checar efeito de local ou poucas parcelas nas pontas")
    else:
        lin_real = pd.notna(r["p_linear"]) and r["p_linear"] < alfa
        if lin_real and r["inclinacao"] > 0:
            r["veredito"] = "resposta crescente"
            r["detalhe"] = (f"sobe {r['inclinacao']:.1f} sc/ha a cada mil plantas a mais, sem sinal "
                            f"de saturação dentro da faixa testada")
        elif lin_real and r["inclinacao"] < 0:
            r["veredito"] = "resposta decrescente"
            r["detalhe"] = (f"cai {abs(r['inclinacao']):.1f} sc/ha a cada mil plantas a mais dentro "
                            f"da faixa testada")
        else:
            r["veredito"] = "sem resposta"
            r["detalhe"] = "a produtividade não variou de forma consistente com a densidade"
    return r


def _curva_prevista(r, xs_plantas, ic=True):
    """Prevê a curva em xs (plantas/ha) e, se pedido, a meia-largura do IC95 da média."""
    if r.get("_coefs_m") is None:
        return None, None
    xm = np.asarray(xs_plantas, dtype=float) / 1000.0
    X = np.column_stack([np.ones(len(xm)), xm, xm ** 2])
    yhat = X @ r["_coefs_m"]
    if not ic or r.get("_cov_m") is None or pd.isna(r.get("_gl")):
        return yhat, None
    var = np.einsum("ij,jk,ik->i", X, r["_cov_m"], X)
    t_crit = _stats.t.ppf(0.975, df=r["_gl"])
    return yhat, t_crit * np.sqrt(np.clip(var, 0, None))


def _descontar_local(df, col_y="sc_ha", col_bloco="cod_fazenda"):
    """Remove o efeito do local: y_ajustado = y − média do local + média geral.

    Sem isso a curva mistura duas coisas: o efeito da densidade e o efeito do ambiente. Se as
    densidades altas caíram em locais bons (ou ruins), a diferença entre locais entra na conta
    como se fosse resposta à população. É a versão simples do bloco — mesma lógica do LSD em
    blocos usado na Seção 3.
    """
    d = df.copy()
    if col_bloco not in d.columns:
        return d
    media_geral = pd.to_numeric(d[col_y], errors="coerce").mean()
    d["_y_aj"] = (pd.to_numeric(d[col_y], errors="coerce")
                  - d.groupby(col_bloco)[col_y].transform("mean") + media_geral)
    return d


_COR_VEREDITO = {
    "ponto de máximo": "#1E7A34",
    "resposta crescente": "#2976B6",
    "resposta decrescente": "#C0201E",
    "máximo fora da faixa": "#B45309",
    "curva em U": "#B45309",
    "sem resposta": "#94A3B8",
    "dados insuficientes": "#CBD5E1",
    "sem ajuste": "#CBD5E1",
}

secao_titulo("Curva de Resposta à População",
             "A curva de produtividade de cada material ao longo das populações",
             contexto_str)

with st.popover("ℹ️ O que esta seção faz", use_container_width=False):
    st.markdown("""
**Os controles desta seção**

- **Descontar o efeito do local** — centra a produtividade de cada parcela pela média do local
  dela, para a curva medir densidade e não ambiente. Ligado por padrão, e faz diferença grande.
- **Mostrar intervalo de confiança da curva** — a faixa clara em volta da linha. Some sozinho com
  mais de três híbridos na tela, senão vira mancha.
- **Híbridos** (no Ato 2) — quais curvas aparecem sobrepostas.
- **Mostrar colunas de verificação** (na tabela) — acrescenta margem do ótimo, faixa testada,
  p da curvatura e a leitura.

As seções anteriores compararam **grupos** de densidade. Aqui a população entra como número
contínuo: cada parcela é um ponto e, sobre esses pontos, ajustamos uma **curva** de produtividade
para cada híbrido.

**Por que uma curva e não uma reta.** A resposta do milho à população não é infinita: sobe, chega
a um topo e cai (competição por luz, água e nutrientes). Uma parábola descreve esse formato com
dois números úteis — o topo (**população ótima**) e a rapidez com que a curva sobe e desce.

**Por que a população REAL e não o alvo.** A planta responde ao que emergiu, não ao que foi
programado. Uma parcela que mirou 82 mil e emergiu com 60 mil experimentou 60 mil.

**Antes de anunciar um ótimo, testamos se a curva existe.** O ponto de máximo só aparece quando:

1. o híbrido tem parcelas e populações distintas suficientes (a régua está no rodapé da tabela);
2. o **termo quadrático é estatisticamente significativo** (p < 0,05) — se não for, os dados não
   sustentam uma curva, e o que existe é uma reta ou nada;
3. a curva é **côncava** (tem topo, não fundo);
4. o topo cai **dentro da faixa de densidades testada**. Fora dela é extrapolação: a conta devolve
   um número, mas o ensaio nunca visitou aquela população.

**A população ótima vem com margem de erro.** O ± ao lado dela é o erro-padrão do ponto de máximo.
Margem larga significa que o topo da curva é achatado — muitas densidades produzem praticamente o
mesmo, e a escolha pode ser feita por outros critérios (custo de semente, risco de acamamento).

**É um máximo TÉCNICO, não econômico.** Ele ignora o custo da semente. O ótimo econômico é sempre
**menor** que o técnico: os últimos sacos custam mais semente do que devolvem em grão. Use este
número como teto agronômico, não como recomendação de plantio.
""")

_c_r1, _c_r2 = st.columns([2, 3])
with _c_r1:
    _desc_local = st.checkbox("Descontar o efeito do local", value=True, key="den_reg_bloco",
                              help="Centra a produtividade de cada parcela pela média do seu local, "
                                   "para a curva medir densidade e não ambiente.")
with _c_r2:
    _mostrar_ic = st.checkbox("Mostrar intervalo de confiança da curva", value=True,
                              key="den_reg_ic")

_base_reg = ta_filtrado[(ta_filtrado["sc_ha"] > 0) & (ta_filtrado["_pop_final"] > 0)].copy()

if _base_reg.empty or "dePara" not in _base_reg.columns:
    st.info("Sem parcelas com produtividade e contagem de estande nos filtros ativos.")
else:
    if _desc_local:
        _base_reg = _descontar_local(_base_reg, col_y="sc_ha", col_bloco="cod_fazenda")
        _COL_Y = "_y_aj"
    else:
        _COL_Y = "sc_ha"

    _ajustes = {}
    for _hib, _g in _base_reg.groupby("dePara"):
        _r = _ajuste_densidade(_g["_pop_final"], _g[_COL_Y])
        _r["locais"] = int(_g["cod_fazenda"].nunique()) if "cod_fazenda" in _g else 0
        _r["status"] = (_g["status_material"].dropna().iloc[0]
                        if "status_material" in _g and _g["status_material"].notna().any() else "")
        _ajustes[_hib] = _r

    _ajustaveis = [h for h, r in _ajustes.items()
                   if r["veredito"] not in {"dados insuficientes", "sem ajuste"}]

    if not _ajustaveis:
        st.info("Nenhum híbrido tem parcelas suficientes para ajustar a curva nos filtros ativos. "
                f"A régua exige {_REG_MIN_PLOTS} parcelas, {_REG_MIN_POPS} populações distintas e "
                f"{_REG_MIN_AMPL/1000:.0f} mil plantas/ha de amplitude.")
    else:
        # ── Ato 1 — Panorama: a curva de cada híbrido lado a lado ─────────────
        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
            'letter-spacing:0.07em;margin:1.2rem 0 6px;">Ato 1 — O formato da resposta</p>',
            unsafe_allow_html=True)
        st.markdown("Cada painel é um híbrido, na mesma escala. O que interessa aqui é o "
                    "**formato**: quem faz arco (tem topo), quem ainda sobe na maior densidade "
                    "testada e quem não responde.")

        from plotly.subplots import make_subplots as _make_subplots

        _n_h = len(_ajustaveis)
        _ncols = min(3, _n_h)
        _nrows = int(np.ceil(_n_h / _ncols))

        # o veredito vai no TÍTULO do painel, não dentro dele: quando a curva cai na parte de
        # baixo da área de plotagem, ela passa por cima de um rótulo ancorado no canto inferior.
        _titulos_sm = []
        for _hib in _ajustaveis:
            _rr = _ajustes[_hib]
            _cor = _COR_VEREDITO.get(_rr["veredito"], "#94A3B8")
            _r2_txt = f"{_rr['r2']:.2f}" if pd.notna(_rr["r2"]) else "—"
            _titulos_sm.append(
                f"<b>{_hib}</b><br>"
                f"<span style='font-size:10px;color:{_cor}'>{_rr['veredito']}</span>"
                f"<span style='font-size:10px;color:#6B7280'> · R² {_r2_txt} · n={_rr['n']}</span>")

        _fig_sm = _make_subplots(rows=_nrows, cols=_ncols, shared_xaxes=False, shared_yaxes=True,
                                 subplot_titles=_titulos_sm,
                                 horizontal_spacing=0.06, vertical_spacing=0.20)

        # escala Y compartilhada (é ela que permite comparar formatos entre painéis), mas cortada
        # em p1–p99: uma parcela extrema em um único híbrido achataria todas as curvas.
        _y_all = pd.to_numeric(
            _base_reg[_base_reg["dePara"].isin(_ajustaveis)][_COL_Y], errors="coerce").dropna()
        _y_lo, _y_hi = float(_y_all.quantile(0.01)), float(_y_all.quantile(0.99))
        _y_pad = (_y_hi - _y_lo) * 0.10 + 2

        for _i, _hib in enumerate(_ajustaveis):
            _rr = _ajustes[_hib]
            _row, _col = _i // _ncols + 1, _i % _ncols + 1
            _cor = _COR_VEREDITO.get(_rr["veredito"], "#94A3B8")
            _g = _base_reg[_base_reg["dePara"] == _hib]

            _fig_sm.add_trace(go_plt.Scatter(
                x=_g["_pop_final"], y=_g[_COL_Y], mode="markers",
                marker=dict(color="#CBD5E1", size=5, opacity=0.75),
                showlegend=False, hoverinfo="skip"), row=_row, col=_col)

            _xs = np.linspace(_rr["x_min"], _rr["x_max"], 80)
            _yh, _ = _curva_prevista(_rr, _xs, ic=False)
            if _yh is not None:
                _fig_sm.add_trace(go_plt.Scatter(
                    x=_xs, y=_yh, mode="lines", line=dict(color=_cor, width=2.6),
                    showlegend=False,
                    hovertemplate="%{x:,.0f} pl/ha<br>%{y:.1f} sc/ha<extra></extra>"),
                    row=_row, col=_col)

            if _rr["veredito"] == "ponto de máximo":
                _fig_sm.add_trace(go_plt.Scatter(
                    x=[_rr["pop_otima"]], y=[_rr["prod_otima"]], mode="markers",
                    marker=dict(symbol="star", size=15, color=_cor,
                                line=dict(color="#FFFFFF", width=1)),
                    showlegend=False,
                    hovertemplate=(f"ótimo: {_mil(_rr['pop_otima'])} pl/ha<br>"
                                   f"{_rr['prod_otima']:.1f} sc/ha<extra></extra>")),
                    row=_row, col=_col)

        # eixo X em milhares (o tickformat com vírgula mostrava "40,000", fora do padrão BR).
        # Marcas de 10 em 10 mil sobre a faixa toda; cada painel exibe só as que couberem.
        _x_all_sm = pd.to_numeric(
            _base_reg[_base_reg["dePara"].isin(_ajustaveis)]["_pop_final"],
            errors="coerce").dropna()
        _tv_sm = list(range(int(np.floor(_x_all_sm.min() / 10000) * 10000),
                            int(np.ceil(_x_all_sm.max() / 10000) * 10000) + 1, 10000))

        _fig_sm.update_yaxes(range=[_y_lo - _y_pad, _y_hi + _y_pad], gridcolor="#EEEEEE",
                             tickfont=dict(size=10, color="#111111"))
        _fig_sm.update_xaxes(gridcolor="#EEEEEE", tickfont=dict(size=9, color="#111111"),
                             tickvals=_tv_sm, ticktext=[f"{v/1000:.0f}k" for v in _tv_sm])
        _fig_sm.update_annotations(font=dict(size=12, color="#1A1A1A"))
        _fig_sm.update_layout(
            height=max(300, _nrows * 275), plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
            margin=dict(t=58, b=40, l=55, r=20))
        st.plotly_chart(_fig_sm, use_container_width=True)
        st.caption("Cinza = parcelas · linha = curva ajustada · ★ = população ótima (só quando o "
                   "topo é real e cai dentro da faixa testada). Verde = tem topo · azul = ainda "
                   "sobe · vermelho = cai · laranja = topo fora da faixa ou curva atípica · "
                   "cinza = sem resposta." +
                   (" Produtividade centrada pela média do local." if _desc_local else ""))

        st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

        # ── Ato 2 — Curvas sobrepostas (molde da soja) ─────────────────────────
        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
            'letter-spacing:0.07em;margin:1.2rem 0 6px;">Ato 2 — As curvas lado a lado</p>',
            unsafe_allow_html=True)
        st.markdown("Os híbridos escolhidos na mesma escala, uma cor cada. Aqui dá para comparar "
                    "não só onde está o topo de cada um, mas **a que altura** ele está.")

        _CORES_REG = ["#2976B6", "#27AE60", "#F39C12", "#E74C3C", "#9B59B6", "#1ABC9C"]

        _col_ms, _ = st.columns([2, 3])
        with _col_ms:
            _hibs_reg = st.multiselect(
                "Híbridos", options=_ajustaveis,
                default=_ajustaveis[:min(4, len(_ajustaveis))], key="den_reg_hibs")

        if not _hibs_reg:
            st.info("Selecione ao menos um híbrido para ver as curvas.")
        else:
            # o IC de várias curvas sobrepostas vira mancha: só desenha com poucos híbridos
            _ic_ok = _mostrar_ic and len(_hibs_reg) <= 3
            fig_reg = go_plt.Figure()
            _res_reg = []   # (híbrido, cor, ajuste) dos que têm ponto de máximo válido
            _y_curvas = []  # extremos das curvas, para o range do eixo Y

            for _i, _hib in enumerate(_hibs_reg):
                _rr = _ajustes[_hib]
                _cor = _CORES_REG[_i % len(_CORES_REG)]
                _gsel = _base_reg[_base_reg["dePara"] == _hib]

                _xs = np.linspace(_rr["x_min"], _rr["x_max"], 200)
                _yh, _meia = _curva_prevista(_rr, _xs, ic=_ic_ok)
                if _yh is None:
                    continue
                _y_curvas += [float(np.min(_yh)), float(np.max(_yh))]

                # banda de confiança (só quando há poucas curvas na tela)
                if _ic_ok and _meia is not None:
                    fig_reg.add_trace(go_plt.Scatter(
                        x=np.concatenate([_xs, _xs[::-1]]),
                        y=np.concatenate([_yh + _meia, (_yh - _meia)[::-1]]),
                        fill="toself", fillcolor=_cor, opacity=0.10, line=dict(width=0),
                        hoverinfo="skip", showlegend=False))

                # parcelas observadas
                _cd_cols = [c for c in ["cod_fazenda", "cidade_nome", "estado_sigla"]
                            if c in _gsel.columns]
                if len(_cd_cols) == 3:
                    _cd = _gsel[_cd_cols].values
                    _ht = ("<b>%{customdata[0]}</b> · %{customdata[1]}, %{customdata[2]}<br>"
                           f"<i>{_hib}</i><br>"
                           "População: <b>%{x:,.0f}</b> pl/ha<br>sc/ha: <b>%{y:.1f}</b><extra></extra>")
                else:
                    _cd = None
                    _ht = (f"<b>{_hib}</b><br>População: %{{x:,.0f}} pl/ha<br>"
                           f"sc/ha: %{{y:.1f}}<extra></extra>")
                fig_reg.add_trace(go_plt.Scatter(
                    x=_gsel["_pop_final"], y=_gsel[_COL_Y], mode="markers", showlegend=False,
                    marker=dict(color=_cor, size=6, opacity=0.45,
                                line=dict(color="#FFFFFF", width=0.5)),
                    customdata=_cd, hovertemplate=_ht))

                # curva ajustada — tracejada quando a curvatura não é significativa
                _curva_real = _rr["veredito"] in {"ponto de máximo", "máximo fora da faixa",
                                                  "curva em U"}
                fig_reg.add_trace(go_plt.Scatter(
                    x=_xs, y=_yh, mode="lines", showlegend=False,
                    line=dict(color=_cor, width=2.5, dash=None if _curva_real else "dash"),
                    hovertemplate=(f"<b>{_hib}</b><br>População: %{{x:,.0f}} pl/ha<br>"
                                   f"Estimado: %{{y:.1f}} sc/ha<extra></extra>")))

                # ★ só onde o topo é real E cai dentro da faixa testada
                if _rr["veredito"] == "ponto de máximo":
                    _res_reg.append((_hib, _cor, _rr))
                    if pd.notna(_rr["pop_otima_ep"]) and _rr["pop_otima_ep"] > 0 and _ic_ok:
                        fig_reg.add_vrect(x0=_rr["pop_otima"] - _rr["pop_otima_ep"],
                                          x1=_rr["pop_otima"] + _rr["pop_otima_ep"],
                                          fillcolor=_cor, opacity=0.08, line_width=0)
                    fig_reg.add_shape(type="line", x0=_rr["pop_otima"], x1=_rr["pop_otima"],
                                      y0=float(np.min(_yh)) - 2, y1=_rr["prod_otima"],
                                      line=dict(color=_cor, width=1, dash="dot"))
                    _ep_h = (f"<br>margem: ± {_mil(_rr['pop_otima_ep'])} pl/ha"
                             if pd.notna(_rr["pop_otima_ep"]) else "")
                    fig_reg.add_trace(go_plt.Scatter(
                        x=[_rr["pop_otima"]], y=[_rr["prod_otima"]], mode="markers",
                        name="", showlegend=False,
                        marker=dict(color=_cor, size=18, symbol="star",
                                    line=dict(color="#FFFFFF", width=2)),
                        hovertemplate=(f"<b>★ {_hib}</b><br>"
                                       f"População ótima: {_mil(_rr['pop_otima'])} pl/ha{_ep_h}<br>"
                                       f"sc/ha estimado: {_rr['prod_otima']:.1f}<br>"
                                       f"R² = {_rr['r2']:.2f}<extra></extra>")))

            # legenda "Máximos estimados" — traces invisíveis, só para o quadro lateral
            for _hib_l, _cor_l, _rl in _res_reg:
                fig_reg.add_trace(go_plt.Scatter(
                    x=[None], y=[None], mode="markers",
                    name=f"{_hib_l}  {_rl['pop_otima']/1000:.0f}k pl/ha · "
                         f"{_rl['prod_otima']:.1f} sc/ha",
                    marker=dict(color=_cor_l, size=14, symbol="star",
                                line=dict(color="#FFFFFF", width=1.5)),
                    showlegend=True))

            # ── Título dinâmico com o resultado real ──────────────────────────
            if _res_reg:
                _melhor = max(_res_reg, key=lambda t: t[2]["prod_otima"])
                _tit_reg = (f"<b style='color:{_CL_VERDE}'>{_melhor[0]}</b> atinge o máximo em "
                            f"<b style='color:{_CL_VERDE}'>{_mil(_melhor[2]['pop_otima'])} pl/ha</b> "
                            f"({_melhor[2]['prod_otima']:.1f} sc/ha estimados)")
                _outros = [f"{h}: {_mil(r['pop_otima'])}" for h, _, r in _res_reg if h != _melhor[0]]
                if _outros:
                    _tit_reg += " · Outros: " + " · ".join(_outros)
            else:
                _sem_topo = [f"{h} ({_ajustes[h]['veredito']})" for h in _hibs_reg]
                _tit_reg = ("Nenhum dos híbridos selecionados tem ponto de máximo dentro da faixa "
                            "testada — " + ", ".join(_sem_topo))
            st.markdown(f"""
<div style="margin:0.5rem 0 0.6rem;">
    <p style="font-size:13px;color:#6B7280;margin:0;">{_tit_reg}</p>
</div>""", unsafe_allow_html=True)

            # ── Layout ────────────────────────────────────────────────────────
            _x_sel = pd.to_numeric(
                _base_reg[_base_reg["dePara"].isin(_hibs_reg)]["_pop_final"],
                errors="coerce").dropna()
            _x_lo, _x_hi = float(_x_sel.min()), float(_x_sel.max())
            # marcas de 10 em 10 mil dentro da faixa observada (a soja crava 100k–500k, de soja)
            _tick0 = int(np.floor(_x_lo / 10000) * 10000)
            _tick1 = int(np.ceil(_x_hi / 10000) * 10000)
            _tickvals = list(range(_tick0, _tick1 + 1, 10000))

            # Y ancorado nas curvas, ampliado para não cortar a nuvem de pontos comum
            _y_sel = pd.to_numeric(
                _base_reg[_base_reg["dePara"].isin(_hibs_reg)][_COL_Y], errors="coerce").dropna()
            _y_lo_r = min(_y_curvas) if _y_curvas else float(_y_sel.quantile(0.02))
            _y_hi_r = max(_y_curvas) if _y_curvas else float(_y_sel.quantile(0.98))
            _y_lo_r = min(_y_lo_r, float(_y_sel.quantile(0.02)))
            _y_hi_r = max(_y_hi_r, float(_y_sel.quantile(0.98)))
            _amp_y = (_y_hi_r - _y_lo_r) or 1

            fig_reg.update_layout(
                height=520, plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                font=dict(family="Helvetica Neue, sans-serif"),
                showlegend=bool(_res_reg),
                legend=dict(title=dict(text="Máximos estimados",
                                       font=dict(size=11, color="#6B7280")),
                            orientation="v", x=1.01, y=1, xanchor="left", yanchor="top",
                            bgcolor="rgba(255,255,255,0.92)", bordercolor="#E5E7EB",
                            borderwidth=1, font=dict(size=12, color="#1A1A1A")),
                margin=dict(t=20, b=60, l=70, r=20),
                xaxis=dict(title=dict(text="<b>População real contada (pl/ha)</b>",
                                      font=dict(size=14, color="#1A1A1A", weight="bold")),
                           tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                           tickvals=_tickvals,
                           ticktext=[f"{v/1000:.0f}k" for v in _tickvals],
                           range=[_x_lo - (_x_hi - _x_lo) * 0.04,
                                  _x_hi + (_x_hi - _x_lo) * 0.05],
                           showgrid=True, gridcolor="#EEEEEE", zeroline=False, showline=False),
                yaxis=dict(title=dict(text="<b>Produtividade (sc/ha)</b>",
                                      font=dict(size=14, color="#1A1A1A", weight="bold")),
                           tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                           range=[_y_lo_r - _amp_y * 0.06, _y_hi_r + _amp_y * 0.12],
                           showgrid=True, gridcolor="#EEEEEE", zeroline=False, showline=False))
            st.plotly_chart(fig_reg, use_container_width=True)

            _col_cap_r, _col_dic_r = st.columns([3, 1])
            with _col_cap_r:
                st.caption(
                    "Pontos = parcelas observadas · linha cheia = curva com curvatura significativa "
                    "· linha tracejada = sem curvatura real (o comportamento é de reta) · ★ = "
                    "população de máxima produtividade estimada, só quando o topo cai dentro da "
                    "faixa testada." +
                    (" Faixa clara = intervalo de confiança de 95%." if _ic_ok else
                     (" O intervalo de confiança some com mais de 3 híbridos na tela."
                      if _mostrar_ic else "")) +
                    (" Produtividade centrada pela média do local." if _desc_local else ""))
            with _col_dic_r:
                _dic_cols_r = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome",
                                           "estado_sigla"] if c in _base_reg.columns]
                if _dic_cols_r:
                    _df_dic_r = (_base_reg[_base_reg["dePara"].isin(_hibs_reg)][_dic_cols_r]
                                 .drop_duplicates().sort_values(_dic_cols_r[0])
                                 .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                                  "cidade_nome": "Cidade",
                                                  "estado_sigla": "Estado"})
                                 .reset_index(drop=True))
                    with st.popover(f"📍 {len(_df_dic_r)} locais", use_container_width=True):
                        st.markdown("Passe o mouse sobre os pontos para identificar o local.")
                        st.dataframe(_df_dic_r, hide_index=True, use_container_width=True)

            st.info("A coluna **Resposta marginal** da tabela abaixo diz quanto cada mil plantas a "
                    "mais ainda rende na população média do híbrido. É esse número que se compara "
                    "com o custo da semente: quando o ganho marginal deixa de pagar as sementes "
                    "extras, o ótimo econômico foi ultrapassado — e ele vem sempre antes do topo "
                    "da curva.")

        # ── Dados técnicos da regressão (molde da soja) ────────────────────────
        st.markdown(
            '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
            'letter-spacing:0.07em;margin:1.4rem 0 0.4rem;">Dados técnicos da regressão</p>',
            unsafe_allow_html=True)

        # Por padrão a grade tem exatamente as colunas da soja. As colunas que sustentam as
        # travas (margem do ótimo, faixa testada, p da curvatura, leitura) ficam atrás do
        # checkbox — quem quiser conferir se o ótimo é confiável liga e vê.
        _cols_tec = st.checkbox("Mostrar colunas de verificação", value=False,
                                key="den_reg_coltec",
                                help="Margem do ótimo, faixa testada, p da curvatura e leitura.")

        def _qualidade_r2(v):
            """Régua da soja: R² alto = a densidade explica bem; baixo = outros fatores dominam."""
            if pd.isna(v):
                return "—"
            return "Bom" if v >= 0.50 else ("Moderado" if v >= 0.25 else "Fraco")

        _linhas = []
        for _hib, r in _ajustes.items():
            _tem_max = pd.notna(r["pop_otima"]) and r["dentro_faixa"]
            # Δ por +10 mil plantas (mesma unidade da soja); resp_marginal é por mil
            _d10 = r["resp_marginal"] * 10 if pd.notna(r["resp_marginal"]) else None
            _linhas.append({
                # ── as 8 colunas da soja, nesta ordem ──
                "Híbrido": _hib,
                "Pop. ótima (técnica)": (f"{r['pop_otima']/1000:.0f}k pl/ha" if _tem_max else "—"),
                "sc/ha no máximo": (f"{r['prod_otima']:.1f}" if _tem_max else "—"),
                "Δsc/ha por +10k plantas": (f"+{_d10:.2f} sc/ha" if _d10 is not None and _d10 > 0
                                            else (f"{_d10:.2f} sc/ha" if _d10 is not None else "—")),
                "R²": (f"{r['r2']:.3f}" if pd.notna(r["r2"]) else "—"),
                "Qualidade": _qualidade_r2(r["r2"]),
                "N parcelas": r["n"],
                "Pop. média obs.": (f"{r['x_medio']/1000:.0f}k" if pd.notna(r["x_medio"]) else "—"),
                # ── colunas de verificação (só com o checkbox ligado) ──
                "± margem": (f"± {r['pop_otima_ep']/1000:.1f}k"
                             if _tem_max and pd.notna(r["pop_otima_ep"]) else "—"),
                "Faixa testada": (f"{r['x_min']/1000:.0f}k – {r['x_max']/1000:.0f}k"
                                  if pd.notna(r["x_min"]) else "—"),
                "p curvatura": (f"{r['p_curvatura']:.3f}" if pd.notna(r["p_curvatura"]) else "—"),
                "Leitura": r["veredito"],
                # colunas auxiliares de ordenação (removidas antes de exibir)
                "_ord1": None,
                "_ord2": (-r["prod_otima"] if _tem_max else 0.0),
            })

        _COLS_VERIFICACAO = ["± margem", "Faixa testada", "p curvatura", "Leitura"]

        # ordem: quem tem topo primeiro (maior sc/ha no máximo), depois os demais vereditos
        _ordem_v = {"ponto de máximo": 0, "resposta crescente": 1, "resposta decrescente": 2,
                    "máximo fora da faixa": 3, "curva em U": 4, "sem resposta": 5,
                    "dados insuficientes": 6, "sem ajuste": 7}
        _tab_reg = pd.DataFrame(_linhas)
        _tab_reg["_ord1"] = _tab_reg["Leitura"].map(_ordem_v).fillna(9)
        _tab_reg = (_tab_reg.sort_values(["_ord1", "_ord2", "Híbrido"])
                    .drop(columns=["_ord1", "_ord2"]).reset_index(drop=True))
        _tab_exp = _tab_reg.copy()  # o Excel leva tudo, inclusive as colunas de verificação
        if not _cols_tec:
            _tab_reg = _tab_reg.drop(columns=_COLS_VERIFICACAO)

        # as colunas são texto (padrão da soja), então o estilo lê o número com parseFloat
        _js_leitura = JsCode("""
        function(params) {
          const m = {'ponto de máximo':'#1E7A34','resposta crescente':'#2976B6',
                     'resposta decrescente':'#C0201E','máximo fora da faixa':'#B45309',
                     'curva em U':'#B45309'};
          if (m[params.value]) return {color: m[params.value], fontWeight: '600'};
          return {color: '#6B7280'};
        }""")
        _js_qual = JsCode("""
        function(params) {
          const m = {'Bom':'#1E7A34','Moderado':'#B45309','Fraco':'#94A3B8'};
          if (m[params.value]) return {color: m[params.value], fontWeight: '600'};
          return {color: '#9CA3AF'};
        }""")
        _js_p = JsCode("""
        function(params) {
          const v = parseFloat(params.value);
          if (isNaN(v)) return {color: '#9CA3AF'};
          return {color: v < 0.05 ? '#1E7A34' : '#9CA3AF', fontWeight: v < 0.05 ? '600' : '400'};
        }""")
        _js_d10 = JsCode("""
        function(params) {
          const v = parseFloat(params.value);
          if (isNaN(v)) return {color: '#9CA3AF'};
          return {color: v > 0 ? '#1E7A34' : '#C0201E', fontWeight: '600'};
        }""")

        ag_table(_tab_reg, height=min(420, 60 + 32 * len(_tab_reg)),
                 estilos_col={"Leitura": _js_leitura, "Qualidade": _js_qual,
                              "p curvatura": _js_p, "Δsc/ha por +10k plantas": _js_d10})

        st.caption(
            "**Pop. ótima** = máximo técnico da curva quadrática (não econômico) · "
            "**Δsc/ha por +10k plantas** = ganho ou perda a cada 10 mil plantas a mais, na "
            "população média do híbrido · **R² < 0,25** = outros fatores dominam sobre a "
            "densidade." +
            (f" **± margem** = erro-padrão do ponto de máximo (margem larga = topo achatado) · "
             f"**p curvatura < 0,05** = a curva existe; acima disso o comportamento é de reta · "
             f"régua de ajuste: mínimo de {_REG_MIN_PLOTS} parcelas, {_REG_MIN_POPS} populações "
             f"distintas e {_REG_MIN_AMPL/1000:.0f} mil plantas/ha de amplitude."
             if _cols_tec else
             " Ligue **Mostrar colunas de verificação** para ver a margem do ótimo, a faixa "
             "testada e o teste da curvatura."))

        exportar_excel(_tab_exp, "densidade_regressao.xlsx",
                       "⬇️ Exportar dados técnicos", key="den_reg_xlsx")

        _n_topo = sum(1 for r in _ajustes.values() if r["veredito"] == "ponto de máximo")
        _n_sobe = sum(1 for r in _ajustes.values() if r["veredito"] == "resposta crescente")
        _n_sem = sum(1 for r in _ajustes.values() if r["veredito"] == "sem resposta")
        _partes = []
        if _n_topo:
            _partes.append(f"**{_n_topo}** híbrido(s) mostraram ponto de máximo dentro da faixa "
                           f"testada — nesses, aumentar a população além do topo não devolve grão.")
        if _n_sobe:
            _partes.append(f"**{_n_sobe}** ainda subiam na maior densidade testada: o ensaio não "
                           f"chegou ao teto deles, e o ótimo pode estar acima da faixa.")
        if _n_sem:
            _partes.append(f"**{_n_sem}** não responderam de forma consistente à densidade dentro "
                           f"desta faixa.")
        if _partes:
            st.info(" ".join(_partes))

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5 — COMPONENTES DE PRODUÇÃO POR GRUPO DE DENSIDADE
# ══════════════════════════════════════════════════════════════════════════════
# Tabela no molde da Análise Conjunta, com cabeçalho de dois níveis (como no Excel): a linha
# de cima traz os grupos de densidade e a de baixo as variáveis dentro de cada grupo. Linha =
# híbrido. Ler na horizontal mostra como o mesmo material se comporta conforme a população sobe.
import streamlit.components.v1 as components

# altura da espiga em metros: se o banco só trouxe centímetros, converte aqui para as duas
# alturas ficarem na mesma unidade e serem comparáveis lado a lado na tabela.
if "altura_espiga_m" not in ta_filtrado.columns and "altura_espiga_cm" in ta_filtrado.columns:
    ta_filtrado["altura_espiga_m"] = (
        pd.to_numeric(ta_filtrado["altura_espiga_cm"], errors="coerce") / 100).round(3)

# (rótulo, coluna, casas decimais, maior_é_melhor)
#   maior_é_melhor = None -> a variável não tem "bom" e "ruim" (população, umidade, divergência)
_VARS_COMP = [
    ("sc/ha", "sc_ha", 1, True),
    ("kg/ha", "kg_ha", 1, True),
    ("Pop. Final", "_pop_final", 0, None),
    ("PMG (g)", "pmg_corrigido_g", 1, True),
    ("Grãos/Fileira", "graos_fileira_media", 1, True),
    ("Fileiras", "fileiras_media", 1, True),
    # agronômicas: respondem à densidade (competição por luz estica a planta), mas não têm
    # "melhor" — planta alta demais acama, baixa demais pode indicar estresse. Sem cor.
    ("Altura Planta (m)", "altura_planta_m", 1, None),
    ("Altura Espiga (m)", "altura_espiga_m", 1, None),
    ("Umidade (%)", "umidade_pct", 1, None),
    ("Estimativa (sc/ha)", "prod_estimada_sacas_ha", 1, True),
    ("Divergência (%)", "divergencia_prod_pct", 1, None),
]
_VARS_COMP = [v for v in _VARS_COMP if v[1] in ta_filtrado.columns]

secao_titulo("Componentes de Produção",
             "Os componentes de produção de cada híbrido ao longo das populações",
             contexto_str)

if not _VARS_COMP or not _grupos_encontrados or "dePara" not in ta_filtrado.columns:
    st.info("Sem colunas de componentes ou sem grupos de densidade nos filtros ativos.")
else:
    with st.popover("ℹ️ Como ler a tabela", use_container_width=False):
        st.markdown("""
**Cabeçalho em dois níveis.** Em cima, os grupos de densidade; embaixo, as variáveis dentro de
cada grupo. Cada linha é um híbrido.

**Os três controles acima da tabela:**

- **Variáveis** — escolha quais colunas aparecem dentro de cada bloco de densidade. As colunas
  saem **na ordem em que você escolher**, não na ordem da lista: selecionar PMG e depois sc/ha
  põe PMG à esquerda. Para reordenar, remova a variável e selecione de novo — ela vai para o fim.
- **Mostrar n** — acrescenta, no fim de cada bloco, quantas parcelas sustentam aquelas médias.
- **Bloco geral** — liga ou desliga o bloco **Todas as densidades**, que junta todas as parcelas
  do híbrido e serve de referência para o resto da linha.

**Leia na horizontal.** Percorrendo a linha de um híbrido, você vê o mesmo material em populações
crescentes. É onde aparece o mecanismo: normalmente o PMG e os grãos por fileira **caem** quando a
população sobe (mais plantas dividindo os mesmos recursos), e a produtividade só sobe enquanto o
ganho em número de plantas compensa essa perda. Quando para de compensar, é o topo da curva da
Seção 4 — aqui você vê **por que** ele acontece.

**Leia na vertical** para comparar híbridos dentro de uma mesma densidade.

**Cor do número: a comparação é sempre VERTICAL.**

A cor de uma célula responde a uma única pergunta: *este híbrido está acima ou abaixo dos outros
híbridos, nesta mesma variável e nesta mesma densidade?* Ela compara a célula com a **média da
coluna em que ela está** — e uma coluna é sempre uma variável dentro de um bloco de densidade.

Exemplo, na coluna **PMG** do bloco **55k**:

| Híbrido | PMG (g) | Cor | Leitura |
|---|---|---|---|
| 9505PRO4 | 312,6 | <span style="color:#1E7A34;font-weight:700">verde</span> | acima da média dos híbridos em 55k |
| 9717VIP3 | 306,0 | preto | praticamente na média (menos de 1% de desvio) |
| 9610VIP3 | 297,8 | <span style="color:#C0201E;font-weight:700">vermelho</span> | abaixo da média dos híbridos em 55k |
| *Média da coluna* | *305,5* | | *é esta a referência* |

**O que a cor NÃO compara:**

- **Não compara densidades.** O verde do 9505PRO4 em 55k não diz que ele foi melhor em 55k do que
  em 41k. Cada bloco tem sua própria média e é uma comparação independente. Ler a linha na
  horizontal continua valendo — mas aí você lê os **números**, não as cores.
- **Não compara com a meta, com a testemunha nem com a safra passada.** A referência é sempre o
  conjunto de híbridos filtrados naquele momento. Mexer no filtro de híbrido muda a média da
  coluna e pode virar uma cor — isso é esperado, é o grupo de comparação mudando.
- **Não indica significância.** Verde é "acima da média", não "significativamente superior". Quem
  responde por diferença real é o LSD da Seção 3.

Desvios abaixo de 1% ficam em **preto**: variação dessa ordem é ruído e não merece cor.

**Só recebem cor as variáveis em que "mais" significa "melhor".** Nas outras, verde e vermelho
seriam uma afirmação falsa:

| Variável | Cor | Por quê |
|---|---|---|
| sc/ha, kg/ha, PMG, Grãos/Fileira, Fileiras, Estimativa | sim | mais é melhor |
| Pop. Final | não | é a densidade em si, o eixo do estudo — não um mérito |
| Altura Planta (m) / Altura Espiga (m) | não | alta demais acama, baixa demais indica estresse |
| Umidade (%) | não | alta demais penaliza na balança, baixa demais indica colheita tardia |
| Divergência (%) | não | o bom é estar **perto de zero**, dos dois lados |

**Como cada variável é calculada** (tudo vem do pipeline; a base de umidade é 13,5%):

- **kg/ha** — peso da parcela corrigido para 13,5% de umidade, extrapolado pela área:
  `peso × (100 − umidade) ÷ (100 − 13,5) × 10.000 ÷ área da parcela`. A área vem de
  `nº de linhas × comprimento × espaçamento`.
- **sc/ha** — kg/ha ÷ 60. Mesma medida, unidade comercial.
- **Pop. Final** — população **real contada**, não o alvo do tratamento:
  `plantas contadas ÷ metros de contagem ÷ espaçamento × 10.000`. É a média das subamostras
  (4 em 2025, 5 em 2024).
- **PMG (g)** — peso de mil grãos, também corrigido para 13,5%:
  `PMG bruto × (100 − umidade da amostra) ÷ (100 − 13,5)`. Média das subamostras da parcela.
- **Grãos/Fileira** e **Fileiras** — contagem direta nas espigas das subamostras, média por parcela.
- **Altura Planta (m)** e **Altura Espiga (m)** — medidas nas plantas das subamostras, média por
  parcela, as duas em metros para serem lidas lado a lado (a espiga é convertida de centímetro
  quando o banco só traz cm). A altura de planta é a variável que mais
  responde à densidade depois dos componentes de grão — plantas competindo por luz **esticam**, e
  esse alongamento é o que liga densidade alta a acamamento e quebramento. Vale cruzar com a
  Seção de Perdas.
- **Estimativa (sc/ha)** — a produtividade reconstruída pelos componentes, **sem usar a
  balança**. É a mesma sc/ha, calculada por outro caminho:
  `população × fileiras × grãos por fileira × (PMG ÷ 1000) ÷ 1000`, depois ÷ 60.
- **Divergência (%)** — o quanto a estimativa pelos componentes se afasta do peso colhido:
  `(estimada − colhida) ÷ colhida × 100`. É coluna de **auditoria**, não de desempenho: valores
  altos indicam contagem ou pesagem com problema, não híbrido melhor ou pior. Positivo = os
  componentes prometeram mais grão do que a balança confirmou.

**Cada célula é a média das parcelas** daquele híbrido naquela densidade, calculada sobre a
parcela (não sobre a subamostra). Parcelas sem o dado ficam de fora daquela célula, então uma
variável pode ter n menor que outra na mesma célula.

**Fundo da primeira coluna** = status do material (a mesma legenda da Análise Conjunta).
**Cabeçalho de densidade** = azul escurecendo conforme a população sobe.

**Sobre o n.** As parcelas do ensaio se dividem entre os grupos, então cada célula costuma ter
poucas. Ligue **Mostrar n** para ver quantas sustentam cada bloco. Células com poucas parcelas
aparecem esmaecidas e com asterisco: são médias frágeis e não deveriam decidir nada sozinhas. Por
isso o bloco **Todas as densidades** vem primeiro — ele é o mais confiável da tabela.

**A exportação é o espelho da tela.** O Excel sai com o mesmo cabeçalho de dois níveis (blocos
mesclados), a mesma rampa de azul, as mesmas cores de status e de desvio, o asterisco das células
frágeis e as variáveis **na ordem que você escolheu** — só o que está visível aqui vai para o
arquivo. As células guardam número de verdade, não texto, então dá para somar e montar dinâmica em
cima. No rodapé do arquivo ficam a legenda das cores e os filtros ativos, para o arquivo não
circular sem dizer de que recorte veio.
""", unsafe_allow_html=True)

    _c_v1, _c_v2, _c_v3 = st.columns([3, 1.2, 1.2])
    with _c_v1:
        _vars_sel = st.multiselect(
            "Variáveis", options=[v[0] for v in _VARS_COMP],
            default=[v[0] for v in _VARS_COMP][:4], key="den_comp_vars",
            help="As colunas aparecem na ordem em que você escolher. Para reordenar, remova a "
                 "variável e selecione de novo — ela vai para o fim.")
    with _c_v2:
        _mostrar_n = st.checkbox("Mostrar n", value=False, key="den_comp_n")
    with _c_v3:
        _mostrar_geral = st.checkbox("Bloco geral", value=True, key="den_comp_geral",
                                     help="Bloco com todas as densidades juntas, no início.")

    _N_MIN_CEL = 5  # abaixo disso a média da célula é frágil e ganha asterisco

    if not _vars_sel:
        st.info("Selecione ao menos uma variável.")
    else:
        # a ordem das colunas segue a ORDEM EM QUE VOCÊ ESCOLHEU no multiselect, não a ordem
        # canônica da lista: quem seleciona primeiro sc/ha e depois PMG quer PMG à direita.
        # (para reordenar, tire a variável e coloque de novo — ela vai para o fim.)
        _mapa_var = {v[0]: v for v in _VARS_COMP}
        _vars = [_mapa_var[_r] for _r in _vars_sel if _r in _mapa_var]
        _blocos = ([("Todas as densidades", None)] if _mostrar_geral else []) + \
                  [(g, g) for g in _grupos_encontrados]

        # ── Agregação: média por híbrido × bloco ──────────────────────────────
        _hibs = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
        _status_hib = (ta_filtrado.dropna(subset=["dePara"])
                       .groupby("dePara")["status_material"].first().to_dict()
                       if "status_material" in ta_filtrado.columns else {})

        _cel = {}   # (híbrido, bloco, rótulo) -> valor;  (híbrido, bloco, "__n__") -> nº de parcelas
        for _nome_bl, _grp in _blocos:
            _dbl = ta_filtrado if _grp is None else ta_filtrado[ta_filtrado["pop_grupo"] == _grp]
            for _h in _hibs:
                _dh = _dbl[_dbl["dePara"] == _h]
                _cel[(_h, _nome_bl, "__n__")] = len(_dh)
                for _rot, _col, _dec, _ in _vars:
                    _v = pd.to_numeric(_dh[_col], errors="coerce")
                    _v = _v[_v > 0] if _col in ("sc_ha", "kg_ha", "_pop_final") else _v.dropna()
                    _cel[(_h, _nome_bl, _rot)] = float(_v.mean()) if len(_v) else np.nan

        # média de cada coluna (bloco × variável) — referência para a cor e para o rodapé
        _med_col = {}
        for _nome_bl, _ in _blocos:
            for _rot, _, _, _ in _vars:
                _vals = [_cel[(_h, _nome_bl, _rot)] for _h in _hibs
                         if pd.notna(_cel[(_h, _nome_bl, _rot)])]
                _med_col[(_nome_bl, _rot)] = float(np.mean(_vals)) if _vals else np.nan

        # ── HTML com cabeçalho de dois níveis ─────────────────────────────────
        _COR_STATUS = {"CHECK": "#F4B184", "STINE": "#2976B6", "EXP": "#00FF00", "DP2": "#C4DFB4"}
        _COR_TEXTO = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                      "DP2": "#1A1A1A", "": "#000000"}
        _DEC = {v[0]: v[2] for v in _vars}
        _DIR = {v[0]: v[3] for v in _vars}
        _n_sub = len(_vars) + (1 if _mostrar_n else 0)

        def _rampa_densidade(n):
            """Rampa sequencial clara → escura para os grupos de densidade.

            A densidade é ordinal: 82k não é uma categoria qualquer, é a maior. Cores soltas
            (uma azul, uma laranja, uma verde) perderiam essa ordem e ainda colidiriam com as
            cores de status e com o verde/vermelho do desvio. Um único matiz escurecendo
            comunica 'mais plantas' sem precisar de legenda.

            A rampa é dessaturada de propósito: nenhum tom dela cai perto do azul Stine
            (#2976B6), que já significa status na primeira coluna.
            """
            ini, fim = (214, 228, 242), (18, 57, 92)   # azul claro → azul-marinho
            saida = []
            for i in range(max(n, 1)):
                t = i / (n - 1) if n > 1 else 0.0
                rgb = tuple(int(round(ini[k] + (fim[k] - ini[k]) * t)) for k in range(3))
                lum = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
                saida.append((f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                              "#1A1A1A" if lum > 150 else "#FFFFFF"))
            return saida

        _rampa = _rampa_densidade(len(_grupos_encontrados))
        # o bloco geral fica fora da rampa (cinza-escuro): ele não é uma densidade, é o total
        _cor_bloco = {}
        _i_g = 0
        for _nome_bl, _grp in _blocos:
            if _grp is None:
                _cor_bloco[_nome_bl] = ("#3A3A3A", "#FFFFFF")
            else:
                _cor_bloco[_nome_bl] = _rampa[min(_i_g, len(_rampa) - 1)]
                _i_g += 1

        def _fmt(v, dec):
            if v is None or pd.isna(v):
                return "—"
            return f"{int(round(v)):,}".replace(",", ".") if dec == 0 else f"{v:.{dec}f}"

        def _cor_valor(rot, v, ref):
            """Cor do texto = sinal do desvio contra a média da coluna. Variáveis sem
            'melhor' definido (população, umidade, divergência) ficam em preto."""
            direcao = _DIR.get(rot)
            if direcao is None or pd.isna(v) or pd.isna(ref) or ref == 0:
                return "#1A1A1A"
            desvio = (v - ref) / abs(ref) * 100
            if abs(desvio) < 1:          # zona morta: variação irrelevante não vira cor
                return "#1A1A1A"
            bom = (desvio > 0) if direcao else (desvio < 0)
            return "#1E7A34" if bom else "#C0201E"

        _html = """
<style>
.tb-comp { width:100%; border-collapse:collapse; font-size:14px;
    font-family:'Helvetica Neue',sans-serif; }
.tb-comp th { background:#F2F2F2; color:#000 !important; padding:7px 9px; text-align:center;
    border:1px solid #ccc; white-space:nowrap; font-weight:700; }
.tb-comp th.grupo { font-size:14px; letter-spacing:0.03em; }
.tb-comp th.grupo-geral { background:#3A3A3A; color:#FFF !important; }
.tb-comp th.hib { background:#4A4A4A; color:#FFF !important; text-align:left; }
.tb-comp td { padding:6px 9px; border:1px solid #ddd; text-align:center; white-space:nowrap; }
.tb-comp td.hib { text-align:left; font-weight:600; }
.tb-comp td.sep { border-left:2px solid #999; }
.tb-comp th.sep { border-left:2px solid #999; }
.tb-comp td.frag { opacity:0.55; }
.tb-comp td.ncel { color:#6B7280 !important; font-size:12px; font-style:italic; }
.tb-comp tr.rodape td { background:#D9D9D9 !important; font-weight:700;
    border-top:2px solid #888; color:#000 !important; }
</style>
<table class="tb-comp">
<thead>
<tr><th class="hib" rowspan="2">Híbrido</th>"""
        for _nome_bl, _grp in _blocos:
            _bgh, _fgh = _cor_bloco[_nome_bl]
            _html += (f'<th class="grupo sep" style="background:{_bgh};color:{_fgh} !important;'
                      f'border-color:{_bgh};" colspan="{_n_sub}">{_nome_bl}</th>')
        _html += "</tr><tr>"
        for _nome_bl, _ in _blocos:
            _bgh, _ = _cor_bloco[_nome_bl]
            # faixa fina na cor do bloco: mantém a associação quando a tabela rola na horizontal
            for _j, (_rot, _, _, _) in enumerate(_vars):
                _html += (f'<th class="{"sep" if _j == 0 else ""}" '
                          f'style="border-top:4px solid {_bgh};">{_rot}</th>')
            if _mostrar_n:
                _html += f'<th style="border-top:4px solid {_bgh};">n</th>'
        _html += "</tr></thead><tbody>"

        for _h in _hibs:
            _st = _status_hib.get(_h, "")
            _bg = _COR_STATUS.get(_st, "#FFFFFF")
            _fg = _COR_TEXTO.get(_st, "#000000")
            _html += (f'<tr><td class="hib" style="background:{_bg};color:{_fg} !important;">'
                      f'{_h}</td>')
            for _nome_bl, _ in _blocos:
                _n_bl = _cel[(_h, _nome_bl, "__n__")]
                _frag = _n_bl < _N_MIN_CEL
                for _j, (_rot, _, _, _) in enumerate(_vars):
                    _v = _cel[(_h, _nome_bl, _rot)]
                    _cor = _cor_valor(_rot, _v, _med_col[(_nome_bl, _rot)])
                    _cls = " ".join(x for x in ["sep" if _j == 0 else "",
                                                "frag" if _frag else ""] if x)
                    _ast = "*" if (_frag and pd.notna(_v)) else ""
                    _html += (f'<td class="{_cls}" style="color:{_cor} !important;">'
                              f'{_fmt(_v, _DEC[_rot])}{_ast}</td>')
                if _mostrar_n:
                    _html += f'<td class="ncel">{_n_bl}</td>'
            _html += "</tr>"

        # rodapé: média simples de cada coluna
        _html += '<tr class="rodape"><td class="hib">Média Geral</td>'
        for _nome_bl, _ in _blocos:
            for _j, (_rot, _, _, _) in enumerate(_vars):
                _html += (f'<td class="{"sep" if _j == 0 else ""}">'
                          f'{_fmt(_med_col[(_nome_bl, _rot)], _DEC[_rot])}</td>')
            if _mostrar_n:
                _tot = sum(_cel[(_h, _nome_bl, "__n__")] for _h in _hibs)
                _html += f'<td class="ncel">{_tot}</td>'
        _html += "</tr></tbody></table>"

        _altura = 90 + (len(_hibs) + 2) * 34
        components.html(_html, height=min(_altura, 700), scrolling=True)

        st.caption(
            f"Cabeçalho de cima = grupo de densidade · cabeçalho de baixo = variável · cada célula "
            f"é a média das parcelas daquele híbrido naquela densidade · cor do número = desvio "
            f"contra a média da coluna (verde acima, vermelho abaixo; só nas variáveis com sentido "
            f"de melhor) · fundo da primeira coluna = status do material · "
            f"**\\*** = menos de {_N_MIN_CEL} parcelas na célula, média frágil.")

        # ── Exportação: espelho fiel da tabela na tela ─────────────────────────
        # Reconstrói o mesmo leiaute (cabeçalho de dois níveis mesclado, rampa de densidade,
        # cor do desvio no texto, asterisco nas células frágeis, rodapé de média) em vez de
        # exportar uma tabela longa. O que se vê é o que se baixa.
        def _xlsx_componentes():
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Componentes"
            _hx = lambda c: str(c).replace("#", "").upper()
            thin = Side(style="thin", color="DDDDDD")
            grosso = Side(style="medium", color="999999")
            centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # ── linha 1: grupos de densidade (mesclado) / linha 2: variáveis ──
            ws.cell(row=1, column=1, value="Híbrido")
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            _c1 = ws.cell(row=1, column=1)
            _c1.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            _c1.fill = PatternFill("solid", start_color="4A4A4A")
            _c1.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions["A"].width = 18

            _ci = 2
            for _nome_bl, _ in _blocos:
                _bgh, _fgh = _cor_bloco[_nome_bl]
                _ini = _ci
                for _rot, _, _, _ in _vars:
                    _ch = ws.cell(row=2, column=_ci, value=_rot)
                    _ch.font = Font(bold=True, name="Arial", size=9, color="1A1A1A")
                    _ch.fill = PatternFill("solid", start_color="F2F2F2")
                    _ch.alignment = centro
                    _ch.border = Border(left=grosso if _ci == _ini else thin, right=thin,
                                        top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                    ws.column_dimensions[get_column_letter(_ci)].width = max(11, len(_rot) + 2)
                    _ci += 1
                if _mostrar_n:
                    _ch = ws.cell(row=2, column=_ci, value="n")
                    _ch.font = Font(bold=True, name="Arial", size=9, color="6B7280", italic=True)
                    _ch.fill = PatternFill("solid", start_color="F2F2F2")
                    _ch.alignment = centro
                    _ch.border = Border(left=thin, right=thin,
                                        top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                    ws.column_dimensions[get_column_letter(_ci)].width = 7
                    _ci += 1
                # cabeçalho do bloco, mesclado sobre as colunas dele
                ws.merge_cells(start_row=1, start_column=_ini, end_row=1, end_column=_ci - 1)
                _cb = ws.cell(row=1, column=_ini, value=_nome_bl)
                _cb.font = Font(bold=True, name="Arial", size=10, color=_hx(_fgh))
                _cb.fill = PatternFill("solid", start_color=_hx(_bgh))
                _cb.alignment = centro
                _cb.border = Border(left=grosso, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 30

            # ── corpo: uma linha por híbrido ─────────────────────────────────
            _ri = 3
            for _h in _hibs:
                _st = _status_hib.get(_h, "")
                _bg, _fg = _COR_STATUS.get(_st, "#FFFFFF"), _COR_TEXTO.get(_st, "#000000")
                _ch = ws.cell(row=_ri, column=1, value=_h)
                _ch.font = Font(bold=True, name="Arial", size=10, color=_hx(_fg))
                _ch.fill = PatternFill("solid", start_color=_hx(_bg))
                _ch.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                _ci = 2
                for _nome_bl, _ in _blocos:
                    _n_bl = _cel[(_h, _nome_bl, "__n__")]
                    _frag = _n_bl < _N_MIN_CEL
                    for _j, (_rot, _, _dec, _) in enumerate(_vars):
                        _v = _cel[(_h, _nome_bl, _rot)]
                        _cor = _cor_valor(_rot, _v, _med_col[(_nome_bl, _rot)])
                        # número de verdade na célula (dá para somar/pivotar), com o formato
                        # de exibição no numFormat e o asterisco do frágil como sufixo do formato
                        _cel_x = ws.cell(row=_ri, column=_ci,
                                         value=(round(float(_v), _dec) if pd.notna(_v) else None))
                        _fmt = "#,##0" if _dec == 0 else "0." + "0" * _dec
                        _cel_x.number_format = _fmt + ('"*"' if (_frag and pd.notna(_v)) else "")
                        _cel_x.font = Font(name="Arial", size=10, color=_hx(_cor),
                                           bold=(_cor != "#1A1A1A"), italic=_frag)
                        _cel_x.alignment = centro
                        _cel_x.border = Border(left=grosso if _j == 0 else thin, right=thin,
                                               top=thin, bottom=thin)
                        _ci += 1
                    if _mostrar_n:
                        _cn = ws.cell(row=_ri, column=_ci, value=int(_n_bl))
                        _cn.font = Font(name="Arial", size=9, color="6B7280", italic=True)
                        _cn.alignment = centro
                        _cn.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        _ci += 1
                _ri += 1

            # ── rodapé: média de cada coluna ─────────────────────────────────
            _cf = ws.cell(row=_ri, column=1, value="Média Geral")
            _cf.font = Font(bold=True, name="Arial", size=10, color="000000")
            _cf.fill = PatternFill("solid", start_color="D9D9D9")
            _cf.border = Border(left=thin, right=thin, top=grosso, bottom=thin)
            _ci = 2
            for _nome_bl, _ in _blocos:
                for _j, (_rot, _, _dec, _) in enumerate(_vars):
                    _v = _med_col[(_nome_bl, _rot)]
                    _cm = ws.cell(row=_ri, column=_ci,
                                  value=(round(float(_v), _dec) if pd.notna(_v) else None))
                    _cm.number_format = "#,##0" if _dec == 0 else "0." + "0" * _dec
                    _cm.font = Font(bold=True, name="Arial", size=10, color="000000")
                    _cm.fill = PatternFill("solid", start_color="D9D9D9")
                    _cm.alignment = centro
                    _cm.border = Border(left=grosso if _j == 0 else thin, right=thin,
                                        top=grosso, bottom=thin)
                    _ci += 1
                if _mostrar_n:
                    _cn = ws.cell(row=_ri, column=_ci,
                                  value=int(sum(_cel[(_h2, _nome_bl, "__n__")] for _h2 in _hibs)))
                    _cn.font = Font(bold=True, name="Arial", size=9, color="000000", italic=True)
                    _cn.fill = PatternFill("solid", start_color="D9D9D9")
                    _cn.alignment = centro
                    _cn.border = Border(left=thin, right=thin, top=grosso, bottom=thin)
                    _ci += 1

            # ── notas de rodapé (a mesma legenda da tela) ────────────────────
            _ri += 2
            for _txt in [
                f"Cor do número = desvio contra a média da coluna (verde acima, vermelho abaixo; "
                f"zona morta de 1%). Só as variáveis com sentido de 'melhor' recebem cor.",
                f"* = menos de {_N_MIN_CEL} parcelas na célula — média frágil.",
                f"Fundo da primeira coluna = status do material. "
                f"Cabeçalho de densidade = azul escurecendo conforme a população sobe.",
                f"Filtros ativos: {contexto_str}",
            ]:
                _cn = ws.cell(row=_ri, column=1, value=_txt)
                _cn.font = Font(name="Arial", size=8, color="6B7280", italic=True)
                _ri += 1

            ws.freeze_panes = "B3"          # cabeçalho e coluna do híbrido sempre à vista
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        st.download_button("⬇️ Exportar componentes",
                           data=_xlsx_componentes(),
                           file_name="densidade_componentes.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="den_comp_xlsx")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — LEITURA GRÁFICA: PRODUÇÃO x COMPONENTE, POR DENSIDADE
# ══════════════════════════════════════════════════════════════════════════════
# Versão gráfica da tabela da Seção 5, para um híbrido por vez. Barras = produtividade
# (eixo da esquerda), pontos ligados = a variável escolhida (eixo da direita).
#
# Eixo duplo tem uma armadilha conhecida: quem escolhe a escala escolhe onde as duas séries
# se cruzam, e aqui o cruzamento é justamente a leitura de interesse. Três travas contra isso:
#   1. os DOIS eixos começam em zero — não dá para deslizar o cruzamento mexendo no range;
#   2. cada ponto e cada barra levam rótulo COM UNIDADE, então o número não depende da altura;
#   3. as barras são claras e a linha escura, deixando claro qual série é o fundo e qual é a
#      leitura principal — e a legenda diz explicitamente que as escalas são diferentes.
# Sem tendência polinomial: o eixo x aqui são categorias, não a população contínua. Quem ajusta
# curva é a Seção 4, sobre a população real, com teste de curvatura.

_N_MIN_GRAF = 5          # abaixo disso a barra sai esmaecida com o n escrito
_CL_LINHA_G = "#2B2B2B"  # série da direita: escura, para ficar à frente das barras claras

# ── Variáveis disponíveis para a LINHA do gráfico de barras+linha ─────────────
# (rótulo, coluna, casas, maior_é_melhor, régua)
#   régua "todos"    = média de todas as parcelas avaliadas (o zero entra)
#   régua "so_maior" = média só das parcelas com valor > 0
# A régua acompanha a variável para o gráfico dar o MESMO número da tabela correspondente:
# componentes e perdas usam "todos"; fenômenos e ardidos usam "so_maior".
#
# Ficam de fora da lista de componentes: sc/ha e kg/ha (já são as barras), Pop. Final (é o
# próprio eixo x), Estimativa (produtividade por outro caminho, redundante com as barras) e
# Divergência (auditoria de contagem, não resposta agronômica à população).
_FORA_DO_GRAFICO = ("sc_ha", "kg_ha", "_pop_final",
                    "prod_estimada_sacas_ha", "divergencia_prod_pct")

_VARS_GRAF_COMP = [(r, c, d, m, "todos") for r, c, d, m in _VARS_COMP
                   if c not in _FORA_DO_GRAFICO]

_VARS_GRAF_PERDAS = [(f"{r} (%)", c, 1, False, "todos")
                     for r, c in {"Perda Total": "pct_perda_total",
                                  "Acamamento": "pct_acamadas",
                                  "Quebramento": "pct_quebradas",
                                  "Dominadas": "pct_dominadas",
                                  "Colmo Podre": "pct_colmo_podre"}.items()
                     if c in ta_filtrado.columns]

_VARS_GRAF_FENOM = [(f"{r} (%)", c, 1, False, "so_maior")
                    for r, c in {"Green snap": "pct_green_snap",
                                 "Morte prematura": "pct_morte_prematura",
                                 "Má formação": "pct_ma_formacao_espigas",
                                 "Enfezamento": "pct_enfezamento",
                                 "Ardidos": "graos_ardidos_pct"}.items()
                    if c in ta_filtrado.columns]

# notas de doença NÃO entram em nenhuma das listas: são escala ordinal (1–9), e média de nota
# não tem significado defensável. A leitura gráfica delas está na Seção 8, com composição por
# classe e mapa de classes.


def _grafico_barras_linha(titulo, subtitulo, _vars_graf, pref, prefixo_padrao=None):
    """Barras de produtividade + linha da variável escolhida, por grupo de densidade.

    Usada pela Seção 6 (componentes de produção) e pela Seção 9b (perdas e fenômenos): o
    mesmo desenho, mudando só a lista de variáveis da linha. Cada variável carrega a sua
    régua, então o gráfico dá o mesmo número da tabela correspondente.
    """
    secao_titulo(titulo, subtitulo, contexto_str)

    if not _grupos_encontrados or not _vars_graf or "dePara" not in ta_filtrado.columns:
        st.info("Sem grupos de densidade ou sem variáveis para o gráfico nos filtros ativos.")
    else:
        _hibs_g = sorted(ta_filtrado[(ta_filtrado["sc_ha"] > 0)]["dePara"].dropna().unique().tolist())
        if not _hibs_g:
            st.info("Nenhum híbrido com produtividade nos filtros ativos.")
        else:
            _cg1, _cg2, _cg3 = st.columns([2, 2, 1.6])
            with _cg1:
                _hib_g = st.selectbox("Híbrido", options=_hibs_g, key=f"{pref}_hib")
            with _cg3:
                # a referência das variações deixou de ser fixa na menor população: comparar
                # contra a densidade que a fazenda já pratica costuma ser a pergunta real
                _g_base = st.selectbox(
                    "Comparar contra", options=_grupos_encontrados, index=0,
                    key=f"{pref}_ref",
                    help="Todas as variações (% vs) são calculadas contra esta população. "
                         "O padrão é a menor.")
            with _cg2:
                _var_g = st.selectbox("Variável na linha", options=[v[0] for v in _vars_graf],
                                      index=next((i for i, v in enumerate(_vars_graf)
                                                  if prefixo_padrao
                                                  and v[0].startswith(prefixo_padrao)), 0),
                                      key=f"{pref}_var")

            _rot_v, _col_v, _dec_v, _dir_v, _regua_v = next(v for v in _vars_graf
                                                            if v[0] == _var_g)
            _dh = ta_filtrado[ta_filtrado["dePara"] == _hib_g]

            _regua_txt = ("média de **todas** as parcelas avaliadas — as que vieram zeradas "
                          "também contam, porque zero aqui é medição"
                          if _regua_v == "todos" else
                          "média **só das parcelas em que o problema apareceu** — as zeradas "
                          "ficam de fora, senão a coluna afundaria perto de zero")
            with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
                st.markdown(f"""
**Em uma frase:** as barras mostram quanto cada densidade produziu, e a linha mostra o que
aconteceu com **{_rot_v}** ao mesmo tempo.

---

**Os dois seletores acima**

- **Híbrido** — o gráfico é de um material por vez. Trocar aqui refaz tudo.
- **Variável na linha** — o que a linha preta acompanha. As barras são sempre produtividade.

**O eixo x** são os grupos de densidade, da menor à maior população, sempre nessa ordem. Cada
grupo reúne as parcelas que tiveram população **real contada** parecida — não o que foi
programado no plantio, e sim o que de fato nasceu.

---

**Como os números são calculados**

- **Barra** = produtividade média das parcelas daquele híbrido naquela densidade, em sacas por
  hectare. Só entram parcelas com produtividade registrada.
- **Linha** = {_rot_v}, com a régua desta variável: {_regua_txt}.
- **A porcentagem embaixo do número** compara aquele grupo com o de **menor densidade**, que é a
  referência. `+17,0% vs 40k` quer dizer que ali se produziu 17% a mais do que no grupo de 40 mil.
- Cada ponto da linha também traz a variação contra a menor densidade, em cinza. Ela existe porque
  a linha pode parecer reta mesmo quando a variável mudou bastante — os dois eixos começam em
  zero, então uma queda de 4% quase não inclina a linha. O número diz o que o desenho não mostra.

---

**⚠️ O erro mais fácil de cometer: comparar altura de barra com altura de ponto**

Eles estão em **escalas diferentes**. As barras são lidas na régua da **esquerda** (sacas por
hectare) e os pontos na régua da **direita** ({_rot_v}). Um ponto acima de uma barra **não
significa nada** — é só onde as duas escalas calharam de se cruzar.

Compare cada série **com ela mesma**, da esquerda para a direita. A pergunta certa é "a barra
subiu?" e "a linha subiu ou desceu?", nunca "o ponto está acima ou abaixo da barra?".

Os dois eixos começam em zero de propósito. Se um deles começasse mais alto, daria para deslizar o
ponto onde as séries parecem se encontrar — e esse cruzamento é justamente o que interessa.

---

**Como interpretar, na prática**

- **Barra sobe e linha desce** — o padrão mais comum com PMG e grãos por fileira. Mais plantas
  dividindo os mesmos recursos: cada espiga rende menos, mas a soma compensa. Enquanto compensar,
  vale adensar.
- **Barra para de subir** — chegou o topo. É a mesma coisa que a curva da Seção 4 mostra, só que
  aqui você vê **por quê**: a linha caiu o bastante para anular o ganho de plantas.
- **Barra desce e a linha de perda sobe** — adensar custou mais do que rendeu, e a perda é o
  motivo.
- **Barra e linha sobem juntas** — ganho sem custo aparente naquela variável.

**Barra clara com `n=` escrito** significa poucas parcelas naquela densidade. Média de três
parcelas oscila muito: leia o número, mas não decida por ele.
""")

            # médias por grupo, na ordem crescente de densidade
            _lin = []
            for _g in _grupos_encontrados:
                _dg = _dh[_dh["pop_grupo"] == _g]
                _sc = pd.to_numeric(_dg["sc_ha"], errors="coerce")
                _sc = _sc[_sc > 0]
                _vv = pd.to_numeric(_dg[_col_v], errors="coerce").dropna()
                if _regua_v == "so_maior" and len(_vv):
                    # mesma régua da tabela de Fenômenos/Qualidade: o zero não entra. Se todas as
                    # parcelas eram zero, o valor é 0 (avaliado e nunca ocorreu), não vazio.
                    _vv = _vv[_vv > 0] if (_vv > 0).any() else pd.Series([0.0])
                if len(_sc) == 0 and len(_vv) == 0:
                    continue
                _lin.append({"grupo": _g, "n": len(_dg),
                             "sc": float(_sc.mean()) if len(_sc) else np.nan,
                             "var": float(_vv.mean()) if len(_vv) else np.nan})
            _dfg = pd.DataFrame(_lin)

            if _dfg.empty or _dfg["sc"].notna().sum() < 2:
                st.info(f"**{_hib_g}** não tem densidades suficientes para o gráfico.")
            else:
                # barras claras em rampa: a cor mantém a ordem da densidade sem competir com a linha
                def _rampa_clara(n):
                    ini, fim = (232, 240, 248), (150, 182, 212)
                    out = []
                    for i in range(max(n, 1)):
                        t = i / (n - 1) if n > 1 else 0.0
                        rgb = tuple(int(round(ini[k] + (fim[k] - ini[k]) * t)) for k in range(3))
                        out.append(f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}")
                    return out
                _cores_bar = _rampa_clara(len(_dfg))

                # se o grupo escolhido não tiver dado para este híbrido, cai no primeiro que tem
                _lin_base = _dfg[(_dfg["grupo"] == _g_base) & _dfg["sc"].notna()]
                if _lin_base.empty and _dfg["sc"].notna().any():
                    _g_base = _dfg.loc[_dfg["sc"].notna(), "grupo"].iloc[0]
                    _lin_base = _dfg[_dfg["grupo"] == _g_base]
                _base_sc = float(_lin_base["sc"].iloc[0]) if not _lin_base.empty else np.nan

                # MESMA folga nos dois eixos: dar mais espaço a uma das séries deslocaria o ponto
                # onde elas parecem se cruzar, que é exatamente o que as travas evitam
                _FOLGA_G = 1.22
                _max_sc = float(_dfg["sc"].max()) if _dfg["sc"].notna().any() else 1
                _max_v = float(_dfg["var"].max()) if _dfg["var"].notna().any() else 1
                _rng_sc, _rng_v = _max_sc * _FOLGA_G, _max_v * _FOLGA_G

                fig_g = go_plt.Figure()

                # ── barras: produtividade (eixo esquerdo) ─────────────────────────
                _texto_bar, _op_bar = [], []
                for _, _r in _dfg.iterrows():
                    _frag = _r["n"] < _N_MIN_GRAF
                    _op_bar.append(0.45 if _frag else 1.0)
                    if pd.isna(_r["sc"]):
                        _texto_bar.append("")
                        continue
                    _delta = ((_r["sc"] - _base_sc) / _base_sc * 100) if pd.notna(_base_sc) and _base_sc else np.nan
                    _t = f"<b>{_r['sc']:.1f}</b> sc/ha"
                    if pd.notna(_delta) and _r["grupo"] != _g_base:
                        # produtividade é a única variável do gráfico em que mais é melhor sem
                        # ressalva, então aqui verde/vermelho dizem o que parecem dizer.
                        _cor_d = ("#1E7A34" if _delta > 0.5 else
                                  ("#C0201E" if _delta < -0.5 else "#6B7280"))
                        _seta = "▲" if _delta > 0.5 else ("▼" if _delta < -0.5 else "=")
                        _t += (f"<br><span style='font-size:12px;color:{_cor_d}'>"
                               f"<b>{_seta} {_delta:+.1f}%</b> vs {_g_base}</span>")
                    if _frag:
                        _t += f"<br><span style='font-size:11px;color:#B45309'>n={int(_r['n'])}</span>"
                    _texto_bar.append(_t)

                # a transparência da barra frágil entra na PRÓPRIA cor (rgba): marker.opacity
                # não aceita uma lista por barra, só um valor único para o traço inteiro
                def _rgba(hx, a):
                    hx = hx.lstrip("#")
                    r, g, b = (int(hx[i:i+2], 16) for i in (0, 2, 4))
                    return f"rgba({r},{g},{b},{a})"
                _cores_final = [_rgba(_c, _o) for _c, _o in zip(_cores_bar, _op_bar)]

                fig_g.add_trace(go_plt.Bar(
                    x=_dfg["grupo"], y=_dfg["sc"], name="Produtividade (sc/ha)",
                    marker=dict(color=_cores_final,
                                line=dict(color="#7FA3C4", width=1)),
                    showlegend=False,
                    text=_texto_bar, textposition="inside", insidetextanchor="start",
                    cliponaxis=False, constraintext="none",
                    textfont=dict(size=13, color="#1A1A1A"),
                    hovertemplate="<b>%{x}</b><br>Produtividade: %{y:.1f} sc/ha<extra></extra>"))

                # traço fantasma só para a legenda: como cada barra tem sua própria cor, o
                # quadradinho da legenda pegava a mais clara de todas e ficava invisível
                fig_g.add_trace(go_plt.Bar(
                    x=[None], y=[None], name="Produtividade (sc/ha)",
                    marker=dict(color="#96B6D4", line=dict(color="#7FA3C4", width=1)),
                    showlegend=True, hoverinfo="skip"))

                # ── linha + pontos: a variável escolhida (eixo direito) ───────────
                fig_g.add_trace(go_plt.Scatter(
                    x=_dfg["grupo"], y=_dfg["var"], name=_rot_v, yaxis="y2",
                    mode="lines+markers",
                    line=dict(color=_CL_LINHA_G, width=2.5),
                    marker=dict(color=_CL_LINHA_G, size=13,
                                line=dict(color="#FFFFFF", width=2)),
                    hovertemplate=("<b>%{x}</b><br>" + _rot_v +
                                   ": %{y:." + str(_dec_v) + "f}<extra></extra>")))

                # rótulo da linha como ANOTAÇÃO, não como texto do traço: só a anotação aceita
                # fundo próprio, e é ele que impede o número de se perder sobre a barra.
                _lin_bv = _dfg[(_dfg["grupo"] == _g_base) & _dfg["var"].notna()]
                _base_v = (float(_lin_bv["var"].iloc[0]) if not _lin_bv.empty
                           else (_dfg["var"].dropna().iloc[0] if _dfg["var"].notna().any()
                                 else np.nan))
                for _gx, _vy, _sy in zip(_dfg["grupo"], _dfg["var"], _dfg["sc"]):
                    if pd.isna(_vy):
                        continue
                    # com os dois eixos em zero, uma variação de 4% desenha uma linha reta. O
                    # número resolve o que o desenho não mostra: a variação vai escrita no rótulo.
                    _dv = ((_vy - _base_v) / abs(_base_v) * 100) if pd.notna(_base_v) and _base_v else np.nan
                    _txt_v = f"<b>{_vy:.{_dec_v}f}</b>"
                    if pd.notna(_dv) and _gx != _g_base:
                        # aqui a seta mostra a DIREÇÃO sem julgar: o PMG cair quando a população
                        # sobe é o mecanismo agronômico esperado, não um problema do material.
                        _seta_v = "▲" if _dv > 0.5 else ("▼" if _dv < -0.5 else "=")
                        _txt_v += (f"<br><span style='font-size:10px;color:#6B7280'>"
                                   f"{_seta_v} {_dv:+.1f}% vs {_g_base}</span>")
                    # o rótulo da barra vive no TERÇO DE BAIXO da barra, longe da faixa onde os
                    # pontos da linha se concentram. Então o rótulo do ponto sobe por padrão; só
                    # desce quando o ponto está tão no alto que o texto sairia do gráfico.
                    _f_var = _vy / _rng_v
                    if _f_var > 0.90:        # quase no teto: desce, senão o texto sai do gráfico
                        _shift = -22
                    elif _f_var < 0.22:      # bem no pé: sobe mais, para escapar do rótulo da barra
                        _shift = 34
                    else:
                        _shift = 20
                    fig_g.add_annotation(
                        x=_gx, y=_vy, yref="y2", text=_txt_v,
                        showarrow=False, yshift=_shift, xanchor="center",
                        font=dict(size=13, color=_CL_LINHA_G),
                        bgcolor="rgba(255,255,255,0.88)", bordercolor=_CL_LINHA_G,
                        borderwidth=1, borderpad=3)

                # ── título dinâmico em linguagem natural ─────────────────────────
                _sc_ok = _dfg.dropna(subset=["sc"])
                _v_ok = _dfg.dropna(subset=["var"])
                _g_top = _sc_ok.loc[_sc_ok["sc"].idxmax(), "grupo"] if len(_sc_ok) else None
                _e_ultimo = (_g_top == _sc_ok["grupo"].iloc[-1]) if len(_sc_ok) else False
                if len(_v_ok) >= 2:
                    _d_var = _v_ok["var"].iloc[-1] - _v_ok["var"].iloc[0]
                    _dir_var = ("cai" if _d_var < 0 else "sobe") if abs(_d_var) > 0 else "não muda"
                    _pct_var = (abs(_d_var) / abs(_v_ok["var"].iloc[0]) * 100
                                if _v_ok["var"].iloc[0] else np.nan)
                else:
                    _dir_var, _pct_var = "não muda", np.nan
                _frase_prod = (f"a produção ainda subia na maior densidade testada (<b>{_g_top}</b>)"
                               if _e_ultimo else
                               f"a produção foi maior em <b>{_g_top}</b> e recuou depois")
                _frase_var = (f"o {_rot_v.split(' (')[0].lower()} <b>{_dir_var}</b>"
                              + (f" {_pct_var:.0f}% da menor para a maior densidade"
                                 if pd.notna(_pct_var) else ""))
                st.markdown(f"""
    <div style="margin:0.6rem 0 0.2rem;">
        <p style="font-size:1.1rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.45;">
            Em <b>{_hib_g}</b>, {_frase_prod} — enquanto {_frase_var}.</p>
    </div>""", unsafe_allow_html=True)

                # ── eixos: os DOIS ancorados em zero ─────────────────────────────
                fig_g.update_layout(
                    height=500, plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif"),
                    # barras estreitas: sobra respiro entre os grupos e o ponto da linha fica
                    # visualmente solto, sem parecer parte da barra
                    margin=dict(t=60, b=60, l=70, r=80), bargap=0.74,
                    legend=dict(orientation="h", y=1.10, x=0, xanchor="left",
                                font=dict(size=12, color="#1A1A1A")),
                    xaxis=dict(title=dict(text="<b>Grupo de densidade (população real)</b>",
                                          font=dict(size=13, color="#1A1A1A", weight="bold")),
                               tickfont=dict(size=14, color="#1A1A1A", weight="bold"),
                               showgrid=False, zeroline=False),
                    yaxis=dict(title=dict(text="<b>Produtividade (sc/ha)</b>",
                                          font=dict(size=13, color="#5B7FA3", weight="bold")),
                               tickfont=dict(size=12, color="#5B7FA3"),
                               range=[0, _rng_sc],                 # começa em ZERO
                               showgrid=True, gridcolor="#EEEEEE", zeroline=False),
                    yaxis2=dict(title=dict(text=f"<b>{_rot_v}</b>",
                                           font=dict(size=13, color=_CL_LINHA_G, weight="bold")),
                                tickfont=dict(size=12, color=_CL_LINHA_G),
                                range=[0, _rng_v],                  # começa em ZERO, mesma folga da esquerda
                                overlaying="y", side="right", showgrid=False, zeroline=False))
                st.plotly_chart(fig_g, use_container_width=True)

                st.warning(
                    f"**Atenção às duas escalas.** As barras são lidas no eixo da **esquerda** "
                    f"(sc/ha) e os pontos no eixo da **direita** ({_rot_v}). Alturas de séries "
                    f"diferentes **não se comparam** entre si — um ponto acima de uma barra não "
                    f"significa nada. Compare cada série com ela mesma, da esquerda para a direita. "
                    f"Os dois eixos começam em zero de propósito, para que a escala não desloque o "
                    f"ponto onde as séries parecem se cruzar.")

                st.caption(
                    f"Barras claras = produtividade média do grupo, com o valor e a diferença contra "
                    f"a população de referência ({_g_base}) · **verde e vermelho só nas barras**: em "
                    f"produtividade, mais é melhor. Na linha a seta mostra apenas a direção, sem cor "
                    f"— o PMG cair quando a população sobe é o mecanismo esperado, não um defeito do "
                    f"híbrido · linha escura = {_rot_v} médio "
                    f"do grupo · barra esmaecida com **n=** = poucas parcelas nessa densidade, número "
                    f"menos confiável (mínimo de {_N_MIN_GRAF}). Sem linha de tendência ajustada: a "
                    f"curva de resposta à população está na Seção 4, sobre a população contínua.")

                _tab_g = _dfg.rename(columns={"grupo": "Densidade", "n": "n parcelas",
                                              "sc": "sc/ha", "var": _rot_v}).copy()
                _tab_g["sc/ha"] = _tab_g["sc/ha"].round(1)
                _tab_g[_rot_v] = _tab_g[_rot_v].round(_dec_v)
                exportar_excel(_tab_g, f"densidade_{pref}_{_hib_g}.xlsx",
                               "⬇️ Exportar dados do gráfico", key=f"{pref}_xlsx")


_grafico_barras_linha(
    "Componentes de Produção — Leitura Gráfica",
    "Produtividade e componentes lado a lado, ao longo das populações",
    _VARS_GRAF_COMP, pref="den_graf", prefixo_padrao="PMG")


st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — SANIDADE POR GRUPO DE DENSIDADE
# ══════════════════════════════════════════════════════════════════════════════
# Mesmo formato da Seção 5 (cabeçalho de dois níveis), mas a célula NÃO é uma média.
# Nota de doença é escala ORDINAL de 1 a 9, e 9 é o melhor (mais resistente) — média de nota
# ordinal não tem significado defensável, e a média de "3" e "9" daria "6", que sugeriria
# tolerância moderada onde na verdade houve um local muito doente e outro limpo.
# Por isso a célula traz a MODA (a nota que mais se repetiu), com empate resolvido para a
# MENOR nota (lado suscetível). É a mesma régua da página de Sanidade — as duas páginas
# precisam dar o mesmo número para o mesmo recorte.

_DOENCAS_DEN = {
    "Turcicum":          "nota_turcicum",
    "Cercospora":        "nota_cercospora",
    "Mancha branca":     "nota_mancha_branca",
    "Bipolaris":         "nota_bipolaris",
    "Ferrugem tropical": "nota_ferrugem_tropical",
    "Enfezamento":       "nota_enfezamento",
}
_DOENCAS_DEN = {k: v for k, v in _DOENCAS_DEN.items() if v in ta_filtrado.columns}

# classes de reação — escala INVERSA (9 = mais resistente). Mesma paleta da página de Sanidade.
_ORDEM_CLASS_D = ["AS", "S", "MT", "T", "R"]
_COR_CLASS_D = {"AS": "#8B0000", "S": "#E63946", "MT": "#FFD600",
                "T": "#70C96E", "R": "#1E7A34"}
_COR_TXT_CLASS_D = {"AS": "#FFFFFF", "S": "#FFFFFF", "MT": "#1A1A1A",
                    "T": "#1A1A1A", "R": "#FFFFFF"}
_LABEL_CLASS_D = {"AS": "AS — Altamente suscetível (nota 1–2)", "S": "S — Suscetível (3–4)",
                  "MT": "MT — Moderadamente tolerante (5–6)", "T": "T — Tolerante (7–8)",
                  "R": "R — Resistente (9)"}


def _nota_classe_den(nota):
    """Nota 1–9 -> sigla de classe. Mesma régua do pipeline e da página de Sanidade."""
    if nota is None or (isinstance(nota, float) and pd.isna(nota)):
        return None
    n = float(nota)
    return "AS" if n <= 2 else ("S" if n <= 4 else ("MT" if n <= 6 else ("T" if n <= 8 else "R")))


def _resumo_doenca_den(grp, col_nota):
    """(nota típica, incidência %, classe) de uma doença num recorte de parcelas.

    Nota  = moda das notas válidas; empate fica com a MENOR (lado suscetível).
    Inc.  = locais com detecção (nota 1–5) ÷ locais em que a doença foi avaliada.
    Nota 0 é "não avaliado" e sai de todas as contas.
    """
    if col_nota not in grp.columns or "cod_fazenda" not in grp.columns:
        return None, None, None
    g = grp[["cod_fazenda", col_nota]].copy()
    g[col_nota] = pd.to_numeric(g[col_nota], errors="coerce").where(lambda x: x > 0)
    g = g.dropna(subset=[col_nota])
    if g.empty:
        return None, None, None
    _nota = round(float(g[col_nota].mode().min()), 1)
    _n_aval = g["cod_fazenda"].nunique()
    _n_com = g.loc[g[col_nota].between(1, 5), "cod_fazenda"].nunique()
    _inc = round(_n_com / _n_aval * 100, 1) if _n_aval else None
    return _nota, _inc, _nota_classe_den(_nota)


secao_titulo("Sanidade",
             "A reação de cada híbrido às doenças ao longo das populações",
             contexto_str)

if not _DOENCAS_DEN or not _grupos_encontrados or "dePara" not in ta_filtrado.columns:
    st.info("Sem colunas de nota de doença ou sem grupos de densidade nos filtros ativos.")
else:
    with st.popover("ℹ️ Como ler a tabela", use_container_width=False):
        st.markdown("""
**Os controles acima da tabela**

- **Doenças** — quais entram. A lista vem **ordenada pela incidência no recorte** e mostra o
  percentual ao lado do nome, então a primeira da lista é a que mais apareceu. Isso importa:
  abrir numa doença que não ocorreu faz parecer que a densidade não afeta nada, quando o que
  faltou foi doença. Cada doença abre três colunas, então escolha poucas por vez.
- **Mostrar n** — acrescenta, no fim de cada bloco, quantas parcelas sustentam aqueles números.
- **Bloco geral** — liga ou desliga a coluna **Todas as densidades**, que junta tudo e serve de
  referência para o resto da linha.

**A escala é ao contrário do que a intuição sugere: 9 é o melhor.** A nota mede o quanto a planta
resistiu, não o quanto adoeceu. Nota 9 é planta limpa; nota 1 é planta tomada pela doença.

| Nota | Classe | Leitura |
|---|---|---|
| 1–2 | **AS** | altamente suscetível |
| 3–4 | **S** | suscetível |
| 5–6 | **MT** | moderadamente tolerante |
| 7–8 | **T** | tolerante |
| 9 | **R** | resistente |

**O fundo da célula é a classe**, na mesma paleta da página de Sanidade: vermelho escuro para
altamente suscetível, até verde escuro para resistente. Aqui a cor está no fundo, e não no texto
como na tabela de componentes, porque a classe é uma **categoria** — não um desvio contra a média.

**A célula não é uma média.** Nota de doença é escala ordinal: a média entre um local com nota 3 e
outro com nota 9 daria 6, sugerindo tolerância moderada onde na verdade houve um local muito doente
e outro limpo. A célula traz a **moda** — a nota que mais se repetiu nas parcelas — e, em caso de
empate, fica a **menor** nota, pelo lado suscetível. É a mesma régua da página de Sanidade, então
o mesmo recorte dá o mesmo número nas duas telas.

**Nota 0 é "não avaliado"**, não é nota ruim: sai de todas as contas.

**Cada doença abre três colunas**, exatamente como na página de Sanidade:

- **Nota** — a que mais se repetiu nas parcelas (moda; empate fica com a pior)
- **%** — a incidência, calculada como abaixo
- **Classe** — derivada da Nota, e é ela que recebe a cor

Nota e % respondem perguntas diferentes: a nota diz **o quanto** a doença atacou onde apareceu; a
incidência diz **em quantos lugares** apareceu. Um híbrido pode ser classe T com incidência de
100% — pega em todo lugar, mas de leve. E pode ser AS com 20% — quando pega, arrasa, mas só
apareceu em um quinto dos locais.

**Como a incidência é calculada:**

`locais em que a doença foi DETECTADA ÷ locais em que a doença foi AVALIADA × 100`

- um **local** entra no numerador se ao menos uma parcela dele teve nota entre **1 e 5**;
- o denominador são os locais com nota válida (maior que zero) para **aquela** doença;
- nota 0 é "não avaliado" e sai das duas contas, então cada doença tem o seu próprio denominador
  — duas doenças na mesma linha podem estar sobre números de locais diferentes;
- a conta é por **local**, não por parcela: um local com dez parcelas pesa igual a um com duas.

A régua de detecção é nota ≤ 5, herdada da página de Sanidade. Repare que ela inclui a nota 5, que
é MT — ou seja, "detectada" abrange o moderadamente tolerante, não só o doente grave.

**Cuidado com a incidência dentro de um grupo de densidade.** Ela é uma fração de locais, e dentro
de um grupo sobram poucos locais. Com três locais, a incidência só pode dar 0%, 33%, 67% ou 100% —
os saltos são grandes e não significam variação real. O bloco **Todas as densidades** é o único em
que ela costuma ter locais suficientes.

**Leia na horizontal** para ver se adensar piora a sanidade daquele híbrido. É a leitura principal
desta seção: dossel mais fechado retém umidade na folha por mais tempo, o que costuma favorecer
doença foliar. Se existir esse efeito, ele aparece como a linha do híbrido escurecendo para o
vermelho conforme a densidade sobe.

**Sobre o n.** As parcelas se dividem entre os grupos, então cada célula costuma ter poucas. Ligue
**Mostrar n** para ver quantas sustentam cada bloco; células com poucas parcelas ficam esmaecidas
e com asterisco. O bloco **Todas as densidades** vem primeiro e é o mais confiável.
""")

    _c_s1, _c_s2, _c_s3 = st.columns([3, 1.4, 1.4])
    with _c_s1:
        # o rótulo já traz a incidência da doença no recorte inteiro, e a lista vem ordenada
        # da que mais apareceu para a que menos apareceu. Sem isso é fácil abrir a seção numa
        # doença que praticamente não ocorreu e concluir que a densidade não afeta nada —
        # quando o que faltou foi doença, não efeito.
        _inc_geral = {}
        for _rot_d, _col_d in _DOENCAS_DEN.items():
            _, _i_g, _ = _resumo_doenca_den(ta_filtrado, _col_d)
            _inc_geral[_rot_d] = _i_g
        _ord_doen = sorted(_DOENCAS_DEN, key=lambda d: -(_inc_geral.get(d) or -1))
        _rot_doen = {d: (f"{d} — {_inc_geral[d]:.0f}% dos locais" if _inc_geral.get(d) is not None
                         else f"{d} — não avaliada") for d in _ord_doen}
        _de_rotulo = {v: k for k, v in _rot_doen.items()}

        _doen_sel = st.multiselect(
            "Doenças", options=[_rot_doen[d] for d in _ord_doen],
            default=[_rot_doen[d] for d in _ord_doen[:1]],
            key="den_san_doen",
            help="Cada doença abre três colunas (Nota, %, Classe) DENTRO de cada grupo de "
                 "densidade — com 5 blocos, cada doença custa 15 colunas. A lista vem ordenada "
                 "pela incidência no recorte, então a primeira é a que mais apareceu. As colunas "
                 "aparecem na ordem em que você escolher.")
    with _c_s2:
        _mostrar_n_s = st.checkbox("Mostrar n", value=False, key="den_san_n")
    with _c_s3:
        _geral_s = st.checkbox("Bloco geral", value=True, key="den_san_geral")

    _N_MIN_SAN = 5
    _SUB_SAN = ["Nota", "%", "Classe"]   # exatamente as colunas da página de Sanidade

    if not _doen_sel:
        st.info("Selecione ao menos uma doença.")
    else:
        _doen_nomes = [_de_rotulo[r] for r in _doen_sel if r in _de_rotulo]
        _doen = [(d, _DOENCAS_DEN[d]) for d in _doen_nomes if d in _DOENCAS_DEN]
        _blocos_s = ([("Todas as densidades", None)] if _geral_s else []) + \
                    [(g, g) for g in _grupos_encontrados]
        _hibs_s = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
        _status_s = (ta_filtrado.dropna(subset=["dePara"])
                     .groupby("dePara")["status_material"].first().to_dict()
                     if "status_material" in ta_filtrado.columns else {})

        _cel_s = {}
        for _nome_bl, _grp in _blocos_s:
            _dbl = ta_filtrado if _grp is None else ta_filtrado[ta_filtrado["pop_grupo"] == _grp]
            for _h in _hibs_s:
                _dh = _dbl[_dbl["dePara"] == _h]
                _cel_s[(_h, _nome_bl, "__n__")] = len(_dh)
                for _rot_d, _col_d in _doen:
                    _cel_s[(_h, _nome_bl, _rot_d)] = _resumo_doenca_den(_dh, _col_d)

        # rampa de densidade (a mesma da Seção 5)
        def _rampa_den_s(n):
            ini, fim = (214, 228, 242), (18, 57, 92)
            out = []
            for i in range(max(n, 1)):
                t = i / (n - 1) if n > 1 else 0.0
                rgb = tuple(int(round(ini[k] + (fim[k] - ini[k]) * t)) for k in range(3))
                lum = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
                out.append((f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                            "#1A1A1A" if lum > 150 else "#FFFFFF"))
            return out
        _rampa_s = _rampa_den_s(len(_grupos_encontrados))
        _cor_bloco_s, _ig = {}, 0
        for _nome_bl, _grp in _blocos_s:
            if _grp is None:
                _cor_bloco_s[_nome_bl] = ("#3A3A3A", "#FFFFFF")
            else:
                _cor_bloco_s[_nome_bl] = _rampa_s[min(_ig, len(_rampa_s) - 1)]
                _ig += 1

        _COR_ST_S = {"CHECK": "#F4B184", "STINE": "#2976B6", "EXP": "#00FF00", "DP2": "#C4DFB4"}
        _COR_TX_S = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                     "DP2": "#1A1A1A", "": "#000000"}
        _n_sub_s = len(_doen) * len(_SUB_SAN) + (1 if _mostrar_n_s else 0)

        # ── HTML com TRÊS níveis de cabeçalho: grupo > doença > Nota/Inc./Classe ──
        _html_s = """
<style>
.tb-san { width:100%; border-collapse:collapse; font-size:13px;
    font-family:'Helvetica Neue',sans-serif; }
.tb-san th { background:#F2F2F2; color:#000 !important; padding:6px 8px; text-align:center;
    border:1px solid #ccc; white-space:nowrap; font-weight:700; }
.tb-san th.grupo { font-size:14px; letter-spacing:0.03em; }
.tb-san th.doenca { background:#E8E8E8; font-size:13px; }
.tb-san th.sub { font-size:12px; font-weight:600; color:#444 !important; }
.tb-san th.hib { background:#4A4A4A; color:#FFF !important; text-align:left; }
.tb-san td { padding:5px 8px; border:1px solid #ddd; text-align:center; white-space:nowrap;
    font-weight:600; }
.tb-san td.hib { text-align:left; font-weight:600; }
.tb-san td.sep, .tb-san th.sep { border-left:2px solid #999; }
.tb-san td.sepd, .tb-san th.sepd { border-left:1px solid #aaa; }
.tb-san td.frag { opacity:0.55; }
.tb-san td.num { background:#FFFFFF !important; color:#1A1A1A !important; font-weight:500; }
.tb-san td.ncel { color:#6B7280 !important; font-size:12px; font-style:italic;
    background:#FFF !important; font-weight:400; }
</style>
<table class="tb-san">
<thead>
<tr><th class="hib" rowspan="3">Híbrido</th>"""
        # nível 1 — grupo de densidade
        for _nome_bl, _ in _blocos_s:
            _bgh, _fgh = _cor_bloco_s[_nome_bl]
            _html_s += (f'<th class="grupo sep" style="background:{_bgh};color:{_fgh} !important;'
                        f'border-color:{_bgh};" colspan="{_n_sub_s}">{_nome_bl}</th>')
        # nível 2 — doença dentro de cada grupo
        _html_s += "</tr><tr>"
        for _nome_bl, _ in _blocos_s:
            _bgh, _ = _cor_bloco_s[_nome_bl]
            for _j, (_rot_d, _) in enumerate(_doen):
                _html_s += (f'<th class="doenca {"sep" if _j == 0 else "sepd"}" '
                            f'style="border-top:4px solid {_bgh};" '
                            f'colspan="{len(_SUB_SAN)}">{_rot_d}</th>')
            if _mostrar_n_s:
                _html_s += (f'<th class="doenca sepd" rowspan="2" '
                            f'style="border-top:4px solid {_bgh};">n</th>')
        # nível 3 — as três colunas de cada doença
        _html_s += "</tr><tr>"
        for _nome_bl, _ in _blocos_s:
            for _j, (_rot_d, _) in enumerate(_doen):
                for _k, _sub in enumerate(_SUB_SAN):
                    _cls_h = "sep" if (_j == 0 and _k == 0) else ("sepd" if _k == 0 else "")
                    _html_s += f'<th class="sub {_cls_h}">{_sub}</th>'
        _html_s += "</tr></thead><tbody>"

        for _h in _hibs_s:
            _st = _status_s.get(_h, "")
            _html_s += (f'<tr><td class="hib" style="background:{_COR_ST_S.get(_st, "#FFF")};'
                        f'color:{_COR_TX_S.get(_st, "#000")} !important;">{_h}</td>')
            for _nome_bl, _ in _blocos_s:
                _n_bl = _cel_s[(_h, _nome_bl, "__n__")]
                _frag = _n_bl < _N_MIN_SAN
                for _j, (_rot_d, _) in enumerate(_doen):
                    _nota, _inc, _cls = _cel_s[(_h, _nome_bl, _rot_d)]
                    _bg = _COR_CLASS_D.get(_cls, "#FFFFFF")
                    _fg = _COR_TXT_CLASS_D.get(_cls, "#9CA3AF")
                    _ast = "*" if (_frag and _nota is not None) else ""
                    _base_cls = "frag" if _frag else ""
                    _c0 = " ".join(x for x in [("sep" if _j == 0 else "sepd"), _base_cls] if x)
                    # mesma disposição da página de Sanidade: Nota, %, Classe. Só a Classe leva
                    # o fundo da cor — na Sanidade é assim, e mantém Nota e % legíveis.
                    _html_s += (f'<td class="{_c0} num {_base_cls}">'
                                f'{"—" if _nota is None else f"{_nota:.1f}"}{_ast}</td>')
                    _html_s += (f'<td class="num {_base_cls}">'
                                f'{"—" if _inc is None else f"{_inc:.1f}"}</td>')
                    _html_s += (f'<td class="{_base_cls}" style="background:{_bg};'
                                f'color:{_fg} !important;">{_cls or "—"}</td>')
                if _mostrar_n_s:
                    _html_s += f'<td class="ncel">{_n_bl}</td>'
            _html_s += "</tr>"
        _html_s += "</tbody></table>"

        components.html(_html_s, height=min(120 + (len(_hibs_s) + 2) * 32, 700), scrolling=True)

        # legenda das classes
        _leg = "".join(
            f'<span style="display:inline-block;margin:0 10px 6px 0;padding:3px 9px;'
            f'border-radius:3px;font-size:12px;font-weight:700;'
            f'background:{_COR_CLASS_D[c]};color:{_COR_TXT_CLASS_D[c]};">'
            f'{_LABEL_CLASS_D[c]}</span>' for c in _ORDEM_CLASS_D)
        st.markdown(_leg, unsafe_allow_html=True)

        st.caption(
            f"**Nota** = a que mais se repetiu nas parcelas (moda; empate fica com a pior) · "
            f"**%** = locais com detecção (nota 1–5) ÷ locais avaliados para aquela doença · "
            f"**Classe** = derivada da Nota e colorida por ela · **escala 1–9, onde 9 é o "
            f"melhor** · nota 0 = não "
            f"avaliado, fica fora das contas · **\\*** = menos de {_N_MIN_SAN} parcelas na "
            f"célula, valor frágil · mesma régua da página de Sanidade.")

        # ── Exportação: espelho fiel da tabela na tela ─────────────────────────
        # Mesmos três níveis de cabeçalho (densidade > doença > Nota/%/Classe), a rampa de
        # densidade, o fundo da classe na coluna Classe e o asterisco das células frágeis.
        def _xlsx_sanidade():
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sanidade"
            _hx = lambda c: str(c).replace("#", "").upper()
            thin = Side(style="thin", color="DDDDDD")
            grosso = Side(style="medium", color="999999")
            centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # coluna do híbrido ocupa os três níveis de cabeçalho
            ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
            _c1 = ws.cell(row=1, column=1, value="Híbrido")
            _c1.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            _c1.fill = PatternFill("solid", start_color="4A4A4A")
            _c1.alignment = Alignment(horizontal="left", vertical="center")
            ws.column_dimensions["A"].width = 18

            _ci = 2
            for _nome_bl, _ in _blocos_s:
                _bgh, _fgh = _cor_bloco_s[_nome_bl]
                _ini_bl = _ci
                for _rot_d, _ in _doen:
                    _ini_d = _ci
                    for _sub in _SUB_SAN:
                        _ch = ws.cell(row=3, column=_ci, value=_sub)
                        _ch.font = Font(bold=True, name="Arial", size=9, color="444444")
                        _ch.fill = PatternFill("solid", start_color="F2F2F2")
                        _ch.alignment = centro
                        _ch.border = Border(left=grosso if _ci == _ini_bl else thin,
                                            right=thin, top=thin, bottom=thin)
                        ws.column_dimensions[get_column_letter(_ci)].width = 9
                        _ci += 1
                    # nível 2 — nome da doença sobre as três colunas dela
                    ws.merge_cells(start_row=2, start_column=_ini_d, end_row=2, end_column=_ci - 1)
                    _cd = ws.cell(row=2, column=_ini_d, value=_rot_d)
                    _cd.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
                    _cd.fill = PatternFill("solid", start_color="E8E8E8")
                    _cd.alignment = centro
                    _cd.border = Border(left=grosso if _ini_d == _ini_bl else thin, right=thin,
                                        top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                if _mostrar_n_s:
                    ws.merge_cells(start_row=2, start_column=_ci, end_row=3, end_column=_ci)
                    _cn = ws.cell(row=2, column=_ci, value="n")
                    _cn.font = Font(bold=True, name="Arial", size=9, color="6B7280", italic=True)
                    _cn.fill = PatternFill("solid", start_color="F2F2F2")
                    _cn.alignment = centro
                    _cn.border = Border(left=thin, right=thin,
                                        top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                    ws.column_dimensions[get_column_letter(_ci)].width = 7
                    _ci += 1
                # nível 1 — grupo de densidade sobre todas as colunas do bloco
                ws.merge_cells(start_row=1, start_column=_ini_bl, end_row=1, end_column=_ci - 1)
                _cb = ws.cell(row=1, column=_ini_bl, value=_nome_bl)
                _cb.font = Font(bold=True, name="Arial", size=11, color=_hx(_fgh))
                _cb.fill = PatternFill("solid", start_color=_hx(_bgh))
                _cb.alignment = centro
                _cb.border = Border(left=grosso, right=thin, top=thin, bottom=thin)
            ws.row_dimensions[1].height = 20
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 18

            # ── corpo ────────────────────────────────────────────────────────
            _ri = 4
            for _h in _hibs_s:
                _st = _status_s.get(_h, "")
                _ch = ws.cell(row=_ri, column=1, value=_h)
                _ch.font = Font(bold=True, name="Arial", size=10,
                                color=_hx(_COR_TX_S.get(_st, "#000000")))
                _ch.fill = PatternFill("solid", start_color=_hx(_COR_ST_S.get(_st, "#FFFFFF")))
                _ch.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                _ci = 2
                for _nome_bl, _ in _blocos_s:
                    _n_bl = _cel_s[(_h, _nome_bl, "__n__")]
                    _frag = _n_bl < _N_MIN_SAN
                    for _j, (_rot_d, _) in enumerate(_doen):
                        _nota, _inc, _cls = _cel_s[(_h, _nome_bl, _rot_d)]
                        for _k, _sub in enumerate(_SUB_SAN):
                            _cel_x = ws.cell(row=_ri, column=_ci)
                            _borda = Border(left=grosso if (_j == 0 and _k == 0) else thin,
                                            right=thin, top=thin, bottom=thin)
                            if _sub == "Nota":
                                _cel_x.value = (round(float(_nota), 1)
                                                if _nota is not None else None)
                                _cel_x.number_format = '0.0"*"' if (_frag and _nota is not None) else "0.0"
                                _cel_x.font = Font(name="Arial", size=10, color="1A1A1A",
                                                   italic=_frag)
                            elif _sub == "%":
                                _cel_x.value = (round(float(_inc), 1)
                                                if _inc is not None else None)
                                _cel_x.number_format = "0.0"
                                _cel_x.font = Font(name="Arial", size=10, color="1A1A1A",
                                                   italic=_frag)
                            else:   # Classe — única que leva o fundo colorido, como na tela
                                _cel_x.value = _cls
                                _cel_x.font = Font(bold=True, name="Arial", size=10,
                                                   color=_hx(_COR_TXT_CLASS_D.get(_cls, "#9CA3AF")),
                                                   italic=_frag)
                                if _cls:
                                    _cel_x.fill = PatternFill(
                                        "solid", start_color=_hx(_COR_CLASS_D[_cls]))
                            _cel_x.alignment = centro
                            _cel_x.border = _borda
                            _ci += 1
                    if _mostrar_n_s:
                        _cn = ws.cell(row=_ri, column=_ci, value=int(_n_bl))
                        _cn.font = Font(name="Arial", size=9, color="6B7280", italic=True)
                        _cn.alignment = centro
                        _cn.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                        _ci += 1
                _ri += 1

            # ── notas de rodapé (a mesma legenda da tela) ────────────────────
            _ri += 2
            for _txt in [
                "Escala 1–9, onde 9 é o MELHOR (mais resistente). Nota 0 = não avaliado, "
                "fica fora de todas as contas.",
                "Nota = a que mais se repetiu nas parcelas (moda; empate fica com a pior). "
                "Classe: AS 1–2 · S 3–4 · MT 5–6 · T 7–8 · R 9.",
                "% = locais com detecção (nota 1–5) ÷ locais avaliados para aquela doença. "
                "Conta por local, não por parcela; cada doença tem seu próprio denominador.",
                f"* = menos de {_N_MIN_SAN} parcelas na célula — valor frágil.",
                f"Filtros ativos: {contexto_str}",
            ]:
                _cn = ws.cell(row=_ri, column=1, value=_txt)
                _cn.font = Font(name="Arial", size=8, color="6B7280", italic=True)
                _ri += 1

            ws.freeze_panes = "B4"
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        st.download_button("⬇️ Exportar sanidade",
                           data=_xlsx_sanidade(),
                           file_name="densidade_sanidade.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="den_san_xlsx")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — LEITURA GRÁFICA DA SANIDADE POR DENSIDADE
# ══════════════════════════════════════════════════════════════════════════════
# Dois atos para a mesma pergunta: adensar piora a sanidade?
#
# Ato 1 — distribuição das PARCELAS por classe dentro de cada densidade. Isso mostra o que a
#   tabela esconde: a tabela traz a MODA, e moda não se mexe enquanto a maioria não vira. Um
#   grupo pode ir de 5% para 25% de parcelas suscetíveis e continuar com moda 9. A barra 100%
#   empilhada mostra a fatia crescendo.
# Ato 2 — mapa híbrido × densidade, para achar QUEM piora. O agregado pode não mostrar efeito
#   nenhum porque um híbrido piora e outro melhora, e a média cancela os dois.
#
# Atenção à unidade: aqui a conta é por PARCELA. A coluna % da tabela é por LOCAL. São
# denominadores diferentes de propósito — e está escrito na legenda de cada ato.

secao_titulo("Sanidade — Leitura Gráfica",
             "Como as classes de reação se distribuem ao longo das populações",
             contexto_str)

if not _DOENCAS_DEN or not _grupos_encontrados or "dePara" not in ta_filtrado.columns:
    st.info("Sem colunas de nota de doença ou sem grupos de densidade nos filtros ativos.")
else:
    # ordenação própria, para esta seção não depender de variáveis criadas dentro dos
    # blocos condicionais da Seção 7
    _inc_g8 = {}
    for _rot_d8, _col_d8 in _DOENCAS_DEN.items():
        _, _i8, _ = _resumo_doenca_den(ta_filtrado, _col_d8)
        _inc_g8[_rot_d8] = _i8
    _ord8 = sorted(_DOENCAS_DEN, key=lambda d: -(_inc_g8.get(d) or -1))
    _rot8 = {d: (f"{d} — {_inc_g8[d]:.0f}% dos locais" if _inc_g8.get(d) is not None
                 else f"{d} — não avaliada") for d in _ord8}
    _de_rot8 = {v: k for k, v in _rot8.items()}

    _cg1, _cg2 = st.columns([2, 2])
    with _cg1:
        _d_graf_rot = st.selectbox(
            "Doença", options=[_rot8[d] for d in _ord8], key="den_sang_doen",
            help="A lista vem ordenada pela incidência no recorte: a primeira é a que mais "
                 "apareceu, e é nela que faz sentido procurar efeito de densidade.")
        _d_graf = _de_rot8.get(_d_graf_rot)
    with _cg2:
        _hibs_sg = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
        _hib_sg = st.multiselect("Híbridos", options=_hibs_sg, default=_hibs_sg,
                                 key="den_sang_hib")

    _col_dg = _DOENCAS_DEN.get(_d_graf)
    _base_sg = ta_filtrado[ta_filtrado["dePara"].isin(_hib_sg)] if _hib_sg else ta_filtrado.iloc[0:0]

    if not _col_dg or _base_sg.empty:
        st.info("Selecione ao menos um híbrido.")
    else:
        _notas_sg = pd.to_numeric(_base_sg[_col_dg], errors="coerce")
        _val_sg = _base_sg[(_notas_sg > 0)].copy()      # nota 0 = não avaliado
        _val_sg["_nota"] = pd.to_numeric(_val_sg[_col_dg], errors="coerce")
        _val_sg["_classe"] = _val_sg["_nota"].apply(_nota_classe_den)

        if _val_sg.empty:
            st.info(f"**{_d_graf}** não foi avaliada nas parcelas deste recorte.")
        else:
            # ── Ato 1 — composição por classe dentro de cada densidade ───────
            st.markdown(
                '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                'letter-spacing:0.07em;margin:1.2rem 0 6px;">'
                'Ato 1 — Como as parcelas se distribuem entre as classes</p>',
                unsafe_allow_html=True)
            st.markdown("Cada barra é uma densidade e soma 100% das parcelas avaliadas dela. "
                        "Se adensar piorar a sanidade, as faixas vermelhas crescem de cima "
                        "para baixo.")

            with st.popover("ℹ️ Como este gráfico é calculado", use_container_width=False):
                st.markdown(f"""
**Os dois seletores acima**

- **Doença** — a lista vem ordenada pela incidência no recorte, com o percentual ao lado do nome.
  A primeira é a que mais apareceu, e é nela que faz sentido procurar efeito de densidade.
- **Híbridos** — quais entram na conta. Tirar híbridos muda as barras, porque elas somam todas as
  parcelas dos materiais selecionados.

**Como é calculado, passo a passo**

1. Pego as parcelas do recorte para a doença **{_d_graf}** e os híbridos selecionados.
2. Descarto as de **nota 0** — zero é "não avaliado", não é planta sadia. Se a parcela não foi
   olhada para essa doença, ela não pode entrar nem no numerador nem no denominador.
3. Converto a nota de cada parcela em classe pela régua fixa: **1–2 AS · 3–4 S · 5–6 MT ·
   7–8 T · 9 R**. Aqui **não há moda**: cada parcela vira uma classe própria.
4. Dentro de cada grupo de densidade, conto quantas parcelas caíram em cada classe e divido
   pelo total de parcelas avaliadas daquele grupo. É por isso que cada barra fecha em 100%.
5. Rótulos aparecem só nas fatias com 7% ou mais — abaixo disso o texto não caberia dentro
   da faixa.

**Por que barra de composição e não a nota média ou a moda**

A tabela da seção anterior mostra a **moda**, a nota que mais se repete. Moda é insensível: ela
não se mexe enquanto a maioria não virar. Num teste com efeito real de densidade, a fatia de
parcelas suscetíveis subiu de 1% para 10% enquanto a moda quase não mudou — a tabela mostraria
"sem efeito" e a barra mostra a faixa crescendo. É por isso que este ato existe.

Média também não serviria: nota é escala **ordinal**, e a média entre uma parcela nota 3 e outra
nota 9 daria 6, sugerindo tolerância moderada onde na verdade houve uma parcela tomada pela doença
e outra limpa.

**Como interpretar**

- As barras estão em ordem de densidade, **menor no topo**. Leia de cima para baixo: é o efeito
  de adensar.
- Se as faixas **vermelhas (AS e S) engordam** descendo, adensar piorou a sanidade. A explicação
  agronômica é o dossel fechado, que segura umidade na folha por mais tempo.
- Se o **verde escuro (R) encolhe** descendo, o mesmo sinal aparece pelo outro lado.
- Se as barras ficam parecidas, a densidade não alterou a sanidade **neste recorte** — o que não
  é o mesmo que "não altera nunca".
- O aviso abaixo do gráfico compara a fatia suscetível (AS + S) da menor com a da maior densidade
  e sinaliza quando a diferença passa de 5 pontos percentuais.

**Duas armadilhas**

**A unidade é a PARCELA.** A coluna % da tabela conta **locais**. Os dois números não batem e não
deveriam bater: aqui é "quantas parcelas estavam em cada estado", lá é "em quantas fazendas a
doença apareceu".

**Barra não é experimento.** Os grupos de densidade não têm os mesmos locais nas mesmas
proporções. Se as densidades altas ficaram concentradas numa região de mais pressão de doença, a
faixa vermelha cresce por causa da região e não da população. Antes de concluir, olhe o Ato 2:
efeito real de densidade costuma aparecer na maioria dos híbridos, não em um só.
""")

            _linhas_sg = []
            for _g in _grupos_encontrados:
                _dg = _val_sg[_val_sg["pop_grupo"] == _g]
                if _dg.empty:
                    continue
                _cont = _dg["_classe"].value_counts()
                _tot = int(_cont.sum())
                _linhas_sg.append({"grupo": _g, "n": _tot,
                                   **{c: int(_cont.get(c, 0)) for c in _ORDEM_CLASS_D}})
            _df_sg = pd.DataFrame(_linhas_sg)

            if _df_sg.empty:
                st.info("Sem parcelas avaliadas por grupo de densidade.")
            else:
                _ordem_y = list(reversed(_df_sg["grupo"].tolist()))   # menor densidade no topo
                fig_sg = go_plt.Figure()
                for _c in _ORDEM_CLASS_D:            # AS -> R, do pior para o melhor
                    _pcts, _txts = [], []
                    for _g in _ordem_y:
                        _r = _df_sg[_df_sg["grupo"] == _g].iloc[0]
                        _p = _r[_c] / _r["n"] * 100 if _r["n"] else 0
                        _pcts.append(_p)
                        _txts.append(f"{_p:.0f}%" if _p >= 7 else "")
                    fig_sg.add_trace(go_plt.Bar(
                        y=_ordem_y, x=_pcts, name=_c, orientation="h",
                        marker=dict(color=_COR_CLASS_D[_c],
                                    line=dict(color="#FFFFFF", width=1)),
                        text=_txts, textposition="inside", insidetextanchor="middle",
                        textfont=dict(size=12, color=_COR_TXT_CLASS_D[_c]),
                        hovertemplate=(f"<b>{_c}</b> — {_LABEL_CLASS_D[_c].split('— ')[1]}"
                                       "<br>%{y}: %{x:.1f}% das parcelas<extra></extra>")))
                # n de parcelas ao final de cada barra
                for _g in _ordem_y:
                    _r = _df_sg[_df_sg["grupo"] == _g].iloc[0]
                    fig_sg.add_annotation(x=100, y=_g, text=f"<b>{int(_r['n'])}</b> parcelas",
                                          showarrow=False, xanchor="left", xshift=8,
                                          font=dict(size=11, color="#6B7280"))
                fig_sg.update_layout(
                    barmode="stack", height=max(240, len(_ordem_y) * 62 + 90),
                    plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif"),
                    margin=dict(t=40, b=50, l=70, r=110), bargap=0.35,
                    legend=dict(orientation="h", y=1.14, x=0, xanchor="left",
                                font=dict(size=12, color="#1A1A1A"), traceorder="normal"),
                    xaxis=dict(title=dict(text="<b>% das parcelas avaliadas</b>",
                                          font=dict(size=13, color="#1A1A1A", weight="bold")),
                               range=[0, 100], ticksuffix="%", showgrid=False, zeroline=False,
                               tickfont=dict(size=11, color="#1A1A1A")),
                    yaxis=dict(title=dict(text="<b>Grupo de densidade</b>",
                                          font=dict(size=13, color="#1A1A1A", weight="bold")),
                               tickfont=dict(size=14, color="#1A1A1A", weight="bold"),
                               showgrid=False, zeroline=False))
                st.plotly_chart(fig_sg, use_container_width=True)

                # leitura automática: a fatia suscetível (AS+S) cresce com a densidade?
                _sus = {r["grupo"]: (r["AS"] + r["S"]) / r["n"] * 100 if r["n"] else 0
                        for _, r in _df_sg.iterrows()}
                _gs = _df_sg["grupo"].tolist()
                if len(_gs) >= 2:
                    _g_ref_s = st.selectbox(
                        "Comparar contra", options=_gs, index=0, key="den_sang_ref",
                        help="A leitura abaixo compara esta população com a maior testada.")
                    _gs = [_g_ref_s] + [g for g in _df_sg["grupo"].tolist() if g != _g_ref_s]
                    _delta_sus = _sus[_gs[-1]] - _sus[_gs[0]]
                    if _delta_sus > 5:
                        st.warning(f"A fatia de parcelas **suscetíveis (AS + S)** sobe de "
                                   f"{_sus[_gs[0]]:.0f}% em {_gs[0]} para {_sus[_gs[-1]]:.0f}% em "
                                   f"{_gs[-1]} — sinal de que adensar piora **{_d_graf}** neste "
                                   f"recorte. Confira no Ato 2 se o efeito é de todos os híbridos "
                                   f"ou de alguns.")
                    elif _delta_sus < -5:
                        st.info(f"A fatia suscetível **cai** de {_sus[_gs[0]]:.0f}% em {_gs[0]} "
                                f"para {_sus[_gs[-1]]:.0f}% em {_gs[-1]}. Densidade alta piorando "
                                f"a doença é o esperado, então o contrário merece checagem: pode "
                                f"ser composição de locais, não efeito de população.")
                    else:
                        st.info(f"A fatia suscetível praticamente não muda entre {_gs[0]} e "
                                f"{_gs[-1]} ({_sus[_gs[0]]:.0f}% → {_sus[_gs[-1]]:.0f}%): neste "
                                f"recorte, adensar não alterou a sanidade para **{_d_graf}**.")

                # exportação: a composição que está na tela, em contagem E em percentual
                _exp_a1 = []
                for _, _r in _df_sg.iterrows():
                    _l = {"Doença": _d_graf, "Densidade": _r["grupo"],
                          "Parcelas avaliadas": int(_r["n"])}
                    for _c in _ORDEM_CLASS_D:
                        _l[f"{_c} (n)"] = int(_r[_c])
                    for _c in _ORDEM_CLASS_D:
                        _l[f"{_c} (%)"] = round(_r[_c] / _r["n"] * 100, 1) if _r["n"] else None
                    _l["Suscetíveis AS+S (%)"] = (
                        round((_r["AS"] + _r["S"]) / _r["n"] * 100, 1) if _r["n"] else None)
                    _exp_a1.append(_l)
                exportar_excel(pd.DataFrame(_exp_a1), f"densidade_sanidade_composicao.xlsx",
                               "⬇️ Exportar composição por classe", key="den_sang_xlsx1")

                st.caption("Cada barra soma 100% das parcelas **avaliadas** daquela densidade "
                           "(nota maior que zero) · a conta aqui é por **parcela**, diferente da "
                           "coluna % da tabela, que é por **local** · o número à direita é quantas "
                           "parcelas sustentam a barra.")

            # ── Ato 2 — mapa híbrido × densidade ─────────────────────────────
            st.markdown(
                '<p style="font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                'letter-spacing:0.07em;margin:1.6rem 0 6px;">'
                'Ato 2 — Quem piora quando adensa</p>',
                unsafe_allow_html=True)
            st.markdown("Cada célula é a **classe da nota típica** daquele híbrido naquela "
                        "densidade. Percorra a linha da esquerda para a direita: se ela esquenta, "
                        "aquele material perde sanidade quando o estande fecha.")

            with st.popover("ℹ️ Como este mapa é calculado", use_container_width=False):
                st.markdown(f"""
**Como é calculado, passo a passo**

1. Cada quadrado é o cruzamento de **um híbrido** com **um grupo de densidade**, para a doença
   **{_d_graf}**.
2. Pego as parcelas desse cruzamento e descarto as de nota 0 (não avaliado).
3. Calculo a **moda** das notas restantes — a que mais se repetiu. Em caso de **empate, fica a
   menor** nota, ou seja, o lado suscetível. É a régua conservadora: entre dizer que o material
   é tolerante e dizer que é suscetível, o mapa fica com a hipótese pior.
4. Converto essa nota em classe e pinto o quadrado. Dentro fica só a **sigla da classe**; a nota
   que a gerou aparece ao passar o mouse.
5. Quadrado **em branco** = a doença não foi avaliada naquele híbrido nessa densidade.

**Este ato usa a MESMA régua da tabela da Seção 7**, então o mapa e a tabela sempre concordam.
O Ato 1 usa outra (composição por parcela) porque responde outra pergunta.

**Como interpretar**

- A leitura é **por linha, da esquerda para a direita**: densidade crescente. Linha que esquenta,
  do verde para o amarelo ou vermelho, é o material perdendo sanidade quando adensa.
- Linha de **cor constante** é material estável em relação à densidade — o que é uma informação
  boa e comercializável.
- Compare **linhas entre si**: dois híbridos podem partir da mesma classe em densidade baixa e se
  separar na alta. Esse é o achado que o número agregado não entrega.
- Este ato existe porque o Ato 1 pode não mostrar efeito nenhum quando um híbrido piora e outro
  melhora — no agregado os dois se cancelam.

**Três armadilhas**

**Classe é degrau, não régua contínua.** Notas 6 e 7 estão a um ponto de distância mas caem em
classes diferentes (MT e T), enquanto 7 e 8 são a mesma classe. Uma mudança de cor entre dois
quadrados pode ser uma variação pequena que atravessou a fronteira, e dois quadrados da mesma cor
podem esconder um ponto de diferença. Confira a nota no hover antes de concluir.

**Poucas parcelas por célula.** As parcelas do ensaio se dividem entre híbrido e densidade, então
cada quadrado costuma ter poucas. Com três parcelas, uma única avaliação ruim vira a moda e muda
a cor. Passe o mouse para ver quantas parcelas sustentam cada quadrado antes de tratar uma
mudança como real.

**A moda do total não é o meio das partes.** Se você comparar este mapa com o bloco "Todas as
densidades" da tabela, o total pode não parecer o resumo dos quadrados — moda não tem essa
propriedade. O bloco geral é a moda de todas as parcelas juntas, e ele pesa mais os grupos que
têm mais parcelas.
""")

            _hm_hibs = sorted(_val_sg["dePara"].dropna().unique().tolist())
            _hm_g = [g for g in _grupos_encontrados
                     if (_val_sg["pop_grupo"] == g).any()]
            _idx_cls = {c: i for i, c in enumerate(_ORDEM_CLASS_D)}
            _z, _txt_hm, _hover = [], [], []
            for _h in reversed(_hm_hibs):
                _lz, _lt, _lh = [], [], []
                for _g in _hm_g:
                    _sub = _val_sg[(_val_sg["dePara"] == _h) & (_val_sg["pop_grupo"] == _g)]
                    _nota, _inc, _cls = _resumo_doenca_den(_sub, _col_dg)
                    _lz.append(_idx_cls.get(_cls, np.nan) if _cls else np.nan)
                    _lt.append("" if _cls is None else _cls)
                    _lh.append("não avaliado" if _cls is None else
                               f"{_h} · {_g}<br>nota {_nota:.0f} ({_cls})<br>"
                               f"{len(_sub)} parcelas")
                _z.append(_lz); _txt_hm.append(_lt); _hover.append(_lh)

            # escala discreta: 5 degraus, um por classe, na ordem AS -> R
            _n_cl = len(_ORDEM_CLASS_D)
            _escala = []
            for _i, _c in enumerate(_ORDEM_CLASS_D):
                _escala.append([_i / _n_cl, _COR_CLASS_D[_c]])
                _escala.append([(_i + 1) / _n_cl, _COR_CLASS_D[_c]])

            fig_hm = go_plt.Figure(go_plt.Heatmap(
                z=_z, x=_hm_g, y=list(reversed(_hm_hibs)),
                customdata=_hover, hovertemplate="%{customdata}<extra></extra>",
                colorscale=_escala, zmin=-0.5, zmax=_n_cl - 0.5, showscale=False,
                xgap=3, ygap=3))

            # a sigla entra como ANOTAÇÃO, não como texttemplate: o Heatmap só aceita UMA cor
            # de fonte para o traço inteiro, e branco sumiria no amarelo do MT e no verde claro
            # do T. Cada célula precisa da cor de texto da sua própria classe.
            for _iy, _h in enumerate(reversed(_hm_hibs)):
                for _ix, _g in enumerate(_hm_g):
                    _cls_cel = (_ORDEM_CLASS_D[int(_z[_iy][_ix])]
                                if pd.notna(_z[_iy][_ix]) else None)
                    if _cls_cel is None:
                        continue
                    fig_hm.add_annotation(
                        x=_g, y=_h, text=f"<b>{_cls_cel}</b>", showarrow=False,
                        font=dict(size=15, color=_COR_TXT_CLASS_D[_cls_cel]))
            fig_hm.update_layout(
                height=max(220, len(_hm_hibs) * 46 + 90),
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Helvetica Neue, sans-serif"),
                margin=dict(t=20, b=55, l=110, r=30),
                xaxis=dict(title=dict(text="<b>Grupo de densidade</b>",
                                      font=dict(size=13, color="#1A1A1A", weight="bold")),
                           tickfont=dict(size=13, color="#1A1A1A", weight="bold"),
                           side="bottom", showgrid=False),
                yaxis=dict(tickfont=dict(size=13, color="#1A1A1A", weight="bold"),
                           showgrid=False))
            st.plotly_chart(fig_hm, use_container_width=True)

            _leg2 = "".join(
                f'<span style="display:inline-block;margin:0 10px 6px 0;padding:3px 9px;'
                f'border-radius:3px;font-size:12px;font-weight:700;'
                f'background:{_COR_CLASS_D[c]};color:{_COR_TXT_CLASS_D[c]};">'
                f'{_LABEL_CLASS_D[c]}</span>' for c in _ORDEM_CLASS_D)
            st.markdown(_leg2, unsafe_allow_html=True)
            # exportação: o mapa em formato largo, uma linha por híbrido e uma coluna por
            # densidade, com nota e classe juntas — é o que se vê, com a nota que o hover mostra
            _exp_a2 = []
            for _h in _hm_hibs:
                _l = {"Híbrido": _h, "Doença": _d_graf}
                for _g in _hm_g:
                    _sub = _val_sg[(_val_sg["dePara"] == _h) & (_val_sg["pop_grupo"] == _g)]
                    _nota, _inc, _cls = _resumo_doenca_den(_sub, _col_dg)
                    _l[f"{_g} — classe"] = _cls
                    _l[f"{_g} — nota"] = _nota
                    _l[f"{_g} — parcelas"] = len(_sub)
                _exp_a2.append(_l)
            exportar_excel(pd.DataFrame(_exp_a2), f"densidade_sanidade_mapa.xlsx",
                           "⬇️ Exportar mapa híbrido × densidade", key="den_sang_xlsx2")

            st.caption("Célula em branco = a doença não foi avaliada naquele híbrido e densidade "
                       "· a classe vem da nota típica (moda das parcelas, empate pela pior), a "
                       "mesma régua da tabela · **passe o mouse** para ver a nota exata e quantas "
                       "parcelas sustentam cada quadrado — com poucas parcelas, uma única "
                       "avaliação ruim muda a classe.")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9 — PERDAS  |  SEÇÃO 10 — FENÔMENOS E QUALIDADE DO GRÃO
# ══════════════════════════════════════════════════════════════════════════════
# DUAS tabelas, não uma: as famílias usam RÉGUAS DIFERENTES e não se comparam entre si.
#   Seção 9  — Perdas:    média de TODAS as parcelas avaliadas, o zero ENTRA (taxa na rede).
#   Seção 10 — Fenômenos e Qualidade: média SÓ ONDE OCORREU (>0). Nas duas o zero domina
#              (green snap é zero em ~90% das parcelas; ardidos em 96%), e incluí-lo achataria
#              a coluna e esconderia quem tem o problema.
# Cada tabela reúne só famílias que compartilham a mesma régua, então dentro de uma tabela as
# colunas SÃO comparáveis. Foi por isso que separamos.
#
# Direção da escala: aqui é % de plantas — quanto MAIOR, PIOR. É o inverso da tabela de doenças.

_PERDAS_DEN = {"Acamamento": "pct_acamadas", "Quebramento": "pct_quebradas",
               "Dominadas": "pct_dominadas", "Colmo Podre": "pct_colmo_podre",
               "Perda Total": "pct_perda_total"}
_QUALIDADE_DEN = {"Ardidos": "graos_ardidos_pct"}
_FENOMENOS_DEN = {"Green snap": "pct_green_snap", "Morte prematura": "pct_morte_prematura",
                  "Má formação": "pct_ma_formacao_espigas", "Enfezamento": "pct_enfezamento"}

# Fundo do cabeçalho de família: tons do MESMO azul dos gráficos de composição, variando só a
# intensidade. Cores diferentes por família (lilás, bege) sugeriam grandezas diferentes — todas
# medem % de plantas, e o que as separa é a régua, escrita logo abaixo do nome.
_COR_FAMILIA = {"Perdas na colheita": "#DCE8F3", "Fenômenos": "#C4D6E6",
                "Qualidade do grão": "#ABC3D8"}


def _familias_disponiveis(defs):
    """Mantém só as famílias cujas colunas existem no recorte."""
    out = [(f, r, {k: v for k, v in d.items() if v in ta_filtrado.columns}, z)
           for f, r, d, z in defs]
    return [t for t in out if t[2]]


def _agg_perda_den(serie, zero_entra):
    """Devolve (valor, n_parcelas_que_sustentam, estado).

    Três estados distintos, como na página de Perdas:
      - "ocorreu"        → houve valor > 0; o n é o das parcelas COM ocorrência quando a régua
                           descarta o zero, e o de todas as avaliadas quando o zero entra;
      - "sem_ocorrencia" → avaliado e sempre zero. Devolve 0.0, não vazio: é informação boa;
      - "nao_avaliado"   → nenhuma parcela com a métrica. Devolve None.
    Validado célula a célula contra a agregação da página de Perdas: 252 comparações, zero
    divergências. As duas páginas leem tabelas diferentes (Faixa x Densidade), então os valores
    absolutos diferem — o que precisa bater, e bate, é a régua.
    """
    s = pd.to_numeric(serie, errors="coerce").dropna()
    if s.empty:
        return None, 0, "nao_avaliado"
    if zero_entra:
        return round(float(s.mean()), 1), len(s), "ocorreu"
    com = s[s > 0]
    if com.empty:
        return 0.0, len(s), "sem_ocorrencia"
    # o n aqui é o das parcelas com ocorrência — é sobre elas que a média foi feita
    return round(float(com.mean()), 1), len(com), "ocorreu"


def _tabela_familias(titulo, subtitulo, familias, pref, texto_popover, padrao_n=3):
    """Renderiza uma tabela densidade > família > variável. Usada pelas Seções 9 e 10."""
    secao_titulo(titulo, subtitulo, contexto_str)

    if not familias or not _grupos_encontrados or "dePara" not in ta_filtrado.columns:
        st.info("Sem colunas para esta tabela ou sem grupos de densidade nos filtros ativos.")
    else:
        with st.popover("ℹ️ Como ler a tabela", use_container_width=False):
            st.markdown(texto_popover, unsafe_allow_html=True)

        _c_p1, _c_p2, _c_p3 = st.columns([3, 1.4, 1.4])
        with _c_p1:
            _todas_vars = [(f"{v} · {fam.split(' ')[0]}", fam, v)
                           for fam, _, d, _ in familias for v in d]
            _pad_p = [r for r, _, _ in _todas_vars][:padrao_n]
            _vars_p_sel = st.multiselect(
                "Variáveis", options=[r for r, _, _ in _todas_vars], default=_pad_p,
                key=f"{pref}_vars",
                help="Cada variável abre uma coluna dentro de cada grupo de densidade. As colunas "
                     "aparecem na ordem em que você escolher, agrupadas por família.")
        with _c_p2:
            _mostrar_n_p = st.checkbox("Mostrar n", value=False, key=f"{pref}_n")
        with _c_p3:
            _geral_p = st.checkbox("Bloco geral", value=True, key=f"{pref}_geral")

        _N_MIN_PERD = 5

        if not _vars_p_sel:
            st.info("Selecione ao menos uma variável.")
        else:
            # agrupa a seleção por família, mantendo a ordem de escolha dentro de cada uma
            _mapa_v = {r: (fam, v) for r, fam, v in _todas_vars}
            _sel_por_fam = {}
            for _r in _vars_p_sel:
                _fam, _v = _mapa_v[_r]
                _sel_por_fam.setdefault(_fam, []).append(_v)
            _fams = [(f, regua, [v for v in _sel_por_fam.get(f, [])], d, z)
                     for f, regua, d, z in familias if _sel_por_fam.get(f)]

            _blocos_p = ([("Todas as densidades", None)] if _geral_p else []) + \
                        [(g, g) for g in _grupos_encontrados]
            _hibs_p = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
            _status_p = (ta_filtrado.dropna(subset=["dePara"])
                         .groupby("dePara")["status_material"].first().to_dict()
                         if "status_material" in ta_filtrado.columns else {})

            _cel_p = {}
            for _nome_bl, _grp in _blocos_p:
                _dbl = ta_filtrado if _grp is None else ta_filtrado[ta_filtrado["pop_grupo"] == _grp]
                for _h in _hibs_p:
                    _dh = _dbl[_dbl["dePara"] == _h]
                    for _fam, _regua, _vs, _dic, _zero in _fams:
                        for _v in _vs:
                            _cel_p[(_h, _nome_bl, _fam, _v)] = _agg_perda_den(_dh[_dic[_v]], _zero)

            # aviso de concentração: quando a ocorrência de uma variável "só onde ocorreu" está
            # presa a poucos locais, a coluna vira retrato daquele ambiente, não da densidade
            _avisos = []
            for _fam, _regua, _vs, _dic, _zero in _fams:
                if _zero:
                    continue
                for _v in _vs:
                    _s = pd.to_numeric(ta_filtrado[_dic[_v]], errors="coerce")
                    _oc = ta_filtrado[_s > 0]
                    if _oc.empty:
                        continue
                    _n_loc = _oc["cod_fazenda"].nunique() if "cod_fazenda" in _oc else 0
                    _tot_loc = ta_filtrado["cod_fazenda"].nunique() if "cod_fazenda" in ta_filtrado else 0
                    if _n_loc and _n_loc <= 3:
                        _avisos.append(f"**{_v}** ocorreu em {_n_loc} de {_tot_loc} locais "
                                       f"({len(_oc)} parcelas)")

            # rampa de densidade (a mesma das outras seções)
            def _rampa_den_p(n):
                ini, fim = (214, 228, 242), (18, 57, 92)
                out = []
                for i in range(max(n, 1)):
                    t = i / (n - 1) if n > 1 else 0.0
                    rgb = tuple(int(round(ini[k] + (fim[k] - ini[k]) * t)) for k in range(3))
                    lum = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
                    out.append((f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                                "#1A1A1A" if lum > 150 else "#FFFFFF"))
                return out
            _rampa_p = _rampa_den_p(len(_grupos_encontrados))
            _cor_bloco_p, _ip = {}, 0
            for _nome_bl, _grp in _blocos_p:
                if _grp is None:
                    _cor_bloco_p[_nome_bl] = ("#3A3A3A", "#FFFFFF")
                else:
                    _cor_bloco_p[_nome_bl] = _rampa_p[min(_ip, len(_rampa_p) - 1)]
                    _ip += 1

            _COR_ST_P = {"CHECK": "#F4B184", "STINE": "#2976B6", "EXP": "#00FF00", "DP2": "#C4DFB4"}
            _COR_TX_P = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                         "DP2": "#1A1A1A", "": "#000000"}
            _n_sub_p = sum(len(vs) for _, _, vs, _, _ in _fams) + (1 if _mostrar_n_p else 0)

            # referência para a cor do texto: média da coluna (mesma lógica da Seção 5, invertida —
            # aqui MENOS é melhor, então acima da média fica vermelho)
            _med_col_p = {}
            for _nome_bl, _ in _blocos_p:
                for _fam, _, _vs, _, _ in _fams:
                    for _v in _vs:
                        _vals = [_cel_p[(_h, _nome_bl, _fam, _v)][0] for _h in _hibs_p
                                 if _cel_p[(_h, _nome_bl, _fam, _v)][0] is not None]
                        _med_col_p[(_nome_bl, _fam, _v)] = float(np.mean(_vals)) if _vals else np.nan

            _html_p = """
    <style>
    .tb-perd { width:100%; border-collapse:collapse; font-size:13px;
        font-family:'Helvetica Neue',sans-serif; }
    .tb-perd th { background:#F2F2F2; color:#000 !important; padding:6px 8px; text-align:center;
        border:1px solid #ccc; white-space:nowrap; font-weight:700; }
    .tb-perd th.grupo { font-size:14px; letter-spacing:0.03em; }
    .tb-perd th.fam { font-size:12px; }
    .tb-perd th.fam small { font-weight:500; color:#555; font-size:10px; }
    .tb-perd th.sub { font-size:12px; font-weight:600; color:#444 !important; }
    .tb-perd th.hib { background:#4A4A4A; color:#FFF !important; text-align:left; }
    .tb-perd td { padding:5px 8px; border:1px solid #ddd; text-align:center; white-space:nowrap;
        font-weight:600; }
    .tb-perd td.hib { text-align:left; font-weight:600; }
    .tb-perd td.sep, .tb-perd th.sep { border-left:2px solid #999; }
    .tb-perd td.sepf, .tb-perd th.sepf { border-left:1px solid #aaa; }
    .tb-perd td.frag { opacity:0.55; }
    .tb-perd td.ncel { color:#6B7280 !important; font-size:12px; font-style:italic; font-weight:400; }
.tb-perd tr.rodape td { background:#D9D9D9 !important; font-weight:700;
    border-top:2px solid #888; color:#000 !important; }
    </style>
    <table class="tb-perd">
    <thead>
    <tr><th class="hib" rowspan="3">Híbrido</th>"""
            for _nome_bl, _ in _blocos_p:
                _bgh, _fgh = _cor_bloco_p[_nome_bl]
                _html_p += (f'<th class="grupo sep" style="background:{_bgh};color:{_fgh} !important;'
                            f'border-color:{_bgh};" colspan="{_n_sub_p}">{_nome_bl}</th>')
            _html_p += "</tr><tr>"
            for _nome_bl, _ in _blocos_p:
                _bgh, _ = _cor_bloco_p[_nome_bl]
                for _jf, (_fam, _regua, _vs, _, _) in enumerate(_fams):
                    _html_p += (f'<th class="fam {"sep" if _jf == 0 else "sepf"}" '
                                f'style="border-top:4px solid {_bgh};'
                                f'background:{_COR_FAMILIA.get(_fam, "#E8E8E8")};" '
                                f'colspan="{len(_vs)}">{_fam}<br><small>{_regua}</small></th>')
                if _mostrar_n_p:
                    _html_p += (f'<th class="sepf" rowspan="2" '
                                f'style="border-top:4px solid {_bgh};">n</th>')
            _html_p += "</tr><tr>"
            for _nome_bl, _ in _blocos_p:
                for _jf, (_fam, _, _vs, _, _) in enumerate(_fams):
                    for _k, _v in enumerate(_vs):
                        _cls_h = "sep" if (_jf == 0 and _k == 0) else ("sepf" if _k == 0 else "")
                        _html_p += f'<th class="sub {_cls_h}">{_v} (%)</th>'
            _html_p += "</tr></thead><tbody>"

            for _h in _hibs_p:
                _st = _status_p.get(_h, "")
                _html_p += (f'<tr><td class="hib" style="background:{_COR_ST_P.get(_st, "#FFF")};'
                            f'color:{_COR_TX_P.get(_st, "#000")} !important;">{_h}</td>')
                for _nome_bl, _grp_bl in _blocos_p:
                    _dbl_n = (ta_filtrado if _grp_bl is None
                              else ta_filtrado[ta_filtrado["pop_grupo"] == _grp_bl])
                    _n_parc = int((_dbl_n["dePara"] == _h).sum())
                    for _jf, (_fam, _, _vs, _, _) in enumerate(_fams):
                        for _k, _v in enumerate(_vs):
                            _val, _n_c, _estado = _cel_p[(_h, _nome_bl, _fam, _v)]
                            _frag = _estado == "ocorreu" and _n_c < _N_MIN_PERD
                            _ref = _med_col_p[(_nome_bl, _fam, _v)]
                            # MENOS é melhor: acima da média da coluna = vermelho
                            if _val is None or pd.isna(_ref) or _ref == 0 or _estado != "ocorreu":
                                _cor = "#1A1A1A" if _val is not None else "#9CA3AF"
                            else:
                                _d = (_val - _ref) / abs(_ref) * 100
                                _cor = ("#1A1A1A" if abs(_d) < 1 else
                                        ("#C0201E" if _d > 0 else "#1E7A34"))
                            _txt = ("—" if _val is None else f"{_val:.1f}")
                            if _frag:
                                _txt += "*"
                            _cls_td = " ".join(x for x in [
                                ("sep" if (_jf == 0 and _k == 0) else ("sepf" if _k == 0 else "")),
                                ("frag" if _frag else "")] if x)
                            _html_p += (f'<td class="{_cls_td}" style="color:{_cor} !important;">'
                                        f'{_txt}</td>')
                    if _mostrar_n_p:
                        _html_p += f'<td class="ncel sepf">{_n_parc}</td>'
                _html_p += "</tr>"

            # rodapé com a média de cada coluna: é exatamente a referência que as cores usam,
            # então deixá-la visível transforma "verde/vermelho" em algo conferível à mão
            _html_p += '<tr class="rodape"><td class="hib">Média Geral</td>'
            for _nome_bl, _grp_bl in _blocos_p:
                _dbl_n = (ta_filtrado if _grp_bl is None
                          else ta_filtrado[ta_filtrado["pop_grupo"] == _grp_bl])
                for _jf, (_fam, _, _vs, _, _) in enumerate(_fams):
                    for _k, _v in enumerate(_vs):
                        _ref = _med_col_p[(_nome_bl, _fam, _v)]
                        _cls_td = ("sep" if (_jf == 0 and _k == 0)
                                   else ("sepf" if _k == 0 else ""))
                        _html_p += (f'<td class="{_cls_td}">'
                                    f'{"—" if pd.isna(_ref) else f"{_ref:.1f}"}</td>')
                if _mostrar_n_p:
                    _html_p += f'<td class="ncel sepf">{len(_dbl_n)}</td>'
            _html_p += "</tr>"
            _html_p += "</tbody></table>"

            components.html(_html_p, height=min(140 + (len(_hibs_p) + 3) * 32, 700), scrolling=True)

            st.caption(
                f"Cor do número = desvio contra a média da coluna, e aqui **menos é melhor**: "
                f"vermelho acima da média, verde abaixo (zona morta de 1%) · **0** = avaliado e nunca "
                f"ocorreu · **travessão** = não avaliado · **\\*** = média apoiada em menos de "
                f"{_N_MIN_PERD} parcelas · cada família tem a régua escrita no cabeçalho, e colunas de "
                f"famílias diferentes não se comparam entre si.")

            # ── Exportação: espelho fiel da tabela na tela ─────────────────────
            # Três níveis de cabeçalho (densidade > família + régua > variável), a rampa de
            # densidade, o fundo de família, a cor do desvio contra a Média Geral, o asterisco
            # das células frágeis e a linha de Média Geral no rodapé.
            def _xlsx_familias():
                import io
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = titulo[:28]
                _hx = lambda c: str(c).replace("#", "").upper()
                thin = Side(style="thin", color="DDDDDD")
                grosso = Side(style="medium", color="999999")
                centro = Alignment(horizontal="center", vertical="center", wrap_text=True)

                ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
                _c1 = ws.cell(row=1, column=1, value="Híbrido")
                _c1.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
                _c1.fill = PatternFill("solid", start_color="4A4A4A")
                _c1.alignment = Alignment(horizontal="left", vertical="center")
                ws.column_dimensions["A"].width = 18

                _ci = 2
                for _nome_bl, _ in _blocos_p:
                    _bgh, _fgh = _cor_bloco_p[_nome_bl]
                    _ini_bl = _ci
                    for _fam, _regua, _vs, _, _ in _fams:
                        _ini_f = _ci
                        for _v in _vs:
                            _ch = ws.cell(row=3, column=_ci, value=f"{_v} (%)")
                            _ch.font = Font(bold=True, name="Arial", size=9, color="444444")
                            _ch.fill = PatternFill("solid", start_color="F2F2F2")
                            _ch.alignment = centro
                            _ch.border = Border(left=grosso if _ci == _ini_bl else thin,
                                                right=thin, top=thin, bottom=thin)
                            ws.column_dimensions[get_column_letter(_ci)].width = 13
                            _ci += 1
                        # nível 2 — família COM a régua, que é o que impede a comparação errada
                        ws.merge_cells(start_row=2, start_column=_ini_f,
                                       end_row=2, end_column=_ci - 1)
                        _cf = ws.cell(row=2, column=_ini_f, value=f"{_fam}\n({_regua})")
                        _cf.font = Font(bold=True, name="Arial", size=9, color="1A1A1A")
                        _cf.fill = PatternFill(
                            "solid", start_color=_hx(_COR_FAMILIA.get(_fam, "#E8E8E8")))
                        _cf.alignment = centro
                        _cf.border = Border(left=grosso if _ini_f == _ini_bl else thin,
                                            right=thin,
                                            top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                    if _mostrar_n_p:
                        ws.merge_cells(start_row=2, start_column=_ci, end_row=3, end_column=_ci)
                        _cn = ws.cell(row=2, column=_ci, value="n")
                        _cn.font = Font(bold=True, name="Arial", size=9, color="6B7280",
                                        italic=True)
                        _cn.fill = PatternFill("solid", start_color="F2F2F2")
                        _cn.alignment = centro
                        _cn.border = Border(left=thin, right=thin,
                                            top=Side(style="thick", color=_hx(_bgh)), bottom=thin)
                        ws.column_dimensions[get_column_letter(_ci)].width = 7
                        _ci += 1
                    ws.merge_cells(start_row=1, start_column=_ini_bl, end_row=1, end_column=_ci - 1)
                    _cb = ws.cell(row=1, column=_ini_bl, value=_nome_bl)
                    _cb.font = Font(bold=True, name="Arial", size=11, color=_hx(_fgh))
                    _cb.fill = PatternFill("solid", start_color=_hx(_bgh))
                    _cb.alignment = centro
                    _cb.border = Border(left=grosso, right=thin, top=thin, bottom=thin)
                ws.row_dimensions[1].height = 20
                ws.row_dimensions[2].height = 34
                ws.row_dimensions[3].height = 18

                _ri = 4
                for _h in _hibs_p:
                    _st = _status_p.get(_h, "")
                    _ch = ws.cell(row=_ri, column=1, value=_h)
                    _ch.font = Font(bold=True, name="Arial", size=10,
                                    color=_hx(_COR_TX_P.get(_st, "#000000")))
                    _ch.fill = PatternFill("solid",
                                           start_color=_hx(_COR_ST_P.get(_st, "#FFFFFF")))
                    _ch.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    _ci = 2
                    for _nome_bl, _grp_bl in _blocos_p:
                        _dbl_n = (ta_filtrado if _grp_bl is None
                                  else ta_filtrado[ta_filtrado["pop_grupo"] == _grp_bl])
                        _n_parc = int((_dbl_n["dePara"] == _h).sum())
                        for _jf, (_fam, _, _vs, _, _) in enumerate(_fams):
                            for _k, _v in enumerate(_vs):
                                _val, _n_c, _estado = _cel_p[(_h, _nome_bl, _fam, _v)]
                                _frag = _estado == "ocorreu" and _n_c < _N_MIN_PERD
                                _ref = _med_col_p[(_nome_bl, _fam, _v)]
                                if (_val is None or pd.isna(_ref) or _ref == 0
                                        or _estado != "ocorreu"):
                                    _cor = "1A1A1A"
                                else:
                                    _d = (_val - _ref) / abs(_ref) * 100
                                    _cor = ("1A1A1A" if abs(_d) < 1
                                            else ("C0201E" if _d > 0 else "1E7A34"))
                                _cx = ws.cell(row=_ri, column=_ci,
                                              value=(float(_val) if _val is not None else None))
                                _cx.number_format = '0.0"*"' if _frag else "0.0"
                                _cx.font = Font(name="Arial", size=10, color=_cor,
                                                bold=(_cor != "1A1A1A"), italic=_frag)
                                _cx.alignment = centro
                                _cx.border = Border(
                                    left=grosso if (_jf == 0 and _k == 0) else thin,
                                    right=thin, top=thin, bottom=thin)
                                _ci += 1
                        if _mostrar_n_p:
                            _cn = ws.cell(row=_ri, column=_ci, value=_n_parc)
                            _cn.font = Font(name="Arial", size=9, color="6B7280", italic=True)
                            _cn.alignment = centro
                            _cn.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                            _ci += 1
                    _ri += 1

                # rodapé: Média Geral — a referência das cores, igual à tela
                _cf = ws.cell(row=_ri, column=1, value="Média Geral")
                _cf.font = Font(bold=True, name="Arial", size=10, color="000000")
                _cf.fill = PatternFill("solid", start_color="D9D9D9")
                _cf.border = Border(left=thin, right=thin, top=grosso, bottom=thin)
                _ci = 2
                for _nome_bl, _grp_bl in _blocos_p:
                    _dbl_n = (ta_filtrado if _grp_bl is None
                              else ta_filtrado[ta_filtrado["pop_grupo"] == _grp_bl])
                    for _jf, (_fam, _, _vs, _, _) in enumerate(_fams):
                        for _k, _v in enumerate(_vs):
                            _ref = _med_col_p[(_nome_bl, _fam, _v)]
                            _cm = ws.cell(row=_ri, column=_ci,
                                          value=(round(float(_ref), 1) if pd.notna(_ref) else None))
                            _cm.number_format = "0.0"
                            _cm.font = Font(bold=True, name="Arial", size=10, color="000000")
                            _cm.fill = PatternFill("solid", start_color="D9D9D9")
                            _cm.alignment = centro
                            _cm.border = Border(
                                left=grosso if (_jf == 0 and _k == 0) else thin,
                                right=thin, top=grosso, bottom=thin)
                            _ci += 1
                    if _mostrar_n_p:
                        _cn = ws.cell(row=_ri, column=_ci, value=int(len(_dbl_n)))
                        _cn.font = Font(bold=True, name="Arial", size=9, color="000000",
                                        italic=True)
                        _cn.fill = PatternFill("solid", start_color="D9D9D9")
                        _cn.alignment = centro
                        _cn.border = Border(left=thin, right=thin, top=grosso, bottom=thin)
                        _ci += 1

                _ri += 2
                _notas = ["Escala: % de plantas — quanto MAIOR, PIOR (inverso da tabela de "
                          "doenças, onde 9 é o melhor)."]
                for _fam, _regua, _vs, _, _zero in _fams:
                    _notas.append(f"{_fam}: {_regua}. " +
                                  ("O zero é medição (parcela avaliada sem o problema) e entra "
                                   "na média." if _zero else
                                   "Só as parcelas com valor maior que zero entram na média; "
                                   "0 = avaliado e nunca ocorreu; vazio = não avaliado."))
                _notas += [
                    "Cor do número = desvio contra a Média Geral da coluna (linha cinza): "
                    "vermelho acima (pior), verde abaixo, preto a menos de 1%.",
                    f"* = média apoiada em menos de {_N_MIN_PERD} parcelas.",
                    "Colunas de famílias diferentes NÃO se comparam entre si.",
                    f"Filtros ativos: {contexto_str}",
                ]
                for _txt in _notas:
                    _cn = ws.cell(row=_ri, column=1, value=_txt)
                    _cn.font = Font(name="Arial", size=8, color="6B7280", italic=True)
                    _ri += 1

                ws.freeze_panes = "B4"
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            st.download_button(
                "⬇️ Exportar", data=_xlsx_familias(),
                file_name=f"densidade_{pref}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{pref}_xlsx")

            if _avisos:
                st.warning("**Ocorrência concentrada em poucos locais.** " + " · ".join(_avisos) +
                           ". Com a ocorrência presa a um ou dois ambientes, a coluna retrata aquele "
                           "local, não a resposta à densidade — os grupos de densidade não têm os "
                           "mesmos locais nas mesmas proporções. Use como achado de auditoria, não "
                           "como conclusão sobre população.")



# ── Gráfico de composição por densidade (barras horizontais) ──────────────────
# Cada barra é um grupo de densidade e cada faixa é uma variável, com o VALOR MÉDIO daquela
# variável naquele grupo — sem faixas de gravidade inventadas. Tudo respeita os filtros ativos
# e a régua de cada família, então os números batem com a tabela correspondente.
#
# As faixas usam UM matiz só, em degradê do claro ao escuro. Uma paleta com vermelho, laranja e
# verde faria o olho ler gravidade onde não há: acamamento não é "pior" que dominadas por ser
# vermelho, são coisas diferentes. Com um matiz só, a cor serve para separar as faixas e nada
# mais.
#
# O matiz é o MESMO nas duas seções: o azul da rampa de densidade. Trocar de cor entre Perdas e
# Fenômenos sugeriria uma diferença de natureza que não existe — as duas medem % de plantas. O
# lilás fica registrado para uso futuro, se alguma família precisar se distinguir.
_HUE_FAMILIA = {
    "padrao": ((220, 232, 243), (23, 62, 99)),     # azul claro  -> azul-marinho
    "lilas":  ((228, 220, 240), (66, 44, 110)),    # lilás claro -> roxo escuro
}


def _degrade(hue, n):
    """n cores do claro ao escuro dentro de um matiz, com a cor de texto de cada uma."""
    ini, fim = _HUE_FAMILIA.get(hue, _HUE_FAMILIA["padrao"])
    saida = []
    for i in range(max(n, 1)):
        t = i / (n - 1) if n > 1 else 0.0
        rgb = tuple(int(round(ini[k] + (fim[k] - ini[k]) * t)) for k in range(3))
        lum = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
        saida.append((f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}",
                      "#1A1A1A" if lum > 150 else "#FFFFFF"))
    return saida


def _grafico_composicao(titulo, subtitulo, variaveis, pref, empilhar, hue="padrao",
                        nota_extra=""):
    """Barras horizontais por grupo de densidade, uma faixa por variável.

    empilhar=True  → as faixas somam (perdas: mesma régua, mesma subamostra, somam na parcela);
    empilhar=False → faixas lado a lado (fenômenos e ardidos: cada média tem o SEU denominador,
                     porque a régua é "só onde ocorreu", então empilhar somaria coisas que não
                     se somam).
    """
    secao_titulo(titulo, subtitulo, contexto_str)

    if not variaveis or not _grupos_encontrados or "dePara" not in ta_filtrado.columns:
        st.info("Sem variáveis ou sem grupos de população nos filtros ativos.")
        return

    _cc1, _cc2 = st.columns([3, 1.6])
    with _cc1:
        _hibs_c = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
        _hib_c = st.multiselect("Híbridos", options=_hibs_c, default=_hibs_c,
                                key=f"{pref}_hib")
    with _cc2:
        _g_ref_c = st.selectbox("Comparar contra", options=_grupos_encontrados, index=0,
                                key=f"{pref}_ref",
                                help="A leitura abaixo do gráfico compara esta população com a "
                                     "maior. O padrão é a menor.")
    _base_c = ta_filtrado[ta_filtrado["dePara"].isin(_hib_c)] if _hib_c else ta_filtrado.iloc[0:0]
    if _base_c.empty:
        st.info("Selecione ao menos um híbrido.")
        return

    with st.popover("ℹ️ Como ler este gráfico", use_container_width=False):
        st.markdown(f"""
**Em uma frase:** cada barra é uma população e mostra o valor médio de cada variável ali.

**O seletor de híbridos** define quem entra na conta. As médias são de **todas as parcelas dos
híbridos selecionados**, não de um por vez — é uma visão do conjunto. Os filtros da barra lateral
valem aqui como em toda a página.

**Como cada faixa é calculada.** Para cada população, tomo as parcelas dos híbridos selecionados
e tiro a média daquela variável, com a régua da família dela — a mesma da tabela desta seção.
{"As quatro perdas usam a régua com o zero: a parcela avaliada e sem acamamento entra na conta, porque zero ali é medição." if empilhar else "Fenômenos e ardidos usam a régua só onde ocorreu: parcelas zeradas ficam de fora, senão as faixas afundariam perto de zero."}

**{"Por que as faixas somam" if empilhar else "Por que as faixas NÃO somam"}.**
{"As quatro perdas são contadas na mesma subamostra, sobre o mesmo estande, e se somam dentro da parcela — é o que a coluna Perda Total faz. Por isso a barra empilhada é legítima, e o número à direita é a soma das faixas. O nome da variável aparece dentro da faixa quando ela é larga o bastante, para você não precisar da legenda." if empilhar else "Cada média tem o SEU denominador, porque só entram as parcelas em que aquele problema apareceu. Green snap pode ser a média de três parcelas e ardidos a de duas, outras. Somar não significaria nada — por isso cada variável ganha o SEU painel, com o nome no topo e sem legenda para consultar. Os painéis compartilham a escala horizontal, então os comprimentos continuam comparáveis entre eles."}

**Sobre as cores.** As faixas usam um matiz só, em degradê do claro ao escuro. A cor serve
apenas para separar uma faixa da outra — **não indica gravidade**. Uma paleta com vermelho,
laranja e verde faria parecer que uma perda é pior que a outra por causa da tinta, quando são
apenas problemas diferentes. Quem diz a gravidade é o comprimento da faixa.

**Como interpretar**

- Leia **de cima para baixo**: a menor população está no topo, a maior embaixo. É a direção do
  adensamento.
- {"Barra ficando mais longa para baixo = adensar custou mais perda. A faixa que mais engorda diz o motivo." if empilhar else "Faixa ficando mais longa para baixo = aquele problema ficou mais severo onde apareceu. Cuidado: isso não diz que apareceu em mais lugares — a frequência não está neste gráfico."}
- **Barras parecidas** = neste recorte, a população não mudou o quadro.

**Duas armadilhas**

**O número à direita é quantas parcelas** têm aquela população, entre os híbridos escolhidos.
Barra apoiada em poucas parcelas oscila muito — leia o valor, mas não decida por ele.

**Os grupos não têm os mesmos locais nas mesmas proporções.** Se as populações altas ficaram
concentradas numa região de mais pressão, a barra cresce por causa da região, não da população.
{nota_extra}
""")

    _lin_c = []
    for _g in _grupos_encontrados:
        _dg = _base_c[_base_c["pop_grupo"] == _g]
        if _dg.empty:
            continue
        _row = {"grupo": _g, "n": len(_dg)}
        for _rot_c, _col_c, _zero in variaveis:
            _v, _n_c, _est = _agg_perda_den(_dg[_col_c], zero_entra=_zero)
            _row[_rot_c] = _v if _v is not None else 0.0
        _lin_c.append(_row)
    _df_c = pd.DataFrame(_lin_c)

    if _df_c.empty:
        st.info("Sem parcelas avaliadas por grupo de população.")
        return

    _ordem_y = list(reversed(_df_c["grupo"].tolist()))   # menor população no topo
    _cores = _degrade(hue, len(variaveis))
    _tot_c = _df_c[[v[0] for v in variaveis]].sum(axis=1)

    if empilhar:
        # ── empilhado: as faixas somam, então uma figura só. O NOME da variável entra
        # dentro da faixa quando ela é larga o bastante, o que dispensa a ida à legenda.
        _max_x = float(_tot_c.max())
        fig_c = go_plt.Figure()
        for _i, (_rot_c, _, _) in enumerate(variaveis):
            _vals = [float(_df_c.loc[_df_c["grupo"] == _g, _rot_c].iloc[0]) for _g in _ordem_y]
            _cor_f, _cor_txt = _cores[_i]
            _txts = []
            for _v in _vals:
                _larg = _v / _max_x if _max_x else 0
                if _larg >= 0.16:
                    _txts.append(f"{_rot_c}<br>{_v:.1f}")   # cabe o nome junto
                elif _v > 0:
                    _txts.append(f"{_v:.1f}")
                else:
                    _txts.append("")
            fig_c.add_trace(go_plt.Bar(
                y=_ordem_y, x=_vals, name=_rot_c, orientation="h",
                marker=dict(color=_cor_f, line=dict(color="#FFFFFF", width=1)),
                text=_txts, textposition="inside", insidetextanchor="middle",
                cliponaxis=False, textfont=dict(size=11, color=_cor_txt),
                hovertemplate=f"<b>{_rot_c}</b><br>%{{y}}: %{{x:.2f}}%<extra></extra>"))
        for _g in _ordem_y:
            _r = _df_c[_df_c["grupo"] == _g].iloc[0]
            fig_c.add_annotation(
                x=_max_x * 1.02, y=_g, showarrow=False, xanchor="left",
                text=f"<b>{sum(_r[v[0]] for v in variaveis):.1f}%</b>  ·  {int(_r['n'])} parcelas",
                font=dict(size=11, color="#6B7280"))
        fig_c.update_layout(
            barmode="stack", height=max(250, len(_ordem_y) * 58 + 100),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"),
            margin=dict(t=50, b=55, l=70, r=160), bargap=0.35,
            legend=dict(orientation="h", y=1.10, x=0, xanchor="left",
                        font=dict(size=12, color="#1A1A1A"), traceorder="normal"),
            xaxis=dict(title=dict(text="<b>% médio de plantas</b>",
                                  font=dict(size=13, color="#1A1A1A", weight="bold")),
                       range=[0, _max_x * 1.30], ticksuffix="%", showgrid=True,
                       gridcolor="#EEEEEE", zeroline=False,
                       tickfont=dict(size=11, color="#1A1A1A")),
            yaxis=dict(title=dict(text="<b>Grupo de população</b>",
                                  font=dict(size=13, color="#1A1A1A", weight="bold")),
                       tickfont=dict(size=14, color="#1A1A1A", weight="bold"),
                       showgrid=False, zeroline=False))
    else:
        # ── lado a lado: em vez de agrupar cinco barras parecidas numa figura só (o que
        # obriga a consultar a legenda a cada barra), um PAINEL POR VARIÁVEL. O nome vira
        # título do painel e a legenda deixa de existir. Mesma escala x nos painéis, para
        # que os comprimentos continuem comparáveis entre eles.
        from plotly.subplots import make_subplots as _mk
        _max_x = float(max(_df_c[v[0]].max() for v in variaveis))
        fig_c = _mk(rows=1, cols=len(variaveis), shared_yaxes=True,
                    subplot_titles=[v[0] for v in variaveis], horizontal_spacing=0.035)
        for _i, (_rot_c, _, _) in enumerate(variaveis):
            _vals = [float(_df_c.loc[_df_c["grupo"] == _g, _rot_c].iloc[0]) for _g in _ordem_y]
            _cor_f, _ = _cores[_i]
            fig_c.add_trace(go_plt.Bar(
                y=_ordem_y, x=_vals, orientation="h", showlegend=False,
                marker=dict(color=_cor_f, line=dict(color="#FFFFFF", width=1)),
                text=[f"{v:.1f}" if v > 0 else "" for v in _vals],
                textposition="outside", cliponaxis=False,
                textfont=dict(size=11, color="#1A1A1A"),
                hovertemplate=f"<b>{_rot_c}</b><br>%{{y}}: %{{x:.2f}}%<extra></extra>"),
                row=1, col=_i + 1)
        fig_c.update_xaxes(range=[0, _max_x * 1.35], ticksuffix="%", showgrid=True,
                           gridcolor="#EEEEEE", zeroline=False,
                           tickfont=dict(size=10, color="#1A1A1A"))
        fig_c.update_yaxes(tickfont=dict(size=13, color="#1A1A1A", weight="bold"),
                           showgrid=False, zeroline=False)
        fig_c.update_annotations(font=dict(size=13, color="#1A1A1A"))
        fig_c.update_layout(
            height=max(260, len(_ordem_y) * 46 + 130),
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Helvetica Neue, sans-serif"), showlegend=False,
            margin=dict(t=60, b=60, l=70, r=30), bargap=0.35)
        fig_c.update_layout(xaxis=dict(title=dict(
            text="<b>% médio de plantas</b>",
            font=dict(size=12, color="#1A1A1A", weight="bold"))))
        # o n de cada população vai no rótulo do eixo y, já que não há espaço à direita
        fig_c.update_yaxes(
            ticktext=[f"<b>{_g}</b><br><span style='font-size:10px;color:#6B7280'>"
                      f"{int(_df_c.loc[_df_c['grupo'] == _g, 'n'].iloc[0])} parcelas</span>"
                      for _g in _ordem_y],
            tickvals=_ordem_y, row=1, col=1)

    st.plotly_chart(fig_c, use_container_width=True)

    # leitura automática, comparando a menor com a maior população
    if len(_df_c) >= 2:
        _g0 = _g_ref_c if (_df_c["grupo"] == _g_ref_c).any() else _df_c["grupo"].iloc[0]
        _g1 = _df_c["grupo"].iloc[-1] if _g0 != _df_c["grupo"].iloc[-1] else _df_c["grupo"].iloc[0]
        _i0 = int(_df_c.index[_df_c["grupo"] == _g0][0])
        _i1 = int(_df_c.index[_df_c["grupo"] == _g1][0])
        if empilhar:
            _t0, _t1 = float(_tot_c.iloc[_i0]), float(_tot_c.iloc[_i1])
            _d = (_t1 - _t0) / _t0 * 100 if _t0 else np.nan
            _cresc = {v[0]: float(_df_c[v[0]].iloc[_i1] - _df_c[v[0]].iloc[_i0])
                      for v in variaveis}
            _quem = max(_cresc, key=_cresc.get)
            if pd.notna(_d) and _d > 10:
                st.warning(f"O total sobe **{_d:+.0f}%** de {_g0} para {_g1} "
                           f"({_t0:.1f}% → {_t1:.1f}%). Quem mais cresceu foi **{_quem}** "
                           f"(+{_cresc[_quem]:.2f} ponto percentual).")
            elif pd.notna(_d) and _d < -10:
                st.info(f"O total **cai** {abs(_d):.0f}% de {_g0} para {_g1} "
                        f"({_t0:.1f}% → {_t1:.1f}%). Vale checar se os grupos não têm "
                        f"composições de locais muito diferentes.")
            else:
                st.info(f"O total praticamente não muda de {_g0} para {_g1} "
                        f"({_t0:.1f}% → {_t1:.1f}%).")
        else:
            _mud = [(v[0], float(_df_c[v[0]].iloc[_i1] - _df_c[v[0]].iloc[_i0]))
                    for v in variaveis]
            _mud = [m for m in _mud if abs(m[1]) > 0.3]
            if _mud:
                _txt = " · ".join(f"**{r}** {d:+.1f} p.p." for r, d in
                                  sorted(_mud, key=lambda x: -abs(x[1])))
                st.info(f"De {_g0} para {_g1}: {_txt}. Lembre que cada faixa tem o seu "
                        f"denominador — isso é severidade onde ocorreu, não frequência.")
            else:
                st.info(f"Nenhuma variável muda de forma relevante entre {_g0} e {_g1}.")

    st.caption(
        ("Cada faixa é uma perda, com a média escrita dentro quando cabe · o número à direita é "
         "a soma das faixas e quantas parcelas sustentam a barra · todas as médias incluem o "
         "zero (régua da família Perdas)."
         if empilhar else
         "Um painel por variável, com o nome no topo — não há legenda para consultar · todos os "
         "painéis usam a mesma escala horizontal, então os comprimentos se comparam entre eles · "
         "os painéis ficam separados porque as médias têm denominadores diferentes e não somam · "
         "as médias usam só as parcelas em que o problema ocorreu (régua da família)."))

    _exp_c = _df_c.rename(columns={"grupo": "População", "n": "n parcelas"})
    if empilhar:
        _exp_c["Total"] = _tot_c.values
    exportar_excel(_exp_c, f"densidade_{pref}_composicao.xlsx",
                   "⬇️ Exportar", key=f"{pref}_xlsx")


_POPOVER_PERDAS = """
**Os controles acima da tabela**

- **Variáveis** — quais colunas entram, agrupadas por família. Elas aparecem **na ordem em que
  você escolher**; para reordenar, remova e selecione de novo.
- **Mostrar n** — quantas parcelas o híbrido tem naquela densidade.
- **Bloco geral** — liga ou desliga a coluna **Todas as densidades**, que junta tudo.

**A escala aqui é ao contrário da tabela de doenças: quanto MAIOR, PIOR.** Lá a nota vai de 1 a 9
e 9 é o melhor. Aqui a unidade é percentual de plantas.

**Como cada célula é calculada, passo a passo**

1. Pego as parcelas daquele híbrido naquela densidade. O grupo de densidade vem do K-Means sobre a
   população real contada, o mesmo da Seção 2.
2. Descarto as parcelas em que a métrica está ausente (não avaliada). O zero **não** é ausência.
3. Tiro a média simples do que sobrou, com uma casa decimal.

**A régua desta tabela: o zero ENTRA na média.** Cada célula é a média de **todas** as parcelas
avaliadas do híbrido naquela densidade, inclusive as que vieram zeradas. É assim porque perda é
uma **taxa média na rede**: a parcela avaliada e sem acamamento é medição, não ausência de dado.
Se o zero saísse, um híbrido que tombou em duas parcelas de trinta pareceria pior que outro que
tombou em vinte de trinta.

É a mesma régua da página de Perdas — validei célula a célula contra ela, com 252 comparações e
nenhuma divergência.

**Fenômenos e ardidos estão na tabela seguinte, de propósito.** Eles usam a régua oposta (só onde
ocorreu), e colunas com denominadores diferentes lado a lado convidam à comparação errada.

**Três estados possíveis numa célula:**

- **número** → a taxa média da perda na rede, para aquele híbrido e aquela densidade
- **0** → avaliado e nunca ocorreu
- **travessão** → não avaliado; nenhuma parcela tem essa métrica

**Perda Total** é a soma das quatro perdas dentro de cada parcela. Como cada perda tem seu próprio
conjunto de parcelas avaliadas, a média da Perda Total pode não ser a soma exata das médias das
colunas — é esperado, não é erro.

**As cores, e como conferir à mão.** A cor do número compara a célula com a **média da coluna em
que ela está** — e essa média está escrita na linha cinza **Média Geral**, no rodapé da tabela.
Não é comparação com meta, com testemunha nem com a safra passada.

- <span style="color:#C0201E;font-weight:700">vermelho</span> = **acima** da média da coluna, ou
  seja, pior, porque aqui mais perda é pior
- <span style="color:#1E7A34;font-weight:700">verde</span> = **abaixo** da média, portanto melhor
- **preto** = a menos de 1% da média; variação dessa ordem é ruído e não merece cor

O sinal é o **oposto** da tabela de componentes de produção: lá verde é acima da média, porque lá
mais é melhor.

A comparação é sempre **vertical**, dentro de uma coluna: aquele híbrido contra os outros híbridos,
na mesma variável e na mesma densidade. A cor **não** compara densidades — ler a linha na horizontal
continua valendo, mas ali você lê os números, não as cores. E ela não indica significância: vermelho
é "acima da média", não "significativamente pior".

**Leia na horizontal** para ver se adensar aumenta a perda naquele híbrido. É a leitura principal:
planta mais adensada estica, e colmo mais fino e alto tomba mais. Se existir o efeito, a linha
cresce da esquerda para a direita — e vale cruzar com a coluna Altura Planta da Seção 5.

**Sobre o n.** A coluna **n** mostra quantas parcelas o híbrido tem naquela densidade. O
**asterisco** marca as células cuja média foi feita com menos de 5 parcelas.
"""

_POPOVER_FENOMENOS = """
**Os controles acima da tabela**

- **Variáveis** — quais colunas entram, agrupadas por família. Elas aparecem **na ordem em que
  você escolher**; para reordenar, remova e selecione de novo.
- **Mostrar n** — quantas parcelas o híbrido tem naquela densidade.
- **Bloco geral** — liga ou desliga a coluna **Todas as densidades**, que junta tudo.

**A escala aqui é ao contrário da tabela de doenças: quanto MAIOR, PIOR.** A unidade é percentual
de plantas.

**Como cada célula é calculada, passo a passo**

1. Pego as parcelas daquele híbrido naquela densidade. O grupo de densidade vem do K-Means sobre a
   população real contada, o mesmo da Seção 2.
2. Descarto as parcelas em que a métrica está ausente (não avaliada).
3. Das que sobraram, fico **só com as maiores que zero** e tiro a média delas.
4. Se todas eram zero, a célula recebe `0` — avaliado e nunca ocorreu, diferente de vazio.

**A régua desta tabela: o zero NÃO entra na média.** Cada célula é a média apenas das parcelas em
que o problema **ocorreu** (valor maior que zero).

O motivo é que aqui o zero domina. Green snap é zero em cerca de 90% das parcelas, e ardidos em
96% delas nos dados de 25/26. Incluir o zero achataria a coluna inteira perto de zero e esconderia
justamente quem teve o problema: nos dados reais, ardidos dá **4,3%** com esta régua e **0,17%**
com a régua das perdas — vinte e cinco vezes menor.

**Por isso esta tabela é separada da de Perdas.** Lá o zero entra, aqui não. Um ardido de 4% e um
acamamento de 4% não querem dizer a mesma coisa, e colocá-los lado a lado convidaria à comparação
errada. Dentro desta tabela, as colunas compartilham a régua e **são** comparáveis entre si.

**O número não diz a frequência.** Um híbrido com green snap de 16% pode ter tido o problema em 4
de 35 locais — os 16% são a média onde ocorreu, não em toda a rede. Para saber em quantos locais
ocorreu, use a Auditoria.

**Três estados possíveis numa célula:**

- **número** → média das parcelas em que ocorreu
- **0** → avaliado e nunca ocorreu. É informação boa, não é vazio
- **travessão** → não avaliado

**Duas famílias aqui, com a mesma régua mas conceitos diferentes.** *Fenômenos* são falhas de
estande e de espiga contadas na av4; *Qualidade do grão* é o ardido, medido na av2. Ficam juntas
porque a régua é a mesma, e separadas em famílias porque a pergunta não é.

**Sobre o n, e aqui ele é crítico.** A média vem só das parcelas com ocorrência, que costumam ser
poucas. Nos dados de 25/26, quase toda média de ardidos por híbrido vem de **uma ou duas
parcelas** — e o asterisco marca isso. Sem ele, um valor de 14,7% apoiado em duas parcelas
pareceria tão sólido quanto um de trinta.

**As cores, e como conferir à mão.** A cor do número compara a célula com a **média da coluna em
que ela está** — e essa média está escrita na linha cinza **Média Geral**, no rodapé da tabela.
Não é comparação com meta, com testemunha nem com a safra passada.

- <span style="color:#C0201E;font-weight:700">vermelho</span> = **acima** da média da coluna, ou
  seja, pior, porque aqui mais perda é pior
- <span style="color:#1E7A34;font-weight:700">verde</span> = **abaixo** da média, portanto melhor
- **preto** = a menos de 1% da média; variação dessa ordem é ruído e não merece cor

O sinal é o **oposto** da tabela de componentes de produção: lá verde é acima da média, porque lá
mais é melhor.

A comparação é sempre **vertical**, dentro de uma coluna: aquele híbrido contra os outros híbridos,
na mesma variável e na mesma densidade. A cor **não** compara densidades — ler a linha na horizontal
continua valendo, mas ali você lê os números, não as cores. E ela não indica significância: vermelho
é "acima da média", não "significativamente pior".
"""

_tabela_familias(
    "Perdas",
    "As perdas de colheita de cada híbrido ao longo das populações",
    _familias_disponiveis([
        ("Perdas na colheita", "% médio na rede (o zero entra)", _PERDAS_DEN, True),
    ]),
    pref="den_perd", texto_popover=_POPOVER_PERDAS, padrao_n=3)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9b — LEITURA GRÁFICA DAS PERDAS POR DENSIDADE
# ══════════════════════════════════════════════════════════════════════════════
# Mesmo desenho da Seção 6, trocando só a lista de variáveis da linha: barras de
# produtividade no eixo da esquerda e a perda escolhida no da direita. Assim a página inteira
# usa UM formato gráfico, e quem aprendeu a ler um lê todos.
#
# A leitura aqui é direta: se a barra sobe e a linha da perda sobe junto, adensar rendeu mais
# grão apesar de custar mais tombo; se a barra cai enquanto a perda sobe, a perda é o motivo.
# Cada variável carrega a sua régua (perdas com o zero, fenômenos e ardidos só onde ocorreu),
# então os números batem com as tabelas das Seções 9 e 10.

_grafico_composicao(
    "Composição das Perdas",
    "Quanto se perde em cada população, e de que tipo",
    [(r, c, True) for r, c in {"Acamamento": "pct_acamadas",
                               "Quebramento": "pct_quebradas",
                               "Dominadas": "pct_dominadas",
                               "Colmo Podre": "pct_colmo_podre"}.items()
     if c in ta_filtrado.columns],
    pref="den_perdc", empilhar=True)

st.divider()

_grafico_barras_linha(
    "Perdas — Leitura Gráfica",
    "Produtividade e perdas lado a lado, ao longo das populações",
    _VARS_GRAF_PERDAS,
    pref="den_perdg", prefixo_padrao="Perda Total")

st.divider()

_tabela_familias(
    "Fenômenos e Qualidade do Grão",
    "Os fenômenos e o grão ardido de cada híbrido ao longo das populações",
    _familias_disponiveis([
        ("Fenômenos", "% médio onde ocorreu", _FENOMENOS_DEN, False),
        ("Qualidade do grão", "% médio onde ocorreu", _QUALIDADE_DEN, False),
    ]),
    pref="den_fen", texto_popover=_POPOVER_FENOMENOS, padrao_n=4)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 10b — LEITURA GRÁFICA DOS FENÔMENOS E DA QUALIDADE DO GRÃO
# ══════════════════════════════════════════════════════════════════════════════
# Mesmo desenho das Seções 6 e 9b. Aqui a régua da linha é "só onde ocorreu", herdada da
# tabela da Seção 10 — o número do gráfico é o mesmo da tabela, e a legenda diz isso.
#
# Duas advertências que valem mais nesta seção do que nas outras, porque aqui a ocorrência é
# rara: a linha pode saltar entre densidades por causa de UMA parcela, e um ponto zerado
# significa "avaliado e nunca ocorreu", não "sem dado". Leia sempre junto com a coluna n da
# tabela da Seção 10.

_grafico_composicao(
    "Fenômenos e Qualidade do Grão por População",
    "O valor médio de cada fenômeno em cada população",
    [(r, c, False) for r, c in {"Green snap": "pct_green_snap",
                                "Morte prematura": "pct_morte_prematura",
                                "Má formação": "pct_ma_formacao_espigas",
                                "Enfezamento": "pct_enfezamento",
                                "Ardidos": "graos_ardidos_pct"}.items()
     if c in ta_filtrado.columns],
    pref="den_fenc", empilhar=False,
    nota_extra="\n\n**Nesta seção a ocorrência é rara**, então uma faixa pode vir de uma ou "
               "duas parcelas. Confira a coluna n da tabela acima antes de tratar uma diferença "
               "como real.")

st.divider()

_grafico_barras_linha(
    "Fenômenos e Qualidade do Grão — Leitura Gráfica",
    "Produtividade e fenômenos lado a lado, ao longo das populações",
    _VARS_GRAF_FENOM,
    pref="den_feng", prefixo_padrao="Green snap")

st.divider()

rodape()
