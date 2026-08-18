"""
pages/3_H2H.py — Head-to-Head · JAUM DTC Milho

Confronto direto entre híbridos, calculado apenas nos locais onde ambos foram
avaliados simultaneamente.

Abas, na ordem em que aparecem:
  · Locais — caracterização do ambiente de cada local; monta o recorte
  · Tabela de Classificação — um Produto 1 contra todos os adversários
  · Análise por Local — par específico, cards + donut + mapa + barras
  · Desvios por Ambiente — reta de desvio vs. média do local, com b, R² e teste t
  · Todos os Materiais — a linha inteira de uma vez, em quatro leituras:
      placar consolidado · matriz material × adversário · por ambiente · local a local

Fonte: tabela_analitica_faixa das safras 2024/25 e 2025/26 (só Faixa; Densidade
tem página própria).
"""

import io

import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_multisafra
from utils.tabelas import (cel, cel_resumo, hdr, tabela_html, tabela_excel,
                           render_tabela as _render_tabela, legenda_cores, MIME_XLSX,
                           BG_RESUMO, FG_RESUMO)
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

st.set_page_config(
    page_title="H2H · JAUM DTC Milho",
    page_icon="⚔️",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

st.markdown("""
<style>
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] { color: #374151 !important; opacity: 1 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────────
STATUS_P1 = ["STINE", "EXP", "DP2"]     # pool de "Produto 1" (materiais Stine)
# PADRÃO de adversário: os comerciais. É a régua do racional de posicionamento — quem disputa
# mercado com a linha são os CHECK. Serve de valor inicial do filtro lateral e de reserva quando
# nada está marcado; NÃO é mais fixo nas abas agregadas, que passaram a seguir a seleção da
# lateral (antes o filtro existia, aparecia na tela e não tinha efeito nelas).
STATUS_ADVERSARIO = ("CHECK",)
EMPATE_MARGEM = 1.0                      # ± sc/ha: dentro disso é empate técnico
SAFRA_PADRAO = ("25/26", "2025/26")      # aceita os dois formatos de rótulo

COR_VITORIA = "#27AE60"
COR_EMPATE = "#FFFF00"
COR_DERROTA = "#FF0000"


def classificar_h2h(pct: float):
    """Classe e cor de fundo a partir do % de vitórias."""
    if pd.isna(pct):
        return "—", "#F3F4F6"
    if pct <= 45:
        return "Restrito", "#FF0000"
    if pct <= 55:
        return "Competitivo", "#FFFF00"
    if pct <= 75:
        return "Superior", "#87CEFF"
    return "Alta Performance", "#90EE90"


COR_STATUS_TITULO = {"CHECK": "#F4B184", "STINE": "#2976B6",
                     "EXP": "#00FF00", "DP2": "#C4DFB4"}


# ─────────────────────────────────────────────────────────────────────────────
# Helper Excel — mesma paleta da classificação
# ─────────────────────────────────────────────────────────────────────────────
def to_excel(df: pd.DataFrame, cols_pct=None) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "H2H"

    df = df.reset_index(drop=True)
    df = df.loc[:, ~df.columns.str.startswith("_")].copy()

    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        c.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="4A4A4A")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    ws.row_dimensions[1].height = 28

    COR_FUNDO_XL = {"Alta Performance": "90EE90", "Superior": "87CEFF",
                    "Competitivo": "FFFF00", "Restrito": "FF0000", "—": "F3F4F6"}
    COR_FONTE_XL = {"Alta Performance": "1A1A1A", "Superior": "1A1A1A",
                    "Competitivo": "1A1A1A", "Restrito": "FFFFFF", "—": "6B7280"}

    idx_classe = df.columns.tolist().index("Classe") if "Classe" in df.columns else None

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        classe_val = str(row_data[idx_classe]) if idx_classe is not None else "—"
        bg = COR_FUNDO_XL.get(classe_val, "FFFFFF")
        fg = COR_FONTE_XL.get(classe_val, "1A1A1A")
        for ci, val in enumerate(row_data, 1):
            try:
                val = None if (val is None or (isinstance(val, float) and np.isnan(val))) else val
            except (TypeError, ValueError):
                pass
            c = ws.cell(row=ri, column=ci, value=val)
            if idx_classe is not None and ci == idx_classe + 1:
                c.font = Font(name="Arial", size=10, color=fg, bold=True)
                c.fill = PatternFill("solid", start_color=bg)
            else:
                c.font = Font(name="Arial", size=10, color="1A1A1A")
                c.fill = PatternFill("solid", start_color="FFFFFF")
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            # percentual: valor continua numérico (dá para ordenar e pivotar), o
            # símbolo entra pelo formato — igual ao que aparece na tela
            if cols_pct and df.columns[ci - 1] in cols_pct and isinstance(val, (int, float)):
                c.number_format = '0.0"%"'
            c.border = brd

    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Helper AgGrid — coluna Classe colorida
# ─────────────────────────────────────────────────────────────────────────────
def ag_table_h2h(df: pd.DataFrame, height: int = 480, estilos_col=None, renderers_col=None):
    """Tabela AgGrid padrão do H2H.
    `estilos_col`: {coluna: JsCode} para cellStyle · `renderers_col`: {coluna: JsCode} para
    cellRenderer (permite desenhar barras dentro da célula)."""
    classe_style = JsCode("""
    function(params) {
        const v = params.value;
        if (v === 'Alta Performance') return {'backgroundColor':'#90EE90','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Superior')         return {'backgroundColor':'#87CEFF','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Competitivo')      return {'backgroundColor':'#FFFF00','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Restrito')         return {'backgroundColor':'#FF0000','color':'#FFFFFF','fontWeight':'700','textAlign':'center'};
        return {};
    }
    """)
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000",
                   "fontFamily": "Helvetica Neue, sans-serif"})
    if "Classe" in df.columns:
        gb.configure_column("Classe", cellStyle=classe_style, minWidth=140)
    for _c, _e in (estilos_col or {}).items():
        if _c in df.columns:
            gb.configure_column(_c, cellStyle=_e)
    for _c, _r in (renderers_col or {}).items():
        if _c in df.columns:
            gb.configure_column(_c, cellRenderer=_r, minWidth=170)
    gb.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal",
                              suppressMenuHide=True, suppressColumnVirtualisation=True,
                              suppressContextMenu=False, enableRangeSelection=True)
    go = gb.build()
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(df, gridOptions=go, height=height, update_mode=GridUpdateMode.NO_UPDATE,
           fit_columns_on_grid_load=False, columns_auto_size_mode=2,
           allow_unsafe_jscode=True, enable_enterprise_modules=True,
           custom_css={
               ".ag-header": {"background-color": "#4A4A4A !important"},
               ".ag-header-row": {"background-color": "#4A4A4A !important"},
               ".ag-header-cell": {"background-color": "#4A4A4A !important"},
               ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
               ".ag-header-cell-text": {"color": "#FFFFFF !important", "font-size": "13px !important",
                                        "font-weight": "700 !important"},
               ".ag-icon": {"color": "#FFFFFF !important", "opacity": "1 !important"},
               ".ag-header-icon": {"color": "#FFFFFF !important", "opacity": "1 !important"},
               ".ag-header-cell-menu-button": {"opacity": "1 !important", "visibility": "visible !important"},
               ".ag-cell": {"font-size": "13px !important"},
               ".ag-row": {"font-size": "13px !important"},
           },
           theme="streamlit", use_container_width=True)


# ─────────────────────────────────────────────────────────────────────────────
# Carregamento — analítica de Faixa das duas safras, já reconciliada pelo
# depara_mestre (sem ele, o mesmo híbrido teria nomes diferentes em cada safra)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def carregar_concat():
    d = carregar_multisafra()
    df = d.get("tabela_analitica_faixa")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    df = df.copy()
    if "produtividade_valida_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_valida_kg_ha"], errors="coerce")
        if "produtividade_kg_ha" in df.columns:
            df["kg_ha"] = df["kg_ha"].fillna(pd.to_numeric(df["produtividade_kg_ha"], errors="coerce"))
    elif "produtividade_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_kg_ha"], errors="coerce")
    if "produtividade_valida_sacas_ha" in df.columns:
        df["sc_ha"] = pd.to_numeric(df["produtividade_valida_sacas_ha"], errors="coerce")
    elif "kg_ha" in df.columns:
        df["sc_ha"] = (df["kg_ha"] / 60).round(1)
    return df


with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

page_header("Head-to-Head",
            "Compare híbridos diretamente nos locais em que ambos foram avaliados simultaneamente.",
            imagem="Data analysis-pana.png")


# ─────────────────────────────────────────────────────────────────────────────
# Sidebar — filtros encadeados
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>', unsafe_allow_html=True)

    if st.button("Limpar filtros", use_container_width=True):
        for key in list(st.session_state.keys()):
            if (key.startswith("h2h_") or key.startswith("__opts_h2h_")
                    or key.startswith("busca_h2h_")):
                del st.session_state[key]
        st.rerun()

    def _podar_keys(prefix, opcoes, molde):
        """Remove estado de checkboxes de opções que saíram da cascata —
        sem isso elas reaparecem marcadas e o filtro se reaplica sozinho."""
        antigas = st.session_state.get(f"__opts_{prefix}", [])
        atuais = set(map(str, opcoes))
        for o in antigas:
            if str(o) not in atuais:
                st.session_state.pop(molde(o), None)
        st.session_state[f"__opts_{prefix}"] = list(opcoes)

    def checkboxes(opcoes, default_all=True, defaults=None, prefix=""):
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_{o}")
        sel = []
        for o in opcoes:
            key = f"{prefix}_{o}"
            if key in st.session_state:
                marcado = st.checkbox(str(o), key=key)
            else:
                checked = (o in defaults) if defaults is not None else default_all
                marcado = st.checkbox(str(o), value=checked, key=key)
            if marcado:
                sel.append(o)
        return sel

    def filtro_busca(opcoes, prefix):
        """Busca textual + seleção persistente. A seleção é lida direto dos checkboxes
        (fonte única), inclusive dos itens ocultos pela busca."""
        if f"{prefix}_reset" not in st.session_state:
            st.session_state[f"{prefix}_reset"] = 0
        r = st.session_state[f"{prefix}_reset"]
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_chk_{r}_{o}")

        busca = st.text_input("Buscar", value="", key=f"busca_{prefix}",
                              placeholder="Digite para filtrar...")
        filtradas = ([c for c in opcoes if busca.strip().lower() in str(c).lower()]
                     if busca.strip() else opcoes)

        if st.button("Limpar seleção", key=f"{prefix}_limpar", use_container_width=True):
            for o in opcoes:
                st.session_state.pop(f"{prefix}_chk_{r}_{o}", None)
            st.session_state.pop(f"__opts_{prefix}", None)
            st.session_state[f"{prefix}_reset"] = r + 1
            st.rerun()

        for c in filtradas:
            st.checkbox(str(c), key=f"{prefix}_chk_{r}_{c}")

        sel = [o for o in opcoes if st.session_state.get(f"{prefix}_chk_{r}_{o}", False)]
        return sel or opcoes

    # 1. Safra — padrão safra atual
    with st.expander("Safra", expanded=True):
        safras_all = sorted(ta_raw["safra"].dropna().unique().tolist())
        safra_default = [s for s in safras_all if str(s) in SAFRA_PADRAO] or safras_all[-1:]
        if "h2h_safra_init" not in st.session_state:
            for o in safras_all:
                st.session_state[f"h2h_safra_{o}"] = (o in safra_default)
            st.session_state["h2h_safra_init"] = True
        safras_sel = checkboxes(safras_all, defaults=safra_default, prefix="h2h_safra")
    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("Região Macro", expanded=False):
        macros_sel = checkboxes(sorted(ta_f1["regiao_macro"].dropna().unique().tolist()), prefix="h2h_macro")
    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("Região Micro", expanded=False):
        micros_sel = checkboxes(sorted(ta_f2["regiao_micro"].dropna().unique().tolist()), prefix="h2h_micro")
    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("Estado", expanded=False):
        estados_sel = filtro_busca(sorted(ta_f3["estado_sigla"].dropna().unique().tolist()), "h2h_estado")
    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("Cidade", expanded=False):
        cidades_sel = filtro_busca(sorted(ta_f4["cidade_nome"].dropna().unique().tolist()), "h2h_cidade")
    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("Fazenda", expanded=False):
        fazendas_sel = filtro_busca(sorted(ta_f5["nomeFazenda"].dropna().unique().tolist()), "h2h_fazenda")
    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Responsável
    with st.expander("Responsável", expanded=False):
        resps_sel = filtro_busca(sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist()), "h2h_resp")
    ta_filtrado = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # 8. Status do adversário (Produto 2) — padrão CHECK
    with st.expander("Status do Adversário (Prod. 2)", expanded=True):
        status_p2_all = sorted(ta_filtrado["status_material"].dropna().unique().tolist())
        status_p2_default = [s for s in ["CHECK"] if s in status_p2_all]
        if "h2h_p2status_init" not in st.session_state and status_p2_all:
            for o in status_p2_all:
                st.session_state[f"h2h_p2status_{o}"] = (o in status_p2_default)
            st.session_state["h2h_p2status_init"] = True
        status_p2_sel = checkboxes(status_p2_all, defaults=status_p2_default, prefix="h2h_p2status")

if ta_filtrado.empty:
    st.warning("Nenhum dado para os filtros selecionados.")
    st.stop()


df_p1 = ta_filtrado[ta_filtrado["status_material"].isin(STATUS_P1)].copy()
df_p2 = (ta_filtrado[ta_filtrado["status_material"].isin(status_p2_sel)].copy()
         if status_p2_sel else pd.DataFrame())


# ─────────────────────────────────────────────────────────────────────────────
# Cruzamento por local (o join pesado — cacheado)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def cruzar_por_local(df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """Cruza df1 × df2 pelo local. Agrega por local antes do merge para não
    duplicar parcelas. Uma linha por (dePara_1, dePara_2, cod_fazenda)."""
    cols = ["dePara", "status_material", "cod_fazenda", "sc_ha", "kg_ha"]
    d1 = df1[[c for c in cols if c in df1.columns]].dropna(subset=["sc_ha"]).copy()
    d2 = df2[[c for c in cols if c in df2.columns]].dropna(subset=["sc_ha"]).copy()
    if d1.empty or d2.empty:
        return pd.DataFrame()
    d1 = d1.groupby(["dePara", "status_material", "cod_fazenda"], as_index=False)[["sc_ha", "kg_ha"]].mean()
    d2 = d2.groupby(["dePara", "status_material", "cod_fazenda"], as_index=False)[["sc_ha", "kg_ha"]].mean()
    merged = d1.merge(d2, on="cod_fazenda", suffixes=("_1", "_2"))
    return merged[merged["dePara_1"] != merged["dePara_2"]].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────────────────────
# Caracterização de locais — ambiente + placar da linha por local
# ─────────────────────────────────────────────────────────────────────────────
# Data de plantio. O nome canônico vem do pipeline: `dataPlantioMilho`, da tabela
# fazenda, propagado em CONTEXTO até as analíticas. Os demais nomes ficam como
# retrocompatibilidade, e a busca por conteúdo cobre o caso de o nome mudar.
COLS_PLANTIO = ["dataPlantioMilho", "data_plantio", "dataPlantio", "dt_plantio",
                "data_semeadura", "dataSemeadura", "plantio"]


def _parece_data(s: pd.Series) -> bool:
    """Metade dos valores não nulos precisa virar data plausível de safra."""
    v = s.dropna()
    if v.empty:
        return False
    d = pd.to_datetime(v, errors="coerce", dayfirst=True)
    if d.notna().mean() < 0.5:
        return False
    anos = d.dropna().dt.year
    return bool(len(anos)) and anos.between(2015, 2035).mean() > 0.8


def _col_plantio(df: pd.DataFrame):
    """1) nomes conhecidos · 2) qualquer coluna cujo nome cite plantio/semeadura e
    cujo conteúdo seja data · 3) desiste."""
    for c in COLS_PLANTIO:
        if c in df.columns and _parece_data(df[c]):
            return c
    for c in df.columns:
        n = str(c).lower()
        if ("plant" in n or "semea" in n or "semei" in n) and _parece_data(df[c]):
            return c
    return None


def candidatas_data(df: pd.DataFrame):
    """Colunas que parecem data — usado no diagnóstico quando nada é encontrado."""
    return [c for c in df.columns if _parece_data(df[c])]


@st.cache_data(show_spinner=False)
def caracterizar_locais(df_base: pd.DataFrame, status_linha: tuple, status_adv: tuple) -> pd.DataFrame:
    """Uma linha por local. Duas famílias de coluna:

    AMBIENTE (média, amplitude, nº de materiais) — calculado sobre a base INTEIRA,
    com todos os híbridos avaliados ali, inclusive os fora do filtro. Assim a
    caracterização do local não muda conforme a seleção, mesmo princípio da Tab 3.

    PLACAR (confrontos, vitórias, aproveitamento, diferença média) — todos os
    duelos linha × adversário naquele local, com a mesma margem de empate do resto
    da página. O adversário é o que a barra lateral tiver marcado.

    CONTRATO: uma linha por `cod_fazenda`, porque o chamador indexa por ele (`set_index`/`map`).
    O pareamento interno usa (local, safra) — ver `cruzar_linha` —, mas a saída volta agregada
    por local. Isso é seguro enquanto o código do local embutir o ano, o que é o caso hoje
    (74 locais nas duas safras, nenhum repetido). Se um código passar a repetir, o `set_index`
    do chamador acusa índice duplicado em vez de misturar em silêncio.
    """
    d = df_base.dropna(subset=["sc_ha"]).copy()
    if d.empty:
        return pd.DataFrame()

    # média por (material, local): mesma agregação de cruzar_por_local. A safra entra na chave
    # pelo mesmo motivo do cruzar_linha: parear só por local depende de o código embutir o ano.
    _chave = ["cod_fazenda", "safra"] if "safra" in d.columns else ["cod_fazenda"]
    med = d.groupby(["dePara", "status_material"] + _chave, as_index=False)["sc_ha"].mean()

    # --- ambiente ---
    amb = med.groupby("cod_fazenda").agg(
        _n_mat=("dePara", "nunique"),
        _melhor=("sc_ha", "max"),
        _pior=("sc_ha", "min"),
    )
    # média do local pela parcela (mesma definição do Panorama), não pela média das médias
    amb["_media"] = d.groupby("cod_fazenda")["sc_ha"].mean()
    amb["_amplitude"] = amb["_melhor"] - amb["_pior"]

    # --- placar da linha ---
    d1 = med[med["status_material"].isin(status_linha)]
    d2 = med[med["status_material"].isin(status_adv)]
    if d1.empty or d2.empty:
        duelos = pd.DataFrame(columns=["cod_fazenda", "_conf", "_vit", "_emp", "_der", "_dif"])
    else:
        m = d1.merge(d2, on=_chave, suffixes=("_1", "_2"))
        m = m[m["dePara_1"] != m["dePara_2"]].copy()
        m["_d"] = m["sc_ha_1"] - m["sc_ha_2"]
        duelos = m.groupby("cod_fazenda").apply(
            lambda g: pd.Series({
                "_conf": len(g),
                "_vit": int((g["_d"] > EMPATE_MARGEM).sum()),
                "_emp": int((g["_d"].abs() <= EMPATE_MARGEM).sum()),
                "_der": int((g["_d"] < -EMPATE_MARGEM).sum()),
                "_dif": g["_d"].mean(),
            }),
            include_groups=False,
        ).reset_index()

    out = amb.reset_index().merge(duelos, on="cod_fazenda", how="left")
    out["_aprov"] = np.where(out["_conf"].fillna(0) > 0, out["_vit"] / out["_conf"] * 100, np.nan)
    return out


def _rotulo_tercil(s: pd.Series) -> pd.Series:
    """Tercil calculado SOBRE O RECORTE ATIVO — responde 'difícil dentro do que
    está na tela', não 'difícil na rede'. Com menos de 6 locais não classifica:
    tercil de 4 valores é rótulo sem conteúdo."""
    if s.notna().sum() < 6:
        return pd.Series(["—"] * len(s), index=s.index)
    try:
        return pd.qcut(s, 3, labels=["Baixa", "Média", "Alta"]).astype(str)
    except ValueError:
        return pd.Series(["—"] * len(s), index=s.index)


@st.cache_data(show_spinner=False)
def cruzar_linha(df_base: pd.DataFrame, status_linha: tuple, status_adv: tuple) -> pd.DataFrame:
    """Todos os duelos linha × adversário, um por (material, adversário, local).

    Uma única passagem produz as quatro tabelas do panorama: basta filtrar os
    locais e agrupar de formas diferentes. Não depende de botão.

    A chave do pareamento é (local, SAFRA). Hoje o código do local já embute o ano — os 74
    locais das duas safras não têm um único código repetido —, então parear só por local
    daria o mesmo resultado. Mas isso é convenção de nome, não garantia: no dia em que um
    código se repetir, parear só por local cruzaria 24/25 com 25/26 sem avisar, e o confronto
    perderia o sentido (mesmo local, ano diferente, clima diferente).
    """
    d = df_base.dropna(subset=["sc_ha"])
    _chave = ["cod_fazenda", "safra"] if "safra" in d.columns else ["cod_fazenda"]
    med = d.groupby(["dePara", "status_material"] + _chave, as_index=False)["sc_ha"].mean()
    d1 = med[med["status_material"].isin(status_linha)]
    d2 = med[med["status_material"].isin(status_adv)]
    if d1.empty or d2.empty:
        return pd.DataFrame(columns=["material", "adversario", "cod_fazenda", "dif", "res"])
    m = d1.merge(d2, on=_chave, suffixes=("_1", "_2"))
    m = m[m["dePara_1"] != m["dePara_2"]].copy()
    m["dif"] = m["sc_ha_1"] - m["sc_ha_2"]
    m["res"] = np.select([m["dif"] > EMPATE_MARGEM, m["dif"] < -EMPATE_MARGEM],
                         ["V", "D"], default="E")
    return m.rename(columns={"dePara_1": "material", "dePara_2": "adversario"})[
        ["material", "adversario", "cod_fazenda", "dif", "res"]]


def agregar_placar(d: pd.DataFrame, por) -> pd.DataFrame:
    """V/E/D, confrontos, aproveitamento e diferença média por chave."""
    if d.empty:
        return pd.DataFrame()
    g = d.groupby(por)
    out = pd.DataFrame({
        "Confrontos": g.size(),
        "V": g["res"].apply(lambda s: int((s == "V").sum())),
        "E": g["res"].apply(lambda s: int((s == "E").sum())),
        "D": g["res"].apply(lambda s: int((s == "D").sum())),
        "Dif": g["dif"].mean(),
    })
    out["Aprov"] = out["V"] / out["Confrontos"] * 100
    return out.reset_index()


def linha_safra(s, g, sufixo_locais="locais"):
    """Uma linha de contexto por safra, com as regiões/estados que ela realmente contém."""
    partes = [f"<b>Safra:</b> {s}"]
    if "regiao_macro" in g.columns and g["regiao_macro"].notna().any():
        partes.append("Macro: " + ", ".join(sorted(g["regiao_macro"].dropna().unique())))
    if "regiao_micro" in g.columns and g["regiao_micro"].notna().any():
        partes.append("Micro: " + ", ".join(sorted(g["regiao_micro"].dropna().unique())))
    if "estado_sigla" in g.columns and g["estado_sigla"].notna().any():
        partes.append(", ".join(sorted(g["estado_sigla"].dropna().unique())))
    partes.append(f"{g['cidade_nome'].nunique()} cidades")
    partes.append(f"{g['cod_fazenda'].nunique()} {sufixo_locais}")
    return " · ".join(partes)


def montar_contexto(base, sufixo_locais="locais"):
    grupos = list(base.groupby("safra")) if "safra" in base.columns else []
    linhas = [linha_safra(s, g, sufixo_locais) for s, g in grupos]
    if len(grupos) > 1:
        linhas = ["<b>Análise multissafra</b>"] + linhas
    return "<br>".join(linhas)


MIME_XLSX = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")


def render_tabela(headers, linhas, nome_arquivo, key, **kw):
    """Wrapper: injeta o sufixo do recorte no nome do arquivo."""
    return _render_tabela(headers, linhas, nome_arquivo, key,
                          sufixo=sufixo_arquivo(), **kw)


def sufixo_arquivo() -> str:
    """Marca o recorte no nome do arquivo — sem isso, três exportações do mesmo
    painel viram três arquivos indistinguíveis na pasta de Downloads."""
    partes = []
    try:
        if 1 <= len(set(estados_sel)) <= 3:
            partes.append("-".join(sorted(set(estados_sel))))
    except (NameError, TypeError):
        pass
    try:
        if len(set(safras_sel)) == 1:
            partes.append(str(list(safras_sel)[0]).replace("/", ""))
    except (NameError, TypeError):
        pass
    return ("_" + "_".join(partes)) if partes else ""


def botao_exportar(df: pd.DataFrame, nome: str, key: str, label: str = "⬇️ Exportar Excel",
                   cols_pct=None):
    """Exporta a tabela como ela está na tela. Colunas técnicas (prefixo _) saem
    no to_excel."""
    if df is None or len(df) == 0:
        return
    # Int64/NA quebram o openpyxl ("Cannot convert <NA> to Excel"): vira None
    df = df.astype(object).where(df.notna(), None)
    st.download_button(label, data=to_excel(df, cols_pct=cols_pct),
                       file_name=f"{nome}{sufixo_arquivo()}.xlsx",
                       mime=MIME_XLSX, key=key)


tab1, tab2, tab3, tab0, tabL = st.tabs(
    ["Tabela de Classificação", "Análise por Local", "Desvios por Ambiente",
     "Locais", "Todos os Materiais"])


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 0 — Locais
# ═══════════════════════════════════════════════════════════════════════════════
with tab0:
    secao_titulo(
        "HEAD-TO-HEAD · LOCAIS",
        "Os locais do recorte, do mais produtivo ao menos produtivo",
        "Referência de ambiente para montar recortes. A cor destaca os extremos de "
        "produtividade e os plantios mais tardios.",
    )

    st.markdown(
        '<div style="background:#F7F7F7;border-left:5px solid #2976B6;border-radius:8px;'
        'padding:12px 18px;margin:2px 0 14px;">'
        '<p style="margin:0;font-size:15px;line-height:1.7;color:#1A1A1A;'
        'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        '<b>O fluxo:</b> &nbsp;<b>1.</b> filtre estado ou região na barra lateral &nbsp;·&nbsp; '
        '<b>2.</b> leia a tabela abaixo e identifique os locais que interessam — os extremos de '
        'produtividade e os plantios tardios já vêm destacados &nbsp;·&nbsp; '
        '<b>3.</b> marque as fazendas deles no filtro <b>Fazenda</b> da barra lateral &nbsp;·&nbsp; '
        '<b>4.</b> volte para qualquer aba: todas já estão respondendo só a esses locais.'
        '<br><span style="color:#6B7280;">Esta tabela não filtra nada — ela mostra o que o filtro '
        'lateral já selecionou. Quem estreita a análise é sempre a barra lateral, e ela vale para '
        'as cinco abas.</span>'
        '</p></div>', unsafe_allow_html=True)

    _cp = _col_plantio(ta_filtrado)
    _base_loc = ta_filtrado.dropna(subset=["sc_ha"])

    if _base_loc.empty:
        st.info("Nenhum local com os filtros atuais.")
    else:
        _cols = ["cod_fazenda", "nomeFazenda", "cidade_nome", "regiao_macro", "regiao_micro",
                 "estado_sigla", "safra"]
        if _cp:
            _cols.append(_cp)
        _meta = (ta_filtrado[[c for c in _cols if c in ta_filtrado.columns]]
                 .drop_duplicates("cod_fazenda"))
        _med = _base_loc.groupby("cod_fazenda")["sc_ha"].mean().round(1).rename("sc/ha")
        _t = _meta.merge(_med, on="cod_fazenda", how="left").sort_values("sc/ha", ascending=False)

        tab_loc = pd.DataFrame({
            "Código": _t["cod_fazenda"].values,
            "Fazenda": _t.get("nomeFazenda", pd.Series(dtype=object)).values,
            "Cidade": _t.get("cidade_nome", pd.Series(dtype=object)).values,
            "Macro": _t.get("regiao_macro", pd.Series(dtype=object)).values,
            "Micro": _t.get("regiao_micro", pd.Series(dtype=object)).values,
            "sc/ha": _t["sc/ha"].round(1).values,
        })
        if _cp:
            _dt = pd.to_datetime(_t[_cp], errors="coerce")
            tab_loc["Plantio"] = _dt.dt.strftime("%d/%m").values
            _dias = (_dt - _dt.min()).dt.days.values
        else:
            _dias = np.full(len(tab_loc), np.nan)

        mostrar_placar = st.toggle(
            "Mostrar placar da linha e amplitude", value=False, key="loc_toggle_placar",
            help="Acrescenta o quanto o local separa os materiais e o aproveitamento da "
                 "linha Stine contra os concorrentes comerciais em cada local.",
        )
        _linha_status = ("STINE",)
        # adversário da lateral também aqui: as colunas de placar deste toggle passam a
        # concordar com as abas seguintes em vez de ficarem presas em CHECK
        _adv_loc = tuple(sorted(status_p2_sel)) or STATUS_ADVERSARIO
        _car = caracterizar_locais(ta_raw, _linha_status,
                                   _adv_loc).set_index("cod_fazenda")
        if mostrar_placar:
            tab_loc["Amplitude"] = tab_loc["Código"].map(_car["_amplitude"]).round(1)
            tab_loc["Confrontos"] = tab_loc["Código"].map(_car["_conf"]).astype("Int64")
            tab_loc["Aprov. %"] = tab_loc["Código"].map(_car["_aprov"]).round(1)
            tab_loc["Dif. sc/ha"] = tab_loc["Código"].map(_car["_dif"]).round(1)

        # ── faixa de cor por quintil dentro do recorte ativo ─────────────────
        CORES_FAIXA = {"alto": "#1E7A34", "medio-alto": "#3E9E52", "neutro": "#1A1A1A",
                       "medio-baixo": "#E06C00", "baixo": "#C0201E"}

        def _faixa(v: pd.Series, invertido=False):
            if v.notna().sum() < 5:
                return pd.Series(["neutro"] * len(v), index=v.index)
            q = v.rank(pct=True, ascending=not invertido)
            return pd.Series(np.select(
                [q >= 0.85, q >= 0.70, q <= 0.10, q <= 0.25],
                ["alto", "medio-alto", "baixo", "medio-baixo"], default="neutro"
            ), index=v.index)

        _f_sc = _faixa(tab_loc["sc/ha"]).tolist()
        _f_pl = (_faixa(pd.Series(_dias), invertido=True).tolist()
                 if _cp else ["neutro"] * len(tab_loc))

        _cpa, _cpb = st.columns([3, 1])
        with _cpb:
            with st.popover("ℹ️ Como entender esta tabela", use_container_width=True):
                st.markdown("""
**📌 A pergunta que esta tabela responde**

> **Que ambientes compõem o recorte que estou olhando, e qual deles serve de base para
> cada afirmação?**

Ela não compara híbridos. Ela descreve os **locais** — para que marcar uma fazenda no filtro
lateral seja uma decisão consciente, e não escolha por nome.

> **Esta tabela não filtra nada.** Ela mostra os locais que o filtro lateral já selecionou. Quem
estreita a análise é a barra lateral, e ela vale para as cinco abas ao mesmo tempo: Tabela de
Classificação, Análise por Local, Desvios por Ambiente, esta e Todos os Materiais.

---

**📋 O que é cada coluna**

| Coluna | O que é | Como é calculado |
|---|---|---|
| **Código** | chave do local (`cod_fazenda`), a mesma usada em todas as páginas | — |
| **Fazenda · Cidade · Macro · Micro** | cadastro do local | vêm da base, sem cálculo |
| **sc/ha** | quanto o ambiente rendeu | média de **todas as parcelas de todos os híbridos** avaliados ali, inclusive os fora do filtro de status. Descreve o ambiente, não a linha |
| **Plantio** | data de semeadura do ensaio | `dataPlantioMilho`, da tabela fazenda. É do local, não do híbrido: todos os materiais daquela fazenda foram plantados no mesmo dia |
| **Amplitude** *(toggle)* | o quanto o local separa os materiais | média por material no local, depois `melhor − pior`. É a medida de exigência, e é independente da média |
| **Confrontos** *(toggle)* | quantos duelos a linha travou ali | `materiais STINE × concorrentes comerciais presentes no local` |
| **Aprov. %** *(toggle)* | quanto a linha Stine venceu naquele local, contra os comerciais | `vitórias ÷ confrontos`, com vitória = diferença acima de +1,0 sc/ha |
| **Dif. sc/ha** *(toggle)* | por quanto a linha ganha ou perde ali | média de `(material Stine − concorrente)` em todos os duelos do local |

Cada linha é um **local**, nunca uma parcela e nunca um híbrido. As colunas do toggle usam
**STINE** contra quem estiver marcado em Status do Adversário (Prod. 2) na barra lateral — o
padrão é CHECK, os concorrentes comerciais. Já a **média** e a **amplitude** do local não seguem
esse filtro: são calculadas sobre todos os híbridos avaliados ali, porque descrevem o ambiente e
não o confronto.

---

**📐 Como ler as cores**

- **Verde** → extremo superior de produtividade do recorte **em tela**. **Vermelho** → extremo
  inferior. As faixas são relativas: um local verde em MG pode ser vermelho em MT.
- **Plantio em vermelho** → os mais tardios do recorte.
- **Código colorido** → repete o destaque da linha para achar o local rápido nos outros gráficos.

> **Média baixa e amplitude alta não são a mesma coisa.** Um local pode render pouco e separar
bem os materiais, render muito e separar bem, ou render muito e não separar nada. Só quando o
local separa é que a diferença entre híbridos ali significa alguma coisa.

---

**🧭 Como usar — o exemplo**

Suponha que você queira defender um material para plantio tardio.

1. Filtre **Estado = MT** na barra lateral. A tabela mostra os locais do estado, do mais para o
   menos produtivo.
2. Olhe a coluna **Plantio**: os mais tardios do recorte saem em vermelho.
3. Marque as **fazendas** desses locais no filtro Fazenda da lateral. As cinco abas passam a
   responder só a eles.
4. Vá a **Todos os Materiais** e compare o aproveitamento com o do recorte cheio. Se subir, o
   argumento tem placar — não só coeficiente. Se ficar igual, o material **perde menos** que os
   concorrentes, mas não ganha deles: a frase de venda muda.
5. Repita com os locais de **menor sc/ha** para separar efeito de época de efeito de ambiente
   pobre. Se um local aparece nos dois recortes, ele está contado duas vezes.

---

**⚠️ Cuidados**

- **Menos de 3 locais não sustenta afirmação.** Com 2 locais, o aproveitamento contra um
  adversário só pode dar 0, 50 ou 100. O aviso aparece sozinho abaixo da tabela.
- **A cor é relativa ao filtro.** Mudou o estado, mudaram as faixas. Ao comparar duas leituras,
  confirme que o recorte é o mesmo.
- **Recortar locais é poderoso e perigoso.** Dá para escolher os três em que a linha ganha e
  apresentar como recorte — por isso o resumo abaixo da tabela sempre diz quantos locais estão
  em tela.
""")

        st.caption(
            f"{len(tab_loc)} locais · média de todas as parcelas do local, todos os híbridos · "
            "verde = extremo superior, vermelho = extremo inferior do recorte em tela"
            + (" · plantio em vermelho = mais tardio" if _cp else
               " · sem coluna de data de plantio na base")
        )
        if not _cp:
            _cand = candidatas_data(ta_filtrado)
            st.info(
                "**Data de plantio não encontrada.** Sem ela não existem a coluna Plantio aqui "
                "nem os terços de época em Todos os Materiais. O nome esperado é "
                "`dataPlantioMilho`, vindo da tabela fazenda pelo pipeline. "
                + (f"Colunas com conteúdo de data nesta base: `{'`, `'.join(_cand[:12])}`."
                   if _cand else
                   "Nenhuma coluna desta base contém datas — verifique se `dataPlantioMilho` "
                   "está em CONTEXTO no pipeline."))

        _headers = list(tab_loc.columns)
        _linhas = []
        for i in range(len(tab_loc)):
            r = tab_loc.iloc[i]
            c_sc, c_pl = CORES_FAIXA[_f_sc[i]], CORES_FAIXA[_f_pl[i]]
            # o código herda o destaque mais forte da linha, para achar o local rápido
            c_cod = c_pl if _f_pl[i] in ("baixo", "medio-baixo") else c_sc
            linha = [cel(r["Código"], cor=c_cod, bold=True),
                     cel(r["Fazenda"]), cel(r["Cidade"]), cel(r["Macro"]), cel(r["Micro"]),
                     cel(r["sc/ha"], "num1", cor=c_sc, bold=_f_sc[i] in ("alto", "baixo"))]
            if _cp:
                linha.append(cel(r["Plantio"], cor=c_pl, align="center",
                                 bold=_f_pl[i] in ("alto", "baixo")))
            if mostrar_placar:
                linha += [cel(r["Amplitude"], "num1"),
                          cel(r["Confrontos"], "num0"),
                          cel(r["Aprov. %"], "pct1", barra=r["Aprov. %"],
                              cor="#2976B6" if pd.notna(r["Aprov. %"]) and r["Aprov. %"] >= 50
                              else "#9AA5B1"),
                          cel(r["Dif. sc/ha"], "sinal1",
                              cor="#1E7A34" if pd.notna(r["Dif. sc/ha"]) and r["Dif. sc/ha"] > 0
                              else "#C0201E")]
            _linhas.append(linha)

        _leg_loc = [("#1E7A34", "extremo superior do recorte", "txt"),
                    ("#3E9E52", "acima da média", "txt"),
                    ("#E06C00", "abaixo da média", "txt"),
                    ("#C0201E", "extremo inferior · e plantio mais tardio", "txt")]
        if mostrar_placar:
            _leg_loc.append(("#2976B6", "aproveitamento da linha no local", "barra"))
        render_tabela(_headers, _linhas, "h2h_locais", "exp_loc", largura_1a=150,
                      legenda=_leg_loc)

        # ── resumo do recorte em tela — automático, sem seleção extra ────────
        _cods = list(tab_loc["Código"])
        _r = _car.loc[_car.index.isin(_cods)]
        _conf = int(_r["_conf"].fillna(0).sum())
        _vit = int(_r["_vit"].fillna(0).sum())
        _emp = int(_r["_emp"].fillna(0).sum())
        _der = int(_r["_der"].fillna(0).sum())
        _placar = (f' · a linha Stine faz <b>{_vit / _conf * 100:.1f}%</b> de aproveitamento '
                   f'({_vit}–{_emp}–{_der} em {_conf} confrontos)') if _conf else ""
        st.markdown(
            f'<div style="background:#F7F7F7;border-left:5px solid #2976B6;border-radius:8px;'
            f'padding:14px 18px;margin:14px 0 4px;">'
            f'<p style="margin:0;font-size:15.5px;line-height:1.6;color:#1A1A1A;'
            f'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
            f'<b>Recorte em tela: {len(_cods)} {"local" if len(_cods) == 1 else "locais"}</b> · '
            f'média {_base_loc["sc_ha"].mean():.1f} sc/ha{_placar}. '
            f'Este é o recorte que as outras quatro abas estão usando.</p></div>',
            unsafe_allow_html=True)
        if len(_cods) < 3:
            st.warning("Menos de 3 locais no filtro. Serve para investigar, não para afirmar.")



# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Tabela de Classificação
# ═══════════════════════════════════════════════════════════════════════════════
with tab1:
    secao_titulo(
        "HEAD-TO-HEAD · TABELA",
        "Classificação vs adversários",
        "Escolha o Produto 1 (STINE / EXP / DP2) e veja como ele se comporta contra cada "
        "adversário nos locais em que ambos foram avaliados.",
    )

    st.markdown(
        '<div style="background:#FFF8E1;border:1px solid #F5D76E;border-left:5px solid #D4A800;'
        'border-radius:8px;padding:14px 18px;margin:6px 0 14px;">'
        '<p style="margin:0;font-size:15.5px;line-height:1.6;color:#4A3B00;'
        'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        '<span style="font-size:19px;vertical-align:-1px;margin-right:8px;">⚠️</span>'
        'O resultado de cada local <b>não muda</b> com os filtros — o que muda é <b>quais locais '
        'entram na conta</b>. Por isso as médias, a % de vitórias e a classificação se recalculam '
        'sozinhos ao filtrar, e não é preciso rodar a análise de novo; o botão só é necessário ao '
        '<b>trocar o híbrido selecionado</b>.<br><br>'
        '<b>Ao comparar recortes diferentes, olhe sempre o nº de comparações.</b> Uma classificação '
        'obtida em 4 locais não tem o mesmo peso que a mesma classificação em 30 — o subtítulo acima '
        'mostra o recorte vigente e a tabela traz o nº de locais de cada confronto.'
        '</p></div>',
        unsafe_allow_html=True,
    )

    if df_p1.empty:
        st.warning("Nenhum híbrido STINE/EXP/DP2 encontrado com os filtros atuais.")
    elif df_p2.empty:
        st.warning("Nenhum adversário disponível. Verifique o filtro 'Status do Adversário' na barra lateral.")
    else:
        hibridos_p1 = sorted(df_p1["dePara"].dropna().unique())
        col_sel, col_btn = st.columns([4, 1])
        with col_sel:
            p1_t1 = st.selectbox("Produto 1 (STINE / EXP / DP2)", hibridos_p1, key="p1_t1")
        with col_btn:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            btn_t1 = st.button("Rodar análise", type="primary", key="btn_t1", use_container_width=True)

        key_t1 = f"res_t1_raw__{p1_t1}"

        if btn_t1:
            with st.spinner("Calculando confrontos..."):
                # roda contra TODOS os adversários da base — os filtros afetam só a exibição
                _p1_full = ta_raw[(ta_raw["status_material"].isin(STATUS_P1)) & (ta_raw["dePara"] == p1_t1)]
                _p2_full = ta_raw[ta_raw["dePara"] != p1_t1]
                df_cross_raw = cruzar_por_local(_p1_full, _p2_full)
                if not df_cross_raw.empty:
                    _meta = (ta_raw[["cod_fazenda", "safra", "regiao_macro", "regiao_micro",
                                     "estado_sigla", "cidade_nome", "nomeFazenda", "nomeResponsavel"]]
                             .drop_duplicates("cod_fazenda"))
                    df_cross_raw = df_cross_raw.merge(_meta, on="cod_fazenda", how="left")
                st.session_state[key_t1] = df_cross_raw

        if key_t1 not in st.session_state:
            st.info("Selecione o Produto 1 e clique em **Rodar análise** para calcular.")
        else:
            _cross_raw = st.session_state[key_t1]

            if _cross_raw.empty:
                st.info("Nenhum confronto encontrado para este híbrido.")
            else:
                # exibição respeita os filtros: locais ativos + status do adversário
                _locais_ativos = set(ta_filtrado["cod_fazenda"].dropna().unique())
                _cross_filt = _cross_raw[_cross_raw["cod_fazenda"].isin(_locais_ativos)]
                if status_p2_sel and "status_material_2" in _cross_filt.columns:
                    _cross_filt = _cross_filt[_cross_filt["status_material_2"].isin(status_p2_sel)]

                _rows = []
                for prod2, grp in _cross_filt.groupby("dePara_2"):
                    n = len(grp)
                    diff = grp["sc_ha_1"] - grp["sc_ha_2"]
                    vit = int((diff > EMPATE_MARGEM).sum())
                    emp = int((diff.abs() <= EMPATE_MARGEM).sum())
                    der = n - vit - emp
                    pct = round(vit / n * 100, 1) if n > 0 else np.nan
                    sc1, sc2 = grp["sc_ha_1"].mean(), grp["sc_ha_2"].mean()
                    kg1 = grp["kg_ha_1"].mean() if "kg_ha_1" in grp.columns else np.nan
                    kg2 = grp["kg_ha_2"].mean() if "kg_ha_2" in grp.columns else np.nan
                    dif_kg = (kg1 - kg2) if not (np.isnan(kg1) or np.isnan(kg2)) else np.nan
                    dif_sc = (sc1 - sc2) if not (np.isnan(sc1) or np.isnan(sc2)) else np.nan
                    dif_pct = ((sc1 / sc2) - 1) * 100 if (sc2 and not np.isnan(sc2)) else np.nan
                    classe, _ = classificar_h2h(pct)
                    _rows.append({
                        "Produto 1": p1_t1,
                        "kg/ha Prod 1": round(kg1, 0) if not np.isnan(kg1) else None,
                        "sc/ha Prod 1": round(sc1, 1),
                        "Produto 2": prod2,
                        "kg/ha Prod 2": round(kg2, 0) if not np.isnan(kg2) else None,
                        "sc/ha Prod 2": round(sc2, 1),
                        "Vitórias": vit,
                        "Empates": emp,
                        "Derrotas": der,
                        "Nº Comparações": n,
                        "Dif. (%)": round(dif_pct, 1) if not np.isnan(dif_pct) else None,
                        "Dif. (kg/ha)": round(dif_kg, 0) if not np.isnan(dif_kg) else None,
                        "Dif. (sc/ha)": round(dif_sc, 1) if not np.isnan(dif_sc) else None,
                        "% Vitórias": pct,
                        "Classe": classe,
                    })

                df_res = pd.DataFrame(_rows)
                if not df_res.empty:
                    df_res = df_res.sort_values("% Vitórias", ascending=False).reset_index(drop=True)

                if df_res.empty:
                    st.info("Nenhum confronto encontrado — o híbrido não compartilha locais com os "
                            "adversários selecionados.")
                else:
                    contexto_str = montar_contexto(_cross_filt)

                    _st_p1 = (df_p1[df_p1["dePara"] == p1_t1]["status_material"].iloc[0]
                              if not df_p1[df_p1["dePara"] == p1_t1].empty else "")
                    st.markdown(
                        f'<div style="margin:0.5rem 0 0.2rem;">'
                        f'<p style="font-size:13px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                        f'letter-spacing:0.05em;margin:0 0 4px;">Análise H2H · Produto 1</p>'
                        f'<h2 style="font-size:1.9rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.2;">'
                        f'<span style="color:#27AE60;">{p1_t1}</span>'
                        f'<span style="font-size:1rem;font-weight:500;color:#6B7280;margin-left:10px;">'
                        f'{_st_p1} · {len(df_res)} adversários</span></h2>'
                        f'<p style="font-size:14px;color:#6B7280;margin:4px 0 0;">{contexto_str}</p></div>',
                        unsafe_allow_html=True)

                    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

                    with st.popover("ℹ️ Como interpretar esta tabela", use_container_width=False):
                        st.markdown(f"""
**📌 O que esta tabela mostra**

Cada linha é um confronto direto entre o **Produto 1** e um adversário (Produto 2), calculado
exclusivamente nos locais onde **ambos foram avaliados simultaneamente**.

---

**📐 Como ler as colunas**

- **kg/ha · sc/ha Prod 1 / Prod 2** → médias *apenas nos locais compartilhados* (não é a média
  geral do híbrido).
- **Vitórias / Empates / Derrotas** → nº de locais em cada resultado.
- **Nº Comparações** → total de locais com ambos avaliados.
- **% Vitórias** → Vitórias ÷ Comparações × 100 — é a base da classificação.
- **Dif. (%)** → quanto o Produto 1 produz a mais ou a menos, em termos relativos.
- **Dif. (kg/ha) e (sc/ha)** → diferença absoluta média entre os dois.

> ⚠️ **Empate**: diferença de até **±{EMPATE_MARGEM:.0f} sc/ha** não conta como vitória nem derrota.

---

**🎨 Legenda das cores — % de vitórias**
""")
                        ca, cb, cc, cd = st.columns(4)
                        ca.markdown('<div style="background:#90EE90;border-radius:6px;padding:8px;text-align:center;">'
                                    '<b style="color:#1A1A1A;">Alta Performance</b><br>'
                                    '<span style="font-size:12px;color:#1A1A1A;">&gt; 75% de vitórias</span></div>',
                                    unsafe_allow_html=True)
                        cb.markdown('<div style="background:#87CEFF;border-radius:6px;padding:8px;text-align:center;">'
                                    '<b style="color:#1A1A1A;">Superior</b><br>'
                                    '<span style="font-size:12px;color:#1A1A1A;">56 – 75% de vitórias</span></div>',
                                    unsafe_allow_html=True)
                        cc.markdown('<div style="background:#FFFF00;border-radius:6px;padding:8px;text-align:center;">'
                                    '<b style="color:#1A1A1A;">Competitivo</b><br>'
                                    '<span style="font-size:12px;color:#1A1A1A;">46 – 55% de vitórias</span></div>',
                                    unsafe_allow_html=True)
                        cd.markdown('<div style="background:#FF0000;border-radius:6px;padding:8px;text-align:center;">'
                                    '<b style="color:#FFFFFF;">Restrito</b><br>'
                                    '<span style="font-size:12px;color:#FFFFFF;">≤ 45% de vitórias</span></div>',
                                    unsafe_allow_html=True)
                        st.markdown("""
---

**💡 Como interpretar**

- **Alta Performance** → vence em mais de 3/4 dos locais: consistentemente superior a esse adversário.
- **Superior** → vence na maioria dos locais.
- **Competitivo** → resultado equilibrado, nenhum se destaca claramente.
- **Restrito** → perde na maioria dos locais frente a esse adversário — atenção ao posicionamento.

> As médias mudam conforme o adversário, porque cada confronto usa só os locais em comum entre os
dois. Por isso o mesmo híbrido pode aparecer com médias diferentes em linhas diferentes.
""")

                    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                    contagem = df_res["Classe"].value_counts()
                    total_cls = len(df_res)
                    c1, c2, c3, c4 = st.columns(4)
                    for col_ui, label, cor_txt in zip(
                            [c1, c2, c3, c4],
                            ["Alta Performance", "Superior", "Competitivo", "Restrito"],
                            ["#27AE60", "#1E40AF", "#F2C811", "#FF0000"]):
                        n_cls = int(contagem.get(label, 0))
                        pct_cl = f"{n_cls / total_cls * 100:.0f}%" if total_cls else "—"
                        col_ui.markdown(
                            f'<div style="border:1px solid #E5E7EB;border-radius:10px;padding:10px 14px;'
                            f'background:#FFFFFF;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,0.07);">'
                            f'<p style="margin:0;font-size:14px;font-weight:600;color:#374151;">{label}</p>'
                            f'<p style="margin:4px 0 0;font-size:2.2rem;font-weight:700;color:{cor_txt};">'
                            f'{n_cls} <span style="font-size:1.2rem;font-weight:500;">({pct_cl})</span></p></div>',
                            unsafe_allow_html=True)

                    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

                    ag_table_h2h(df_res, height=min(680, int((36 + 32 * len(df_res) + 20) * 1.3)))
                    st.download_button("⬇️ Exportar Excel", data=to_excel(df_res),
                                       file_name=f"h2h_{p1_t1}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="dl_t1")
                    st.caption(f"ℹ️ Confrontos calculados só nos locais compartilhados por cada par. "
                               f"Empate = diferença de até ±{EMPATE_MARGEM:.0f} sc/ha.")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Análise por Local
# ═══════════════════════════════════════════════════════════════════════════════
with tab2:
    secao_titulo(
        "HEAD-TO-HEAD · POR LOCAL",
        "Diferença de produtividade por local",
        "Selecione um par específico e veja a diferença de sc/ha em cada local compartilhado.",
    )

    st.markdown(
        '<div style="background:#FFF8E1;border:1px solid #F5D76E;border-left:5px solid #D4A800;'
        'border-radius:8px;padding:14px 18px;margin:6px 0 14px;">'
        '<p style="margin:0;font-size:15.5px;line-height:1.6;color:#4A3B00;'
        'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        '<span style="font-size:19px;vertical-align:-1px;margin-right:8px;">⚠️</span>'
        'O resultado de cada local <b>não muda</b> com os filtros — o que muda é <b>quais locais '
        'entram na conta</b>. Por isso as médias, a % de vitórias e a classificação se recalculam '
        'sozinhos ao filtrar, e não é preciso rodar a análise de novo; o botão só é necessário ao '
        '<b>trocar o híbrido selecionado</b>.<br><br>'
        '<b>Ao comparar recortes diferentes, olhe sempre o nº de comparações.</b> Uma classificação '
        'obtida em 4 locais não tem o mesmo peso que a mesma classificação em 30 — o subtítulo acima '
        'mostra o recorte vigente e a tabela traz o nº de locais de cada confronto.'
        '</p></div>',
        unsafe_allow_html=True,
    )

    if df_p1.empty:
        st.warning("Dados insuficientes com os filtros atuais.")
    else:
        hibridos_p1_t2 = sorted(df_p1["dePara"].dropna().unique())
        col_a2, col_b2, col_c2 = st.columns([2, 2, 1])

        with col_a2:
            p1_t2 = st.selectbox("Produto 1 (STINE / EXP / DP2)", hibridos_p1_t2, key="p1_t2")

        locais_p1_t2 = set(df_p1[df_p1["dePara"] == p1_t2]["cod_fazenda"].dropna().unique())
        # ADVERSÁRIO AQUI NÃO PASSA PELO FILTRO DE STATUS DA LATERAL, e é de propósito: esta aba
        # é um confronto 1 contra 1 que você escolhe a dedo, então limitar a lista a CHECK
        # atrapalharia justamente a pergunta comum "qual dos meus dois posicionar neste ambiente".
        # O filtro da lateral continua valendo onde ele faz sentido — nas abas que agregam contra
        # um universo de adversários (Locais e Todos os Materiais). Aqui vale a regra simples:
        # qualquer material avaliado nos mesmos locais, menos o próprio Produto 1.
        _pool_adv = ta_filtrado[(ta_filtrado["cod_fazenda"].isin(locais_p1_t2))
                                & (ta_filtrado["dePara"] != p1_t2)]
        _status_adv = _pool_adv.groupby("dePara")["status_material"].first().to_dict()
        adv_disp = sorted(_status_adv)

        with col_b2:
            if adv_disp:
                p2_t2 = st.selectbox("Produto 2 (adversário)", adv_disp, key="p2_t2")
            else:
                st.warning("Nenhum outro material foi avaliado nos mesmos locais deste híbrido.")
                p2_t2 = None

        with col_c2:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            btn_t2 = st.button("Rodar análise", type="primary", key="btn_t2", use_container_width=True)

        key_t2 = f"res_t2_raw__{p1_t2}__{p2_t2}"

        if btn_t2 and p2_t2:
            with st.spinner("Calculando..."):
                # cruza o par em TODOS os locais da base — os filtros afetam só a exibição
                d1_loc = (ta_raw[ta_raw["dePara"] == p1_t2][["cod_fazenda", "sc_ha", "kg_ha"]]
                          .dropna(subset=["sc_ha"])
                          .groupby("cod_fazenda", as_index=False)[["sc_ha", "kg_ha"]].mean())
                d2_loc = (ta_raw[ta_raw["dePara"] == p2_t2][["cod_fazenda", "sc_ha", "kg_ha"]]
                          .dropna(subset=["sc_ha"])
                          .groupby("cod_fazenda", as_index=False)[["sc_ha", "kg_ha"]].mean())
                st.session_state[key_t2] = d1_loc.merge(d2_loc, on="cod_fazenda",
                                                        suffixes=("_1", "_2")).copy()

        if key_t2 not in st.session_state or not p2_t2:
            st.info("Selecione os dois híbridos e clique em **Rodar análise** para calcular.")
        else:
            # aplica os filtros vigentes na exibição (mesma lógica da Tab 1)
            _loc_ativos_t2 = set(ta_filtrado["cod_fazenda"].dropna().unique())
            df_loc = st.session_state[key_t2]
            df_loc = df_loc[df_loc["cod_fazenda"].isin(_loc_ativos_t2)].copy()
            if not df_loc.empty:
                df_loc["diff_sc"] = df_loc["sc_ha_1"] - df_loc["sc_ha_2"]
                df_loc["resultado"] = df_loc["diff_sc"].apply(
                    lambda x: "Vitória" if x > EMPATE_MARGEM
                    else ("Empate" if abs(x) <= EMPATE_MARGEM else "Derrota"))
                df_loc = df_loc.sort_values("diff_sc").reset_index(drop=True)

            if df_loc.empty:
                st.info("Nenhum local compartilhado encontrado para este par.")
            else:
                n_loc = len(df_loc)
                n_vit = int((df_loc["resultado"] == "Vitória").sum())
                n_emp = int((df_loc["resultado"] == "Empate").sum())
                n_der = int((df_loc["resultado"] == "Derrota").sum())
                vit_sc = df_loc.loc[df_loc["resultado"] == "Vitória", "diff_sc"]
                der_sc = df_loc.loc[df_loc["resultado"] == "Derrota", "diff_sc"]
                max_vit = float(vit_sc.max()) if len(vit_sc) else np.nan
                med_vit = float(vit_sc.mean()) if len(vit_sc) else np.nan
                min_der = float(der_sc.min()) if len(der_sc) else np.nan
                med_der = float(der_sc.mean()) if len(der_sc) else np.nan

                _base_t2 = ta_filtrado[ta_filtrado["cod_fazenda"].isin(df_loc["cod_fazenda"])]
                contexto_t2 = montar_contexto(_base_t2, "locais compartilhados")

                _st1 = (df_p1[df_p1["dePara"] == p1_t2]["status_material"].iloc[0]
                        if not df_p1[df_p1["dePara"] == p1_t2].empty else "")
                # do pool desta aba: o adversário pode estar fora do filtro de status da lateral,
                # e aí `df_p2` não o teria — o rótulo sairia vazio no cabeçalho
                _st2 = _status_adv.get(p2_t2, "")

                st.markdown(
                    f'<div style="margin:0.5rem 0 0.2rem;">'
                    f'<p style="font-size:13px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                    f'letter-spacing:0.05em;margin:0 0 4px;">Análise H2H · Confronto direto</p>'
                    f'<h2 style="font-size:1.9rem;font-weight:700;color:#1A1A1A;margin:0;line-height:1.2;">'
                    f'<span style="color:#27AE60;">{p1_t2}</span>'
                    f'<span style="font-size:0.85rem;font-weight:500;color:#6B7280;margin-left:6px;">{_st1}</span>'
                    f'<span style="font-size:1.1rem;font-weight:500;color:#6B7280;margin:0 12px;">vs</span>'
                    f'<span style="color:#1A1A1A;">{p2_t2}</span>'
                    f'<span style="font-size:0.85rem;font-weight:500;color:#6B7280;margin-left:6px;">{_st2}</span>'
                    f'</h2><p style="font-size:14px;color:#6B7280;margin:4px 0 0;">{contexto_t2}</p></div>',
                    unsafe_allow_html=True)

                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                col_pop1, col_pop2, _ = st.columns([2, 2, 4])
                with col_pop1:
                    with st.popover("ℹ️ Como interpretar", use_container_width=True):
                        st.markdown(f"""
**📌 O que este painel mostra**

Confronto direto entre **Produto 1** e **Produto 2** nos locais onde **ambos foram avaliados
simultaneamente**.

O adversário não precisa ser um concorrente: a lista traz **todos os materiais avaliados nos
mesmos locais** do Produto 1, com o status ao lado do nome — CHECK, STINE, EXP ou DP2. Dá para
confrontar dois materiais do próprio portfólio, útil para decidir posicionamento entre eles ou
para medir um experimental contra o híbrido que ele pretende substituir. O próprio Produto 1 fica
fora da lista.

Esta aba **não** usa o filtro Status do Adversário da barra lateral: aqui você escolhe o par a
dedo, e limitar a lista atrapalharia. O filtro continua valendo nas abas que agregam contra um
universo de adversários (Locais e Todos os Materiais). Os filtros de local, safra e responsável
valem aqui normalmente.

> Ao comparar dois materiais Stine, lembre que "vitória" e "derrota" continuam sendo do ponto de
vista do Produto 1. Não é disputa de mercado, é escolha de qual dos dois posicionar naquele
ambiente — e empate, aqui, é resposta legítima: significa que a decisão pode ser tomada por outro
critério, como sanidade, estande ou disponibilidade de semente.

---

**📐 Como ler os cards**

- **Locais avaliados** → total de locais com os dois híbridos presentes.
- **Vitórias** → locais em que o Produto 1 superou por mais de **+{EMPATE_MARGEM:.0f} sc/ha**
  (*Max* = maior diferença positiva; *Média* = média das vitórias).
- **Empates** → diferença dentro de ±{EMPATE_MARGEM:.0f} sc/ha.
- **Derrotas** → locais em que ficou abaixo por mais de {EMPATE_MARGEM:.0f} sc/ha
  (*Min* = pior diferença; *Média* = média das derrotas).

---

**🗺️ Mapa e barras**

- No **mapa**, a cor do ponto indica a magnitude da diferença naquele local, e o símbolo distingue
  a safra quando há mais de uma ativa.
- Nas **barras**, cada linha é um local, ordenado da maior derrota à maior vitória. Verde =
  vitória, amarelo = empate, vermelho = derrota.

> Uma leitura útil: muitos locais próximos de zero indicam materiais equivalentes; poucos locais
com diferenças grandes em ambos os sentidos indicam que o ambiente decide o vencedor.
""")
                with col_pop2:
                    with st.popover("Dicionário de locais", use_container_width=True):
                        df_dic_t2 = (ta_filtrado[ta_filtrado["cod_fazenda"].isin(df_loc["cod_fazenda"])]
                                     [["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                                     .drop_duplicates()
                                     .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                                     .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                                      "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                                     .reset_index(drop=True))
                        st.markdown(f"Referência dos **{len(df_dic_t2)} locais** do confronto.")
                        st.dataframe(df_dic_t2, hide_index=True, use_container_width=True)

                st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

                pct_vit = f"{n_vit / n_loc * 100:.0f}%" if n_loc else "—"
                pct_emp = f"{n_emp / n_loc * 100:.0f}%" if n_loc else "—"
                pct_der = f"{n_der / n_loc * 100:.0f}%" if n_loc else "—"
                COR_EMPATE_CARD = "#D4A800"
                card = ("border:1px solid #E5E7EB;border-radius:10px;padding:12px 16px;"
                        "background:#FFFFFF;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,0.07);")

                k1, k2, k3, k4 = st.columns(4)
                k1.markdown(f'<div style="{card}">'
                            f'<p style="margin:0;font-size:12px;color:#6B7280;">Locais avaliados</p>'
                            f'<p style="margin:6px 0 0;font-size:1.9rem;font-weight:700;color:#1A1A1A;">{n_loc}</p>'
                            f'</div>', unsafe_allow_html=True)

                sub_v = (f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_VITORIA};">Max: {max_vit:+.1f} sc/ha</p>'
                         f'<p style="margin:0;font-size:14px;font-weight:600;color:{COR_VITORIA};">Média: {med_vit:+.1f} sc/ha</p>'
                         ) if not np.isnan(max_vit) else '<p style="margin:2px 0;font-size:14px;">&nbsp;</p><p style="margin:0;font-size:14px;">&nbsp;</p>'
                k2.markdown(f'<div style="{card}border-top:3px solid {COR_VITORIA};">'
                            f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">Vitórias</p>{sub_v}'
                            f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_VITORIA};">'
                            f'{n_vit} <span style="font-size:1rem;font-weight:600;">({pct_vit})</span></p></div>',
                            unsafe_allow_html=True)

                k3.markdown(f'<div style="{card}border-top:3px solid {COR_EMPATE_CARD};">'
                            f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">Empates</p>'
                            f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_EMPATE_CARD};">'
                            f'Entre ±{EMPATE_MARGEM:.0f} sc/ha</p>'
                            f'<p style="margin:0;font-size:14px;">&nbsp;</p>'
                            f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_EMPATE_CARD};">'
                            f'{n_emp} <span style="font-size:1rem;font-weight:600;">({pct_emp})</span></p></div>',
                            unsafe_allow_html=True)

                sub_d = (f'<p style="margin:2px 0;font-size:14px;font-weight:600;color:{COR_DERROTA};">Min: {min_der:+.1f} sc/ha</p>'
                         f'<p style="margin:0;font-size:14px;font-weight:600;color:{COR_DERROTA};">Média: {med_der:+.1f} sc/ha</p>'
                         ) if not np.isnan(min_der) else '<p style="margin:2px 0;font-size:14px;">&nbsp;</p><p style="margin:0;font-size:14px;">&nbsp;</p>'
                k4.markdown(f'<div style="{card}border-top:3px solid {COR_DERROTA};">'
                            f'<p style="margin:0;font-size:15px;font-weight:700;color:#1A1A1A;">Derrotas</p>{sub_d}'
                            f'<p style="margin:6px 0;font-size:1.9rem;font-weight:700;color:{COR_DERROTA};">'
                            f'{n_der} <span style="font-size:1rem;font-weight:600;">({pct_der})</span></p></div>',
                            unsafe_allow_html=True)

                st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

                col_donut, col_mapa = st.columns([1, 2])

                with col_donut:
                    fig_d = go_plt.Figure(go_plt.Pie(
                        labels=["Vitórias", "Empates", "Derrotas"], values=[n_vit, n_emp, n_der],
                        hole=0.55, marker_colors=[COR_VITORIA, COR_EMPATE, COR_DERROTA],
                        textinfo="label+percent", textposition="outside",
                        textfont=dict(size=12, family="Helvetica Neue, sans-serif", color="#111111"),
                        hovertemplate="%{label}: %{value} local(is) (%{percent})<extra></extra>",
                        sort=False, pull=[0.03, 0.03, 0.03],
                        domain=dict(x=[0.15, 0.85], y=[0.05, 0.90])))
                    fig_d.update_layout(
                        title=dict(text="Resultado geral do confronto",
                                   font=dict(size=13, color="#111111"),
                                   x=0.5, xanchor="center", y=0.99, yanchor="top"),
                        showlegend=False, height=420, margin=dict(t=80, b=20, l=60, r=60),
                        paper_bgcolor="#FFFFFF", font=dict(family="Helvetica Neue, sans-serif"))
                    st.plotly_chart(fig_d, use_container_width=True)

                with col_mapa:
                    _cols_coord = ["cod_fazenda", "latitude", "longitude", "nomeFazenda",
                                   "cidade_nome", "estado_sigla", "safra"]
                    if all(c in ta_filtrado.columns for c in ["latitude", "longitude"]):
                        df_coords = (ta_filtrado[_cols_coord].dropna(subset=["latitude", "longitude"])
                                     .sort_values("safra").drop_duplicates("cod_fazenda", keep="last"))
                        df_map = df_loc.merge(df_coords, on="cod_fazenda", how="left") \
                                       .dropna(subset=["latitude", "longitude"])
                    else:
                        df_map = pd.DataFrame()

                    folium = None
                    if df_map.empty:
                        st.info("Coordenadas não disponíveis para os locais deste confronto. "
                                "Atualize o pipeline e limpe o cache para habilitar o mapa.")
                    else:
                        try:
                            import folium
                            from streamlit_folium import st_folium
                        except ModuleNotFoundError:
                            folium = None
                            st.warning("Mapa indisponível: as bibliotecas de mapa não estão "
                                       "instaladas. Rode `pip install folium==0.20.0 "
                                       "streamlit-folium==0.27.1` e reinicie o app. "
                                       "O restante da análise funciona normalmente.")

                    if not df_map.empty and folium is not None:
                        def _cor_diff(d):
                            if d >= 10:  return "#27AE60"
                            if d >= 5:   return "#52C97A"
                            if d >= 2:   return "#A8E6BC"
                            if d >= -1:  return "#FFFF00"
                            if d >= -5:  return "#FF9900"
                            if d >= -10: return "#FF4400"
                            return "#FF0000"

                        _safras_u = sorted(df_map["safra"].dropna().unique().tolist())
                        _icones = ["circle", "star", "square", "diamond"]
                        _simb = {s: _icones[i % len(_icones)] for i, s in enumerate(_safras_u)}

                        def _svg_circulo(c):
                            return (f'<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24">'
                                    f'<circle cx="12" cy="12" r="10" fill="{c}" stroke="white" stroke-width="2"/></svg>')

                        def _svg_quadrado(c):
                            return (f'<svg xmlns="http://www.w3.org/2000/svg" width="24" height="24">'
                                    f'<rect x="2" y="2" width="20" height="20" rx="3" fill="{c}" stroke="white" stroke-width="2"/></svg>')

                        def _svg_estrela(c):
                            return (f'<svg xmlns="http://www.w3.org/2000/svg" width="26" height="26" viewBox="0 0 26 26">'
                                    f'<polygon points="13,1 16.5,9.5 26,10.5 19.5,17 21.5,26 13,21.5 4.5,26 6.5,17 0,10.5 9.5,9.5" '
                                    f'fill="{c}" stroke="white" stroke-width="1.5"/></svg>')

                        _SVG_FN = {"circle": _svg_circulo, "square": _svg_quadrado, "star": _svg_estrela}

                        # marcadores na MESMA coordenada (ex.: mesma fazenda em duas safras, ou
                        # locais que herdaram a coordenada da cidade) ficariam empilhados e só o
                        # último apareceria — espalha em círculo pequeno para todos ficarem visíveis
                        import math
                        _RAIO_JITTER = 0.03          # ~3 km: separa visualmente sem tirar do município
                        df_map = df_map.reset_index(drop=True)
                        df_map["_dlat"] = 0.0
                        df_map["_dlon"] = 0.0
                        _chave_pos = (df_map["latitude"].round(4).astype(str) + "_"
                                      + df_map["longitude"].round(4).astype(str))
                        _n_empilhados = 0
                        for _, _idx in df_map.groupby(_chave_pos).groups.items():
                            _idx = list(_idx)
                            if len(_idx) > 1:
                                _n_empilhados += len(_idx)
                                for _k, _i in enumerate(_idx):
                                    _ang = 2 * math.pi * _k / len(_idx)
                                    df_map.at[_i, "_dlat"] = _RAIO_JITTER * math.cos(_ang)
                                    df_map.at[_i, "_dlon"] = _RAIO_JITTER * math.sin(_ang)

                        m = folium.Map(location=[df_map["latitude"].mean(), df_map["longitude"].mean()],
                                       zoom_start=5, tiles="OpenStreetMap")
                        for _, row in df_map.iterrows():
                            _cor = _cor_diff(row["diff_sc"])
                            _safra = str(row.get("safra", ""))
                            _icon = _simb.get(_safra, "circle")
                            _svg = _SVG_FN.get(_icon, _svg_circulo)(_cor)
                            _size = 26 if _icon == "star" else 24
                            _deslocado = (row["_dlat"] != 0) or (row["_dlon"] != 0)
                            _nota_pos = ("<br><i style='font-size:11px;color:#6B7280;'>ponto deslocado "
                                         "para não sobrepor outro local na mesma coordenada</i>"
                                         if _deslocado else "")
                            folium.Marker(
                                location=[row["latitude"] + row["_dlat"],
                                          row["longitude"] + row["_dlon"]],
                                popup=folium.Popup(
                                    f"<b>{row['nomeFazenda']}</b><br>{row['cidade_nome']} — {row['estado_sigla']}<br>"
                                    f"Safra: {_safra}<br><b>{p1_t2}:</b> {row['sc_ha_1']:.1f} sc/ha<br>"
                                    f"<b>{p2_t2}:</b> {row['sc_ha_2']:.1f} sc/ha<br>"
                                    f"<b>Diferença:</b> {row['diff_sc']:+.1f} sc/ha{_nota_pos}", max_width=280),
                                tooltip=f"{row['nomeFazenda']} · {_safra} · {row['diff_sc']:+.1f} sc/ha",
                                icon=folium.DivIcon(html=_svg, icon_size=(_size, _size),
                                                    icon_anchor=(_size // 2, _size // 2))).add_to(m)

                        _leg_cores = [("≥ +10 sc/ha", "#27AE60"), ("≥ +5 sc/ha", "#52C97A"),
                                      ("≥ +2 sc/ha", "#A8E6BC"), (f"±{EMPATE_MARGEM:.0f} sc/ha (empate)", "#FFFF00"),
                                      ("< −2 sc/ha", "#FF9900"), ("< −5 sc/ha", "#FF4400"),
                                      ("≤ −10 sc/ha", "#FF0000")]
                        _leg_html = "".join(
                            f'<div style="display:flex;align-items:center;gap:6px;margin:2px 0;">'
                            f'<div style="width:14px;height:14px;border-radius:50%;background:{c};'
                            f'border:1px solid #ccc;flex-shrink:0;"></div>'
                            f'<span style="font-size:11px;color:#374151;">{lbl}</span></div>'
                            for lbl, c in _leg_cores)
                        _SIMB_LEG = {"circle": "●", "square": "■", "star": "★", "diamond": "◆"}
                        _leg_safras = "".join(
                            f'<div style="display:flex;align-items:center;gap:6px;margin:2px 0;">'
                            f'<span style="font-size:16px;color:#374151;">{_SIMB_LEG.get(_simb.get(sf, "circle"), "●")}</span>'
                            f'<span style="font-size:11px;color:#374151;">{sf}</span></div>'
                            for sf in _safras_u)

                        st.markdown('<p style="font-size:13px;font-weight:600;color:#4A4A4A;margin:0 0 6px;">'
                                    'Locais por diferença</p>', unsafe_allow_html=True)
                        _cm, _cl = st.columns([4, 1])
                        with _cm:
                            st_folium(m, use_container_width=True, height=420, returned_objects=[])
                            if _n_empilhados:
                                st.caption(
                                    f"ℹ️ {_n_empilhados} locais dividem a mesma coordenada (mesma fazenda "
                                    "em safras diferentes, ou coordenada herdada da cidade). Eles foram "
                                    "afastados alguns quilômetros entre si só para ficarem visíveis — a "
                                    "posição real é a mesma, e o popup avisa quando o ponto está deslocado.")
                        with _cl:
                            st.markdown(
                                f"<div style='background:white;padding:10px 14px;border-radius:8px;"
                                f"border:1px solid #E5E7EB;box-shadow:0 1px 4px rgba(0,0,0,0.08);"
                                f"display:inline-block;'>"
                                f"<p style='font-size:12px;font-weight:700;color:#1A1A1A;margin:0 0 6px;'>"
                                f"Diferença (sc/ha)</p>{_leg_html}"
                                + (f"<p style='font-size:12px;font-weight:700;color:#1A1A1A;margin:8px 0 4px;'>"
                                   f"Safra</p>{_leg_safras}" if len(_safras_u) > 1 else "")
                                + "</div>", unsafe_allow_html=True)

                st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

                cores_bar = df_loc["resultado"].map({"Vitória": COR_VITORIA, "Empate": COR_EMPATE,
                                                     "Derrota": COR_DERROTA}).tolist()
                fig_b = go_plt.Figure(go_plt.Bar(
                    x=df_loc["diff_sc"].round(1), y=df_loc["cod_fazenda"], orientation="h",
                    marker_color=cores_bar, text=df_loc["diff_sc"].round(1), textposition="outside",
                    textfont=dict(size=11, color="#111111"),
                    hovertemplate="<b>%{y}</b><br>Diferença: %{x:+.1f} sc/ha<extra></extra>"))
                fig_b.add_vline(x=0, line_color="#333333", line_width=2)
                fig_b.update_layout(
                    title=dict(text=f"Diferença de produtividade por local — {p1_t2} × {p2_t2}",
                               font=dict(size=13, color="#111111")),
                    xaxis=dict(title="Diferença (sc/ha)", tickfont=dict(size=11, color="#111111"),
                               zerolinecolor="#CCCCCC"),
                    yaxis=dict(title="Local", tickfont=dict(size=11, color="#111111")),
                    height=max(380, n_loc * 28 + 100), margin=dict(t=50, b=50, l=130, r=100),
                    plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"))
                st.plotly_chart(fig_b, use_container_width=True)

                st.markdown("**Dados por local**")
                df_exp = df_loc[["cod_fazenda", "sc_ha_1", "sc_ha_2", "diff_sc", "resultado"]].copy()
                df_exp.columns = ["Local", f"sc/ha — {p1_t2}", f"sc/ha — {p2_t2}",
                                  "Diferença (sc/ha)", "Resultado"]
                df_exp = df_exp.round(1)
                _est_dif = JsCode("""
                function(params) {
                    const v = Number(params.value);
                    if (isNaN(v)) return {'textAlign':'center'};
                    if (v > 0) return {'color':'#15803D','fontWeight':'800','textAlign':'center'};
                    if (v < 0) return {'color':'#B91C1C','fontWeight':'800','textAlign':'center'};
                    return {'textAlign':'center'};
                }
                """)
                _est_res = JsCode("""
                function(params) {
                    const v = params.value;
                    if (v === 'Vitória') return {'backgroundColor':'rgba(39,174,96,0.18)','fontWeight':'700','textAlign':'center'};
                    if (v === 'Derrota') return {'backgroundColor':'rgba(255,0,0,0.15)','fontWeight':'700','textAlign':'center'};
                    if (v === 'Empate')  return {'backgroundColor':'rgba(255,255,0,0.25)','fontWeight':'700','textAlign':'center'};
                    return {'textAlign':'center'};
                }
                """)
                ag_table_h2h(df_exp, height=min(520, 40 + 32 * min(len(df_exp), 14) + 20),
                             estilos_col={"Diferença (sc/ha)": _est_dif, "Resultado": _est_res})
                st.download_button("⬇️ Exportar Excel", data=to_excel(df_exp),
                                   file_name=f"h2h_local_{p1_t2}_vs_{p2_t2}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_t2")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Desvios por Ambiente
# ═══════════════════════════════════════════════════════════════════════════════
with tab3:
    secao_titulo(
        "HEAD-TO-HEAD · DESVIOS",
        "Como cada híbrido se comporta em ambientes favoráveis vs. adversos?",
        "Desvio em relação à média do local — pontos acima de zero indicam que o híbrido superou a "
        "média daquele ambiente.",
    )

    st.markdown(
        '<div style="background:#FFF8E1;border:1px solid #F5D76E;border-left:5px solid #D4A800;'
        'border-radius:8px;padding:14px 18px;margin:6px 0 14px;">'
        '<p style="margin:0;font-size:15.5px;line-height:1.6;color:#4A3B00;'
        'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
        '<span style="font-size:19px;vertical-align:-1px;margin-right:8px;">⚠️</span>'
        'A média do local considera <b>todos os híbridos avaliados ali</b>, inclusive os fora do '
        'filtro — assim o desvio de cada local <b>não muda</b> conforme a seleção. Os filtros '
        'definem <b>quais locais</b> entram, e a reta e o R² <b>se recalculam '
        'sozinhos</b>; o botão só é necessário ao <b>trocar os híbridos</b>.<br><br>'
        '<b>Cuidado ao comparar recortes:</b> uma inclinação calculada com 5 locais é muito menos '
        'confiável que a mesma com 30. Confira o nº de locais e o R² antes de concluir.'
        '</p></div>',
        unsafe_allow_html=True,
    )

    if df_p1.empty:
        st.warning("Dados insuficientes com os filtros atuais.")
    else:
        hibridos_p1_t3 = sorted(df_p1["dePara"].dropna().unique())
        col_a3, col_b3, col_c3 = st.columns([2, 2, 1])

        with col_a3:
            p1_t3 = st.selectbox("Produto 1 (STINE / EXP / DP2)", hibridos_p1_t3, key="p1_t3")

        _loc_p1_t3 = set(df_p1[df_p1["dePara"] == p1_t3]["cod_fazenda"].dropna().unique())
        adv_disp_t3 = sorted(ta_filtrado[(ta_filtrado["cod_fazenda"].isin(_loc_p1_t3))
                                         & (ta_filtrado["dePara"] != p1_t3)]["dePara"].dropna().unique())

        with col_b3:
            if adv_disp_t3:
                p2_t3 = st.selectbox("Produto 2 (adversário)", adv_disp_t3, key="p2_t3")
            else:
                st.warning("Nenhum adversário com locais em comum.")
                p2_t3 = None

        with col_c3:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            btn_t3 = st.button("Rodar análise", type="primary", key="btn_t3", use_container_width=True)

        key_t3 = f"res_t3_raw__{p1_t3}__{p2_t3}"

        if btn_t3 and p2_t3:
            with st.spinner("Calculando desvios..."):
                # média do local sobre TODOS os híbridos avaliados ali (não muda com o filtro)
                _base = ta_raw[ta_raw["sc_ha"] > 0]
                _media_loc = _base.groupby("cod_fazenda")["sc_ha"].mean().rename("media_local")

                _l1 = set(_base[_base["dePara"] == p1_t3]["cod_fazenda"].dropna())
                _l2 = set(_base[_base["dePara"] == p2_t3]["cod_fazenda"].dropna())
                _comuns = _l1 & _l2

                def _desvios(hib):
                    return (_base[(_base["dePara"] == hib) & (_base["cod_fazenda"].isin(_comuns))]
                            [["cod_fazenda", "sc_ha"]].dropna()
                            .groupby("cod_fazenda", as_index=False)["sc_ha"].mean()
                            .join(_media_loc, on="cod_fazenda")
                            .assign(desvio=lambda d: d["sc_ha"] - d["media_local"], hibrido=hib))

                st.session_state[key_t3] = (_desvios(p1_t3), _desvios(p2_t3))

        if key_t3 not in st.session_state or not p2_t3:
            st.info("Selecione os dois híbridos e clique em **Rodar análise** para calcular.")
        else:
            _df1, _df2 = st.session_state[key_t3]
            # filtros vigentes agem só na exibição
            _loc_ativos_t3 = set(ta_filtrado["cod_fazenda"].dropna().unique())
            _df1 = _df1[_df1["cod_fazenda"].isin(_loc_ativos_t3)].copy()
            _df2 = _df2[_df2["cod_fazenda"].isin(_loc_ativos_t3)].copy()

            if _df1.empty or _df2.empty:
                st.info("Nenhum local compartilhado dentro do recorte filtrado.")
            else:
                with st.popover("ℹ️ Como interpretar · Desvios por Ambiente", use_container_width=False):
                    st.markdown("""
**📌 O que este gráfico mostra**

Cada ponto é o desempenho de um híbrido **em relação à média de todos os híbridos** daquele local.

- **Eixo X** → produtividade média do local (o quanto o ambiente foi favorável).
- **Eixo Y** → desvio do híbrido: `sc/ha do híbrido − média do local`.

---

**📐 Como ler**

- **Ponto acima de zero** → produziu acima da média naquele local.
- **Ponto abaixo de zero** → ficou abaixo da média.
- **Reta subindo (b > 0)** → ganha vantagem em ambientes melhores: é **responsivo**.
- **Reta descendo (b < 0)** → perde vantagem quando o ambiente melhora, mas se segura nos adversos.
- **Reta horizontal (b ≈ 0)** → comportamento independente do ambiente.

O **R²** indica o quanto o ambiente explica o desvio: alto significa comportamento previsível ao
longo dos ambientes; baixo, que o híbrido reage de forma irregular.

---

**🔍 Diferença para a aba Por Local**

Lá as barras mostram a diferença **entre os dois híbridos**. Aqui cada um é medido contra o
**conjunto** de todos os híbridos do local — então dá para ver se ambos vão bem, ambos vão mal, ou
se um sobe enquanto o outro desce conforme o ambiente muda.

> O **p-valor** vem de um teste t pareado entre os dois híbridos nos locais em comum: abaixo de
0,05 indica que a diferença média entre eles dificilmente é acaso.
""")

                _COR_MAP = {"CHECK": "#E67E22", "STINE": "#2976B6", "EXP": "#009900", "DP2": "#7AAF6A"}
                _st1_t3 = (df_p1[df_p1["dePara"] == p1_t3]["status_material"].iloc[0]
                           if not df_p1[df_p1["dePara"] == p1_t3].empty else "")
                _st2_t3 = (ta_filtrado[ta_filtrado["dePara"] == p2_t3]["status_material"].iloc[0]
                           if not ta_filtrado[ta_filtrado["dePara"] == p2_t3].empty else "")
                _cor1 = _COR_MAP.get(_st1_t3, "#2976B6")
                _cor2 = _COR_MAP.get(_st2_t3, "#E67E22")
                if _cor1 == _cor2:                      # mesmo status: diferencia o adversário
                    _cor2 = "#8E44AD"

                _sc1 = _df1.set_index("cod_fazenda")["sc_ha"]
                _sc2 = _df2.set_index("cod_fazenda")["sc_ha"]
                _idx = _sc1.index.intersection(_sc2.index)
                _dif = _sc1[_idx] - _sc2[_idx]
                _dif_media = float(_dif.mean()) if len(_dif) else 0.0
                _n_vit_t3 = int((_dif > EMPATE_MARGEM).sum())
                _pct_vit_t3 = round(_n_vit_t3 / len(_idx) * 100, 1) if len(_idx) else 0

                _pval_str = ""
                try:
                    from scipy import stats as _stats
                    if len(_idx) >= 2:
                        _, _pv = _stats.ttest_rel(_sc1[_idx], _sc2[_idx])
                        _pval_str = f"p={_pv:.3f}"
                except Exception:
                    _pval_str = ""

                _sinal = "+" if _dif_media >= 0 else ""
                _sub = (f"{_pct_vit_t3:.0f}% de vitórias · diferença média {_sinal}{_dif_media:.1f} sc/ha"
                        + (f" · {_pval_str}" if _pval_str else "")
                        + f" · {len(_idx)} locais")

                st.markdown(
                    f'<p style="font-size:22px;font-weight:700;color:#1A1A1A;margin:0;">'
                    f'<span style="color:{_cor1};">{p1_t3}</span>'
                    f' <span style="font-size:16px;font-weight:400;color:#6B7280;">vs</span> '
                    f'<span style="color:{_cor2};">{p2_t3}</span></p>'
                    f'<p style="font-size:15px;color:#374151;margin:4px 0 14px;">{_sub}</p>',
                    unsafe_allow_html=True)

                _meta_loc = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                             .drop_duplicates("cod_fazenda").set_index("cod_fazenda"))

                fig_dev = go_plt.Figure()
                _retas = []
                for _d, _hib, _cor in [(_df1, p1_t3, _cor1), (_df2, p2_t3, _cor2)]:
                    _d = _d.join(_meta_loc, on="cod_fazenda", how="left")
                    _x = _d["media_local"].values
                    _y = _d["desvio"].values
                    fig_dev.add_trace(go_plt.Scatter(
                        x=_x, y=_y, mode="markers", name=_hib,
                        marker=dict(color=_cor, size=10, opacity=0.80,
                                    line=dict(color="#FFFFFF", width=1.2)),
                        customdata=_d[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]].values,
                        hovertemplate=(f"<b>{_hib}</b><br><b>%{{customdata[0]}}</b> — %{{customdata[1]}}<br>"
                                       "%{customdata[2]}, %{customdata[3]}<br>"
                                       "Média local: %{x:.1f} sc/ha<br>"
                                       "Desvio: %{y:+.1f} sc/ha<extra></extra>")))
                    if len(_x) >= 2:
                        try:
                            _Xr = np.column_stack([np.ones(len(_x)), _x])
                            _beta, _, _, _ = np.linalg.lstsq(_Xr, _y, rcond=None)
                            _xl = np.linspace(_x.min(), _x.max(), 100)
                            _yl = _beta[0] + _beta[1] * _xl
                            _ss_res = np.sum((_y - _Xr @ _beta) ** 2)
                            _ss_tot = np.sum((_y - _y.mean()) ** 2)
                            _r2 = 1 - _ss_res / _ss_tot if _ss_tot > 0 else np.nan
                            fig_dev.add_trace(go_plt.Scatter(
                                x=_xl, y=_yl, mode="lines", line=dict(color=_cor, width=2.5),
                                showlegend=False, hoverinfo="skip"))
                            _retas.append({"hib": _hib, "cor": _cor, "x_end": _xl[-1],
                                           "y_end": _yl[-1], "slope": _beta[1], "r2": _r2})
                        except Exception:
                            pass

                if len(_retas) == 2 and abs(_retas[0]["y_end"] - _retas[1]["y_end"]) < 4:
                    _ay = [-35, 20] if _retas[0]["y_end"] >= _retas[1]["y_end"] else [20, -35]
                else:
                    _ay = [10] * len(_retas)
                for _i, _r in enumerate(_retas):
                    fig_dev.add_annotation(
                        x=_r["x_end"], y=_r["y_end"],
                        text=f"<b>{_r['hib']}</b><br>b={_r['slope']:+.3f} · R²={_r['r2']:.2f}",
                        showarrow=True, arrowhead=0, arrowcolor=_r["cor"], arrowwidth=1.5,
                        ax=25, ay=_ay[_i], xanchor="left",
                        font=dict(size=12, color=_r["cor"], weight="bold"),
                        bgcolor="rgba(255,255,255,0.85)", bordercolor=_r["cor"],
                        borderwidth=1, borderpad=3)

                fig_dev.add_hline(y=0, line=dict(color="#444444", width=1.5, dash="dot"))
                fig_dev.update_layout(
                    height=520, plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                    font=dict(family="Helvetica Neue, sans-serif", color="#1A1A1A"),
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="left", x=0,
                                font=dict(size=13, color="#1A1A1A", weight="bold")),
                    xaxis=dict(title=dict(text="<b>Produtividade média do local (sc/ha)</b>",
                                          font=dict(size=14, color="#1A1A1A", weight="bold")),
                               tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                               showgrid=True, gridcolor="#E5E5E5", zeroline=False),
                    yaxis=dict(title=dict(text="<b>Desvio em relação à média do local (sc/ha)</b>",
                                          font=dict(size=14, color="#1A1A1A", weight="bold")),
                               tickfont=dict(size=12, color="#1A1A1A", weight="bold"),
                               showgrid=True, gridcolor="#E5E5E5", zeroline=False),
                    margin=dict(t=60, b=60, l=80, r=220))

                st.plotly_chart(fig_dev, use_container_width=True)
                st.caption("ℹ️ Desvio = sc/ha do híbrido − média de todos os híbridos no local · "
                           "b > 0 = ganha vantagem em ambientes favoráveis · b < 0 = perde vantagem "
                           "em ambientes melhores · b ≈ 0 = comportamento independente do ambiente.")

                st.markdown("**Desvios por local**")
                _tab3 = (_df1[["cod_fazenda", "media_local", "sc_ha", "desvio"]]
                         .merge(_df2[["cod_fazenda", "sc_ha", "desvio"]], on="cod_fazenda",
                                suffixes=("_1", "_2"))
                         .sort_values("media_local"))
                _tab3.columns = ["Local", "Média do local (sc/ha)", f"sc/ha — {p1_t3}",
                                 f"Desvio — {p1_t3}", f"sc/ha — {p2_t3}", f"Desvio — {p2_t3}"]
                _tab3 = _tab3.round(1)
                def _js_sinal3():
                    return JsCode("""
                    function(params) {
                        const v = Number(params.value);
                        if (isNaN(v)) return {'textAlign':'center'};
                        if (v > 0) return {'color':'#15803D','fontWeight':'800','textAlign':'center'};
                        if (v < 0) return {'color':'#B91C1C','fontWeight':'800','textAlign':'center'};
                        return {'textAlign':'center'};
                    }
                    """)
                ag_table_h2h(_tab3, height=min(520, 40 + 32 * min(len(_tab3), 14) + 20),
                             estilos_col={f"Desvio — {p1_t3}": _js_sinal3(),
                                          f"Desvio — {p2_t3}": _js_sinal3()})
                st.download_button("⬇️ Exportar Excel", data=to_excel(_tab3),
                                   file_name=f"h2h_desvios_{p1_t3}_vs_{p2_t3}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="dl_t3")


# ═══════════════════════════════════════════════════════════════════════════════
# TAB L — Todos os Materiais
# ═══════════════════════════════════════════════════════════════════════════════
with tabL:
    secao_titulo(
        "HEAD-TO-HEAD · TODOS OS MATERIAIS",
        "A linha inteira contra o campo, em quatro leituras",
        "Mesmo cálculo das outras abas, agregado para todos os materiais da linha de uma vez. "
        "Escolha a leitura abaixo.",
    )

    _st_disp = [s for s in STATUS_P1 if s in ta_raw["status_material"].unique()]
    _status_linha = st.multiselect(
        "Status na coluna Material", _st_disp,
        default=[s for s in ["STINE"] if s in _st_disp] or _st_disp[:1],
        key="tabL_status_linha",
        help="Quem entra como 'linha'. Do outro lado ficam os status marcados em Status do "
             "Adversário (Prod. 2) na barra lateral — por padrão, os comerciais (CHECK).")
    if not _status_linha:
        _status_linha = [s for s in ["STINE"] if s in _st_disp] or _st_disp[:1]
    _tup_linha = tuple(sorted(_status_linha))

    # ADVERSÁRIO: quem a lateral escolheu, não uma constante. Antes esta aba fixava CHECK e o
    # filtro "Status do Adversário (Prod. 2)" não a alterava — as outras abas mudavam e esta não,
    # e a mensagem de vazio ainda mandava conferir justamente esse filtro. `cruzar_linha` é
    # cacheada por argumento, então passar a seleção mantém o cache (uma entrada por combinação).
    _tup_adv = tuple(sorted(status_p2_sel))
    _cruz_all = cruzar_linha(ta_raw, _tup_linha, _tup_adv) if _tup_adv else pd.DataFrame()
    _ativos = set(ta_filtrado["cod_fazenda"].dropna().unique())
    _cruz = (_cruz_all[_cruz_all["cod_fazenda"].isin(_ativos)].copy()
             if not _cruz_all.empty else _cruz_all)

    if _cruz.empty:
        if not _tup_adv:
            st.info("Nenhum status marcado em **Status do Adversário (Prod. 2)** na barra lateral: "
                    "sem adversário não há confronto. Marque ao menos um.")
        else:
            st.info(f"Nenhum confronto entre a linha ({', '.join(_tup_linha)}) e "
                    f"{', '.join(_tup_adv)} nos locais do recorte atual. Os dois lados precisam "
                    f"ter sido avaliados no MESMO local para existir duelo.")
    else:
        # ── quem entra na tabela ────────────────────────────────────────────
        # Vazio = todos, mesma convenção dos filtros da barra lateral.
        _cruz_cheio = _cruz.copy()
        _mats_all = sorted(_cruz_cheio["material"].unique())
        _advs_all = sorted(_cruz_cheio["adversario"].unique())

        _cf1, _cf2 = st.columns(2)
        with _cf1:
            _sel_mat = st.multiselect(
                f"Materiais da linha ({len(_mats_all)})", _mats_all, key="tabL_mat",
                placeholder="Todos — desmarcado significa todos")
        with _cf2:
            _sel_adv = st.multiselect(
                f"Adversários ({len(_advs_all)})", _advs_all, key="tabL_adv",
                placeholder="Todos — desmarcado significa todos")
        _sel_mat = _sel_mat or _mats_all
        _sel_adv = _sel_adv or _advs_all
        _cruz = _cruz_cheio[_cruz_cheio["material"].isin(_sel_mat)
                            & _cruz_cheio["adversario"].isin(_sel_adv)]

        if _cruz.empty:
            st.warning("A combinação escolhida não tem nenhum confronto.")
            st.stop()

        # ── contexto do recorte ─────────────────────────────────────────────
        _n_loc = _cruz["cod_fazenda"].nunique()
        _n_adv = _cruz["adversario"].nunique()
        _n_mat = _cruz["material"].nunique()
        _ap_geral = (_cruz["res"] == "V").mean() * 100

        # Recortar adversário muda o denominador de todos os materiais: com os mais
        # duros de fora, o aproveitamento sobe sozinho. O aviso existe para que essa
        # escolha nunca fique implícita.
        if len(_sel_adv) < len(_advs_all):
            _ap_cheio = (_cruz_cheio[_cruz_cheio["material"].isin(_sel_mat)]["res"] == "V").mean() * 100
            _fora = [a for a in _advs_all if a not in _sel_adv]
            st.markdown(
                f'<div style="background:#FFF8E1;border:1px solid #F5D76E;'
                f'border-left:5px solid #D4A800;border-radius:8px;padding:12px 18px;'
                f'margin:6px 0 10px;">'
                f'<p style="margin:0;font-size:15px;line-height:1.6;color:#4A3B00;'
                f'font-family:\'Helvetica Neue\',Helvetica,Arial,sans-serif;">'
                f'<b>Universo de adversários reduzido:</b> {len(_sel_adv)} de {len(_advs_all)}. '
                f'Ficaram de fora {", ".join(_fora[:6])}{" ..." if len(_fora) > 6 else ""}. '
                f'Com todos, o aproveitamento seria <b>{_ap_cheio:.1f}%</b>; com esta seleção é '
                f'<b>{_ap_geral:.1f}%</b>. Números de recortes diferentes não se comparam entre si '
                f'nem com os das outras abas.</p></div>',
                unsafe_allow_html=True)
        elif len(_sel_mat) < len(_mats_all):
            st.caption(f"Mostrando {len(_sel_mat)} de {len(_mats_all)} materiais. O universo de "
                       "adversários está inteiro, então os percentuais continuam comparáveis com "
                       "as outras abas.")

        _conf_mat = _cruz.groupby("material").size()
        _conf_max = int(_conf_mat.max()) if len(_conf_mat) else 0
        _POUCO = 0.5 * _conf_max          # metade da cobertura do mais testado
        _magros = sorted(_conf_mat[_conf_mat < _POUCO].index)
        if _magros:
            st.caption(
                f"⚠️ Cobertura desigual: {', '.join(_magros[:8])}"
                f"{' ...' if len(_magros) > 8 else ''} "
                f"{'têm' if len(_magros) > 1 else 'tem'} menos da metade dos confrontos do "
                f"material mais testado ({_conf_max}). O percentual deles oscila muito — leia "
                "sempre junto com a coluna Confrontos, destacada em laranja.")

        _leitura = st.radio("Leitura", ["Placar", "Matriz", "Por ambiente", "Local a local"],
                            horizontal=True, key="tabL_leitura", label_visibility="collapsed")

        st.caption(f"Adversários: **{', '.join(_tup_adv)}** — quem está marcado em Status do "
                   f"Adversário (Prod. 2) na barra lateral. Locais, safra e demais filtros da "
                   f"lateral também valem aqui.")

        _COMUM = """
**📌 O que este painel responde**

> **A linha inteira ganha de quem, perde de quem, e em que tipo de ambiente isso muda?**

As outras abas olham um material por vez. Esta olha a linha como portfólio.

---

**🔢 O cálculo, comum às quatro leituras**

1. Em cada **local**, tira-se a média das parcelas de cada material.
2. Cada par (material da linha × adversário) que dividiu um local é **um confronto**. Local e
   safra formam a chave: o mesmo local em safras diferentes não é o mesmo ambiente.
3. `diferença = média do material − média do adversário`, naquele local.
4. Acima de **+1,0 sc/ha** é vitória · entre −1,0 e +1,0 é **empate** · abaixo é derrota.
5. **Aproveitamento = vitórias ÷ confrontos.** O empate **não vale meio ponto**: fica fora do
   numerador e dentro do denominador.
6. **Quem entra de cada lado vem dos filtros:** o seletor acima define a coluna Material e o
   **Status do Adversário (Prod. 2)**, na barra lateral, define o outro lado. O padrão é CHECK,
   o universo de disputa de mercado — mas se você marcar EXP ou DP2, o confronto passa a ser
   contra eles, aqui e nas outras abas.

> Por isso o total de confrontos é muito maior que o de locais: são materiais × adversários ×
locais em comum.

---

**🎚️ O que segue os filtros e o que não segue**

Tudo que é **resultado** segue: os locais do recorte, a safra, o status do adversário, e os dois
seletores de material e adversário desta aba. Se um número muda de valor conforme a seleção, ele
seguiu o filtro.

Duas coisas **não** seguem, de propósito, porque são **régua** e não resultado:

| Referência | Como é calculada | Por quê |
|---|---|---|
| **Média e amplitude do local** | todos os híbridos avaliados ali, inclusive os fora do filtro | é atributo do ambiente. Recalculada dentro do recorte, mudaria conforme quem você está olhando |
| **Faixas de classificação** (>75 · 56–75 · 46–55 · até 45) | limites fixos do painel | são o mesmo critério em todas as telas; virar quantil do recorte tornaria "Superior" relativo à seleção |

Os **terços de ambiente** da leitura "Por ambiente" são a exceção que confirma a regra: eles são
recortados **dentro** da seleção ativa, e é por isso que "Produtividade Baixa" significa o terço
inferior daquele recorte, não da rede.

---
"""

        _CUIDADOS = """
---

**⚠️ Cuidados válidos para as quatro leituras**

- **Aproveitamento não é produtividade.** Vencer 51 de 100 por 1,5 sc/ha e vencer 51 por 15 sc/ha
  dão o mesmo número — por isso a diferença média está sempre ao lado.
- **Recortes diferentes não se comparam.** Mudou o filtro, mudaram os adversários e os locais.
- **Os locais vêm da barra lateral**, os mesmos das outras quatro abas.
- **O universo de adversários também vem da lateral** (Status do Adversário, padrão CHECK) e vale
  igual em todas as abas. Trocar CHECK por EXP muda a pergunta: deixa de ser disputa de mercado e
  passa a ser comparação interna de pipeline. Registre qual estava marcado ao citar um número.
- **Média e amplitude do local são régua, não resultado**: saem da base inteira e não mudam com o
  filtro, de propósito. Aproveitamento, confrontos e diferença média são resultado, e mudam.
- **Os dois seletores acima da tabela são desta aba.** Escolher quais materiais aparecem é
  inofensivo — muda só quais linhas são desenhadas. **Escolher adversários muda o denominador de
  todos**: com os mais duros de fora, o aproveitamento sobe sozinho. Por isso aparece um aviso
  amarelo com o número que sairia se o universo estivesse inteiro.
- **Adversário com poucos locais em comum** entra no agregado, mas não deve ser citado pelo nome.
  A coluna de confrontos existe para isso.
"""

        _DIC = {
            "Placar": """
**📋 O que é cada coluna**

| Coluna | O que é | Como é calculado |
|---|---|---|
| **Material** | híbrido da linha | — |
| **Aprov. %** | quanto dos confrontos ele venceu | `V ÷ Confrontos × 100`. A barra é o próprio valor, de 0 a 100 |
| **V · E · D** | vitórias, empates e derrotas | contagem dos confrontos pela regra de ±1,0 sc/ha |
| **Confrontos** | quantos duelos entraram na conta | `adversários × locais em comum com cada um`. Some V+E+D e tem que bater |
| **Dif. média sc/ha** | por quanto ele ganha ou perde | média de todas as diferenças dos confrontos dele. Verde positivo, vermelho negativo |
| **Alta perf. · Superior · Competitivo · Restrito** | contra quantos **adversários** ele está em cada faixa | para cada adversário calcula-se o % de vitórias contra ele; a faixa é a do painel (>75 · 56–75 · 46–55 · até 45). São contagens de adversários, **não de confrontos** |

---

**🧭 Exemplo de uso**

Dois materiais aparecem com aproveitamento parecido, digamos 31% e 30%.

1. Olhe **V–E–D**: se um tem 40–9–79 e o outro 42–4–92, eles chegaram ao mesmo lugar por
   caminhos diferentes — o primeiro empata mais, o segundo perde mais.
2. Olhe **Dif. média**: se as duas forem próximas, a diferença entre eles é de consistência, não
   de patamar.
3. Olhe as **quatro últimas colunas**: um material com 2 adversários em Superior tem duelo
   favorável para vender; um com zero em Superior e 10 em Restrito não tem argumento de placar
   neste recorte — o argumento dele terá de vir de ambiente, e a leitura seguinte mostra isso.
""",
            "Matriz": """
**📋 O que é cada coluna**

| Coluna | O que é | Como é calculado |
|---|---|---|
| **Material** | híbrido da linha, do melhor para o pior aproveitamento geral | — |
| **Uma coluna por adversário** | % de vitórias daquele material contra aquele adversário | `vitórias ÷ locais em comum do par × 100`. Célula vazia = nunca dividiram local |
| **APROVEITAMENTO DA LINHA** (penúltima linha) | quanto a linha inteira venceu daquele adversário | soma das vitórias de **todos** os materiais contra ele ÷ soma dos confrontos |
| **DIF. MÉDIA sc/ha** (última linha) | por quanto a linha ganha ou perde dele | média das diferenças de todos os duelos contra aquele adversário |

A cor é a classificação do painel: verde acima de 75 · azul 56 a 75 · amarelo 46 a 55 ·
vermelho até 45.

**A ordenação é o argumento.** Colunas do adversário mais duro para o mais fácil, medido pelo
aproveitamento da linha inteira. Reordenar por fabricante destrói a leitura.

---

**🧭 Exemplo de uso**

1. **Por coluna**: se as três primeiras colunas forem vermelhas de cima a baixo, esses são os
   concorrentes que a linha não enfrenta neste recorte. Contra eles não existe discurso de
   produtividade — existe discurso de atributo. É informação de portfólio, e não aparece em
   nenhuma outra tela.
2. **Por linha**: procure o ponto em que a linha de cada material vira de vermelho para verde.
   Quem vira cedo tem portfólio amplo; quem não vira em nenhuma coluna é material de nicho, e o
   nicho terá de ser definido por outro critério que não o placar.
3. **Antes de citar um duelo pelo nome**, exporte o "par a par" e confira o nº de confrontos
   daquela célula.
""",
            "Por ambiente": """
**📋 O que é cada coluna**

| Coluna | O que é | Como é calculado |
|---|---|---|
| **Material** | híbrido da linha · a última linha é a **linha inteira** | — |
| **Geral %** | aproveitamento em todos os locais do recorte | `vitórias ÷ confrontos` |
| **Produtividade Baixa · Média · Alta** | aproveitamento só nos locais daquele terço | os locais em tela são divididos em **três grupos de tamanho igual** pela média do local (`qcut` em 3). "Baixa" é o terço inferior **deste recorte**, não da rede |
| **Plantio Cedo · Meio · Tardio** | aproveitamento só nos locais daquele terço de época | mesma divisão em três, pela data de semeadura do ensaio. Só aparece se a base tiver a data |
| **(NL)** no cabeçalho | quantos locais entraram naquele terço | — |

**Os dois cortes são independentes e não somam.** Um local pode ser de produtividade baixa e de
plantio tardio ao mesmo tempo — ele aparece nas duas colunas, e somá-las conta o local duas vezes.

---

**🧭 Exemplo de uso**

Um material tem aproveitamento geral baixo e você precisa decidir se ele tem argumento.

1. Se o número **sobe** da Produtividade Alta para a Baixa, ele é material de ambiente
   restritivo: perde onde sobra tudo e ganha onde falta.
2. Se **sobe** de Plantio Cedo para Tardio, ele é material de janela apertada.
3. Se subir nos dois, confira se são os mesmos locais — exporte "quais locais entraram em cada
   terço". Se forem os mesmos, é **um** achado, não dois.
4. A afirmação sobre a marca sai da última linha, não do material isolado.
""",
            "Local a local": """
**📋 O que é cada coluna**

| Coluna | O que é | Como é calculado |
|---|---|---|
| **Material** | híbrido da linha | — |
| **Uma coluna por local** | `vitórias/confrontos` daquele material naquele local | confrontos = quantos adversários ele enfrentou ali. Célula vazia = ele não foi avaliado no local |
| **MÉDIA DO LOCAL sc/ha** | quanto o ambiente rendeu | média de **todas** as parcelas de **todos** os híbridos do local |
| **APROV. DA LINHA %** | quanto a linha inteira venceu ali | vitórias de todos os materiais ÷ confrontos daquele local |

Colunas ordenadas da **menor** para a **maior** média do local. Verde a partir de 70% de vitórias
no local, vermelho até 20%.

---

**🧭 Exemplo de uso**

1. **Leia a tendência, não a célula.** Verde à esquerda e vermelho à direita é material de
   ambiente restritivo; o contrário é material de alto investimento; disperso é material cujo
   desempenho não é explicado pela produtividade do local — procure em época, população ou
   sanidade.
2. Um `0/10` num local não é erro nem falha do material: é um ambiente em que ele não venceu
   nenhum dos dez adversários presentes. Cruze com a média do local antes de concluir.
3. Compare duas linhas em busca de **inversão**: dois materiais que ganham em locais opostos são
   complementares no portfólio, e essa é uma recomendação diferente de "um é melhor que o outro".
""",
        }

        _c1, _c2 = st.columns([3, 1])
        with _c1:
            st.caption(
                f"{_n_mat} materiais da linha · {_n_adv} adversários · {_n_loc} locais · "
                f"{len(_cruz)} confrontos · aproveitamento geral {_ap_geral:.1f}% · "
                f"empate = ±{EMPATE_MARGEM:.1f} sc/ha"
            )
        with _c2:
            with st.popover(f"ℹ️ Como entender · {_leitura}", use_container_width=True):
                st.markdown(_COMUM + _DIC[_leitura] + _CUIDADOS)


        VERDE, VERM, AZUL, CINZA = "#1E7A34", "#C0201E", "#2976B6", "#9AA5B1"
        _LEG_CLASSE = [("#90EE90", "Alta Performance · acima de 75%", "bg"),
                       ("#87CEFF", "Superior · 56 a 75%", "bg"),
                       ("#FFFF00", "Competitivo · 46 a 55%", "bg"),
                       ("#FF0000", "Restrito · até 45%", "bg")]
        BG_CLASSE = [(75, "#90EE90", "#1A1A1A"), (55, "#87CEFF", "#1A1A1A"),
                     (45, "#FFFF00", "#1A1A1A"), (-1, "#FF0000", "#FFFFFF")]

        def _cel_classe(v, tipo="pct0"):
            """Célula com fundo pela classificação do painel."""
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return cel(None, tipo, cor=CINZA, align="center")
            for corte, bg, fg in BG_CLASSE:
                if v > corte:
                    return cel(v, tipo, cor=fg, bg=bg, bold=True, align="center")
            return cel(v, tipo, align="center")

        _ordem_mat = (agregar_placar(_cruz, "material").sort_values("Aprov", ascending=False)
                      ["material"].tolist())
        _col_adv = agregar_placar(_cruz, "adversario").sort_values("Aprov")

        # ── PLACAR ───────────────────────────────────────────────────────────
        if _leitura == "Placar":
            _p = agregar_placar(_cruz, "material").set_index("material")
            _pa = agregar_placar(_cruz, ["material", "adversario"])
            _pa["classe"] = _pa["Aprov"].apply(lambda v: classificar_h2h(v)[0])
            _cl = _pa.pivot_table(index="material", columns="classe", values="adversario",
                                  aggfunc="count").fillna(0).astype(int)
            for _k in ["Alta Performance", "Superior", "Competitivo", "Restrito"]:
                if _k not in _cl.columns:
                    _cl[_k] = 0

            st.caption(
                "Ordenado do melhor para o pior aproveitamento. "
                f"**As quatro últimas colunas contam concorrentes, não confrontos:** dizem contra "
                f"quantos dos {_n_adv} concorrentes do recorte o material está em cada faixa. "
                f"Somadas, dão quantos concorrentes ele chegou a enfrentar — no máximo {_n_adv}, "
                "menos que isso se não dividiu local com algum deles.")
            # V, E e D com as mesmas cores da Análise por Local. No cabeçalho vai a
            # cor cheia; no número, a variante legível — COR_EMPATE é amarelo puro e
            # não se lê como texto, por isso existe COR_EMPATE_CARD.
            _TXT_V, _TXT_E, _TXT_D = COR_VITORIA, "#D4A800", COR_DERROTA
            _headers = ["Material", "Aprov.",
                        hdr("V", COR_VITORIA, "#FFFFFF"),
                        hdr("E", COR_EMPATE, "#1A1A1A"),
                        hdr("D", COR_DERROTA, "#FFFFFF"),
                        "Confrontos", "Dif. média sc/ha",
                        hdr("Alta perf.", "#90EE90", "#1A1A1A"),
                        hdr("Superior", "#87CEFF", "#1A1A1A"),
                        hdr("Competitivo", "#FFFF00", "#1A1A1A"),
                        hdr("Restrito", "#FF0000", "#FFFFFF")]
            _linhas = []
            for m in _ordem_mat:
                r = _p.loc[m]
                _linhas.append([
                    cel(m, bold=True),
                    cel(r["Aprov"], "pct1", barra=r["Aprov"],
                        cor=AZUL if r["Aprov"] >= 50 else CINZA),
                    cel(r["V"], "num0", align="center", cor=_TXT_V, bold=True),
                    cel(r["E"], "num0", align="center", cor=_TXT_E),
                    cel(r["D"], "num0", align="center", cor=_TXT_D),
                    cel(r["Confrontos"], "num0", align="center",
                        cor="#E06C00" if r["Confrontos"] < _POUCO else "#1A1A1A",
                        bold=r["Confrontos"] < _POUCO),
                    cel(r["Dif"], "sinal1", cor=VERDE if r["Dif"] > 0 else VERM, bold=True),
                    cel(_cl.loc[m, "Alta Performance"], "num0", align="center"),
                    cel(_cl.loc[m, "Superior"], "num0", align="center"),
                    cel(_cl.loc[m, "Competitivo"], "num0", align="center"),
                    cel(_cl.loc[m, "Restrito"], "num0", align="center"),
                ])
            render_tabela(_headers, _linhas, "h2h_placar_linha", "exp_placar", largura_1a=200,
                          legenda=[(AZUL, "aproveitamento de 50% ou mais", "barra"),
                                   (CINZA, "abaixo de 50%", "barra"),
                                   (VERDE, "diferença média positiva", "txt"),
                                   (VERM, "diferença média negativa", "txt"),
                                   ("#E06C00", "menos da metade dos confrontos do mais "
                                    "testado — percentual instável", "txt")])
            st.caption("As quatro últimas colunas usam as mesmas cores da Tabela de "
                       "Classificação e da Matriz: são as faixas do painel, aplicadas ao "
                       "cabeçalho porque a coluna inteira é daquela classe.")

        # ── MATRIZ ───────────────────────────────────────────────────────────
        elif _leitura == "Matriz":
            _ordem_adv = _col_adv["adversario"].tolist()
            _pa = agregar_placar(_cruz, ["material", "adversario"])
            _mx = _pa.pivot(index="material", columns="adversario", values="Aprov")
            _nx = _pa.pivot(index="material", columns="adversario", values="Confrontos")
            _ca = _col_adv.set_index("adversario")
            _min_n = int(np.nanmin(_nx.values)) if _nx.size else 0

            st.caption(
                f"% de vitórias. Colunas do adversário **mais duro** (esquerda) para o **mais "
                f"fácil** (direita); linhas do melhor para o pior aproveitamento. "
                f"Menor nº de locais em comum de um par: {_min_n}. "
                "Cor pela classificação do painel — verde acima de 75%, azul 56 a 75%, "
                "amarelo 46 a 55%, vermelho até 45%."
            )
            _headers = ["Material"] + _ordem_adv
            _linhas = []
            for m in _ordem_mat:
                _linhas.append([cel(m, bold=True)] + [
                    _cel_classe(_mx.loc[m, a] if (m in _mx.index and a in _mx.columns) else np.nan)
                    for a in _ordem_adv])
            _linhas.append([cel_resumo("APROVEITAMENTO DA LINHA")] +
                           [cel_resumo(float(_ca.loc[a, "Aprov"]), "pct0") for a in _ordem_adv])
            _linhas.append([cel_resumo("DIF. MÉDIA sc/ha")] +
                           [cel_resumo(float(_ca.loc[a, "Dif"]), "sinal1") for a in _ordem_adv])
            render_tabela(_headers, _linhas, "h2h_matriz", "exp_matriz", largura_1a=210,
                          legenda=_LEG_CLASSE + [(BG_RESUMO, "linha de resumo — a linha inteira "
                                                  "contra aquele adversário", "bg")])
            st.caption("As duas últimas linhas são a leitura por coluna: a linha inteira contra "
                       "aquele adversário. Coluna vermelha de ponta a ponta é adversário que a "
                       "linha não enfrenta neste recorte.")
            _exp_pares = (_pa.rename(columns={"material": "Material", "adversario": "Adversário",
                                              "Aprov": "Aprov. %", "Dif": "Dif. média sc/ha"})
                          .round(1))
            botao_exportar(_exp_pares, "h2h_pares_detalhe", "exp_pares",
                           "⬇️ Exportar par a par (com nº de confrontos)",
                           cols_pct=["Aprov. %"])

        # ── POR AMBIENTE ─────────────────────────────────────────────────────
        elif _leitura == "Por ambiente":
            _mloc = (ta_filtrado.dropna(subset=["sc_ha"]).groupby("cod_fazenda")["sc_ha"].mean())
            _mloc = _mloc[_mloc.index.isin(_cruz["cod_fazenda"].unique())]
            _recortes = {}
            if len(_mloc) >= 6:
                _q = pd.qcut(_mloc, 3, labels=["Baixa", "Média", "Alta"])
                for _r in ["Baixa", "Média", "Alta"]:
                    _recortes[f"Produtividade {_r}"] = set(_mloc.index[_q == _r])
            _q_prod = _q if len(_mloc) >= 6 else None
            _q_ep, _dtl = None, None
            _cpL = _col_plantio(ta_filtrado)
            if _cpL:
                _dtl = (ta_filtrado[["cod_fazenda", _cpL]].drop_duplicates("cod_fazenda")
                        .set_index("cod_fazenda")[_cpL])
                _dtl = pd.to_datetime(_dtl, errors="coerce").dropna()
                _dtl = _dtl[_dtl.index.isin(_cruz["cod_fazenda"].unique())]
                if len(_dtl) >= 6:
                    _q_ep = pd.qcut(_dtl.rank(method="first"), 3, labels=["Cedo", "Meio", "Tardio"])
                    for _r in ["Cedo", "Meio", "Tardio"]:
                        _recortes[f"Plantio {_r}"] = set(_dtl.index[_q_ep == _r])

            if not _recortes:
                st.info("Recorte pequeno demais para dividir em terços — são necessários ao menos "
                        "6 locais. Use o filtro Fazenda da barra lateral.")
            else:
                st.caption(
                    "Aproveitamento em cada terço do recorte. Os terços são calculados **sobre os "
                    "locais em tela** — 'Baixa' é o terço inferior deste recorte, não da rede. "
                    "(NL) é o nº de locais do terço. Um mesmo local pode estar num terço de "
                    "produtividade e num de época: são cortes independentes, não somam."
                )
                _headers = ["Material", "Geral"] + [f"{k} ({len(v)}L)" for k, v in _recortes.items()]
                _linhas = []
                _rot_total = ("LINHA INTEIRA" if _tup_linha == ("STINE",)
                              else "TOTAL · " + " + ".join(_tup_linha))
                for m in _ordem_mat + [_rot_total]:
                    _resumo = (m == _rot_total)
                    _d = _cruz if _resumo else _cruz[_cruz["material"] == m]
                    _fabr = cel_resumo if _resumo else None
                    linha = [cel_resumo(m) if _resumo else cel(m, bold=True),
                             cel_resumo((_d["res"] == "V").mean() * 100, "pct1")]
                    for _locs in _recortes.values():
                        _s = _d[_d["cod_fazenda"].isin(_locs)]
                        _v = (_s["res"] == "V").mean() * 100 if len(_s) else np.nan
                        linha.append(_fabr(_v, "pct1") if _fabr else _cel_classe(_v, "pct1"))
                    _linhas.append(linha)
                render_tabela(_headers, _linhas, "h2h_por_ambiente", "exp_amb", largura_1a=210,
                              legenda=_LEG_CLASSE + [(BG_RESUMO, "coluna Geral e linha da marca — "
                                                      "resumo, não comparável célula a célula", "bg")])
                st.caption("Leia na horizontal: material cujo aproveitamento sobe da esquerda para "
                           "a direita responde ao ambiente. A última linha agrega os materiais "
                           "selecionados — só é afirmação sobre a marca quando a seleção é a "
                           "linha comercial.")
                if not _cpL:
                    st.info("**Terços de época indisponíveis:** a analítica não trouxe "
                            "`dataPlantioMilho`. Só os terços de produtividade estão na tabela.")
                elif _q_ep is None:
                    st.caption("Menos de 6 locais com data de plantio válida: os terços de época "
                               "não foram calculados.")

                with st.expander("Quais locais entraram em cada terço, e por quê", expanded=False):
                    st.markdown(
                        "Os terços são calculados **sobre os locais em tela**, dividindo-os em "
                        "três grupos de tamanho igual — por média do local (produtividade) e "
                        "pela ordem da data de semeadura (época). Não há corte fixo: mudou o "
                        "filtro lateral, mudam os grupos. Por isso a mesma fazenda pode ser "
                        "'Alta' num recorte de MT e 'Média' num recorte da rede inteira.")
                    _fz = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome"]]
                           .drop_duplicates("cod_fazenda").set_index("cod_fazenda"))
                    _ordem_det = list(_mloc.sort_values().index)
                    _hd = ["Local", "Fazenda", "Cidade", "Média sc/ha", "Terço de produtividade"]
                    if _q_ep is not None:
                        _hd += ["Plantio", "Terço de época"]
                    _ld = []
                    for _l in _ordem_det:
                        _tp = str(_q_prod[_l]) if _q_prod is not None else "—"
                        _cor_tp = {"Baixa": VERM, "Alta": VERDE}.get(_tp, "#1A1A1A")
                        linha = [cel(_l, bold=True),
                                 cel(_fz["nomeFazenda"].get(_l, "")),
                                 cel(_fz["cidade_nome"].get(_l, "")),
                                 cel(float(_mloc[_l]), "num1"),
                                 cel(_tp, cor=_cor_tp, bold=_tp in ("Baixa", "Alta"),
                                     align="center")]
                        if _q_ep is not None:
                            _te = str(_q_ep[_l]) if _l in _q_ep.index else "—"
                            _cor_te = {"Tardio": VERM, "Cedo": VERDE}.get(_te, "#1A1A1A")
                            linha += [cel(_dtl[_l].strftime("%d/%m") if _l in _dtl.index else "",
                                          cor=_cor_te, align="center"),
                                      cel(_te, cor=_cor_te, bold=_te in ("Cedo", "Tardio"),
                                          align="center")]
                        _ld.append(linha)
                    render_tabela(_hd, _ld, "h2h_ambiente_locais", "exp_amb_loc",
                                  largura_1a=150, altura_max=420,
                                  legenda=[(VERDE, "terço mais produtivo · plantio mais cedo", "txt"),
                                           (VERM, "terço menos produtivo · plantio mais tardio", "txt")])
                    if _q_ep is not None and _q_prod is not None:
                        _dois = [l for l in _ordem_det
                                 if str(_q_prod[l]) == "Baixa" and l in _q_ep.index
                                 and str(_q_ep[l]) == "Tardio"]
                        if _dois:
                            st.warning(
                                f"{len(_dois)} local(is) está(ão) nos dois recortes ao mesmo "
                                f"tempo — produtividade baixa E plantio tardio: "
                                f"{', '.join(_dois)}. O ganho que aparece numa coluna é em parte "
                                f"o mesmo da outra. Não somar as duas leituras.")

        # ── LOCAL A LOCAL ────────────────────────────────────────────────────
        else:
            _mloc = (ta_filtrado.dropna(subset=["sc_ha"]).groupby("cod_fazenda")["sc_ha"].mean()
                     .sort_values())
            _locs = [c for c in _mloc.index if c in set(_cruz["cod_fazenda"])]
            _v = _cruz.pivot_table(index="material", columns="cod_fazenda", values="res",
                                   aggfunc=lambda s: (s == "V").sum())
            _n = _cruz.pivot_table(index="material", columns="cod_fazenda", values="res",
                                   aggfunc="size")
            st.caption(
                "Vitórias sobre confrontos em cada local, ordenado da **menor** para a **maior** "
                "média do local. A leitura é a tendência da esquerda para a direita, não a célula: "
                "verde à esquerda e vermelho à direita é material de ambiente restritivo."
            )
            _headers = ["Material"] + _locs
            _linhas = []
            for m in _ordem_mat:
                linha = [cel(m, bold=True)]
                for c in _locs:
                    if m in _v.index and c in _v.columns and pd.notna(_n.loc[m, c]):
                        vv, nn = int(_v.loc[m, c]), int(_n.loc[m, c])
                        f = vv / nn
                        cor = VERDE if f >= 0.70 else (VERM if f <= 0.20 else "#1A1A1A")
                        linha.append(cel(f"{vv}/{nn}", cor=cor, bold=f >= 0.70 or f <= 0.20,
                                         align="center"))
                    else:
                        linha.append(cel("", align="center"))
                _linhas.append(linha)
            _linhas.append([cel_resumo("MÉDIA DO LOCAL sc/ha")] +
                           [cel_resumo(float(_mloc[c]), "num1") for c in _locs])
            _linhas.append([cel_resumo("APROV. DA LINHA")] +
                           [cel_resumo((_cruz[_cruz["cod_fazenda"] == c]["res"] == "V").mean() * 100,
                                       "pct0") for c in _locs])
            render_tabela(_headers, _linhas, "h2h_local_a_local", "exp_local", largura_1a=210,
                          legenda=[(VERDE, "venceu 70% ou mais dos confrontos do local", "txt"),
                                   (VERM, "venceu 20% ou menos", "txt"),
                                   (BG_RESUMO, "linhas de resumo — contexto do local, não "
                                    "desempenho de material", "bg")])
            st.caption("As duas últimas linhas dão o contexto do local: quanto rendeu e quanto a "
                       "linha inteira venceu ali. Célula vazia = os dois não dividiram aquele local.")

        # ── conferência entre abas ──────────────────────────────────────────
        st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
        with st.expander("Conferência: os números desta aba batem com os das outras?",
                         expanded=False):
            st.markdown(
                "Cada aba chega ao placar por um caminho de código diferente: a **Tabela de "
                "Classificação** usa `cruzar_por_local` a partir de um Produto 1 por vez, a aba "
                "**Locais** usa `caracterizar_locais` agregando por local, e esta usa "
                "`cruzar_linha` de uma vez só. Os três deveriam dar o mesmo número no mesmo "
                "recorte — a conferência abaixo recalcula pelos três caminhos e compara.\n\n"
                "> Isto verifica **coerência interna do painel**, não a correção do dado. Se a "
                "base estiver errada, os três erram junto.")

            _locais_tela = set(_cruz["cod_fazenda"].unique())
            _res_conf, _tudo_ok = [], True

            for _m in _sel_mat:
                _a = _cruz[_cruz["material"] == _m]
                _va = int((_a["res"] == "V").sum())
                _na = len(_a)
                # caminho da Tabela de Classificação, replicado
                _p1f = ta_raw[(ta_raw["status_material"].isin(STATUS_P1)) & (ta_raw["dePara"] == _m)]
                _p2f = ta_raw[ta_raw["dePara"] != _m]
                _cr = cruzar_por_local(_p1f, _p2f)
                if not _cr.empty:
                    _cr = _cr[_cr["cod_fazenda"].isin(_locais_tela)
                              & _cr["status_material_2"].isin(_tup_adv)
                              & _cr["dePara_2"].isin(_sel_adv)]
                    _d = _cr["sc_ha_1"] - _cr["sc_ha_2"]
                    _vb, _nb = int((_d > EMPATE_MARGEM).sum()), len(_cr)
                else:
                    _vb = _nb = 0
                _ok = (_va == _vb) and (_na == _nb)
                _tudo_ok = _tudo_ok and _ok
                _res_conf.append((_m, _va, _na, _vb, _nb, _ok))

            _hdc = ["Material", "Vitórias aqui", "Confrontos aqui",
                    "Vitórias na Tab. Classificação", "Confrontos lá", "Situação"]
            _ldc = []
            for _m, _va, _na, _vb, _nb, _ok in _res_conf:
                _c = VERDE if _ok else VERM
                _ldc.append([cel(_m, bold=True),
                             cel(_va, "num0", align="center"), cel(_na, "num0", align="center"),
                             cel(_vb, "num0", align="center"), cel(_nb, "num0", align="center"),
                             cel("confere" if _ok else "DIFERE", cor=_c, bold=True,
                                 align="center")])

            # linha inteira pelo caminho da aba Locais
            # mesmo adversário da tela: com a constante, a auditoria acusaria divergência
            # sempre que a lateral tivesse algo diferente de CHECK marcado
            _carc = caracterizar_locais(ta_raw, _tup_linha, _tup_adv)
            _carc = _carc[_carc["cod_fazenda"].isin(_locais_tela)]
            _v_loc, _n_loc_conf = int(_carc["_vit"].fillna(0).sum()), int(_carc["_conf"].fillna(0).sum())
            _v_aqui, _n_aqui = int((_cruz["res"] == "V").sum()), len(_cruz)
            _comparavel = ((len(_sel_mat) == len(_mats_all)) and (len(_sel_adv) == len(_advs_all))
                           and _tup_linha == ("STINE",))
            _ok_linha = _comparavel and (_v_loc == _v_aqui) and (_n_loc_conf == _n_aqui)
            _ldc.append([cel_resumo("TOTAL · vs aba Locais"),
                         cel_resumo(_v_aqui, "num0"), cel_resumo(_n_aqui, "num0"),
                         cel_resumo(_v_loc, "num0"), cel_resumo(_n_loc_conf, "num0"),
                         cel_resumo("confere" if _ok_linha else
                                    ("seleção parcial" if not _comparavel else "DIFERE"))])
            render_tabela(_hdc, _ldc, "h2h_conferencia", "exp_conf", largura_1a=250,
                          altura_max=380,
                          legenda=[(VERDE, "os dois caminhos deram o mesmo número", "txt"),
                                   (VERM, "divergência — investigar antes de usar", "txt")])

            if not _comparavel:
                st.caption("O total só é comparável com a aba Locais quando os dois seletores "
                           "estão vazios: lá o cálculo é sempre com todos os materiais e todos "
                           "os adversários do recorte.")
            if _tup_linha != ("STINE",):
                st.caption("A aba Locais calcula o placar sempre com status STINE. Com outro "
                           "status selecionado aqui, a linha de total compara recortes diferentes "
                           "de propósito — o que vale conferir é a parte de cima, material a "
                           "material.")
            if _tudo_ok and (_ok_linha or not _comparavel):
                st.success("Todos os caminhos conferem neste recorte.")
            else:
                st.error("Há divergência entre caminhos de cálculo. Não use os números até "
                         "entender a causa — comece conferindo se o filtro Status do Adversário "
                         "e o recorte de locais são os mesmos nas duas abas.")



rodape()
