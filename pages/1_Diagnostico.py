"""
pages/1_Diagnostico.py — Diagnóstico de dados do painel de milho

Estrutura:
  1. Status de carregamento das duas safras
  2. Visão consolidada (retrato geral)
  3. Integridade estrutural — problemas que apontam para o pipeline ou a base
     (órfãos, sem responsável, duplicados, sem de-para, sem tipo de ensaio,
      sem cidade/região, de-para sem status, plot sem avaliação)
  4. [próximo] Checagem por avaliação (av1–av4)

A integridade estrutural vem ANTES das avaliações de propósito: valida se a
fundação dos dados está sã antes de olhar a qualidade da coleta.
"""
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_2024, carregar_2025, aplicar_mestre, _mapas_mestre
from pipeline_milho_2025 import (NOTAS_AV1, RENAME_AV1, DOENCAS_AV2, RENAME_AV2,
                                 FLOR_MIN, FLOR_MAX, UMID_MIN, UMID_MAX, PROD_TETO)


st.set_page_config(
    page_title="Diagnóstico · JAUM DTC",
    page_icon="🔄",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()


# ── Tabela de detalhe: st.dataframe nativo (renderiza sempre dentro de expander) ──
def ag_table(df, height=340, key=None):
    """Usa st.dataframe nativo em vez do AgGrid — o AgGrid ficava em branco dentro de
    expander. Ordenação e busca são nativas do Streamlit; header estilizado via config."""
    st.dataframe(df, use_container_width=True, hide_index=True, height=height)


def exportar_excel(df, nome_arquivo="registros.xlsx", label="⬇️ Exportar Excel", key=None):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
        cell.fill = PatternFill("solid", start_color="F2F2F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[1].height = 26
    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
            cell.border = border
    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


# Rótulos amigáveis para as colunas técnicas (usados nas tabelas de detalhe e no WhatsApp)
ROTULOS = {
    "safra": "Safra", "cod_fazenda": "Cód. Local", "nomeFazenda": "Fazenda",
    "cidade_nome": "Cidade", "estado_sigla": "UF", "regiao_macro": "Região Macro",
    "regiao_micro": "Região Micro", "nomeResponsavel": "Responsável",
    "dePara": "Híbrido", "status_material": "Status", "tipoTeste": "Tipo de ensaio",
    "indexTratamento": "Trat.", "pop_tratamento": "População-alvo", "nome": "Nome (banco)",
    "regional": "Regional", "_motivo": "Motivo", "_problema": "Problema",
    "pesoParcela": "Peso parcela (kg)", "umidade_pct": "Umidade (%)",
    "area_parcela_m2": "Área parcela (m²)", "plantas_10m_media": "Estande (plantas)",
    "espacamento": "Espaçamento (m)", "produtividade_kg_ha": "Produtividade (kg/ha)",
    "pmg_corrigido_g": "PMG (g)", "fileiras_media": "Fileiras", "graos_fileira_media": "Grãos/fileira",
    "flags_produtividade": "Problemas",
    "altura_planta_cm": "Altura planta (cm)", "altura_espiga_cm": "Altura espiga (cm)",
    "dias_flor_masculino": "Flor. masculino (dias)", "dias_flor_feminino": "Flor. feminino (dias)",
}


def _rotular(df):
    """Renomeia as colunas técnicas para rótulos amigáveis (dePara → Material, etc.)."""
    return df.rename(columns={c: ROTULOS.get(c, c) for c in df.columns})


def _sanitizar(df):
    """Prepara um DataFrame para a tabela renderizar sem falhar (bug da tabela em branco).
    Corrige as causas comuns: índice não-contíguo, tipos nullable (Int64/boolean), datetimes,
    e NaN em colunas de texto. Aplicado de forma central a TODAS as tabelas de detalhe."""
    d = df.copy().reset_index(drop=True)
    for col in d.columns:
        dt = str(d[col].dtype)
        if dt in ("Int64", "Int32", "Int16", "boolean", "Float64"):
            # tipos nullable do pandas → object com string, NaN vira "—"
            d[col] = d[col].apply(lambda x: "—" if pd.isna(x) else (
                str(int(x)) if float(x).is_integer() else str(x)) if isinstance(x, (int, float)) else x)
        elif "datetime" in dt:
            d[col] = d[col].astype(str).replace({"NaT": "", "nan": "", "NaN": ""})
        elif dt == "object":
            # object pode ter NaN/None/NaN-string misturado — normaliza para texto limpo
            d[col] = d[col].apply(
                lambda x: "" if (x is None or (isinstance(x, float) and pd.isna(x))
                                 or (isinstance(x, str) and x.lower() == "nan")) else x)
        else:
            # qualquer outro tipo residual (category, etc.) → string
            try:
                if d[col].isna().any():
                    d[col] = d[col].apply(lambda x: "" if pd.isna(x) else x)
            except (TypeError, ValueError):
                pass
    return d


# CSS: tabelas nativas e captions mais legíveis (padrão da soja)
st.markdown("""
<style>
[data-testid="stDataFrame"] td, [data-testid="stDataFrame"] th,
[data-testid="stDataFrame"] [role="columnheader"] span {
    font-size: 13px !important; font-weight: 600 !important; color: #000000 !important;
}
[data-testid="stCaptionContainer"] p { color: #374151 !important; }
</style>
""", unsafe_allow_html=True)

page_header("Diagnóstico de Dados",
            "Status de carregamento e checagem de integridade antes de usar as análises.",
            imagem="App development-bro.png")

# ── Botão de atualização ──────────────────────────────────────────────────────
_c1, _c2 = st.columns([1, 5])
with _c1:
    if st.button("🔄 Recarregar dados", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 1 — STATUS DE CARREGAMENTO
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Dados", "Status de carregamento", "Verifique se as duas safras foram carregadas.")

with st.spinner("Carregando as duas safras..."):
    d24 = carregar_2024()
    d25 = carregar_2025()

def _status_card(safra, d):
    if d.get("ok"):
        return (f"<div style='background:#F4FBF6;border:1px solid #CDEBD8;border-radius:8px;"
                f"padding:14px 18px;'><span style='font-size:15px;font-weight:700;color:#1A1A1A;'>"
                f"Safra {safra}</span><br><span style='color:#1E8449;font-weight:600;font-size:13px;'>"
                f"✓ carregada</span></div>")
    return (f"<div style='background:#FDF3F3;border:1px solid #F3C9C9;border-radius:8px;"
            f"padding:14px 18px;'><span style='font-size:15px;font-weight:700;color:#1A1A1A;'>"
            f"Safra {safra}</span><br><span style='color:#B02A37;font-weight:600;font-size:13px;'>"
            f"✗ falha ao carregar</span></div>")

col1, col2 = st.columns(2)
with col1:
    st.markdown(_status_card("2024/25", d24), unsafe_allow_html=True)
    if not d24.get("ok") and d24.get("traceback"):
        with st.expander("Detalhe técnico — 2024/25"):
            st.code(d24["traceback"], language="text")
with col2:
    st.markdown(_status_card("2025/26", d25), unsafe_allow_html=True)
    if not d25.get("ok") and d25.get("traceback"):
        with st.expander("Detalhe técnico — 2025/26"):
            st.code(d25["traceback"], language="text")

# Escolha da safra a diagnosticar (integridade é por safra)
safras_ok = [(s, d) for s, d in [("2024/25", d24), ("2025/26", d25)] if d.get("ok")]
if not safras_ok:
    st.warning("Nenhuma safra carregada — não há o que diagnosticar.")
    st.stop()

# ── Filtros globais na sidebar (valem para TODAS as seções) ───────────────────
def _empilhar(chave, safras_dict):
    """Empilha uma tabela (ex: 'fazendas') das safras selecionadas, garantindo a coluna safra.
    Aplica o de-para mestre para o mesmo híbrido ter um nome único entre safras
    (ex.: 9505VTPRO4 em 24/25 e 9505PRO4 em 25/26 viram 9505PRO4)."""
    _mapa_canon, _mapa_status = _mapas_mestre()
    frames = []
    for s, d in safras_dict:
        t = d.get(chave)
        if isinstance(t, pd.DataFrame) and not t.empty:
            t = t.copy()
            if "safra" not in t.columns:
                t["safra"] = s
            t = aplicar_mestre(t, _mapa_canon, _mapa_status)
            frames.append(t)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


with st.sidebar:
    st.markdown('<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
                'letter-spacing:0.05em;padding:0.5rem 0;">Filtros do diagnóstico</p>',
                unsafe_allow_html=True)

    # safra — checkbox; 2025/26 já vem marcada por padrão; nenhuma marcada = todas
    st.markdown("**Safra**")
    nomes_safra = [s for s, _ in safras_ok]
    SAFRA_PADRAO = "2025/26"
    marcadas = [s for s in nomes_safra
                if st.checkbox(s, value=(s == SAFRA_PADRAO), key=f"diag_sf_{s}")]
    safras_sel_nomes = marcadas or nomes_safra   # nada marcado → todas
    if not marcadas:
        st.caption("Nenhuma marcada = todas as safras.")

# subconjunto de safras selecionadas (lista de tuplas, como safras_ok)
safras_sel = [(s, d) for s, d in safras_ok if s in safras_sel_nomes]

# tabelas de apoio empilhadas das safras selecionadas
def _analitica(safras_dict):
    frames = []
    for s, d in safras_dict:
        for ch in ["tabela_analitica_faixa", "tabela_analitica_densidade"]:
            t = d.get(ch)
            if isinstance(t, pd.DataFrame) and not t.empty:
                t = t.copy()
                if "safra" not in t.columns:
                    t["safra"] = s
                frames.append(t)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

plots_all = _analitica(safras_sel)
fazendas_all = _empilhar("fazendas", safras_sel)
tratamento_base = _empilhar("tratamento_base", safras_sel)

# responsável — na sidebar, opções vêm das safras selecionadas
TODOS_RESP = "(todos os responsáveis)"
with st.sidebar:
    st.markdown("**Responsável (DTC)**")
    resps = sorted(plots_all["nomeResponsavel"].dropna().unique().tolist()) \
        if ("nomeResponsavel" in plots_all.columns and not plots_all.empty) else []
    resp_sel = st.selectbox("Responsável", [TODOS_RESP] + resps, key="diag_resp",
                            label_visibility="collapsed")

filtra_resp = resp_sel != TODOS_RESP
if filtra_resp:
    plots = plots_all[plots_all["nomeResponsavel"] == resp_sel] if "nomeResponsavel" in plots_all.columns else plots_all
    fazendas = fazendas_all[fazendas_all["nomeResponsavel"] == resp_sel] if "nomeResponsavel" in fazendas_all.columns else fazendas_all
else:
    plots = plots_all
    fazendas = fazendas_all

# fazenda — na sidebar; as opções seguem o responsável já selecionado
TODAS_FAZ = "(todas as fazendas)"
with st.sidebar:
    st.markdown("**Fazenda**")
    _faz_opts = sorted(plots["nomeFazenda"].dropna().unique().tolist()) \
        if ("nomeFazenda" in plots.columns and not plots.empty) else []
    faz_sel = st.selectbox("Fazenda", [TODAS_FAZ] + _faz_opts, key="diag_faz",
                           label_visibility="collapsed")

filtra_faz = faz_sel != TODAS_FAZ
if filtra_faz:
    if "nomeFazenda" in plots.columns:
        plots = plots[plots["nomeFazenda"] == faz_sel]
    if "nomeFazenda" in fazendas.columns:
        fazendas = fazendas[fazendas["nomeFazenda"] == faz_sel]

# Vínculo confiável fazenda->responsável vem dos PLOTS (o plot carrega o responsável de fato).
# A dimensão `fazendas` às vezes traz esse campo vazio/divergente, e aí o filtro por responsável
# a deixava passar — fazendas de outro técnico vazavam para a mensagem de WhatsApp. Aqui o
# responsável de cada fazenda é reidratado a partir dos plots antes das verificações estruturais.
if "cod_fazenda" in plots_all.columns and "nomeResponsavel" in plots_all.columns:
    _resp_por_faz = (plots_all.dropna(subset=["cod_fazenda"])
                     .groupby("cod_fazenda")["nomeResponsavel"]
                     .agg(lambda s: s.dropna().mode().iloc[0] if s.dropna().any() else np.nan))
    if "cod_fazenda" in fazendas.columns:
        fazendas = fazendas.copy()
        fazendas["nomeResponsavel"] = fazendas["cod_fazenda"].map(_resp_por_faz)
        # com o vínculo reidratado, aplica o filtro de responsável que antes podia ter sido pulado
        if filtra_resp:
            fazendas = fazendas[fazendas["nomeResponsavel"] == resp_sel]

# tipo de ensaio — na sidebar; "todos" mostra Faixa + Densidade juntos
TODOS_TIPO = "(Faixa e Densidade)"
with st.sidebar:
    st.markdown("**Tipo de ensaio**")
    tipos_disp = sorted(plots["tipoTeste"].dropna().unique().tolist()) \
        if ("tipoTeste" in plots.columns and not plots.empty) else []
    tipo_sel = st.radio("Tipo", [TODOS_TIPO] + tipos_disp, key="diag_tipo",
                        label_visibility="collapsed")

filtra_tipo = tipo_sel != TODOS_TIPO
if filtra_tipo:
    if "tipoTeste" in plots.columns:
        plots = plots[plots["tipoTeste"] == tipo_sel]
    if "tipoTeste" in tratamento_base.columns:
        tratamento_base = tratamento_base[tratamento_base["tipoTeste"] == tipo_sel]

# Alguns itens são estruturais (de-para, cadastro de tratamento) e não têm vínculo com plot,
# então não respondem aos filtros de responsável nem de fazenda. A nota avisa quais estão ativos.
_filtros_ativos = [n for n, on in [("responsável", filtra_resp), ("fazenda", filtra_faz)] if on]
# NOTA_SEM_RESP: itens ESTRUTURAIS (cadastro de material, de-para) não têm vínculo com plot,
# então não respondem aos filtros — a nota avisa isso.
NOTA_SEM_RESP = (" (não filtrado por " + " nem por ".join(_filtros_ativos)
                 + " — este item não tem vínculo com o plot)") if _filtros_ativos else ""
# NOTA_FAZENDA: itens de FAZENDA (coordenada, datas) TÊM responsável (reidratado dos plots) e
# são filtrados normalmente — aqui a nota é vazia, porque o filtro se aplica de fato.
NOTA_FAZENDA = ""

# Acumulador de divergências acionáveis (alimentado pelas checagens) → texto do WhatsApp no fim.
# Cada item: {"fazenda", "ensaio" (Faixa/Densidade), "tipo" (do problema), "material", "trat", "obs"}
DIVERGENCIAS = []


def _registrar_div(df, tipo, obs_col=None):
    """Adiciona as linhas de um DataFrame de divergência ao acumulador do WhatsApp."""
    if df is None or df.empty:
        return
    for _, r in df.iterrows():
        DIVERGENCIAS.append({
            "fazenda": str(r.get("nomeFazenda") or r.get("cod_fazenda") or "—"),
            "ensaio": str(r.get("tipoTeste") or "—"),
            "tipo": tipo,
            "material": str(r.get("dePara") or "—"),
            "trat": str(r.get("indexTratamento") or "—"),
            "pop": r.get("pop_tratamento"),
            "obs": str(r.get(obs_col)) if (obs_col and pd.notna(r.get(obs_col))) else "",
        })

# ══════════════════════════════════════════════════════════════════════════════
# 2 — VISÃO CONSOLIDADA
# ══════════════════════════════════════════════════════════════════════════════
st.divider()
secao_titulo("Consolidação", "Como estão os dados hoje",
             "Retrato atual dos ensaios, fazendas e materiais por safra.")

# empilha as analíticas (faixa + densidade) das safras carregadas, para o retrato geral
_frames = []
for s, d in safras_ok:
    for chave in ["tabela_analitica_faixa", "tabela_analitica_densidade"]:
        t = d.get(chave)
        if isinstance(t, pd.DataFrame) and not t.empty:
            _frames.append(t)
consolidado = pd.concat(_frames, ignore_index=True) if _frames else pd.DataFrame()

if not consolidado.empty:
    def _metrica(label, valor):
        return (f"<div style='background:#FFFFFF;border:1px solid #EAECEF;border-radius:8px;"
                f"padding:14px 18px;'>"
                f"<div style='font-size:12px;font-weight:600;color:#6B7280;text-transform:uppercase;"
                f"letter-spacing:0.04em;'>{label}</div>"
                f"<div style='font-size:28px;font-weight:700;color:#1A1A1A;margin-top:2px;'>{valor}</div></div>")

    total = f"{len(consolidado):,}".replace(",", ".")
    n_safras = consolidado["safra"].nunique() if "safra" in consolidado.columns else len(safras_ok)
    n_faz = consolidado["cod_fazenda"].nunique() if "cod_fazenda" in consolidado.columns else "—"
    n_mat = consolidado["dePara"].nunique() if "dePara" in consolidado.columns else "—"
    mc = st.columns(4)
    for col, (lab, val) in zip(mc, [("Total de plots", total), ("Safras", n_safras),
                                    ("Fazendas", n_faz), ("Materiais", n_mat)]):
        col.markdown(_metrica(lab, val), unsafe_allow_html=True)

    if "safra" in consolidado.columns:
        resumo = (consolidado.groupby("safra", dropna=False)
                  .agg(Plots=("safra", "count"),
                       Fazendas=("cod_fazenda", "nunique"),
                       Materiais=("dePara", "nunique"))
                  .reset_index().rename(columns={"safra": "Safra"}))
        st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
        ag_table(resumo, height=min(200, 60 + 30 * len(resumo)), key="ag_resumo_safra")

# ══════════════════════════════════════════════════════════════════════════════
# 3 — INTEGRIDADE ESTRUTURAL
# ══════════════════════════════════════════════════════════════════════════════
st.divider()
secao_titulo("Integridade", "Integridade estrutural dos dados",
             "Problemas que geralmente apontam para o pipeline ou a base — não para a coleta de "
             "campo. Convém resolver estes antes de confiar nas análises.")

COL_CTX = ["safra", "cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
           "nomeResponsavel", "dePara", "status_material", "tipoTeste", "indexTratamento"]


def _ctx(df):
    return [c for c in COL_CTX if c in df.columns]


def _ordenar_trat(df):
    """Ordena a tabela: primeiro Faixa, depois Densidade; dentro de cada ensaio, pelo índice do
    tratamento (a população-alvo desempata quando o índice repete). Numéricos para 10 vir após 2."""
    d = df.copy()
    chaves = []
    if "tipoTeste" in d.columns:
        # Faixa (0) antes de Densidade (1); outros ao fim
        d["__ens"] = d["tipoTeste"].astype(str).str.lower().map(
            lambda t: 0 if t.startswith("faixa") else (1 if t.startswith("dens") else 2))
        chaves.append("__ens")
    if "indexTratamento" in d.columns:
        d["__trat"] = pd.to_numeric(d["indexTratamento"], errors="coerce")
        chaves.append("__trat")
    if "pop_tratamento" in d.columns:
        d["__pop"] = pd.to_numeric(d["pop_tratamento"], errors="coerce")
        chaves.append("__pop")
    if chaves:
        d = d.sort_values(chaves, na_position="last").drop(
            columns=[c for c in ["__ens", "__trat", "__pop"] if c in d.columns])
    return d


def _corte_dinamico(serie, piso, pct=0.95):
    """Corte que acompanha a safra: max(percentil da carga atual, piso agronômico).

    A safra está em andamento, então um número fixo envelhece — cedo pega quase tudo, tarde
    quase nada. O percentil recalcula a cada carga ("os X% piores da rede de agora"). O piso
    evita o efeito colateral do percentil puro: em safra boa, sem ele o alerta apontaria plots
    sadios só por serem o topo de uma distribuição baixa. Devolve (corte, base_percentil)."""
    v = pd.to_numeric(serie, errors="coerce").dropna()
    v = v[v > 0]                       # o percentil é sobre quem TEVE ocorrência
    p = float(v.quantile(pct)) if len(v) >= 20 else piso   # amostra pequena: fica no piso
    return max(round(p, 1), piso), round(p, 1)


def _histograma_corte(serie, corte, unidade="%", cor_barra="#1E8449", key=""):
    """Histograma horizontal em SVG puro (sem Plotly) da distribuição de uma métrica,
    com a linha de corte marcada. As barras à direita do corte saem em vermelho — são os
    plots que o cartão lista. Serve para ver o FORMATO: se o corte separa poucos casos no
    alto ou fatia uma nuvem inteira. Só desenha com dado suficiente."""
    v = pd.to_numeric(serie, errors="coerce").dropna()
    v = v[v >= 0]
    if len(v) < 12:
        return   # amostra pequena: o histograma engana mais do que ajuda
    vmax = float(v.max())
    if vmax <= 0:
        return
    # 12 faixas do 0 ao máximo (arredonda o topo para um número redondo)
    import math
    topo = max(corte * 1.2, vmax)
    passo = topo / 12
    bins = [i * passo for i in range(13)]
    contagens = []
    for i in range(12):
        lo, hi = bins[i], bins[i + 1]
        sel = (v >= lo) & (v < hi) if i < 11 else (v >= lo) & (v <= hi)
        contagens.append(int(sel.sum()))
    max_c = max(contagens) or 1

    # geometria do SVG
    W, H = 680, 200
    ml, mr, mt, mb = 34, 12, 22, 34
    plot_w, plot_h = W - ml - mr, H - mt - mb
    bw = plot_w / 12
    x_corte = ml + (corte / topo) * plot_w

    partes = [f'<svg viewBox="0 0 {W} {H}" width="100%" '
              f'style="font-family:Helvetica Neue,sans-serif;">']
    # eixo y (contagem) — só a linha e o rótulo do topo
    partes.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt+plot_h}" '
                  f'stroke="#CCC" stroke-width="1"/>')
    partes.append(f'<text x="{ml-6}" y="{mt+8}" text-anchor="end" font-size="10" '
                  f'fill="#6B7280">{max_c}</text>')
    partes.append(f'<text x="{ml-6}" y="{mt+plot_h}" text-anchor="end" font-size="10" '
                  f'fill="#6B7280">0</text>')
    # barras
    for i, c in enumerate(contagens):
        h = (c / max_c) * plot_h
        x = ml + i * bw
        y = mt + plot_h - h
        acima = bins[i + 1] > corte
        cor = "#DC2626" if acima else cor_barra   # verde abaixo do corte, vermelho acima
        op = "0.85" if acima else "0.55"
        partes.append(f'<rect x="{x+1:.1f}" y="{y:.1f}" width="{bw-2:.1f}" height="{h:.1f}" '
                      f'fill="{cor}" fill-opacity="{op}" rx="1"/>')
        if c > 0:
            partes.append(f'<text x="{x+bw/2:.1f}" y="{y-2:.1f}" text-anchor="middle" '
                          f'font-size="9" fill="#374151">{c}</text>')
    # eixo x — marcas em 0, corte e máximo
    partes.append(f'<line x1="{ml}" y1="{mt+plot_h}" x2="{ml+plot_w}" y2="{mt+plot_h}" '
                  f'stroke="#CCC" stroke-width="1"/>')
    for val in [0, corte, round(topo)]:
        xx = ml + (val / topo) * plot_w
        partes.append(f'<text x="{xx:.1f}" y="{mt+plot_h+14}" text-anchor="middle" '
                      f'font-size="10" fill="#6B7280">{val:g}{unidade}</text>')
    # linha de corte
    partes.append(f'<line x1="{x_corte:.1f}" y1="{mt}" x2="{x_corte:.1f}" y2="{mt+plot_h}" '
                  f'stroke="#DC2626" stroke-width="1.5" stroke-dasharray="4 3"/>')
    partes.append(f'<text x="{x_corte:.1f}" y="{mt-1}" text-anchor="middle" font-size="10" '
                  f'font-weight="700" fill="#DC2626">corte {corte:g}{unidade}</text>')
    partes.append('</svg>')
    st.markdown("".join(partes), unsafe_allow_html=True)


def _legenda_corte():
    """Explica a regra do corte dinâmico uma vez, abaixo do(s) histograma(s) do cartão."""
    st.caption(
        "A linha de corte é **dinâmica**: acompanha a carga porque a safra está em andamento. "
        "Para cada métrica, o corte é o **maior valor** entre o **p95 dos plots que tiveram "
        "ocorrência** (o percentil 95 — só 5% ficam acima) e um **piso mínimo** definido por "
        "boa prática agronômica. O piso evita alarme falso em safra boa, quando o p95 fica muito "
        "baixo; o p95 evita a enxurrada em safra ruim. As barras à direita do corte, em vermelho, "
        "são os plots que a tabela lista. Com poucos dados no início da safra, o corte fica no "
        "piso até haver amostra suficiente.")


def _card(titulo, n, descricao, detalhe_df=None, ok_texto="nenhuma ocorrência", ajuda=None,
          chave=None):
    """Card de verificação no estilo limpo (Schwabish): faixa lateral de cor, tipografia
    clara, ruído mínimo. Verde se n==0; âmbar com tabela ag_table + export se houver."""
    ok = (n == 0)
    cor = "#1E8449" if ok else "#B9770E"
    fundo = "#F4FBF6" if ok else "#FFFBF2"
    selo = f"✓ {ok_texto}" if ok else f"⚠ {n} ocorrência(s)"
    st.markdown(
        f"<div style='border-left:3px solid {cor};background:{fundo};border-radius:0 6px 6px 0;"
        f"padding:12px 16px;margin:8px 0;'>"
        f"<div style='display:flex;justify-content:space-between;align-items:baseline;'>"
        f"<span style='font-size:15px;font-weight:700;color:#1A1A1A;'>{titulo}</span>"
        f"<span style='font-size:13px;font-weight:700;color:{cor};'>{selo}</span></div>"
        f"<div style='font-size:12px;color:#6B7280;margin-top:3px;'>{descricao}</div></div>",
        unsafe_allow_html=True)
    if ajuda:
        with st.popover("ℹ️ Como interpretar", use_container_width=False):
            st.markdown(ajuda)
    if not ok and detalhe_df is not None and not detalhe_df.empty:
        # a chave dos widgets vem do título, mas títulos com prefixo comum colidem depois do
        # corte em 20 caracteres — por isso quem gera cartões em série passa `chave` explícita
        chave = chave or "".join(ch for ch in titulo.lower() if ch.isalnum())[:20]
        det_ord = _ordenar_trat(detalhe_df)
        det_rot = _sanitizar(_rotular(det_ord))
        with st.expander(f"Ver os {n} registro(s)"):
            altura = min(340, 60 + 30 * min(len(det_rot), 12))
            ag_table(det_rot, height=altura, key=f"ag_{chave}")
            exportar_excel(det_rot, nome_arquivo=f"{chave}.xlsx",
                           label="⬇️ Exportar estes registros", key=f"exp_{chave}")


if plots.empty:
    st.info("Sem plots analíticos nesta safra para diagnosticar.")
    st.stop()

# ── V1: registros órfãos (sem fazenda ou sem material) ────────────────────────
orf_faz = plots[plots["cod_fazenda"].isna()] if "cod_fazenda" in plots.columns else plots.iloc[0:0]
orf_mat = plots[plots["dePara"].isna()] if "dePara" in plots.columns else plots.iloc[0:0]
n_orf = len(orf_faz) + len(orf_mat)
det_orf = pd.concat([orf_faz.assign(_problema="sem fazenda"), orf_mat.assign(_problema="sem material")],
                    ignore_index=True) if n_orf else pd.DataFrame()
if not det_orf.empty:
    det_orf = det_orf[["_problema"] + _ctx(det_orf)]
_card("Registros órfãos", n_orf,
      "Plots sem fazenda ou sem material vinculado. Costuma indicar join que perdeu o vínculo.",
      det_orf)

# ── V2: fazendas sem responsável ──────────────────────────────────────────────
if not fazendas.empty and "nomeResponsavel" in fazendas.columns:
    faz_sem_resp = fazendas[fazendas["nomeResponsavel"].isna()]
    cols_faz = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"] if c in fazendas.columns]
    _card("Fazendas sem responsável", len(faz_sem_resp),
          "Fazendas sem técnico responsável definido.",
          faz_sem_resp[cols_faz] if len(faz_sem_resp) else None)
else:
    _card("Fazendas sem responsável", 0, "Tabela de fazendas não disponível.", None,
          ok_texto="não verificável")

# ── V3: duplicados (mesmo plot repetido) ──────────────────────────────────────
# Usa a CHAVE real do pipeline (fazendaRef + idBaseRef + tipoTeste + indexTratamento).
# NÃO usar dePara aqui: na Densidade o mesmo material aparece em várias populações com o
# mesmo nome (dePara), então dePara não identifica um plot único — idBaseRef sim.
CHAVE_PLOT = ["fazendaRef", "idBaseRef", "tipoTeste", "indexTratamento"]
chave_plot = [c for c in CHAVE_PLOT if c in plots.columns]
if len(chave_plot) >= 3:
    mask_dup = plots.duplicated(subset=chave_plot, keep=False)
    dups = plots[mask_dup].sort_values(chave_plot)
    _card("Plots duplicados", int(mask_dup.sum()),
          "Mesmo plot (fazenda + tratamento + tipo + índice) aparecendo mais de uma vez. "
          "Pode indicar merge que multiplicou linhas.",
          dups[_ctx(dups)] if mask_dup.sum() else None)
else:
    _card("Plots duplicados", 0,
          "Chave única do plot (fazendaRef/idBaseRef) não disponível nesta tabela.",
          None, ok_texto="não verificável")

# ── V4: material sem de-para (nome cru não bateu) ─────────────────────────────
if not tratamento_base.empty and "dePara" in tratamento_base.columns:
    mat_sem = tratamento_base[tratamento_base["dePara"].isna()]
    cols_m = [c for c in ["nome", "tipoTeste", "indexTratamento", "regional"] if c in tratamento_base.columns]
    det_m = mat_sem[cols_m].drop_duplicates() if len(mat_sem) else None
    if det_m is not None and "nome" in det_m.columns:
        # nº de responsáveis que usam o material — número, não a lista (a lista seria a rede
        # inteira para híbrido comum). Serve para distinguir material raro de material comum.
        _cont_m = {}
        if "nome" in plots_all.columns and "nomeResponsavel" in plots_all.columns:
            _bm = plots_all.dropna(subset=["nome", "nomeResponsavel"])
            _cont_m = _bm.groupby("nome")["nomeResponsavel"].nunique().to_dict()
        det_m = det_m.assign(**{"Usado por (nº resp.)":
                                det_m["nome"].map(lambda n: _cont_m.get(n, 0))})
    _card("Híbrido sem de-para", len(mat_sem),
          "Nome do material no banco que não bateu com o de-para da safra — fica sem nome canônico. "
          "É um ajuste de cadastro central (do de-para), não de campo." + NOTA_SEM_RESP,
          det_m)
else:
    _card("Híbrido sem de-para", 0, "Catálogo de tratamentos não disponível.", None, ok_texto="não verificável")

# ── V4b: híbrido fora do de-para MESTRE (nome não é reconciliado entre safras) ──
_mapa_canon_diag, _ = _mapas_mestre()
if not tratamento_base.empty and "dePara" in tratamento_base.columns and _mapa_canon_diag:
    # dePara_original guarda o canônico da safra; se ele não está no mestre, o nome
    # daquele material não é reconciliado e pode divergir entre safras
    _col_orig = "dePara_original" if "dePara_original" in tratamento_base.columns else "dePara"
    _chaves_mestre = set(_mapa_canon_diag.keys())
    _fora = tratamento_base[
        tratamento_base[_col_orig].notna()
        & ~tratamento_base[_col_orig].astype(str).str.strip().isin(_chaves_mestre)
    ]
    n_fora = _fora[_col_orig].nunique()
    if n_fora:
        cols_fm = [c for c in [_col_orig, "nome", "status_material", "safra", "tipoTeste"]
                   if c in _fora.columns]
        det_fm = (_fora[cols_fm].drop_duplicates(subset=[_col_orig, "safra"])
                  .rename(columns={_col_orig: "Híbrido (nome da safra)"})
                  .sort_values("Híbrido (nome da safra)"))
        # nº de responsáveis que usam (número, não lista — a lista seria a rede inteira)
        _cmap = {}
        if "dePara" in plots_all.columns and "nomeResponsavel" in plots_all.columns:
            _cmap = (plots_all.dropna(subset=["dePara", "nomeResponsavel"])
                     .groupby("dePara")["nomeResponsavel"].nunique().to_dict())
        det_fm = det_fm.assign(**{"Usado por (nº resp.)":
                                  det_fm["Híbrido (nome da safra)"].map(lambda n: _cmap.get(n, 0))})
        reg = _fora.drop_duplicates(subset=[_col_orig, "safra"]).copy()
        reg["_obs"] = "fora do de-para mestre"
        _registrar_div(reg, "Híbrido fora do de-para mestre", obs_col="_obs")
    else:
        det_fm = None
    _card("Híbridos fora do de-para mestre", int(n_fora),
          "O material não está no depara_mestre.csv, então o nome e o status usados são os da própria "
          "safra. Se ele aparecer em mais de uma safra com grafias diferentes, será tratado como dois "
          "híbridos distintos nas análises. Acrescente ao mestre para unificar." + NOTA_SEM_RESP,
          det_fm)
elif not _mapa_canon_diag:
    _card("Híbridos fora do de-para mestre", 0,
          "Arquivo depara_mestre.csv não encontrado em config/ — nenhuma reconciliação entre safras "
          "está sendo aplicada.", None, ok_texto="não verificável")

# ── V5: plots sem tipo de ensaio ──────────────────────────────────────────────
if "tipoTeste" in plots.columns:
    sem_tipo = plots[plots["tipoTeste"].isna() | (plots["tipoTeste"].astype(str).str.strip() == "")]
    _card("Plots sem tipo de ensaio", len(sem_tipo),
          "Plots sem Faixa/Densidade definido — não entram nas análises que dependem do tipo.",
          sem_tipo[_ctx(sem_tipo)] if len(sem_tipo) else None)

# ── V6: fazendas sem cidade/região (de-para de município falhou) ──────────────
if not fazendas.empty:
    cols_geo = [c for c in ["cidade_nome", "regiao_macro", "regiao_micro"] if c in fazendas.columns]
    if cols_geo:
        mask_geo = fazendas[cols_geo].isna().any(axis=1)
        faz_sem_geo = fazendas[mask_geo]
        cols_show = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
                                 "regiao_macro", "regiao_micro", "nomeResponsavel"] if c in fazendas.columns]
        _card("Fazendas sem cidade/região", int(mask_geo.sum()),
              "De-para de município não preencheu cidade ou região — afeta o mapa e os cortes regionais.",
              faz_sem_geo[cols_show] if mask_geo.sum() else None)
    else:
        _card("Fazendas sem cidade/região", 0, "Colunas de região não disponíveis.", None,
              ok_texto="não verificável")

# ── V6b: fazendas sem data de plantio (afeta a marcha de plantio e a análise de época) ──
if not fazendas.empty and "dataPlantioMilho" in fazendas.columns:
    faz_u = fazendas.drop_duplicates(subset="cod_fazenda") if "cod_fazenda" in fazendas.columns else fazendas
    mask_pl = faz_u["dataPlantioMilho"].isna()
    n_pl = int(mask_pl.sum())
    cols_dt = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
                           "nomeResponsavel"] if c in faz_u.columns]
    if n_pl:
        det_pl = faz_u[mask_pl][cols_dt]
        reg = faz_u[mask_pl].copy()
        reg["_obs"] = "sem data de plantio"
        _registrar_div(reg, "Fazenda sem data de plantio", obs_col="_obs")
    else:
        det_pl = None
    _card("Fazendas sem data de plantio", n_pl,
          "A data de plantio não foi lançada no cadastro da fazenda. Sem ela, o local não aparece na "
          "marcha de plantio nem na análise de época (mas entra nas demais tabelas)." + NOTA_SEM_RESP,
          det_pl)

# ── V6c: fazendas sem data de colheita (afeta a marcha de colheita) ───────────
if not fazendas.empty and "dataColheitaMilho" in fazendas.columns:
    faz_u = fazendas.drop_duplicates(subset="cod_fazenda") if "cod_fazenda" in fazendas.columns else fazendas
    mask_co = faz_u["dataColheitaMilho"].isna()
    n_co = int(mask_co.sum())
    cols_dt = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
                           "nomeResponsavel"] if c in faz_u.columns]
    if n_co:
        det_co = faz_u[mask_co][cols_dt]
        reg = faz_u[mask_co].copy()
        reg["_obs"] = "sem data de colheita"
        _registrar_div(reg, "Fazenda sem data de colheita", obs_col="_obs")
    else:
        det_co = None
    _card("Fazendas sem data de colheita", n_co,
          "A data de colheita não foi lançada no cadastro da fazenda. Sem ela, o local não aparece na "
          "marcha de colheita (mas entra nas demais tabelas)." + NOTA_FAZENDA,
          det_co)

# ── V6d: fazendas sem coordenadas (afeta os mapas) ───────────────────────────
if not fazendas.empty and {"latitude", "longitude"}.issubset(fazendas.columns):
    faz_u = fazendas.drop_duplicates(subset="cod_fazenda") if "cod_fazenda" in fazendas.columns else fazendas
    _lat = pd.to_numeric(faz_u["latitude"], errors="coerce")
    _lon = pd.to_numeric(faz_u["longitude"], errors="coerce")
    # sem coordenada = ausente, zerada, ou fora dos limites plausíveis do Brasil
    _fora = (_lat.isna() | _lon.isna() | ((_lat == 0) & (_lon == 0))
             | (~_lat.between(-34, 6)) | (~_lon.between(-74, -34)))
    n_geo = int(_fora.sum())
    cols_geo = [c for c in ["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
                            "nomeResponsavel", "latitude", "longitude"] if c in faz_u.columns]
    if n_geo:
        det_geo = faz_u[_fora][cols_geo]
        reg = faz_u[_fora].copy()
        _motivo = np.where(_lat[_fora].isna() | _lon[_fora].isna(), "sem coordenada",
                           "coordenada fora do Brasil")
        reg["_obs"] = _motivo
        _registrar_div(reg, "Fazenda sem coordenada", obs_col="_obs")
    else:
        det_geo = None
    _card("Fazendas sem coordenada válida", n_geo,
          "Latitude/longitude ausentes, zeradas ou fora dos limites do Brasil. Sem coordenada o local "
          "não aparece nos mapas (H2H por local e página de Mapa), embora entre normalmente nas "
          "tabelas e gráficos." + NOTA_FAZENDA,
          det_geo)

# ── V7: de-para sem status_material ───────────────────────────────────────────
if "dePara" in plots.columns and "status_material" in plots.columns:
    mask_st = plots["dePara"].notna() & (
        plots["status_material"].isna() | (plots["status_material"].astype(str).str.strip() == ""))
    sem_status = plots[mask_st].drop_duplicates(subset=["dePara"]) if "dePara" in plots.columns else plots[mask_st]
    _card("Material com de-para mas sem status", int(mask_st.sum()),
          "Material que bateu no de-para mas ficou sem classificação (STINE/CHECK/EXP) — de-para incompleto.",
          sem_status[[c for c in ["dePara", "status_material", "cod_fazenda", "tipoTeste"] if c in plots.columns]]
          if mask_st.sum() else None)

# ── V8: tratamento cadastrado e nunca avaliado ───────────────────────────────
# O tratamentoBase é o catálogo de materiais; não tem fazenda (o vínculo tratamento->fazenda
# só nasce na avaliação, quando idBaseRef encontra a fazendaRef via plot). Então:
#  - "sem avaliação" = uuid do tratamento que NÃO aparece como idBaseRef em NENHUM plot da rede
#    inteira (plots_all — não os filtrados, senão os tratamentos de outros técnicos vazam);
#  - para esses, fazenda/responsável não existem por definição, e a tabela diz isso;
#  - para os que TÊM avaliação, o cruzamento traz a fazenda e o responsável de cada um.
if not tratamento_base.empty and "idBaseRef" in plots_all.columns:
    tb_id = "uuid" if "uuid" in tratamento_base.columns else None
    if tb_id:
        _ids_avaliados = set(plots_all["idBaseRef"].dropna().unique())
        _sem = tratamento_base[~tratamento_base[tb_id].isin(_ids_avaliados)].copy()
        # deduplica pelo próprio uuid (cada linha do catálogo é um tratamento)
        _sem = _sem.drop_duplicates(subset=[tb_id])
        n_sem = len(_sem)
        # regional sai: para o órfão ele é redundante e você pediu para tirar. Fazenda/cidade/
        # estado/responsável não entram porque não existem para quem nunca foi avaliado
        # (o vínculo com a fazenda só nasce na avaliação — confirmado no schema do pipeline).
        cols_tb = [c for c in ["safra", "nome", "dePara", "tipoTeste", "indexTratamento",
                               "pop_tratamento"] if c in _sem.columns]
        det_sem = _sem[cols_tb] if (n_sem and cols_tb) else None
        _card("Tratamentos cadastrados sem avaliação", n_sem,
              "Materiais no catálogo cujo cadastro nunca apareceu em nenhuma avaliação da rede — "
              "cadastrado e nunca medido, ou o vínculo se perdeu no pipeline. Fazenda e responsável "
              "não aparecem porque esse vínculo só nasce com uma avaliação, que estes não têm. "
              "Casos típicos: cadastros de teste (ex.: 'TESTE MILHO') e materiais lançados fora de "
              "protocolo. (comparado contra a rede inteira, não muda com os filtros)",
              det_sem)
    else:
        _card("Tratamentos cadastrados sem avaliação", 0,
              "Chave única do tratamento (uuid) não disponível para cruzar.",
              None, ok_texto="não verificável")

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 4 — CHECAGEM POR AVALIAÇÃO · av1 (qualidade inicial)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Avaliações", "Qualidade da coleta — av1 (qualidade inicial)",
             "Problemas na coleta das 8 notas da av1. A escala do protocolo é de 1 a 5 (0 = não avaliado).")

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
**O que é a av1**

Na av1 (qualidade inicial), o técnico dá uma nota de **1 a 5** para cada uma das **8 categorias**
do plot (uniformidade, densidade, vigor, daninhas, pragas, doenças, homogeneidade e solo). A
**média dessas 8 notas** resume a qualidade inicial do plot. O protocolo sempre foi de 1 a 5;
a nota **0** significa "não avaliado" e não entra na média.

---

**O que o diagnóstico verifica**

- **Notas fora do intervalo (1–5):** qualquer nota que não seja 0 a 5. Como o protocolo é 1 a 5 e o
  app não permite negativos, na prática isso significa nota **6 a 9** — erro de digitação. Se
  entrasse na conta, distorceria a média, então o pipeline **zera a média** do plot inteiro.
- **Média não calculada:** plots que ficaram sem média, com o **motivo** (nota fora do intervalo,
  ou todas as 8 notas iguais a 0 — plot não avaliado).

---

**Como interpretar / o que fazer**

- Se um plot aparece em **notas fora do intervalo**, procure a categoria com o valor 6–9 e
  **corrija a digitação no aplicativo** (a nota certa está entre 1 e 5).
- Se a média não foi calculada por **todas as notas 0**, o plot simplesmente não foi avaliado —
  não é erro, é dado faltante (avaliar em campo, se ainda fizer sentido).
- Use os filtros de **responsável** e **fazenda** na sidebar para estreitar o recorte e o botão
  **Exportar** para mandar a lista de correções ao time.
""")

# nomes legíveis das 8 categorias
LABEL_NOTA = {
    "nota_uniformidade": "Uniformidade de emergência", "nota_densidade": "Densidade de plantas",
    "nota_vigor": "Vigor das plantas", "nota_daninhas": "Presença de daninhas",
    "nota_pragas": "Presença de pragas", "nota_doencas": "Presença de doenças",
    "nota_homogeneidade": "Homogeneidade de crescimento", "nota_solo": "Estado geral do solo",
}

av1 = _empilhar("av1_gold", safras_sel)
# aplica os filtros globais (responsável e tipo de ensaio)
if filtra_resp and not av1.empty and "nomeResponsavel" in av1.columns:
    av1 = av1[av1["nomeResponsavel"] == resp_sel]
if filtra_tipo and not av1.empty and "tipoTeste" in av1.columns:
    av1 = av1[av1["tipoTeste"] == tipo_sel]

if av1.empty:
    st.info("Sem dados de av1 para este filtro.")
else:
    notas_presentes = [c for c in NOTAS_AV1 if c in av1.columns]
    # matriz numérica das 8 notas (uma vez, reaproveitada nas 3 checagens)
    M = av1[notas_presentes].apply(pd.to_numeric, errors="coerce")

    ctx_av1 = [c for c in ["safra", "cod_fazenda", "nomeFazenda", "nomeResponsavel", "dePara",
                           "tipoTeste", "indexTratamento"] if c in av1.columns]

    def _plots_com(mask_bool, extra_cols=None):
        """Monta o detalhe dos plots que casam com a máscara, com contexto + as notas."""
        sub = av1[mask_bool]
        cols = ctx_av1 + (extra_cols or []) + notas_presentes
        cols = [c for c in cols if c in sub.columns]
        return sub[cols]

    # ── C1: notas fora do intervalo válido (1–5), tratando 0 como "não avaliado" ──
    # Como o protocolo é 1–5 e o app não permite negativos, "fora do intervalo" na prática
    # significa nota ≥ 6 (erro de digitação). Uma checagem só, guarda-chuva.
    fora_por_nota = M.apply(lambda col: col.notna() & (col != 0) & ((col < 1) | (col > 5)))
    mask_fora = fora_por_nota.any(axis=1)
    n_fora = int(mask_fora.sum())
    det_fora = _plots_com(mask_fora) if n_fora else None

    # monta a observação específica: quais categorias estão fora e com que valor
    # (ex.: "Vigor das plantas=7; Densidade de plantas=6") — para o WhatsApp ficar acionável
    if n_fora:
        det_fora = det_fora.copy()
        obs_list = []
        for idx in det_fora.index:
            partes = []
            for c in notas_presentes:
                v = M.at[idx, c] if idx in M.index else None
                if pd.notna(v) and (v != 0) and (v < 1 or v > 5):
                    val = int(v) if float(v).is_integer() else v
                    partes.append(f"{LABEL_NOTA.get(c, c)}={val}")
            obs_list.append("; ".join(partes))
        det_fora["_detalhe"] = obs_list
        _registrar_div(det_fora, "Nota fora do intervalo (1-5)", obs_col="_detalhe")
        det_fora = det_fora.drop(columns=["_detalhe"])   # não exibir na tabela (fica nas notas)
    _card("Notas fora do intervalo (1–5)", n_fora,
          "Plots com alguma das 8 notas fora da escala do protocolo (1 a 5) — na prática, nota 6 a 9. "
          "O 0 é válido (não avaliado). Foi digitada errada e, se entrasse na conta, distorceria a "
          "média — então o plot inteiro tem a média zerada." + NOTA_SEM_RESP,
          det_fora)

    # mask_escala reaproveitada na C3 (motivo da média não calculada) — mesmo critério
    mask_escala = mask_fora

    # ── C3: média não calculada, com o motivo ──
    if "media_categorias" in av1.columns:
        mask_sem_media = av1["media_categorias"].isna()
        # motivo: nota fora do intervalo (≥6), ou todas as notas 0 (não avaliado)
        todas_zero = M.apply(lambda col: col.fillna(0) == 0).all(axis=1)
        motivo = pd.Series("outro", index=av1.index)
        motivo[todas_zero] = "todas as notas 0 (não avaliado)"
        motivo[mask_escala] = "nota fora do intervalo (6 a 9)"
        n_sem_media = int(mask_sem_media.sum())
        if n_sem_media:
            det_sm = av1[mask_sem_media].copy()
            det_sm["_motivo"] = motivo[mask_sem_media]
            # registra no WhatsApp só os "tudo zero" (não avaliados) — os de nota fora do
            # intervalo já foram registrados na checagem anterior, para não duplicar.
            det_zero = det_sm[det_sm["_motivo"] == "todas as notas 0 (não avaliado)"]
            _registrar_div(det_zero, "Plot sem avaliação (todas as notas 0)")
            cols_sm = ["_motivo"] + [c for c in ctx_av1 if c in det_sm.columns] + notas_presentes
            det_sm = det_sm[[c for c in cols_sm if c in det_sm.columns]]
        else:
            det_sm = None
        _card("Média não calculada", n_sem_media,
              "Plots sem media_categorias. O motivo aparece na coluna _motivo: nota fora do intervalo "
              "(corrigir a digitação) ou todas as notas 0 (plot não avaliado)." + NOTA_SEM_RESP,
              det_sm)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 4b — CHECAGEM POR AVALIAÇÃO · av2 (sanidade / doenças)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Avaliações", "Qualidade da coleta — av2 (sanidade)",
             "Consistência das notas de doença. A escala é de 1 a 9 (0 = doença ausente na fazenda).")

LABEL_DOENCA = {
    "nota_turcicum": "Mancha de Turcicum", "nota_cercospora": "Mancha de Cercospora",
    "nota_mancha_branca": "Mancha branca", "nota_bipolaris": "Mancha de Bipolaris",
    "nota_ferrugem_tropical": "Ferrugem tropical", "nota_enfezamento": "Enfezamento",
}

with st.popover("Como interpretar", use_container_width=False):
    st.markdown("""
**O que é a av2**

Na av2 (sanidade), o técnico dá uma nota de **1 a 9** para cada doença, onde **quanto maior a
nota, mais sadia** a planta (mais resistente). A nota **0** tem um significado especial: a doença
está **ausente naquela fazenda**.

---

**A regra do protocolo**

Pelo protocolo, a nota **0 significa que a doença está ausente no local (fazenda)**. E a ausência
é uma característica do ambiente, não do plot: se a doença não ocorre naquela fazenda, ela não
ocorre em **nenhum** dos plots dali. Portanto, se um plot da fazenda tem 0 para uma doença, todos
os outros plots da mesma fazenda deveriam ter 0 para essa mesma doença. Caso contrário — se alguns
têm 0 e outros têm nota — é preciso **corrigir**: acertar a nota dos que estão errados.

---

**O que o diagnóstico verifica**

- **Ausência inconsistente na fazenda:** aplica a regra acima. Se uma doença aparece com 0 em
  alguns plots e com nota (1 a 9) em outros, na mesma fazenda, marca a inconsistência e mostra
  todos os plots daquela doença ali — para você ver quais destoam.
- **Plot sem nenhuma doença avaliada:** plots onde todas as 6 doenças estão em branco (nem 0 nem
  nota) — sanidade não avaliada.

---

**Como interpretar / o que fazer**

- Na inconsistência, decida qual é o caso real da fazenda: se a doença **ocorreu**, os plots com 0
  estão errados e precisam receber a nota; se **não ocorreu**, os plots com nota estão errados e
  devem ir a 0.
- Use os filtros de **responsável** e **fazenda** na sidebar e o botão **Exportar** para acionar o time.
""")

av2 = _empilhar("av2_gold", safras_sel)
if filtra_resp and not av2.empty and "nomeResponsavel" in av2.columns:
    av2 = av2[av2["nomeResponsavel"] == resp_sel]
if filtra_tipo and not av2.empty and "tipoTeste" in av2.columns:
    av2 = av2[av2["tipoTeste"] == tipo_sel]

if av2.empty:
    st.info("Sem dados de av2 para este filtro.")
else:
    doencas_presentes = [d for d in DOENCAS_AV2 if d in av2.columns]
    # população-alvo entra na tabela só quando o filtro está em Densidade
    _dens = filtra_tipo and str(tipo_sel).lower().startswith("dens")
    ctx_av2 = [c for c in ["safra", "cod_fazenda", "nomeFazenda", "nomeResponsavel", "dePara",
                           "tipoTeste", "indexTratamento"]
               + (["pop_tratamento"] if _dens else [])
               if c in av2.columns]

    # ── C1: ausência inconsistente na fazenda (mistura de 0 e não-0 na mesma fazenda/doença) ──
    # Regra do protocolo: 0 = doença ausente NA FAZENDA → deveria ser 0 em TODOS os plots dela.
    # Um cartão POR DOENÇA, cada um listando os plots das fazendas afetadas por aquela doença.
    # A coluna Situação marca o lado com menos plots ("conferir"), que costuma ser o engano.
    chave_faz = "cod_fazenda" if "cod_fazenda" in av2.columns else None
    if chave_faz:
        for d in doencas_presentes:
            _blocos, _resumos = [], []
            for faz, grupo in av2.groupby(chave_faz):
                vals = pd.to_numeric(grupo[d], errors="coerce")
                m_zero, m_nota = (vals == 0), (vals > 0)
                if not (m_zero.any() and m_nota.any()):
                    continue                      # consistente: só 0 ou só nota
                n_zero, n_nota = int(m_zero.sum()), int(m_nota.sum())
                _min_mask = m_nota if n_nota <= n_zero else m_zero
                sub = grupo.copy()
                sub["Nota"] = vals.values
                sub["Situação"] = np.where(_min_mask.values, "conferir", "maioria")
                sub["_min"] = min(n_zero, n_nota)
                _blocos.append(sub)
                _nome_faz = (grupo["nomeFazenda"].iloc[0]
                             if "nomeFazenda" in grupo.columns else str(faz))
                _resumos.append(f"<b>{_nome_faz}</b>: {n_zero} com 0 e {n_nota} com nota")

            _rot_d = LABEL_DOENCA.get(d, d)
            _chave_d = "ausinc" + "".join(ch for ch in str(d).lower() if ch.isalnum())
            if not _blocos:
                _card(f"Ausência inconsistente — {_rot_d}", 0,
                      "O 0 significa que a doença não ocorreu no local, então deveria valer para "
                      "todos os plots da fazenda." + NOTA_SEM_RESP,
                      None, chave=_chave_d)
                continue

            det_full = pd.concat(_blocos, ignore_index=True)
            # nota como texto (evita 7.0 e problema de serialização com NaN misto)
            det_full["Nota"] = det_full["Nota"].apply(
                lambda x: "0 (ausente)" if x == 0 else ("—" if pd.isna(x) else str(int(x))))
            # fazendas com menos divergência primeiro; dentro delas, o lado a conferir no topo
            det_full["_ordem_sit"] = (det_full["Situação"] != "conferir").astype(int)
            det_full = det_full.sort_values(
                ["_min", "nomeFazenda", "_ordem_sit"] if "nomeFazenda" in det_full.columns
                else ["_min", "_ordem_sit"]).reset_index(drop=True)

            cols_inc = ["Situação", "Nota"] + ctx_av2
            det_inc = det_full[[c for c in cols_inc if c in det_full.columns]]
            n_inc = len(det_inc)

            _reg = det_full[det_full["Situação"] == "conferir"].copy()
            if "dePara" not in _reg.columns:
                _reg["dePara"] = "—"
            _reg["_wpp_obs"] = _rot_d + " = " + _reg["Nota"].astype(str) + " (destoa da fazenda)"
            _registrar_div(_reg, f"Ausência inconsistente — {_rot_d}", obs_col="_wpp_obs")

            _card(f"Ausência inconsistente — {_rot_d}", n_inc,
                  "Fazendas em que esta doença aparece com 0 em alguns plots e com nota em "
                  "outros. Como o 0 significa ausência no local, um dos dois lados está errado. "
                  "A coluna <b>Situação</b> marca com <b>conferir</b> o lado com menos plots. "
                  "Resumo: " + " · ".join(_resumos) + "." + NOTA_SEM_RESP,
                  det_inc, chave=_chave_d)

    # ── C2: plot sem nenhuma doença avaliada (todas em branco) ──
    Mv2 = av2[doencas_presentes].apply(pd.to_numeric, errors="coerce")
    # "não avaliado" = todas as doenças NaN (nem 0 nem nota). O 0 conta como avaliado (= ausente).
    mask_vazio = Mv2.isna().all(axis=1)
    n_vazio = int(mask_vazio.sum())
    if n_vazio:
        det_vazio = av2[mask_vazio][[c for c in ctx_av2 if c in av2.columns]]
        _registrar_div(av2[mask_vazio], "Plot sem sanidade avaliada")
    else:
        det_vazio = None
    _card("Plot sem nenhuma doença avaliada", n_vazio,
          "Plots onde nenhuma das 6 doenças foi avaliada (todas em branco). Sanidade não coletada."
          + NOTA_SEM_RESP,
          det_vazio)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 4c — CHECAGEM POR AVALIAÇÃO · av3 (caracterização: altura e florescimento)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Avaliações", "Qualidade da coleta — av3 (caracterização)",
             "Consistência da altura (planta e espiga) e do florescimento (dias do plantio).")

ASI_MAX = 10   # dias; diferença máxima plausível entre florescimento masculino e feminino

with st.popover("Como interpretar", use_container_width=False):
    st.markdown(f"""
**O que é a av3**

Na av3 (caracterização), mede-se a **altura da planta** e a **altura da espiga** (média de 5
plantas, em cm) e as **datas de florescimento** masculino (pendão) e feminino (espiga), das quais
o pipeline calcula os **dias do plantio até o florescimento**.

---

**O que o diagnóstico verifica**

- **Altura não disponível:** a altura de planta ou de espiga ficou sem valor — não medida,
  registrada como 0, ou um valor absurdo (acima de 350 cm) que o sistema descartou.
- **Espiga mais alta que a planta:** a espiga fica no meio do colmo, então sua altura é sempre
  menor que a da planta. Se a espiga aparece mais alta, houve troca dos campos ou erro de digitação.
- **Florescimento fora de {FLOR_MIN}–{FLOR_MAX} dias:** o florescimento costuma ocorrer nessa
  janela após o plantio. Fora dela, a data de plantio ou de florescimento provavelmente está errada.
- **Intervalo de florescimento maior que {ASI_MAX} dias:** o masculino e o feminino podem florescer
  em qualquer ordem, mas a diferença entre eles costuma ser pequena. Um intervalo muito grande (em
  qualquer direção) pode indicar estresse da lavoura ou erro em uma das datas.

---

**Como interpretar / o que fazer**

- Confira a medição ou a data no aplicativo e corrija o que estiver errado.
- Use os filtros de **responsável** e **fazenda** na sidebar e o botão **Exportar** para acionar o time.
""")

av3 = _empilhar("av3_gold", safras_sel)
if filtra_resp and not av3.empty and "nomeResponsavel" in av3.columns:
    av3 = av3[av3["nomeResponsavel"] == resp_sel]
if filtra_tipo and not av3.empty and "tipoTeste" in av3.columns:
    av3 = av3[av3["tipoTeste"] == tipo_sel]

if av3.empty:
    st.info("Sem dados de av3 para este filtro.")
else:
    ctx_av3 = [c for c in ["safra", "cod_fazenda", "nomeFazenda", "nomeResponsavel", "dePara",
                           "tipoTeste", "indexTratamento"] if c in av3.columns]
    alt_p = pd.to_numeric(av3["altura_planta_cm"], errors="coerce") if "altura_planta_cm" in av3.columns else pd.Series(dtype=float)
    alt_e = pd.to_numeric(av3["altura_espiga_cm"], errors="coerce") if "altura_espiga_cm" in av3.columns else pd.Series(dtype=float)

    # C1: altura não disponível
    mask_alt = pd.Series(False, index=av3.index)
    if not alt_p.empty:
        mask_alt = mask_alt | alt_p.isna()
    if not alt_e.empty:
        mask_alt = mask_alt | alt_e.isna()
    n_alt = int(mask_alt.sum())
    if n_alt:
        cols = ctx_av3 + [c for c in ["altura_planta_cm", "altura_espiga_cm"] if c in av3.columns]
        det_alt = av3[mask_alt][cols]
        _registrar_div(av3[mask_alt], "Altura não disponível")
    else:
        det_alt = None
    _card("Altura não disponível", n_alt,
          "Altura de planta ou espiga sem valor — não medida, registrada como 0, ou valor absurdo "
          "(acima de 350 cm) que o sistema descartou." + NOTA_SEM_RESP,
          det_alt)

    # C2: espiga mais alta que a planta
    if not alt_p.empty and not alt_e.empty:
        mask_esp = alt_p.notna() & alt_e.notna() & (alt_e > alt_p)
        n_esp = int(mask_esp.sum())
        if n_esp:
            cols = ctx_av3 + ["altura_planta_cm", "altura_espiga_cm"]
            det_esp = av3[mask_esp][[c for c in cols if c in av3.columns]]
            reg = av3[mask_esp].copy()
            reg["_obs"] = ("espiga " + alt_e[mask_esp].round(0).astype("Int64").astype(str)
                           + " cm > planta " + alt_p[mask_esp].round(0).astype("Int64").astype(str) + " cm")
            _registrar_div(reg, "Espiga mais alta que a planta", obs_col="_obs")
        else:
            det_esp = None
        _card("Espiga mais alta que a planta", n_esp,
              "A altura da espiga ficou maior que a da planta — impossível, já que a espiga fica no "
              "meio do colmo. Provável troca de campos ou erro de digitação." + NOTA_SEM_RESP,
              det_esp)

    # C3: florescimento fora da faixa
    mask_flor = pd.Series(False, index=av3.index)
    for bruto_col, val_col in [("dias_flor_masculino", "dias_flor_masculino_valido"),
                               ("dias_flor_feminino", "dias_flor_feminino_valido")]:
        if bruto_col in av3.columns and val_col in av3.columns:
            bruto = pd.to_numeric(av3[bruto_col], errors="coerce")
            val = pd.to_numeric(av3[val_col], errors="coerce")
            mask_flor = mask_flor | (bruto.notna() & val.isna())
    n_flor = int(mask_flor.sum())
    if n_flor:
        cols = ctx_av3 + [c for c in ["dias_flor_masculino", "dias_flor_feminino"] if c in av3.columns]
        det_flor = av3[mask_flor][[c for c in cols if c in av3.columns]]
        _registrar_div(av3[mask_flor], f"Florescimento fora de {FLOR_MIN}-{FLOR_MAX} dias")
    else:
        det_flor = None
    _card(f"Florescimento fora de {FLOR_MIN}–{FLOR_MAX} dias", n_flor,
          f"Dias do plantio ao florescimento fora da janela plausível ({FLOR_MIN} a {FLOR_MAX} dias). "
          "A data de plantio ou de florescimento provavelmente foi digitada errada." + NOTA_SEM_RESP,
          det_flor)

    # C4: intervalo grande entre florescimento masculino e feminino (qualquer ordem)
    if "dias_flor_masculino" in av3.columns and "dias_flor_feminino" in av3.columns:
        dm = pd.to_numeric(av3["dias_flor_masculino"], errors="coerce")
        df_ = pd.to_numeric(av3["dias_flor_feminino"], errors="coerce")
        intervalo = (df_ - dm).abs()
        mask_asi = dm.notna() & df_.notna() & (intervalo > ASI_MAX)
        n_asi = int(mask_asi.sum())
        if n_asi:
            cols = ctx_av3 + ["dias_flor_masculino", "dias_flor_feminino"]
            det_asi = av3[mask_asi][[c for c in cols if c in av3.columns]]
            reg = av3[mask_asi].copy()
            reg["_obs"] = ("intervalo de " + intervalo[mask_asi].astype("Int64").astype(str)
                           + " dias entre masculino e feminino")
            _registrar_div(reg, f"Intervalo de florescimento maior que {ASI_MAX} dias", obs_col="_obs")
        else:
            det_asi = None
        _card(f"Intervalo de florescimento maior que {ASI_MAX} dias", n_asi,
              "A diferença entre o florescimento masculino e o feminino ficou grande (em qualquer "
              "ordem). Um intervalo muito longo pode indicar estresse ou erro na data." + NOTA_SEM_RESP,
              det_asi)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 4d — CHECAGEM POR AVALIAÇÃO · av4 (produtividade)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Avaliações", "Qualidade da coleta — av4 (produtividade)",
             "Consistência da produtividade e dos componentes de rendimento (colheita).")

PMG_MIN, PMG_MAX = 100, 500        # g, faixa plausível do peso de mil grãos
FIL_MIN, FIL_MAX = 10, 22          # fileiras por espiga
GRF_MIN, GRF_MAX = 15, 50          # grãos por fileira
PROD_PISO = 1000                   # kg/ha; abaixo disso, provável erro

with st.popover("Como interpretar", use_container_width=False):
    st.markdown(f"""
**O que é a av4**

Na av4 (produtividade), colhe-se a parcela e mede-se o **peso**, a **umidade** e a **geometria**
(linhas, comprimento, espaçamento), de onde o pipeline calcula a **produtividade (kg/ha)**
corrigida para umidade padrão. Também se contam o **estande final** (para a população real) e os
**componentes de rendimento** (fileiras, grãos por fileira, peso de mil grãos).

Na mesma subamostra contam-se ainda as **perdas** (plantas acamadas, quebradas, dominadas e com
colmo podre) e os **fenômenos** (green snap, morte prematura, má formação de espigas e
enfezamento). Os dois viram percentual pela mesma régua: a média das contagens dividida pelo
estande final. Contagem **0 é medição** — o avaliador percorreu os 10 metros e não encontrou;
**campo vazio** é subamostra não avaliada e fica fora da conta.

---

**O que o diagnóstico verifica**

- **Produtividade bloqueada:** o pipeline não conseguiu validar a produtividade. O motivo aparece:
  *não colhido* (sem peso), *sem geometria* (faltam as medidas da parcela) ou *produtividade
  impossível* (acima de {PROD_TETO:,} kg/ha).
- **Umidade fora de {UMID_MIN}–{UMID_MAX}%:** umidade de colheita fora da faixa sã — provável erro
  de medição ou digitação.
- **População real não contada:** plot sem estande final — a população real (plantas/ha) não foi
  calculada.
- **PMG ou componentes absurdos:** peso de mil grãos, fileiras ou grãos por fileira fora de faixas
  amplas ({PMG_MIN}–{PMG_MAX} g / {FIL_MIN}–{FIL_MAX} fileiras / {GRF_MIN}–{GRF_MAX} grãos).
- **Produtividade suspeita:** muito baixa (abaixo de {PROD_PISO:,} kg/ha) ou muito alta (acima de
  {PROD_TETO:,} kg/ha) — vale conferir peso, área e umidade.

---

**Como interpretar / o que fazer**

- Confira no aplicativo o peso, a umidade, a geometria e as contagens; corrija o que estiver errado.
- Use os filtros de **responsável** e **fazenda** na sidebar e o botão **Exportar** para acionar o time.
""".replace(",", "."))

av4 = _empilhar("av4_gold", safras_sel)
if filtra_resp and not av4.empty and "nomeResponsavel" in av4.columns:
    av4 = av4[av4["nomeResponsavel"] == resp_sel]
if filtra_tipo and not av4.empty and "tipoTeste" in av4.columns:
    av4 = av4[av4["tipoTeste"] == tipo_sel]

if av4.empty:
    st.info("Sem dados de av4 para este filtro.")
else:
    ctx_av4 = [c for c in ["safra", "cod_fazenda", "nomeFazenda", "nomeResponsavel", "dePara",
                           "tipoTeste", "indexTratamento"] if c in av4.columns]

    def _fmt_flag(f):
        m = {"nao_colhido": "não colhido", "sem_geometria": "sem geometria",
             "prod_impossivel": "produtividade impossível", "umidade_baixa": "umidade baixa",
             "umidade_alta": "umidade alta"}
        return "; ".join(m.get(x.strip(), x.strip()) for x in str(f).split(";") if x.strip())

    # C1: produtividade bloqueada (flags bloqueantes)
    if "flags_produtividade" in av4.columns:
        flags = av4["flags_produtividade"].fillna("").astype(str)
        bloqueantes = ["nao_colhido", "sem_geometria", "prod_impossivel"]
        mask_bloq = flags.apply(lambda f: any(b in f for b in bloqueantes))
        n_bloq = int(mask_bloq.sum())
        if n_bloq:
            cols = ctx_av4 + [c for c in ["produtividade_kg_ha", "pesoParcela", "umidade_pct",
                                          "area_parcela_m2", "flags_produtividade"] if c in av4.columns]
            det_bloq = av4[mask_bloq][[c for c in cols if c in av4.columns]]
            reg = av4[mask_bloq].copy()
            reg["_obs"] = flags[mask_bloq].apply(_fmt_flag)
            _registrar_div(reg, "Produtividade bloqueada", obs_col="_obs")
        else:
            det_bloq = None
        _card("Produtividade bloqueada", n_bloq,
              "O pipeline não validou a produtividade: não colhido (sem peso), sem geometria "
              "(faltam medidas da parcela) ou produtividade impossível." + NOTA_SEM_RESP,
              det_bloq)

    # C2: umidade fora de faixa
    if "umidade_pct" in av4.columns:
        u = pd.to_numeric(av4["umidade_pct"], errors="coerce")
        mask_umid = u.notna() & ((u < UMID_MIN) | (u > UMID_MAX))
        n_umid = int(mask_umid.sum())
        if n_umid:
            cols = ctx_av4 + ["umidade_pct"]
            det_umid = av4[mask_umid][[c for c in cols if c in av4.columns]]
            reg = av4[mask_umid].copy()
            reg["_obs"] = "umidade " + u[mask_umid].round(1).astype(str) + "%"
            _registrar_div(reg, f"Umidade fora de {UMID_MIN}-{UMID_MAX}%", obs_col="_obs")
        else:
            det_umid = None
        _card(f"Umidade fora de {UMID_MIN}–{UMID_MAX}%", n_umid,
              "Umidade de colheita fora da faixa sã — provável erro de medição ou digitação."
              + NOTA_SEM_RESP,
              det_umid)

    # C3: população real não contada (com o motivo: sem contagem, ou sem espaçamento)
    if "populacao_real_plantas_ha" in av4.columns:
        pop = pd.to_numeric(av4["populacao_real_plantas_ha"], errors="coerce")
        mask_pop = pop.isna()
        n_pop = int(mask_pop.sum())
        if n_pop:
            estande = pd.to_numeric(av4["plantas_10m_media"], errors="coerce") if "plantas_10m_media" in av4.columns else pd.Series(np.nan, index=av4.index)
            espac = pd.to_numeric(av4["espacamento"], errors="coerce") if "espacamento" in av4.columns else pd.Series(np.nan, index=av4.index)
            espac_ok = espac.notna() & (espac > 0) & (espac <= 2)
            # motivo: sem contagem de estande, ou estande contado mas sem espaçamento válido
            motivo = pd.Series("outro", index=av4.index)
            motivo[mask_pop & estande.isna()] = "sem contagem de estande"
            motivo[mask_pop & estande.notna() & ~espac_ok] = "sem espaçamento (estande contado)"
            det_pop = av4[mask_pop].copy()
            det_pop["_motivo"] = motivo[mask_pop]
            cols = ["_motivo"] + ctx_av4 + [c for c in ["plantas_10m_media", "espacamento"] if c in av4.columns]
            det_pop = det_pop[[c for c in cols if c in det_pop.columns]]
            reg = av4[mask_pop].copy()
            reg["_obs"] = motivo[mask_pop]
            _registrar_div(reg, "População real não contada", obs_col="_obs")
        else:
            det_pop = None
        _card("População real não contada", n_pop,
              "Plot sem população real (plantas/ha). O motivo aparece na coluna _motivo: sem contagem "
              "de estande, ou estande contado mas sem o espaçamento para calcular." + NOTA_SEM_RESP,
              det_pop)

    # C4: PMG ou componentes absurdos
    faixas = [("pmg_corrigido_g", PMG_MIN, PMG_MAX, "PMG", "g"),
              ("fileiras_media", FIL_MIN, FIL_MAX, "fileiras", ""),
              ("graos_fileira_media", GRF_MIN, GRF_MAX, "grãos/fileira", "")]
    mask_comp = pd.Series(False, index=av4.index)
    obs_comp = pd.Series("", index=av4.index)
    for col, vmin, vmax, nome, un in faixas:
        if col in av4.columns:
            v = pd.to_numeric(av4[col], errors="coerce")
            fora = v.notna() & ((v < vmin) | (v > vmax))
            mask_comp = mask_comp | fora
            obs_comp[fora] = obs_comp[fora].where(obs_comp[fora] == "", obs_comp[fora] + "; ") + \
                nome + "=" + v[fora].round(1).astype(str) + (f" {un}" if un else "")
    n_comp = int(mask_comp.sum())
    if n_comp:
        cols = ctx_av4 + [c for c, *_ in faixas if c in av4.columns]
        det_comp = av4[mask_comp][[c for c in cols if c in av4.columns]]
        reg = av4[mask_comp].copy()
        reg["_obs"] = obs_comp[mask_comp]
        _registrar_div(reg, "PMG ou componentes absurdos", obs_col="_obs")
    else:
        det_comp = None
    _card("PMG ou componentes absurdos", n_comp,
          f"Peso de mil grãos ({PMG_MIN}–{PMG_MAX} g), fileiras ({FIL_MIN}–{FIL_MAX}) ou grãos por "
          f"fileira ({GRF_MIN}–{GRF_MAX}) fora de faixas amplas — provável erro." + NOTA_SEM_RESP,
          det_comp)

    # C5: produtividade suspeita (muito baixa ou muito alta)
    prod_col = "produtividade_kg_ha" if "produtividade_kg_ha" in av4.columns else None
    if prod_col:
        p = pd.to_numeric(av4[prod_col], errors="coerce")
        mask_susp = p.notna() & ((p < PROD_PISO) | (p > PROD_TETO))
        n_susp = int(mask_susp.sum())
        if n_susp:
            cols = ctx_av4 + [c for c in [prod_col, "pesoParcela", "umidade_pct",
                                          "area_parcela_m2"] if c in av4.columns]
            det_susp = av4[mask_susp][[c for c in cols if c in av4.columns]]
            reg = av4[mask_susp].copy()
            reg["_obs"] = p[mask_susp].apply(
                lambda x: ("muito baixa " if x < PROD_PISO else "muito alta ") + f"{x:.0f} kg/ha")
            _registrar_div(reg, "Produtividade suspeita", obs_col="_obs")
        else:
            det_susp = None
        _card("Produtividade suspeita", n_susp,
              f"Produtividade muito baixa (< {PROD_PISO} kg/ha) ou muito alta (> {PROD_TETO} kg/ha) "
              "— vale conferir peso, área e umidade.".replace(str(PROD_TETO), f"{PROD_TETO:,}".replace(",", "."))
              + NOTA_SEM_RESP,
              det_susp)

    # C5b: perda de colheita não avaliada ou avaliada pela metade
    # Só faz sentido depois da correção do pipeline: antes, contagem 0 era descartada e um plot
    # "contado e sem ocorrência" ficava vazio igual a um plot que ninguém contou. Agora vazio
    # significa exclusivamente que a contagem não foi feita, e isso é integridade de dado.
    COMPONENTES_PERDA_CHK = [("pct_acamadas", "acamamento"), ("pct_colmo_podre", "colmo podre"),
                             ("pct_quebradas", "quebramento"), ("pct_dominadas", "dominadas")]
    _cols_pc = [(c, n) for c, n in COMPONENTES_PERDA_CHK if c in av4.columns]
    if _cols_pc:
        _preench = pd.DataFrame(
            {c: pd.to_numeric(av4[c], errors="coerce").notna() for c, _ in _cols_pc},
            index=av4.index)
        _n_ok = _preench.sum(axis=1)
        mask_sem = _n_ok == 0                              # ninguém contou
        mask_parcial = (_n_ok > 0) & (_n_ok < len(_cols_pc))   # contou parte
        mask_pc = mask_sem | mask_parcial
        n_pc = int(mask_pc.sum())
        if n_pc:
            cols_pc = ctx_av4 + [c for c, _ in _cols_pc] + ["pct_perda_total"]
            det_pc = av4[mask_pc][[c for c in cols_pc if c in av4.columns]]
            reg = av4[mask_pc].copy()

            def _obs_pc(r):
                faltam = [n for c, n in _cols_pc
                          if pd.isna(pd.to_numeric(r.get(c), errors="coerce"))]
                if len(faltam) == len(_cols_pc):
                    return "nenhuma das quatro perdas foi contada — perda total fica vazia"
                return ("faltou contar: " + ", ".join(faltam)
                        + " — a perda total soma só o que foi contado e sai subestimada")
            reg["_obs"] = reg.apply(_obs_pc, axis=1)
            _registrar_div(reg, "Perda não avaliada", obs_col="_obs")
        else:
            det_pc = None
        _card("Perda de colheita não avaliada", n_pc,
              "Plots em que nenhuma das quatro perdas foi contada, ou em que só parte delas foi. "
              "No primeiro caso a perda total fica vazia; no segundo ela soma apenas o que foi "
              "contado e sai menor que a real. A coluna de observação diz qual é o caso e o que "
              "faltou." + NOTA_SEM_RESP,
              det_pc)

    # C6: perda total alta — corte DINÂMICO, acompanha a safra em andamento.
    # corte = max(p95 dos plots com perda > 0 na carga atual, piso agronômico de 15%).
    # O p95 recalcula a cada carga; o piso segura falso alarme em safra boa. Na rede 25/26
    # cheia (1457 plots) o p95 dá ~31%, então em fim de safra o corte tende a subir do piso.
    PISO_PERDA = 15
    PERDA_ALERTA, _p95_perda = _corte_dinamico(av4.get("pct_perda_total", pd.Series(dtype=float)),
                                               PISO_PERDA)
    COMPONENTES_PERDA = [("pct_acamadas", "Acamamento"), ("pct_colmo_podre", "Colmo podre"),
                         ("pct_quebradas", "Quebramento"), ("pct_dominadas", "Dominadas")]
    if "pct_perda_total" in av4.columns:
        v_perda = pd.to_numeric(av4["pct_perda_total"], errors="coerce")
        mask_perda = v_perda.notna() & (v_perda > PERDA_ALERTA)
        n_perda = int(mask_perda.sum())
        if n_perda:
            cols_p = ctx_av4 + ["pct_perda_total"] + [c for c, _ in COMPONENTES_PERDA if c in av4.columns]
            det_perda = av4[mask_perda][[c for c in cols_p if c in av4.columns]]
            reg = av4[mask_perda].copy()
            # obs: lista os componentes que puxaram (o maior primeiro)
            def _obs_perda(r):
                partes = []
                for c, nome in COMPONENTES_PERDA:
                    val = pd.to_numeric(r.get(c), errors="coerce")
                    if pd.notna(val) and val > 0:
                        partes.append((val, f"{nome} {val:.1f}%"))
                partes.sort(reverse=True)
                tot = pd.to_numeric(r.get("pct_perda_total"), errors="coerce")
                return f"perda total {tot:.1f}% (" + ", ".join(p for _, p in partes) + ")"
            reg["_obs"] = reg.apply(_obs_perda, axis=1)
            _registrar_div(reg, "Perda total alta", obs_col="_obs")
        else:
            det_perda = None
        _base_txt = (f"p95 da carga = {_p95_perda:g}%" if PERDA_ALERTA > PISO_PERDA
                     else f"piso de {PISO_PERDA}% — p95 da carga ({_p95_perda:g}%) ficou abaixo")
        _card(f"Perda total acima de {PERDA_ALERTA:g}%", n_perda,
              f"Soma das perdas de colheita (acamamento + colmo podre + quebramento + dominadas) "
              f"acima de {PERDA_ALERTA:g}% — corte dinâmico ({_base_txt}). A tabela mostra os "
              f"componentes que compõem cada caso." + NOTA_SEM_RESP,
              det_perda)
        _histograma_corte(av4["pct_perda_total"], PERDA_ALERTA, "%",
                          cor_barra="#1E8449", key="hist_perda")
        _legenda_corte()

        # componentes da perda: um histograma por componente, grade 2 por linha, cada um com
        # seu corte dinâmico. Ajuda a responder "perda de quê?" quando a perda total dispara.
        COMP_PISO = {"pct_acamadas": 8, "pct_colmo_podre": 8,
                     "pct_quebradas": 8, "pct_dominadas": 8}
        COMP_LABEL = {"pct_acamadas": "Acamamento", "pct_colmo_podre": "Colmo podre",
                      "pct_quebradas": "Quebramento", "pct_dominadas": "Dominadas"}
        _comp_com_dado = [c for c in COMP_PISO
                          if c in av4.columns and (pd.to_numeric(av4[c], errors="coerce") > 0).any()]
        if _comp_com_dado:
            st.markdown("<div style='font-size:12px;color:#6B7280;margin:6px 0 -2px 34px;'>"
                        "Componentes da perda — cada um com seu corte dinâmico:</div>",
                        unsafe_allow_html=True)
            _cols_pc = st.columns(min(2, len(_comp_com_dado)))
            for _i, c in enumerate(_comp_com_dado):
                _corte_c, _ = _corte_dinamico(av4[c], COMP_PISO[c])
                with _cols_pc[_i % len(_cols_pc)]:
                    st.markdown(f"<div style='font-size:12px;font-weight:700;color:#374151;"
                                f"margin:10px 0 2px 34px;'>{COMP_LABEL[c]}</div>",
                                unsafe_allow_html=True)
                    _histograma_corte(av4[c], _corte_c, "%",
                                      cor_barra="#1E8449", key=f"hist_comp_{c}")

    # C7: fenômenos da colheita acima do corte de cada um
    # Corte por fenômeno, perto do p95 de cada (régua nova, rede 25/26, 1163 plots avaliados):
    # cada fenômeno tem escala própria — green snap e enfezamento sobem muito, morte prematura
    # e má formação raramente passam de 8%. Um corte único achataria essa diferença e esconderia
    # o green snap anômalo no ruído do enfezamento comum. Valores redondos perto do p95:
    # Piso agronômico por fenômeno (baixo, raramente muda). O corte de fato é
    # max(p95 da carga, piso) — calculado abaixo, então acompanha a safra.
    FENOMENO_PISO = {
        "pct_green_snap":          10,
        "pct_morte_prematura":      5,
        "pct_ma_formacao_espigas":  5,
        "pct_enfezamento":         10,
    }
    FEN_LABEL = {"pct_green_snap": "Green snap", "pct_morte_prematura": "Morte prematura",
                 "pct_ma_formacao_espigas": "Má formação de espigas", "pct_enfezamento": "Enfezamento"}
    fenom_existentes = [(c, FEN_LABEL[c]) for c in FENOMENO_PISO if c in av4.columns]
    # corte dinâmico por fenômeno: {coluna: (corte, p95)}
    FENOMENO_ALERTA = {}
    _fen_p95 = {}
    for c, _ in fenom_existentes:
        FENOMENO_ALERTA[c], _fen_p95[c] = _corte_dinamico(av4[c], FENOMENO_PISO[c])
    if fenom_existentes:
        # cada fenômeno é comparado ao SEU corte
        mask_fen = pd.Series(False, index=av4.index)
        for c, _ in fenom_existentes:
            v = pd.to_numeric(av4[c], errors="coerce")
            mask_fen = mask_fen | (v.notna() & (v > FENOMENO_ALERTA[c]))
        n_fen = int(mask_fen.sum())
        if n_fen:
            cols_f = ctx_av4 + [c for c, _ in fenom_existentes]
            det_fen = av4[mask_fen][[c for c in cols_f if c in av4.columns]]
            reg = av4[mask_fen].copy()
            def _obs_fen(r):
                partes = []
                for c, nome in fenom_existentes:
                    val = pd.to_numeric(r.get(c), errors="coerce")
                    if pd.notna(val) and val > FENOMENO_ALERTA[c]:
                        partes.append((val, f"{nome} {val:.1f}% (corte {FENOMENO_ALERTA[c]:g}%)"))
                partes.sort(reverse=True)
                return ", ".join(p for _, p in partes)
            reg["_obs"] = reg.apply(_obs_fen, axis=1)
            _registrar_div(reg, "Fenômeno de colheita alto", obs_col="_obs")
        else:
            det_fen = None
        _cortes_txt = ", ".join(f"{FEN_LABEL[c]} {v:g}%" for c, v in FENOMENO_ALERTA.items()
                                if c in av4.columns)
        _card("Fenômeno de colheita alto", n_fen,
              "Corte dinâmico por fenômeno (o maior entre o p95 da carga e um piso), porque as "
              "escalas são diferentes e a safra está em andamento. Cortes atuais: "
              + _cortes_txt + ". A tabela mostra quais passaram e o percentual." + NOTA_SEM_RESP,
              det_fen)
        # um histograma por fenômeno — as escalas diferem, então não cabe um gráfico só.
        # só desenha os que têm ao menos um plot com ocorrência no recorte.
        _fen_com_dado = [(c, nome) for c, nome in fenom_existentes
                         if (pd.to_numeric(av4[c], errors="coerce") > 0).any()]
        if _fen_com_dado:
            _cols_h = st.columns(min(2, len(_fen_com_dado)))
            for _i, (c, nome) in enumerate(_fen_com_dado):
                with _cols_h[_i % len(_cols_h)]:
                    st.markdown(f"<div style='font-size:12px;font-weight:700;color:#374151;"
                                f"margin:10px 0 2px 34px;'>{nome}</div>", unsafe_allow_html=True)
                    _histograma_corte(av4[c], FENOMENO_ALERTA[c], "%",
                                      cor_barra="#1E8449", key=f"hist_fen_{c}")
            _legenda_corte()


st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# 5 — COMPILADO PARA WHATSAPP
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo("Compilado", "Mensagem para o time",
             "Divergências do filtro atual em texto pronto para copiar e colar no WhatsApp. "
             "Selecione uma safra e um responsável na barra lateral para gerar a mensagem.")

# a mensagem só faz sentido direcionada: exige UMA safra e UM responsável
_uma_safra = len(safras_sel_nomes) == 1
if not (_uma_safra and filtra_resp):
    faltando = []
    if not _uma_safra:
        faltando.append("marque **uma única safra**")
    if not filtra_resp:
        faltando.append("escolha **um responsável**")
    st.info("Para gerar a mensagem direcionada, na barra lateral: " + " e ".join(faltando) +
            ". Assim a mensagem fica endereçada à pessoa certa, com só os dados dela.")
elif not DIVERGENCIAS:
    st.success("Nenhuma divergência acionável para este responsável nesta safra — nada a enviar.")
else:
    df_div = pd.DataFrame(DIVERGENCIAS)

    # ── Seletor: quais tipos de problema entram na mensagem ──────────────────
    # Com muitas divergências a mensagem vira parede de texto. O responsável quase sempre
    # quer mandar um assunto por vez ("fazendas sem coordenada", "inconsistência de doença"),
    # então filtra-se por TIPO antes de montar. Cada opção mostra a contagem.
    _cont_tipo = df_div["tipo"].value_counts()
    _tipos_todos = sorted(_cont_tipo.index.tolist())
    _label_tipo = {t: f"{t} ({_cont_tipo[t]})" for t in _tipos_todos}

    st.markdown("**O que incluir na mensagem**")
    # inicializa o estado UMA vez com tudo marcado; atalhos abaixo só reescrevem esse estado
    for _t in _tipos_todos:
        st.session_state.setdefault(f"diag_wpp_t_{_t}", True)

    _bt1, _bt2, _sp = st.columns([2, 2, 3])
    if _bt1.button("Marcar todos", use_container_width=True, key="diag_wpp_todos"):
        for _t in _tipos_todos:
            st.session_state[f"diag_wpp_t_{_t}"] = True
        st.rerun()
    if _bt2.button("Limpar", use_container_width=True, key="diag_wpp_limpar"):
        for _t in _tipos_todos:
            st.session_state[f"diag_wpp_t_{_t}"] = False
        st.rerun()

    # um checkbox por tipo, em duas colunas, com a contagem ao lado
    _tipos_sel = []
    _cols_chk = st.columns(2)
    for _i, _t in enumerate(_tipos_todos):
        if _cols_chk[_i % 2].checkbox(_label_tipo[_t], key=f"diag_wpp_t_{_t}"):
            _tipos_sel.append(_t)

    if not _tipos_sel:
        st.info("Selecione ao menos um tipo de problema para montar a mensagem.")
        st.stop()

    df_div = df_div[df_div["tipo"].isin(_tipos_sel)]
    if df_div.empty:
        st.info("Nenhuma divergência nos tipos selecionados.")
        st.stop()

    # cabeçalho da mensagem
    quem = resp_sel if filtra_resp else "equipe"
    safras_txt = ", ".join(safras_sel_nomes)
    linhas = [f"Ola, {quem}. Segue o que precisa de conferencia nos dados (safra {safras_txt}):", ""]

    # agrupa por fazenda → tipo de problema (título); materiais separados por Faixa/Densidade
    # (na Densidade, mostra a população-alvo ao lado do material, e cada material+pop é único)
    def _fmt_pop(p):
        try:
            return f"{int(float(p)):,}".replace(",", ".")
        except (ValueError, TypeError):
            return None

    for fazenda in sorted(df_div["fazenda"].unique()):
        bloco_faz = df_div[df_div["fazenda"] == fazenda]
        for tipo in sorted(bloco_faz["tipo"].unique()):
            bloco = bloco_faz[bloco_faz["tipo"] == tipo]
            linhas.append(f"*{fazenda} — {tipo}*")
            # Faixa antes de Densidade (ordem de ensaio explícita, não alfabética)
            def _ord_ensaio(e):
                el = str(e).lower()
                return 0 if el.startswith("faixa") else (1 if el.startswith("dens") else 2)
            for ensaio in sorted(bloco["ensaio"].unique(), key=_ord_ensaio):
                sub = bloco[bloco["ensaio"] == ensaio]
                linhas.append(f"{ensaio}:")

                def _pref_trat(r):
                    t = r.get("trat")
                    return f"{t} - " if (t and str(t) != "—") else ""

                def _trat_num(r):
                    # índice do tratamento como número, para ordenar certo (10 depois de 2)
                    try:
                        return int(float(r.get("trat")))
                    except (ValueError, TypeError):
                        return 10**9   # sem índice vai para o fim

                if ensaio.lower().startswith("dens"):
                    # índice + material + população; deduplica; ordena por (população, índice)
                    vistos = []
                    itens = []
                    for _, r in sub.iterrows():
                        pop_txt = _fmt_pop(r.get("pop"))
                        base = f"{_pref_trat(r)}{r['material']}"
                        rotulo = f"{base} · {pop_txt} pl/ha" if pop_txt else base
                        chave = (r.get("trat"), r["material"], pop_txt)
                        if chave not in vistos:
                            vistos.append(chave)
                            itens.append((_trat_num(r), r.get("pop"), rotulo))
                    # índice do tratamento primeiro; população desempata quando o índice repete
                    itens.sort(key=lambda x: (x[0], x[1] is None, x[1] if x[1] is not None else 0))
                    for *_ignora, rotulo in itens:
                        linhas.append(f"  {rotulo}")
                else:
                    # Faixa: índice + material, sem repetir; ordena por índice do tratamento
                    vistos = set()
                    itens = []
                    for _, r in sub.iterrows():
                        rotulo = f"{_pref_trat(r)}{r['material']}"
                        if rotulo not in vistos:
                            vistos.add(rotulo)
                            itens.append((_trat_num(r), rotulo))
                    itens.sort(key=lambda x: x[0])
                    for _, rotulo in itens:
                        linhas.append(f"  {rotulo}")
            linhas.append("")

    linhas.append("Podem verificar e corrigir no aplicativo? Qualquer duvida, me chamem. Obrigado!")
    texto_wpp = "\n".join(linhas)

    total = len(df_div)
    n_faz = df_div["fazenda"].nunique()
    _n_todos = len(DIVERGENCIAS)
    _recorte = "" if total == _n_todos else f" (de {_n_todos} no total)"
    st.caption(f"{total} divergência(s){_recorte} em {n_faz} local(is). "
               "Confira a mensagem e use o botão para copiar.")

    # mensagem num card destacado (contorno + fundo suave, estilo bolha) e botão de copiar embaixo
    import html as _html
    import json as _json
    texto_html = _html.escape(texto_wpp).replace("\n", "<br>")
    texto_js = _json.dumps(texto_wpp)   # string segura para o JS
    components.html(f"""
    <div style="font-family:'Helvetica Neue',sans-serif;">
      <div style="border:1.5px solid #C9D6DF;background:#F7FAFC;border-radius:12px;
                  padding:16px 20px;max-height:420px;overflow-y:auto;
                  font-size:13.5px;line-height:1.55;color:#1A1A1A;white-space:normal;
                  box-shadow:0 1px 4px rgba(0,0,0,0.04);">{texto_html}</div>
      <button id="btn-copiar-wpp" onclick='navigator.clipboard.writeText({texto_js}).then(
                 () => {{ const b=document.getElementById("btn-copiar-wpp");
                          b.innerText="Copiado!"; b.style.background="#1E8449";
                          setTimeout(()=>{{b.innerText="Copiar mensagem"; b.style.background="#27AE60";}},1800); }})'
              style="margin-top:12px;padding:10px 22px;border:none;border-radius:8px;
                     background:#27AE60;color:#FFFFFF;font-size:14px;font-weight:700;
                     cursor:pointer;font-family:'Helvetica Neue',sans-serif;">
        Copiar mensagem
      </button>
    </div>
    """, height=min(560, 130 + 18 * min(texto_wpp.count(chr(10)) + 1, 22)))

rodape()
