"""
pages/4_Sanidade.py — Sanidade (milho)

Adaptada da página de Doenças do painel de soja, seguindo Better Data Visualization
(Schwabish). Fonte: tabela_analitica_faixa das safras 2024/25 e 2025/26 (só Faixa).

Cobre av2 (doenças) e av4 (perdas e fenômenos da colheita):
  - doenças: nota 1-9 (9 = mais resistente), classe AS/S/MT/T/R, incidência recalculada aqui;
  - perdas: acamadas, quebradas, dominadas, colmo podre e o total;
  - fenômenos: green snap, morte prematura, má formação de espigas, enfezamento-contagem.

Cuidados de leitura (valem para a página inteira):
  - a nota de doença é INVERSA (9 é bom); perdas e fenômenos são percentuais (alto é ruim).
    Nunca usar a mesma paleta nos dois blocos;
  - enfezamento é medido DUAS vezes (nota na av2, contagem na av4). São escalas e momentos
    diferentes: não somar, não plotar na mesma série;
  - fenômenos só existem em 2025 (ajuste de protocolo). Em 24/25 as colunas não vêm;
  - perda e fenômeno usam a MESMA régua e vêm da MESMA subamostra no app: média das contagens
    ÷ estande final × 100. A contagem 0 conta (o avaliador percorreu os 10 m e não achou); só a
    subamostra não avaliada fica de fora. O percentual é a taxa do plot.

Seções (construídas incrementalmente):
  1. Auditoria — tabela analítica de sanidade por ensaio
"""
import numpy as np
import pandas as pd
import streamlit as st
import plotly.graph_objects as go_plt

# ── Cores por status do material (milho: CHECK / STINE / EXP / DP2) ──────────
COR_STATUS_PLOT = {
    "CHECK": "#F4B184",   # testemunha externa (laranja)
    "STINE": "#2976B6",   # comercial Stine (azul)
    "EXP":   "#00FF00",   # experimental / em avaliação (verde vibrante)
    "DP2":   "#C4DFB4",   # duplo propósito / segundo ano (verde claro)
}
COR_BORDA = {
    "CHECK": "#C46A3A",
    "STINE": "#1A4F7A",
    "EXP":   "#009900",
    "DP2":   "#7AAF6A",
}

# ── Classes de reação a doenças (escala INVERSA: 9 = mais resistente) ────────
ORDEM_CLASS = ["AS", "S", "MT", "T", "R"]
COR_CLASS = {
    "AS": "#8B0000",   # vermelho escuro — altamente suscetível
    "S":  "#E63946",   # vermelho        — suscetível
    "MT": "#FFD600",   # amarelo         — moderadamente tolerante
    "T":  "#70C96E",   # verde claro     — tolerante
    "R":  "#1E7A34",   # verde escuro    — resistente
}
COR_TEXTO_CLASS = {"AS": "#FFFFFF", "S": "#FFFFFF", "MT": "#1A1A1A",
                   "T": "#1A1A1A", "R": "#FFFFFF"}

# Mesma paleta no formato que o openpyxl aceita (RRGGBB, sem "#"). Derivadas dos dicionários
# acima de propósito: a cor da classe no Excel tem que ser a MESMA da tela, e a única forma de
# garantir isso é não redigitar o hex. Mudar COR_CLASS muda tela e export juntos.
COR_CLASS_XL = {k: v.lstrip("#") for k, v in COR_CLASS.items()}
COR_TEXTO_CLASS_XL = {k: v.lstrip("#") for k, v in COR_TEXTO_CLASS.items()}


def _eh_col_classe(nome) -> bool:
    """A coluna carrega sigla de classe? Cobre os dois padrões usados na página:
    'Classe Ferrugem' (tabelas exibidas) e 'FER_classe' (base da apresentação)."""
    n = str(nome).strip().lower()
    return n == "classe" or n.startswith("classe ") or n.endswith("_classe")


# ── Cor por FAIXA de valor (colunas sem sigla, mas coloridas na tela) ──────────
# Escritas uma vez como dado: o cellStyle do AgGrid é GERADO daqui e o Excel lê a mesma tabela.
# Cada regra é (operador, limite, cor), avaliada em ordem; a primeira que casa vence.
#   operador: ">=", ">", "<=", "<", "==", "contem" (substring, para colunas de texto),
#             ou None para "vale sempre" (regra final, sem limite)
#   cor:      sigla de classe (usa COR_CLASS/COR_TEXTO_CLASS) ou par literal (fundo, texto)
# Sem regra final, valor que não casa fica SEM cor — é o caso das contagens do delta, em que
# zero não é bom nem ruim e pintar seria dar significado a ausência de ocorrência.
FAIXAS_COR = {
    "Top 25% do local (%)": [(">=", 75, "R"), (">=", 50, "T"), (">=", 25, "MT"),
                             (None, None, "S")],
    "Nota":                 [(">=", 7, "T"), (">=", 5, "MT"), (None, None, "S")],
    # delta vs referência: contagens em tom pastel (leitura de apoio) e o delta médio em
    # cor cheia (é o número que ordena a tabela)
    "▲ Acima ref":          [(">", 0, ("#D5F5D5", "#1A1A1A"))],
    "▼ Abaixo ref":         [(">", 0, ("#FDDCDE", "#1A1A1A"))],
    "Delta médio":          [(">", 0, "R"), ("<", 0, "S"), (None, None, "MT")],
    # evolução por safra: consistência em cor cheia; tendência por TEXTO ("↑ melhora" /
    # "↓ piora"), em pastel — "→ estável" e "—" ficam sem cor de propósito
    "Consistência %":       [(">=", 75, "R"), (">=", 50, "T"), (">=", 25, "MT"),
                             (None, None, "S")],
    "Tendência":            [("contem", "melhora", ("#D5F5D5", "#1A1A1A")),
                             ("contem", "piora", ("#FDDCDE", "#1A1A1A"))],
    # mapa de colapso: as contagens usam o verde pastel como "nada a relatar" e a cor cheia
    # quando há ocorrência; os dois percentuais têm a mesma escala de 20/50/80
    "S / AS (≤4)":          [(">", 0, "S"), (None, None, ("#D5F5D5", "#1A1A1A"))],
    "MT (5–6)":             [(">", 0, "MT")],
    "T / R (≥7)":           [(">", 0, "T")],
    "% Colapso":            [(">=", 80, "AS"), (">=", 50, "S"), (">=", 20, "MT"),
                             (None, None, ("#D5F5D5", "#1A1A1A"))],
    "% Incidência":         [(">=", 80, "AS"), (">=", 50, "S"), (">=", 20, "MT"),
                             (None, None, ("#D5F5D5", "#1A1A1A"))],
}

_OPS_FAIXA = {">=": lambda v, l: v >= l, ">": lambda v, l: v > l,
              "<=": lambda v, l: v <= l, "<": lambda v, l: v < l,
              "==": lambda v, l: v == l}


def _par_cor(cor):
    """Resolve a cor da regra em (fundo, texto), com '#'. Aceita sigla de classe ou par literal."""
    if isinstance(cor, (tuple, list)):
        return cor[0], cor[1]
    return COR_CLASS[cor], COR_TEXTO_CLASS.get(cor, "#1A1A1A")


def cor_da_faixa(valor, faixas):
    """(fundo, texto) para o valor, ou None se nenhuma regra casar.
    Regras numéricas exigem valor numérico; a regra `contem` trabalha sobre o texto."""
    if valor is None:
        return None
    try:
        v = float(valor)
        if v != v:                  # NaN
            return None
    except (TypeError, ValueError):
        v = None
    _txt = str(valor)
    for op, limite, cor in faixas:
        if op is None:
            return _par_cor(cor)
        if op == "contem":
            if str(limite) in _txt:
                return _par_cor(cor)
            continue
        if v is not None and _OPS_FAIXA[op](v, limite):
            return _par_cor(cor)
    return None


def js_faixa(faixas):
    """cellStyle do AgGrid gerado a partir de `faixas` — NOVA instância a cada chamada,
    porque reusar o mesmo objeto JsCode em duas colunas perde o estilo."""
    _linhas = []
    for op, limite, cor in faixas:
        _bg, _fg = _par_cor(cor)
        _ret = f"return {{ background: '{_bg}', color: '{_fg}', fontWeight: '700' }};"
        if op is None:
            _cond = None
        elif op == "contem":
            _cond = f"String(v).indexOf('{limite}') >= 0"
        else:
            _cond = f"v {op} {limite}"
        _linhas.append(f"                {_ret}" if _cond is None
                       else f"                if ({_cond}) {_ret}")
    return JsCode("            function(p) {\n"
                  "                var v = p.value;\n"
                  "                if (v === null || v === undefined || v === '') return {};\n"
                  + "\n".join(_linhas) + "\n"
                  "                return {};\n"
                  "            }\n")


LABEL_CLASS = {
    "AS": "AS — Altamente suscetível (nota 1–2)",
    "S":  "S — Suscetível (3–4)",
    "MT": "MT — Moderadamente tolerante (5–6)",
    "T":  "T — Tolerante (7–8)",
    "R":  "R — Resistente (9)",
}


def nota_para_classe(nota):
    """Converte nota 1-9 em sigla de classe. Mesma régua do pipeline (_classificar_doenca)."""
    if nota is None or (isinstance(nota, float) and np.isnan(nota)):
        return None
    n = float(nota)
    if n <= 2:
        return "AS"
    if n <= 4:
        return "S"
    if n <= 6:
        return "MT"
    if n <= 8:
        return "T"
    return "R"


def resumo_doenca(grp: pd.DataFrame, col_nota: str) -> tuple:
    """Nota típica, incidência e classe de UMA doença num recorte de plots.

    Fonte única das Seções 2 e 4 — se cada uma calculasse por conta, uma hora divergiriam.
      - Nota  = moda das notas válidas; em caso de EMPATE fica a MENOR nota (lado suscetível);
      - Inc.  = locais com detecção (nota 1–5) ÷ locais em que a doença foi avaliada;
      - Classe = régua do pipeline aplicada à nota.
    Nota 0 é "não avaliado" e sai de todas as contas. Devolve (None, None, None) sem dado."""
    if col_nota not in grp.columns:
        return None, None, None
    g = grp[["cod_fazenda", col_nota]].copy()
    g[col_nota] = pd.to_numeric(g[col_nota], errors="coerce").where(lambda x: x > 0)
    g = g.dropna(subset=[col_nota])
    if g.empty:
        return None, None, None
    nota = round(float(g[col_nota].mode().min()), 1)
    n_aval = g["cod_fazenda"].nunique()
    # o local conta como "teve" se ao menos um plot dele detectou a doença
    n_com = g.loc[g[col_nota].between(1, 5), "cod_fazenda"].nunique()
    inc = round(n_com / n_aval * 100, 1) if n_aval else None
    return nota, inc, nota_para_classe(nota)


# ── Doenças da av2 (6) ───────────────────────────────────────────────────────
# `inc` não vem do pipeline (não está em MET_AV2): é recalculado aqui a partir da nota.
DOENCAS = {
    "Turcicum":          {"nota": "nota_turcicum",          "class": "class_nota_turcicum"},
    "Cercospora":        {"nota": "nota_cercospora",        "class": "class_nota_cercospora"},
    "Mancha branca":     {"nota": "nota_mancha_branca",     "class": "class_nota_mancha_branca"},
    "Bipolaris":         {"nota": "nota_bipolaris",         "class": "class_nota_bipolaris"},
    "Ferrugem tropical": {"nota": "nota_ferrugem_tropical", "class": "class_nota_ferrugem_tropical"},
    "Enfezamento":       {"nota": "nota_enfezamento",       "class": "class_nota_enfezamento"},
}

# ── Perdas da av4 (média das subamostras avaliadas ÷ estande final; somam no total) ──
PERDAS = {
    "Acamamento":  "pct_acamadas",
    "Quebramento": "pct_quebradas",
    "Dominadas":   "pct_dominadas",
    "Colmo Podre": "pct_colmo_podre",
}
PERDA_TOTAL = "pct_perda_total"

# ── Fenômenos da av4 (mesma régua e mesma subamostra das perdas; só 2025) ──
# NÃO somam com pct_perda_total — denominador diferente.
FENOMENOS = {
    "Green snap":             {"pct": "pct_green_snap",           "n": "green_snap_plantas"},
    "Morte prematura":        {"pct": "pct_morte_prematura",      "n": "morte_prematura_plantas"},
    "Má formação de espigas": {"pct": "pct_ma_formacao_espigas",  "n": "ma_formacao_espigas_plantas"},
    "Enfezamento (contagem)": {"pct": "pct_enfezamento",          "n": "enfezamento_plantas"},
}

# ── Sanidade medida só em 2024 (mudou de instrumento em 2025) ────────────────
# tombamento verde virou a contagem de green snap na av4; grãos ardidos saiu do protocolo.
LEGADO_2024 = {
    "Nota tombamento verde (24/25)": "nota_tombamento_verde",
    "Grãos ardidos % (24/25)":       "graos_ardidos_pct",
}

from utils.theme import aplicar_tema, page_header, secao_titulo, rodape
from utils.loader import carregar_multisafra
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode

st.set_page_config(
    page_title="Sanidade · JAUM DTC",
    page_icon="🩺",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_tema()

# ── Helper AgGrid (mesma configuração da soja: header escuro, menu funcionando) ─
def ag_table(df, height=400, estilos_col=None, renderers_col=None):
    """Tabela AgGrid padrão. `estilos_col`: {coluna: JsCode} para cellStyle ·
    `renderers_col`: {coluna: JsCode} para cellRenderer (permite barras na célula)."""
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "14px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    for _col, _estilo in (estilos_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellStyle=_estilo)
    for _col, _rend in (renderers_col or {}).items():
        if _col in df.columns:
            gb.configure_column(_col, cellRenderer=_rend, minWidth=170)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        suppressContextMenu=False, enableRangeSelection=True,
    )
    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    go["onFirstDataRendered"] = JsCode("function(params) { params.api.sizeColumnsToFit(); }")
    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                   {"background-color": "#4A4A4A !important"},
            ".ag-header-row":               {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":              {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":        {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":         {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                     {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":  {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-icon-menu":                {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # sem "color" aqui: um !important sobreporia a cor inline do cellStyle
            # (era o que impedia o verde/vermelho dos desvios de aparecer)
            ".ag-cell":                     {"font-size": "13px !important"},
            ".ag-row":                      {"font-size": "13px !important"},
        },
        theme="streamlit", use_container_width=True,
    )


# ── Helper exportar Excel (mesmo padrão da soja) ──────────────────────────────
def exportar_excel(df, nome_arquivo="tabela.xlsx", label="⬇️ Exportar Excel", key=None,
                   faixas_cor=None):
    """Export com as cores da tela. `faixas_cor`: {coluna: faixas} para colunas numéricas
    coloridas por limite (ver FAIXAS_COR); as colunas de classe são detectadas pelo nome."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)

    # O AgGrid ACRESCENTA `::auto_unique_id::` ao DataFrame que recebe, no próprio objeto. Como
    # o export vem depois da tabela na página, a coluna interna vazava para o Excel. Descarta
    # aqui, no único lugar por onde todos os exports passam, em vez de em cada chamada.
    df = df.drop(columns=[c for c in df.columns if str(c).startswith("::")], errors="ignore")

    faixas_cor = faixas_cor or {}

    # ALINHAMENTO como no AgGrid: texto à esquerda, número à direita — e o cabeçalho segue a
    # coluna, senão o rótulo "Consistência %" fica solto sobre números encostados na borda.
    _num = [bool(pd.api.types.is_numeric_dtype(df[c])) for c in df.columns]
    _alin = ["right" if n else "left" for n in _num]

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        # "= Empate" começa com "=" e o openpyxl o gravava como FÓRMULA (data_type 'f'): no Excel
        # virava fórmula inválida e, ao reabrir, o cabeçalho aparecia vazio ("Unnamed"). Forçar
        # texto resolve para "=", "+", "-" e "@", que são os inícios que o Excel interpreta.
        if str(col).lstrip()[:1] in ("=", "+", "-", "@"):
            cell.data_type = "s"
        # cabeçalho escuro com texto branco: é o #4A4A4A do custom_css de TODAS as tabelas do
        # painel (.ag-header), então o arquivo abre com a mesma cara da tela
        cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="4A4A4A")
        cell.alignment = Alignment(horizontal=_alin[ci - 1], vertical="center", wrap_text=True)
        cell.border = border
        # largura pelo conteúdo real (cabeçalho e valores), com piso e teto para não desmontar
        # o layout com um nome de híbrido longo
        _maior = max([len(str(col))] +
                     [len(str(v)) for v in df[col].head(200).tolist() if v is not None])
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(40, max(11, _maior + 3))
    ws.row_dimensions[1].height = 30

    # quais colunas carregam sigla de classe — resolvido uma vez, fora do laço de células
    _col_classe = [_eh_col_classe(c) for c in df.columns]

    for ri, row_data in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, 1):
            try:
                if val is None or (isinstance(val, float) and np.isnan(val)):
                    val = None
                elif type(val).__name__ in ('NAType', 'NaTType'):
                    val = None
                elif pd.isna(val):
                    val = None
            except (TypeError, ValueError):
                pass
            cell = ws.cell(row=ri, column=ci, value=val)
            if isinstance(val, str) and val.lstrip()[:1] in ("=", "+", "-", "@"):
                cell.data_type = "s"          # mesma blindagem de fórmula do cabeçalho
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal=_alin[ci - 1], vertical="center")
            cell.border = border
            # classe com a MESMA cor da tela: fundo = categoria (AS vermelho-escuro → R verde),
            # negrito e cor de texto do dicionário, como no cellStyle do AgGrid. Só pinta quando
            # o valor é uma das cinco siglas — assim "—" e vazio continuam sem fundo.
            _cls = str(val).strip() if val is not None else ""
            if _col_classe[ci - 1] and _cls in COR_CLASS_XL:
                cell.fill = PatternFill("solid", start_color=COR_CLASS_XL[_cls])
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=COR_TEXTO_CLASS_XL.get(_cls, "1A1A1A"))
            # colunas coloridas por faixa de valor (Top 25%, Nota, Delta médio, contagens): as
            # regras vêm da MESMA tabela que gera o cellStyle do AgGrid, então não podem divergir
            elif df.columns[ci - 1] in faixas_cor:
                _par = cor_da_faixa(val, faixas_cor[df.columns[ci - 1]])
                if _par:
                    cell.fill = PatternFill("solid", start_color=_par[0].lstrip("#"))
                    cell.font = Font(name="Arial", size=10, bold=True,
                                     color=_par[1].lstrip("#"))

    # o painel fixa o híbrido à esquerda (pinned) e o cabeçalho no topo; e o filtro do AgGrid
    # vira o autofiltro do Excel, para a tabela ser navegável do mesmo jeito
    ws.freeze_panes = "B2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(buf)
    buf.seek(0)
    st.download_button(label=label, data=buf, file_name=nome_arquivo,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)


# ── Header ────────────────────────────────────────────────────────────────────
page_header(
    "Sanidade",
    "Doenças na lavoura (av2) e o que apareceu na colheita (av4) — perdas e fenômenos. "
    "Comece conferindo os dados por ensaio na Auditoria; as análises agregadas vêm em seguida.",
    imagem="Researchers-pana.png",
)

# ── Carregamento: analítica de Faixa das duas safras, já reconciliada ─────────
# Usa carregar_multisafra(), que aplica o depara_mestre — sem ele, o mesmo híbrido
# aparece com nomes diferentes em cada safra.
@st.cache_data(show_spinner=False)
def carregar_concat():
    d = carregar_multisafra()
    df = d.get("tabela_analitica_faixa")
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    df = df.copy()
    # produtividade canônica: usa a válida (kg/ha e sacas), com fallback para a bruta
    if "produtividade_valida_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_valida_kg_ha"], errors="coerce")
        if "produtividade_kg_ha" in df.columns:
            df["kg_ha"] = df["kg_ha"].fillna(pd.to_numeric(df["produtividade_kg_ha"], errors="coerce"))
    elif "produtividade_kg_ha" in df.columns:
        df["kg_ha"] = pd.to_numeric(df["produtividade_kg_ha"], errors="coerce")
    if "produtividade_valida_sacas_ha" in df.columns:
        df["sc_ha"] = pd.to_numeric(df["produtividade_valida_sacas_ha"], errors="coerce")
    elif "kg_ha" in df.columns:
        df["sc_ha"] = (df["kg_ha"] / 60).round(1)   # 1 saca = 60 kg
    # altura em metros (o pipeline entrega em cm) — mesma régua da Análise Conjunta
    if "altura_planta_cm" in df.columns:
        df["altura_planta_m"] = (pd.to_numeric(df["altura_planta_cm"], errors="coerce") / 100).round(1)
    if "altura_espiga_cm" in df.columns:
        df["altura_espiga_m"] = (pd.to_numeric(df["altura_espiga_cm"], errors="coerce") / 100).round(1)
    return df


with st.spinner("Carregando dados..."):
    ta_raw = carregar_concat()

if ta_raw.empty:
    st.error("Nenhum dado disponível. Verifique a página de Diagnóstico.")
    st.stop()

# Só interessam os plots com ALGUMA medida de sanidade (doença, perda ou fenômeno).
# Um plot só com produtividade não tem o que auditar aqui.
COLS_SANIDADE = ([d["nota"] for d in DOENCAS.values()]
                 + list(PERDAS.values()) + [PERDA_TOTAL]
                 + [f["pct"] for f in FENOMENOS.values()]
                 + list(LEGADO_2024.values()))
_cols_san_existentes = [c for c in COLS_SANIDADE if c in ta_raw.columns]
if _cols_san_existentes:
    ta_raw = ta_raw[ta_raw[_cols_san_existentes].notna().any(axis=1)].copy()

if ta_raw.empty:
    st.error("Nenhum registro com dados de sanidade. Verifique a página de Diagnóstico.")
    st.stop()


# ── Sidebar — filtros encadeados (prefixo `sn_`, independente das outras páginas) ─
with st.sidebar:
    st.markdown(
        '<p style="font-size:11px;font-weight:600;color:#6B7280;text-transform:uppercase;'
        'letter-spacing:0.05em;padding:0.5rem;">Filtros</p>', unsafe_allow_html=True)

    if st.button("Limpar filtros", use_container_width=True, key="sn_btn_limpar"):
        for key in list(st.session_state.keys()):
            if any(key.startswith(p) for p in ["sn_safra_", "sn_macro_", "sn_micro_", "sn_estado_",
                                               "sn_cidade_", "sn_fazenda_", "sn_resp_", "sn_status_",
                                               "sn_cult_", "sn_busca_", "__opts_sn_"]):
                del st.session_state[key]
        st.rerun()

    def _podar_keys(prefix, opcoes, molde):
        """Remove o estado de checkboxes de opções que saíram da cascata.
        Sem isso, ao reaparecerem elas voltam marcadas e o filtro se reaplica sozinho."""
        antigas = st.session_state.get(f"__opts_{prefix}", [])
        atuais = set(map(str, opcoes))
        for o in antigas:
            if str(o) not in atuais:
                st.session_state.pop(molde(o), None)
        st.session_state[f"__opts_{prefix}"] = list(opcoes)

    def checkboxes(opcoes, default_all=True, defaults=None, prefix=""):
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_{o}")
        sel = []
        for o in opcoes:
            key = f"{prefix}_{o}"
            if key in st.session_state:          # já existe: o widget é a verdade
                marcado = st.checkbox(str(o), key=key)
            else:                                 # primeira criação: aplica o padrão
                checked = (o in defaults) if defaults is not None else default_all
                marcado = st.checkbox(str(o), value=checked, key=key)
            if marcado:
                sel.append(o)
        return sel

    def filtro_busca(opcoes, prefix):
        """Busca textual + seleção persistente. A seleção é lida direto dos checkboxes
        (fonte única), inclusive dos itens ocultos pela busca."""
        if f"{prefix}_reset" not in st.session_state:
            st.session_state[f"{prefix}_reset"] = 0
        r = st.session_state[f"{prefix}_reset"]
        _podar_keys(prefix, opcoes, lambda o: f"{prefix}_chk_{r}_{o}")

        busca = st.text_input("Buscar", value="", key=f"sn_busca_{prefix}",
                              placeholder="Digite para filtrar...")
        filtradas = [c for c in opcoes if busca.strip().lower() in str(c).lower()] if busca.strip() else opcoes

        if st.button("Limpar seleção", key=f"{prefix}_limpar", use_container_width=True):
            for o in opcoes:
                st.session_state.pop(f"{prefix}_chk_{r}_{o}", None)
            st.session_state.pop(f"__opts_{prefix}", None)
            st.session_state[f"{prefix}_reset"] = r + 1
            st.rerun()

        for c in filtradas:
            st.checkbox(str(c), key=f"{prefix}_chk_{r}_{c}")

        sel = [o for o in opcoes if st.session_state.get(f"{prefix}_chk_{r}_{o}", False)]
        return sel or opcoes

    # 1. Safra — padrão safra atual (na base o valor é "25/26")
    with st.expander("Safra", expanded=True):
        safras_all = sorted(ta_raw["safra"].dropna().unique().tolist())
        SAFRA_PADRAO = ("25/26", "2025/26")   # aceita os dois formatos
        safra_default = [s for s in safras_all if str(s) in SAFRA_PADRAO] or safras_all[-1:]
        # inicializa o estado UMA vez: garante só a safra padrão marcada
        if "sn_safra_init" not in st.session_state:
            for o in safras_all:
                st.session_state[f"sn_safra_{o}"] = (o in safra_default)
            st.session_state["sn_safra_init"] = True
        safras_sel = checkboxes(safras_all, defaults=safra_default, prefix="sn_safra")
    ta_f1 = ta_raw[ta_raw["safra"].isin(safras_sel)] if safras_sel else ta_raw.iloc[0:0]

    # 2. Região Macro
    with st.expander("Região Macro", expanded=False):
        macros_sel = checkboxes(sorted(ta_f1["regiao_macro"].dropna().unique().tolist()), prefix="sn_macro")
    ta_f2 = ta_f1[ta_f1["regiao_macro"].isin(macros_sel)] if macros_sel else ta_f1.iloc[0:0]

    # 3. Região Micro
    with st.expander("Região Micro", expanded=False):
        micros_sel = checkboxes(sorted(ta_f2["regiao_micro"].dropna().unique().tolist()), prefix="sn_micro")
    ta_f3 = ta_f2[ta_f2["regiao_micro"].isin(micros_sel)] if micros_sel else ta_f2.iloc[0:0]

    # 4. Estado
    with st.expander("Estado", expanded=False):
        estados_sel = filtro_busca(sorted(ta_f3["estado_sigla"].dropna().unique().tolist()), "sn_estado")
    ta_f4 = ta_f3[ta_f3["estado_sigla"].isin(estados_sel)] if estados_sel else ta_f3.iloc[0:0]

    # 5. Cidade
    with st.expander("Cidade", expanded=False):
        cidades_sel = filtro_busca(sorted(ta_f4["cidade_nome"].dropna().unique().tolist()), "sn_cidade")
    ta_f5 = ta_f4[ta_f4["cidade_nome"].isin(cidades_sel)] if cidades_sel else ta_f4.iloc[0:0]

    # 6. Fazenda
    with st.expander("Fazenda", expanded=False):
        fazendas_sel = filtro_busca(sorted(ta_f5["nomeFazenda"].dropna().unique().tolist()), "sn_fazenda")
    ta_f6 = ta_f5[ta_f5["nomeFazenda"].isin(fazendas_sel)] if fazendas_sel else ta_f5.iloc[0:0]

    # 7. Responsável
    with st.expander("Responsável", expanded=False):
        resps_sel = filtro_busca(sorted(ta_f6["nomeResponsavel"].dropna().unique().tolist()), "sn_resp")
    ta_f7 = ta_f6[ta_f6["nomeResponsavel"].isin(resps_sel)] if resps_sel else ta_f6.iloc[0:0]

    # 8. Status do híbrido
    with st.expander("Status do Híbrido", expanded=False):
        status_sel = checkboxes(sorted(ta_f7["status_material"].dropna().unique().tolist()), prefix="sn_status")
    ta_f8 = ta_f7[ta_f7["status_material"].isin(status_sel)] if status_sel else ta_f7.iloc[0:0]

    # 9. Híbrido (dePara)
    with st.expander("Híbrido", expanded=False):
        materiais_sel = filtro_busca(sorted(ta_f8["dePara"].dropna().unique().tolist()), "sn_cult")
    ta_filtrado = ta_f8[ta_f8["dePara"].isin(materiais_sel)] if materiais_sel else ta_f8.iloc[0:0]


def _aplicar_filtros_local(df):
    """Aplica só os filtros de tempo e geografia (nunca híbrido/status) — a referência de um
    ensaio existe independentemente de quais híbridos foram selecionados. Base da produção
    relativa POR LOCAL com referência fixa, igual à Análise Conjunta."""
    if df.empty:
        return df
    d = df
    for _col, _sel in [("safra", safras_sel), ("regiao_macro", macros_sel),
                       ("regiao_micro", micros_sel), ("estado_sigla", estados_sel),
                       ("cidade_nome", cidades_sel), ("nomeFazenda", fazendas_sel),
                       ("nomeResponsavel", resps_sel)]:
        if _sel and _col in d.columns:
            d = d[d[_col].isin(_sel)]
    return d


if ta_filtrado.empty:
    st.warning("Nenhum dado para os filtros selecionados.")
    st.stop()


# ── Contexto: filtros ativos (usado nos subtítulos das seções) ────────────────
def _linha_safra(s, g):
    """Uma linha de contexto por safra, com as regiões/estados que ela realmente contém."""
    partes = [f"<b>Safra:</b> {s}"]
    if "regiao_macro" in g.columns and g["regiao_macro"].notna().any():
        partes.append("Macro: " + ", ".join(sorted(g["regiao_macro"].dropna().unique())))
    if "regiao_micro" in g.columns and g["regiao_micro"].notna().any():
        partes.append("Micro: " + ", ".join(sorted(g["regiao_micro"].dropna().unique())))
    if "estado_sigla" in g.columns and g["estado_sigla"].notna().any():
        partes.append(", ".join(sorted(g["estado_sigla"].dropna().unique())))
    partes.append(f"{g['cidade_nome'].nunique()} cidades")
    partes.append(f"{g['cod_fazenda'].nunique()} locais")
    return " · ".join(partes)


_grupos_safra = list(ta_filtrado.groupby("safra"))
_linhas_safra = [_linha_safra(s, g) for s, g in _grupos_safra]
if len(_grupos_safra) > 1:
    _linhas_safra = ["<b>Análise multissafra</b>"] + _linhas_safra
contexto_str = "<br>".join(_linhas_safra)


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 1 — AUDITORIA (mesmo mapa da Análise Conjunta + densidade + sanidade)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Auditoria",
    "Quais são os dados por ensaio?",
    "Visão individual de cada observação — produção, população, arquitetura, doenças, perdas e "
    "fenômenos na mesma linha. Use para conferência antes das análises agregadas.",
)

with st.popover("ℹ️ Como ler esta tabela", use_container_width=False):
    st.markdown("""
**📌 As famílias de variável, e por que não se misturam**

**Doenças (av2)** — nota de **1 a 9**, onde **9 é o melhor**: a planta reagiu bem, quase não
mostrou sintoma. Nota baixa é planta doente. A escala é o contrário do que a intuição sugere,
então vale conferir sempre antes de concluir.

- **Classe**: AS altamente suscetível (1–2) · S suscetível (3–4) · MT moderadamente tolerante
  (5–6) · T tolerante (7–8) · R resistente (9).
- **Teve a doença?** "Sim" quando a nota ficou entre 1 e 5, "Não" entre 6 e 9. O corte entre 5 e
  6 é uma convenção — perto da fronteira, olhe a nota, não o Sim/Não.
- **Nota 0 significa "não avaliado"**, não "sem doença". O pipeline já transforma em vazio.

**Perdas e fenômenos (av4)** — vêm da mesma subamostra no aplicativo e usam a mesma régua: em
cada trecho de 10 metros conta-se quantas plantas têm o problema, isso vira um percentual sobre
o estande daquele trecho, e o valor do plot é a **média** desses percentuais. As quatro perdas
somam no **Perda Total**; fenômeno não entra nessa soma.

**Zero e vazio são coisas diferentes.** Zero é medição: o avaliador percorreu os 10 metros e não
encontrou nenhuma planta com o problema. Vazio é subamostra não avaliada, e fica fora da média.

**Enfezamento aparece duas vezes** — como nota na av2 e como contagem na av4. São medidas
diferentes do mesmo problema, em momentos diferentes.

**População** — `Pop. Real` é o estande contado no campo. População muito abaixo da esperada
para o ensaio é falha de estande, e costuma explicar tanto perda quanto nota de doença.
""")

# ── Seletor de base da Produção Relativa (idêntico ao da Análise Conjunta) ────
col_ref, col_test, _ = st.columns([2, 2, 3])
with col_ref:
    base_rel = st.selectbox(
        "Base da Produção Relativa",
        options=["Média geral do ensaio", "Maior produtividade", "Testemunha"],
        index=0, key="sn_base_rel",
    )
with col_test:
    if base_rel == "Testemunha":
        testemunhas = sorted(
            ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"].dropna().unique().tolist())
        testemunha_sel = st.selectbox("Selecione a testemunha", options=testemunhas,
                                      key="sn_testemunha") if testemunhas else None
        if not testemunhas:
            st.warning("Nenhuma testemunha disponível nos filtros atuais.")
    else:
        testemunha_sel = None

# ── Produção relativa (%) — POR LOCAL, com referência FIXA do ensaio inteiro ──
#    A referência de cada local (média/máximo/testemunha) é calculada sobre TODOS os híbridos do
#    ensaio (ta_raw + só filtros de local), NÃO sobre os filtrados. Filtrar por status/híbrido
#    muda só o que aparece, não o "100% do local". Mesma regra da Análise Conjunta.
#    LOCAL = (safra, cod_fazenda) para não juntar safras de um mesmo local num recorte multissafra.
df_tabela = ta_filtrado.copy()
LOCAL = [c for c in ["safra", "cod_fazenda"] if c in df_tabela.columns] or ["cod_fazenda"]

_ref_scope = _aplicar_filtros_local(ta_raw)
_ref_scope = _ref_scope[pd.to_numeric(_ref_scope["kg_ha"], errors="coerce") > 0]

if base_rel == "Maior produtividade":
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].max()
elif base_rel == "Testemunha" and testemunha_sel:
    _ref = (_ref_scope[_ref_scope["dePara"] == testemunha_sel].groupby(LOCAL)["kg_ha"].mean())
else:  # "Média geral do ensaio" (padrão)
    _ref = _ref_scope.groupby(LOCAL)["kg_ha"].mean()

_chave = df_tabela.set_index(LOCAL).index
ref_por_local = pd.Series(_chave.map(_ref).to_numpy(), index=df_tabela.index)
df_tabela["prod_relativa_pct"] = ((df_tabela["kg_ha"] / ref_por_local) * 100).round(1)

# ── Doenças: nota (1 casa) + "teve?" recalculado + classe ────────────────────
# A incidência não vem do pipeline (não está em MET_AV2); é derivada da nota aqui.
rename_doencas = {}
for _nome, _cols in DOENCAS.items():
    _cn, _cc = _cols["nota"], _cols["class"]
    if _cn not in df_tabela.columns:
        continue
    _nota = pd.to_numeric(df_tabela[_cn], errors="coerce").where(lambda x: x > 0)  # 0 = não avaliado
    df_tabela[_cn] = _nota.round(1)
    df_tabela[f"inc_{_cn}"] = np.where(_nota.isna(), None,
                                       np.where(_nota.between(1, 5), "Sim", "Não"))
    if _cc not in df_tabela.columns:
        df_tabela[_cc] = _nota.apply(nota_para_classe)
    rename_doencas.update({_cn: f"Nota {_nome}",
                           f"inc_{_cn}": f"Teve {_nome}?",
                           _cc: f"Classe {_nome}"})

# ── Fenômenos da av4 (contagem + percentual; não existem em 24/25) ───────────
rename_fenomenos = {}
for _nome, _cols in FENOMENOS.items():
    _cp, _cq = _cols["pct"], _cols["n"]
    if _cq in df_tabela.columns:
        df_tabela[_cq] = pd.to_numeric(df_tabela[_cq], errors="coerce").round(0).astype("Int64")
        rename_fenomenos[_cq] = f"{_nome} (plantas)"
    if _cp in df_tabela.columns:
        df_tabela[_cp] = pd.to_numeric(df_tabela[_cp], errors="coerce").round(1)
        rename_fenomenos[_cp] = f"{_nome} (%)"

# ── Sanidade só medida em 24/25 (mudou de instrumento em 2025) ───────────────
if "nota_tombamento_verde" in df_tabela.columns:
    df_tabela["nota_tombamento_verde"] = (pd.to_numeric(df_tabela["nota_tombamento_verde"],
                                                        errors="coerce").where(lambda x: x > 0).round(1))

# ── Mapa de colunas: base da Análise Conjunta + densidade + sanidade ─────────
col_map = {
    # ── base idêntica à Auditoria da Análise Conjunta (mesma ordem, mesmos rótulos) ──
    "safra":                      "Safra",
    "cod_fazenda":                "Cód. Local",
    "nomeFazenda":                "Fazenda",
    "cidade_nome":                "Cidade",
    "estado_sigla":               "Estado",
    "regiao_macro":               "Região Macro",
    "regiao_micro":               "Região Micro",
    "nomeResponsavel":            "Responsável",
    "dePara":                     "Híbrido",
    "status_material":            "Status",
    "indexTratamento":            "Trat.",
    "dataPlantioMilho":           "Plantio",
    "dataColheitaMilho":          "Colheita",
    "kg_ha":                      "kg/ha",
    "sc_ha":                      "sc/ha",
    "prod_relativa_pct":          "Prod. Relativa (%)",
    "umidade_pct":                "Umidade (%)",
    "populacao_real_plantas_ha":  "Pop. Real (pl/ha)",
    "altura_planta_m":            "Alt. Planta (m)",
    "altura_espiga_m":            "Alt. Espiga (m)",
    "pmg_corrigido_g":            "PMG (g)",
    "fileiras_media":             "Fileiras",
    "graos_fileira_media":        "Grãos/Fileira",
    "graos_ardidos_pct":          "Ardidos (%)",
    "pct_acamadas":               "Acamamento (%)",
    "pct_colmo_podre":            "Colmo Podre (%)",
    "pct_quebradas":              "Quebramento (%)",
    "pct_dominadas":              "Dominadas (%)",
    "pct_perda_total":            "Perda Total (%)",
}
# doenças: nota → teve? → classe, na ordem de DOENCAS
for _nome, _cols in DOENCAS.items():
    _cn, _cc = _cols["nota"], _cols["class"]
    for _c in [_cn, f"inc_{_cn}", _cc]:
        if _c in rename_doencas:
            col_map[_c] = rename_doencas[_c]
# fenômenos: plantas → %
for _nome, _cols in FENOMENOS.items():
    for _c in [_cols["n"], _cols["pct"]]:
        if _c in rename_fenomenos:
            col_map[_c] = rename_fenomenos[_c]
# sanidade só medida em 24/25 (tombamento verde virou green snap na av4 de 2025)
col_map["nota_tombamento_verde"] = "Nota Tombamento Verde (24/25)"

cols_disp = [c for c in col_map if c in df_tabela.columns]

# ordenação hierárquica: Macro → Micro → Estado → Cidade → Fazenda; dentro da fazenda, maior→menor kg/ha
_hier = [c for c in ["regiao_macro", "regiao_micro", "estado_sigla", "cidade_nome", "nomeFazenda"]
         if c in df_tabela.columns]
if _hier and "kg_ha" in df_tabela.columns:
    df_tabela = df_tabela.sort_values(
        _hier + ["kg_ha"],
        ascending=[True] * len(_hier) + [False],
        na_position="last",
    ).reset_index(drop=True)

# ── Esconde colunas sem nenhum dado no recorte, mas avisa quais foram ────────
# Em 24/25 os fenômenos não existem; em 25/26 o legado sai vazio. Coluna vazia só
# ocupa espaço — o aviso na legenda preserva a informação de que ela existe.
_CTX_FIXO = ["safra", "cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla",
             "regiao_macro", "regiao_micro", "nomeResponsavel", "dePara", "status_material"]
_vazias = [c for c in cols_disp if c not in _CTX_FIXO and df_tabela[c].isna().all()]
cols_disp = [c for c in cols_disp if c not in _vazias]

df_show = df_tabela[cols_disp].rename(columns=col_map)
# população como número inteiro (não 60942.0)
if "Pop. Real (pl/ha)" in df_show.columns:
    df_show["Pop. Real (pl/ha)"] = pd.to_numeric(
        df_show["Pop. Real (pl/ha)"], errors="coerce").round(0).astype("Int64")
# datas em dd/mm/aaaa (sem hora)
for _c in ["Plantio", "Colheita"]:
    if _c in df_show.columns:
        df_show[_c] = pd.to_datetime(df_show[_c], errors="coerce").dt.strftime("%d/%m/%Y").fillna("—")

ag_table(df_show, height=min(560, 40 + 32 * min(len(df_show), 15) + 20))
exportar_excel(df_show, nome_arquivo="auditoria_sanidade.xlsx",
               label="⬇️ Exportar Auditoria", key="exp_audit_sn")

_cap = (f"{len(df_show)} observações · {df_tabela['dePara'].nunique()} híbridos · "
        f"{df_tabela['cod_fazenda'].nunique()} locais.")
if _vazias:
    _cap += (" Colunas ocultas por não terem nenhum dado neste recorte: "
             + ", ".join(sorted({col_map[c] for c in _vazias})) + ".")
st.caption(_cap)

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 2 — RESUMO POR DOENÇA (matriz híbrido × doença, adaptada da soja)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Resumo por doença",
    "Como cada híbrido se comporta sanitariamente?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Nota mais frequente, incidência e classificação de cada doença por híbrido, sobre todos os
locais dos filtros ativos. Notas 0 (não avaliado) ficam de fora de todas as contas.

**Colunas**
- **Nota** → a nota que mais se repetiu nos plots avaliados (escala 1–9, **9 é o melhor**)
- **Inc.** → % dos locais em que a doença foi detectada (nota 1–5)
- **Classe** → derivada da Nota
""")
    cols_leg = st.columns(len(COR_CLASS))
    for i, (cls, cor) in enumerate(COR_CLASS.items()):
        cols_leg[i].markdown(
            f'<div style="background:{cor};color:{COR_TEXTO_CLASS[cls]};border-radius:6px;'
            f'padding:8px;text-align:center;font-size:12px;font-weight:600;">{LABEL_CLASS[cls]}</div>',
            unsafe_allow_html=True,
        )
    st.markdown("""
**Dois cuidados na leitura**

- **Empate na nota mais frequente cai para o lado pior.** Se um híbrido teve tantas notas 7
  quanto 5, a tabela mostra 5. É proposital: em recomendação, errar para o lado suscetível
  custa menos que errar para o lado resistente.
- **A incidência tem denominador próprio.** A coluna *Locais* conta todos os locais do híbrido
  no recorte; a *Inc.* de cada doença considera só os locais em que **aquela** doença foi
  avaliada, que podem ser bem menos. Um híbrido com 80 locais pode ter ferrugem avaliada em 12.

Compare os lançamentos (STINE / EXP) com os checks para avaliar o diferencial sanitário.
""")

# ── Resumo por híbrido × doença ───────────────────────────────────────────────
# Nota  = moda dos plots avaliados (empate → menor nota, ver popover);
# Inc.  = locais com detecção (nota 1–5) ÷ locais em que a doença foi avaliada;
# Classe = mesma régua do pipeline aplicada à moda.
ORDEM_STATUS = ["CHECK", "STINE", "EXP", "DP2"]

resumo_rows = []
for hibrido, grp in ta_filtrado.groupby("dePara", dropna=True):
    _modo_status = grp["status_material"].mode()
    row = {
        "Híbrido": hibrido,
        "Status":  _modo_status.iloc[0] if not _modo_status.empty else "",
        "Locais":  grp["cod_fazenda"].nunique(),
        "kg/ha":   round(grp["kg_ha"].dropna().mean(), 1) if "kg_ha" in grp.columns and grp["kg_ha"].notna().any() else None,
        "sc/ha":   round(grp["sc_ha"].dropna().mean(), 1) if "sc_ha" in grp.columns and grp["sc_ha"].notna().any() else None,
    }
    for doenca, cols in DOENCAS.items():
        nota, inc, classe = resumo_doenca(grp, cols["nota"])   # fonte única (ver helper)
        row[f"Nota {doenca}"] = nota
        row[f"Inc. {doenca} (%)"] = inc
        row[f"Classe {doenca}"] = classe
    resumo_rows.append(row)

df_resumo = pd.DataFrame(resumo_rows)

# doenças sem nenhum dado no recorte saem da tabela (3 colunas vazias cada)
_doencas_vazias = [d for d in DOENCAS if df_resumo[f"Nota {d}"].isna().all()]
if _doencas_vazias:
    df_resumo = df_resumo.drop(columns=[f"{p} {d}{s}" for d in _doencas_vazias
                                        for p, s in [("Nota", ""), ("Inc.", " (%)"), ("Classe", "")]])

# ordem: status na gramática do painel (CHECK → STINE → EXP → DP2), depois maior sc/ha
_ordem = [s for s in ORDEM_STATUS if s in set(df_resumo["Status"])]
_ordem += sorted(set(df_resumo["Status"]) - set(_ordem))
df_resumo["Status"] = pd.Categorical(df_resumo["Status"], categories=_ordem, ordered=True)
df_resumo = (df_resumo.sort_values(["Status", "sc/ha"], ascending=[True, False], na_position="last")
             .reset_index(drop=True))
df_resumo["Status"] = df_resumo["Status"].astype(str)


# ── AgGrid: classe com fundo colorido (fundo = categoria, regra do painel) ────
def ag_table_doencas(df, height=400):
    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000", "fontFamily": "Helvetica Neue, sans-serif"},
    )
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        enableRangeSelection=True,
    )

    _centro = {"fontSize": "13px", "color": "#000000",
               "fontFamily": "Helvetica Neue, sans-serif", "textAlign": "center"}

    # O cellStyle da classe é GERADO da paleta (COR_CLASS/COR_TEXTO_CLASS), não digitado à mão:
    # os hexes ficavam repetidos aqui, no HTML e no Excel, e foi essa repetição que fez o export
    # sair com cor diferente da tela. Agora as três leituras vêm do mesmo dicionário.
    def _js_classe():
        _regras = "\n".join(
            f"                if (v === '{k}') return {{ background: '{COR_CLASS[k]}', "
            f"color: '{COR_TEXTO_CLASS.get(k, '#1A1A1A')}', "
            f"fontWeight: '700', textAlign: 'center' }};"
            for k in ORDEM_CLASS)
        # uma instância de JsCode POR COLUNA — reusar o mesmo objeto perde o estilo
        return JsCode("            function(params) {\n"
                      "                var v = params.value;\n"
                      f"{_regras}\n"
                      "                return { textAlign: 'center' };\n"
                      "            }\n")

    for col in df.columns:
        if col.startswith("Classe "):
            gb.configure_column(col, width=90, headerClass="ag-header-center",
                                cellStyle=_js_classe())
        elif col.startswith("Nota "):
            gb.configure_column(col, width=80, cellStyle=_centro, headerClass="ag-header-center")
        elif col.startswith("Inc. "):
            gb.configure_column(col, width=90, cellStyle=_centro, headerClass="ag-header-center")

    gb.configure_column("Híbrido", pinned="left", width=170)
    gb.configure_column("Status", width=90)
    gb.configure_column("Locais", width=80, cellStyle=_centro, headerClass="ag-header-center")
    gb.configure_column("kg/ha", width=90, cellStyle=_centro, headerClass="ag-header-center")
    gb.configure_column("sc/ha", width=90, cellStyle=_centro, headerClass="ag-header-center")

    go = gb.build()
    go["defaultColDef"]["headerClass"] = "ag-header-black"
    # SEM sizeColumnsToFit: com mais de 20 colunas ele espreme tudo e o texto do
    # cabeçalho some (é o que acontece na página de soja). Aqui a tabela rola na
    # horizontal e o híbrido fica fixo à esquerda.

    AgGrid(
        df, gridOptions=go, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-header":                      {"background-color": "#4A4A4A !important"},
            ".ag-header-row":                  {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":                 {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":           {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":            {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":                        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-cell-menu-button":     {"opacity": "1 !important", "visibility": "visible !important"},
            ".ag-header-cell-menu-button span": {"color": "#FFFFFF !important"},
            ".ag-icon-menu":                   {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-icon-filter":                 {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # sem "color" no .ag-cell: um !important sobreporia a cor inline do cellStyle
            ".ag-row":                         {"font-size": "13px !important"},
            ".ag-header-center .ag-header-cell-label": {"justify-content": "center !important"},
        },
        theme="streamlit", use_container_width=True,
    )


ag_table_doencas(df_resumo, height=min(680, 36 + 32 * len(df_resumo) + 20))
exportar_excel(df_resumo, nome_arquivo="resumo_doencas.xlsx",
               label="⬇️ Exportar Resumo por Doença", key="exp_resumo_sn")

_cap_res = ("**Nota** = nota que mais se repetiu nos plots avaliados, escala 1–9 (9 = melhor "
            "sanidade); empate fica com a nota pior. **Inc.** = % dos locais em que a doença foi "
            "detectada (nota 1–5), sobre os locais em que ela foi avaliada. **Classe** = derivada "
            "da Nota · AS (pior) → S → MT → T → R (melhor).")
if _doencas_vazias:
    _cap_res += (" Sem avaliação neste recorte, fora da tabela: "
                 + ", ".join(sorted(_doencas_vazias)) + ".")
st.caption(_cap_res)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 3 — APRESENTAÇÃO (visão consolidada de sanidade por híbrido)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Apresentação",
    "Visão consolidada de sanidade por híbrido",
    contexto_str,
)

# Siglas para o cabeçalho não estourar a largura com 6 doenças × 3 colunas
SIGLAS = {
    "Turcicum":          "TUR",
    "Cercospora":        "CER",
    "Mancha branca":     "MBR",
    "Bipolaris":         "BIP",
    "Ferrugem tropical": "FTR",
    "Enfezamento":       "ENF",
}

# Fundo da linha = status do material (mesma paleta do painel); texto claro só no azul
COR_TEXTO_STATUS = {"CHECK": "#1A1A1A", "STINE": "#FFFFFF", "EXP": "#1A1A1A",
                    "DP2": "#1A1A1A", "": "#000000"}

doencas_disp = [
    d for d, cols in DOENCAS.items()
    if cols["nota"] in ta_filtrado.columns
    and pd.to_numeric(ta_filtrado[cols["nota"]], errors="coerce").gt(0).any()
]

with st.popover("ℹ️ Como interpretar esta tabela", use_container_width=False):
    st.markdown("""
Mesma informação da Seção 2, no formato de apresentação: uma linha por híbrido, fundo da linha
pelo status do material e as doenças escolhidas lado a lado. É a tabela para levar para reunião.

**Colunas por doença**
- **Nota** → nota que mais se repetiu nos plots avaliados, escala 1–9 (**9 é o melhor**)
- **%** → percentual dos locais em que a doença foi detectada (nota 1–5)
- **Classe** → derivada da Nota

**Fundo da linha** → status do híbrido: CHECK, STINE, EXP, DP2.

**Alerta automático** → abaixo da tabela, a lista dos híbridos que ficaram em MT, S ou AS em
alguma doença. É o resumo do que precisa de atenção na hora de recomendar.

O empate na nota mais frequente cai para a nota pior, e a incidência considera só os locais em
que aquela doença foi avaliada — os mesmos critérios da Seção 2, calculados pela mesma função.
""")
    cols_leg = st.columns(len(COR_CLASS))
    for i, (cls, cor) in enumerate(COR_CLASS.items()):
        cols_leg[i].markdown(
            f'<div style="background:{cor};color:{COR_TEXTO_CLASS[cls]};border-radius:6px;'
            f'padding:8px;text-align:center;font-size:12px;font-weight:600;">{LABEL_CLASS[cls]}</div>',
            unsafe_allow_html=True,
        )

if not doencas_disp:
    st.info("Nenhuma doença avaliada neste recorte.")
else:
    with st.expander("Selecionar doenças", expanded=True):
        _cols_sel = st.columns(3)
        doencas_sel = []
        for i, d in enumerate(doencas_disp):
            if _cols_sel[i % 3].checkbox(f"{SIGLAS.get(d, d)} — {d}", value=True, key=f"sn_apres_{d}"):
                doencas_sel.append(d)

    if not doencas_sel:
        st.info("Selecione ao menos uma doença para gerar a tabela.")
    else:
        apres_rows = []
        for hibrido, grp in ta_filtrado.groupby("dePara", dropna=True):
            _md = grp["status_material"].mode()
            row = {
                "Híbrido": hibrido,
                "status_material": _md.iloc[0] if not _md.empty else "",
                "Locais": grp["cod_fazenda"].nunique(),
                "sc/ha": round(grp["sc_ha"].dropna().mean(), 1) if "sc_ha" in grp.columns and grp["sc_ha"].notna().any() else None,
                "kg/ha": round(grp["kg_ha"].dropna().mean(), 1) if "kg_ha" in grp.columns and grp["kg_ha"].notna().any() else None,
            }
            for doenca in doencas_sel:
                sig = SIGLAS.get(doenca, doenca)
                nota, inc, classe = resumo_doenca(grp, DOENCAS[doenca]["nota"])   # fonte única
                row[f"{sig}_nota"] = nota
                row[f"{sig}_pct"] = inc
                row[f"{sig}_classe"] = classe
            apres_rows.append(row)

        df_apres = pd.DataFrame(apres_rows)
        _ordem_a = [x for x in ORDEM_STATUS if x in set(df_apres["status_material"])]
        _ordem_a += sorted(set(df_apres["status_material"]) - set(_ordem_a))
        df_apres["status_material"] = pd.Categorical(df_apres["status_material"],
                                                     categories=_ordem_a, ordered=True)
        df_apres = (df_apres.sort_values(["status_material", "sc/ha"],
                                         ascending=[True, False], na_position="last")
                    .reset_index(drop=True))
        df_apres["status_material"] = df_apres["status_material"].astype(str)

        # ── Tabela HTML com cabeçalho de dois níveis ────────────────────────
        _th1 = ('<th rowspan="2" style="text-align:left;">Híbrido</th>'
                '<th rowspan="2">Locais</th>'
                '<th rowspan="2">sc/ha</th>'
                '<th rowspan="2">kg/ha</th>')
        _th2 = ""
        for doenca in doencas_sel:
            _th1 += (f'<th colspan="3" style="text-align:center;'
                     f'border-left:2px solid #AAAAAA;">{doenca}</th>')
            _th2 += ('<th style="border-left:2px solid #AAAAAA;text-align:center;">Nota</th>'
                     '<th>%</th><th>Classe</th>')

        html = """
<style>
.tb-sn { width:100%; border-collapse:collapse; font-size:14px; font-family:'Helvetica Neue',sans-serif; }
.tb-sn th {
    background:#E8E8E8; color:#1A1A1A !important; padding:7px 8px;
    text-align:center; border:1px solid #BBBBBB; white-space:nowrap; font-weight:700;
}
.tb-sn th:first-child { text-align:left; }
.tb-sn td { padding:6px 8px; border:1px solid #ddd; text-align:center; white-space:nowrap; }
.tb-sn td:first-child { text-align:left; font-weight:600; }
.tb-sn td[data-fg="white"] { color:#FFFFFF !important; }
.tb-sn td[data-fg="dark"]  { color:#1A1A1A !important; }
.tb-sn td.sep { border-left:2px solid #AAAAAA !important; }
</style>
""" + f"""
<table class="tb-sn">
<thead><tr>{_th1}</tr><tr>{_th2}</tr></thead>
<tbody>
"""

        def _fmt(v):
            return "—" if v is None or (isinstance(v, float) and np.isnan(v)) else str(v)

        alertas = {}
        for _, r in df_apres.iterrows():
            status = r.get("status_material", "")
            bg = COR_STATUS_PLOT.get(status, "#FFFFFF")
            fg = COR_TEXTO_STATUS.get(status, "#000000")
            dfg = "white" if fg == "#FFFFFF" else "dark"
            html += "<tr>"
            for campo in ["Híbrido", "Locais", "sc/ha", "kg/ha"]:
                html += f'<td data-fg="{dfg}" style="background:{bg};">{_fmt(r.get(campo))}</td>'
            for doenca in doencas_sel:
                sig = SIGLAS.get(doenca, doenca)
                classe = r.get(f"{sig}_classe") or "—"
                pct = r.get(f"{sig}_pct")
                if classe in ("AS", "S", "MT"):
                    pct_txt = "—" if pct is None or (isinstance(pct, float) and np.isnan(pct)) else f"{pct:.0f}%"
                    alertas[(r["Híbrido"], doenca)] = (pct_txt, classe)
                html += f'<td class="sep" data-fg="{dfg}" style="background:{bg};">{_fmt(r.get(f"{sig}_nota"))}</td>'
                html += f'<td data-fg="{dfg}" style="background:{bg};">{_fmt(pct)}</td>'
                # célula de Classe: cor da CLASSE (T/R verde, MT amarelo, S/AS vermelho), não do status
                if classe in COR_CLASS:
                    _bg_cls = COR_CLASS[classe]
                    _fg_cls = COR_TEXTO_CLASS.get(classe, "#1A1A1A")
                    html += (f'<td style="background:{_bg_cls};color:{_fg_cls};'
                             f'font-weight:700;text-align:center;">{classe}</td>')
                else:
                    html += f'<td data-fg="{dfg}" style="background:{bg};">{classe}</td>'
            html += "</tr>"
        html += "</tbody></table>"

        st.markdown(html, unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Alertas: quem ficou em MT, S ou AS ──────────────────────────────
        if alertas:
            _linhas = [f"**{h}** → {d}: incidência {p} · classe {c}"
                       for (h, d), (p, c) in alertas.items()]
            st.warning("⚠️ **Atenção — híbridos em Moderadamente Tolerante (MT), Suscetível (S) "
                       "ou Altamente Suscetível (AS):**\n\n" + "  \n".join(_linhas))

        # ── Rodapé: cobertura de cada doença na rede ────────────────────────
        _n_locais_tot = ta_filtrado["cod_fazenda"].nunique()
        _cob = []
        for doenca in doencas_sel:
            _cn = DOENCAS[doenca]["nota"]
            if _cn not in ta_filtrado.columns:
                continue
            _s = pd.to_numeric(ta_filtrado[_cn], errors="coerce")
            _nl = ta_filtrado.loc[_s > 0, "cod_fazenda"].nunique()
            if _nl < _n_locais_tot:
                _cob.append(f"**{SIGLAS.get(doenca, doenca)}**: {_nl} de {_n_locais_tot} locais")
        if _cob:
            st.caption("Locais com nota registrada por doença (nos demais, aquela doença não foi "
                       "avaliada): " + " · ".join(_cob))

        # ── Exportação com as cores da tela ─────────────────────────────────
        def exportar_apresentacao(df, doencas):
            import io
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            COR_XL = {k: v.lstrip("#") for k, v in COR_STATUS_PLOT.items()}
            TXT_XL = {k: v.lstrip("#") for k, v in COR_TEXTO_STATUS.items()}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sanidade"
            thin = Side(style="thin", color="CCCCCC")
            medium = Side(style="medium", color="AAAAAA")
            borda = Border(left=thin, right=thin, top=thin, bottom=thin)

            base = ["Híbrido", "Locais", "sc/ha", "kg/ha"]
            for ci, col in enumerate(base, 1):
                ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)
                c = ws.cell(row=1, column=ci, value=col)
                c.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
                c.fill = PatternFill("solid", start_color="E8E8E8")
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                c.border = borda
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) + 2)

            ci = len(base) + 1
            for doenca in doencas:
                ws.merge_cells(start_row=1, start_column=ci, end_row=1, end_column=ci + 2)
                c = ws.cell(row=1, column=ci, value=doenca)
                c.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
                c.fill = PatternFill("solid", start_color="E8E8E8")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = Border(left=medium, right=medium, top=thin, bottom=thin)
                for j, sub in enumerate(["Nota", "%", "Classe"]):
                    c2 = ws.cell(row=2, column=ci + j, value=sub)
                    c2.font = Font(bold=True, name="Arial", size=10, color="1A1A1A")
                    c2.fill = PatternFill("solid", start_color="E8E8E8")
                    c2.alignment = Alignment(horizontal="center", vertical="center")
                    c2.border = Border(left=medium if j == 0 else thin, right=thin,
                                       top=thin, bottom=thin)
                    ws.column_dimensions[get_column_letter(ci + j)].width = 10
                ci += 3

            for ri, (_, r) in enumerate(df.iterrows(), start=3):
                status = r.get("status_material", "")
                fill = PatternFill("solid", start_color=COR_XL.get(status, "FFFFFF"))
                cor_txt = TXT_XL.get(status, "000000")
                valores = [r.get(c) for c in base]
                for doenca in doencas:
                    sig = SIGLAS.get(doenca, doenca)
                    valores += [r.get(f"{sig}_nota"), r.get(f"{sig}_pct"),
                                r.get(f"{sig}_classe") or "—"]
                for cj, val in enumerate(valores, 1):
                    try:
                        if val is None or (isinstance(val, float) and np.isnan(val)):
                            val = "—"
                    except (TypeError, ValueError):
                        pass
                    c = ws.cell(row=ri, column=cj, value=val)
                    c.font = Font(name="Arial", size=10, color=cor_txt)
                    c.fill = fill
                    c.alignment = Alignment(horizontal="left" if cj == 1 else "center",
                                            vertical="center")
                    c.border = borda
                    # A célula de Classe é a EXCEÇÃO da linha: na tela ela é pintada pela classe
                    # (T/R verde, MT amarelo, S/AS vermelho), não pelo status do material — e é
                    # ela que o leitor procura. Cada doença ocupa três colunas (Nota, %, Classe),
                    # então a de classe é a terceira de cada trio depois das colunas base.
                    _cls = str(val).strip()
                    if cj > len(base) and (cj - len(base)) % 3 == 0 and _cls in COR_CLASS_XL:
                        c.fill = PatternFill("solid", start_color=COR_CLASS_XL[_cls])
                        c.font = Font(name="Arial", size=10, bold=True,
                                      color=COR_TEXTO_CLASS_XL.get(_cls, "1A1A1A"))

            ws.freeze_panes = "B3"
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        st.download_button(
            "⬇️ Exportar Apresentação (com cores)",
            data=exportar_apresentacao(df_apres, doencas_sel),
            file_name="apresentacao_sanidade.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="exp_apres_sn",
        )

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 4 — RANKING POR QUARTIL SANITÁRIO
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Ranking por Quartil Sanitário",
    "Quais híbridos se mantêm no topo da sanidade na maioria dos locais?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
**Top 25% do local (%)**
Em cada local, os híbridos são comparados entre si. Os que ficam entre os 25% mais sadios daquele
ambiente "entram no top". Esta coluna mostra em quantos % dos locais o híbrido conseguiu isso.

Um híbrido com 80% significa que em 8 de cada 10 locais ele foi mais sadio que a maioria — mesmo
nos locais onde a doença foi severa para todos.

*(Tecnicamente: híbrido com nota ≥ 3º quartil — Q3 — do local. Como a nota alta é a mais sadia,
o Q3 marca o corte dos 25% melhores. O Q3 é calculado sobre TODOS os híbridos avaliados no local,
não só os filtrados — o mérito do híbrido não muda conforme a seleção da tela.)*

---

**Nota**
A nota do meio — ordenando as notas médias por local do menor para o maior, é o valor central.
Representa o comportamento habitual do híbrido, sem ser puxada por um local excepcionalmente bom
ou ruim.

*(Tecnicamente: mediana das notas médias por local.)*

---

**Como ler as duas colunas juntas**

| Top 25% | Nota típica | O que significa |
|---|---|---|
| Alto | Alta | Consistentemente sadio — destaque real |
| Alto | Baixa | Irregular — sadio em alguns locais, sofre em outros |
| Baixo | Alta | Mediano em todo lugar — nunca decepciona, nunca surpreende |
| Baixo | Baixa | Perfil sanitário fraco na maioria dos ambientes |

---

**sc/ha médio** → produtividade média nos locais avaliados — para cruzar sanidade com produção.
""")

if not doencas_sel:
    st.info("Selecione ao menos uma doença na Apresentação para exibir o ranking.")
elif "sc_ha" not in ta_filtrado.columns:
    st.info("Coluna sc_ha não encontrada na base — ranking indisponível.")
else:
    doenca_rank = st.selectbox(
        "Doença para o ranking",
        options=doencas_sel,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
        key="sn_doenca_rank",
    )
    col_nota_rank = DOENCAS[doenca_rank]["nota"]

    if col_nota_rank not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        # Q3 POR LOCAL sobre o ENSAIO INTEIRO (todos os híbridos avaliados no local, via
        # _aplicar_filtros_local sobre ta_raw). O filtro de material não move o corte: o "top 25%"
        # é uma propriedade do híbrido no ambiente, não da seleção da tela. Mesmo princípio da
        # referência por local da Análise Conjunta.
        _base_rank = _aplicar_filtros_local(ta_raw)[
            ["dePara", "status_material", "cod_fazenda", col_nota_rank, "sc_ha"]].copy()
        _base_rank[col_nota_rank] = pd.to_numeric(_base_rank[col_nota_rank], errors="coerce")
        _base_rank = _base_rank[_base_rank[col_nota_rank] > 0].dropna(subset=[col_nota_rank])

        # Q3 de cada local, fixo, sobre todos os híbridos
        _q3_local = _base_rank.groupby("cod_fazenda")[col_nota_rank].quantile(0.75).rename("_q3")

        # a EXIBIÇÃO é dos filtrados: junta o corte fixo do local aos plots filtrados
        df_rank = ta_filtrado[
            ["dePara", "status_material", "cod_fazenda", col_nota_rank, "sc_ha"]].copy()
        df_rank[col_nota_rank] = pd.to_numeric(df_rank[col_nota_rank], errors="coerce")
        df_rank["sc_ha"] = pd.to_numeric(df_rank["sc_ha"], errors="coerce")
        df_rank = df_rank[df_rank[col_nota_rank] > 0].dropna(subset=[col_nota_rank])
        df_rank = df_rank.merge(_q3_local, on="cod_fazenda", how="left")
        df_rank["_top"] = (df_rank[col_nota_rank] >= df_rank["_q3"]).astype(int)

        # média por híbrido × local antes de agregar
        df_rank_agg = (df_rank.groupby(["dePara", "status_material", "cod_fazenda"])
                       .agg(nota_media=(col_nota_rank, "mean"),
                            _top_local=("_top", "max"),
                            sc_ha_media=("sc_ha", "mean"))
                       .reset_index())

        rank_rows = []
        for hibrido, grp in df_rank_agg.groupby("dePara"):
            status = grp["status_material"].mode()[0] if not grp["status_material"].mode().empty else ""
            n_locais = len(grp)
            n_top = int(grp["_top_local"].sum())
            pct_top = round(n_top / n_locais * 100, 1) if n_locais > 0 else 0
            sc_med = round(grp["sc_ha_media"].mean(), 1) if grp["sc_ha_media"].notna().any() else None
            rank_rows.append({
                "Híbrido":              hibrido,
                "Status":               status,
                "Locais":               n_locais,
                "Top 25% do local (%)": pct_top,
                "Nota":                 round(grp["nota_media"].median(), 2),
                "sc/ha médio":          sc_med,
            })

        df_ranking = (pd.DataFrame(rank_rows)
                      .sort_values("Top 25% do local (%)", ascending=False)
                      .reset_index(drop=True))

        gb_rk = GridOptionsBuilder.from_dataframe(df_ranking)
        gb_rk.configure_default_column(
            resizable=True, sortable=True, filter=True,
            cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
        gb_rk.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
        gb_rk.configure_column("Híbrido", pinned="left", width=170)
        gb_rk.configure_column("Status", width=90)
        gb_rk.configure_column("Locais", width=80)
        gb_rk.configure_column("Top 25% do local (%)", width=160,
                               cellStyle=js_faixa(FAIXAS_COR["Top 25% do local (%)"]))
        gb_rk.configure_column("Nota", width=90,
                               cellStyle=js_faixa(FAIXAS_COR["Nota"]))
        gb_rk.configure_column("sc/ha médio", width=110)

        go_rk = gb_rk.build()
        go_rk["defaultColDef"]["headerClass"] = "ag-header-black"

        AgGrid(
            df_ranking, gridOptions=go_rk,
            height=min(500, 36 + 32 * len(df_ranking) + 20),
            update_mode=GridUpdateMode.NO_UPDATE,
            fit_columns_on_grid_load=False,
            allow_unsafe_jscode=True, enable_enterprise_modules=True,
            custom_css={
                ".ag-header":                  {"background-color": "#4A4A4A !important"},
                ".ag-header-row":              {"background-color": "#4A4A4A !important"},
                ".ag-header-cell":             {"background-color": "#4A4A4A !important"},
                ".ag-header-cell-label":       {"color": "#FFFFFF !important", "font-weight": "700"},
                ".ag-header-cell-text":        {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                ".ag-icon":                    {"color": "#FFFFFF !important", "opacity": "1 !important"},
                ".ag-row":                     {"font-size": "13px !important"},
            },
            theme="streamlit", use_container_width=True,
        )

        _n_loc_total = ta_filtrado["cod_fazenda"].nunique()
        _n_loc_rk = df_rank_agg["cod_fazenda"].nunique()
        _cap_rk = (f"**Top 25% do local (%)** = % de locais em que o híbrido ficou entre os 25% mais "
                   f"sadios daquele local (nota ≥ Q3, calculado sobre todos os híbridos do local). "
                   f"Ordenado do maior para o menor. {len(df_ranking)} híbridos avaliados para {doenca_rank}")
        _cap_rk += (f" ({_n_loc_rk} de {_n_loc_total} locais com nota registrada)."
                    if _n_loc_rk < _n_loc_total else ".")
        st.caption(_cap_rk)

        exportar_excel(df_ranking, nome_arquivo="ranking_quartil_sanitario.xlsx",
                       label="⬇️ Exportar Ranking por Quartil", key="exp_rank_sn",
                       faixas_cor=FAIXAS_COR)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 5 — DESEMPENHO POR LOCAL (gráfico de linhas)
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Desempenho por Local",
    "Como cada híbrido se comportou em cada local?",
    contexto_str,
)

col_interp, col_dic = st.columns([1, 9])
with col_interp:
    with st.popover("ℹ️ Como interpretar", use_container_width=True):
        st.markdown("""
Cada linha é um híbrido. O eixo Y mostra a nota da doença **do pior (topo) ao melhor (base)** —
quanto mais baixa a linha, melhor a sanidade do híbrido naquele local. Como a nota vai de 1 (mais
doente) a 9 (mais sadio), o eixo é invertido de propósito: assim "linha lá embaixo" quer dizer
"híbrido sadio", que é a leitura intuitiva.

**Elementos visuais**

- **Linha** — a trajetória do híbrido entre os locais, uma cor por híbrido.
- **Passe o mouse** sobre a linha para ver: híbrido, local, status, nota, incidência % e classe.
- **Faixa sombreada em vermelho** — o local do **pior caso**: onde algum híbrido teve a nota mais
  baixa de toda a doença. É por isso que a faixa cai exatamente onde uma linha atinge o topo do
  gráfico (nota baixa = linha no alto). O rótulo mostra a nota e qual híbrido foi. Isto é o pior
  ponto individual — diferente do card de alerta acima, que mede em quantos híbridos a doença
  apareceu (espalhamento). Um mede "onde ficou mais forte", o outro "onde se espalhou mais".

**Referência no gráfico**

- **Linha vermelha tracejada (nota 6)** — o limite entre Moderadamente Tolerante e Tolerante.
  Acima dela (notas 1 a 5), o híbrido está em zona de atenção para aquela doença.

**Escala da nota**

| Nota | Classe | Significado |
|---|---|---|
| 9 | R | Resistente |
| 7–8 | T | Tolerante |
| 5–6 | MT | Moderadamente Tolerante |
| 3–4 | S | Susceptível |
| 1–2 | AS | Altamente Susceptível |

**Dica de leitura**

Compare a altura das linhas entre híbridos no mesmo local. Um híbrido com a linha sempre perto da
base tem o melhor perfil sanitário naquela doença. Cruzamentos entre linhas mostram que a resposta
depende do ambiente — o híbrido é bom num local e sofre em outro.

**Card de alerta — como a incidência do local é calculada**

O card aparece quando um local tem **incidência média acima de 50%**. O cálculo é:

```
% incidência do local = avaliações com nota 1–5 (todos os híbridos do local)
                        ────────────────────────────────────────────────────  × 100
                        total de avaliações com nota > 0 naquele local
```

**É a incidência do ambiente (local), não de um híbrido.** Todos os híbridos avaliados ali entram
na conta — o número reflete a pressão da doença no ambiente, independentemente do material. Por
exemplo: se num local foram feitas 15 avaliações e 10 tiveram nota entre 1 e 5, a incidência do
local é **67%** — a doença apareceu com severidade na maioria dos híbridos daquele lugar.
""")

with col_dic:
    _df_dic = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
               .drop_duplicates()
               .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
               .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                "cidade_nome": "Cidade", "estado_sigla": "Estado"})
               .reset_index(drop=True))
    with st.popover(f"📍 Dicionário de locais ({len(_df_dic)})", use_container_width=False):
        st.markdown("Referência dos códigos exibidos no eixo X do gráfico.")
        st.dataframe(_df_dic, hide_index=True, use_container_width=True)

if not doencas_disp:
    st.info("Nenhuma doença com nota registrada nos filtros ativos.")
else:
    doenca_graf = st.selectbox(
        "Selecione a doença",
        options=doencas_disp,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}",
        key="sn_doenca_graf",
    )
    col_nota_g = DOENCAS[doenca_graf]["nota"]

    if col_nota_g not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        df_g = ta_filtrado[["dePara", "status_material", "cod_fazenda", "cidade_nome",
                            "estado_sigla", col_nota_g]].copy()
        df_g[col_nota_g] = pd.to_numeric(df_g[col_nota_g], errors="coerce")
        df_g = df_g[df_g[col_nota_g] > 0].dropna(subset=[col_nota_g])

        if df_g.empty:
            st.info("Nenhuma avaliação disponível para esta doença nos filtros ativos.")
        else:
            # aviso de legibilidade: muitas linhas viram novelo. Sugere filtrar híbridos.
            _n_hib_graf = df_g["dePara"].nunique()
            if _n_hib_graf > 12:
                st.info(
                    f"São **{_n_hib_graf} híbridos** neste recorte — o gráfico de linhas fica mais "
                    "legível com menos materiais. Use o filtro de híbrido na barra lateral para "
                    "comparar um grupo menor (ex.: só os CHECK, ou só os seus materiais de interesse).")

            # nota por híbrido × local (faixa não tem repetição — mean = valor único)
            df_g_agg = (df_g.groupby(["dePara", "status_material", "cod_fazenda",
                                      "cidade_nome", "estado_sigla"])
                        [col_nota_g].mean().round(1).reset_index())

            # incidência por híbrido × local (nota 1-5)
            df_inc = df_g.copy()
            df_inc["_inc"] = df_inc[col_nota_g].between(1, 5).astype(int)
            df_inc_agg = (df_inc.groupby(["dePara", "cod_fazenda"])
                          .agg(_n=("_inc", "count"), _inc=("_inc", "sum")).reset_index())
            df_inc_agg["inc_pct"] = (df_inc_agg["_inc"] / df_inc_agg["_n"] * 100).round(1)
            df_g_agg = df_g_agg.merge(df_inc_agg[["dePara", "cod_fazenda", "inc_pct"]],
                                      on=["dePara", "cod_fazenda"], how="left")
            df_g_agg["classe"] = df_g_agg[col_nota_g].apply(nota_para_classe)

            # ordena locais: estado → cidade → código
            locais_ord = (df_g_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                          .drop_duplicates()
                          .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                          ["cod_fazenda"].tolist())

            hibridos_g = sorted(df_g_agg["dePara"].unique().tolist())
            palette = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                       "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
                       "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
                       "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5",
                       "#393B79", "#637939", "#8C6D31", "#843C39", "#7B4173"]
            cor_hibrido = {c: palette[i % len(palette)] for i, c in enumerate(hibridos_g)}

            # card de alerta — locais com alta incidência (pressão do ambiente)
            inc_por_local = (df_inc_agg.groupby("cod_fazenda")
                             .apply(lambda g: round(g["_inc"].sum() / g["_n"].sum() * 100, 1))
                             .reset_index().rename(columns={0: "inc_media"}))
            locais_alerta = inc_por_local[inc_por_local["inc_media"] > 50]["cod_fazenda"].tolist()
            if locais_alerta:
                dic_local = (ta_filtrado[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                             .drop_duplicates().set_index("cod_fazenda"))
                detalhes = []
                for local in locais_alerta:
                    inc_val = inc_por_local.loc[inc_por_local["cod_fazenda"] == local, "inc_media"].values[0]
                    cidade = dic_local.loc[local, "cidade_nome"] if local in dic_local.index else ""
                    estado = dic_local.loc[local, "estado_sigla"] if local in dic_local.index else ""
                    detalhes.append(f"**{local}** ({cidade} — {estado}): {inc_val:.0f}%")
                st.warning(
                    f"⚠️ **{len(locais_alerta)} local(is) onde {doenca_graf} apareceu em mais da "
                    f"metade dos híbridos avaliados:**  \n" + "  \n".join(detalhes)
                    + "  \n_Incidência = % de híbridos com a doença; não mede a severidade._")

            fig = go_plt.Figure()
            for hibrido in hibridos_g:
                df_c = df_g_agg[df_g_agg["dePara"] == hibrido].sort_values(
                    "cod_fazenda", key=lambda s: s.map(lambda x: locais_ord.index(x)))
                if df_c.empty:
                    continue
                cor = cor_hibrido[hibrido]
                fig.add_trace(go_plt.Scatter(
                    x=df_c["cod_fazenda"], y=df_c[col_nota_g], mode="lines", name=hibrido,
                    line=dict(color=cor, width=2, shape="spline", smoothing=0.8),
                    legendgroup=hibrido, showlegend=True,
                    customdata=df_c[["status_material", "inc_pct", "classe"]].values,
                    hovertemplate=(f"<b>{hibrido}</b> · %{{x}}<br>Status: %{{customdata[0]}}<br>"
                                   f"Nota: %{{y}}<br>Incidência: %{{customdata[1]:.0f}}%<br>"
                                   f"Classe: %{{customdata[2]}}<extra></extra>"),
                ))

            fig.add_hline(y=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
                          annotation_text="Limite MT/T (nota 6)", annotation_position="top right",
                          annotation_font=dict(size=11, color="#E74C3C"))

            # destaca o local de maior pressão (mesmo critério do card)
            # DESTAQUE = local com a PIOR nota individual (a linha mais alta que o olho vê).
            # Não a média do local: o olho procura a linha que mais sobe, que é o pior híbrido
            # daquele local. Assim o vermelho cai exatamente onde a linha atinge o topo.
            _pior_local = df_g_agg.groupby("cod_fazenda")[col_nota_g].min().rename("pior_nota")
            if not _pior_local.empty:
                local_critico = _pior_local.idxmin()         # menor nota individual = pior caso
                nota_critica = _pior_local.min()
                if local_critico in locais_ord:
                    dic_local_g = (ta_filtrado[["cod_fazenda", "cidade_nome"]]
                                   .drop_duplicates().set_index("cod_fazenda"))
                    cidade_critica = dic_local_g.loc[local_critico, "cidade_nome"] if local_critico in dic_local_g.index else ""
                    # qual híbrido teve a pior nota ali (para o rótulo)
                    _pior_hib = (df_g_agg[(df_g_agg["cod_fazenda"] == local_critico) &
                                          (df_g_agg[col_nota_g] == nota_critica)]["dePara"].tolist())
                    _hib_txt = f" · {_pior_hib[0]}" if _pior_hib else ""
                    fig.add_shape(type="rect", xref="x", yref="paper",
                                  x0=locais_ord.index(local_critico) - 0.5,
                                  x1=locais_ord.index(local_critico) + 0.5,
                                  y0=0, y1=1, fillcolor="rgba(231,76,60,0.07)", line=dict(width=0))
                    fig.add_annotation(
                        x=local_critico, y=0.7, yref="y",
                        text=f"pior caso · nota {nota_critica:.0f}<br>{cidade_critica}{_hib_txt}",
                        showarrow=False, xanchor="center", yanchor="top",
                        font=dict(size=10, color="#E74C3C"), bgcolor="rgba(255,255,255,0.85)",
                        bordercolor="#E74C3C", borderwidth=1, borderpad=3)

            n_locais_g = len(locais_ord)
            altura_graf = max(450, min(700, 350 + n_locais_g * 8))
            fig.update_layout(
                height=altura_graf,
                yaxis=dict(title=dict(text="<b>Nota (1 = pior · 9 = melhor)</b>",
                                      font=dict(size=14, color="#111111", weight="bold")),
                           range=[9.3, 0.7], tickvals=list(range(1, 10)),
                           tickfont=dict(size=12, color="#111111", weight="bold"),
                           gridcolor="#EEEEEE", zeroline=False),
                xaxis=dict(title=dict(text="<b>Local (código)</b>",
                                      font=dict(size=14, color="#111111", weight="bold")),
                           tickangle=-45, tickfont=dict(size=11, color="#111111", weight="bold"),
                           categoryorder="array", categoryarray=locais_ord, gridcolor="#EEEEEE"),
                legend=dict(title=dict(text="<b>Híbrido</b>", font=dict(size=12, color="#111111")),
                            font=dict(size=12, color="#111111", weight="bold"),
                            itemsizing="constant", bgcolor="rgba(255,255,255,0.85)",
                            bordercolor="#DDDDDD", borderwidth=1),
                plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=100, l=60, r=20),
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                hovermode="closest")

            n_locais_total = ta_filtrado["cod_fazenda"].nunique()
            st.plotly_chart(fig, use_container_width=True)
            _cap_loc = (f"Eixo Y invertido — pior nota no topo, melhor na base. Linha tracejada "
                        f"vermelha = limite nota 6 (zona de atenção acima). {len(hibridos_g)} "
                        f"híbridos · {n_locais_g} locais com avaliação de {doenca_graf}")
            _cap_loc += (f" (de {n_locais_total} locais ativos — os demais não têm nota para esta doença)."
                         if n_locais_g < n_locais_total else ".")
            st.caption(_cap_loc)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 6 — DELTA VS REFERÊNCIA
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Delta vs Referência",
    "Como cada híbrido se compara ao CHECK ou STINE de referência?",
    contexto_str,
)

_col_interp_delta, _col_dic_delta = st.columns([1, 9])
with _col_interp_delta:
    with st.popover("ℹ️ Como interpretar", use_container_width=False):
        st.markdown("""
Escolha uma **doença** e um **híbrido de referência** (um CHECK ou STINE). Cada barra mostra a
diferença de nota entre um híbrido e essa referência **no mesmo local** — assim você compara os
materiais contra a testemunha, ambiente por ambiente, tirando o efeito do local.

**Como ler**

- **Barra verde (delta positivo)** — o híbrido teve nota mais alta que a referência naquele local,
  ou seja, ficou **mais sadio** que a testemunha ali.
- **Barra vermelha (delta negativo)** — o híbrido ficou com nota mais baixa, **mais doente** que a
  referência naquele local.
- Só entram **locais onde os dois foram avaliados** — a comparação é sempre no mesmo ambiente.

**O cálculo**

```
delta = nota do híbrido no local − nota da referência no mesmo local
```

Um híbrido consistentemente verde em vários locais tem vantagem sanitária real sobre a testemunha,
e não por sorte de ter pego ambientes fáceis — porque a comparação é local a local.

**A tabela abaixo do gráfico** resume tudo: em quantos locais cada híbrido ficou acima, abaixo ou
empatado com a referência, e o delta médio. É a leitura mais rápida quando há muitos híbridos.
""")

with _col_dic_delta:
    with st.popover(f"📍 Dicionário de locais ({ta_filtrado['cod_fazenda'].nunique()})",
                    use_container_width=False):
        _df_dic_delta = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                         .drop_duplicates()
                         .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                         .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                          "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                         .reset_index(drop=True))
        st.markdown("Referência dos códigos exibidos no eixo X do gráfico.")
        st.dataframe(_df_dic_delta, hide_index=True, use_container_width=True)

_col_d_delta, _col_ref_delta = st.columns(2)
with _col_d_delta:
    doenca_delta = st.selectbox(
        "Doença", options=doencas_disp,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}", key="sn_doenca_delta")

_refs_disponiveis = sorted(
    ta_filtrado[ta_filtrado["status_material"].isin(["CHECK", "STINE"])]["dePara"]
    .dropna().unique().tolist())

if not doencas_disp:
    st.info("Nenhuma doença com nota registrada nos filtros ativos.")
elif not _refs_disponiveis:
    st.info("Nenhum híbrido CHECK ou STINE disponível nos filtros ativos para usar como referência.")
else:
    with _col_ref_delta:
        ref_delta = st.selectbox("Referência (CHECK / STINE)", options=_refs_disponiveis,
                                 key="sn_ref_delta")
    col_nota_delta = DOENCAS[doenca_delta]["nota"]

    if col_nota_delta not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        df_delta = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                                "cidade_nome", "estado_sigla", col_nota_delta]].copy()
        df_delta[col_nota_delta] = pd.to_numeric(df_delta[col_nota_delta], errors="coerce")
        df_delta = df_delta[df_delta[col_nota_delta] > 0].dropna(subset=[col_nota_delta])

        df_delta_agg = (df_delta.groupby(["dePara", "status_material", "cod_fazenda",
                                          "cidade_nome", "estado_sigla"])
                        [col_nota_delta].mean().round(1).reset_index())

        df_ref = df_delta_agg[df_delta_agg["dePara"] == ref_delta][
            ["cod_fazenda", col_nota_delta]].rename(columns={col_nota_delta: "_nota_ref"})

        if df_ref.empty:
            st.info(f"**{ref_delta}** não tem avaliações para {doenca_delta} nos filtros ativos.")
        else:
            df_delta_agg = df_delta_agg.merge(df_ref, on="cod_fazenda", how="inner")
            df_delta_agg["delta"] = (df_delta_agg[col_nota_delta] - df_delta_agg["_nota_ref"]).round(1)
            df_plot = df_delta_agg[df_delta_agg["dePara"] != ref_delta].copy()

            if df_plot.empty:
                st.info("Nenhum outro híbrido foi avaliado nos mesmos locais que a referência.")
            else:
                # aviso: muitas barras por local ficam ilegíveis. A tabela abaixo cobre todos.
                _n_hib_delta = df_plot["dePara"].nunique()
                if _n_hib_delta > 8:
                    st.info(
                        f"São **{_n_hib_delta} híbridos** comparados — o gráfico de barras fica "
                        "cheio. Use o filtro de híbrido na barra lateral para ver poucos materiais "
                        "de cada vez; a **tabela resumo** abaixo do gráfico mostra todos de forma "
                        "compacta, independentemente da quantidade.")

                locais_delta = (df_plot[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                                .drop_duplicates()
                                .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                                ["cod_fazenda"].tolist())
                ordem_hibridos = (df_plot.groupby("dePara")["delta"].mean()
                                  .sort_values(ascending=False).index.tolist())

                fig_delta = go_plt.Figure()
                for hibrido in ordem_hibridos:
                    df_c = df_plot[df_plot["dePara"] == hibrido].set_index("cod_fazenda")
                    dx, dy, hover = [], [], []
                    for local in locais_delta:
                        if local in df_c.index:
                            d_val = df_c.loc[local, "delta"]
                            dx.append(local); dy.append(d_val)
                            hover.append(f"<b>{hibrido}</b> · {local}<br>"
                                         f"Nota: {df_c.loc[local, col_nota_delta]} · "
                                         f"Ref ({ref_delta}): {df_c.loc[local, '_nota_ref']}<br>"
                                         f"Delta: {d_val:+.1f}")
                    if not dx:
                        continue
                    cores = ["#1E7A34" if v >= 0 else "#E63946" for v in dy]
                    fig_delta.add_trace(go_plt.Bar(
                        name=hibrido, x=dx, y=dy, marker_color=cores,
                        text=[f"{v:+.1f}" for v in dy], textposition="outside",
                        textfont=dict(size=10, color="#333333"),
                        hovertext=hover, hoverinfo="text",
                        legendgroup=hibrido, showlegend=True, offsetgroup=hibrido))

                fig_delta.add_hline(y=0, line_color="#333333", line_width=1.5)
                fig_delta.add_hrect(y0=0, y1=8, fillcolor="rgba(30,122,52,0.04)", line_width=0)
                fig_delta.add_hrect(y0=-8, y1=0, fillcolor="rgba(230,57,70,0.04)", line_width=0)

                altura_delta = max(420, min(680, 350 + len(locais_delta) * 12))
                fig_delta.update_layout(
                    height=altura_delta,
                    title=dict(text=f"Delta de nota vs <b>{ref_delta}</b> — {doenca_delta}",
                               font=dict(size=14, color="#111111"), x=0, xanchor="left"),
                    barmode="group", bargap=0.15, bargroupgap=0.05,
                    xaxis=dict(title=dict(text="<b>Local (código)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickangle=-45, tickfont=dict(size=11, color="#111111", weight="bold"),
                               categoryorder="array", categoryarray=locais_delta, gridcolor="#EEEEEE"),
                    yaxis=dict(title=dict(text="<b>Delta (nota)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickfont=dict(size=12, color="#111111", weight="bold"),
                               gridcolor="#EEEEEE", zeroline=False),
                    legend=dict(title=dict(text="<b>Híbrido</b>",
                                           font=dict(size=12, color="#111111", weight="bold")),
                                font=dict(size=11, color="#111111", weight="bold"),
                                bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                    plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                    margin=dict(t=50, b=120, l=60, r=20),
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                    hovermode="x unified")

                _n_loc_total_delta = ta_filtrado["cod_fazenda"].nunique()
                st.plotly_chart(fig_delta, use_container_width=True)
                _cap_delta = (f"Delta = nota do híbrido − nota de {ref_delta} no mesmo local. "
                              f"Verde = mais sadio que a referência · vermelho = mais doente. "
                              f"{len(ordem_hibridos)} híbridos · {len(locais_delta)} locais com "
                              f"avaliação conjunta")
                _cap_delta += (f" (de {_n_loc_total_delta} locais ativos — os demais não têm "
                               f"avaliação conjunta para esta doença)."
                               if len(locais_delta) < _n_loc_total_delta else ".")
                st.caption(_cap_delta)

                # tabela resumo — funciona bem com qualquer nº de híbridos
                resumo_delta_rows = []
                for hibrido in ordem_hibridos:
                    df_c = df_plot[df_plot["dePara"] == hibrido]
                    status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                    resumo_delta_rows.append({
                        "Híbrido":       hibrido,
                        "Status":        status,
                        "Locais":        len(df_c),
                        "▲ Acima ref":   int((df_c["delta"] > 0).sum()),
                        "▼ Abaixo ref":  int((df_c["delta"] < 0).sum()),
                        "= Empate":      int((df_c["delta"] == 0).sum()),
                        "Delta médio":   round(df_c["delta"].mean(), 2),
                    })
                df_resumo_delta = pd.DataFrame(resumo_delta_rows)

                gb_dt = GridOptionsBuilder.from_dataframe(df_resumo_delta)
                gb_dt.configure_default_column(
                    resizable=True, sortable=True, filter=True,
                    cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
                gb_dt.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
                gb_dt.configure_column("Híbrido", pinned="left", width=170)
                gb_dt.configure_column("Status", width=90)
                gb_dt.configure_column("Locais", width=80)
                gb_dt.configure_column("▲ Acima ref", width=110,
                                       cellStyle=js_faixa(FAIXAS_COR["▲ Acima ref"]))
                gb_dt.configure_column("▼ Abaixo ref", width=110,
                                       cellStyle=js_faixa(FAIXAS_COR["▼ Abaixo ref"]))
                gb_dt.configure_column("= Empate", width=90)
                gb_dt.configure_column("Delta médio", width=110,
                                       cellStyle=js_faixa(FAIXAS_COR["Delta médio"]))

                go_dt = gb_dt.build()
                go_dt["defaultColDef"]["headerClass"] = "ag-header-black"
                AgGrid(
                    df_resumo_delta, gridOptions=go_dt,
                    height=min(500, 36 + 32 * len(df_resumo_delta) + 20),
                    update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
                    allow_unsafe_jscode=True, enable_enterprise_modules=True,
                    custom_css={
                        ".ag-header":            {"background-color": "#4A4A4A !important"},
                        ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                        ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                        ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-row":               {"font-size": "13px !important"},
                    },
                    theme="streamlit", use_container_width=True)
                st.caption(
                    f"**▲ Acima ref** = nº de locais em que o híbrido superou {ref_delta} em sanidade. "
                    f"**Delta médio** = média dos deltas em todos os locais avaliados juntos.")
                exportar_excel(df_resumo_delta, nome_arquivo="delta_referencia.xlsx",
                               label="⬇️ Exportar Delta vs Referência", key="exp_delta_sn",
                               faixas_cor=FAIXAS_COR)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 7 — EVOLUÇÃO POR SAFRA
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Evolução por Safra",
    "O perfil sanitário dos híbridos melhorou ou piorou ao longo das safras?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Compara a nota média de cada híbrido **entre as safras disponíveis na base**, independente do
filtro de safra ativo na barra lateral — assim todas as safras aparecem no gráfico. Os demais
filtros (macro, micro, estado, status) continuam valendo.

**Gráfico**

- **Linha subindo** → híbrido melhorando a sanidade (nota maior = mais sadio).
- **Linha descendo** → híbrido com piora no perfil sanitário.
- **Linha cinza tracejada** → média geral do grupo naquela safra (a referência).
- Híbridos avaliados em **apenas uma safra** ficam de fora do gráfico (não há o que ligar).

---

**Tabela de consistência — como cada coluna é calculada**

- **Safras aval.** → em quantas safras distintas o híbrido tem nota > 0 para a doença escolhida.
- **Acima da média** → em cada safra, calcula a média geral do grupo; se a nota do híbrido for
  maior ou igual a essa média, conta 1. A coluna soma em quantas safras isso aconteceu.
- **Consistência %** → Acima da média ÷ Safras aval. × 100.
- **Tendência** → inclinação da reta entre as safras, em ordem cronológica. Com duas safras, é
  simplesmente a diferença entre a nota da mais recente e a da anterior:
  - **↑ melhora** → nota subindo (mais sadio a cada safra);
  - **↓ piora** → nota caindo;
  - **→ estável** → variação pequena (entre −0,1 e +0,1).

---

**Atenção ao ler Consistência e Tendência juntas**

Um híbrido pode ter **100% de consistência e ↓ piora** ao mesmo tempo: ele está caindo, mas o grupo
inteiro caiu junto, então ele continua acima da média. A consistência mede a posição **relativa ao
grupo**; a tendência mede a evolução **absoluta** da nota. Para ver a variação crua, olhe as
colunas de nota por safra diretamente.
""")

_col_d_ev, _col_status_ev = st.columns(2)
with _col_d_ev:
    doenca_ev = st.selectbox(
        "Doença", options=doencas_disp,
        format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}", key="sn_doenca_ev")
with _col_status_ev:
    _status_ev_opts = sorted(ta_raw["status_material"].dropna().unique().tolist())
    status_ev_sel = st.multiselect("Status", options=_status_ev_opts,
                                   default=_status_ev_opts, key="sn_status_ev")

if not doencas_disp:
    st.info("Nenhuma doença com nota registrada nos filtros ativos.")
else:
    col_nota_ev = DOENCAS[doenca_ev]["nota"]
    if col_nota_ev not in ta_raw.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        # todos os filtros EXCETO safra — para todas as safras aparecerem. Usa os valores já
        # presentes em ta_filtrado (que reflete os filtros de local ativos) como whitelist.
        df_ev = ta_raw.copy()
        for _c in ["regiao_macro", "regiao_micro", "estado_sigla", "cidade_nome",
                   "nomeFazenda", "nomeResponsavel"]:
            if _c in ta_filtrado.columns and _c in df_ev.columns:
                _vals = ta_filtrado[_c].dropna().unique().tolist()
                if _vals:
                    df_ev = df_ev[df_ev[_c].isin(_vals)]
        if status_ev_sel:
            df_ev = df_ev[df_ev["status_material"].isin(status_ev_sel)]

        df_ev[col_nota_ev] = pd.to_numeric(df_ev[col_nota_ev], errors="coerce")
        df_ev = df_ev[df_ev[col_nota_ev] > 0].dropna(subset=[col_nota_ev, "safra", "dePara"])

        df_ev_agg = (df_ev.groupby(["dePara", "status_material", "safra"])
                     [col_nota_ev].mean().round(2).reset_index())
        safras_ev = sorted(df_ev_agg["safra"].unique().tolist())

        if len(safras_ev) < 2:
            st.info("São necessárias ao menos 2 safras na base para exibir a evolução. "
                    "Com os dados atuais só há uma safra com avaliação desta doença.")
        else:
            contagem_safras = df_ev_agg.groupby("dePara")["safra"].nunique()
            hibridos_ev = contagem_safras[contagem_safras >= 2].index.tolist()
            n_single = int((contagem_safras == 1).sum())

            if not hibridos_ev:
                st.info("Nenhum híbrido foi avaliado em mais de uma safra com os filtros ativos.")
            else:
                df_ev_plot = df_ev_agg[df_ev_agg["dePara"].isin(hibridos_ev)].copy()
                palette_ev = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                              "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
                              "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
                              "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5"]
                hib_ord_ev = sorted(hibridos_ev)
                cor_ev = {c: palette_ev[i % len(palette_ev)] for i, c in enumerate(hib_ord_ev)}

                if len(hib_ord_ev) > 12:
                    st.info(
                        f"São **{len(hib_ord_ev)} híbridos** com duas ou mais safras — o gráfico "
                        "fica cheio. Use o filtro de híbrido na barra lateral para ver menos linhas; "
                        "a tabela de consistência abaixo cobre todos.")

                fig_ev = go_plt.Figure()
                for hibrido in hib_ord_ev:
                    df_c = df_ev_plot[df_ev_plot["dePara"] == hibrido].sort_values("safra")
                    status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                    fig_ev.add_trace(go_plt.Scatter(
                        x=df_c["safra"].astype(str).tolist(), y=df_c[col_nota_ev].tolist(),
                        mode="lines+markers", name=hibrido,
                        line=dict(color=cor_ev[hibrido], width=2),
                        marker=dict(size=8, color=cor_ev[hibrido]),
                        hovertemplate=(f"<b>{hibrido}</b> ({status})<br>Safra: %{{x}}<br>"
                                       f"Nota média: %{{y:.2f}}<extra></extra>")))

                media_geral_ev = df_ev_plot.groupby("safra")[col_nota_ev].mean().round(2)
                fig_ev.add_trace(go_plt.Scatter(
                    x=[str(s) for s in media_geral_ev.index], y=media_geral_ev.values.tolist(),
                    mode="lines+markers", name="Média do grupo",
                    line=dict(color="#888888", width=2, dash="dash"),
                    marker=dict(size=7, color="#888888", symbol="diamond"),
                    hovertemplate="<b>Média do grupo</b><br>Safra: %{x}<br>Nota: %{y:.2f}<extra></extra>"))

                fig_ev.update_layout(
                    height=480,
                    yaxis=dict(title=dict(text="<b>Nota média (9 = mais sadio)</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               range=[0.7, 9.3], tickvals=list(range(1, 10)),
                               tickfont=dict(size=12, color="#111111", weight="bold"),
                               gridcolor="#EEEEEE", zeroline=False),
                    xaxis=dict(title=dict(text="<b>Safra</b>",
                                          font=dict(size=14, color="#111111", weight="bold")),
                               tickfont=dict(size=12, color="#111111", weight="bold"),
                               type="category", gridcolor="#EEEEEE"),
                    legend=dict(title=dict(text="<b>Híbrido</b>", font=dict(size=12, color="#111111")),
                                font=dict(size=11, color="#111111", weight="bold"),
                                bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                    plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                    margin=dict(t=40, b=50, l=60, r=20),
                    font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                    hovermode="closest")
                st.plotly_chart(fig_ev, use_container_width=True)
                _cap_ev = (f"Nota média por safra · {len(hib_ord_ev)} híbridos avaliados em 2+ safras "
                           f"para {doenca_ev}.")
                if n_single:
                    _cap_ev += f" {n_single} híbrido(s) com uma safra só ficaram fora do gráfico."
                st.caption(_cap_ev)

                # tabela de consistência
                media_por_safra = df_ev_plot.groupby("safra")[col_nota_ev].mean().to_dict()
                cons_rows = []
                for hibrido in hib_ord_ev:
                    df_c = df_ev_plot[df_ev_plot["dePara"] == hibrido].sort_values("safra")
                    status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                    notas_por_safra = dict(zip(df_c["safra"], df_c[col_nota_ev]))
                    n_safras = len(notas_por_safra)
                    n_acima = sum(1 for sf, nt in notas_por_safra.items()
                                  if nt >= media_por_safra.get(sf, 0))
                    consist = round(n_acima / n_safras * 100, 0) if n_safras else 0
                    # tendência = slope da reta (com 2 safras = diferença entre as notas)
                    vals = [notas_por_safra[sf] for sf in sorted(notas_por_safra)]
                    tend = "—"
                    if len(vals) >= 2:
                        xs = list(range(len(vals))); n_ = len(xs)
                        mx, my = sum(xs) / n_, sum(vals) / n_
                        den = sum((x - mx) ** 2 for x in xs)
                        slope = (sum((x - mx) * (y - my) for x, y in zip(xs, vals)) / den) if den else 0
                        tend = "↑ melhora" if slope > 0.1 else ("↓ piora" if slope < -0.1 else "→ estável")
                    linha = {"Híbrido": hibrido, "Status": status, "Safras aval.": n_safras}
                    for sf in safras_ev:
                        linha[str(sf)] = round(notas_por_safra[sf], 1) if sf in notas_por_safra else None
                    linha["Acima da média"] = n_acima
                    linha["Consistência %"] = consist
                    linha["Tendência"] = tend
                    cons_rows.append(linha)
                df_cons = pd.DataFrame(cons_rows).sort_values("Consistência %", ascending=False).reset_index(drop=True)

                gb_ev = GridOptionsBuilder.from_dataframe(df_cons)
                gb_ev.configure_default_column(
                    resizable=True, sortable=True, filter=True,
                    cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
                gb_ev.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
                gb_ev.configure_column("Híbrido", pinned="left", width=170)
                gb_ev.configure_column("Status", width=90)
                gb_ev.configure_column("Consistência %", width=130,
                                       cellStyle=js_faixa(FAIXAS_COR["Consistência %"]))
                gb_ev.configure_column("Tendência", width=120,
                                       cellStyle=js_faixa(FAIXAS_COR["Tendência"]))
                for sf in safras_ev:
                    gb_ev.configure_column(str(sf), width=90)

                go_ev = gb_ev.build()
                go_ev["defaultColDef"]["headerClass"] = "ag-header-black"
                AgGrid(
                    df_cons, gridOptions=go_ev,
                    height=min(500, 36 + 32 * len(df_cons) + 20),
                    update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
                    allow_unsafe_jscode=True, enable_enterprise_modules=True,
                    custom_css={
                        ".ag-header":            {"background-color": "#4A4A4A !important"},
                        ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                        ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                        ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                        ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                        ".ag-row":               {"font-size": "13px !important"},
                    },
                    theme="streamlit", use_container_width=True)
                st.caption(
                    "**Consistência %** = em quantas safras o híbrido ficou na média do grupo ou acima. "
                    "**Tendência** = direção da nota entre safras (com 2 safras, é a diferença entre elas).")
                exportar_excel(df_cons, nome_arquivo="evolucao_por_safra.xlsx",
                               faixas_cor=FAIXAS_COR,
                               label="⬇️ Exportar Evolução por Safra", key="exp_ev_sn")

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 8 — PERFIL MULTIDOENÇA
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Perfil Multidoença",
    "Qual é o perfil sanitário completo de cada híbrido?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Mostra o perfil sanitário de um híbrido em **todas as doenças ao mesmo tempo**, em dois visuais
lado a lado. A nota usada é a **menor observada** (o pior caso) — de propósito: revela a
vulnerabilidade que a média ou a moda escondem. Um híbrido pode ser tipicamente resistente e ainda
assim ter desabado numa doença num local; é isso que o perfil expõe. O hover mostra quantas
observações com nota > 0 entraram.

**Radar (esquerda)**
Cada eixo é uma doença. Quanto maior a área, melhor o perfil geral. Serve para ver de relance se o
híbrido é equilibrado ou tem um ponto fraco — uma área "amassada" num eixo é o calcanhar de aquiles.

*Atenção: a área do radar muda conforme a ordem das doenças nos eixos. Use para impressão geral,
não para comparar números com precisão — para isso, as barras.*

**Barras horizontais (direita)**
As mesmas doenças e valores, sem distorção geométrica. Cada barra é a nota (pior caso), colorida
pela classe (AS vermelho-escuro → R verde). Use para ler os valores exatos e comparar híbridos.

**Usando os dois juntos**
O radar dá a impressão rápida; as barras confirmam os números. Se o radar parecer grande mas as
barras tiverem células amarelas, o híbrido é equilibrado mas não excelente.

**Escala:** nota 9 = resistente, 1 = altamente suscetível. A linha vermelha nas barras (nota 6) é o
limite entre zona de atenção e zona segura.
""")

_hibridos_pm = sorted(ta_filtrado["dePara"].dropna().unique().tolist())
hibridos_pm_sel = st.multiselect(
    "Selecione híbridos para comparar (máx. 6 recomendado)",
    options=_hibridos_pm,
    default=_hibridos_pm[:min(4, len(_hibridos_pm))],
    key="sn_hibridos_pm")

if not hibridos_pm_sel:
    st.info("Selecione ao menos um híbrido para exibir o perfil.")
else:
    pm_rows = []
    for doenca in DOENCAS:
        col_nota_pm = DOENCAS[doenca]["nota"]
        if col_nota_pm not in ta_filtrado.columns:
            continue
        s_all = pd.to_numeric(ta_filtrado[col_nota_pm], errors="coerce")
        if s_all[s_all > 0].empty:
            continue
        for hibrido in hibridos_pm_sel:
            mask = (ta_filtrado["dePara"] == hibrido)
            s_c = s_all[mask]
            s_c = s_c[s_c > 0].dropna()
            if s_c.empty:
                continue
            nota_min = float(s_c.min())
            pm_rows.append({
                "hibrido": hibrido, "doenca": SIGLAS.get(doenca, doenca),
                "nota": nota_min, "classe": nota_para_classe(nota_min),
                "n_obs": int(len(s_c)), "n_total": int(mask.sum())})

    if not pm_rows:
        st.info("Sem dados suficientes para os híbridos selecionados.")
    else:
        df_pm = pd.DataFrame(pm_rows)
        doencas_pm = sorted(df_pm["doenca"].unique().tolist())
        palette_pm = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                      "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF"]
        cor_pm = {c: palette_pm[i % len(palette_pm)] for i, c in enumerate(hibridos_pm_sel)}

        col_radar, col_bar = st.columns(2)

        with col_radar:
            fig_radar = go_plt.Figure()
            for hibrido in hibridos_pm_sel:
                df_c = df_pm[df_pm["hibrido"] == hibrido]
                notas_radar = []
                for d in doencas_pm:
                    row = df_c[df_c["doenca"] == d]
                    notas_radar.append(float(row["nota"].values[0]) if not row.empty else 0)
                theta = doencas_pm + [doencas_pm[0]]
                r = notas_radar + [notas_radar[0]]
                hexc = cor_pm[hibrido].lstrip("#")
                rc, gc, bc = int(hexc[0:2], 16), int(hexc[2:4], 16), int(hexc[4:6], 16)
                fig_radar.add_trace(go_plt.Scatterpolar(
                    r=r, theta=theta, fill="toself", name=hibrido,
                    line=dict(color=cor_pm[hibrido], width=2),
                    fillcolor=f"rgba({rc},{gc},{bc},0.15)",
                    hovertemplate="<b>%{fullData.name}</b><br>%{theta}: %{r:.1f}<extra></extra>"))
            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(visible=True, range=[0, 9], tickvals=[1, 3, 5, 7, 9],
                                    tickfont=dict(size=10, color="#555555", weight="bold"),
                                    gridcolor="#DDDDDD"),
                    angularaxis=dict(tickfont=dict(size=11, color="#111111", weight="bold"),
                                     gridcolor="#DDDDDD"),
                    bgcolor="#FAFAFA"),
                legend=dict(font=dict(size=11, color="#111111", weight="bold"),
                            bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                paper_bgcolor="#FFFFFF", margin=dict(t=40, b=40, l=40, r=40), height=420,
                font=dict(family="Helvetica Neue, sans-serif"))
            st.plotly_chart(fig_radar, use_container_width=True)
            st.caption("Área varia com a ordem das doenças — use para impressão geral. "
                       "Nota = menor valor observado (pior caso).")

        with col_bar:
            fig_bar_pm = go_plt.Figure()
            ordem_doencas_pm = (df_pm.groupby("doenca")["nota"].mean()
                                .sort_values(ascending=True).index.tolist())
            for hibrido in hibridos_pm_sel:
                df_c = df_pm[df_pm["hibrido"] == hibrido].set_index("doenca")
                notas_bar, hover_bar, textos = [], [], []
                for d in ordem_doencas_pm:
                    if d in df_c.index:
                        n = df_c.loc[d, "nota"]; cls = df_c.loc[d, "classe"]
                        n_obs = int(df_c.loc[d, "n_obs"]); n_tot = int(df_c.loc[d, "n_total"])
                        notas_bar.append(n)
                        hover_bar.append(f"<b>{hibrido}</b><br>{d}: {n:.1f} ({cls})<br>"
                                         f"{n_obs}/{n_tot} obs com nota > 0")
                        textos.append(f"{n:.1f} ({n_obs}/{n_tot})")
                    else:
                        notas_bar.append(None); hover_bar.append(f"<b>{hibrido}</b><br>{d}: sem dado")
                        textos.append("—")
                fig_bar_pm.add_trace(go_plt.Bar(
                    name=hibrido, y=ordem_doencas_pm, x=notas_bar, orientation="h",
                    marker_color=cor_pm[hibrido], text=textos, textposition="outside",
                    textfont=dict(size=11, weight="bold"), hovertext=hover_bar, hoverinfo="text"))
            fig_bar_pm.add_vline(x=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
                                 annotation_text="nota 6", annotation_position="top",
                                 annotation_font=dict(size=13, color="#E74C3C", weight="bold"))
            fig_bar_pm.update_layout(
                barmode="group", height=420,
                xaxis=dict(title=dict(text="<b>Nota (1 = pior · 9 = melhor)</b>",
                                      font=dict(size=14, color="#111111", weight="bold")),
                           range=[0, 10.5], tickvals=list(range(0, 10)),
                           tickfont=dict(size=11, color="#111111", weight="bold"), gridcolor="#EEEEEE"),
                yaxis=dict(tickfont=dict(size=11, color="#111111", weight="bold"),
                           gridcolor="#EEEEEE", categoryorder="array", categoryarray=ordem_doencas_pm),
                legend=dict(font=dict(size=11, color="#111111", weight="bold"),
                            bgcolor="rgba(255,255,255,0.85)", bordercolor="#DDDDDD", borderwidth=1),
                plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=40, l=10, r=60),
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                hovermode="y unified")
            st.plotly_chart(fig_bar_pm, use_container_width=True)
            st.caption("Nota = menor valor observado (pior caso). Rótulo: nota (obs com nota > 0 / "
                       "total de registros). Doenças ordenadas da pior para a melhor nota.")

            n_locais_total_pm = ta_filtrado["cod_fazenda"].nunique()
            _linhas_rodape_pm = []
            for doenca in DOENCAS:
                col_nota_pm2 = DOENCAS[doenca]["nota"]
                if col_nota_pm2 not in ta_filtrado.columns:
                    continue
                s_pm2 = pd.to_numeric(ta_filtrado[col_nota_pm2], errors="coerce")
                n_loc_pm2 = ta_filtrado[s_pm2 > 0]["cod_fazenda"].nunique()
                if n_loc_pm2 < n_locais_total_pm:
                    _linhas_rodape_pm.append(f"**{SIGLAS.get(doenca, doenca)}**: {n_loc_pm2}/{n_locais_total_pm}")
            if _linhas_rodape_pm:
                st.caption("Locais com nota registrada (de " + str(n_locais_total_pm) + " ativos): "
                           + " · ".join(_linhas_rodape_pm))

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 9 — HEATMAP HÍBRIDO × LOCAL
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Heatmap por Local",
    "Como cada híbrido se saiu em cada local para uma doença?",
    contexto_str,
)

_col_interp_hm2, _col_dic_hm2 = st.columns([1, 9])
with _col_interp_hm2:
    with st.popover("ℹ️ Como interpretar", use_container_width=False):
        st.markdown("""
Para a doença escolhida, cada célula mostra a nota do híbrido naquele local. O ponto **•** indica
que a doença foi detectada ali (incidência > 0%). A célula é colorida pela classe da nota.
""")
        _cls_cols = st.columns(5)
        for i, (cls, cor) in enumerate(COR_CLASS.items()):
            fg = COR_TEXTO_CLASS[cls]
            _cls_cols[i].markdown(
                f'<div style="background:{cor};color:{fg};border-radius:6px;padding:6px;'
                f'text-align:center;font-size:12px;font-weight:700;">{cls}<br>'
                f'<span style="font-weight:400;font-size:11px;">{LABEL_CLASS[cls].split(" — ")[1].split(" (")[0]}</span></div>',
                unsafe_allow_html=True)
        st.markdown("""
**Destaques automáticos**
- **Borda vermelha + ⚠️** → local com maior pressão da doença (a nota mais frequente do grupo ali
  é baixa, abaixo de 6).
- **★ verde** → híbrido com melhor perfil geral no filtro ativo (a nota mais frequente entre os
  locais é a mais alta).

Híbrido não avaliado num local aparece como célula cinza (—). A linha preta separa grupos de status.
""")

with _col_dic_hm2:
    with st.popover(f"📍 Dicionário de locais ({ta_filtrado['cod_fazenda'].nunique()})",
                    use_container_width=False):
        _df_dic_hm2 = (ta_filtrado[["cod_fazenda", "nomeFazenda", "cidade_nome", "estado_sigla"]]
                       .drop_duplicates()
                       .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                       .rename(columns={"cod_fazenda": "Código", "nomeFazenda": "Local",
                                        "cidade_nome": "Cidade", "estado_sigla": "Estado"})
                       .reset_index(drop=True))
        st.markdown("Referência dos códigos exibidos nas colunas do heatmap.")
        st.dataframe(_df_dic_hm2, hide_index=True, use_container_width=True)

if not doencas_disp:
    st.info("Nenhuma doença com nota registrada nos filtros ativos.")
else:
    _c_hm_sel, _c_hm_chk = st.columns([3, 1])
    with _c_hm_sel:
        doenca_hm2 = st.selectbox("Selecione a doença", options=doencas_disp,
                                  format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}", key="sn_doenca_hm2")
    with _c_hm_chk:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        mostrar_nota_hm2 = st.checkbox("Mostrar nota", value=False, key="sn_hm2_mostrar_nota",
                                       help="A classe (T, R, MT…) aparece sempre; marque para ver "
                                            "também a nota numérica em cada célula.")
    col_nota_hm2 = DOENCAS[doenca_hm2]["nota"]

    if col_nota_hm2 not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        df_hm2 = ta_filtrado[["dePara", "status_material", "cod_fazenda",
                              "cidade_nome", "estado_sigla", col_nota_hm2]].copy()
        df_hm2[col_nota_hm2] = pd.to_numeric(df_hm2[col_nota_hm2], errors="coerce")
        df_hm2 = df_hm2[df_hm2[col_nota_hm2] > 0].dropna(subset=[col_nota_hm2])

        if df_hm2.empty:
            st.info("Nenhuma avaliação disponível para esta doença nos filtros ativos.")
        else:
            df_hm2_agg = (df_hm2.groupby(["dePara", "status_material", "cod_fazenda",
                                          "cidade_nome", "estado_sigla"])
                          [col_nota_hm2].mean().round(1).reset_index())
            locais_hm2 = (df_hm2_agg[["cod_fazenda", "cidade_nome", "estado_sigla"]]
                          .drop_duplicates()
                          .sort_values(["estado_sigla", "cidade_nome", "cod_fazenda"])
                          ["cod_fazenda"].tolist())
            cult_status_hm2 = (df_hm2_agg[["dePara", "status_material"]].drop_duplicates()
                               .assign(_ord=lambda d: d["status_material"].apply(
                                   lambda s: ORDEM_STATUS.index(s) if s in ORDEM_STATUS else 99))
                               .sort_values(["_ord", "dePara"]))
            hibridos_hm2 = cult_status_hm2["dePara"].tolist()
            status_map_hm2 = cult_status_hm2.set_index("dePara")["status_material"].to_dict()

            pivot_hm2 = df_hm2_agg.pivot_table(index="dePara", columns="cod_fazenda",
                                               values=col_nota_hm2, aggfunc="mean").reindex(
                index=hibridos_hm2, columns=locais_hm2)

            df_hm2_inc = df_hm2.copy()
            df_hm2_inc["_inc"] = df_hm2_inc[col_nota_hm2].between(1, 5).astype(int)
            pivot_inc_hm2 = df_hm2_inc.groupby(["dePara", "cod_fazenda"]).apply(
                lambda g: round(g["_inc"].sum() / len(g) * 100, 1)).unstack(fill_value=np.nan)

            # matriz de notas (z) — fundo colorido pelo go.Heatmap NATIVO (leve, não trava com
            # muitas células). Os números vêm de annotations, com a cor da classe (COR_TEXTO_CLASS).
            _z_hm2 = [[(None if (v is None or (isinstance(v, float) and np.isnan(v))) else float(v))
                       for v in (pivot_hm2.loc[h].tolist() if h in pivot_hm2.index
                                 else [None] * len(locais_hm2))]
                      for h in hibridos_hm2]
            # colorscale por CLASSE, casando EXATO com nota_para_classe (que usa <=): n<=2 AS,
            # n<=4 S, n<=6 MT, n<=8 T, >8 R. A cor muda logo ACIMA de 2,4,6,8 (epsilon), então
            # cada nota — inteira ou média fracionária — pega a cor da sua classe, sem borrar.
            def _p(n):
                return (n - 1) / 8
            _eps = 0.0001
            _fr = [_p(2) + _eps, _p(4) + _eps, _p(6) + _eps, _p(8) + _eps]
            colorscale_hm2 = [
                [0.0, COR_CLASS["AS"]], [_fr[0], COR_CLASS["AS"]],
                [_fr[0], COR_CLASS["S"]], [_fr[1], COR_CLASS["S"]],
                [_fr[1], COR_CLASS["MT"]], [_fr[2], COR_CLASS["MT"]],
                [_fr[2], COR_CLASS["T"]], [_fr[3], COR_CLASS["T"]],
                [_fr[3], COR_CLASS["R"]], [1.0, COR_CLASS["R"]]]

            # duas matrizes de texto (números escuros nas células claras, brancos nas escuras),
            # cada uma num heatmap com texttemplate — renderiza tudo numa passada, SEM annotation
            # por célula (que travava com muitas células). Cor do texto vem de COR_TEXTO_CLASS.
            _txt_esc_hm2, _txt_bra_hm2 = [], []
            for hibrido in hibridos_hm2:
                lin_e, lin_b = [], []
                for local in locais_hm2:
                    v = pivot_hm2.loc[hibrido, local] if (hibrido in pivot_hm2.index and local in pivot_hm2.columns) else None
                    if v is None or (isinstance(v, float) and np.isnan(v)):
                        lin_e.append(""); lin_b.append("")
                    else:
                        cls = nota_para_classe(v)
                        inc_v = None
                        try:
                            inc_v = pivot_inc_hm2.loc[hibrido, local]
                        except Exception:
                            pass
                        ast = " •" if inc_v is not None and not (isinstance(inc_v, float) and np.isnan(inc_v)) and inc_v > 0 else ""
                        # classe é o padrão; a nota entra só se o usuário marcar "Mostrar nota"
                        if mostrar_nota_hm2:
                            s = f"{v:.1f}{ast}<br><b>{cls}</b>"
                        else:
                            s = f"<b>{cls}</b>{ast}"
                        if COR_TEXTO_CLASS.get(cls, "#1A1A1A") == "#FFFFFF":
                            lin_b.append(s); lin_e.append("")
                        else:
                            lin_e.append(s); lin_b.append("")
                _txt_esc_hm2.append(lin_e); _txt_bra_hm2.append(lin_b)

            fig_hm2 = go_plt.Figure()
            # trace 1: fundo colorido por classe + números escuros (classes MT/T, fundo claro)
            fig_hm2.add_trace(go_plt.Heatmap(
                z=_z_hm2, x=locais_hm2, y=hibridos_hm2, text=_txt_esc_hm2, texttemplate="%{text}",
                textfont=dict(size=10, color="#1A1A1A"),
                zmin=1, zmax=9, colorscale=colorscale_hm2,
                xgap=1, ygap=1, hoverongaps=False,
                colorbar=dict(title=dict(text="nota", side="right"), thickness=12, len=0.6),
                hovertemplate="%{y} · %{x}<br>nota: %{z:.1f}<extra></extra>"))
            # trace 2: transparente + números brancos (classes AS/S/R, fundo escuro)
            fig_hm2.add_trace(go_plt.Heatmap(
                z=_z_hm2, x=locais_hm2, y=hibridos_hm2, text=_txt_bra_hm2, texttemplate="%{text}",
                textfont=dict(size=10, color="#FFFFFF"),
                zmin=1, zmax=9, colorscale=[[0, "rgba(0,0,0,0)"], [1, "rgba(0,0,0,0)"]],
                showscale=False, hoverinfo="skip"))

            for i in range(len(hibridos_hm2) - 1):
                if status_map_hm2.get(hibridos_hm2[i], "") != status_map_hm2.get(hibridos_hm2[i + 1], ""):
                    fig_hm2.add_shape(type="line", xref="paper", x0=0, x1=1,
                                      y0=i + 0.5, y1=i + 0.5, line=dict(color="#333333", width=2))

            # destaque do local mais crítico (menor nota mais frequente do grupo, < 6)
            moda_local_hm2 = (df_hm2_agg.groupby("cod_fazenda")[col_nota_hm2]
                              .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else np.nan))
            if not moda_local_hm2.empty and len(locais_hm2) > 0:
                local_crit_hm2 = moda_local_hm2.idxmin()
                nota_crit_hm2 = round(moda_local_hm2.min(), 1)
                if local_crit_hm2 in locais_hm2 and nota_crit_hm2 < 6:
                    j_crit = locais_hm2.index(local_crit_hm2)
                    fig_hm2.add_shape(type="rect", xref="x", yref="paper",
                                      x0=j_crit - 0.5, x1=j_crit + 0.5, y0=0, y1=1,
                                      fillcolor="rgba(0,0,0,0)", line=dict(color="#E74C3C", width=2.5))
                    fig_hm2.add_annotation(x=local_crit_hm2, xref="x", y=1.0, yref="paper",
                                           text=f"⚠️ {nota_crit_hm2}",
                                           showarrow=False, xanchor="center", yanchor="bottom", yshift=4,
                                           font=dict(size=10, color="#E74C3C"),
                                           bgcolor="rgba(255,255,255,0.9)", bordercolor="#E74C3C",
                                           borderwidth=1, borderpad=2)

            # híbrido com melhor perfil geral (maior nota mais frequente)
            moda_cult_hm2 = (df_hm2_agg.groupby("dePara")[col_nota_hm2]
                             .apply(lambda s: s.mode().iloc[0] if not s.mode().empty else np.nan))
            if not moda_cult_hm2.empty and len(hibridos_hm2) > 0:
                melhor_cult = moda_cult_hm2.idxmax()
                nota_melhor = round(moda_cult_hm2.max(), 1)
                if melhor_cult in hibridos_hm2:
                    fig_hm2.add_annotation(x=1.0, xref="paper", y=melhor_cult, yref="y",
                                           text=f"★ melhor nota {nota_melhor}", showarrow=False,
                                           xanchor="left", yanchor="middle", xshift=8,
                                           font=dict(size=11, color="#1E7A34", weight="bold"),
                                           bgcolor="rgba(255,255,255,0.85)", bordercolor="#1E7A34",
                                           borderwidth=1, borderpad=2)

            for hibrido in hibridos_hm2:
                cor_s = COR_STATUS_PLOT.get(status_map_hm2.get(hibrido, ""), "#333333")
                # texto do híbrido: usa a cor do status, mas escurece o verde-vibrante do EXP p/ legibilidade
                if status_map_hm2.get(hibrido, "") == "EXP":
                    cor_s = "#1A7A1A"
                fig_hm2.add_annotation(x=-0.01, xref="paper", y=hibrido, yref="y", text=f"<b>{hibrido}</b>",
                                       showarrow=False, xanchor="right", yanchor="middle",
                                       font=dict(size=11, color=cor_s, weight="bold"))

            altura_hm2 = max(400, len(hibridos_hm2) * 28 + 80)
            fig_hm2.update_layout(
                height=altura_hm2,
                xaxis=dict(side="bottom", tickfont=dict(size=11, color="#111111", weight="bold"),
                           tickangle=-45, title=dict(text="<b>Local (cod_fazenda)</b>",
                                                     font=dict(size=14, color="#111111", weight="bold")),
                           categoryorder="array", categoryarray=locais_hm2),
                yaxis=dict(tickfont=dict(size=11, color="#111111", weight="bold"),
                           autorange="reversed", showticklabels=False),
                margin=dict(t=30, b=100, l=200, r=140),
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"))

            st.plotly_chart(fig_hm2, use_container_width=True)
            n_locais_total_hm2 = ta_filtrado["cod_fazenda"].nunique()
            _cap_hm2 = (f"Cor da célula = classe da nota do híbrido no local. **•** = doença presente "
                        f"(incidência > 0%). Cinza = não avaliado. Linha preta = divisão de status. "
                        f"**⚠️** = local de maior pressão. **★** = híbrido de melhor nota geral. "
                        f"{len(hibridos_hm2)} híbridos · {len(locais_hm2)} locais para {doenca_hm2}")
            _cap_hm2 += (f" (de {n_locais_total_hm2} locais ativos — os demais não têm nota para esta doença)."
                         if len(locais_hm2) < n_locais_total_hm2 else ".")
            st.caption(_cap_hm2)

st.divider()


# ══════════════════════════════════════════════════════════════════════════════
# SEÇÃO 10 — ANÁLISE DE SOBREVIVÊNCIA + MAPA DE COLAPSO
# ══════════════════════════════════════════════════════════════════════════════
secao_titulo(
    "Análise de Sobrevivência",
    "Qual híbrido mantém a sanidade sob pressão crescente?",
    contexto_str,
)

with st.popover("ℹ️ Como interpretar", use_container_width=False):
    st.markdown("""
Cada curva mostra, para um nível de exigência de nota (eixo X), **em quantos % dos locais o híbrido
atingiu ao menos aquela nota**. Um híbrido que se mantém alto mesmo com a exigência subindo é o mais
resistente — aguenta a pressão da doença sem "cair".

A ideia é emprestada das curvas de sobrevivência: em vez de "quantos sobrevivem ao longo do tempo",
é "em quantos locais o híbrido se mantém sadio conforme a régua aperta".

**Eixos**
- **X — Nota** → régua de exigência, de 1 (fácil) a 9 (máxima).
- **Y — Sobrevivência** → % de locais em que o híbrido atingiu ao menos aquela nota.

**Destaques**
- **▲ melhor / ▼ pior** → híbridos com maior e menor área sob a curva (ASC).
- **Faixa amarela (notas 4–6)** → zona de atenção, onde a resistência começa a ceder.
- **Linha vermelha (nota 6)** → fronteira MT/T.

**Tabela ASC** → resume o desempenho num único número: quanto maior a área sob a curva, melhor o
perfil sanitário geral do híbrido naquela doença.

**Mapa de Colapso** (abaixo) → em quantos locais cada híbrido teve nota ≤ 4 (situação crítica, S ou
AS), com a lista dos locais onde isso aconteceu.
""")

if not doencas_disp:
    st.info("Nenhuma doença com nota registrada nos filtros ativos.")
else:
    doenca_surv = st.selectbox("Selecione a doença", options=doencas_disp,
                               format_func=lambda d: f"{SIGLAS.get(d, d)} — {d}", key="sn_doenca_surv")
    col_nota_sv = DOENCAS[doenca_surv]["nota"]

    if col_nota_sv not in ta_filtrado.columns:
        st.warning("Coluna de nota não encontrada para esta doença.")
    else:
        df_sv = ta_filtrado[["dePara", "status_material", "cod_fazenda", col_nota_sv]].copy()
        df_sv[col_nota_sv] = pd.to_numeric(df_sv[col_nota_sv], errors="coerce")
        df_sv = df_sv[df_sv[col_nota_sv] > 0].dropna(subset=[col_nota_sv])
        df_sv_agg = (df_sv.groupby(["dePara", "status_material", "cod_fazenda"])
                     [col_nota_sv].mean().round(1).reset_index())
        hibridos_sv = sorted(df_sv_agg["dePara"].unique().tolist())

        if df_sv_agg.empty or not hibridos_sv:
            st.info("Sem dados suficientes para esta doença nos filtros ativos.")
        else:
            thresholds = [t / 10 for t in range(10, 91)]  # 1.0 → 9.0
            palette_sv = ["#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
                          "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
                          "#AEC7E8", "#FFBB78", "#98DF8A", "#FF9896", "#C5B0D5",
                          "#C49C94", "#F7B6D2", "#C7C7C7", "#DBDB8D", "#9EDAE5",
                          "#393B79", "#637939", "#8C6D31", "#843C39", "#7B4173"]
            cor_sv = {c: palette_sv[i % len(palette_sv)] for i, c in enumerate(hibridos_sv)}

            curvas, asc_vals = {}, {}
            for hibrido in hibridos_sv:
                notas = df_sv_agg[df_sv_agg["dePara"] == hibrido][col_nota_sv].tolist()
                n = len(notas)
                sobrev = [sum(1 for v in notas if v >= t) / n * 100 for t in thresholds]
                curvas[hibrido] = sobrev
                asc_vals[hibrido] = float(np.trapezoid(sobrev, thresholds))

            df_asc = (pd.DataFrame.from_dict(asc_vals, orient="index", columns=["Área sob a curva"])
                      .reset_index().rename(columns={"index": "Híbrido"})
                      .sort_values("Área sob a curva", ascending=False).reset_index(drop=True))
            df_asc["Área sob a curva"] = df_asc["Área sob a curva"].round(2)

            if len(hibridos_sv) > 12:
                st.info(
                    f"São **{len(hibridos_sv)} híbridos** — muitas curvas juntas. Use o filtro de "
                    "híbrido na barra lateral para comparar menos; a tabela ASC (no gráfico) e o "
                    "Mapa de Colapso (abaixo) cobrem todos.")

            fig_sv = go_plt.Figure()
            for hibrido in hibridos_sv:
                status = df_sv_agg[df_sv_agg["dePara"] == hibrido]["status_material"].mode()
                status = status[0] if not status.empty else ""
                fig_sv.add_trace(go_plt.Scatter(
                    x=thresholds, y=curvas[hibrido], mode="lines", name=hibrido,
                    line=dict(color=cor_sv[hibrido], width=2, shape="spline", smoothing=0.5),
                    customdata=[[status, round(asc_vals[hibrido], 2)]] * len(thresholds),
                    hovertemplate=(f"<b>{hibrido}</b><br>Status: %{{customdata[0]}}<br>"
                                   f"Nota ≥ %{{x:.1f}}: %{{y:.1f}}% dos locais<br>"
                                   f"ASC: %{{customdata[1]}}<extra></extra>")))

            fig_sv.add_vline(x=6, line_dash="dot", line_color="#E74C3C", line_width=1.5,
                             annotation_text="Limite MT/T (nota 6)", annotation_position="top right",
                             annotation_font=dict(size=13, color="#E74C3C", weight="bold"))
            fig_sv.add_vrect(x0=4, x1=6, fillcolor="rgba(255,214,0,0.10)", layer="below",
                             line_width=0, annotation_text="zona de atenção",
                             annotation_position="top left",
                             annotation_font=dict(size=13, color="#B8860B", weight="bold"))

            melhor_sv = df_asc.iloc[0]["Híbrido"]
            pior_sv = df_asc.iloc[-1]["Híbrido"]
            for hibrido, label, cor_l, anchor in [(melhor_sv, "▲ melhor", "#1E7A34", "bottom"),
                                                  (pior_sv, "▼ pior", "#8B0000", "top")]:
                if hibrido in curvas:
                    fig_sv.add_annotation(
                        x=9, y=curvas[hibrido][-1],
                        text=f"<b>{hibrido}</b><br><span style='font-size:10px'>{label}</span>",
                        showarrow=False, xanchor="left", yanchor=anchor, xshift=8,
                        font=dict(size=10, color=cor_l), bgcolor="rgba(255,255,255,0.85)",
                        bordercolor=cor_l, borderwidth=1, borderpad=3)

            top10 = df_asc.head(10)
            _fill = ["#D5F5D5" if i < 3 else "#F9F9F9" for i in range(len(top10))]
            fig_sv.add_trace(go_plt.Table(
                domain=dict(x=[0.63, 0.88], y=[0.40, 1.0]), columnwidth=[100, 50],
                header=dict(values=["<b>Híbrido</b>", "<b>ASC</b>"], fill_color="#4A4A4A",
                            font=dict(color="white", size=11), align="center", height=26),
                cells=dict(values=[top10["Híbrido"].tolist(), top10["Área sob a curva"].tolist()],
                           fill_color=[_fill, _fill], font=dict(color="#111111", size=11),
                           align=["left", "center"], height=22)))

            fig_sv.update_layout(
                height=520,
                xaxis=dict(title=dict(text="<b>Nota</b>", font=dict(size=14, color="#111111", weight="bold")),
                           tickvals=list(range(1, 10)), tickfont=dict(size=12, color="#111111", weight="bold"),
                           range=[0.9, 9.1], autorange=False, gridcolor="#CCCCCC", griddash="dot",
                           gridwidth=1, domain=[0, 0.60]),
                yaxis=dict(title=dict(text="<b>Sobrevivência (% de locais)</b>",
                                      font=dict(size=14, color="#111111", weight="bold")),
                           tickformat=".0f", ticksuffix="%", range=[-5, 105], autorange=False,
                           tickfont=dict(size=12, color="#111111", weight="bold"),
                           gridcolor="#CCCCCC", griddash="dot", gridwidth=1, zeroline=False),
                legend=dict(title=dict(text="<b>Híbrido</b>", font=dict(size=12, color="#111111", weight="bold")),
                            font=dict(size=11, color="#111111", weight="bold"), x=1.02, y=1,
                            xanchor="left", bgcolor="rgba(255,255,255,0.85)",
                            bordercolor="#DDDDDD", borderwidth=1),
                plot_bgcolor="#FAFAFA", paper_bgcolor="#FFFFFF",
                margin=dict(t=40, b=60, l=60, r=160),
                font=dict(family="Helvetica Neue, sans-serif", size=12, color="#111111"),
                hovermode="x unified")

            n_locais_sv = df_sv_agg["cod_fazenda"].nunique()
            n_locais_total_sv = ta_filtrado["cod_fazenda"].nunique()
            st.plotly_chart(fig_sv, use_container_width=True)
            _cap_sv = f"{doenca_surv} · {len(hibridos_sv)} híbridos · {n_locais_sv} locais avaliados"
            _cap_sv += (f" (de {n_locais_total_sv} ativos — os demais não têm nota para esta doença)."
                        if n_locais_sv < n_locais_total_sv else ".")
            st.caption(_cap_sv)

            st.divider()
            st.markdown("#### Mapa de Colapso — locais com S ou AS (nota ≤ 4)")

            colapso_rows = []
            for hibrido in hibridos_sv:
                df_c = df_sv_agg[df_sv_agg["dePara"] == hibrido]
                status = df_c["status_material"].mode()[0] if not df_c["status_material"].mode().empty else ""
                n_total = len(df_c)
                n_colapso = int((df_c[col_nota_sv] <= 4).sum())
                n_mt = int(((df_c[col_nota_sv] > 4) & (df_c[col_nota_sv] <= 6)).sum())
                n_ok = int((df_c[col_nota_sv] > 6).sum())
                n_inc = int(df_c[col_nota_sv].between(1, 5).sum())
                locais_col = df_c[df_c[col_nota_sv] <= 4]["cod_fazenda"].tolist()
                locais_inc = df_c[df_c[col_nota_sv].between(1, 5)]["cod_fazenda"].tolist()
                colapso_rows.append({
                    "Híbrido": hibrido, "Status": status, "Locais Aval.": n_total,
                    "S / AS (≤4)": n_colapso, "MT (5–6)": n_mt, "T / R (≥7)": n_ok,
                    "% Colapso": round(n_colapso / n_total * 100, 1) if n_total else 0,
                    "% Incidência": round(n_inc / n_total * 100, 1) if n_total else 0,
                    "Locais em Colapso": ", ".join(locais_col) if locais_col else "—",
                    "Locais c/ Incidência": ", ".join(locais_inc) if locais_inc else "—"})

            df_colapso = (pd.DataFrame(colapso_rows)
                          .sort_values(["S / AS (≤4)", "% Colapso"], ascending=[False, False])
                          .reset_index(drop=True))

            gb_col = GridOptionsBuilder.from_dataframe(df_colapso)
            gb_col.configure_default_column(
                resizable=True, sortable=True, filter=True,
                cellStyle={"fontSize": "13px", "fontFamily": "Helvetica Neue, sans-serif"})
            gb_col.configure_grid_options(headerHeight=36, rowHeight=32, domLayout="normal")
            gb_col.configure_column("Híbrido", pinned="left", width=170)
            gb_col.configure_column("Status", width=90)
            gb_col.configure_column("Locais Aval.", width=100)
            gb_col.configure_column("S / AS (≤4)", width=100,
                                    cellStyle=js_faixa(FAIXAS_COR["S / AS (≤4)"]))
            gb_col.configure_column("MT (5–6)", width=90,
                                    cellStyle=js_faixa(FAIXAS_COR["MT (5–6)"]))
            gb_col.configure_column("T / R (≥7)", width=90,
                                    cellStyle=js_faixa(FAIXAS_COR["T / R (≥7)"]))
            gb_col.configure_column("% Colapso", width=100,
                                    cellStyle=js_faixa(FAIXAS_COR["% Colapso"]))
            gb_col.configure_column("% Incidência", width=105,
                                    cellStyle=js_faixa(FAIXAS_COR["% Incidência"]))
            gb_col.configure_column("Locais em Colapso", width=280)
            gb_col.configure_column("Locais c/ Incidência", width=280)

            go_col = gb_col.build()
            go_col["defaultColDef"]["headerClass"] = "ag-header-black"
            AgGrid(
                df_colapso, gridOptions=go_col,
                height=min(620, 36 + 32 * len(df_colapso) + 20),
                update_mode=GridUpdateMode.NO_UPDATE, fit_columns_on_grid_load=False,
                allow_unsafe_jscode=True, enable_enterprise_modules=True,
                custom_css={
                    ".ag-header":            {"background-color": "#4A4A4A !important"},
                    ".ag-header-row":        {"background-color": "#4A4A4A !important"},
                    ".ag-header-cell":       {"background-color": "#4A4A4A !important"},
                    ".ag-header-cell-label": {"color": "#FFFFFF !important", "font-weight": "700"},
                    ".ag-header-cell-text":  {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
                    ".ag-icon":              {"color": "#FFFFFF !important", "opacity": "1 !important"},
                    ".ag-row":               {"font-size": "13px !important"},
                },
                theme="streamlit", use_container_width=True)
            st.caption(
                "**S / AS (≤4)** = nº de locais em colapso sanitário · **% Colapso** = % de locais "
                "com nota ≤ 4 · **% Incidência** = % de locais com doença detectada (notas 1–5) · "
                "**MT (5–6)** = zona de atenção · **T / R (≥7)** = sanidade adequada. Ordenado pelo "
                "maior nº de colapsos.")
            exportar_excel(df_colapso, nome_arquivo="mapa_colapso.xlsx",
                           faixas_cor=FAIXAS_COR,
                           label="⬇️ Exportar Mapa de Colapso", key="exp_colapso_sn")

st.divider()

rodape()
